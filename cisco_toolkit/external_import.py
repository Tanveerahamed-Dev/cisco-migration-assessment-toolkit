"""Offline external source-of-truth import + intent-vs-observed reconcile (roadmap B).

`ssot.reconcile` (ssot.py:149) checks the snapshot against its OWN raw basis — internal consistency only.
This is the complementary reconcile NetClaw's `netbox-reconcile` and the SoT-copilots do: ingest a DECLARED
inventory from a file drop (a CMDB / NetBox / Nautobot CSV/XLSX export — read-only, air-gapped, never a
live API) and diff it against the collected evidence, emitting a named drift taxonomy:

  MISSING_DEVICE      — declared in the SoT, never observed in the collection
  UNDOCUMENTED_DEVICE — observed on the wire, absent from the SoT
  MODEL_MISMATCH      — declared model != observed model
  IP_DRIFT            — declared mgmt IP != observed mgmt IP
  UNVERIFIABLE        — declared AND in the inventory, but never collected (a blind spot): we refuse to
                        claim a match OR a miss — the coverage-honest 5th state NetClaw lacks.

Opt-in (mirrors the `rest_collect` posture): the engine never reads an external file unless asked.
"""
from __future__ import annotations

import csv
import io
from typing import Any, Dict, List

# external column header -> our canonical field. Lower-cased, stripped before lookup.
_ALIASES = {
    "host": "host", "hostname": "host", "name": "host", "device": "host", "device_name": "host",
    "model": "model", "device_type": "model", "type": "model", "platform": "model", "part_number": "model",
    "serial": "serial", "serial_number": "serial", "sn": "serial",
    "mgmt_ip": "mgmt_ip", "ip": "mgmt_ip", "ip_address": "mgmt_ip", "management_ip": "mgmt_ip",
    "primary_ip": "mgmt_ip", "primary_ip4": "mgmt_ip",
}


def normalize_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Map arbitrary inventory-export columns onto {host, model, serial, mgmt_ip} (only the fields present)."""
    out: List[Dict[str, Any]] = []
    for r in rows or []:
        norm: Dict[str, Any] = {}
        for k, v in (r or {}).items():
            canon = _ALIASES.get(str(k).strip().lower())
            if canon and v not in (None, "") and canon not in norm:
                norm[canon] = str(v).strip()
        if norm.get("host"):
            out.append(norm)
    return out


def read_inventory_csv(text: str) -> List[Dict[str, Any]]:
    """Parse CSV text (a NetBox/CMDB export) -> normalized inventory rows. Tolerant; never raises."""
    try:
        reader = csv.DictReader(io.StringIO((text or "").lstrip("﻿")))   # strip a UTF-8 BOM (Excel 'CSV UTF-8')
        return normalize_rows(list(reader))
    except (csv.Error, ValueError):
        return []


def read_inventory_file(path: str) -> List[Dict[str, Any]]:
    """Read a declared inventory from a CSV or XLSX file -> normalized rows. Tolerant; never raises."""
    p = str(path or "")
    if p.lower().endswith((".xlsx", ".xlsm")):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True, data_only=True)
            grid = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
            wb.close()
            if not grid:
                return []
            header = [str(h or "").strip() for h in grid[0]]
            return normalize_rows([dict(zip(header, r)) for r in grid[1:]])
        except Exception:
            return []
    try:
        with open(p, encoding="utf-8-sig") as f:   # utf-8-sig also strips a BOM
            return read_inventory_csv(f.read())
    except OSError:
        return []


def _norm_host(h: Any) -> str:
    return str(h or "").strip().lower()


def observed_devices(snap: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """The collected estate, keyed by lower-cased host: {host, model, mgmt_ip, collected:bool}.

    Built from lifecycle_risk.per_device (model/platform) + health_scores (the full collected set); a host
    that the blind-spot list marks 'not collected' is flagged collected=False (its facts are UNVERIFIABLE)."""
    out: Dict[str, Dict[str, Any]] = {}
    for r in ((snap or {}).get("lifecycle_risk") or {}).get("per_device") or []:
        h = r.get("host")
        if not h:
            continue
        out[_norm_host(h)] = {"host": h, "model": r.get("model") or r.get("platform"),
                              "mgmt_ip": r.get("mgmt_ip"), "collected": True}
    for r in (snap or {}).get("health_scores") or []:
        h = r.get("switch")
        k = _norm_host(h)
        if k and k not in out:
            out[k] = {"host": h, "model": None, "mgmt_ip": None, "collected": True}
    for d in ((snap or {}).get("collection_completeness") or {}).get("devices") or []:
        if str(d.get("status", "")).strip().lower() in ("not collected", "partial"):   # partial = facts unverifiable
            k = _norm_host(d.get("host"))
            if k in out:
                out[k]["collected"] = False
            elif k:
                out[k] = {"host": d.get("host"), "model": None, "mgmt_ip": None, "collected": False}
    return out


_TYPES = ("MISSING_DEVICE", "UNDOCUMENTED_DEVICE", "MODEL_MISMATCH", "IP_DRIFT", "UNVERIFIABLE")


def reconcile_external(snap: Dict[str, Any], declared: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Diff a declared inventory against the collected evidence -> {rows:[…], summary:{…}}.

    `declared` is normalized rows (from `normalize_rows`/`read_inventory_csv`). A clean, fully-matching,
    collected device emits no row. Pure read; writes nothing to any device."""
    observed = observed_devices(snap)
    declared = normalize_rows(declared or [])                 # always normalize (idempotent) -> raw alias columns canonicalize
    deduped: Dict[str, dict] = {}                             # one reconciliation per host (first-seen wins)
    for d in declared:
        k = _norm_host(d.get("host"))
        if k and k not in deduped:
            deduped[k] = d
    declared = list(deduped.values())
    rows: List[dict] = []
    declared_keys = set()

    for d in declared:
        k = _norm_host(d.get("host"))
        if not k:
            continue
        declared_keys.add(k)
        obs = observed.get(k)
        if obs is None:
            rows.append({"type": "MISSING_DEVICE", "host": d.get("host"),
                         "detail": "declared in the source-of-truth but never observed in the collection",
                         "declared": d, "observed": None})
            continue
        if not obs.get("collected"):
            rows.append({"type": "UNVERIFIABLE", "host": d.get("host"),
                         "detail": "declared and in inventory, but never collected — cannot confirm or refute",
                         "declared": d, "observed": obs})
            continue
        dm, om = (d.get("model") or "").strip().lower(), (obs.get("model") or "").strip().lower()
        if dm and om and dm != om:
            rows.append({"type": "MODEL_MISMATCH", "host": d.get("host"),
                         "detail": "declared model %r != observed model %r" % (d.get("model"), obs.get("model")),
                         "declared": d, "observed": obs})
        di = (d.get("mgmt_ip") or "").strip().split("/")[0]   # NetBox stores CIDR; compare the host address only
        oi = (obs.get("mgmt_ip") or "").strip().split("/")[0]
        if di and oi and di != oi:
            rows.append({"type": "IP_DRIFT", "host": d.get("host"),
                         "detail": "declared mgmt IP %s != observed %s" % (di, oi),
                         "declared": d, "observed": obs})

    for k, obs in observed.items():
        if k not in declared_keys and obs.get("collected"):
            rows.append({"type": "UNDOCUMENTED_DEVICE", "host": obs.get("host"),
                         "detail": "observed in the collection but absent from the source-of-truth",
                         "declared": None, "observed": obs})

    rows.sort(key=lambda r: (_TYPES.index(r["type"]) if r["type"] in _TYPES else 9, str(r.get("host"))))
    summary = {t: sum(1 for r in rows if r["type"] == t) for t in _TYPES}
    summary.update(n_declared=len(declared_keys), n_observed=len(observed), n_rows=len(rows))
    return {"rows": rows, "summary": summary}
