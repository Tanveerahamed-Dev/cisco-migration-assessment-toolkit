"""Portable, decision-neutral projections of the single-snapshot protocol receipt.

Renderers in this module never mint source custody from a parsed snapshot.  They accept only the
receipt/export pair produced by the Protocol Assurance portfolio owner and validated by the shared
document projection.  Direct/legacy callers therefore receive an explicit ``NOT VERIFIED`` surface.
"""

from __future__ import annotations

from typing import Any

from cisco_toolkit.docmeta import protocol_assurance_receipt_view
from cisco_toolkit.excel import _xls_cell_value


RECEIPT_SCHEMA = "protocol_single_snapshot_receipt/1"
EXPORT_SCHEMA = "protocol_single_snapshot_export/1"
SHEET_NAME = "Protocol Assurance"
DEFAULT_EXPORT_REFERENCE = (
    "AssessHub /api/snapshots/{id}/protocol-assurance/export after persisted-byte import"
)


def protocol_assurance_surface_payload(
        bundle: Any, *, complete_export_reference: str = "") -> dict:
    """Return a closed renderer payload without re-hashing or reinterpreting owner evidence."""
    view = protocol_assurance_receipt_view(bundle)
    reference = complete_export_reference.strip() if isinstance(
        complete_export_reference, str) else ""
    if not view["valid"]:
        return {
            "schema": RECEIPT_SCHEMA,
            "receipt_valid": False,
            "status": "NOT VERIFIED",
            "reason": view["reason"],
            "owns_score": False,
            "owns_verdict": False,
            "custody_status": "not_verified",
            "source_binding": {},
            "receipt_sha256": "",
            "summary": {"n_families": 0, "n_subjects_total": 0},
            "families": [],
            "subject_cap": {"rendered": 0, "total": 0, "omitted": 0},
            "complete_export": {
                "schema": EXPORT_SCHEMA,
                "sha256": "",
                "media_type": "application/json",
                "reference": reference or DEFAULT_EXPORT_REFERENCE,
                "available": False,
            },
            "operator_note": (
                "This renderer did not hash the parsed snapshot or mint source custody. "
                "Import the exact snapshot into AssessHub to obtain the source-bound receipt and "
                "complete uncapped export."
            ),
        }

    export = dict(view["complete_export"])
    return {
        "schema": view["schema"],
        "owner_version": view["owner_version"],
        "receipt_valid": True,
        "status": str(view["custody_status"] or "not_verified").upper(),
        "reason": "",
        "owns_score": False,
        "owns_verdict": False,
        "custody_status": view["custody_status"],
        "custody_failures": list(view["custody_failures"]),
        "source_binding": dict(view["source_binding"]),
        "script_owner": dict(view["script_owner"]),
        "receipt_sha256": view["receipt_sha256"],
        "summary": dict(view["summary"]),
        "families": [dict(row) for row in view["families"]],
        "subject_cap": dict(view["subject_totals"]),
        "complete_export": {
            **export,
            "reference": (
                reference
                or str(export.get("default_reference") or "")
                or "local export reference not supplied"
            ),
            "available": True,
        },
        "operator_note": (
            "This portfolio owns no score and no verdict. Evidence status is a bounded coverage "
            "classification, not cutover authorization. Decision code does not consume capped "
            "presentation arrays."
        ),
    }


def _write_row(sheet, row: int, label: str, value: Any) -> None:
    sheet.cell(row=row, column=1, value=_xls_cell_value(label))
    sheet.cell(row=row, column=2, value=_xls_cell_value(value))


def write_protocol_assurance_receipt_sheet(
        workbook,
        bundle: Any = None,
        *,
        complete_export_reference: str = "") -> dict:
    """Add the assessment-workbook receipt sheet and return its validated surface payload."""
    from openpyxl.styles import Alignment, Font, PatternFill

    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    sheet = workbook.create_sheet(SHEET_NAME)
    payload = protocol_assurance_surface_payload(
        bundle,
        complete_export_reference=complete_export_reference,
    )

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A13"
    sheet.column_dimensions["A"].width = 31
    sheet.column_dimensions["B"].width = 84
    sheet.merge_cells("A1:F1")
    title = sheet["A1"]
    title.value = "Protocol Assurance · single-snapshot receipt"
    title.font = Font(bold=True, color="FFFFFF", size=14)
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24

    status = payload["status"]
    _write_row(sheet, 3, "Receipt status", status)
    _write_row(sheet, 4, "Receipt contract", payload["schema"])
    _write_row(sheet, 5, "Decision ownership", "No score · no verdict")
    _write_row(sheet, 6, "Custody", payload["custody_status"])
    if not payload["receipt_valid"]:
        _write_row(
            sheet,
            7,
            "Why NOT VERIFIED",
            f"{payload['reason']}. {payload['operator_note']}",
        )
        export = payload["complete_export"]
        _write_row(
            sheet,
            8,
            "Complete export",
            f"{export['schema']} · uncapped JSON · NOT ATTACHED · {export['reference']}",
        )
        _write_row(
            sheet,
            9,
            "Cap disclosure",
            "Rendered / total / omitted: 0 / 0 / 0 because no source-bound receipt was supplied.",
        )
        sheet["B3"].fill = PatternFill("solid", fgColor="FFF2CC")
        sheet["B3"].font = Font(bold=True, color="9C6500")
        return payload

    source = payload["source_binding"]
    cap = payload["subject_cap"]
    export = payload["complete_export"]
    if source["source"] == "persisted snapshots.snapshot_json blob":
        source_label = "Exact persisted source"
        source_value = (
            f"engagement {source['engagement_id']} · campaign {source['campaign_id']} · "
            f"snapshot {source['snapshot_id']} · {source['bytes']} bytes · {source['sha256']}"
        )
    else:
        source_label = "Exact emitted source"
        source_value = (
            f"local snapshot {source['label']} · {source['bytes']} bytes · {source['sha256']}"
        )
    _write_row(
        sheet,
        7,
        source_label,
        source_value,
    )
    _write_row(sheet, 8, "Receipt SHA-256", payload["receipt_sha256"])
    _write_row(
        sheet,
        9,
        "Receipt subject cap",
        f"Rendered / total / omitted: {cap['rendered']} / {cap['total']} / {cap['omitted']}",
    )
    _write_row(
        sheet,
        10,
        "Complete export",
        f"{export['schema']} · uncapped JSON · {export['sha256']} · {export['reference']}",
    )
    _write_row(sheet, 11, "Operator boundary", payload["operator_note"])
    sheet["B3"].fill = PatternFill(
        "solid", fgColor="E2F0D9" if status == "BOUND" else "FFF2CC")
    sheet["B3"].font = Font(
        bold=True, color="548235" if status == "BOUND" else "9C6500")

    headers = (
        "Protocol family",
        "Assurance level",
        "Evidence status",
        "Rendered",
        "Total",
        "Omitted",
    )
    for column, label in enumerate(headers, start=1):
        cell = sheet.cell(row=13, column=column, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
    for row_number, family in enumerate(payload["families"], start=14):
        values = (
            family["family"],
            family["assurance_level"],
            family["evidence_status"],
            family["rendered"],
            family["total"],
            family["omitted"],
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_number, column=column, value=_xls_cell_value(value))
    if payload["families"]:
        sheet.auto_filter.ref = f"A13:F{13 + len(payload['families'])}"
    sheet.column_dimensions["C"].width = 20
    for column in ("D", "E", "F"):
        sheet.column_dimensions[column].width = 12
    return payload


__all__ = [
    "DEFAULT_EXPORT_REFERENCE",
    "EXPORT_SCHEMA",
    "RECEIPT_SCHEMA",
    "SHEET_NAME",
    "protocol_assurance_surface_payload",
    "write_protocol_assurance_receipt_sheet",
]
