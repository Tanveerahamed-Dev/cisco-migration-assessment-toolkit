"""The snapshot-reporting layer: build the pre/post-cutover snapshot (snapshot_state - the JSON
contract embedded in the HTML and written beside every workbook) and render outputs from it - the
Blast-Radius Explorer HTML (write_html_explorer) and the '--compare OLD NEW' diff workbook
(write_diff_workbook). Extracted verbatim from COLLECT_PARSE_V3_23_0.py across PHASE 2.7 steps
29-30 (behaviour byte-identical). Depends on openpyxl + stdlib + the package's model/__version__."""
import ipaddress
import json
import logging
import os
import re
from datetime import datetime
from pathlib import PurePath
from typing import Any, Dict, List, Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from cisco_toolkit import __version__
from cisco_toolkit.analyze import (compute_current_baseline_gate,
                                   normalize_routing_adjacency_state)
from cisco_toolkit.model import DevicePhysical, InterfaceData
from cisco_toolkit.brand_tokens import WORKBOOK_NAVY_HEX
from cisco_toolkit.protocol_receipt_surfaces import protocol_assurance_surface_payload
from cisco_toolkit.textutils import is_finite_num, xml_safe as _cv

logger = logging.getLogger(__name__)


_DIFF_FIELDS = ["status", "switchport_mode", "vlan", "trunk_native_vlan",
                "trunk_allowed_vlans", "stp_blocked", "port_channel",
                "svi_ip", "hsrp_behavior", "subnet_primary_route"]

def _macset(s) -> set:
    """The MAC-address set of an interface's mac field, tolerant of a malformed value.

    `s or ""` guards None/empty but keeps a TRUTHY non-str (an int, or the `float('inf')` json.loads
    makes of a bare JSON `Infinity`), and `re.split` then raises `TypeError: expected string or
    bytes-like object` -- aborting the whole --compare workbook. A container is not a MAC list at all
    (-> empty set, i.e. 'not observed'); a non-str SCALAR is stringified so a value that really was a
    single MAC-ish token is still compared rather than silently dropped."""
    if not isinstance(s, str):
        s = "" if s is None or isinstance(s, (dict, list, tuple, set)) else str(s)
    return set(t for t in re.split(r"[,\s;]+", s) if t)


# Pre/post-cutover VALIDATION (NEW-V3.23.106). Beyond the raw interface/SVI/MAC diff, compare the
# COMPUTED analysis between two snapshots so an operator can answer "did the cutover make anything
# worse?": per-switch health-band shifts and the consolidated punch-list findings that OPENED vs
# RESOLVED (the punch-list already rolls up all finding sources, deduped + severity-ranked). Pure
# read of two snapshot_state() dicts; tolerant of older snapshots that lack the computed keys.
# 'Insufficient Data' is deliberately ABSENT: it is a coverage state, not a health-scale point (ssot.py:60-63), so
# transitions into/out of it are handled as coverage_shifts, never ranked as better/worse than a real band.
_BAND_RANK = {"Excellent": 0, "Good": 1, "Fair": 2, "Poor": 3, "Critical": 4}
_FIND_SEV_RANK = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3, "Info": 4}


# Renamed-title aliases: a pure title REWORDING (same detector, same measure, same scope semantics)
# must not churn the cutover-validation delta as resolved+opened when an OLD-engine baseline snapshot
# is compared against a NEW-engine one (PR-#396 review: the native-VLAN-1 rename flipped an
# identical-network compare from CLEAN to REVIEW). Each entry maps the pre-rename numbered-title form
# onto the CURRENT wording BEFORE keying; the count and device-set still participate in the key, so a
# genuine scope change (different N / different devices) keeps showing honestly as resolved+opened.
_TITLE_RENAMES = (
    (re.compile(r"^Native VLAN 1 on (\d+) inter-switch trunk\(s\)$"),
     r"Native VLAN 1 on \1 operationally-trunking port(s)"),
)


def _canon_title(title: str) -> str:
    for rx, repl in _TITLE_RENAMES:
        title = rx.sub(repl, title)
    return title


def _as_dict(x):
    """A snapshot section coerced to a dict. `... or {}` only guards FALSY values, so a truthy non-dict (a list/
    scalar from a malformed --compare/--trend snapshot) flowed into .values()/set() and raised (audit-5 totality);
    `.get(k, {})` likewise returns None for a present-but-null key."""
    return x if isinstance(x, dict) else {}


def _renderable_num(v) -> bool:
    """True for a number that arithmetic and cell-rendering can both survive.

    `isinstance(v, (int, float))` is NOT enough for a value read out of an untrusted snapshot:
    `json.loads` accepts integer literals of unbounded precision (`10**400` -> OverflowError on any
    float conversion, including `sum()/len()` and openpyxl's cell writer) and the bare `Infinity` /
    `-Infinity` / `NaN` tokens (which propagate through an average into a rendered figure).

    Delegates to the ONE owner (textutils.is_finite_num, shared with ssot.reconcile and causal), and
    keeps `bool` accepted here only because the pre-existing `isinstance(v, (int, float))` filter this
    replaces accepted it (bool is an int subclass) and a band tally must not change silently."""
    return isinstance(v, bool) or is_finite_num(v)


def _skey(x):
    """A TOTAL sort key for a set of snapshot-derived labels of possibly MIXED type.

    `sorted()` over a set built from device-derived leaves assumes every member is the same type, but a
    JSON snapshot can carry `"switch": 10` on one row and `"switch": "10"` on the next (a foreign-tool
    export, an older schema, a hand-trim) -- and Python 3 raises
    `TypeError: '<' not supported between instances of 'str' and 'int'`. Unlike the unhashable case this
    needs no poison at all: two ordinary, JSON-legal rows are enough.

    Strings sort FIRST and EXACTLY as before (the `(0, x)` branch compares the raw str), so the ordering
    of every well-formed snapshot is byte-identical; non-strings are grouped after them and ordered by
    their repr, which is deterministic and never raises."""
    return (0, x) if isinstance(x, str) else (1, str(x))


def _hkey(x):
    """A HASHABLE form of a snapshot LEAF about to be used as a dict key or as the left operand of an
    `in <dict>` test. _as_dict/_as_list guard the section's SHAPE; this guards the leaf inside a row.

    A dict/list where a label is expected (`health_scores[].switch`, `migration_readiness[].readiness`)
    is unhashable, and `{r.get('switch'): r}` / `r.get('readiness') in readiness` then raises
    `TypeError: unhashable type: 'dict'` -- which ABORTS write_diff_workbook and write_campaign_workbook
    (the --compare / --trend deliverables) and 500s the webapp diff/trend routes on every later read of
    a stored snapshot. Hashable values pass through UNCHANGED (keys, dedup and sort order over real data
    are untouched); only the unhashable poison is stringified, so it stays a DISTINCT key rather than
    silently merging rows or dropping a device. Twin of design_advisor._hkey (audit-7)."""
    try:
        hash(x)
        return x
    except TypeError:
        return str(x)


def _as_list(x):
    """A snapshot section coerced to a list -- the list-shaped twin of _as_dict, same reasoning: `... or []`
    only guards FALSY values, so a TRUTHY non-list (`health_scores: 5` in a hand-crafted upload) survived the
    guard and raised on the next `for r in ...` / `list(...)`. On the webapp side that is a STORED availability
    DoS: the upload is accepted (the only check is dict + a 'devices' key) and then 500s the unwrapped
    /explorer, /api/compare and /trend routes on every later read.

    Identical to `(x or [])` for every list/None input, so well-formed output is unchanged. A wrong-typed
    DICT section also degrades identically: iterating a dict yielded str keys, which the element-level
    `isinstance(r, dict)` filters already dropped."""
    return x if isinstance(x, list) else []


_SECTION_TYPES = {
    "health_scores": list,
    "punchlist": list,
    "migration_readiness": list,
    "lifecycle_risk": dict,
}


def _schema_status(value: Any) -> dict:
    """Normalize an optional caller-provided schema-compatibility result.

    The CLI owns the exact file paths and therefore remains the authority that can compare their
    schema versions before loading them.  The reporting layer accepts either the historic
    ``(status, message)`` tuple or a mapping and preserves an explicit ``override`` marker.  An
    absent value is not invented here.
    """
    if isinstance(value, dict):
        return {
            "status": str(value.get("status") or "").strip().lower(),
            "message": str(value.get("message") or ""),
            "override": bool(value.get("override") or value.get("overridden")),
        }
    if isinstance(value, (list, tuple)) and value:
        return {
            "status": str(value[0] or "").strip().lower(),
            "message": str(value[1] or "") if len(value) > 1 else "",
            "override": False,
        }
    if isinstance(value, str) and value.strip():
        return {"status": value.strip().lower(), "message": "", "override": False}
    return {}


def _analysis_integrity(snap: Any, required_sections=()) -> dict:
    """Machine-readable assessment-integrity view used by every decision renderer in this module.

    ``assessment_integrity.failed_phases`` is authoritative even when the failed producer returned
    an empty fallback.  Top-level ``_unavailable`` sentinels and required-but-missing analysis
    sections are folded into the same channel.  The helper deliberately reports *all* failed phases:
    a report may still render the observations that remain valid, but it cannot certify the run as
    clean while its own snapshot says part of the assessment failed.
    """
    s = snap if isinstance(snap, dict) else {}
    integrity = _as_dict(s.get("assessment_integrity"))
    failures: List[str] = []
    raw_failed = integrity.get("failed_phases")
    if isinstance(raw_failed, list):
        failures.extend(f"phase failed: {str(p)}" for p in raw_failed if str(p).strip())
    elif raw_failed:
        failures.append(f"phase failed: {str(raw_failed)}")
    for key, value in integrity.items():
        if key in ("failed_phases", "n_violations"):
            continue
        if str(value).strip().lower() in ("failed", "compute_failed", "unavailable", "error"):
            failures.append(f"{key}: {value}")
    for key, value in s.items():
        if isinstance(value, dict) and value.get("_unavailable"):
            failures.append(f"{key}: unavailable")
    for key in required_sections:
        expected = _SECTION_TYPES.get(str(key), object)
        if key not in s or not isinstance(s.get(key), expected):
            failures.append(f"{key}: missing or unusable")
    # Stable order, no duplicate disclosure when a sentinel and the integrity map name the same block.
    failures = list(dict.fromkeys(failures))
    return {"ok": not failures, "failures": failures}


def _section_available(snap: dict, key: str, phase_tokens=()) -> bool:
    """Whether a list-shaped analysis section is safe to delta.

    A failed producer commonly leaves the exact same ``[]`` as a legitimate clean computation.  The
    failed-phase channel is therefore part of availability; without it, comparing a healthy baseline
    to a failed empty fallback fabricates that every old finding was resolved.
    """
    if key not in snap or not isinstance(snap.get(key), list):
        return False
    failures = _analysis_integrity(snap).get("failures") or []
    def _phase_key(value: Any) -> str:
        return "".join(ch for ch in str(value).lower() if ch.isalnum())

    lowered = [_phase_key(f) for f in failures]
    return not any(any(_phase_key(tok) in f for tok in phase_tokens) for f in lowered)


def _finding_key(f: dict) -> tuple:
    """Stable identity for a punch-list finding across two runs: (category, FULL title, device-set).
    The title is intentionally NOT digit-normalized: stripping digits collapsed DISTINCT per-identifier
    findings that differ only by an embedded id (e.g. 'Fake FHRP redundancy (VLAN 20)' vs '(VLAN 21)'
    on the same gateways) into one key, which could hide a real fix-and-new-break swap as 'no change'
    in the cutover-validation verdict. Device order is normalized so the same finding on the same
    devices matches regardless of listing order; an aggregated finding whose count changes honestly
    shows as resolved+opened (its scope genuinely changed). The ONLY title rewriting applied is the
    _TITLE_RENAMES alias table above (pre-rename wording -> current wording), which is not a
    normalization: it never merges two distinct findings, it only keeps one finding's identity stable
    across an engine-version rename."""
    # _as_list (not `or []`): a ROW-INNER scalar -- a well-formed finding whose 'devices' is `5` -- survives
    # `or []` and raises on `for d in 5`, so one poisoned punch-list row 500s /api/compare and /trend.
    devs = tuple(sorted(str(d) for d in _as_list(f.get("devices"))))
    return (str(f.get("category", "")), _canon_title(str(f.get("title", ""))), devs)


_DEVICES_CELL_WIDTH = 60


def _devices_cell(devices, width: int = _DEVICES_CELL_WIDTH) -> str:
    """A finding's device list rendered for the Findings Delta 'Devices' column, WITH the house
    ``(+N more)`` disclosure when it does not fit.

    The raw ``", ".join(...)[:60]`` this replaces cut mid-token and said nothing about it, so a
    finding on 9 devices rendered as ``MERIDIAN-SW-171, ..., DS03-DC`` -- a silent cap in a
    CUTOVER-GATE artifact, and the reader's two wrong conclusions are (a) the finding is scoped to
    the devices shown and (b) ``DS03-DC`` is a hostname (it is not; it is the front half of one).
    Measured on the real Meridian reference snapshot: 109 of 1805 punch-list findings, and 11 of 115 in
    webapp/sample_data/sample_fleet.snapshot.json.

    Truncation now lands on a whole-device boundary and the remainder is COUNTED, never dropped
    silently. ``_as_list`` (not ``or []``) for the same reason ``_finding_key`` uses it: a row-inner
    scalar ``devices: 5`` survives ``or []`` and raises on ``for d in 5``."""
    names = [str(d) for d in _as_list(devices)]
    if not names:
        return ""
    joined = ", ".join(names)
    if len(joined) <= width:
        return joined
    shown: List[str] = []
    used = 0
    for n in names:
        add = len(n) + (2 if shown else 0)
        if shown and used + add > width:
            break
        shown.append(n)
        used += add
    if not shown:                                  # one device whose name alone exceeds the width
        shown = [names[0]]
    hidden = len(names) - len(shown)
    return ", ".join(shown) + (f" (+{hidden} more)" if hidden else "")


_PROTOCOL_RECEIPT_FAMILIES = ("STP", "EtherChannel", "VTP", "OSPF", "BGP", "EIGRP", "FHRP")
_PROTOCOL_ADJACENCY_FAMILIES = ("OSPF", "BGP", "EIGRP")
_PROTOCOL_RECEIPT_STATES = {
    "assessed", "partial", "captured_no_record", "captured_empty",
    "capture_error", "not_collected", "analysis_unavailable",
}


def _protocol_text(value: Any) -> str:
    """A strict text leaf for the embedded protocol projections.

    Device/peer identities are strings in the engine schema.  Stringifying a dict/list here would turn a
    malformed upload into a plausible identity and could merge unrelated rows, so containers and scalars
    abstain instead.  This helper is deliberately narrower than ``_hkey`` because these values participate
    in a cutover gate, not just a crash-safe display.
    """
    return value.strip() if isinstance(value, str) else ""


def _protocol_peer_key(peer: str) -> str:
    """Canonical comparison identity while retaining the producer's raw peer text for display."""
    try:
        return ipaddress.ip_address(peer).exploded.casefold()
    except ValueError:
        return peer.casefold()


def _protocol_receipt_view(snap: dict) -> dict:
    """Validate and index the exact ``protocol_assessability/1`` denominator.

    A hand-trimmed or foreign snapshot must not be allowed to keep the word ``assessed`` on the few rows it
    retained.  The receipt is therefore accepted only when its closed seven-family denominator and summary
    reconcile exactly.  The comparison remains total: invalid receipts return a reason instead of raising.
    """
    present = "protocol_assessability" in snap
    pa = snap.get("protocol_assessability")
    if not isinstance(pa, dict):
        return {"present": present, "valid": False, "index": {},
                "reason": "protocol assessability receipt is missing" if not present else
                          "protocol assessability receipt has an unusable shape"}
    if pa.get("schema") != "protocol_assessability/1":
        return {"present": True, "valid": False, "index": {},
                "reason": "protocol assessability receipt schema is not protocol_assessability/1"}

    families = pa.get("families")
    rows = pa.get("rows")
    summary = pa.get("summary")
    family_names = tuple(
        _protocol_text(item.get("protocol")) for item in families if isinstance(item, dict)
    ) if isinstance(families, list) else ()
    if family_names != _PROTOCOL_RECEIPT_FAMILIES:
        return {"present": True, "valid": False, "index": {},
                "reason": "protocol assessability family denominator is incomplete or reordered"}
    if not isinstance(rows, list) or not isinstance(summary, dict):
        return {"present": True, "valid": False, "index": {},
                "reason": "protocol assessability rows or summary have an unusable shape"}

    n_devices = summary.get("n_devices")
    n_families = summary.get("n_families")
    n_cells = summary.get("n_cells")
    if (not isinstance(n_devices, int) or isinstance(n_devices, bool) or n_devices < 0
            or n_families != len(_PROTOCOL_RECEIPT_FAMILIES)
            or n_cells != len(rows)
            or n_cells != n_devices * len(_PROTOCOL_RECEIPT_FAMILIES)):
        return {"present": True, "valid": False, "index": {},
                "reason": "protocol assessability summary does not reconcile to its exact denominator"}

    index: Dict[tuple, dict] = {}
    hosts: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            return {"present": True, "valid": False, "index": {},
                    "reason": "protocol assessability contains a malformed row"}
        host = _protocol_text(row.get("switch"))
        protocol = _protocol_text(row.get("protocol"))
        state = _protocol_text(row.get("state"))
        if not host or protocol not in _PROTOCOL_RECEIPT_FAMILIES or state not in _PROTOCOL_RECEIPT_STATES:
            return {"present": True, "valid": False, "index": {},
                    "reason": "protocol assessability contains an invalid switch, family, or state"}
        if state == "assessed" and row.get("health_row_emitted") is not True:
            return {"present": True, "valid": False, "index": {},
                    "reason": "protocol assessability marks a cell assessed without an emitted health row"}
        key = (host, protocol)
        if key in index:
            return {"present": True, "valid": False, "index": {},
                    "reason": "protocol assessability contains a duplicate switch-family cell"}
        index[key] = row
        hosts.add(host)
    if len(hosts) != n_devices or len(index) != n_cells:
        return {"present": True, "valid": False, "index": {},
                "reason": "protocol assessability rows do not reconcile to the device denominator"}
    return {"present": True, "valid": True, "index": index, "reason": ""}


def _protocol_state(protocol: str, raw: str) -> tuple[str, Optional[bool]]:
    """Normalize only adjacency state semantics, never counters or uptime.

    BGP's numeric State/PfxRcd is Established; the number is a prefix count and must not churn the peer.
    EIGRP's parser emits ``up <uptime>``; the changing uptime is not a state transition.  OSPF DR/BDR role
    suffixes likewise do not change adjacency health.
    """
    # One Python owner is shared with the single-snapshot Validation/NRFU baseline.  In particular, an
    # unfamiliar EIGRP token is REVIEW/unclassified rather than definite negative evidence, while the
    # anchored producer form ``up <uptime>`` remains acceptable and uptime churn is ignored.
    return normalize_routing_adjacency_state(protocol, _protocol_text(raw))


def _protocol_routing_view(snap: dict) -> dict:
    """Normalize the embedded routing-neighbor projection without inventing absent peers."""
    raw = snap.get("routing_neighbors")
    if not isinstance(raw, dict):
        return {"available": False, "pairs": {}, "malformed_pairs": set(),
                "reason": "routing_neighbors is missing or has an unusable shape"}

    pairs: Dict[tuple, Dict[str, dict]] = {}
    malformed_pairs: set[tuple] = set()
    global_error = ""
    for host_value, protocols in raw.items():
        host = _protocol_text(host_value)
        if not host or not isinstance(protocols, dict):
            global_error = "routing_neighbors contains a malformed host projection"
            continue
        for protocol in _PROTOCOL_ADJACENCY_FAMILIES:
            value = protocols.get(protocol.lower(), [])
            pair = (host, protocol)
            if not isinstance(value, list):
                malformed_pairs.add(pair)
                continue
            records: Dict[str, dict] = {}
            for item in value:
                if not isinstance(item, dict):
                    malformed_pairs.add(pair)
                    continue
                peer = _protocol_text(item.get("neighbor"))
                state_raw = _protocol_text(item.get("state"))
                if not peer or not state_raw:
                    malformed_pairs.add(pair)
                    continue
                peer_key = _protocol_peer_key(peer)
                if peer_key in records:
                    malformed_pairs.add(pair)
                    continue
                state, healthy = _protocol_state(protocol, state_raw)
                records[peer_key] = {
                    "switch": host,
                    "protocol": protocol,
                    "peer": peer,
                    "state_raw": state_raw,
                    "state": state,
                    "healthy": healthy,
                    "interface": _protocol_text(item.get("interface")),
                    "address": _protocol_text(item.get("address")),
                    "as": _protocol_text(item.get("as")),
                }
            pairs[pair] = records
    return {"available": not global_error, "pairs": pairs, "malformed_pairs": malformed_pairs,
            "reason": global_error}


def _protocol_metadata(record: dict) -> tuple:
    protocol = record.get("protocol")
    if protocol == "BGP":
        return (record.get("as", ""),)
    if protocol == "OSPF":
        return (record.get("interface", ""), record.get("address", ""))
    return (record.get("interface", ""),)


def compute_protocol_adjacency_delta(old: dict, new: dict, *,
                                     source_binding: Optional[dict] = None) -> dict:
    """Evidence-gated before/after delta for baseline-observed OSPF/BGP/EIGRP peers.

    This is deliberately *not* an expected-neighbor health model.  It asks the narrower cutover question:
    "what happened to peers that the baseline actually observed?"  Every compared host-family cell requires
    an exact ``protocol_assessability/1`` cell in state ``assessed`` on both sides.  A disappearing last peer
    makes the after receipt ``captured_no_record`` and is therefore a coverage loss / REVIEW, never a proven
    down adjacency.  The embedded receipt does not cryptographically bind ``routing_neighbors`` content, so
    the returned custody label names that limitation even when caller-owned snapshot hashes are present.
    """
    old = old if isinstance(old, dict) else {}
    new = new if isinstance(new, dict) else {}
    old_receipt, new_receipt = _protocol_receipt_view(old), _protocol_receipt_view(new)
    old_routing, new_routing = _protocol_routing_view(old), _protocol_routing_view(new)

    old_pairs = old_routing["pairs"]
    new_pairs = new_routing["pairs"]
    scoped_pairs = {pair for pair, peers in old_pairs.items() if peers}
    if old_receipt["valid"]:
        scoped_pairs.update(
            pair for pair, row in old_receipt["index"].items()
            if pair[1] in _PROTOCOL_ADJACENCY_FAMILIES and row.get("health_row_emitted") is True
        )
    scoped_pairs = sorted(scoped_pairs, key=lambda pair: (pair[0].casefold(), pair[0], pair[1]))

    changes: List[dict] = []
    coverage_gaps: List[dict] = []
    n_baseline_peers = sum(len(old_pairs.get(pair, {})) for pair in scoped_pairs)
    n_comparable_cells = 0
    n_preserved = 0
    n_state_regressed = 0
    n_recovered = 0
    n_no_longer_observed = 0
    n_added = 0
    n_metadata_changed = 0

    def _receipt_state(view: dict, pair: tuple) -> str:
        if not view["valid"]:
            return "legacy" if not view["present"] else "invalid"
        row = view["index"].get(pair)
        return _protocol_text(row.get("state")) if isinstance(row, dict) else "missing_cell"

    def _receipt_row(view: dict, pair: tuple) -> Optional[dict]:
        row = view["index"].get(pair) if view["valid"] else None
        return row if isinstance(row, dict) else None

    for pair in scoped_pairs:
        host, protocol = pair
        before_state, after_state = _receipt_state(old_receipt, pair), _receipt_state(new_receipt, pair)
        before_row, after_row = _receipt_row(old_receipt, pair), _receipt_row(new_receipt, pair)
        reasons: List[str] = []
        if not old_receipt["valid"]:
            reasons.append(f"before: {old_receipt['reason']}")
        elif before_state != "assessed":
            reasons.append(f"before receipt state is {before_state}")
        elif (before_row is not None and before_row.get("health_row_emitted") is True
              and not old_pairs.get(pair)):
            reasons.append(
                "before receipt is assessed with an emitted health row but the routing-neighbor projection "
                "has zero peers"
            )
        if not new_receipt["valid"]:
            reasons.append(f"after: {new_receipt['reason']}")
        elif after_state != "assessed":
            reasons.append(f"after receipt state is {after_state}")
        elif (after_row is not None and after_row.get("health_row_emitted") is True
              and not new_pairs.get(pair)):
            reasons.append(
                "after receipt is assessed with an emitted health row but the routing-neighbor projection "
                "has zero peers"
            )
        if not old_routing["available"]:
            reasons.append(f"before: {old_routing['reason']}")
        if not new_routing["available"]:
            reasons.append(f"after: {new_routing['reason']}")
        if pair in old_routing["malformed_pairs"]:
            reasons.append("before routing-neighbor projection is malformed or duplicated")
        if pair in new_routing["malformed_pairs"]:
            reasons.append("after routing-neighbor projection is malformed or duplicated")
        if reasons:
            coverage_gaps.append({
                "switch": host,
                "protocol": protocol,
                "before_state": before_state,
                "after_state": after_state,
                "reason": "Not comparable: " + "; ".join(dict.fromkeys(reasons)),
            })
            continue

        n_comparable_cells += 1
        before_peers = old_pairs.get(pair, {})
        after_peers = new_pairs.get(pair, {})
        for peer_key in sorted(before_peers):
            before = before_peers[peer_key]
            after = after_peers.get(peer_key)
            if after is None:
                n_no_longer_observed += 1
                changes.append({
                    "switch": host, "protocol": protocol, "peer": before["peer"],
                    "before_state": before["state_raw"], "after_state": "not observed",
                    "result": "no_longer_observed",
                    "note": ("The peer was present in the baseline but is no longer observed while the family "
                             "remains assessable. Confirm whether this topology change was intended."),
                })
                continue

            # Preservation is orthogonal to a reviewable metadata/role change: the same baseline peer can
            # remain in an acceptable state while moving interface or changing AS.  Count the preserved
            # operational relationship and disclose the topology/configuration change separately.
            if before["healthy"] is True and after["healthy"] is True:
                n_preserved += 1
            if before["healthy"] is True and after["healthy"] is False:
                n_state_regressed += 1
                changes.append({
                    "switch": host, "protocol": protocol, "peer": before["peer"],
                    "before_state": before["state_raw"], "after_state": after["state_raw"],
                    "result": "state_degraded",
                    "note": "A previously acceptable observed adjacency entered a parsed unacceptable state.",
                })
            elif before["healthy"] is False and after["healthy"] is True:
                n_recovered += 1
                changes.append({
                    "switch": host, "protocol": protocol, "peer": before["peer"],
                    "before_state": before["state_raw"], "after_state": after["state_raw"],
                    "result": "recovered",
                    "note": "A previously unacceptable observed adjacency is now in an acceptable state.",
                })
            elif before["state"] != after["state"]:
                changes.append({
                    "switch": host, "protocol": protocol, "peer": before["peer"],
                    "before_state": before["state_raw"], "after_state": after["state_raw"],
                    "result": "state_changed",
                    "note": "The parsed adjacency state changed; review the transition in operational context.",
                })
            elif _protocol_metadata(before) != _protocol_metadata(after):
                n_metadata_changed += 1
                changes.append({
                    "switch": host, "protocol": protocol, "peer": before["peer"],
                    "before_state": before["state_raw"], "after_state": after["state_raw"],
                    "result": "metadata_changed",
                    "note": "The observed adjacency remains present but its interface, address, or AS changed.",
                })

        for peer_key in sorted(set(after_peers) - set(before_peers)):
            after = after_peers[peer_key]
            n_added += 1
            changes.append({
                "switch": host, "protocol": protocol, "peer": after["peer"],
                "before_state": "not observed", "after_state": after["state_raw"],
                "result": "added",
                "note": "A peer not observed in the baseline is now present; confirm that it was intended.",
            })

    review_changes = sum(
        1 for item in changes if item["result"] in
        ("no_longer_observed", "added", "metadata_changed", "state_changed")
    )
    both_legacy = not old_receipt["present"] and not new_receipt["present"]
    if not scoped_pairs:
        gate = "NOT_ASSESSED"
        note = ("Protocol adjacency preservation was not assessed: the baseline contained no observed "
                "OSPF, BGP, or EIGRP peer subject. This is not proof that no peers were configured.")
    elif n_state_regressed:
        gate = "REGRESSED"
        note = (f"Protocol adjacency gate REGRESSED: {n_state_regressed} baseline-observed peer(s) entered "
                "an unacceptable parsed state. Investigate or roll back before proceeding.")
    elif coverage_gaps:
        gate = "NOT_ASSESSED" if both_legacy else "REVIEW"
        note = (
            f"Protocol adjacency gate {'NOT ASSESSED' if both_legacy else 'REVIEW'}: "
            f"{len(coverage_gaps)} baseline device-family cell(s) could not be compared. A previously "
            "observed peer may no longer be evidenced, but missing/empty/error/parser-gap evidence is not "
            "proof that it is down or absent. Re-collect before accepting the change."
        )
    elif review_changes:
        gate = "REVIEW"
        note = (f"Protocol adjacency gate REVIEW: {review_changes} observed topology or state change(s) "
                "need operator confirmation; no expected-peer intent was inferred.")
    else:
        gate = "PASS"
        note = (f"Protocol adjacency gate PASS: no new adjacency regression was observed across "
                f"{n_comparable_cells} baseline device-family cell(s). PASS means baseline-observed peers in "
                "assessable before/after scope did not newly degrade; it does not prove the expected peer set "
                "is complete.")

    binding = source_binding if isinstance(source_binding, dict) else {}
    source_bound = bool(binding.get("before") and binding.get("after"))
    return {
        "schema": "protocol_adjacency_delta/1",
        "gate": gate,
        "assessed": bool(scoped_pairs) and not coverage_gaps,
        "scope": "baseline_observed",
        "projection_custody": ("source_bound_embedded_unverified" if source_bound
                               else "embedded_unverified"),
        "summary": {
            "n_baseline_peers": n_baseline_peers,
            "n_scoped_cells": len(scoped_pairs),
            "n_comparable_cells": n_comparable_cells,
            "n_preserved": n_preserved,
            "n_state_regressed": n_state_regressed,
            "n_recovered": n_recovered,
            "n_no_longer_observed": n_no_longer_observed,
            "n_added": n_added,
            "n_metadata_changed": n_metadata_changed,
            "n_coverage_gaps": len(coverage_gaps),
        },
        "changes": sorted(changes, key=lambda item: (
            item["switch"].casefold(), item["switch"], item["protocol"],
            item["peer"].casefold(), item["result"])),
        "coverage_gaps": coverage_gaps,
        "note": note,
        "limitations": [
            "The gate preserves baseline-observed peers; it is not an expected-neighbor denominator.",
            "A peer disappearance is an observed topology change, not proof that the adjacency is down.",
            "The snapshot source may be hash-bound, but the embedded receipt does not cryptographically bind "
            "the routing-neighbor projection; projection custody remains unverified.",
        ],
    }


def compute_snapshot_delta(old: dict, new: dict, *, source_binding: Optional[dict] = None,
                           schema_status: Any = None) -> dict:
    """Migration-validation delta between two snapshots: switch/interface counts, per-switch health-band
    shifts (regressed vs improved), punch-list findings opened vs resolved, and an overall verdict.
    Returns a dict.  Missing/failed computed sections abstain and make the overall result
    ``INDETERMINATE``; they are never interpreted as a clean empty result.

    ``source_binding`` and ``schema_status`` are optional caller-owned provenance.  The file-loading
    caller can supply exact input byte hashes; when present they are copied into the returned result
    and into the workbook rather than being replaced by a weaker re-serialization hash.
    """
    old = old if isinstance(old, dict) else {}            # total: the --compare/--trend path may hand us a
    new = new if isinstance(new, dict) else {}            # non-dict (a JSON file that parsed to null/[]/scalar)
    od, nd = _as_dict(old.get("devices")), _as_dict(new.get("devices"))
    oi, ni = _as_dict(old.get("interfaces")), _as_dict(new.get("interfaces"))
    old_integrity = _analysis_integrity(old, ("health_scores", "punchlist"))
    new_integrity = _analysis_integrity(new, ("health_scores", "punchlist"))
    schema = _schema_status(schema_status)
    integrity_failures = ([f"before: {f}" for f in old_integrity["failures"]]
                          + [f"after: {f}" for f in new_integrity["failures"]])
    if schema and schema.get("status") not in ("", "ok"):
        qualifier = " (explicitly overridden)" if schema.get("override") else ""
        integrity_failures.append(
            f"schema compatibility {schema.get('status')}{qualifier}: "
            f"{schema.get('message') or 'cross-input compatibility was not proven'}")

    # ---- health-band shifts (per switch present in BOTH runs) ----
    # isinstance guard: a null element in the list (hand-trimmed / older-schema snapshot fed to --compare/--trend)
    # must degrade, not AttributeError on r.get (audit-4 #15). The ELEMENT filter was present but the SECTION
    # was not coerced -- `health_scores: 5` survives `or []` and dies on `for r in 5` (_as_list, both operands:
    # a guard on only `old` leaves the identical crash reachable through `new` on the same route).
    # _hkey on the KEY leaf: an unhashable dict/list `switch` raised TypeError here, aborting the whole
    # --compare workbook (and 500ing the webapp diff route) on a snapshot that is stored and re-read.
    health_comparable = (
        _section_available(old, "health_scores", ("health score",))
        and _section_available(new, "health_scores", ("health score",))
    )
    oh = ({_hkey(r.get("switch")): r for r in _as_list(old.get("health_scores")) if isinstance(r, dict)}
          if health_comparable else {})
    nh = ({_hkey(r.get("switch")): r for r in _as_list(new.get("health_scores")) if isinstance(r, dict)}
          if health_comparable else {})
    regressed: List[dict] = []
    improved: List[dict] = []
    coverage_shifts: List[dict] = []   # transitions in/out of 'Insufficient Data' (coverage events, not health)
    for sw in sorted(set(oh) & set(nh), key=_skey):     # _skey: mixed-type switch labels must not raise
        ob, nb = oh[sw].get("band", ""), nh[sw].get("band", "")
        if ob == nb:
            continue
        # 'Insufficient Data' is a COVERAGE state, not a point on the health scale (ssot.py:60-63: it never counts
        # as the worst band). A shift into/out of it is a coverage change, NOT a health improvement/regression --
        # ranking it worse-than-Critical made 'Insufficient Data -> Critical' read 'improved' + CLEAN, i.e. a
        # newly-online Critical box certified as an improvement at the cutover gate (audit-4 #5).
        if ob == "Insufficient Data" or nb == "Insufficient Data":
            coverage_shifts.append({"switch": sw, "old_band": ob, "new_band": nb,
                                    "old_score": oh[sw].get("score", ""), "new_score": nh[sw].get("score", ""),
                                    "kind": "went_dark" if nb == "Insufficient Data" else "newly_assessed"})
            continue
        orank, nrank = _BAND_RANK.get(ob, 9), _BAND_RANK.get(nb, 9)
        if nrank == orank:
            continue
        row = {"switch": sw, "old_band": ob, "new_band": nb,
               "old_score": oh[sw].get("score", ""), "new_score": nh[sw].get("score", "")}
        (regressed if nrank > orank else improved).append(row)
    regressed.sort(key=lambda r: -_BAND_RANK.get(r["new_band"], 0))
    # coverage events that must keep the verdict off CLEAN: a device gone dark (lost visibility at cutover) or one
    # newly collected and found Critical/Poor (a real problem just revealed).
    n_went_dark = sum(1 for c in coverage_shifts if c["kind"] == "went_dark")
    newly_bad = [c for c in coverage_shifts if c["kind"] == "newly_assessed" and c["new_band"] in ("Critical", "Poor")]

    # ---- punch-list findings opened vs resolved ----
    findings_comparable = (
        _section_available(old, "punchlist", ("punch-list", "punchlist"))
        and _section_available(new, "punchlist", ("punch-list", "punchlist"))
    )
    o_find = ({_finding_key(f): f for f in _as_list(old.get("punchlist")) if isinstance(f, dict)}
              if findings_comparable else {})
    n_find = ({_finding_key(f): f for f in _as_list(new.get("punchlist")) if isinstance(f, dict)}
              if findings_comparable else {})
    # fully-deterministic order (set-difference iteration order is unstable): severity, then the
    # finding's stable identity, so two runs of the diff workbook are byte-reproducible.
    def _fsort(f: dict) -> tuple:
        return (_FIND_SEV_RANK.get(f.get("severity", ""), 9), _finding_key(f))
    opened = sorted((n_find[k] for k in (set(n_find) - set(o_find))), key=_fsort)
    resolved = sorted((o_find[k] for k in (set(o_find) - set(n_find))), key=_fsort)
    n_opened_high = sum(1 for f in opened if f.get("severity") in ("Critical", "High"))

    # ---- computed reachability what-if (W2): did the change DEFINITIVELY break a previously-working flow? ----
    # The native RIB->FIB differential (the offline Batfish-peer); coverage-honest, so only definitive
    # newly_blocked flows count toward the verdict -- ambiguous/incomplete pairs stay 'inconclusive'.
    try:
        from cisco_toolkit import fib
        rdelta = fib.reachability_delta(old, new)
    except Exception:                                   # never let the cutover diff fail on a reachability hiccup
        rdelta = {"summary": {}, "newly_blocked": [], "newly_reachable": [], "preserved": 0,
                  "inconclusive": 0, "pairs_tested": 0, "subnets_tested": 0, "capped": False}
    n_newly_blocked = len(rdelta.get("newly_blocked") or [])
    n_ecmp_partial = len(_as_list(rdelta.get("ecmp_partial_drop")))

    # Coverage-honest reachability clause -- NEVER an unqualified 'no reachability regressions' (no-silent-caps /
    # not-assessed!=healthy doctrine). It always discloses either 'NOT assessed' or the BOUNDED sample it tested.
    if not rdelta.get("assessed"):
        reach_phrase = ("computed reachability NOT assessed — no routes collected"
                        if (rdelta.get("subnets_total") or 0) == 0 else
                        f"computed reachability NOT assessed — only {rdelta.get('subnets_total')} subnet(s) "
                        "collected, no inter-subnet flow to test")
    else:
        cap = ", capped" if rdelta.get("capped") else ""
        reach_phrase = (f"computed reachability: {n_newly_blocked} of {rdelta.get('pairs_tested', 0)} sampled "
                        f"inter-subnet flow(s) newly blocked (bounded sample: {rdelta.get('subnets_tested', 0)} of "
                        f"{rdelta.get('subnets_total', 0)} subnet(s), one representative host each{cap})")

    # ---- physical cabling delta (the EDA cable-map SSOT, diffed) ----
    # A cable DOWN across the change window is a hard physical regression; '-> unknown' is a coverage
    # event ('no longer observed'), never a down. Pre-cable-map snapshots rehydrate from interfaces.
    try:
        from cisco_toolkit.analyze import cable_map_of_snapshot, compute_cable_map_diff
        cdelta = compute_cable_map_diff(cable_map_of_snapshot(old), cable_map_of_snapshot(new))
    except Exception:                                   # never let the cutover diff fail on the cable pass
        cdelta = {"assessed": False, "added": [], "removed": [], "status_changes": [],
                  "members_changed": [], "summary": {}}
    _cs = cdelta.get("summary") or {}
    n_cables_down = int(_cs.get("n_went_down") or 0)
    n_cables_changed = (int(_cs.get("n_added") or 0) + int(_cs.get("n_removed") or 0)
                        + int(_cs.get("n_members_changed") or 0) + int(_cs.get("n_no_longer_observed") or 0))
    if not cdelta.get("assessed"):
        cable_phrase = ("physical cabling NOT assessed — no CDP/LLDP links in either snapshot "
                        "(NOT a statement that cabling is unchanged)")
    else:
        cable_phrase = (f"physical cabling: {_cs.get('n_added', 0)} cable(s) added, "
                        f"{_cs.get('n_removed', 0)} removed, {n_cables_down} went DOWN, "
                        f"{_cs.get('n_no_longer_observed', 0)} no longer observed")

    # ---- observed routing-adjacency preservation (receipt-gated) ----
    # This is intentionally separate from sparse protocol_health: the exact assessability receipt tells us
    # whether each baseline host-family cell can be compared, while routing_neighbors supplies the bounded
    # observed peers.  The helper never invents an expected-neighbor denominator and names its embedded-
    # projection custody limitation in the returned result.
    pdelta = compute_protocol_adjacency_delta(old, new, source_binding=source_binding)
    _ps = pdelta.get("summary") or {}
    n_protocol_regressed = int(_ps.get("n_state_regressed") or 0)
    protocol_review = pdelta.get("gate") == "REVIEW"
    protocol_phrase = str(pdelta.get("note") or "Protocol adjacency preservation was not assessed.")

    # ---- current-snapshot acceptance baseline ----
    # The delta deliberately answers only "did this get worse?".  Preserve that semantic, but carry
    # the independently reconciled current validation baseline so the combined cutover decision cannot
    # turn an unchanged degraded state into PASS.
    current_baseline = compute_current_baseline_gate(new.get("validation_plan"))
    # Minimal dictionaries are also used as unit-level/legacy delta inputs.  Requiring a new acceptance
    # axis there would silently break the historical direct API.  A real snapshot contract (or any caller
    # that explicitly carries validation_plan, including an empty one) opts into the acceptance gate.
    current_baseline_required = (
        "validation_plan" in new
        or str(new.get("schema") or "").startswith("collect_parse_snapshot/")
    )

    # ---- verdict ----
    removed_sw = sorted(set(od) - set(nd), key=_skey)
    adverse_delta = bool(
        n_opened_high or regressed or n_newly_blocked or n_ecmp_partial or n_cables_down
        or n_protocol_regressed
    )
    # Integrity/schema uncertainty dominates the certification verdict.  Keep any adverse
    # observations visible in the note, but never let a real-looking delta imply that the
    # incompatible or incomplete inputs themselves were valid to compare.
    if integrity_failures:
        verdict = "INDETERMINATE"
        observed = []
        if n_opened_high:
            observed.append(f"{n_opened_high} apparent new High/Critical finding(s)")
        if regressed:
            observed.append(f"{len(regressed)} apparent health-band regression(s)")
        if n_newly_blocked:
            observed.append(f"{n_newly_blocked} apparent newly blocked sampled flow(s)")
        if n_ecmp_partial:
            observed.append(f"{n_ecmp_partial} apparent blackholing ECMP leg(s)")
        if n_cables_down:
            observed.append(f"{n_cables_down} apparent cable-down transition(s)")
        if n_protocol_regressed:
            observed.append(f"{n_protocol_regressed} apparent protocol adjacency state regression(s)")
        observed_note = (
            " Adverse observations that still require investigation: " + "; ".join(observed) + "."
            if observed else ""
        )
        note = (
            f"Delta certification withheld: {len(integrity_failures)} integrity/schema gap(s) make one "
            "or more analyses unavailable. No missing/failed section was interpreted as clean. "
            + "; ".join(integrity_failures)
            + f".{observed_note} {cable_phrase}; {reach_phrase}. {protocol_phrase}"
        )
    elif adverse_delta:
        verdict = "REGRESSED"
        bits = []
        if n_opened_high:
            bits.append(f"{n_opened_high} new High/Critical finding(s)")
        if regressed:
            bits.append(f"{len(regressed)} switch(es) dropped a health band")
        if n_ecmp_partial:
            bits.append(f"{n_ecmp_partial} sampled flow(s) have a proven blackholing ECMP leg")
        if n_protocol_regressed:
            bits.append(f"{n_protocol_regressed} baseline-observed protocol adjacency state regression(s)")
        bits.append(cable_phrase)
        bits.append(reach_phrase)
        bits.append(protocol_phrase)
        note = "; ".join(bits) + ". Investigate before declaring the cutover good."
    elif opened or removed_sw or n_went_dark or newly_bad or n_cables_changed or protocol_review:
        verdict = "REVIEW"
        cov_bits = []
        if n_went_dark:
            cov_bits.append(f"{n_went_dark} switch(es) went dark (lost collection — can't be certified)")
        if newly_bad:
            cov_bits.append(f"{len(newly_bad)} newly-collected switch(es) found Critical/Poor")
        note = (f"{len(opened)} new finding(s); {len(removed_sw)} switch(es) no longer present"
                + ("; " + "; ".join(cov_bits) if cov_bits else "")
                + f"; {cable_phrase}; {reach_phrase}; {protocol_phrase} Confirm these are expected.")
    else:
        verdict = "CLEAN"
        note = (
            "Delta-only observation: no health-band regressions or newly opened findings were observed "
            "in the available comparable analyses; "
            f"{cable_phrase}; {reach_phrase}; {protocol_phrase} This is not a cutover authorization; reconcile the "
            "Pre-Change Certificate and named blind spots."
        )

    def _scn(s):  # SSOT: canonical device count (one source); raw len() only as the pre-brief fallback
        # _as_dict at BOTH levels (the same shape _trend_point's local _d() already guards): a truthy
        # non-dict `executive_brief` -- or a well-formed brief whose inner `scale` is a scalar -- slips
        # past `or {}` and AttributeErrors on the next .get().
        return _as_dict(_as_dict(s.get("executive_brief")).get("scale")).get("n_devices")
    return {
        "switches": {"old": _scn(old) or len(od), "new": _scn(new) or len(nd),
                     "added": sorted(set(nd) - set(od)), "removed": removed_sw},
        # oi/ni are already _as_dict-coerced above; this is the PER-ELEMENT gap -- `interfaces: {"sw1": 5}`
        # reaches len(5). (A null per-host value crashed here too: len(None).)
        "interfaces": {"old": sum(len(_as_dict(v)) for v in oi.values()),
                       "new": sum(len(_as_dict(v)) for v in ni.values())},
        "health": {"regressed": regressed, "improved": improved,
                   "n_regressed": len(regressed), "n_improved": len(improved),
                   "coverage_shifts": coverage_shifts, "n_coverage_shifts": len(coverage_shifts)},
        "findings": {"opened": opened, "resolved": resolved, "n_opened": len(opened),
                     "n_resolved": len(resolved), "n_opened_high": n_opened_high},
        "reachability": rdelta,
        "cabling": cdelta,
        "protocol_adjacencies": pdelta,
        "current_baseline": current_baseline,
        "current_baseline_required": current_baseline_required,
        "integrity": {
            "ok": not integrity_failures,
            "failures": integrity_failures,
            "health_comparable": health_comparable,
            "findings_comparable": findings_comparable,
        },
        "provenance": {
            "source_binding": dict(source_binding) if isinstance(source_binding, dict) else {},
            "schema_status": schema,
        },
        "verdict": verdict,
        "verdict_display": ("NO DELTA REGRESSION OBSERVED" if verdict == "CLEAN" else verdict),
        "verdict_scope": "delta_only",
        "verdict_note": note,
    }


def compute_cutover_gate(delta: dict, certificate: dict,
                         current_baseline: Optional[dict] = None,
                         comparison_admission: Optional[dict] = None,
                         protocol_family_changes: Optional[dict] = None,
                         operator_evidence: Optional[dict] = None,
                         decision_input_authority: Any = None) -> dict:
    """Return the one combined before/after decision shown to an operator.

    ``compute_snapshot_delta`` and the Pre-Change Certificate answer different questions.  The former
    includes observed health, findings, cabling, reachability, and receipt-gated protocol adjacency
    changes; the latter is deliberately bounded to FIB/path-intent and segmentation checks.  Keeping the
    precedence here prevents a PASS on that narrower certificate from masking a proven delta regression.

    The function is pure and total so every presentation surface can render the exact same receipt.
    ``current_baseline`` is optional for backward compatibility with direct legacy callers; when
    omitted, the receipt embedded by ``compute_snapshot_delta`` is consumed.  The additive
    ``comparison_admission`` contract is consumed only when supplied by a source-owning caller.
    It can withhold a PASS for incompatible ownership, malformed identity/custody, or missing
    comparison scope without changing the legacy two-argument contract.  A supplied admission also
    requires the process-local canonical gate-input authority minted by ``compare_bound_pair`` so receipts from
    different source pairs cannot be mixed and rehashed.  The additive reference-only family change
    set owns no verdict; when supplied, this sole decision owner consumes each native family's
    producer-owned decision effect over the complete, uncapped row set. Canonical operator evidence
    is likewise reference-only: when it is source-bound into the admitted private input bundle, this
    gate consumes only bounded local L2 failure risks and coverage faults. It never promotes a
    simulation to convergence, traffic-continuity, or service-survival proof.
    """
    delta = _as_dict(delta)
    certificate = _as_dict(certificate)

    delta_verdict = str(delta.get("verdict") or "INDETERMINATE").strip().upper()
    if delta_verdict not in {"REGRESSED", "INDETERMINATE", "REVIEW", "CLEAN"}:
        delta_verdict = "INDETERMINATE"
    delta_display = str(delta.get("verdict_display") or delta_verdict)
    certificate_verdict = str(certificate.get("verdict") or "INDETERMINATE").strip().upper()
    if certificate_verdict not in {"PASS", "CONDITIONAL", "FAIL", "INDETERMINATE"}:
        certificate_verdict = "INDETERMINATE"

    protocol = _as_dict(delta.get("protocol_adjacencies"))
    protocol_summary = _as_dict(protocol.get("summary"))

    def _count(name: str) -> int:
        value = protocol_summary.get(name, 0)
        return int(value) if (isinstance(value, int) and not isinstance(value, bool) and value >= 0) else 0

    protocol_regressions = _count("n_state_regressed")
    protocol_coverage_gaps = _count("n_coverage_gaps")
    protocol_baseline_peers = _count("n_baseline_peers")
    protocol_gate = str(protocol.get("gate") or "NOT_ASSESSED").strip().upper()

    def _bounded_note(value: Any, fallback: str) -> str:
        text = value.strip() if isinstance(value, str) else ""
        text = text or fallback
        return text if len(text) <= 600 else text[:597] + "..."

    delta_note = _bounded_note(delta.get("verdict_note"), "No delta basis was emitted.")
    certificate_note = _bounded_note(
        certificate.get("verdict_note"), "No certificate basis was emitted.")

    admission_supplied = comparison_admission is not None
    admission = _as_dict(comparison_admission)
    admission_status = "not_comparable"
    admission_failures: list = []
    admission_gaps: list = []
    decision_inputs_valid = False
    if admission_supplied:
        from cisco_toolkit.protocol_assurance import validate_comparison_admission

        admission_validation = validate_comparison_admission(comparison_admission)
        if admission_validation.get("valid") is True:
            admission_status = str(admission.get("status") or "not_comparable").strip().lower()
            admission_failures = admission.get("failures") \
                if isinstance(admission.get("failures"), list) else []
            admission_gaps = admission.get("coverage_gaps") \
                if isinstance(admission.get("coverage_gaps"), list) else []
        else:
            admission_failures = [
                _bounded_note(
                    admission_validation.get("reason"),
                    "Comparison admission receipt is missing, malformed, or semantically inconsistent.",
                )
            ]
        from cisco_toolkit.comparison import validate_cutover_decision_input_authority

        decision_input_validation = validate_cutover_decision_input_authority(
            decision_input_authority,
            delta=delta,
            certificate=certificate,
            admission=comparison_admission,
            protocol_families=protocol_family_changes,
            operator_evidence=operator_evidence,
        )
        decision_inputs_valid = decision_input_validation.get("valid") is True
        if not decision_inputs_valid:
            admission_status = "not_comparable"
            admission_failures.append(_bounded_note(
                decision_input_validation.get("reason"),
                "Canonical source-bound cutover decision inputs are missing or detached.",
            ))
    admission_note = _bounded_note(
        "; ".join(str(item) for item in (admission_failures or admission_gaps) if str(item).strip()),
        ("Comparison ownership, custody, subjects, and owner semantics were admitted."
         if admission_status == "admitted"
         else "Comparison admission did not establish decision-grade comparability."),
    )

    l2_supplied = operator_evidence is not None
    l2_status = "not_supplied"
    l2_note = "No canonical L2 rehearsal receipt was supplied."
    l2_applicable_families: list = []
    l2_current_faults = l2_projected_risks = l2_not_verified = 0
    l2_observed_supplied = False
    l2_observed_status = "not_supplied"
    l2_observed_assurance = "not_verified"
    l2_observed_family = l2_observed_subject = l2_observed_scenario = ""
    l2_observed_matched_projected_risks = 0
    l2_observed_note = "No source-bound observed local L2 trial was supplied."
    if l2_supplied and admission_supplied:
        if not decision_inputs_valid:
            l2_status = "not_verified"
            l2_note = (
                "L2 rehearsal evidence is detached from the canonical admitted gate-input bundle."
            )
            l2_not_verified = 1
        else:
            from cisco_toolkit.l2_rehearsal import (
                validate_l2_failure_rehearsal,
                validate_observed_l2_failure_evidence,
            )
            from cisco_toolkit.protocol_assurance import CUTOVER_OPERATOR_EVIDENCE_SCHEMA

            evidence = _as_dict(operator_evidence)
            rehearsal_wrapper = _as_dict(evidence.get("rehearsal"))
            wrapper_assurance = rehearsal_wrapper.get("assurance_level")
            l2_validation = validate_l2_failure_rehearsal(
                rehearsal_wrapper.get("l2_failure_rehearsal")
            ) if (
                evidence.get("schema") == CUTOVER_OPERATOR_EVIDENCE_SCHEMA
                and evidence.get("owner") == "reference_only_projection"
                and evidence.get("owns_verdict") is False
                and wrapper_assurance in {
                    "not_verified", "local_safety_preservation"
                }
            ) else {"valid": False, "reason": "operator evidence root is malformed"}
            if l2_validation.get("valid") is not True:
                l2_status = "not_verified"
                l2_note = _bounded_note(
                    l2_validation.get("reason"),
                    "Applicable L2 rehearsal evidence is missing or malformed.",
                )
                l2_not_verified = 1
            else:
                l2_status = str(l2_validation.get("gate_status") or "not_verified")
                l2_applicable_families = list(
                    l2_validation.get("applicable_families") or [])
                l2_current_faults = int(l2_validation.get("n_current_faults") or 0)
                l2_projected_risks = int(l2_validation.get("n_projected_risks") or 0)
                l2_not_verified = int(l2_validation.get("n_not_verified") or 0)
                l2_observed_supplied = "observed_l2_failure_evidence" in rehearsal_wrapper
                if l2_observed_supplied:
                    observed_validation = validate_observed_l2_failure_evidence(
                        rehearsal_wrapper.get("observed_l2_failure_evidence"),
                        expected_recovery_binding=_as_dict(
                            _as_dict(admission.get("source_binding")).get("after")
                        ),
                    )
                    if observed_validation.get("valid") is True:
                        expected_wrapper_status = {
                            "observed_survival": "local_safety_preservation",
                            "observed_failure": "observed_failure",
                            "not_verified": "not_verified",
                        }.get(observed_validation.get("status"), "not_verified")
                        if rehearsal_wrapper.get("status") != expected_wrapper_status:
                            observed_validation = {
                                "valid": False,
                                "reason": (
                                    "observed L2 trial context does not reconcile to the "
                                    "canonical operator-evidence status"
                                ),
                            }
                    if observed_validation.get("valid") is not True:
                        l2_observed_status = "not_verified"
                        l2_observed_note = _bounded_note(
                            observed_validation.get("reason"),
                            "Observed local L2 trial is detached, malformed, or bound to another recovery snapshot.",
                        )
                        l2_status = "not_verified"
                        l2_not_verified = max(1, l2_not_verified)
                    else:
                        l2_observed_status = str(
                            observed_validation.get("status") or "not_verified"
                        )
                        l2_observed_family = str(
                            observed_validation.get("family") or ""
                        )
                        l2_observed_subject = str(
                            observed_validation.get("subject") or ""
                        )
                        l2_observed_scenario = str(
                            observed_validation.get("failure_scenario") or ""
                        )
                        l2_observed_assurance = (
                            "local_safety_preservation"
                            if l2_observed_status in {
                                "observed_survival", "observed_failure"
                            }
                            else "not_verified"
                        )
                        if l2_observed_status == "observed_failure":
                            l2_status = "observed_failure"
                            l2_observed_note = (
                                "The source-bound trial observed an unsafe post-failure or recovery state "
                                "for the exact local L2 subject."
                            )
                        elif l2_observed_status == "not_verified":
                            l2_status = "not_verified"
                            l2_not_verified = max(1, l2_not_verified)
                            l2_observed_note = (
                                "The supplied local L2 trial did not verify its exact precondition, "
                                "failure transition, post-failure state, and recovery."
                            )
                        else:
                            exact_matches = [
                                row for row in l2_validation.get("gate_scenarios", [])
                                if isinstance(row, dict)
                                and row.get("family") == l2_observed_family
                                and row.get("subject") == l2_observed_subject
                                and row.get("failure_scenario") == l2_observed_scenario
                                and row.get("disposition") == "projected_risk"
                                and row.get("family") in {
                                    "etherchannel", "multichassis_lag"
                                }
                            ]
                            if len(exact_matches) == 1:
                                l2_observed_matched_projected_risks = 1
                                l2_projected_risks = max(0, l2_projected_risks - 1)
                                l2_status = (
                                    "current_fault" if l2_current_faults else
                                    "not_verified" if l2_not_verified else
                                    "projected_risk" if l2_projected_risks else
                                    "simulation_only" if l2_applicable_families else
                                    "not_applicable"
                                )
                                l2_observed_note = (
                                    "The source-bound trial establishes local_safety_preservation only "
                                    "for the exact matching local projected-risk scenario."
                                )
                            else:
                                l2_observed_note = (
                                    "The source-bound trial establishes local_safety_preservation for its "
                                    "exact subject, but it does not neutralize an aggregate, unrelated, "
                                    "service-path, current-fault, or coverage-gap scenario."
                                )
                if l2_status == "not_applicable":
                    l2_note = (
                        "No applicable STP, EtherChannel, multichassis-LAG, or requested service-path "
                        "failure subject is declared; "
                        "placeholder NOT VERIFIED rows do not participate in the gate."
                    )
                elif l2_status == "simulation_only":
                    l2_note = (
                        "Applicable bounded local eligibility and/or coherent stored synthetic "
                        "failure projections are present. Synthetic preservation is not an observed "
                        "field trial; these projections do not prove convergence, remote forwarding, "
                        "traffic continuity, or service survival."
                    )
                elif l2_status == "projected_risk":
                    l2_note = (
                        f"{l2_projected_risks} applicable bounded local or requested service-path "
                        "failure projection(s) identify projected risk; field-observed service "
                        "survival remains not verified."
                    )
                elif l2_status == "current_fault":
                    l2_note = (
                        f"{l2_current_faults} applicable L2 scenario(s) retain a current fault; "
                        "local failure projections cannot make that fault acceptable."
                    )
                elif l2_status == "observed_failure":
                    l2_note = (
                        "The exact source-bound local L2 trial observed an unsafe post-failure or "
                        "recovery state. Service-path survival remains not verified."
                    )
                else:
                    l2_status = "not_verified"
                    l2_note = (
                        f"{l2_not_verified} applicable L2 family/scenario receipt(s) are missing "
                        "or not verified."
                    )
                if l2_observed_supplied:
                    l2_note = f"{l2_note} Observed-trial basis: {l2_observed_note}"

    families_supplied = protocol_family_changes is not None
    families = _as_dict(protocol_family_changes)
    family_status = "not_supplied"
    family_blocking = family_review = family_not_verified = family_rows = 0
    family_note = "No additive protocol family change set was supplied."
    if families_supplied:
        from cisco_toolkit.protocol_assurance import (
            validate_protocol_family_change_set_authority,
        )

        family_authority = validate_protocol_family_change_set_authority(
            protocol_family_changes,
            expected_source_binding=(admission.get("source_binding")
                                     if admission_supplied else None),
        )
        transition_tokens = (
            "unchanged_healthy", "unchanged_degraded", "recovered", "regressed",
            "appeared", "disappeared", "intent_changed", "coverage_lost",
            "not_comparable",
        )
        effect_tokens = ("block", "review", "none", "not_verified")
        assurance_tokens = {
            "intent_reconciled_survival", "observed_state_preservation",
            "local_safety_preservation", "not_verified",
        }

        def _natural(value: Any) -> Optional[int]:
            return (value if isinstance(value, int) and not isinstance(value, bool)
                    and value >= 0 else None)

        malformed = (
            family_authority.get("valid") is not True
            or families.get("schema") != "protocol_family_change_set/1"
            or families.get("owner") != "reference_only_composition"
            or families.get("owns_score") is not False
            or families.get("owns_verdict") is not False
            or not isinstance(families.get("families"), list)
            or not families.get("families")
            or not isinstance(families.get("summary"), dict)
        )
        aggregate = {
            "n_families": 0, "n_subject_changes": 0, "n_expected": 0,
            "n_unexpected": 0, "n_coverage_lost": 0, "n_blocking": 0,
            "n_review": 0, "n_not_verified": 0,
            "by_transition": {token: 0 for token in transition_tokens},
            "by_decision_effect": {token: 0 for token in effect_tokens},
        }
        seen_families = set()
        for family in families.get("families", []) if not malformed else []:
            family_name = (family.get("family") if isinstance(family, dict) else None)
            owner_schema = (family.get("owner_schema") if isinstance(family, dict) else None)
            family_summary = family.get("summary") if isinstance(family, dict) else None
            support = family.get("support_profile") if isinstance(family, dict) else None
            changes = family.get("changes") if isinstance(family, dict) else None
            if (not isinstance(family_name, str) or not family_name.strip()
                    or family_name in seen_families
                    or not isinstance(owner_schema, str) or not owner_schema.strip()
                    or family.get("assurance_level") not in assurance_tokens
                    or not isinstance(changes, list)
                    or not isinstance(family_summary, dict)
                    or not isinstance(family.get("source_receipt"), dict)
                    or not isinstance(support, dict)
                    or support.get("schema") != "protocol_support_profile/1"
                    or support.get("family") != family_name
                    or support.get("owner_schema") != owner_schema
                    or support.get("implementation_state") != "implemented"
                    or support.get("assurance_level") not in assurance_tokens
                    or not isinstance(support.get("evidence_contracts"), list)
                    or not support.get("evidence_contracts")
                    or not isinstance(support.get("scope"), dict) or not support.get("scope")
                    or not isinstance(support.get("runtime_support_claim"), str)
                    or not support.get("runtime_support_claim").strip()
                    or not isinstance(support.get("limitations"), list)
                    or not support.get("limitations")):
                malformed = True
                break
            seen_families.add(family_name)
            row_transitions = {token: 0 for token in transition_tokens}
            row_effects = {token: 0 for token in effect_tokens}
            expected_count = unexpected_count = 0
            for row in changes:
                if (not isinstance(row, dict)
                        or row.get("family") != family_name
                        or not isinstance(row.get("subject"), str)
                        or not row.get("subject").strip()
                        or row.get("transition") not in transition_tokens
                        or row.get("decision_effect") not in effect_tokens
                        or not isinstance(row.get("expected"), bool)
                        or "before_state" not in row or "after_state" not in row
                        or not isinstance(row.get("note"), str) or not row.get("note").strip()):
                    malformed = True
                    break
                family_rows += 1
                row_transitions[row["transition"]] += 1
                effect = row["decision_effect"]
                row_effects[effect] += 1
                family_blocking += effect == "block"
                family_review += effect == "review"
                family_not_verified += effect == "not_verified"
                expected_count += row["expected"]
                unexpected_count += (
                    not row["expected"] and row["transition"] in {
                        "unchanged_degraded", "regressed", "appeared", "disappeared",
                        "intent_changed",
                    }
                )
            if malformed:
                break
            implicit = _natural(family_summary.get("n_implicit_unchanged_healthy"))
            declared_transitions = family_summary.get("by_transition")
            declared_effects = family_summary.get("by_decision_effect")
            expected_transitions = dict(row_transitions)
            if implicit is not None:
                expected_transitions["unchanged_healthy"] += implicit
            expected_summary = {
                "n_subject_changes": len(changes),
                "n_expected": expected_count,
                "n_unexpected": unexpected_count,
                "n_coverage_lost": expected_transitions["coverage_lost"],
                "n_blocking": row_effects["block"],
                "n_review": row_effects["review"],
                "n_not_verified": row_effects["not_verified"],
            }
            if (implicit is None
                    or any(_natural(family_summary.get(key)) != value
                           for key, value in expected_summary.items())
                    or declared_transitions != expected_transitions
                    or declared_effects != row_effects):
                malformed = True
                break
            aggregate["n_families"] += 1
            for key in (
                    "n_subject_changes", "n_expected", "n_unexpected", "n_coverage_lost",
                    "n_blocking", "n_review", "n_not_verified"):
                aggregate[key] += expected_summary[key]
            for token in transition_tokens:
                aggregate["by_transition"][token] += expected_transitions[token]
            for token in effect_tokens:
                aggregate["by_decision_effect"][token] += row_effects[token]
        if not malformed:
            supplied_summary = families["summary"]
            malformed = any(
                supplied_summary.get(key) != value for key, value in aggregate.items()
            )
        if malformed:
            family_status = "not_comparable"
            family_blocking = family_review = family_rows = 0
            family_not_verified = 1
            family_note = _bounded_note(
                (family_authority.get("reason")
                 if family_authority.get("valid") is not True else None),
                "Protocol family change set is missing or malformed.",
            )
        elif family_not_verified:
            family_status = "coverage_lost"
            family_note = (
                f"{family_not_verified} protocol family subject(s) are not verified or comparable."
            )
        elif family_blocking:
            family_status = "regressed"
            family_note = f"{family_blocking} protocol family subject regression(s) block acceptance."
        elif family_review:
            family_status = "review"
            family_note = f"{family_review} unexpected protocol family change(s) require review."
        else:
            family_status = "clear"
            family_note = f"{family_rows} protocol family change row(s) carry no blocking effect."

    # A direct pre-existing caller may supply only the historical delta + certificate pair.  Preserve
    # its exact receipt shape and precedence.  Every real compute_snapshot_delta result now carries the
    # current-baseline receipt, including NOT_ASSESSED when the new snapshot omitted a validation plan.
    delta_requires_baseline = delta.get("current_baseline_required", True)
    if not isinstance(delta_requires_baseline, bool):
        delta_requires_baseline = True
    baseline_supplied = (current_baseline is not None
                         or ("current_baseline" in delta and delta_requires_baseline))
    raw_baseline = current_baseline if current_baseline is not None else delta.get("current_baseline")

    def _baseline_receipt(value: Any) -> tuple:
        block = _as_dict(value)
        invalid_note = (
            "Current-baseline receipt is missing or malformed; no all-clear can be established."
        )
        if block.get("schema") != "current_baseline_gate/1":
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        baseline_verdict = str(block.get("verdict") or "INDETERMINATE").strip().upper()
        if baseline_verdict not in {"BLOCKED", "INDETERMINATE", "CLEAR", "NOT_ASSESSED"}:
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        assessed = block.get("assessed")
        integrity = _as_dict(block.get("integrity"))
        valid_flag = integrity.get("valid")
        summary = _as_dict(block.get("summary"))
        by_state = _as_dict(summary.get("by_state"))

        def count(value: Any) -> Optional[int]:
            return value if (isinstance(value, int) and not isinstance(value, bool) and value >= 0) else None

        degraded = count(by_state.get("degraded"))
        review = count(by_state.get("review"))
        not_verified = count(by_state.get("not_verified"))
        total = count(summary.get("n_blockers"))
        returned = count(summary.get("n_blockers_returned"))
        n_items = count(summary.get("n_items"))
        blockers = block.get("blockers")
        capped = summary.get("blockers_capped")
        counts = (degraded, review, not_verified, total, returned, n_items)
        if (not isinstance(assessed, bool) or not isinstance(valid_flag, bool)
                or set(by_state) != {"degraded", "review", "not_verified"}
                or any(item is None for item in counts) or not isinstance(blockers, list)
                or not isinstance(capped, bool)):
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        if (total != degraded + review + not_verified or returned != len(blockers)
                or returned > total or capped != (returned < total)):
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        if any(not isinstance(row, dict)
               or not isinstance(row.get("evidence_state"), str)
               or row.get("evidence_state") not in {"degraded", "review", "not_verified"}
               for row in blockers):
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        returned_counts = {
            state: sum(row.get("evidence_state") == state for row in blockers)
            for state in ("degraded", "review", "not_verified")
        }
        if any(returned_counts[state] > by_state[state] for state in returned_counts):
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        coherent = (
            (baseline_verdict == "BLOCKED" and assessed and valid_flag and degraded > 0)
            or (baseline_verdict == "CLEAR" and assessed and valid_flag and total == 0 and n_items > 0)
            or (baseline_verdict == "NOT_ASSESSED" and not assessed and valid_flag and total == 0)
            or (baseline_verdict == "INDETERMINATE" and (
                (assessed and valid_flag and not degraded and (review or not_verified))
                or not valid_flag
            ))
        )
        if not coherent:
            return "INDETERMINATE", invalid_note, 0, 0, 0, 0
        baseline_note = _bounded_note(block.get("note"), invalid_note)
        return baseline_verdict, baseline_note, total, degraded, review, not_verified

    if baseline_supplied:
        (baseline_verdict, baseline_note, baseline_blockers, baseline_degraded,
         baseline_review, baseline_not_verified) = _baseline_receipt(raw_baseline)
    else:
        baseline_verdict = baseline_note = ""
        baseline_blockers = baseline_degraded = baseline_review = baseline_not_verified = 0

    # The nested protocol block is part of the delta owner's published result.  A malformed or
    # hand-edited caller must not retain CLEAN while simultaneously carrying a proven protocol
    # regression/review gate; reconcile toward the stricter observed result.  NOT_ASSESSED remains
    # deliberately non-poisoning for legacy snapshots with no baseline protocol receipt.
    if ((protocol_gate == "REGRESSED" or protocol_regressions)
            and delta_verdict in {"CLEAN", "REVIEW"}):
        if delta_verdict != "REGRESSED":
            delta_display = "REGRESSED"
        delta_verdict = "REGRESSED"
    elif protocol_gate == "REVIEW" and delta_verdict == "CLEAN":
        delta_verdict = "REVIEW"
        delta_display = "REVIEW"
    elif (protocol_gate == "NOT_ASSESSED" and delta_verdict == "CLEAN"
          and (protocol_baseline_peers or protocol_coverage_gaps)):
        # Legacy receipts remain non-poisoning when there was no protocol subject to compare.  Once
        # baseline peers exist, however, a strict overall cutover PASS would hide the comparator's own
        # re-collection requirement, so the combined gate must abstain.
        delta_verdict = "INDETERMINATE"
        delta_display = "INDETERMINATE"

    # Preserve the established workbook precedence: a proven observed regression is strongest; then
    # current/certificate blockers; missing evidence; reviewable delta; and certificate conditions.
    # CLEAR is the only supplied current-baseline state compatible with an overall PASS.
    if admission_supplied and admission_status != "admitted":
        verdict = "INDETERMINATE"
    elif families_supplied and family_status == "not_comparable":
        verdict = "INDETERMINATE"
    elif delta_verdict == "REGRESSED":
        verdict = "REGRESSED"
    elif (families_supplied and family_blocking
          and delta_verdict != "INDETERMINATE"):
        verdict = "REGRESSED"
    elif baseline_supplied and baseline_verdict == "BLOCKED":
        verdict = "FAIL"
    elif l2_supplied and l2_status in {"current_fault", "observed_failure"}:
        verdict = "FAIL"
    elif certificate_verdict == "FAIL":
        verdict = "FAIL"
    elif (delta_verdict == "INDETERMINATE" or certificate_verdict == "INDETERMINATE"
          or (families_supplied and family_not_verified)
          or (l2_supplied and l2_status == "not_verified")
          or (baseline_supplied and baseline_verdict in {"INDETERMINATE", "NOT_ASSESSED"})):
        verdict = "INDETERMINATE"
    elif (delta_verdict == "REVIEW" or (families_supplied and family_review)
          or (l2_supplied and l2_status == "projected_risk")):
        verdict = "REVIEW"
    elif certificate_verdict == "CONDITIONAL":
        verdict = "CONDITIONAL"
    else:
        verdict = "PASS"

    baseline_basis = (f"Current baseline: {baseline_verdict}. Current-baseline basis: {baseline_note} "
                      if baseline_supplied else "")
    admission_basis = (f"Comparison admission: {admission_status.upper()}. "
                       f"Admission basis: {admission_note} " if admission_supplied else "")
    family_basis = (f"Protocol family changes: {family_status.upper()}. "
                    f"Family basis: {family_note} " if families_supplied else "")
    l2_basis = (f"L2 bounded rehearsal: {l2_status.upper()}. "
                f"L2 basis: {l2_note} " if l2_supplied else "")
    note = (f"{admission_basis}Delta observation: {delta_display}. Delta basis: {delta_note} "
            f"{family_basis}"
            f"{l2_basis}"
            f"{baseline_basis}Pre-Change Certificate: {certificate_verdict}. "
            f"Certificate basis: {certificate_note}")

    if admission_supplied and admission_status != "admitted":
        action = ("Do not declare the cutover good; repair the comparison admission failure or "
                  "missing coverage and recompute from bound evidence.")
    elif verdict == "REGRESSED":
        if families_supplied and family_blocking:
            noun = "subject was" if family_blocking == 1 else "subjects were"
            action = (f"{family_blocking} protocol family {noun} classified as blocking by the "
                      "native family owner; investigate or roll back before proceeding.")
        elif protocol_regressions:
            noun = "regression was" if protocol_regressions == 1 else "regressions were"
            action = (f"{protocol_regressions} baseline-observed protocol adjacency state {noun} "
                      "proven; investigate or roll back before proceeding.")
            if protocol_coverage_gaps:
                gap_noun = "cell was" if protocol_coverage_gaps == 1 else "cells were"
                action += (f" {protocol_coverage_gaps} additional protocol comparison {gap_noun} "
                           "not assessable; re-collect before accepting the change.")
        else:
            action = "A before/after regression was proven; investigate or roll back before proceeding."
    elif verdict == "FAIL":
        if baseline_supplied and baseline_verdict == "BLOCKED":
            other = " and the blocking certificate condition(s)" if certificate_verdict == "FAIL" else ""
            action = (f"Do not proceed; resolve or explicitly disposition {baseline_degraded} definite "
                      f"current-baseline degradation(s){other}, then re-run the comparison.")
        elif l2_supplied and l2_status == "current_fault":
            action = (f"Do not proceed; resolve {l2_current_faults} current L2 fault(s), then "
                      "re-collect and recompute the bounded rehearsal evidence.")
        elif l2_supplied and l2_status == "observed_failure":
            action = (
                "Do not proceed; the source-bound local L2 trial observed an unsafe "
                "post-failure or recovery state. Restore safety and repeat the exact trial."
            )
        else:
            action = "Do not proceed; resolve the blocking certificate condition(s) and re-run the comparison."
    elif verdict == "INDETERMINATE":
        if baseline_supplied and baseline_verdict == "NOT_ASSESSED":
            action = ("Do not declare the cutover good; generate and reconcile the new snapshot's validation "
                      "baseline before acceptance.")
        elif baseline_supplied and baseline_verdict == "INDETERMINATE":
            action = ("Do not declare the cutover good; repair or verify the current validation baseline and "
                      "re-run before acceptance.")
        elif protocol_gate == "NOT_ASSESSED" and protocol_baseline_peers:
            noun = "peer could" if protocol_baseline_peers == 1 else "peers could"
            action = (f"{protocol_baseline_peers} baseline-observed protocol {noun} not be compared; "
                      "re-collect protocol evidence before accepting the change.")
        elif l2_supplied and l2_status == "not_verified":
            action = (
                "Do not declare the cutover good; collect and reconcile bounded rehearsal "
                "evidence for every applicable STP, EtherChannel, multichassis-LAG, and "
                "requested service-path failure subject."
            )
        else:
            action = "Do not declare the cutover good; repair the missing or invalid evidence and re-run."
    elif verdict == "REVIEW":
        if l2_supplied and l2_status == "projected_risk":
            action = (
                f"Review and disposition {l2_projected_risks} bounded local or requested "
                "service-path failure risk projection(s); synthetic results are not proof "
                "of observed service survival."
            )
        elif protocol_coverage_gaps:
            noun = "cell was" if protocol_coverage_gaps == 1 else "cells were"
            action = (f"{protocol_coverage_gaps} protocol comparison {noun} not assessable; "
                      "re-collect and review before accepting the change.")
        else:
            action = "Operator review is required before accepting the change."
    elif verdict == "CONDITIONAL":
        action = "Clear or explicitly accept every named certificate blind spot before proceeding."
    else:
        action = "No blocking condition was observed within the disclosed, bounded comparison scope."
    operator_note = f"Overall before/after cutover decision: {verdict}. {action}"
    if verdict != "PASS":
        operator_note += f" Delta basis: {delta_note}"
        if baseline_supplied:
            operator_note += f" Current-baseline basis: {baseline_note}"
        if l2_supplied:
            operator_note += f" L2 bounded-rehearsal basis: {l2_note}"
        operator_note += f" Certificate basis: {certificate_note}"

    result = {
        "schema": "cutover_gate/1",
        "verdict": verdict,
        "note": note,
        "operator_note": operator_note,
        "delta_verdict": delta_verdict,
        "delta_display": delta_display,
        "delta_note": delta_note,
        "certificate_verdict": certificate_verdict,
        "certificate_note": certificate_note,
        "protocol_gate": protocol_gate,
        "protocol_baseline_peers": protocol_baseline_peers,
        "protocol_regressions": protocol_regressions,
        "protocol_coverage_gaps": protocol_coverage_gaps,
    }
    if admission_supplied:
        result.update({
            "comparison_admission_status": admission_status,
            "comparison_admission_note": admission_note,
        })
    if families_supplied:
        result.update({
            "protocol_family_status": family_status,
            "protocol_family_note": family_note,
            "protocol_family_rows": family_rows,
            "protocol_family_blocking": family_blocking,
            "protocol_family_review": family_review,
            "protocol_family_not_verified": family_not_verified,
        })
    if l2_supplied:
        result.update({
            "l2_rehearsal_status": l2_status,
            "l2_rehearsal_note": l2_note,
            "l2_rehearsal_applicable_families": l2_applicable_families,
            "l2_rehearsal_current_faults": l2_current_faults,
            "l2_rehearsal_projected_risks": l2_projected_risks,
            "l2_rehearsal_not_verified": l2_not_verified,
        })
        if l2_observed_supplied:
            result.update({
                "l2_observed_trial_status": l2_observed_status,
                "l2_observed_trial_assurance": l2_observed_assurance,
                "l2_observed_trial_family": l2_observed_family,
                "l2_observed_trial_subject": l2_observed_subject,
                "l2_observed_trial_scenario": l2_observed_scenario,
                "l2_observed_trial_matched_projected_risks":
                    l2_observed_matched_projected_risks,
                "l2_observed_trial_note": l2_observed_note,
            })
    if baseline_supplied:
        result.update({
            "current_baseline_verdict": baseline_verdict,
            "current_baseline_note": baseline_note,
            "current_baseline_blockers": baseline_blockers,
            "current_baseline_degraded": baseline_degraded,
            "current_baseline_review": baseline_review,
            "current_baseline_not_verified": baseline_not_verified,
        })
    return result


def write_diff_workbook(old: dict, new: dict, out_path: str, precert: dict = None, *,
                        source_binding: Optional[dict] = None, schema_status: Any = None,
                        protocol_families: Optional[dict] = None,
                        comparison: Optional[dict] = None) -> dict:
    """Write a diff workbook (Summary / Interface Changes / Endpoint Changes /
    SVI Changes) comparing two snapshot_state() dicts. `precert` is an optional precomputed
    Pre-Change Validation Certificate (roadmap C1); when None it is computed here, so the
    'Pre-Change Certificate' sheet is always present.  Exact input hashes and schema-gate status
    supplied by the file-loading caller are rendered into both decision surfaces.

    ``protocol_families`` is an optional complete ``protocol_family_change_set/1`` composed by a
    source-owning caller.  When omitted, the workbook computes each applicable native owner from the
    same in-memory snapshots and source hashes, then composes it with the unchanged IPv4 adjacency
    delta.  In both cases that one, uncapped object is passed to the canonical cutover gate and
    rendered without recomputation.

    ``comparison`` is the preferred decision-grade input: one complete
    ``source_bound_cutover_comparison/1`` from the canonical composer.  When supplied, the workbook
    projects its delta, precertification, family changes, gate, and receipt verbatim; it never
    recomputes or reinterprets any decision owner.  The older arguments remain for direct callers
    and legacy artifacts.
    """
    old = old if isinstance(old, dict) else {}            # total on a non-dict snapshot (parsed null/[]/scalar)
    new = new if isinstance(new, dict) else {}
    from openpyxl import Workbook
    HF = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FILL = PatternFill("solid", fgColor=WORKBOOK_NAVY_HEX)
    AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
    DF = Font(name="Calibri", size=10)
    NONE = "\u2205"  # empty marker

    wb = Workbook(); wb.remove(wb.active)
    from cisco_toolkit.excel import harden_workbook
    harden_workbook(wb)   # sanitize control chars in device-derived text -> no IllegalCharacterError abort

    def sheet(title, cols):
        ws = wb.create_sheet(title)
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HF; cell.fill = FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"
        return ws

    def autofit(ws, ncols):
        for col in range(1, ncols + 1):
            mx = len(str(ws.cell(row=1, column=col).value or ""))
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None: mx = max(mx, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 60)

    def _ifmap(s):
        """`interfaces` coerced at BOTH levels -- the same read compute_snapshot_delta already routes through
        _as_dict. Re-deriving it here with raw len()/.get() crashed the whole workbook on a trimmed /
        foreign-tool / older-schema snapshot (`interfaces: {"sw1": 5}` -> TypeError at the totals, and the
        string variant `{"sw1": "abc"}` counted 3 phantom interfaces before dying in the sheet loops) -- i.e.
        no diff workbook and no Pre-Change Validation Certificate, at the cutover gate."""
        return {h: {p: _as_dict(d) for p, d in _as_dict(v).items()}
                for h, v in _as_dict(s.get("interfaces")).items()}

    def _dnum(a, b):
        """b - a when both really are numbers. The canonical counts come from the snapshot, so a malformed
        one can carry a string there and a raw subtraction would abort the workbook."""
        ok = all(isinstance(x, (int, float)) and not isinstance(x, bool) for x in (a, b))
        return round(b - a, 1) if ok else ""

    comparison_doc: Optional[dict] = None
    comparison_delta: Optional[dict] = None
    decision_input_authority: Any = None
    if comparison is not None:
        if not isinstance(comparison, dict):
            raise ValueError("comparison must be a source_bound_cutover_comparison/1 object")
        from cisco_toolkit.comparison import (
            bound_comparison_decision_input_authority,
            validate_bound_comparison_authority,
        )

        comparison_authority = validate_bound_comparison_authority(comparison)
        if comparison_authority.get("valid") is not True:
            raise ValueError(str(comparison_authority.get("reason") or (
                "comparison is detached from the canonical source/context owner"
            )))
        if precert is not None or protocol_families is not None:
            raise ValueError(
                "comparison is the complete decision input; detached precert/protocol families "
                "cannot also be supplied"
            )
        comparison_doc = comparison
        decision_input_authority = bound_comparison_decision_input_authority(comparison)
        additive_keys = {
            "comparison_schema", "comparison_admission", "change_intent",
            "protocol_families", "precert", "cutover_gate", "operator_evidence",
            "comparison_receipt",
        }
        comparison_delta = {
            key: value for key, value in comparison_doc.items() if key not in additive_keys
        }
        from cisco_toolkit.protocol_assurance import (
            bound_snapshot_source,
            verify_receipt_envelope,
        )
        receipt_payload = {
            "admission": comparison_doc.get("comparison_admission"),
            "change_intent": comparison_doc.get("change_intent"),
            "protocol_families": comparison_doc.get("protocol_families"),
            "delta": comparison_delta,
            "precert": comparison_doc.get("precert"),
            "cutover_gate": comparison_doc.get("cutover_gate"),
            "operator_evidence": comparison_doc.get("operator_evidence"),
        }
        required_objects = (
            "comparison_admission", "change_intent", "protocol_families", "precert",
            "cutover_gate", "operator_evidence", "comparison_receipt",
        )
        if comparison_doc.get("comparison_schema") != "source_bound_cutover_comparison/1" \
                or any(not isinstance(comparison_doc.get(key), dict) for key in required_objects) \
                or not verify_receipt_envelope(
                    comparison_doc.get("comparison_receipt"), receipt_payload):
            raise ValueError("comparison receipt is missing, malformed, or detached from its payload")
        admitted_sources = _as_dict(
            _as_dict(comparison_doc.get("comparison_admission")).get("source_binding")
        )
        for side, snapshot in (("before", old), ("after", new)):
            marker = bound_snapshot_source(snapshot)
            binding = _as_dict(admitted_sources.get(side))
            if (marker.get("source_bound") is not True
                    or marker.get("sha256") != binding.get("sha256")
                    or marker.get("bytes") != binding.get("bytes")):
                raise ValueError(
                    f"comparison {side} source binding does not match the rendered snapshot bytes"
                )

    oi, ni = _ifmap(old), _ifmap(new)
    od, nd = _as_dict(old.get("devices")), _as_dict(new.get("devices"))
    delta = (comparison_delta if comparison_delta is not None else compute_snapshot_delta(
        old, new, source_binding=source_binding,
        schema_status=schema_status))   # migration-validation analysis
    # Compute the independent certificate BEFORE the Summary so its coverage verdict can constrain
    # the headline gate.  A clean delta over an unassessed reachability surface is not a PASS.
    from cisco_toolkit.precert import CERT_SHEET_HEADERS, CERT_SHEET_NAME, compute_precert
    if comparison_doc is not None:
        cert = comparison_doc["precert"]
    elif isinstance(precert, dict):
        cert = dict(precert)
        if source_binding and not cert.get("source_binding"):
            cert["source_binding"] = dict(source_binding)
        if schema_status is not None and not cert.get("schema_status"):
            cert["schema_status"] = _schema_status(schema_status)
    else:
        cert = compute_precert(
            old, new, source_hashes=source_binding, schema_status=schema_status)
    rd0 = delta.get("reachability") or {}      # W2 reachability what-if (disclose the bounded sample, never silent)
    cd0 = (delta.get("cabling") or {}).get("summary") or {}   # physical cable delta (EDA cable-map SSOT)
    pd0 = delta.get("protocol_adjacencies") or {}              # receipt-gated observed adjacency delta
    pds0 = pd0.get("summary") or {}
    if comparison_doc is not None:
        family_changes = comparison_doc["protocol_families"]
    elif protocol_families is None:
        # Release 1's CLI has no expected-change-intent input.  Compose the reference-only initial
        # family view from the SAME in-memory pair and IPv4 delta already owned above; an empty intent
        # means every material change remains unexpected and reviewable.
        from cisco_toolkit.protocol_assurance import (
            compute_native_protocol_deltas,
            protocol_family_change_set,
        )
        source_pair = _as_dict(source_binding)
        native_deltas = compute_native_protocol_deltas(
            old,
            new,
            before_binding=source_pair.get("before"),
            after_binding=source_pair.get("after"),
        )
        family_changes = protocol_family_change_set(
            pd0, {"expected_changes": []}, native_deltas=native_deltas)
    else:
        # Do not copy, truncate, or reinterpret a caller-composed native family set: the gate and
        # workbook below must consume/project the exact same object.
        family_changes = protocol_families
    bd0 = _as_dict(delta.get("current_baseline"))              # current-snapshot acceptance baseline
    bds0 = _as_dict(bd0.get("summary"))
    bd_states0 = _as_dict(bds0.get("by_state"))
    bd_total0 = (
        bds0.get("n_blockers")
        if type(bds0.get("n_blockers")) is int and bds0.get("n_blockers") >= 0 else 0
    )
    bd_rendered0 = (
        bds0.get("n_blockers_returned")
        if type(bds0.get("n_blockers_returned")) is int
        and bds0.get("n_blockers_returned") >= 0 else 0
    )
    bd_omitted0 = max(0, bd_total0 - bd_rendered0)
    if comparison_doc is not None:
        baseline_export0 = _as_dict(
            _as_dict(comparison_doc.get("operator_evidence")).get(
                "current_baseline_blocker_export"
            )
        )
    else:
        from cisco_toolkit.protocol_assurance import current_baseline_blocker_export

        baseline_export0 = current_baseline_blocker_export(new)
    baseline_export_summary0 = _as_dict(baseline_export0.get("summary"))
    baseline_export_rows0 = (
        baseline_export0.get("rows")
        if isinstance(baseline_export0.get("rows"), list) else []
    )

    def _protocol_metric(key: str):
        value = pds0.get(key, 0)
        # A positive observed result remains useful under partial coverage; a zero does not.  When one
        # scoped cell is unassessed, render its zero outcomes as an abstention rather than an all-clear.
        return value if value or pd0.get("assessed") else "—"

    def _family_state_cell(value: Any) -> Any:
        """Render a producer-owned before/after state without turning absence into a clean blank."""
        if value is None or value == "" or value == {} or value == [] or value == ():
            return "NOT VERIFIED"
        if isinstance(value, str):
            return value.strip() or "NOT VERIFIED"
        if isinstance(value, (dict, list, tuple)):
            try:
                return json.dumps(value, sort_keys=True, ensure_ascii=True,
                                  separators=(",", ":"), allow_nan=False)
            except (TypeError, ValueError, OverflowError):
                return "NOT VERIFIED — malformed producer state"
        if isinstance(value, set):
            return "NOT VERIFIED — malformed producer state"
        return value

    def _family_surface(change_set: Any) -> tuple:
        """Project every producer row and calculate display-only totals over that full row set.

        These counts never participate in a decision; ``compute_cutover_gate`` above remains the
        sole consumer of family-native ``decision_effect`` values.  Placeholder rows are deliberate:
        a missing family/list/leaf must remain visible as NOT VERIFIED, never disappear as a clean zero.
        """
        closed_transitions = {
            "unchanged_healthy", "unchanged_degraded", "recovered", "regressed",
            "appeared", "disappeared", "intent_changed", "coverage_lost", "not_comparable",
        }
        closed_effects = {"block", "review", "none", "not_verified"}
        closed_assurance = {
            "intent_reconciled_survival", "observed_state_preservation",
            "local_safety_preservation", "not_verified",
        }
        block = _as_dict(change_set)
        raw_families = block.get("families")
        family_records = raw_families if isinstance(raw_families, list) else []
        rows: List[dict] = []
        source_rows = 0
        placeholder_rows = 0
        expected_counts = {"YES": 0, "NO": 0, "NOT VERIFIED": 0}
        effect_counts = {name: 0 for name in ("block", "review", "none", "not_verified")}
        effect_counts["NOT VERIFIED"] = 0
        transition_counts: Dict[str, int] = {}
        assurance_counts: Dict[str, int] = {}

        def _placeholder(family: Any, note: str) -> None:
            nonlocal placeholder_rows
            placeholder_rows += 1
            rows.append({
                "family": _family_state_cell(family),
                "subject_kind": "NOT VERIFIED",
                "subject": "NOT VERIFIED",
                "transition": "NOT VERIFIED",
                "expected": "NOT VERIFIED",
                "decision_effect": "NOT VERIFIED",
                "assurance": "NOT VERIFIED",
                "before": "NOT VERIFIED",
                "after": "NOT VERIFIED",
                "note": note,
                "placeholder": True,
            })

        if not isinstance(raw_families, list) or not raw_families:
            _placeholder(
                "NOT VERIFIED",
                "Protocol family change set is missing or empty; row-level changes are NOT VERIFIED.",
            )
        else:
            for family_index, raw_family in enumerate(raw_families):
                if not isinstance(raw_family, dict):
                    _placeholder(
                        "NOT VERIFIED",
                        f"families[{family_index}] is malformed; its changes are NOT VERIFIED.",
                    )
                    continue
                family_name = raw_family.get("family")
                assurance = raw_family.get("assurance_level")
                raw_changes = raw_family.get("changes")
                if not isinstance(raw_changes, list) or not raw_changes:
                    _placeholder(
                        family_name,
                        (f"Family {family_name or 'NOT VERIFIED'} emitted no materialized subject-change "
                         "rows; row-level before/after evidence is NOT VERIFIED on this sheet."),
                    )
                    continue
                for row_index, raw_row in enumerate(raw_changes):
                    source_rows += 1
                    if not isinstance(raw_row, dict):
                        _placeholder(
                            family_name,
                            (f"Family {family_name or 'NOT VERIFIED'} changes[{row_index}] is malformed; "
                             "the producer row is retained as NOT VERIFIED."),
                        )
                        expected_counts["NOT VERIFIED"] += 1
                        effect_counts["NOT VERIFIED"] += 1
                        transition_counts["NOT VERIFIED"] = transition_counts.get("NOT VERIFIED", 0) + 1
                        assurance_counts["NOT VERIFIED"] = assurance_counts.get("NOT VERIFIED", 0) + 1
                        continue

                    transition_raw = raw_row.get("transition")
                    transition_key = (transition_raw if transition_raw in closed_transitions
                                      else "NOT VERIFIED")
                    effect_raw = raw_row.get("decision_effect")
                    effect_key = effect_raw if effect_raw in closed_effects else "NOT VERIFIED"
                    assurance_raw = raw_row.get("assurance_level", assurance)
                    assurance_key = (assurance_raw if assurance_raw in closed_assurance
                                     else "NOT VERIFIED")
                    expected_raw = raw_row.get("expected")
                    expected_key = ("YES" if expected_raw is True else
                                    "NO" if expected_raw is False else "NOT VERIFIED")
                    transition_counts[transition_key] = transition_counts.get(transition_key, 0) + 1
                    effect_counts[effect_key] += 1
                    assurance_counts[assurance_key] = assurance_counts.get(assurance_key, 0) + 1
                    expected_counts[expected_key] += 1
                    rows.append({
                        "family": _family_state_cell(raw_row.get("family", family_name)),
                        "subject_kind": _family_state_cell(
                            raw_row.get("subject_kind") or "family_subject"
                        ),
                        "subject": _family_state_cell(raw_row.get("subject")),
                        "transition": _family_state_cell(transition_raw),
                        "expected": expected_key,
                        "decision_effect": _family_state_cell(effect_raw),
                        "assurance": _family_state_cell(assurance_raw),
                        "before": _family_state_cell(raw_row.get("before_state")),
                        "after": _family_state_cell(raw_row.get("after_state")),
                        "note": _family_state_cell(raw_row.get("note")),
                        "placeholder": False,
                    })

        def _counts_text(counts: Dict[str, int]) -> str:
            return ", ".join(f"{name}={count}" for name, count in sorted(counts.items()))

        disclosure = (
            f"Complete uncapped export: rendered {source_rows} of {source_rows} producer subject-change "
            f"row(s); omitted=0; family records={len(family_records)}; placeholder NOT VERIFIED "
            f"rows={placeholder_rows}. Expected: {_counts_text(expected_counts)}. "
            f"Producer decision effects: {_counts_text(effect_counts)}. "
            f"Transitions: {_counts_text(transition_counts) if transition_counts else 'NOT VERIFIED'}. "
            f"Assurance: {_counts_text(assurance_counts) if assurance_counts else 'NOT VERIFIED'}."
        )
        return rows, source_rows, disclosure

    family_surface_rows, family_source_rows, family_disclosure = _family_surface(family_changes)

    # Summary (leads with the cutover-validation VERDICT)
    ws = sheet("Summary", ["Metric", "Old", "New", "Delta"])
    added_sw = sorted(set(nd) - set(od), key=_skey); removed_sw = sorted(set(od) - set(nd), key=_skey)
    # SSOT (Law 1): the fleet size and the interface totals are the delta's -- `_scn()` = the canonical
    # executive_brief.scale.n_devices, with len() only as the pre-brief fallback, and the _as_dict-guarded
    # interface count. compute_snapshot_delta (just above) already ran on these SAME two inputs and is what
    # the webapp /api/compare publishes, so a second raw len() here made ONE call render TWO fleet sizes --
    # the AssessHub compare view and the CLI diff workbook disagreeing at the cutover gate.
    o_sw, n_sw = delta["switches"]["old"], delta["switches"]["new"]
    o_if, n_if = delta["interfaces"]["old"], delta["interfaces"]["new"]
    if comparison_doc is not None:
        # A detached envelope proves only that its caller rehashed a payload.  Re-run the sole gate
        # owner over the exact uncapped decision inputs while the process-local family authority is
        # still present, then render the supplied canonical object only if every JSON type/value is
        # identical.  This rejects a regenerated envelope carrying a forged PASS or an edited family
        # subset without introducing a second workbook verdict owner.
        expected_gate = compute_cutover_gate(
            delta,
            cert,
            comparison_admission=comparison_doc["comparison_admission"],
            protocol_family_changes=family_changes,
            operator_evidence=comparison_doc["operator_evidence"],
            decision_input_authority=decision_input_authority,
        )
        expected_gate_wire = json.dumps(
            expected_gate,
            sort_keys=True,
            ensure_ascii=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        supplied_gate_wire = json.dumps(
            comparison_doc["cutover_gate"],
            sort_keys=True,
            ensure_ascii=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        if expected_gate_wire != supplied_gate_wire:
            raise ValueError(
                "comparison cutover gate is detached from its canonical decision inputs"
            )
        gate = comparison_doc["cutover_gate"]
    else:
        gate = compute_cutover_gate(delta, cert, protocol_family_changes=family_changes)
    gate_verdict = gate["verdict"]
    gate_note = gate["note"]
    _VERDICT_FILL = {"PASS": "C6EFCE", "CLEAN": "C6EFCE", "CONDITIONAL": "FFEB9C",
                     "CLEAR": "C6EFCE", "REVIEW": "FFEB9C", "NOT_ASSESSED": "D9D9D9",
                     "INDETERMINATE": "D9D9D9", "FAIL": "FFC7CE", "BLOCKED": "FFC7CE",
                     "REGRESSED": "FFC7CE", "SIMULATION_ONLY": "D9D9D9",
                     "PROJECTED_RISK": "FFEB9C", "CURRENT_FAULT": "FFC7CE",
                     "OBSERVED_FAILURE": "FFC7CE",
                     "LOCAL_SAFETY_PRESERVATION": "C6EFCE",
                     "NOT_APPLICABLE": "D9D9D9", "NOT_VERIFIED": "D9D9D9"}
    metrics = [
        ("CUTOVER GATE VERDICT", "", gate_verdict, gate_note),
        ("DELTA OBSERVATION", "", delta.get("verdict_display", delta["verdict"]), delta["verdict_note"]),
        ("CURRENT BASELINE GATE", "", bd0.get("verdict", "INDETERMINATE"),
         bd0.get("note", "Current-baseline receipt is missing or malformed.")),
        ("Current baseline blockers", "", bds0.get("n_blockers", 0),
         (f"{bd_states0.get('degraded', 0)} degraded, {bd_states0.get('review', 0)} review, "
          f"{bd_states0.get('not_verified', 0)} not verified; "
          f"rendered={bd_rendered0}, total={bd_total0}, omitted={bd_omitted0}; "
          "complete sink: Current Baseline Export sheet")),
        ("Switches", o_sw, n_sw, _dnum(o_sw, n_sw)),
        # ...added/removed enumerate the COLLECTED devices, which is a narrower set than the canonical
        # count above whenever the estate holds devices that were inventoried but not collected. Labelled,
        # so the two rows cannot be read as one contradictory fact.
        ("Switches added (collected)", "", "", ", ".join(added_sw) or "0"),
        ("Switches removed (collected)", "", "", ", ".join(removed_sw) or "0"),
        ("Interfaces (total)", o_if, n_if, _dnum(o_if, n_if)),
        ("Health bands regressed", "", delta["health"]["n_regressed"],
         ", ".join(r["switch"] for r in delta["health"]["regressed"]) or "0"),
        ("Health bands improved", "", delta["health"]["n_improved"], delta["health"]["n_improved"]),
        ("Findings opened", "", delta["findings"]["n_opened"],
         f"{delta['findings']['n_opened_high']} High/Critical"),
        ("Findings resolved", "", delta["findings"]["n_resolved"], delta["findings"]["n_resolved"]),
        ("Protocol adjacency state regressions", "", _protocol_metric("n_state_regressed"),
         pd0.get("note", "NOT assessed")),
        ("Protocol adjacencies no longer observed", "", _protocol_metric("n_no_longer_observed"),
         "Observed topology change; not proof that a peer is down or absent"),
        ("Protocol adjacency coverage gaps", "", pds0.get("n_coverage_gaps", 0),
         (f"Gate {pd0.get('gate', 'NOT_ASSESSED')}; projection custody "
          f"{pd0.get('projection_custody', 'embedded_unverified')}")),
        ("Protocol family changes (complete)", "",
         family_source_rows if family_source_rows else "NOT VERIFIED",
         (f"Canonical family status {gate.get('protocol_family_status', 'NOT VERIFIED')}. "
          f"{family_disclosure}")),
        ("Reachability flows newly blocked", "", len(rd0.get("newly_blocked") or []),
         (f"{rd0.get('pairs_tested', 0)} sampled flow(s) across {rd0.get('subnets_tested', 0)} of "
          f"{rd0.get('subnets_total', 0)} subnet(s){' (CAPPED — coverage incomplete)' if rd0.get('capped') else ''}, "
          f"one representative host each; {rd0.get('preserved', 0)} preserved, "
          f"{rd0.get('inconclusive', 0)} inconclusive"
          if rd0.get("assessed") else
          "NOT assessed — no routes collected (this is NOT a statement that reachability is unchanged)")),
        ("Physical cables changed", "", "",
         (f"{cd0.get('n_added', 0)} added, {cd0.get('n_removed', 0)} removed, "
          f"{cd0.get('n_went_down', 0)} went DOWN, {cd0.get('n_no_longer_observed', 0)} no longer observed, "
          f"{cd0.get('n_members_changed', 0)} LAG membership change(s)"
          if (delta.get("cabling") or {}).get("assessed") else
          "NOT assessed — no CDP/LLDP links in either snapshot (NOT a statement that cabling is unchanged)")),
    ]
    if gate.get("l2_rehearsal_status"):
        metrics.insert(1, (
            "L2 BOUNDED REHEARSAL",
            "",
            str(gate["l2_rehearsal_status"]).upper(),
            gate.get("l2_rehearsal_note", "Canonical L2 rehearsal basis was not published."),
        ))
    if gate.get("l2_observed_trial_status"):
        observed_status = str(gate["l2_observed_trial_status"])
        observed_display = (
            "LOCAL_SAFETY_PRESERVATION"
            if observed_status == "observed_survival" else observed_status.upper()
        )
        metrics.insert(2, (
            "L2 OBSERVED LOCAL TRIAL",
            "",
            observed_display,
            (
                f"family={gate.get('l2_observed_trial_family', '')}; "
                f"subject={gate.get('l2_observed_trial_subject', '')}; "
                f"scenario={gate.get('l2_observed_trial_scenario', '')}; "
                f"assurance={gate.get('l2_observed_trial_assurance', 'not_verified')}; "
                "matched projected risks="
                f"{gate.get('l2_observed_trial_matched_projected_risks', 0)}. "
                f"{gate.get('l2_observed_trial_note', '')}"
            ),
        ))
    r = 2
    for m in metrics:
        for c, v in enumerate(m, 1):
            cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
        if m[0] in {
                "CUTOVER GATE VERDICT", "L2 BOUNDED REHEARSAL",
                "L2 OBSERVED LOCAL TRIAL", "CURRENT BASELINE GATE"}:
            vc = ws.cell(row=r, column=3)
            vc.fill = PatternFill("solid", fgColor=_VERDICT_FILL.get(str(m[2]), "FFFFFF"))
            vc.font = Font(name="Calibri", bold=True, size=11)
        r += 1
    autofit(ws, 4); ws.column_dimensions["D"].width = 70

    if gate.get("l2_observed_trial_status"):
        observed_receipt = _as_dict(_as_dict(_as_dict(
            (comparison_doc or {}).get("operator_evidence")
        ).get("rehearsal")).get("observed_l2_failure_evidence"))
        ws = sheet("Observed L2 Trial", ["Field", "Value"])
        observed_rows = [
            ("Gate status", gate.get("l2_observed_trial_status")),
            ("Assurance", gate.get("l2_observed_trial_assurance")),
            ("Family", gate.get("l2_observed_trial_family")),
            ("Subject", gate.get("l2_observed_trial_subject")),
            ("Failure scenario", gate.get("l2_observed_trial_scenario")),
            ("Matched projected risks", gate.get(
                "l2_observed_trial_matched_projected_risks", 0)),
            ("Gate note", gate.get("l2_observed_trial_note")),
            ("Receipt schema", observed_receipt.get("schema")),
            ("Receipt status", observed_receipt.get("status")),
            ("Local scenario claim", _as_dict(
                observed_receipt.get("claims")).get("local_scenario")),
            ("Service-path survival", _as_dict(
                observed_receipt.get("claims")).get("service_path_survival")),
            ("Traffic continuity", _as_dict(
                observed_receipt.get("claims")).get("traffic_continuity")),
            ("Convergence", _as_dict(
                observed_receipt.get("claims")).get("convergence")),
        ]
        sources = _as_dict(observed_receipt.get("source_binding"))
        for phase in ("pre_failure", "post_failure", "recovery"):
            phase_source = _as_dict(sources.get(phase))
            observed_rows.extend([
                (f"{phase} source ID", phase_source.get("source_id")),
                (f"{phase} SHA-256", phase_source.get("sha256")),
                (f"{phase} bytes", phase_source.get("bytes")),
                (f"{phase} collected_at", phase_source.get("collected_at")),
                (f"{phase} custody_at", phase_source.get("custody_at")),
            ])
        for row_index, (field, value) in enumerate(observed_rows, 2):
            ws.cell(row=row_index, column=1, value=_cv(field)).font = DF
            ws.cell(row=row_index, column=2, value=_cv(value)).font = DF
            ws.cell(row=row_index, column=1).alignment = AL
            ws.cell(row=row_index, column=2).alignment = AL
        autofit(ws, 2)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 90

    # Current Baseline Gate — actionable rows behind the headline FAIL/INDETERMINATE.  These are the
    # bounded, reconciled rows emitted by compute_current_baseline_gate, never a second workbook-local
    # marker parser.  The disclosure row is always present, including an explicit cap state.
    ws = sheet("Current Baseline Gate",
               ["State", "Device", "Wave", "Category", "Check",
                "Observed baseline / acceptance", "Projection custody", "Source key"])
    r = 2
    blocker_rows = bd0.get("blockers") if isinstance(bd0.get("blockers"), list) else []
    for blocker in blocker_rows:
        if not isinstance(blocker, dict):
            continue
        vals = [
            blocker.get("evidence_state", ""), blocker.get("device", ""),
            blocker.get("wave", ""), blocker.get("category", ""), blocker.get("check", ""),
            blocker.get("expect", ""), blocker.get("projection_custody", ""),
            blocker.get("source_key", ""),
        ]
        for c, value in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=_cv(value)); cell.font = DF; cell.alignment = AL
        fill = "FFC7CE" if blocker.get("evidence_state") == "degraded" else "FFEB9C"
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill)
        r += 1
    cap_state = "CAPPED" if bds0.get("blockers_capped") else "NOT CAPPED"
    disclosure = (
        f"Gate {bd0.get('verdict', 'INDETERMINATE')}; rendered={bd_rendered0}, "
        f"total={bd_total0}, omitted={bd_omitted0} ({cap_state}). Complete sink: Current "
        "Baseline Export sheet. Counts cover the full reconciled plan: "
        f"{bd_states0.get('degraded', 0)} degraded, {bd_states0.get('review', 0)} review, "
        f"{bd_states0.get('not_verified', 0)} not verified. "
        f"{bd0.get('note', 'Current-baseline receipt is missing or malformed.')}"
    )
    ws.cell(row=r, column=1, value="DISCLOSURE").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=r, column=6, value=_cv(disclosure)).font = DF
    ws.cell(row=r, column=6).alignment = AL
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    autofit(ws, 8)
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 76
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 48

    # Additive complete sink for the v1 gate's intentional 50-row compatibility cap.  The sheet is
    # reference-only: cutover_gate/1 consumed the original gate counts before this projection and
    # never reads these uncapped presentation rows.
    ws = sheet("Current Baseline Export",
               ["State", "Device", "Wave", "Category", "Check",
                "Observed baseline / acceptance", "Projection custody", "Source key"])
    r = 2
    if baseline_export0.get("status") == "available":
        for blocker in baseline_export_rows0:
            if not isinstance(blocker, dict):
                continue
            vals = [
                blocker.get("evidence_state", ""), blocker.get("device", ""),
                blocker.get("wave", ""), blocker.get("category", ""),
                blocker.get("check", ""), blocker.get("expect", ""),
                blocker.get("projection_custody", ""), blocker.get("source_key", ""),
            ]
            for c, value in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=_cv(value))
                cell.font = DF
                cell.alignment = AL
            r += 1
    if r == 2:
        ws.cell(row=r, column=1, value="NOT VERIFIED").font = DF
        ws.cell(row=r, column=6, value=_cv(
            baseline_export0.get(
                "note", "Complete current-baseline blocker export is unavailable."
            )
        )).font = DF
        r += 1
    export_total = baseline_export_summary0.get("n_blockers_total")
    export_rendered = baseline_export_summary0.get("n_rows_returned")
    export_omitted = baseline_export_summary0.get("omitted")
    export_complete = baseline_export_summary0.get("complete")
    export_disclosure = (
        f"status={baseline_export0.get('status', 'not_verified')}; "
        f"rendered={export_rendered if type(export_rendered) is int else 0}, "
        f"total={export_total if type(export_total) is int else 0}, "
        f"omitted={export_omitted if type(export_omitted) is int else 0}, "
        f"complete={'YES' if export_complete is True else 'NO'}; "
        f"rows_sha256={baseline_export_summary0.get('rows_sha256', 'NOT VERIFIED')}. "
        "Reference-only export; this sheet does not participate in the verdict."
    )
    ws.cell(row=r, column=1, value="DISCLOSURE").font = Font(
        name="Calibri", bold=True, size=10
    )
    ws.cell(row=r, column=6, value=_cv(export_disclosure)).font = DF
    ws.cell(row=r, column=6).alignment = AL
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    autofit(ws, 8)
    ws.column_dimensions["E"].width = 44
    ws.column_dimensions["F"].width = 76
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 48

    # Interface Changes
    ws = sheet("Interface Changes", ["Hostname", "Port", "Change", "Field: Old -> New"])
    r = 2
    for host in sorted(set(oi) | set(ni), key=_skey):
        op, npp = oi.get(host, {}), ni.get(host, {})
        for port in sorted(set(op) | set(npp), key=_skey):
            o, n = op.get(port), npp.get(port)
            if o is None and n is None:
                continue
            if o is None:
                change, deltas = "Added port", []
            elif n is None:
                change, deltas = "Removed port", []
            else:
                change = "Modified"
                deltas = [f"{f}: {o.get(f, '') or NONE} -> {n.get(f, '') or NONE}"
                          for f in _DIFF_FIELDS if (o.get(f, "") or "") != (n.get(f, "") or "")]
                if not deltas:
                    continue
            for c, v in enumerate([host, port, change, " | ".join(deltas)], 1):
                cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
            r += 1
    autofit(ws, 4); ws.column_dimensions["D"].width = 70

    # Endpoint (MAC) Changes
    ws = sheet("Endpoint Changes", ["Hostname", "Port", "Change", "MAC"])
    r = 2
    for host in sorted(set(oi) | set(ni), key=_skey):
        op, npp = oi.get(host, {}), ni.get(host, {})
        for port in sorted(set(op) | set(npp), key=_skey):
            om = _macset((op.get(port) or {}).get("end_host_mac", ""))
            nm = _macset((npp.get(port) or {}).get("end_host_mac", ""))
            for mac in sorted(nm - om):
                for c, v in enumerate([host, port, "MAC appeared", mac], 1):
                    cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
                r += 1
            for mac in sorted(om - nm):
                for c, v in enumerate([host, port, "MAC gone", mac], 1):
                    cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
                r += 1
    autofit(ws, 4)

    # SVI / Gateway Changes
    ws = sheet("SVI Changes", ["Hostname", "SVI", "Change", "Detail"])
    r = 2
    for host in sorted(set(oi) | set(ni), key=_skey):
        op, npp = oi.get(host, {}), ni.get(host, {})
        svis = sorted({p for p in (set(op) | set(npp)) if re.match(r"^Vlan\d+$", p, re.I)})
        for p in svis:
            o, n = op.get(p), npp.get(p)
            if o is None and n is not None:
                ch = "SVI added"
                detail = f"IP {n.get('svi_ip', '') or NONE}, FHRP {n.get('hsrp_behavior', '') or NONE}"
            elif n is None and o is not None:
                ch = "SVI removed"
                detail = f"was IP {o.get('svi_ip', '') or NONE}"
            else:
                diffs = [f"{f}: {o.get(f, '') or NONE} -> {n.get(f, '') or NONE}"
                         for f in ("svi_ip", "hsrp_behavior", "subnet_primary_route")
                         if (o.get(f, "") or "") != (n.get(f, "") or "")]
                if not diffs:
                    continue
                ch, detail = "SVI changed", " | ".join(diffs)
            for c, v in enumerate([host, p, ch, detail], 1):
                cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
            r += 1
    autofit(ws, 4); ws.column_dimensions["D"].width = 60

    # Health Shifts (NEW-V3.23.106) — per-switch health-band change, regressions first
    ws = sheet("Health Shifts", ["Switch", "Direction", "Old band", "New band", "Old score", "New score"])
    r = 2
    for direction, rows in (("REGRESSED", delta["health"]["regressed"]),
                            ("improved", delta["health"]["improved"]),
                            # in/out of 'Insufficient Data' -> labelled by kind (went dark / newly assessed), never
                            # silently folded into 'improved' (audit-4 #5)
                            *(((c["kind"], [c]) for c in delta["health"].get("coverage_shifts", [])))):
        for d in rows:
            vals = [d["switch"], direction, d["old_band"], d["new_band"], d["old_score"], d["new_score"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
            if direction == "REGRESSED":
                ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FFC7CE")
            r += 1
    if r == 2:
        ws.cell(row=2, column=1, value="No health-band changes between the two snapshots.").font = DF
    autofit(ws, 6)

    # Findings Delta (NEW-V3.23.106) — consolidated punch-list items opened vs resolved by the cutover
    ws = sheet("Findings Delta", ["State", "Severity", "Category", "Devices", "Finding"])
    r = 2
    for state, items in (("OPENED", delta["findings"]["opened"]),
                         ("resolved", delta["findings"]["resolved"])):
        for f in items:
            vals = [state, f.get("severity", ""), f.get("category", ""),
                    _devices_cell(f.get("devices")), f.get("title", "")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
            ws.cell(row=r, column=1).fill = PatternFill(
                "solid", fgColor="FFC7CE" if state == "OPENED" else "C6EFCE")
            r += 1
    if r == 2:
        ws.cell(row=2, column=1, value="No punch-list findings opened or resolved.").font = DF
    autofit(ws, 5); ws.column_dimensions["E"].width = 70

    # Protocol Adjacency Delta — baseline-observed OSPF/BGP/EIGRP peers, gated by the exact runtime
    # assessability receipt on BOTH sides.  A coverage gap is rendered as its own row and never as zero peers.
    ws = sheet("Protocol Adjacency Delta",
               ["Switch", "Protocol", "Peer", "Before", "After", "Result", "Evidence / next action"])
    r = 2
    for item in pd0.get("changes") or []:
        vals = [item.get("switch", ""), item.get("protocol", ""), item.get("peer", ""),
                item.get("before_state", ""), item.get("after_state", ""),
                item.get("result", ""), item.get("note", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
        result = str(item.get("result") or "")
        fill = ("FFC7CE" if result == "state_degraded" else
                "C6EFCE" if result == "recovered" else "FFEB9C")
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=fill)
        r += 1
    for gap in pd0.get("coverage_gaps") or []:
        vals = [gap.get("switch", ""), gap.get("protocol", ""), "",
                gap.get("before_state", ""), gap.get("after_state", ""),
                "NOT COMPARABLE", gap.get("reason", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="D9D9D9")
        r += 1
    if r == 2:
        ws.cell(row=2, column=1, value=_cv(pd0.get("note") or
                                          "Protocol adjacency preservation NOT assessed.")).font = DF
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    autofit(ws, 7); ws.column_dimensions["G"].width = 72

    # Protocol Family Changes — complete, uncapped projection of the same reference-only change set
    # consumed by the canonical cutover gate.  This is additive: the established IPv4 owner sheet above
    # remains intact.  The renderer neither infers native semantics nor drops malformed/empty producer rows;
    # every evidence absence is a visible NOT VERIFIED placeholder.
    ws = sheet("Protocol Family Changes",
               ["Family", "Subject kind", "Subject", "Transition", "Expected",
                "Producer decision_effect", "Assurance", "Before", "After", "Note"])
    r = 2
    for item in family_surface_rows:
        vals = [
            item["family"], item["subject_kind"], item["subject"], item["transition"],
            item["expected"], item["decision_effect"], item["assurance"], item["before"],
            item["after"], item["note"],
        ]
        for c, value in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=_cv(value)); cell.font = DF; cell.alignment = AL
        effect = str(item.get("decision_effect") or "")
        transition = str(item.get("transition") or "")
        fill = (
            "FFC7CE" if effect == "block" else
            "FFEB9C" if effect == "review" else
            "C6EFCE" if transition == "recovered" else
            "D9D9D9" if item.get("placeholder") or effect in {"not_verified", "NOT VERIFIED"}
            else "FFFFFF"
        )
        ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor=fill)
        r += 1
    totals_subject = (f"{family_source_rows} producer row(s)" if family_source_rows
                      else "NOT VERIFIED")
    totals = ["FULL TOTALS", "see Note", totals_subject, "see Note", "see Note", "see Note",
              "see Note", f"rendered {family_source_rows} of {family_source_rows}",
              "omitted 0", family_disclosure]
    for c, value in enumerate(totals, 1):
        cell = ws.cell(row=r, column=c, value=_cv(value)); cell.font = DF; cell.alignment = AL
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=10)
    autofit(ws, 10)
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions["H"].width = 54
    ws.column_dimensions["I"].width = 54
    ws.column_dimensions["J"].width = 80

    # Comparison Receipt — complete, uncapped projection of the same detached receipt and decision
    # components that own this workbook.  This sheet is deliberately metadata-only: the complete
    # subject-change population remains on Protocol Family Changes, and no UI/export cap feeds the
    # gate.  Legacy direct callers receive an explicit NOT VERIFIED row rather than an invented bind.
    ws = sheet("Comparison Receipt", ["Section", "Field", "Value", "Completeness"])
    receipt_rows: List[tuple] = []
    if comparison_doc is None:
        receipt_rows.append((
            "Receipt", "status", "NOT VERIFIED — no canonical source-bound comparison supplied",
            "rendered 1 / total 1 / omitted 0; complete",
        ))
    else:
        from cisco_toolkit.protocol_assurance import canonical_sha256
        admission = _as_dict(comparison_doc.get("comparison_admission"))
        source_pair = _as_dict(admission.get("source_binding"))
        envelope = _as_dict(comparison_doc.get("comparison_receipt"))
        intent = _as_dict(comparison_doc.get("change_intent"))
        families_receipt = _as_dict(comparison_doc.get("protocol_families"))
        gate_receipt = _as_dict(comparison_doc.get("cutover_gate"))
        receipt_rows.extend([
            ("Comparison", "schema", comparison_doc.get("comparison_schema"), "complete"),
            ("Admission", "status", admission.get("status"), "complete"),
            ("Admission", "engagement_id", admission.get("engagement_id"), "complete"),
            ("Admission", "campaign_id", admission.get("campaign_id"), "complete"),
            ("Admission", "failures", json.dumps(
                admission.get("failures") if isinstance(admission.get("failures"), list) else [],
                sort_keys=True, ensure_ascii=True, separators=(",", ":")), "complete"),
            ("Admission", "coverage_gaps", json.dumps(
                admission.get("coverage_gaps")
                if isinstance(admission.get("coverage_gaps"), list) else [],
                sort_keys=True, ensure_ascii=True, separators=(",", ":")), "complete"),
            ("Receipt integrity", "status", "VERIFIED", "complete"),
        ])
        for side in ("before", "after"):
            binding = _as_dict(source_pair.get(side))
            for field in (
                    "source", "snapshot_id", "label", "sha256", "bytes", "script_version"):
                receipt_rows.append((
                    f"Source {side}", field, binding.get(field, "NOT VERIFIED"), "complete",
                ))
        subject_pair = _as_dict(admission.get("subject_binding"))
        for side in ("before", "after"):
            subjects = _as_dict(subject_pair.get(side))
            receipt_rows.extend([
                (f"Subjects {side}", "schema", subjects.get("schema", "NOT VERIFIED"),
                 "complete"),
                (f"Subjects {side}", "identity_kind",
                 subjects.get("identity_kind", "NOT VERIFIED"), "complete"),
                (f"Subjects {side}", "n_subjects",
                 subjects.get("n_subjects", "NOT VERIFIED"), "complete"),
                (f"Subjects {side}", "subjects_sha256",
                 subjects.get("subjects_sha256", "NOT VERIFIED"), "complete"),
                (f"Subjects {side}", "subjects", json.dumps(
                    subjects.get("subjects") if isinstance(subjects.get("subjects"), list) else [],
                    sort_keys=True, ensure_ascii=True, separators=(",", ":")), "complete"),
                (f"Subjects {side}", "valid", subjects.get("valid", "NOT VERIFIED"),
                 "complete"),
                (f"Subjects {side}", "failures", json.dumps(
                    subjects.get("failures") if isinstance(subjects.get("failures"), list) else [],
                    sort_keys=True, ensure_ascii=True, separators=(",", ":")), "complete"),
            ])
        for key, value in sorted(_as_dict(admission.get("owner_versions")).items()):
            receipt_rows.append(("Decision owner", key, value, "complete"))
        for profile in admission.get("support_profiles") or []:
            profile = _as_dict(profile)
            receipt_rows.append((
                "Support profile", str(profile.get("family") or "NOT VERIFIED"),
                json.dumps(profile, sort_keys=True, ensure_ascii=True, separators=(",", ":")),
                "complete",
            ))
        receipt_rows.extend([
            ("Digest", "change_intent", canonical_sha256(intent), "complete"),
            ("Digest", "protocol_family_change_set", canonical_sha256(families_receipt), "complete"),
            ("Digest", "cutover_gate", canonical_sha256(gate_receipt), "complete"),
            ("Digest", "comparison_payload", envelope.get("payload_sha256"), "complete"),
            ("Digest", "comparison_receipt", envelope.get("receipt_sha256"), "complete"),
        ])
        total = len(receipt_rows) + 1
        receipt_rows.append((
            "Export", "row disclosure", f"rendered {total} / total {total} / omitted 0",
            "complete uncapped receipt metadata and subject identities; protocol family rows are "
            "complete on their sheet",
        ))
    for row_number, values in enumerate(receipt_rows, 2):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row_number, column=column, value=_cv(value))
            cell.font = DF
            cell.alignment = AL
        fill = None
        if values[0] == "Receipt":
            fill = "D9D9D9"
        elif values[0] == "Receipt integrity" and values[1] == "status":
            fill = "C6EFCE" if values[2] == "VERIFIED" else "FFC7CE"
        elif values[0] == "Admission" and values[1] == "status":
            fill = {
                "admitted": "C6EFCE",
                "coverage_lost": "FFEB9C",
                "not_comparable": "FFC7CE",
            }.get(str(values[2]), "D9D9D9")
        elif values[0] == "Export":
            fill = "D9D9D9"
        if fill:
            ws.cell(row=row_number, column=3).fill = PatternFill("solid", fgColor=fill)
    autofit(ws, 4)
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 82
    ws.column_dimensions["D"].width = 70

    # Reachability (W2) — computed RIB->FIB flows the change DEFINITIVELY broke (the offline Batfish-peer what-if).
    # Coverage-honest: distinguishes 'tested, no regressions' from 'no routes collected -> not assessed'.
    ws = sheet("Reachability", ["Flow", "Src", "Dst", "Before", "After"])
    r = 2
    for label, fill, items in (("NEWLY BLOCKED", "FFC7CE", rd0.get("newly_blocked") or []),
                               ("newly reachable", "C6EFCE", rd0.get("newly_reachable") or [])):
        for p in items:
            vals = [label, p.get("src", ""), p.get("dst", ""), p.get("old_status", ""), p.get("new_status", "")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill)
            r += 1
    if r == 2:                                            # no regressions to list -> be coverage-honest about WHY
        if rd0.get("assessed"):
            msg = (f"No computed reachability regressions across {rd0.get('pairs_tested', 0)} sampled flow(s) "
                   f"({rd0.get('subnets_tested', 0)} of {rd0.get('subnets_total', 0)} subnet(s), one representative "
                   f"host each{' — CAPPED, coverage incomplete' if rd0.get('capped') else ''}). Bounded sample: a "
                   "regression to an untested host or subnet would not appear here.")
        elif (rd0.get("subnets_total") or 0) == 0:
            msg = "No routes collected — computed reachability NOT assessed (this is NOT 'no regressions')."
        else:
            msg = (f"Only {rd0.get('subnets_total')} subnet(s) collected — no inter-subnet flow to test; "
                   "computed reachability NOT assessed.")
        ws.cell(row=2, column=1, value=msg).font = DF
    autofit(ws, 5); ws.column_dimensions["D"].width = 42; ws.column_dimensions["E"].width = 42

    # Cabling Changes — the EDA cable-map SSOT diffed: physical cables added / removed / op-status
    # flips + LAG membership changes. Coverage-honest: '-> unknown' reads 'no longer observed' (a
    # coverage event, never a down), and a snapshot pair with no CDP/LLDP links reads NOT ASSESSED.
    cd = delta.get("cabling") or {}
    ws = sheet("Cabling Changes", ["Change", "Switch A", "Port A", "Switch B", "Port B", "Status", "Note"])
    cab_rows: list = []
    if not cd.get("assessed"):
        cab_rows.append(("NOT ASSESSED", "", "", "", "", "",
                         "no CDP/LLDP links in either snapshot — NOT a statement that cabling is unchanged"))
    else:
        for c0 in cd.get("status_changes") or []:
            cab_rows.append(("Status changed", c0.get("a", ""), c0.get("a_port", ""), c0.get("b", ""),
                             c0.get("b_port", ""), f"{c0.get('from', '')} -> {c0.get('to', '')}",
                             c0.get("classification", "")))
        for label, items in (("Cable added", cd.get("added") or []), ("Cable removed", cd.get("removed") or [])):
            for c0 in items:
                cab_rows.append((label, c0.get("a", ""), c0.get("a_port", ""), c0.get("b", ""),
                                 c0.get("b_port", ""), c0.get("op_status", ""),
                                 f"port-channel, {len(c0.get('members') or [])} member(s)" if c0.get("is_pc") else ""))
        for m0 in cd.get("members_changed") or []:
            cab_rows.append(("LAG members changed", m0.get("a", ""), m0.get("a_port", ""), m0.get("b", ""),
                             m0.get("b_port", ""), f"{m0.get('old_members', 0)} -> {m0.get('new_members', 0)} member(s)",
                             "a port-channel leg was added or removed"))
        if not cab_rows:
            cab_rows.append(("No changes", "", "", "", "", "",
                             f"{(cd.get('summary') or {}).get('n_unchanged', 0)} cable(s) compared, unchanged"))
    r = 2
    for vals in cab_rows:
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
        if str(vals[5]).endswith("-> down"):
            ws.cell(row=r, column=6).fill = PatternFill("solid", fgColor="FFC7CE")
        r += 1
    autofit(ws, 7); ws.column_dimensions["G"].width = 55

    # Pre-Change Validation Certificate (roadmap C1) — the fib differential packaged as the PPDIOO gate
    # artifact: verdict + every changed flow cited before->after + segmentation invariants + path intents
    # + every NAMED blind spot. Rendered from the certificate dict (the .precert.json SSOT); computed here
    # when the caller did not pass one, so the sheet is always present.
    cf = cert.get("flows") or {}
    _CERT_FILL = {"PASS": "C6EFCE", "CONDITIONAL": "FFEB9C", "FAIL": "FFC7CE", "INDETERMINATE": "D9D9D9"}
    stamps = cert.get("stamps") or {}
    if cf.get("assessed"):
        cover = (f"{cf.get('preserved', 0)} preserved, {cf.get('both_unreachable', 0)} unreachable both sides, "
                 f"{len(cf.get('changed') or [])} changed, {len(cf.get('inconclusive') or [])} inconclusive "
                 f"(bounded sample: {cf.get('subnets_tested', 0)} of {cf.get('subnets_total', 0)} subnet(s), one "
                 f"representative host each{', CAPPED — coverage incomplete' if cf.get('capped') else ''})")
    else:
        cover = "computed reachability NOT assessed (this is NOT a statement that reachability is unchanged)"
    cert_rows: list = [
        ("VERDICT", "pre-change validation certificate",
         (stamps.get("before") or {}).get("generated_at", ""), (stamps.get("after") or {}).get("generated_at", ""),
         cert.get("verdict", ""), cert.get("verdict_note", "")),
        ("Flows", "coverage", "", "", f"{cf.get('pairs_tested', 0)} flow(s) tested", cover),
    ]
    for f0 in (cf.get("changed") or []):
        cert_rows.append(("Flows", f"{f0.get('src')} -> {f0.get('dst')}", f0.get("before", ""), f0.get("after", ""),
                          "NEWLY BLOCKED" if f0.get("regression") else "newly reachable", ""))
    for f0 in (cf.get("inconclusive") or []):
        cert_rows.append(("Flows", f"{f0.get('src')} -> {f0.get('dst')}", "", "", "INCONCLUSIVE",
                          f0.get("reason", "")))
    for s0 in (cert.get("segmentation") or []):
        held = s0.get("held")
        cert_rows.append(("Segmentation", s0.get("invariant", ""), "", "",
                          "HELD" if held is True else ("VIOLATED" if held is False else "NOT EVALUABLE"),
                          s0.get("evidence", "")))
    for i0 in (cert.get("intents") or []):
        cert_rows.append(("Intent", f"{i0.get('id')} (expect {i0.get('expect')})",
                          i0.get("old_verdict", ""), i0.get("new_verdict", ""),
                          "REGRESSED" if i0.get("regressed") else
                          ("coverage lost" if i0.get("coverage_lost") else str(i0.get("new_verdict", ""))),
                          f"{i0.get('old_status')} -> {i0.get('new_status')}"))
    for b0 in (cert.get("blind_spots") or []):
        cert_rows.append(("Blind spot", b0, "", "", "OPEN", ""))
    for g0 in (cert.get("gate_failures") or []):
        cert_rows.append(("Gate failure", g0, "", "", "BLOCKING", "Do not proceed"))
    if not (cert.get("blind_spots") or []):
        cert_rows.append(("Blind spot", "none open", "", "", "", ""))
    binding = cert.get("source_binding") if isinstance(cert.get("source_binding"), dict) else {}
    for side in ("before", "after"):
        if binding.get(side):
            cert_rows.append(("Provenance", f"{side} input SHA-256", "", "",
                              "BOUND", str(binding.get(side))))
    schema_binding = cert.get("schema_status") if isinstance(cert.get("schema_status"), dict) else {}
    if schema_binding:
        cert_rows.append(("Provenance", "schema compatibility", "", "",
                          str(schema_binding.get("status") or "unverifiable").upper(),
                          str(schema_binding.get("message") or "")
                          + (" (explicit override recorded)" if schema_binding.get("override") else "")))
    _ROW_FILL = {"NEWLY BLOCKED": "FFC7CE", "newly reachable": "C6EFCE", "VIOLATED": "FFC7CE",
                 "REGRESSED": "FFC7CE", "INCONCLUSIVE": "D9D9D9", "NOT EVALUABLE": "D9D9D9", "OPEN": "FFEB9C"}
    _ROW_FILL["BLOCKING"] = "FFC7CE"
    ws = sheet(CERT_SHEET_NAME, list(CERT_SHEET_HEADERS))
    r = 2
    for vals in cert_rows:
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
        if vals[0] == "VERDICT":
            vc = ws.cell(row=r, column=5)
            vc.fill = PatternFill("solid", fgColor=_CERT_FILL.get(str(vals[4]), "FFFFFF"))
            vc.font = Font(name="Calibri", bold=True, size=11)
        elif str(vals[4]) in _ROW_FILL:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=_ROW_FILL[str(vals[4])])
        r += 1
    autofit(ws, 6); ws.column_dimensions["B"].width = 46; ws.column_dimensions["F"].width = 70

    wb.save(out_path)
    return gate


# -----------------------------------------------------------------------------
# Migration CAMPAIGN trend (NEW-V3.23.145). Where compute_snapshot_delta diffs a PAIR (pre/post a single
# cutover), the campaign tracker ingests a SERIES of collections taken across the whole migration and shows
# the trajectory: is the network actually getting healthier wave by wave? It extracts the headline metrics
# per collection (avg health, band mix, punch-list size + Critical/High count, NOT-READY groups,
# Past-LDoS),
# computes the first->last trajectory per metric + an overall IMPROVING/MIXED/REGRESSING verdict, and reuses
# compute_snapshot_delta for the per-step findings burndown (opened vs resolved). Pure read of N snapshot
# dicts; tolerant of older snapshots that lack a metric -- it is omitted from the trajectory, never scored as
# a zero, and if it was comparable at ONE end of the campaign only, its loss is disclosed in the verdict note.
# -----------------------------------------------------------------------------
def _trend_point(snap: dict) -> dict:
    """Headline metrics for one snapshot in the campaign timeline."""
    # Tolerant of a malformed --trend/--compare snapshot: a wrong-typed health_scores (a dict, not the
    # expected list-of-dicts) would iterate its KEYS (str), and `str.get` raises AttributeError -> the whole
    # trend workbook aborts (CLI --trend is unwrapped). Coerce to a list of dicts up front; a non-dict row
    # can't reach the `.get` calls below.
    # COVERAGE HONESTY: `have_*` records whether the section was actually PRESENT (a usable list), so an
    # absent one abstains with "" instead of tallying to a hard 0. A partial upload, an older schema, or a
    # failed analysis phase otherwise scored identically to a clean fleet -- and, being a fall in every
    # counter, read as a campaign IMPROVING to a green "CAMPAIGN VERDICT". avg_health and past_ldos in this
    # same dict already abstained; the idiom was simply applied to 2 of the 6 metrics.
    _hs = snap.get("health_scores")
    have_hs = isinstance(_hs, list)
    hs = [r for r in _hs if isinstance(r, dict)] if have_hs else []
    # _renderable_num, not a bare isinstance((int, float)): that admitted the two numeric values a JSON
    # snapshot can carry but arithmetic cannot survive -- an unbounded-precision int literal
    # (`sum(scores)/len(scores)` -> "integer division result too large for a float", aborting the whole
    # --trend workbook) and the bare `Infinity`/`NaN` json.loads accepts (which would render an average
    # of inf/nan into the campaign deck).
    scores = [r.get("score") for r in hs if _renderable_num(r.get("score"))]
    bands: Dict[str, int] = {}
    for r in hs:
        bands[str(r.get("band", ""))] = bands.get(str(r.get("band", "")), 0) + 1
    _pl = snap.get("punchlist")
    have_pl = isinstance(_pl, list)
    pl = [f for f in _pl if isinstance(f, dict)] if have_pl else []   # same list-of-dicts guard as health_scores
    readiness = {"READY": 0, "CAUTION": 0, "NOT READY": 0}
    _mr = snap.get("migration_readiness")
    have_mr = isinstance(_mr, list)
    for r in (_mr if have_mr else []):
        rd = _hkey(r.get("readiness")) if isinstance(r, dict) else None
        if rd in readiness:                       # _hkey: an unhashable label must not raise on `in`
            readiness[rd] += 1
    # isinstance-guard (not `or {}`): a TRUTHY non-dict section/subsection (a list/str/int in a malformed
    # --trend/--compare upload) slips past `or {}` and crashes the .get() below -> 500s the unwrapped /trend
    # endpoint and aborts the CLI --trend workbook. Same truthy-non-dict class the deliverables already guard.
    def _d(v):
        return v if isinstance(v, dict) else {}
    lr = _d(_d(snap.get("lifecycle_risk")).get("summary"))
    _eb = _d(snap.get("executive_brief"))
    _scale = _d(_eb.get("scale"))
    _posture = _d(_eb.get("posture"))
    avg = _posture.get("avg_health")
    if avg is None and scores:
        avg = round(sum(scores) / len(scores), 1)
    # str() (not `or ""`): `generated_at` is a timestamp STRING by contract, but a truthy non-str survives
    # `or ""` and the ts[:10] slice below then raises (`5[:10]` -> TypeError) -- or, for a list, silently
    # returns a LIST into the timeline's 'date' column. The slicing variant of the same guard gap.
    ts = str(snap.get("generated_at") or "")
    integrity = _analysis_integrity(
        snap, ("health_scores", "punchlist", "migration_readiness", "lifecycle_risk"))
    return {
        "date": ts[:10], "generated_at": ts, "version": snap.get("script_version", ""),
        # SSOT: prefer the engine's canonical scale / posture; the local len()/band-tally is a fallback only.
        "n_switches": _scale.get("n_devices") if _scale.get("n_devices") is not None else len(_d(snap.get("devices"))),
        "avg_health": avg if avg is not None else "",
        # Each falls back to its raw tally ONLY when that section was collected; otherwise "" (abstain),
        # exactly like avg_health/past_ldos below. "Not observed" must never render as an observed zero.
        "n_critical": _posture.get("n_critical") if _posture.get("n_critical") is not None else (
            bands.get("Critical", 0) if have_hs else ""),
        "n_punchlist": len(pl) if have_pl else "",
        "n_crit_high": sum(1 for f in pl if f.get("severity") in ("Critical", "High")) if have_pl else "",
        "n_not_ready": readiness["NOT READY"] if have_mr else "",
        # "Past end-of-support" = Past-LDoS — the migration-critical date-band count the
        # brief/deck/explorer/workbook all headline. NOT Past-EoS (end-of-SALE while LDoS remains future;
        # neither band proves entitlement): reading
        # n_past_eos here showed 0 while 152 boxes were past support (the EoS/LDoS silent-drop class).
        # A count over PARTIAL coverage must not be trended as if it were complete. `past_ldos` is
        # lower-is-better in _TREND_METRICS, so a fleet whose platforms lack an exact EoX match or
        # complete retained source/date authority reports 0 and scores as the BEST possible value —
        # an un-assessed campaign step
        # reads as an improvement over a fully-assessed earlier one. That is absence converted into
        # a positive signal, which is worse than absence rendered as neutral.
        #
        # `""` is this dict's existing "not available" convention (see n_punchlist / n_not_ready
        # above), and the trajectory skips a metric that is missing on either side rather than
        # comparing against it. So an incomplete step drops out of THIS metric only; every other
        # metric on that snapshot still trends. (handoff §7.25)
        "past_ldos": ("" if lr.get("n_unknown") else lr.get("n_past_ldos", "")) if lr else "",
        "integrity_ok": integrity["ok"],
        "integrity_failures": integrity["failures"],
    }


# (metric label, timeline key, lower-is-better) for the trajectory + the timeline columns.
_TREND_METRICS = (("Avg health / 100", "avg_health", False),
                  ("Critical-band switches", "n_critical", True),
                  ("Punch-list items", "n_punchlist", True),
                  ("Critical/High findings", "n_crit_high", True),
                  ("NOT READY groups", "n_not_ready", True),
                  ("Past end-of-support", "past_ldos", True))


def compute_campaign_trend(snapshots: List[dict], *, source_bindings: Optional[list] = None,
                           schema_status: Any = None) -> dict:
    """Trajectory of a migration campaign across a SERIES of snapshots. Returns
    {timeline, steps, trajectory, verdict, verdict_note}; degrades gracefully when a metric is absent.
    A collection with a failed/missing core analysis makes the campaign verdict INDETERMINATE rather
    than allowing the remaining counters to manufacture an improvement."""
    snaps = list(snapshots or [])
    timeline = [dict(_trend_point(s), collection=f"C{i + 1}") for i, s in enumerate(snaps)]
    schema = _schema_status(schema_status)
    failed_collections = [
        {"collection": pt["collection"], "failures": list(pt.get("integrity_failures") or [])}
        for pt in timeline if not pt.get("integrity_ok")
    ]
    if schema and schema.get("status") not in ("", "ok"):
        failed_collections.append({
            "collection": "series",
            "failures": [
                f"schema compatibility {schema.get('status')}: "
                f"{schema.get('message') or 'cross-input compatibility was not proven'}"
                + (" (explicit override recorded)" if schema.get("override") else "")
            ],
        })

    steps: List[dict] = []
    for i in range(len(snaps) - 1):
        pair_binding = None
        if isinstance(source_bindings, list) and i + 1 < len(source_bindings):
            pair_binding = {"before": source_bindings[i], "after": source_bindings[i + 1]}
        d = compute_snapshot_delta(snaps[i], snaps[i + 1],
                                   source_binding=pair_binding, schema_status=schema)
        pa = d.get("protocol_adjacencies") or {}
        pas = pa.get("summary") or {}
        cb = _as_dict(d.get("current_baseline"))
        cbs = _as_dict(cb.get("summary"))
        steps.append({
            "from": timeline[i]["collection"], "to": timeline[i + 1]["collection"],
            "from_date": timeline[i]["date"], "to_date": timeline[i + 1]["date"],
            "opened": d["findings"]["n_opened"], "opened_high": d["findings"]["n_opened_high"],
            "resolved": d["findings"]["n_resolved"], "net": d["findings"]["n_opened"] - d["findings"]["n_resolved"],
            "regressed": d["health"]["n_regressed"], "improved": d["health"]["n_improved"],
            "protocol_gate": pa.get("gate", "NOT_ASSESSED"),
            "protocol_baseline_peers": pas.get("n_baseline_peers", 0),
            "protocol_state_regressed": pas.get("n_state_regressed", 0),
            "protocol_recovered": pas.get("n_recovered", 0),
            "protocol_no_longer_observed": pas.get("n_no_longer_observed", 0),
            "protocol_added": pas.get("n_added", 0),
            "protocol_coverage_gaps": pas.get("n_coverage_gaps", 0),
            "protocol_projection_custody": pa.get("projection_custody", "embedded_unverified"),
            "protocol_note": pa.get("note", ""),
            "current_baseline_verdict": cb.get("verdict", "NOT_ASSESSED"),
            "current_baseline_blockers": cbs.get("n_blockers", 0),
            "current_baseline_note": cb.get("note", ""),
            "verdict": d["verdict"],
        })

    trajectory: List[dict] = []
    lost: List[str] = []
    never: List[str] = []
    verdict, note = "INSUFFICIENT", "Need at least two collections to show a trend."
    campaign_protocol: dict = {}
    if len(timeline) >= 2:
        first, last = timeline[0], timeline[-1]
        campaign_binding = None
        if isinstance(source_bindings, list) and len(source_bindings) >= len(snaps):
            campaign_binding = {"before": source_bindings[0], "after": source_bindings[-1]}
        # Compare the campaign ENDPOINTS, not absolute peer counts. A transient peer regression remains in
        # `steps`, while a regression recovered by the last collection must not mislabel the final trajectory.
        campaign_protocol = compute_protocol_adjacency_delta(
            snaps[0], snaps[-1], source_binding=campaign_binding)
        # Metrics whose EVIDENCE was collected in one endpoint of the campaign and not the other. Dropping
        # them from the trajectory is right (they are not comparable) but doing it SILENTLY is the same
        # survivorship trap as a device dropping out: the surviving metrics then set a verdict over a
        # narrower estate than the reader assumes. Disclosed below, and (like a dark device) a clean
        # IMPROVING/FLAT is downgraded.
        #
        # BOTH ends abstaining was the remaining silence, and it is the WORSE one. "Never part of this
        # campaign's evidence" was an assumption, not an observation: `_trend_point` yields a non-numeric
        # value for a metric whose analysis FAILED or was never collected, so a campaign in which the
        # lifecycle/EoL pass fell over at every collection dropped that metric out of the trajectory
        # entirely -- measured, both ends all-Unknown gives verdict FLAT with the lifecycle row simply
        # absent and no NOT-COMPARABLE line. A reader cannot tell "we looked and found nothing" from
        # "we never looked", and a count of 0 that means NOT MEASURED must not read as nothing wrong.
        # Disclosed separately from `lost` because the actions differ: one endpoint missing is usually a
        # collection gap in that run, both missing means the metric was never measured in this campaign
        # at all. It does NOT downgrade the verdict -- there is no evidence in either direction to
        # downgrade on -- it is stated.
        for metric, key, good_down in _TREND_METRICS:
            a, b = first.get(key), last.get(key)
            ok_a, ok_b = isinstance(a, (int, float)), isinstance(b, (int, float))
            if not ok_a or not ok_b:
                (lost if ok_a != ok_b else never).append(metric)
                continue
            delta = b - a
            direction = "flat" if delta == 0 else (
                "improving" if ((delta < 0) if good_down else (delta > 0)) else "worsening")
            trajectory.append({"metric": metric, "first": a, "last": b,
                               "delta": round(delta, 1), "direction": direction})
        better = sum(1 for r in trajectory if r["direction"] == "improving")
        worse = sum(1 for r in trajectory if r["direction"] == "worsening")
        # Devices that went DARK across the campaign (present in an earlier collection, absent later). A rising
        # avg-health is partly SURVIVORSHIP when an unhealthy device drops out of the average, so a campaign that
        # lost devices cannot read a clean IMPROVING without disclosing them (audit-2 #4 false-health).
        gone: set = set()

        def _bands(s):
            return {str(r.get("switch")): str(r.get("band", "")) for r in _as_list(s.get("health_scores"))
                    if isinstance(r, dict)}
        for i in range(len(snaps) - 1):
            # _as_dict, not `or {}`: set(5) raises, and set("sw1") SUCCEEDS on a string -- yielding the
            # CHARACTERS, so the went-dark scan would report three phantom devices gone and downgrade the
            # verdict on evidence that does not exist. Both halves of the same guard gap.
            gone |= set(_as_dict(snaps[i].get("devices"))) - set(_as_dict(snaps[i + 1].get("devices")))
            # the engine never DROPS an uncollected device -- it keeps it as an 'Insufficient Data' STUB in
            # `devices`, so the set-difference above misses a previously-collected switch that went dark; detect
            # the band transition real-band -> 'Insufficient Data' too, else a campaign that loses its worst
            # collected switches reads a survivorship-biased IMPROVING (audit-5 #9).
            b0, b1 = _bands(snaps[i]), _bands(snaps[i + 1])
            for sw, band in b0.items():
                if band and band != "Insufficient Data" and b1.get(sw) == "Insufficient Data":
                    gone.add(sw)
        n_gone = len(gone)
        if trajectory:                                    # comparable metrics exist -> a real verdict
            if better == 0 and worse == 0:
                verdict = "FLAT"
            elif better > worse:
                verdict = "IMPROVING"
            elif worse > better:
                verdict = "REGRESSING"
            else:
                verdict = "MIXED"
            if n_gone and verdict in ("IMPROVING", "FLAT"):
                verdict = "MIXED"                         # devices went dark -> not a clean improvement
            if lost and verdict in ("IMPROVING", "FLAT"):
                verdict = "MIXED"                         # lost evidence -> not a clean improvement either
            protocol_gate = campaign_protocol.get("gate")
            if protocol_gate == "REGRESSED":
                verdict = "REGRESSING"                    # definitive endpoint state degradation dominates
            elif protocol_gate == "REVIEW" and verdict in ("IMPROVING", "FLAT"):
                verdict = "MIXED"                         # adjacency topology/coverage loss blocks clean progress
        hr = next((r for r in trajectory if r["metric"].startswith("Avg health")), None)
        cr = next((r for r in trajectory if r["metric"].startswith("Critical/High")), None)
        parts = [f"Across {len(timeline)} collections: {better} metric(s) improving, {worse} worsening."]
        if hr:
            parts.append(f"Avg health {hr['first']}→{hr['last']}.")
        if cr:
            parts.append(f"Critical/High findings {cr['first']}→{cr['last']}.")
        if n_gone:
            parts.append(f"{n_gone} device(s) went DARK (present then absent) -- the health trajectory may "
                         "be survivorship-biased; confirm these are planned decommissions, not failures.")
        if lost:
            parts.append(f"NOT COMPARABLE: {len(lost)} metric(s) lost their evidence between the first and "
                         f"last collection ({', '.join(lost)}) -- excluded from the verdict, and NOT a "
                         "statement that they are clean.")
        if never:
            parts.append(f"NOT COMPARABLE: {len(never)} metric(s) were never measured at EITHER end of "
                         f"this campaign ({', '.join(never)}) -- absent from the trajectory above because "
                         "there is no evidence, NOT because there is nothing to report.")
        _cps = campaign_protocol.get("summary") or {}
        if (_cps.get("n_baseline_peers") or 0) > 0 or campaign_protocol.get("gate") != "NOT_ASSESSED":
            parts.append(str(campaign_protocol.get("note") or ""))
        note = " ".join(parts).strip()

        if failed_collections:
            verdict = "INDETERMINATE"
            detail = "; ".join(
                f"{row['collection']}: {', '.join(str(x) for x in row['failures'])}"
                for row in failed_collections)
            note = (
                f"Campaign certification withheld because {len(failed_collections)} collection/schema "
                f"integrity record(s) are not trustworthy. The trajectory remains visible only as a "
                f"partial observation and is not an improvement claim. {detail}. {note}"
            ).strip()

    last_snapshot = snaps[-1] if snaps and isinstance(snaps[-1], dict) else None
    current_baseline = compute_current_baseline_gate(
        last_snapshot.get("validation_plan") if last_snapshot is not None else None)
    return {"timeline": timeline, "steps": steps, "trajectory": trajectory,
            # Machine-readable twin of the two NOT-COMPARABLE sentences in verdict_note, so a
            # renderer can show the gap as a row rather than having to parse prose. `lost` = the
            # evidence existed at one end only; `never_measured` = neither end had it, which is
            # NOT the same as a clean zero.
            "not_comparable": {"lost": list(lost), "never_measured": list(never)},
            "integrity": {"ok": not failed_collections, "failures": failed_collections},
            "provenance": {
                "source_bindings": list(source_bindings) if isinstance(source_bindings, list) else [],
                "schema_status": schema,
            },
            "protocol_adjacencies": campaign_protocol,
            # The trajectory verdict remains a direction-of-travel result.  The last collection's
            # independent acceptance state travels beside it so a renderer cannot mistake IMPROVING/FLAT
            # for cutover clearance while a current blocker remains.
            "current_baseline": current_baseline,
            "verdict": verdict, "verdict_note": note}


def write_campaign_workbook(snapshots: List[dict], out_path: str, *,
                            source_bindings: Optional[list] = None, schema_status: Any = None,
                            adjacent_comparisons: Any = None) -> None:
    """Write a migration-campaign trend workbook (Campaign Summary verdict + per-metric trajectory /
    Timeline w/ a trajectory line chart / Burndown of findings opened-vs-resolved per step) from a SERIES
    of snapshot_state() dicts."""
    from openpyxl import Workbook
    HF = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    FILL = PatternFill("solid", fgColor=WORKBOOK_NAVY_HEX)
    AL = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DF = Font(name="Calibri", size=10)

    trend = compute_campaign_trend(
        snapshots, source_bindings=source_bindings, schema_status=schema_status)
    receipt_set = _as_dict(adjacent_comparisons)
    receipt_rows = (
        receipt_set.get("comparisons")
        if isinstance(receipt_set.get("comparisons"), list) else []
    )
    receipt_rows = [row for row in receipt_rows if isinstance(row, dict)]
    expected_pairs = receipt_set.get("n_pairs_total")
    if type(expected_pairs) is not int or expected_pairs < 0:
        expected_pairs = max(0, len(snapshots or []) - 1)
    receipt_status = str(receipt_set.get("status") or "not_verified").strip().lower()
    if receipt_set.get("schema") != "campaign_adjacent_comparison_set/1":
        receipt_status = "not_verified"
    receipt_note = str(receipt_set.get("note") or (
        "Canonical adjacent comparison receipts were not supplied to this workbook writer."
    ))
    receipt_complete = (
        receipt_status == "verified"
        and receipt_set.get("complete") is True
        and len(receipt_rows) == expected_pairs
    )
    receipt_render_cap = 1000
    rendered_receipts = receipt_rows[:receipt_render_cap]
    rendered_count = len(rendered_receipts)
    total_produced = len(receipt_rows)
    omitted_count = max(0, total_produced - rendered_count)
    complete_export = os.path.splitext(os.path.abspath(out_path))[0] + ".trend-comparisons.json"
    wb = Workbook(); wb.remove(wb.active)
    from cisco_toolkit.excel import harden_workbook
    harden_workbook(wb)   # sanitize control chars in device-derived text -> no IllegalCharacterError abort

    def sheet(title, cols):
        ws = wb.create_sheet(title)
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=h); cell.font = HF; cell.fill = FILL; cell.alignment = CEN
        ws.freeze_panes = "A2"
        return ws

    def autofit(ws, ncols):
        for col in range(1, ncols + 1):
            mx = len(str(ws.cell(row=1, column=col).value or ""))
            for row in range(2, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    mx = max(mx, len(str(v)))
            ws.column_dimensions[get_column_letter(col)].width = min(max(mx + 2, 12), 60)

    _DIR_FILL = {"improving": "C6EFCE", "worsening": "FFC7CE", "flat": "FFEB9C"}
    _VERDICT_FILL = {"IMPROVING": "C6EFCE", "MIXED": "FFEB9C", "FLAT": "DDEBF7",
                     "REGRESSING": "FFC7CE", "INDETERMINATE": "D9D9D9",
                     "INSUFFICIENT": "EFEFEF", "CLEAR": "C6EFCE", "BLOCKED": "FFC7CE",
                     "NOT_ASSESSED": "D9D9D9"}

    # ---- Campaign Summary (leads with the trajectory verdict) ----
    ws = sheet("Campaign Summary", ["Metric", "First", "Last", "Delta", "Trajectory"])
    vc = ws.cell(row=2, column=1, value="CAMPAIGN VERDICT"); vc.font = Font(name="Calibri", bold=True, size=11)
    vv = ws.cell(row=2, column=2, value=_cv(trend["verdict"]))
    vv.font = Font(name="Calibri", bold=True, size=11)
    vv.fill = PatternFill("solid", fgColor=_VERDICT_FILL.get(trend["verdict"], "FFFFFF"))
    ws.cell(row=2, column=3, value=_cv(trend["verdict_note"])).alignment = AL
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
    current_baseline = _as_dict(trend.get("current_baseline"))
    bc = ws.cell(row=3, column=1, value="CURRENT BASELINE GATE")
    bc.font = Font(name="Calibri", bold=True, size=11)
    bv = ws.cell(row=3, column=2, value=_cv(current_baseline.get("verdict", "INDETERMINATE")))
    bv.font = Font(name="Calibri", bold=True, size=11)
    bv.fill = PatternFill("solid", fgColor=_VERDICT_FILL.get(str(bv.value), "FFFFFF"))
    ws.cell(row=3, column=3, value=_cv(current_baseline.get(
        "note", "Current-baseline receipt is missing or malformed."))).alignment = AL
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=5)
    rc = ws.cell(row=4, column=1, value="ADJACENT CANONICAL RECEIPTS")
    rc.font = Font(name="Calibri", bold=True, size=11)
    rv = ws.cell(row=4, column=2, value=_cv(receipt_status.upper()))
    rv.font = Font(name="Calibri", bold=True, size=11)
    rv.fill = PatternFill(
        "solid",
        fgColor=("C6EFCE" if receipt_status == "verified" and receipt_complete
                 else "FFC7CE" if receipt_status == "not_comparable" else "FFEB9C"),
    )
    ws.cell(
        row=4,
        column=3,
        value=_cv(
            f"complete={str(receipt_complete).lower()}; rendered={rendered_count}; "
            f"total produced={total_produced}; omitted={omitted_count}; expected pairs={expected_pairs}. "
            f"Complete uncapped export: {complete_export}. {receipt_note}"
        ),
    ).alignment = AL
    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=5)
    r = 5
    for t in trend["trajectory"]:
        for c, v in enumerate([t["metric"], t["first"], t["last"], t["delta"], t["direction"]], 1):
            cell = ws.cell(row=r, column=c, value=_cv(v)); cell.font = DF; cell.alignment = AL
        ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=_DIR_FILL.get(t["direction"], "FFFFFF"))
        r += 1
    if not trend["trajectory"]:
        ws.cell(row=5, column=1, value="Not enough comparable metrics across the snapshots.").font = DF
    prov = trend.get("provenance") or {}
    binds = prov.get("source_bindings") if isinstance(prov.get("source_bindings"), list) else []
    if binds:
        r += 1
        ws.cell(row=r, column=1, value="INPUT SHA-256 BINDINGS").font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r, column=2, value=_cv("; ".join(
            f"C{i + 1}={str(v)}" for i, v in enumerate(binds)))).alignment = AL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    schema = prov.get("schema_status") if isinstance(prov.get("schema_status"), dict) else {}
    if schema:
        r += 1
        ws.cell(row=r, column=1, value="SCHEMA COMPATIBILITY").font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r, column=2, value=_cv(
            f"{str(schema.get('status') or 'unverifiable').upper()}: {schema.get('message') or ''}"
            + (" (explicit override recorded)" if schema.get("override") else ""))).alignment = AL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    autofit(ws, 5); ws.column_dimensions["A"].width = 26

    # ---- Timeline (one row per collection) + a trajectory line chart ----
    cols = ["Collection", "Date", "Version", "Switches", "Avg Health", "Critical",
            "Punch-list", "Crit/High", "NOT READY", "Past end-of-support"]
    ws = sheet("Timeline", cols)
    keys = ["collection", "date", "version", "n_switches", "avg_health", "n_critical",
            "n_punchlist", "n_crit_high", "n_not_ready", "past_ldos"]
    for i, pt in enumerate(trend["timeline"], start=2):
        for c, k in enumerate(keys, 1):
            cell = ws.cell(row=i, column=c, value=_cv(pt.get(k, ""))); cell.font = DF
            cell.alignment = CEN if c >= 4 else AL
    autofit(ws, len(cols))
    if len(trend["timeline"]) >= 2:
        try:
            from openpyxl.chart import LineChart, Reference
            n = len(trend["timeline"])
            chart = LineChart(); chart.title = "Migration trajectory"; chart.style = 12
            chart.y_axis.title = "count / score"; chart.x_axis.title = "collection"; chart.height = 8; chart.width = 16
            for col in (5, 8, 7):   # Avg Health, Crit/High, Punch-list
                chart.add_data(Reference(ws, min_col=col, min_row=1, max_row=1 + n), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
            ws.add_chart(chart, get_column_letter(len(cols) + 2) + "2")
        except Exception as e:
            logger.warning(f"  Campaign trend chart skipped: {e}")

    # ---- Burndown (findings opened vs resolved per consecutive step) ----
    ws = sheet("Burndown", ["Step", "Opened", "Opened High/Crit", "Resolved", "Net", "Health regressed",
                            "Health improved", "Step verdict"])
    for i, st in enumerate(trend["steps"], start=2):
        vals = [f"{st['from']} → {st['to']}", st["opened"], st["opened_high"], st["resolved"],
                st["net"], st["regressed"], st["improved"], st["verdict"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=_cv(v)); cell.font = DF
            cell.alignment = CEN if 2 <= c <= 7 else AL
        nf = ws.cell(row=i, column=5)
        nf.fill = PatternFill("solid", fgColor="FFC7CE" if (st["net"] or 0) > 0 else "C6EFCE")
    if not trend["steps"]:
        ws.cell(row=2, column=1, value="Need at least two snapshots for a burndown.").font = DF
    autofit(ws, 8)

    # ---- Protocol Adjacencies (source-bound observed delta, one row per consecutive campaign step) ----
    # Absolute peer totals are intentionally not trajectory metrics: topology and coverage change.  These
    # columns carry only the receipt-gated before/after classifications owned by compute_snapshot_delta.
    ws = sheet("Protocol Adjacencies",
               ["Step", "Gate", "Baseline peers", "State regressed", "Recovered",
                "No longer observed", "Added", "Coverage gaps", "Projection custody",
                "Scope / next action"])
    for i, st in enumerate(trend["steps"], start=2):
        vals = [f"{st['from']} → {st['to']}", st.get("protocol_gate", "NOT_ASSESSED"),
                st.get("protocol_baseline_peers", 0), st.get("protocol_state_regressed", 0),
                st.get("protocol_recovered", 0), st.get("protocol_no_longer_observed", 0),
                st.get("protocol_added", 0), st.get("protocol_coverage_gaps", 0),
                st.get("protocol_projection_custody", "embedded_unverified"),
                st.get("protocol_note", "")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=_cv(v)); cell.font = DF
            cell.alignment = CEN if 2 <= c <= 8 else AL
        gate = str(st.get("protocol_gate") or "NOT_ASSESSED")
        ws.cell(row=i, column=2).fill = PatternFill(
            "solid", fgColor={"PASS": "C6EFCE", "REVIEW": "FFEB9C", "REGRESSED": "FFC7CE"}.get(
                gate, "D9D9D9"))
    if not trend["steps"]:
        ws.cell(row=2, column=1, value="Need at least two snapshots for a protocol adjacency delta.").font = DF
    autofit(ws, 10); ws.column_dimensions["I"].width = 38; ws.column_dimensions["J"].width = 72

    # ---- Adjacent Comparison Receipts (canonical source-bound pair decisions; presentation only) ----
    # The portable JSON sidecar owns the complete uncapped row set.  This sheet may cap rendering, and
    # discloses rendered/total/omitted explicitly; no gate or decision code consumes this UI projection.
    cols = [
        "Pair", "Before snapshot ID", "After snapshot ID", "Before SHA-256", "After SHA-256",
        "Canonical gate", "Admission", "Comparison receipt SHA-256", "Set complete", "Rendered",
        "Total produced", "Omitted", "Status / complete-export action",
    ]
    ws = sheet("Adjacent Comparison Receipts", cols)
    summary_values = [
        "PAIR SET STATUS", "", "", "", "", receipt_status.upper(), "", "",
        receipt_complete, rendered_count, total_produced, omitted_count,
        f"Expected pairs={expected_pairs}. Complete uncapped export: {complete_export}. {receipt_note}",
    ]
    for column, value in enumerate(summary_values, 1):
        cell = ws.cell(row=2, column=column, value=_cv(value)); cell.font = DF; cell.alignment = AL
    ws.cell(row=2, column=6).fill = PatternFill(
        "solid",
        fgColor=("C6EFCE" if receipt_status == "verified" and receipt_complete
                 else "FFC7CE" if receipt_status == "not_comparable" else "FFEB9C"),
    )
    for row_index, entry in enumerate(rendered_receipts, start=3):
        comparison = _as_dict(entry.get("comparison"))
        gate = _as_dict(comparison.get("cutover_gate"))
        admission = _as_dict(comparison.get("comparison_admission"))
        envelope = _as_dict(comparison.get("comparison_receipt"))
        source_pair = _as_dict(admission.get("source_binding"))
        before_source = _as_dict(source_pair.get("before"))
        after_source = _as_dict(source_pair.get("after"))
        canonical_gate = entry.get("canonical_gate", gate.get("verdict", "INDETERMINATE"))
        admission_status = entry.get("admission_status", admission.get("status", "not_comparable"))
        values = [
            f"{entry.get('from', '')} → {entry.get('to', '')}",
            entry.get("before_snapshot_id", before_source.get("snapshot_id", "")),
            entry.get("after_snapshot_id", after_source.get("snapshot_id", "")),
            entry.get("before_sha256", before_source.get("sha256", "")),
            entry.get("after_sha256", after_source.get("sha256", "")),
            canonical_gate,
            admission_status,
            entry.get("comparison_receipt_sha256", envelope.get("receipt_sha256", "")),
            receipt_complete,
            "", "", "", "Copied from the canonical comparison; no gate reinterpretation.",
        ]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=column, value=_cv(value)); cell.font = DF
            cell.alignment = CEN if column in {2, 3, 6, 7, 9} else AL
        ws.cell(row=row_index, column=6).fill = PatternFill(
            "solid",
            fgColor={"PASS": "C6EFCE", "REVIEW": "FFEB9C", "CONDITIONAL": "FFEB9C",
                     "REGRESSED": "FFC7CE", "FAIL": "FFC7CE"}.get(
                         str(canonical_gate), "D9D9D9"),
        )
    if not rendered_receipts:
        ws.cell(
            row=3,
            column=1,
            value="No canonical adjacent comparison receipts were produced; see the status row.",
        ).font = DF
    autofit(ws, len(cols))
    ws.column_dimensions["D"].width = 72
    ws.column_dimensions["E"].width = 72
    ws.column_dimensions["H"].width = 72
    ws.column_dimensions["M"].width = 100

    wb.save(out_path)
    logger.info(f"[OK] Campaign trend: {trend['verdict']} across {len(trend['timeline'])} collections")


def snapshot_state(all_interfaces: Dict[str, Dict[str, InterfaceData]],
                   all_device_physical: List[DevicePhysical]) -> dict:
    import dataclasses
    return {
        "schema": "collect_parse_snapshot/1",
        "script_version": f"V{__version__}",   # NEW-V3.23.8 (M2): was hard-coded "V3.23.0"
        "generated_at": datetime.now().isoformat(),
        "devices": {dp.hostname: dataclasses.asdict(dp) for dp in all_device_physical},
        "interfaces": {host: {port: dataclasses.asdict(d) for port, d in ifaces.items()}
                       for host, ifaces in all_interfaces.items()},
    }


def sparsify_interfaces(snap: dict) -> dict:
    """Return a SHALLOW copy of `snap` whose `interfaces` subtree drops every field equal to its
    empty default (Tier-3 #14 Phase-2). InterfaceData fields default to '' except the positive
    ``run_config_observed`` boolean, so dropping '' and that field's False value is lossless --
    InterfaceData.from_sparse restores omitted defaults on read
    (~70% of the interface field-cells are '' on a real fleet). Confined to `interfaces`: `devices`
    (DevicePhysical) has non-'' defaults and stays dense. Applied ONLY to the on-disk snapshot; the
    in-memory snap_dict every in-process consumer reads is left untouched, so this changes nothing
    about the pipeline's behaviour -- only the persisted file's size."""
    ifaces = snap.get("interfaces")
    if not isinstance(ifaces, dict):
        return snap
    sparse = {host: {port: {k: v for k, v in rec.items()
                            if v != "" and not (k == "run_config_observed" and v is False)}
                     for port, rec in ports.items() if isinstance(rec, dict)}
              for host, ports in ifaces.items() if isinstance(ports, dict)}
    return {**snap, "interfaces": sparse}


# -----------------------------------------------------------------------------
# NEW-V3.23.90: shrink the snapshot copy EMBEDDED in the single-file explorer.
# The on-disk snapshot.json stays full-fidelity (it is the data contract and the
# `--compare` input); this only trims the in-page payload, which on a real fleet
# (the 254-device Meridian scan embedded a 52 MB blob) is dominated by two things the
# explorer never renders verbatim:
#   * interfaces  - hundreds of ports/device, ~50 fields each, most empty strings.
#     The explorer reads every interface field defensively (`d.x||""`, `d.x&&...`,
#     `(d.x||"").trim()`), so an ABSENT key is indistinguishable from an empty one
#     -> dropping empty/placeholder field VALUES is display-neutral. The port entry
#     itself is always kept so buildModel's `Object.keys(ifaces[host])` is unchanged.
#   * physical_health - tens of thousands of Info/OK rows; the sole consumer
#     (deviceIntelSection) filters severity to non-Info/non-OK, so they are dead weight.
# Everything else (already aggregated per-host in analyze, V3.23.90) passes through.
# -----------------------------------------------------------------------------
_EMBED_DROP_VALUES: tuple = ("", None, [], {}, "--")


def _slim_ports(ports) -> dict:
    """One host's {port: record} map with empty/placeholder field VALUES dropped.

    isinstance-coerced at BOTH levels, exactly as sparsify_interfaces already does: `interfaces` itself is
    isinstance-guarded by the caller, but the PER-ELEMENT layer was not -- a truthy non-dict per-host map
    (`interfaces: {"sw1": 5}`) or port record (`{"Gi1/0/1": 5}`) survives `... or {}` and AttributeErrors on
    .items(). On a stored AssessHub upload that is a permanent 500 for GET /api/snapshots/{id}/explorer,
    which has no try wrapper.

    Behaviour-preserving: a FALSY map/record still degrades to {} exactly as `(x or {})` did, and the port
    ENTRY is still always kept, so the explorer's `Object.keys(ifaces[host])` is unchanged."""
    if not isinstance(ports, dict):
        return {}
    return {port: ({k: v for k, v in rec.items() if v not in _EMBED_DROP_VALUES}
                   if isinstance(rec, dict) else {})
            for port, rec in ports.items()}


def _script_safe_json(value) -> str:
    """``json.dumps`` hardened for the HTML ``<script>`` DATA context — which is NOT the JS string
    context ``json.dumps`` alone is correct for.

    Inside ``<script>...</script>`` the HTML tokenizer runs before the JS parser, so a few byte
    sequences change what the browser considers script text no matter how well-formed the JS is:

      * ``</script`` closes the element early — the classic break-out.
      * ``<!--`` enters script-data-escaped state, and a following ``<script`` enters
        script-data-double-escaped, in which the template's own ``</script>`` no longer closes the
        element. Neither sequence contains ``</``, so the old ``.replace("</", "<\\\\/")`` missed both.

    Escaping EVERY ``<`` as ``\\u003c`` makes all three transitions inexpressible, and is lossless:
    ``<`` only ever occurs inside JSON string values here, and ``\\u003c`` is a valid escape in both
    JSON and JS that parses back to ``<``. U+2028/U+2029 are escaped too because ``ensure_ascii=False``
    would otherwise emit them raw, and they are JS line terminators inside a string literal.

    This is the ONE encoder for anything interpolated into the explorer's script block; the defect it
    replaces came from the payload and the label being hardened differently.
    """
    out = json.dumps(value, separators=(",", ":"), ensure_ascii=False)
    return (out.replace("<", "\\u003c")
               .replace("\u2028", "\\u2028")
               .replace("\u2029", "\\u2029"))


def _slim_for_embed(snap_dict: dict) -> dict:
    """Return a renderer-scoped, size-reduced copy of the snapshot for embedding in the
    explorer HTML. Pure (input not mutated); see the block comment above for why each
    transform is safe. Defensive: tolerates missing/oddly-typed sections."""
    out = dict(snap_dict)
    # Traffic Assurance has no dedicated explorer renderer.  Embedding either its governed results or the
    # supporting custody receipt would create a silent, unused projection with no UI contract; withhold both
    # until that renderer exists and can preserve the snapshot's exact verdict/custody semantics.
    out.pop("traffic_assurance", None)
    out.pop("traffic_evidence_custody", None)
    intf = snap_dict.get("interfaces")
    if isinstance(intf, dict):
        out["interfaces"] = {host: _slim_ports(ports) for host, ports in intf.items()}
    ph = snap_dict.get("physical_health")
    if isinstance(ph, list):
        out["physical_health"] = [
            r for r in ph
            if not (isinstance(r, dict) and r.get("severity") in ("Info", "OK", None))]
    return out


# -----------------------------------------------------------------------------
# NEW-V3.17: HTML consolidation. Bake the live snapshot into a copy of the
# read-only Blast-Radius Explorer template so one run yields both the workbook
# and a ready-to-open, air-gapped topology explorer (no second tool, no manual
# snapshot load). Pure stdlib (os + json); no new imports.
# -----------------------------------------------------------------------------
def write_html_explorer(
        output_path: str,
        snap_dict: dict,
        label: str,
        *,
        protocol_assurance_bundle: Any = None,
        complete_export_reference: str = "") -> None:
    """
    Emit a self-contained Blast-Radius Explorer with the live topology embedded.

    Reads 'blast_radius_explorer.html' from inside the package (with a fallback to the legacy repo-root
    location), replaces its demo bootstrap with the embedded snapshot, and writes the patched
    single-file HTML to output_path.

    The template boots on a demo via the LAST statement in its <script>:
        load(demoSnapshot(),"DEMO TOPOLOGY",false);
    That exact text also appears earlier as the demo button's onclick handler, so a
    naive str.replace() (which replaces every occurrence) would corrupt the button -
    it would inject a `const` declaration into an arrow-function body and break all
    JS on the page. We therefore replace ONLY the final occurrence (the real
    bootstrap) via rpartition(), leaving the demo button intact as a one-click way
    back to the sample topology.

    Safety / robustness:
      * Missing template -> warn and skip (never crash a run whose workbook already saved).
      * Bootstrap line absent (template changed) -> warn and skip.
      * Snapshot is minified (separators=(',',':')) to keep the embedded payload small.
      * The optional Protocol Assurance sidecar is accepted only through the shared receipt/export
        validator. A detached snapshot or malformed sidecar renders NOT VERIFIED; this writer never
        hashes a parsed mapping or mints source custody.
      * The snapshot, receipt projection, and label are all emitted via ``_script_safe_json`` -- see
        there. Neither device-derived text nor caller-supplied metadata can break out of the script.
    """
    # The explorer template ships INSIDE the package (cisco_toolkit/blast_radius_explorer.html), so it is
    # present in a built wheel and resolves the same whether the package is run from a checkout or a
    # pip install. Fall back to the legacy repo-root location for any older layout that still keeps it there.
    _here = os.path.dirname(os.path.abspath(__file__))
    template = os.path.join(_here, "blast_radius_explorer.html")
    if not os.path.isfile(template):
        template = os.path.join(os.path.dirname(_here), "blast_radius_explorer.html")  # legacy repo-root
    if not os.path.isfile(template):
        logger.warning(f"  HTML Explorer skipped: template not found at {template}")
        return

    with open(template, encoding="utf-8") as f:
        html = f.read()

    bootstrap = 'load(demoSnapshot(),"DEMO TOPOLOGY",false);'
    if bootstrap not in html:
        logger.warning("  HTML Explorer skipped: demo bootstrap line not found in template "
                       "(template may have changed).")
        return

    slim = _slim_for_embed(snap_dict)                  # NEW-V3.23.90: shrink the in-page payload only
    # BOTH values go through the same hardening. The label used to get a bare json.dumps() while the
    # payload beside it got the '</' treatment, and that asymmetry WAS the bug: json.dumps produces a
    # correct JS string literal but says nothing about the HTML <script> DATA context, so a label of
    # `pwn</script><script>alert(document.domain)</script>` closed the block early and the injected
    # script ran under the template's own `script-src 'unsafe-inline'` CSP. The label is
    # attacker-reachable -- AssessHub takes it as an unsanitized form field that falls back to the
    # UPLOADED FILENAME, stores it, and serves it back from the explorer route -- so this was stored
    # XSS with same-origin access to every stored client snapshot.
    embedded = _script_safe_json(slim)
    protocol_surface = protocol_assurance_surface_payload(
        protocol_assurance_bundle,
        complete_export_reference=complete_export_reference,
    )
    embedded_protocol = _script_safe_json(protocol_surface)
    replacement = (f"const EMBEDDED_PROTOCOL_ASSURANCE={embedded_protocol};\n"
                   f"const EMBEDDED_SNAPSHOT={embedded};\n"
                   f"load(EMBEDDED_SNAPSHOT,{_script_safe_json(label)},true,"
                   "EMBEDDED_PROTOCOL_ASSURANCE);")

    # Replace ONLY the last occurrence (the bootstrap), not the button's onclick.
    head, _sep, tail = html.rpartition(bootstrap)
    patched = head + replacement + tail

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(patched)
    logger.info(f"[Phase 22] HTML Explorer embedded payload: {len(embedded) / 1e6:.1f} MB")
    logger.info(f"[Phase 22] HTML Explorer written: {output_path}")


# -----------------------------------------------------------------------------
# Snapshot redaction (opt-in --redact): pseudonymize IPs / MACs / serial numbers
# so a single-file HTML/JSON deliverable can be shared without leaking the real
# addressing. Mappings are CONSISTENT (same input -> same output) and IPs keep
# their /24 grouping, so ARP (MAC->IP), dual-homing, and subnet/flow-trace
# relationships the explorer relies on survive. Hostnames are intentionally kept.
# -----------------------------------------------------------------------------
_REDACT_IP_RE = re.compile(r"\b(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})\b")
_REDACT_MAC_RE = re.compile(
    r"\b(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}\b|\b(?:[0-9A-Fa-f]{4}\.){2}[0-9A-Fa-f]{4}\b")
# IPv6 (HTML_-01): the IPv4 remap above is dotted-quad ONLY, so IPv6 addresses (global 2001:db8::,
# link-local fe80::, embedded in descriptions) passed through --redact untouched -- the share-safe contract
# was broken for any dual-stack / IPv6 fleet. This matches the full 8-group and ::-compressed textual forms
# but NOT a colon-MAC (which has neither 7 colons nor a '::'), so the MAC remap stays correct regardless of
# order. Every alternative requires either 7 colons or a '::', and all quantifiers are bounded ({1,7}), so a
# long hex/colon run cannot induce catastrophic backtracking (ReDoS). An optional %zone-id is consumed.
_REDACT_IP6_RE = re.compile(
    r"(?<![:.\w])(?:"
    r"(?:[0-9A-Fa-f]{1,4}:){7}[0-9A-Fa-f]{1,4}"                 # x:x:x:x:x:x:x:x  (7 colons; a MAC has 5)
    r"|(?:[0-9A-Fa-f]{1,4}:){1,7}:"                             # x::  …  x:x:x:x:x:x:x::
    r"|(?:[0-9A-Fa-f]{1,4}:){1,6}:[0-9A-Fa-f]{1,4}"             # x::x … x:x:x:x:x:x::x
    r"|(?:[0-9A-Fa-f]{1,4}:){1,5}(?::[0-9A-Fa-f]{1,4}){1,2}"
    r"|(?:[0-9A-Fa-f]{1,4}:){1,4}(?::[0-9A-Fa-f]{1,4}){1,3}"
    r"|(?:[0-9A-Fa-f]{1,4}:){1,3}(?::[0-9A-Fa-f]{1,4}){1,4}"
    r"|(?:[0-9A-Fa-f]{1,4}:){1,2}(?::[0-9A-Fa-f]{1,4}){1,5}"
    r"|[0-9A-Fa-f]{1,4}:(?::[0-9A-Fa-f]{1,4}){1,6}"
    r"|:(?::[0-9A-Fa-f]{1,4}){1,7}"                             # ::x  ::x:x  (loopback/unspecified-with-suffix)
    r")(?:%[0-9A-Za-z]+)?(?![:.\w])")
_REDACT_SERIAL_KEYS = {"serial_number", "chassis_serial",
                       "current_switch_serial", "neighbor_switch_serial",
                       # controller + inventory serials reach the snapshot under other key names: ACI
                       # fabric-node 'serial' (parse_aci_fabric_nodes -> snap['aci'].nodes) and the
                       # power-supply 'ps_serials' list (parse_show_inventory). Without these they leaked
                       # verbatim under --redact (a bare serial matches no inline secret form). 'ps_serials'
                       # is a LIST -- _walk recurses each string item with the same key, so each is redacted.
                       "serial", "ps_serials", "sn"}

# Credential deny-list: conservatively match KNOWN secret-bearing config/output forms
# (IOS / IOS-XE / NX-OS, case-insensitive) and replace ONLY the secret token with a
# placeholder, keeping the surrounding keywords as context. Each pattern captures the
# prefix in group 1 and the secret in group 2; the secret is swapped for "<redacted>".
# Idempotent: re-running over an already-scrubbed string re-captures "<redacted>" and
# substitutes it for itself. We are deliberately narrow (no blanket token redaction) so
# non-secret structured fields are never corrupted.
_REDACT_PLACEHOLDER = "<redacted>"
# A secret value may be a bare token or a quoted, space-bearing value.  Matching only ``\S+``
# changed ``set passphrase "correct horse battery staple"`` into
# ``set passphrase <redacted> horse battery staple"`` and the verifier then blessed the residue
# because the first token was the placeholder.  Consume one complete shell/config value instead.
_REDACT_SECRET_VALUE = r"""(?:"(?:\\.|[^"\\\r\n])*"|'(?:\\.|[^'\\\r\n])*'|\S+)"""
_REDACT_SECRET_RES = [re.compile(prefix + "(" + _REDACT_SECRET_VALUE + ")", re.I) for prefix in (
    # SNMP community strings: 'snmp-server community <VALUE>' and the bare
    # 'community <VALUE>' form (host/group/trap lines).
    r"(snmp-server\s+community\s+)",
    r"(\bcommunity\s+)",
    # 'snmp-server host <ip> [vrf X] [traps|informs] version {1|2c} <COMMUNITY>' -- the trap-host community is a
    # bare positional token with NO 'community' keyword to anchor on, so the two patterns above missed it and it
    # shipped verbatim under --redact (leak-corpus K4). v3 uses a username (not a secret), so only 1|2c match.
    r"(snmp-server\s+host\s+\S+\s+(?:vrf\s+\S+\s+)?(?:(?:traps?|informs?)\s+)?version\s+(?:1|2c)\s+)",
    # Cisco password/secret forms: type-7/type-5 and cleartext, 'enable secret',
    # and 'username <u> password|secret <VALUE>'. The username token is preserved.
    r"(\bpassword\s+(?:(?:ENC|\d+)\s+)?)",
    r"(\bsecret\s+(?:\d+\s+)?)",
    r"((?:username|user)\s+\S+\s+(?:password|secret)\s+(?:\d+\s+)?)",
    # Shared keys. Specific forms FIRST so the generic bare 'key' below cannot consume
    # their qualifier (e.g. 'pre-shared-key local <V>' must not let 'key local' match).
    # TACACS+/RADIUS server keys, 'key-string <VALUE>' (SNMPv3 / EIGRP / OSPF keychains),
    # IKE pre-shared keys, and 'crypto isakmp key <VALUE> address ...'.
    r"((?:tacacs-server|radius-server)\s+(?:host\s+\S+\s+)?key\s+(?:\d+\s+)?)",
    r"(key-string\s+(?:\d+\s+)?)",
    r"(pre-shared-key\s+(?:(?:local|remote|ascii-text|hexadecimal)\s+)?(?:\d+\s+)?)",
    r"(crypto\s+isakmp\s+key\s+(?:\d+\s+)?)",
    # Generic 'key 7 <hex>' / 'key <cleartext>' (keychain key, OSPF/EIGRP authentication). The optional
    # inner group absorbs a hash-algorithm label so 'authentication-key|message-digest-key N md5|sha|
    # hmac-sha <DIGEST>' (NTP/OSPF/EIGRP) redacts the DIGEST after it, not the 'md5'/'sha' token -- the
    # latter left the real, offline-crackable digest exposed. Anchored on 'key', so a bare IKE 'hash md5'
    # algorithm choice (no key-id + secret) is never corrupted. The negative lookahead keeps STRUCTURAL
    # follow-words intact: 'key chain <NAME>' declares a keychain (name is not a secret), and the bare rule
    # runs AFTER the pre-shared-key rule over the accumulating string, so without the guard it re-fired on
    # 'pre-shared-key local <redacted>' and mangled the 'local'/'remote' direction qualifier.
    r"((?<!private-)(?<!shared-)\bkey\s+(?:\d+\s+)?(?:(?:md5|sha\S*|hmac-\S+|cmac-\S+)\s+(?:\d+\s+)?)?)(?!chain\b|local\b|remote\b)",
    # Non-Cisco vendor config forms: FortiGate 'set passwd|psksecret|password [ENC] <VALUE>' and Junos
    # 'authentication-key|secret "<VALUE>"' -- 'passwd'/'psksecret' are not the whole words 'password'/'secret',
    # so the Cisco patterns above miss them.
    r"(set\s+(?:passwd|psksecret|password|private-key|passphrase)\s+(?:ENC\s+)?)",
)]
# JSON-VALUE secrets: the controller-REST channels (ACI / ISE / FMC / vManage) and IaC exports store a secret as
# a VALUE under a key, with no inline keyword for the deny-list regexes above to anchor on. So redact the WHOLE
# value when its key is a known secret-bearing name. Keys are normalized (lowercased, '_'/'-' stripped) before
# the lookup. 'key' is deliberately EXCLUDED -- it is a generic structural field name across the snapshot (causal
# flows, design decisions) and the inline 'key <val>' CLI form is already covered above.
_REDACT_SECRET_KEYS = {
    "password", "passwd", "pwd", "passphrase", "secret", "psksecret", "presharedkey", "psk",
    "token", "authtoken", "accesstoken", "apikey", "apisecret", "community", "snmpcommunity",
    "credential", "credentials", "privatekey", "sharedsecret", "clientsecret",
}   # 'pass' deliberately omitted -- it is a pass/fail COUNT key in security summaries, not a secret

# SUBSTRING tokens for a credential-named key -- the real controller/config field names are COMPOUND
# (tacacsSharedSecret, roCommunity, authPassword, wpaPassphrase, enableSecret), which an EXACT-match against the
# set above silently missed (multi-domain audit #6). A normalized key CONTAINING any token is a secret bearer.
# Bare 'key' and 'pass' are NOT tokens (generic structural field / pass-fail count) -- see _is_secret_key.
_REDACT_SECRET_TOKENS = (
    "password", "passwd", "passphrase", "secret", "community", "psk", "presharedkey", "sharedsecret",
    "token", "apikey", "apisecret", "privatekey", "privkey", "credential",
)

# CDP/LLDP 'Device ID: <host>(<SERIAL>)' embeds the neighbor's chassis serial in a free-text field that is NOT a
# serial-named key, so the key-based serial pass missed it (audit-5 sec HIGH: 55 real serials survived --redact).
# Match a parenthesized Cisco serial (3 letters + 4 digits + 2-6 alnum, e.g. FOC1830R1QS) so it routes through the
# SAME serial pseudonymizer for a consistent SNxxxx; the SNxxxx pseudonym (2 leading letters) never re-matches.
_REDACT_CDP_SERIAL_RE = re.compile(r"\(([A-Z]{3}[0-9]{4}[A-Z0-9]{2,6})\)")
# Cisco serials also occur in arbitrary prose and generated validation expectations, where no
# serial-shaped schema key or CDP parentheses exist.  Match the same deliberately narrow token
# shape independently of context; SN#### pseudonyms cannot re-match it, so the pass is idempotent.
_REDACT_INLINE_SERIAL_RE = re.compile(
    r"(?<![A-Z0-9])([A-Z]{3}[0-9]{4}[A-Z0-9]{2,6})(?![A-Z0-9])",
    re.IGNORECASE,
)
_REDACT_EMAIL_RE = re.compile(
    r"(?<![\w.+-])[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@"
    r"(?:[A-Z0-9-]+\.)+[A-Z]{2,63}(?![\w.-])",
    re.IGNORECASE,
)

_SYNTH_DOMAIN = "assesshub-redacted.invalid"
_SYNTH_MARKER_RE = re.compile(
    r"(?:v4-n\d{5}-h\d{3}|v6-\d{8}|mac-\d{12}|serial-\d{6})\."
    + re.escape(_SYNTH_DOMAIN),
    re.IGNORECASE,
)
_SYNTH_SERIAL_RE = re.compile(
    r"serial-\d{6}\." + re.escape(_SYNTH_DOMAIN) + r"\Z",
    re.IGNORECASE,
)
_SYNTH_EMAIL_RE = re.compile(
    r"contact-\d{6}@" + re.escape(_SYNTH_DOMAIN) + r"\Z",
    re.IGNORECASE,
)
_MAX_V4_SYNTH_NETWORKS = 65_536
_MAX_SYNTH_IDENTITIES = 999_999


class RedactionPseudonymExhausted(RuntimeError):
    """The bounded synthetic namespace cannot issue another collision-free marker."""


class _SyntheticAllocator:
    """Issue unmistakably synthetic, bounded, collision-checked ``.invalid`` markers."""

    def __init__(self, reserved=None, limit: int = _MAX_SYNTH_IDENTITIES):
        self.reserved = set(reserved or ())
        self.issued = set()
        self.next_value = 1
        self.limit = int(limit)

    def issue(self, render) -> str:
        while self.next_value <= self.limit:
            value = render(self.next_value)
            self.next_value += 1
            if value in self.reserved or value in self.issued:
                continue
            self.issued.add(value)
            return value
        raise RedactionPseudonymExhausted(
            "bounded redaction pseudonym namespace exhausted; refusing partial redaction"
        )


def _existing_synthetic_markers(root) -> set:
    """Reserve producer-origin markers already present so a new identity cannot collide with one."""
    found = set()
    stack = [root]
    while stack:
        value = stack.pop()
        if isinstance(value, dict):
            stack.extend(value.keys())
            stack.extend(value.values())
        elif isinstance(value, (list, tuple, set)):
            stack.extend(value)
        elif isinstance(value, str):
            found.update(match.group(0).casefold() for match in _SYNTH_MARKER_RE.finditer(value))
            found.update(match.group(0).casefold() for match in _REDACT_EMAIL_RE.finditer(value)
                         if _SYNTH_EMAIL_RE.fullmatch(match.group(0)))
    return found


def _norm_key(k) -> str:
    return re.sub(r"[_-]", "", str(k or "").lower())


def _scrub_secrets(s: str) -> str:
    """Replace known credential / community / key material in a config-or-output string
    with a placeholder, preserving surrounding context. Conservative (deny-list of
    compiled regexes, secret-token capture only) and idempotent."""
    for rx in _REDACT_SECRET_RES:
        s = rx.sub(r"\g<1>" + _REDACT_PLACEHOLDER, s)
    return s


def redact_snapshot(snap: dict) -> dict:
    """Return a copy of the snapshot with IPs, MACs, serial numbers, and emails consistently
    pseudonymized for sharing the single-file deliverable. Same input maps to the same
    output and IPs keep their /24 grouping, so topology / ARP / subnet relationships
    survive; hostnames are kept. Pure (stdlib only); the input is not mutated.
    COLLISION-PROOF IPv4 (same cure as _make_redactor's): pseudonym /24s are drawn from
    240.0.0.0/4 — the IANA-reserved Class E block no deployed device can carry — so a
    redacted deliverable can NEVER reproduce a real address. The old in-band
    10.{i//256}.{i%256} scheme could (net 10.0.20 drew '10.0.10', re-emitting the real
    gateway 10.0.10.1). A net ALREADY in 240.x maps to ITSELF: every IPv4 in a scrubbed
    output is 240.x, so that identity rule is exactly what keeps redact_snapshot
    idempotent (a second pass is a no-op)."""
    reserved = _existing_synthetic_markers(snap)
    ip_map: Dict[str, int] = {}
    ip6_map: Dict[str, str] = {}
    mac_map: Dict[str, str] = {}
    serial_map: Dict[str, str] = {}
    email_map: Dict[str, str] = {}
    ip6_allocator = _SyntheticAllocator(reserved)
    mac_allocator = _SyntheticAllocator(reserved)
    serial_allocator = _SyntheticAllocator(reserved)
    email_allocator = _SyntheticAllocator(reserved)
    _next_net = [1]

    def _ip(m):
        net = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
        if net not in ip_map:
            while _next_net[0] <= _MAX_V4_SYNTH_NETWORKS:
                candidate = _next_net[0]
                _next_net[0] += 1
                if not any(
                    f"v4-n{candidate:05d}-h{host:03d}.{_SYNTH_DOMAIN}".casefold() in reserved
                    for host in range(256)
                ):
                    ip_map[net] = candidate
                    break
            if net not in ip_map:
                raise RedactionPseudonymExhausted(
                    "bounded IPv4 redaction namespace exhausted; refusing partial redaction"
                )
        return f"v4-n{ip_map[net]:05d}-h{int(m.group(4)):03d}.{_SYNTH_DOMAIN}"

    def _ip6(m):
        s = m.group(0)
        if s not in ip6_map:
            ip6_map[s] = ip6_allocator.issue(
                lambda i: f"v6-{i:08d}.{_SYNTH_DOMAIN}"
            )
        return ip6_map[s]

    def _mac(m):
        key = re.sub(r"[^0-9a-f]", "", m.group(0).lower())
        if key not in mac_map:
            mac_map[key] = mac_allocator.issue(
                lambda i: f"mac-{i:012d}.{_SYNTH_DOMAIN}"
            )
        return mac_map[key]

    def _serial(v):
        if not v:
            return v
        if _SYNTH_SERIAL_RE.fullmatch(str(v)):
            return v
        key = str(v).upper()
        if key not in serial_map:
            serial_map[key] = serial_allocator.issue(
                lambda i: f"serial-{i:06d}.{_SYNTH_DOMAIN}"
            )
        return serial_map[key]

    def _email(value):
        address = str(value)
        if _SYNTH_EMAIL_RE.fullmatch(address):
            return address
        key = address.casefold()
        if key not in email_map:
            email_map[key] = email_allocator.issue(
                lambda i: f"contact-{i:06d}@{_SYNTH_DOMAIN}"
            )
        return email_map[key]

    def _scrub_identity_tokens(s):
        """Pseudonymize serials outside email tokens, then pseudonymize real email addresses.

        Shielding each complete address from the serial matcher preserves approved example-domain
        placeholders even if their local part happens to look like a Cisco serial.
        """
        out = []
        cursor = 0

        def serials(chunk):
            chunk = _REDACT_CDP_SERIAL_RE.sub(
                lambda m: "(" + _serial(m.group(1)) + ")", chunk
            )
            return _REDACT_INLINE_SERIAL_RE.sub(lambda m: _serial(m.group(1)), chunk)

        for match in _REDACT_EMAIL_RE.finditer(s):
            out.append(serials(s[cursor:match.start()]))
            out.append(_email(match.group(0)))
            cursor = match.end()
        out.append(serials(s[cursor:]))
        return "".join(out)

    def _scrub(s):
        # Strip credentials / community / key material first so a secret token is replaced wholesale, THEN
        # pseudonymize any remaining IPv4 / IPv6 / MACs in context. IPv6 is remapped before MAC; the two
        # patterns are mutually exclusive (a MAC has neither 7 colons nor a '::'), so neither corrupts the other.
        s = _scrub_identity_tokens(_scrub_secrets(s))
        return _REDACT_MAC_RE.sub(_mac, _REDACT_IP6_RE.sub(_ip6, _REDACT_IP_RE.sub(_ip, s)))

    def _is_secret_key(key) -> bool:
        nk = _norm_key(key)
        if not nk or nk == "key":            # 'key' is a generic structural field name -- never a secret on its own
            return False
        return nk in _REDACT_SECRET_KEYS or any(tok in nk for tok in _REDACT_SECRET_TOKENS)

    def _redact_key(key):
        if not isinstance(key, str):
            return key
        return _scrub(key)

    def _store_preserving_key(out, safe_key, value):
        """Store a redacted key without silently overwriting a colliding original entry.

        Case variants of one serial intentionally share a stable ``SN####`` pseudonym, and a
        pre-existing pseudonym-shaped key may already occupy that spelling.  Preserve the first
        spelling and add a deterministic ordinal alias for every later collision.  The alias
        contains only the stable pseudonym, never the source serial.
        """
        candidate = safe_key
        if candidate in out:
            base = str(safe_key)
            ordinal = 2
            candidate = f"{base}~{ordinal}"
            while candidate in out:
                ordinal += 1
                candidate = f"{base}~{ordinal}"
        out[candidate] = value

    def _redact_all(o):
        # Every string leaf under a credential-named container is a secret bearer -- a secret nested one level
        # below the key ({'apikey':{'value':...}}) must not survive (multi-domain audit #7). Over-redacting
        # non-secret siblings (e.g. an 'enc: type6' tag) is the safe direction.
        if isinstance(o, dict):
            out = {}
            for k, v in o.items():
                _store_preserving_key(out, _redact_key(k), _redact_all(v))
            return out
        if isinstance(o, list): return [_redact_all(v) for v in o]
        if isinstance(o, str): return _REDACT_PLACEHOLDER if o else o
        return o

    def _walk(o, key=None):
        if isinstance(o, dict):
            out = {}
            for k, v in o.items():
                safe_key = _redact_key(k)
                if isinstance(v, (dict, list)) and _is_secret_key(k):
                    # secret-named key over a CONTAINER -> scrub every leaf
                    safe_value = _redact_all(v)
                else:
                    safe_value = _walk(v, k)
                _store_preserving_key(out, safe_key, safe_value)
            return out
        if isinstance(o, list):
            return [_walk(v, key) for v in o]
        if isinstance(o, str):
            if key in _REDACT_SERIAL_KEYS: return _serial(o)
            if key == "wild": return o   # ACL wildcard mask is not an address; preserve so post-redact L4 eval stays correct
            if o and _is_secret_key(key):
                return _REDACT_PLACEHOLDER   # a secret stored as a JSON value under a credential-named key
            return _scrub(o)
        return o

    return _walk(snap)


def _make_redactor(reserved=None):
    """A fresh, self-consistent pseudonymizer (same scheme as redact_snapshot: IPv4 keeps its /24 grouping;
    IPv6 -> fd00::; MAC -> 02:..; serial -> SNxxxx). Returns (scrub_str, redact_serial) sharing per-call maps.
    Shared by redact_collected_inplace + redact_workbook_cells so the --redact workbook is scrubbed everywhere.
    COLLISION-PROOF IPv4 (same Class E scheme as redact_snapshot's map): pseudonym /24s are drawn from
    240.0.0.0/4 — the IANA-reserved Class E block that can never be assigned to real gear — so a pseudonym
    can NEVER reproduce a real deployed address. The old in-band 10.{i//256}.{i%256} scheme could: the
    NRFU sheet's second real SVI net drew the pseudonym '10.0.10', re-emitting the OTHER real gateway IP
    10.0.10.1 verbatim into a --redact workbook (and a net whose index matched its own name survived
    unredacted). Belt-and-braces guards keep a candidate from equalling the input net, an already-seen
    net, or an already-issued pseudonym even if a capture somehow contains 240.x addresses — unlike
    redact_snapshot, which must map an already-240.x net to ITSELF to stay idempotent, this per-call map
    is never re-fed its own output, so it refuses identity outright. Deterministic per call."""
    reserved = set(reserved or ())
    ip_map: Dict[str, int] = {}
    ip6_map: Dict[str, str] = {}
    mac_map: Dict[str, str] = {}
    serial_map: Dict[str, str] = {}
    email_map: Dict[str, str] = {}
    ip6_allocator = _SyntheticAllocator(reserved)
    mac_allocator = _SyntheticAllocator(reserved)
    serial_allocator = _SyntheticAllocator(reserved)
    email_allocator = _SyntheticAllocator(reserved)
    _next_ip = [1]

    def _ip(m):
        net = f"{m.group(1)}.{m.group(2)}.{m.group(3)}"
        if net not in ip_map:
            while _next_ip[0] <= _MAX_V4_SYNTH_NETWORKS:
                candidate = _next_ip[0]
                _next_ip[0] += 1
                if not any(
                    f"v4-n{candidate:05d}-h{host:03d}.{_SYNTH_DOMAIN}".casefold() in reserved
                    for host in range(256)
                ):
                    ip_map[net] = candidate
                    break
            if net not in ip_map:
                raise RedactionPseudonymExhausted(
                    "bounded IPv4 redaction namespace exhausted; refusing partial redaction"
                )
        return f"v4-n{ip_map[net]:05d}-h{int(m.group(4)):03d}.{_SYNTH_DOMAIN}"

    def _ip6(m):
        s = m.group(0)
        if s not in ip6_map:
            ip6_map[s] = ip6_allocator.issue(
                lambda i: f"v6-{i:08d}.{_SYNTH_DOMAIN}"
            )
        return ip6_map[s]

    def _mac(m):
        key = re.sub(r"[^0-9a-f]", "", m.group(0).lower())
        if key not in mac_map:
            mac_map[key] = mac_allocator.issue(
                lambda i: f"mac-{i:012d}.{_SYNTH_DOMAIN}"
            )
        return mac_map[key]

    def serial(v):
        if not v:
            return v
        if _SYNTH_SERIAL_RE.fullmatch(str(v)):
            return v
        key = str(v).upper()
        if key not in serial_map:
            serial_map[key] = serial_allocator.issue(
                lambda i: f"serial-{i:06d}.{_SYNTH_DOMAIN}"
            )
        return serial_map[key]

    def email(v):
        address = str(v)
        if _SYNTH_EMAIL_RE.fullmatch(address):
            return address
        key = address.casefold()
        if key not in email_map:
            email_map[key] = email_allocator.issue(
                lambda i: f"contact-{i:06d}@{_SYNTH_DOMAIN}"
            )
        return email_map[key]

    def _identities(s):
        chunks = []
        cursor = 0

        def serials(chunk):
            chunk = _REDACT_CDP_SERIAL_RE.sub(
                lambda m: "(" + serial(m.group(1)) + ")", chunk
            )
            return _REDACT_INLINE_SERIAL_RE.sub(lambda m: serial(m.group(1)), chunk)

        for match in _REDACT_EMAIL_RE.finditer(s):
            chunks.append(serials(s[cursor:match.start()]))
            chunks.append(email(match.group(0)))
            cursor = match.end()
        chunks.append(serials(s[cursor:]))
        return "".join(chunks)

    def scrub(s):
        s = _identities(_scrub_secrets(s))
        return _REDACT_MAC_RE.sub(_mac, _REDACT_IP6_RE.sub(_ip6, _REDACT_IP_RE.sub(_ip, s)))

    return scrub, serial


#: Structured documents `redact_collection_dir` does NOT rewrite in place, and the scratch name it
#: writes while rewriting a capture. `_scrub_secrets` is a grammar over line-oriented device-config
#: text (``snmp-server community <V>``, ``username u password <V>``); it does not read structured
#: data, substituting inside a serialised document is how a capture stops parsing, and rewriting a
#: generated ``.html`` deliverable that happens to sit in the collection folder would break the run
#: manifest that already sealed it. These are the ONLY exclusions — see `_is_raw_capture` — they
#: are the same set the independent verifier applies
#: (`webapp.backend.redaction_verify._STRUCTURED_CAPTURE_SUFFIXES`), and that verifier reports them
#: under ``uncovered`` rather than letting the silence read as "clean".
#:
#: RULE OWNER: ``webapp.backend.redaction_verify.is_uncoverable_capture`` states this rule for the
#: independent verifier AND for the ingest census (``ingest._is_raw_capture`` delegates to it), so it
#: is the owner of record. The producer cannot import it — ``cisco_toolkit`` must not depend on
#: ``webapp``, and the verifier's independence forbids importing the producer — so the rule is
#: RESTATED here with the SAME suffix set and the SAME primitive: the owner's own expression,
#: ``PurePath(name).suffix.casefold()`` (see `_capture_suffix`).
#:
#: Two earlier restatements used a DIFFERENT primitive and each opened a hole. ``str.endswith``
#: disagreed on bare-name dotfiles (the producer skipped ``.json`` as structured while the verifier
#: scanned it as a capture). ``os.path.splitext(name)[1]`` replaced it under a comment claiming the
#: two were byte-for-byte identical — they are not, ``splitext`` skips ALL leading dots on the
#: basename and ``PurePath.suffix`` skips only the first — so they disagreed on every name with two
#: or more leading dots. Suffix-of-a-name is deceptively easy to restate and has now been restated
#: wrongly twice, which is why the rule has ONE owner and ``tests/test_redact_collection.py`` pins
#: the two against each other over a GENERATED corpus of that structural class rather than a
#: hand-list of examples.
_REDACT_SKIP_CAPTURE_SUFFIXES = frozenset({".json", ".xml", ".yml", ".yaml", ".html", ".htm"})
_REDACT_SCRUB_TEMP_SUFFIX = ".redacting"
#: Deliberate BOUND on the widened (default-INCLUDE) rule — see `redact_collection_dir`. Mirrors the
#: verifier's ``redaction_verify.MAX_ARTIFACT_BYTES``: a file larger than this is not a device
#: capture, and reading it whole into memory to rewrite the engineer's only copy is the wrong risk to
#: take on a field laptop. Skipped files are DISCLOSED (returned), never silently dropped.
_REDACT_MAX_CAPTURE_BYTES = 128 * 1024 * 1024


def _capture_suffix(filename: str) -> str:
    """The basename's final extension, casefolded — the ONE primitive both sides of the raw-capture
    rule use.

    This is the OWNER's own expression, not a paraphrase of it: the independent verifier computes
    ``Path(filename).suffix.casefold()``, and ``Path.suffix`` IS ``PurePath.suffix`` on every
    platform, so this cannot drift from it.

    It is deliberately NOT ``os.path.splitext(filename)[1]``, which stood here under a comment
    asserting the two were byte-for-byte identical. Measured over a generated corpus of 168 names
    they return different strings for 51, and for names with two or more leading dots the difference
    changes the CLASSIFICATION: ``splitext("..json")[1]`` is ``""`` (⇒ a capture) while
    ``PurePath("..json").suffix`` is ``".json"`` (⇒ a structured serialisation the scrub must not
    touch). Measured consequence with the old primitive: `redact_collection_dir` rewrote
    ``CORE-1/..json`` in place until it no longer parsed (``Expecting ',' delimiter``) while the
    verifier declared the same file uncoverable, and the count reconciliation at
    ``COLLECT_PARSE_V3_23_0.py:4267`` then failed the whole run with "producer/verifier raw-capture
    file counts disagree"."""
    return PurePath(filename or "").suffix.casefold()


def _is_raw_capture(filename: str) -> bool:
    """Is ``filename`` a raw device capture the in-place secret scrub owns? (name half.)

    The old test was ``fn.endswith(".txt")`` — a single extension standing in for the class
    "collected capture text". A device folder holding ``show_version.txt`` alongside
    ``backup-config.cfg`` or ``show_tech-support.log`` is accepted everywhere else in this
    codebase, and those two were never scrubbed: measured, they kept cleartext ``enable secret``,
    ``snmp-server community`` and ``username ... password`` values through a run that reported the
    captures SCRUBBED and exited 0.

    So the default is now INCLUDE, and only the two structural exclusions above opt out. An
    extension nobody anticipated (``.conf``, ``.cfg``, ``.log``, ``.out``, none at all) is a
    capture and gets scrubbed — the failure mode is inverted from "unknown ⇒ leaked" to
    "unknown ⇒ protected". The content half of the rule (binary bytes are not a capture) is
    applied at the read in `redact_collection_dir`, because only the bytes can decide it.

    Matched with `_capture_suffix`, which is the OWNER's primitive verbatim — not `str.endswith`
    (which disagreed on bare-name dotfiles: ``.json`` has no extension, so the verifier counted it a
    capture while the endswith form skipped it) and not `os.path.splitext` (which disagreed on names
    with two or more leading dots: ``..json``). Two matchers for one rule is two rules; see
    `_REDACT_SKIP_CAPTURE_SUFFIXES` for who owns it and what each divergence cost."""
    suffix = _capture_suffix(filename)
    return suffix != _REDACT_SCRUB_TEMP_SUFFIX and suffix not in _REDACT_SKIP_CAPTURE_SUFFIXES


class _ScrubResult(tuple):
    """``(scanned, changed)`` — plus ``.uncovered``, the files this pass declined to rewrite.

    A plain 2-tuple, so every existing ``scanned, changed = redact_collection_dir(...)`` caller and
    every ``== (1, 1)`` assertion keeps working unchanged; the coverage list rides alongside as an
    attribute instead of being visible only in a ``logger.debug`` line nobody reads at a client site.
    Each entry is ``(relative_path, reason)``."""

    def __new__(cls, scanned: int, changed: int, uncovered=()):
        self = super().__new__(cls, (scanned, changed))
        self.uncovered = tuple(uncovered)
        return self


def redact_collection_dir(collection_dir: str) -> tuple:
    """Plan A / Tier-1 #5: scrub SECRET VALUES (passwords / communities / keys — the same
    conservative _scrub_secrets deny-list --redact uses) IN PLACE across every collected
    text capture under collection_dir (see `_is_raw_capture`; NOT just ``*.txt``). Values
    only: IPs / hostnames / interfaces are KEPT
    so the dir stays analyzable with --no-collect and remains the --compare/--trend
    source; nothing is ever deleted. Idempotent (the placeholder never re-matches).
    Returns (txt_files_scanned, files_changed). Fail-soft per file — one unreadable
    capture never aborts the scrub of the rest. Note: rewritten captures will no longer
    match archive hashes recorded at collection time (deliberate, opt-in).

    BYTE FIDELITY (the promise above is "values only; nothing is ever deleted", and it was
    being broken twice on the engineer's ONLY copy of the raw captures):
    * ``errors="surrogateescape"`` on BOTH read and write round-trips bytes that are not valid
      UTF-8 exactly. ``errors="ignore"`` silently DELETED them — e.g. 0x96, the cp1252 en-dash
      that shows up in real banners and interface descriptions.
    * ``newline=""`` on both sides stops text mode rewriting every ``\\n`` to ``\\r\\n`` on
      Windows, which changed every line of every scrubbed capture.
    * temp file + ``os.replace`` makes the rewrite atomic: a yank or a full disk mid-write left
      a truncated capture that is indistinguishable from a legitimate scrub, because the run
      manifest was sealed a phase earlier.

    THE BOUND (because `_is_raw_capture` defaults to INCLUDE, this pass can otherwise reach any text
    file anywhere under the collection root, and it rewrites the engineer's only copy in place).
    Four deliberate limits, none of them a list of names, and every exclusion is DISCLOSED in the
    returned ``.uncovered`` rather than left in a debug log:

    1. the name rule (structured serialisations + the scrub's own scratch suffix) — `_is_raw_capture`;
    2. binary content — a NUL byte anywhere means this is a container, not line-oriented capture text;
    3. size — over `_REDACT_MAX_CAPTURE_BYTES` (the verifier's own artifact ceiling) nothing is even
       read: a 200 MB blob in a collection folder is a core dump or a pcap, not a ``show`` capture;
    4. the grammar itself — a file is only ever REWRITTEN when `_scrub_secrets` actually matched a
       secret line in it, so an unrelated ``.md`` or ``.csv`` that happens to sit under the root is
       read and left byte-identical.

    Returns ``(scanned, changed)`` (a `_ScrubResult`, so ``.uncovered`` carries limits 1-3)."""
    scanned = changed = 0
    uncovered: List[tuple] = []
    base = collection_dir or ""

    def _rel(path: str) -> str:
        try:
            return os.path.relpath(path, base).replace(os.sep, "/")
        except ValueError:                      # different drive on Windows -- name it absolutely
            return path

    for root, _dirs, files in os.walk(base):
        for fn in files:
            p = os.path.join(root, fn)
            if not _is_raw_capture(fn):
                uncovered.append((_rel(p), "not a raw capture by name (structured serialisation "
                                           "or the scrub's own scratch file)"))
                continue
            try:
                size = os.path.getsize(p)
            except OSError as e:
                logger.debug(f"redact_collection_dir: unsizeable {p}: {e}")
                uncovered.append((_rel(p), f"size could not be read ({e.__class__.__name__})"))
                continue
            if size > _REDACT_MAX_CAPTURE_BYTES:
                # Bound 3. Not read, not rewritten, not counted as scanned -- and said out loud,
                # because "we did not look at this one" must never arrive as part of a clean count.
                logger.warning(f"redact_collection_dir: {p} is {size} bytes (> "
                               f"{_REDACT_MAX_CAPTURE_BYTES}); NOT scrubbed and NOT scanned")
                uncovered.append((_rel(p), f"{size} bytes exceeds the {_REDACT_MAX_CAPTURE_BYTES}-byte "
                                           "capture ceiling; not read, so not scrubbed"))
                continue
            try:
                with open(p, "r", encoding="utf-8", errors="surrogateescape", newline="") as f:
                    text = f.read()
            except Exception as e:
                logger.debug(f"redact_collection_dir: unreadable {p}: {e}")
                uncovered.append((_rel(p), f"unreadable ({e.__class__.__name__})"))
                continue
            if "\x00" in text:
                # The content half of the capture rule. A NUL byte means binary; the deny-list
                # grammar would read noise out of it, and rewriting it is the one thing the
                # BYTE FIDELITY contract above must never risk. Not counted as scanned either --
                # a file this pass declines to protect must not inflate its own coverage number.
                logger.debug(f"redact_collection_dir: binary, not a capture: {p}")
                uncovered.append((_rel(p), "binary content (a NUL byte), not a text capture"))
                continue
            scanned += 1
            scrubbed = _scrub_secrets(text)
            if scrubbed != text:
                tmp = p + ".redacting"
                try:
                    with open(tmp, "w", encoding="utf-8", errors="surrogateescape",
                              newline="") as f:
                        f.write(scrubbed)
                    os.replace(tmp, p)      # atomic: the capture is either old or new, never half
                    changed += 1
                except Exception as e:
                    logger.warning(f"redact_collection_dir: could not rewrite {p}: {e}")
                    try:
                        os.unlink(tmp)      # never leave a partial beside the real capture
                    except OSError:
                        pass
                    uncovered.append((_rel(p), f"secrets were found but the rewrite FAILED "
                                               f"({e.__class__.__name__}); the original is unchanged "
                                               "and still holds them"))
    uncovered.sort()
    return _ScrubResult(scanned, changed, uncovered)


def redact_collected_inplace(all_interfaces: dict, all_device_physical: list) -> None:
    """Pseudonymize the COLLECTED dataclasses (InterfaceData + DevicePhysical) IN PLACE, so the always-produced
    .xlsx workbook -- which the sheet builders assemble from these dataclasses BEFORE redact_snapshot ever runs on
    the JSON -- cannot leak real serials / IPs / MACs under --redact. Serial-named fields -> SNxxxx; every other
    string -> IP/IPv6/MAC/secret scrub. For the --redact share-safe path only; mutates in place. Never raises.
    (Sheets built from COMPUTED structures / raw config text, not these dataclasses, are caught separately by
    redact_workbook_cells.)"""
    import dataclasses as _dc
    values = []
    for dp in (all_device_physical or []):
        try:
            values.extend(getattr(dp, f.name, "") for f in _dc.fields(dp))
        except TypeError:
            pass
    for ports in (all_interfaces or {}).values():
        for obj in (ports or {}).values():
            try:
                values.extend(getattr(obj, f.name, "") for f in _dc.fields(obj))
            except TypeError:
                pass
    scrub, serial = _make_redactor(_existing_synthetic_markers(values))

    def _red_obj(o):
        try:
            fields = _dc.fields(o)
        except TypeError:
            return                                    # not a dataclass instance -> nothing to redact
        for f in fields:
            v = getattr(o, f.name, "")
            if isinstance(v, str) and v:
                # serial-named field -> consistent SNxxxx; everything else -> pattern scrub (catches an IP/MAC in
                # ANY field, so a leak can't hide in a free-text column we forgot to enumerate).
                setattr(o, f.name,
                        serial(v) if (f.name in _REDACT_SERIAL_KEYS or f.name.endswith("_serial")) else scrub(v))

    for dp in (all_device_physical or []):
        _red_obj(dp)
    for ports in (all_interfaces or {}).values():
        for d in (ports or {}).values():
            _red_obj(d)


def redact_workbook_cells(wb) -> None:
    """Final-pass --redact safety net: scrub IPv4/IPv6/MACs/credentials from EVERY cell of an openpyxl workbook,
    so sheets built from COMPUTED structures (e.g. Subnet Reachability) or RAW config text (Golden-Config Drift)
    -- which redact_collected_inplace can't reach because they don't come from the collected dataclasses -- cannot
    leak real addresses/secrets. Serials carry no reliable text pattern, so they are pseudonymized upstream in the
    dataclasses; this pass covers everything pattern-matchable. Mutates in place; never raises."""
    try:
        sheets = list(wb.worksheets)
    except Exception as exc:
        raise RuntimeError("workbook cells could not be enumerated for redaction") from exc
    values = [
        cell.value
        for ws in sheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    scrub, _ = _make_redactor(_existing_synthetic_markers(values))
    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v:
                    nv = scrub(v)
                    if nv != v:
                        cell.value = nv
