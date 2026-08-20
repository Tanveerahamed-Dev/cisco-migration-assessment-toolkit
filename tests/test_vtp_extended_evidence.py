"""Release-1 VTP/VLAN source, delta, intent, and canonical-gate counterexamples."""
from __future__ import annotations

import copy
import json
from pathlib import Path

import pytest

from cisco_toolkit import vtp_extended as vtp_ext
from cisco_toolkit.capture_integrity import compute_capture_integrity_from_paths
from cisco_toolkit.html import compute_cutover_gate, compute_snapshot_delta, write_diff_workbook
from cisco_toolkit.protocol_assurance import normalize_change_intent, protocol_family_change_set
from cisco_toolkit.protocol_deltas import compute_vtp_safety_delta
from cisco_toolkit.vtp_extended import (
    VTP_EXTENDED_EVIDENCE_SCHEMA,
    compute_vtp_extended_evidence,
    embedded_vtp_extended_evidence,
    validate_vtp_extended_evidence,
)
from cisco_toolkit.vtp_safety import compute_vtp_safety_baseline


IOS_STATUS = """VTP Version capable             : 1 to 3
VTP version running             : {version}
VTP Domain Name                 : {domain}
VTP Pruning Mode                : Disabled
VTP Operating Mode              : {mode}
Configuration Revision          : {revision}
MD5 digest                      : 0xDE 0xAD 0xBE 0xEF
"""

NXOS_STATUS = """VTP Version capable             : 1 to 3
VTP version running             : {version}
VTP Domain Name                 : {domain}
VTP Operating Mode              : {mode}
Configuration Revision          : {revision}
Feature VLAN:
--------------
Number of existing VLANs        : 3
MD5 digest                      : 0x01 0x02 0x03 0x04
"""

VLAN_BRIEF = """VLAN Name                             Status    Ports
---- -------------------------------- --------- -------------------------------
1    default                          active    Gi1/0/1
10   USERS                            active    Gi1/0/2
1002 fddi-default                     act/unsup
"""


def _sources(tmp_path: Path, specs: dict[str, dict]) -> tuple[dict, dict, dict, dict]:
    paths: dict[str, dict[str, str]] = {}
    devices = {}
    for host, spec in specs.items():
        host_dir = tmp_path / host
        host_dir.mkdir(parents=True, exist_ok=True)
        paths[host] = {}
        for command, filename, body in (
            ("show vtp status", "show-vtp-status.txt", spec["status"]),
            ("show vlan brief", "show-vlan-brief.txt", spec["vlan"]),
            ("show running-config", "show-running-config.txt", spec["run"]),
        ):
            path = host_dir / filename
            path.write_text(body, encoding="utf-8")
            paths[host][command] = str(path)
        devices[host] = {"platform": spec.get("platform", "ios-xe")}
    integrity = compute_capture_integrity_from_paths(paths)
    protected = compute_vtp_safety_baseline(paths, integrity, devices)
    extended = compute_vtp_extended_evidence(paths, integrity, devices)
    return protected, extended, paths, integrity


def _spec(*, revision: int = 7, mode: str = "Server", domain: str = "CAMPUS",
          version: str = "2", vlan: str = VLAN_BRIEF, password: str = "",
          pruning: str = "", platform: str = "ios-xe", nxos: bool = False) -> dict:
    status_format = NXOS_STATUS if nxos else IOS_STATUS
    run = (
        "!Command: show running-config\n"
        "!Running configuration last done at: Wed Aug 20 09:05:00 2026\n"
        "!Time: Wed Aug 20 09:06:00 2026\n"
        "version 10.5(2) Bios:version 08.39\n"
        "hostname fixture\n"
        if nxos else "Building configuration...\nhostname fixture\n"
    )
    if pruning:
        run += pruning + "\n"
    if password:
        run += f"vtp password {password} hidden\n"
    run += "!\n" if nxos else "!\nend\n"
    return {
        "status": status_format.format(
            version=version, domain=domain, mode=mode, revision=revision),
        "vlan": vlan,
        "run": run,
        "platform": platform,
    }


def _snapshot(protected: dict, extended: dict, hosts: list[str]) -> dict:
    return {
        "devices": {host: {} for host in hosts},
        "vtp_safety_baseline": protected,
        "vtp_extended_evidence": extended,
    }


def _clean_ipv4() -> dict:
    return {
        "schema": "protocol_adjacency_delta/1",
        "summary": {"n_preserved": 1},
        "changes": [],
        "coverage_gaps": [],
    }


def _clean_snapshot_delta() -> dict:
    return {
        "verdict": "CLEAN",
        "verdict_display": "NO DELTA REGRESSION OBSERVED",
        "verdict_note": "clean",
        "protocol_adjacencies": {
            "gate": "PASS",
            "summary": {"n_state_regressed": 0, "n_coverage_gaps": 0,
                        "n_baseline_peers": 1},
        },
    }


def test_real_ios_and_nxos_shapes_are_source_bound_and_never_publish_secret(tmp_path: Path):
    secret = "DO-NOT-SERIALIZE-THIS-VTP-PASSWORD"
    protected, evidence, _paths, _integrity = _sources(tmp_path, {
        "ios-edge": _spec(password=secret, pruning="vtp pruning"),
        "nxos-edge": _spec(password=secret, pruning="vtp pruning", nxos=True,
                           platform="nxos"),
    })

    assert protected["rows"]
    assert evidence["schema"] == VTP_EXTENDED_EVIDENCE_SCHEMA
    assert evidence["owns_score"] is False and evidence["owns_verdict"] is False
    assert validate_vtp_extended_evidence(evidence, require_current_run=True)["valid"] is True
    assert evidence["summary"]["by_status"] == {
        "healthy": 2, "unsafe": 0, "not_verified": 0,
    }
    for row in evidence["rows"]:
        assert row["database_identity"] == "domain=CAMPUS;version=2"
        assert row["vlan_database_digest"].startswith("sha256:")
        assert row["vlan_count"] == 3
        assert row["pruning_state"] == "configured_enabled"
        assert row["authentication_configured"] is True
    serialized = json.dumps(evidence)
    assert secret not in serialized
    assert str(tmp_path) not in serialized
    assert "MD5 digest" not in serialized


def test_file_backed_nxos_full_config_without_ios_end_is_complete_and_unique(
        tmp_path: Path):
    fixture = (
        Path(__file__).parent / "fixtures" / "vtp" / "nxos" / "show_running_config.txt"
    ).read_text(encoding="utf-8")
    assert "!Command: show running-config" in fixture
    assert not any(line.strip().casefold() == "end" for line in fixture.splitlines())

    spec = _spec(nxos=True, platform="nxos")
    spec["run"] = fixture
    _protected, evidence, _paths, integrity = _sources(
        tmp_path / "single", {"nxos-edge": spec})

    inspection = next(
        row for row in integrity["inspections"]
        if row["host"] == "nxos-edge" and row["command"] == "show running-config"
    )
    assert inspection["status"] == "ok"
    row = evidence["rows"][0]
    assert row["status"] == "healthy"
    assert row["pruning_state"] == "configured_enabled"
    assert row["authentication_configured"] is True
    assert evidence["coverage"][0]["commands"]["show running-config"][
        "parser_status"] == "complete"

    duplicate = _spec(nxos=True, platform="nxos")
    duplicate["run"] = fixture + "\n" + fixture
    _protected, duplicated, _paths, _integrity = _sources(
        tmp_path / "duplicate", {"nxos-edge": duplicate})
    assert duplicated["rows"][0]["status"] == "not_verified"
    assert duplicated["coverage"][0]["commands"]["show running-config"][
        "parser_status"] == "rejected"


@pytest.mark.parametrize(
    ("platform", "nxos_shape"),
    (("ios-xe", True), ("nxos", False)),
    ids=("ios-cannot-borrow-nxos-envelope", "nxos-cannot-borrow-ios-envelope"),
)
def test_running_config_envelope_is_bound_to_declared_platform(
        tmp_path: Path, platform: str, nxos_shape: bool):
    spec = _spec(platform=platform, nxos=nxos_shape)
    _protected, evidence, _paths, _integrity = _sources(
        tmp_path, {"edge": spec})

    assert evidence["rows"][0]["status"] == "not_verified"
    assert "running_config_parser_not_verified" in {
        item["code"] for item in evidence["rows"][0]["findings"]
    }


def test_nxos_envelope_rejects_ios_terminal_end(tmp_path: Path):
    spec = _spec(platform="nxos", nxos=True)
    spec["run"] += "end\n"
    _protected, evidence, _paths, _integrity = _sources(
        tmp_path, {"nxos-edge": spec})

    assert evidence["rows"][0]["status"] == "not_verified"
    assert evidence["coverage"][0]["commands"]["show running-config"][
        "parser_status"] == "rejected"


def test_same_database_auth_and_vlan_digest_contradictions_are_current_unsafe(
        tmp_path: Path):
    changed_vlan = VLAN_BRIEF.replace("10   USERS", "10   GUESTS")
    _protected, evidence, _paths, _integrity = _sources(tmp_path, {
        "dist-a": _spec(password="configured", vlan=VLAN_BRIEF),
        "dist-b": _spec(password="", vlan=changed_vlan),
    })

    assert evidence["summary"]["n_unsafe"] == 2
    assert evidence["summary"]["n_authentication_contradictions"] == 2
    assert evidence["summary"]["n_vlan_digest_contradictions"] == 2
    for row in evidence["rows"]:
        assert row["status"] == "unsafe"
        assert {item["code"] for item in row["findings"]} >= {
            "authentication_contradiction", "vlan_database_digest_contradiction",
        }


@pytest.mark.parametrize("platform", ("eos", "ios-xr", "iosxr", "bios-appliance"))
def test_unsupported_platform_never_inherits_cisco_source_parity(
        tmp_path: Path, platform: str):
    _protected, evidence, _paths, _integrity = _sources(
        tmp_path, {"unsupported-edge": _spec(platform=platform)})

    row = evidence["rows"][0]
    assert row["platform"] == platform
    assert row["status"] == "not_verified"
    assert {item["code"] for item in row["findings"]} == {"platform_unsupported"}
    assert validate_vtp_extended_evidence(
        evidence, require_current_run=True)["valid"] is True


@pytest.mark.parametrize(
    "running_config",
    (
        "hostname edge\nvtp pruning\nvtp pruning\nend\n",
        (
            "hostname edge\nvtp pruning\nvtp password FIRST hidden\nend\n"
            "hostname edge\nvtp pruning\nvtp password SECOND hidden\nend\n"
        ),
    ),
    ids=("duplicate-directive", "concatenated-config-envelopes"),
)
def test_duplicate_or_concatenated_running_config_is_not_source_evidence(
        tmp_path: Path, running_config: str) -> None:
    spec = _spec()
    spec["run"] = running_config
    _protected, evidence, _paths, _integrity = _sources(
        tmp_path, {"edge": spec})

    assert evidence["rows"][0]["status"] == "not_verified"
    assert "running_config_parser_not_verified" in {
        item["code"] for item in evidence["rows"][0]["findings"]
    }


@pytest.mark.parametrize("mutation", ["missing", "truncated", "renamed", "malformed_vlan"])
def test_missing_truncated_renamed_and_malformed_source_leaves_withhold_evidence(
        tmp_path: Path, mutation: str):
    spec = _spec()
    if mutation == "truncated":
        spec["run"] = "hostname edge\nvtp pruning\n"  # no terminal ``end``
    elif mutation == "renamed":
        spec["status"] = spec["status"].replace("VTP Operating Mode", "VTP Operation Mood")
    elif mutation == "malformed_vlan":
        spec["vlan"] = spec["vlan"].replace("10   USERS", "10")
    _protected, _extended, paths, integrity = _sources(tmp_path, {"edge": spec})
    if mutation == "missing":
        paths["edge"].pop("show vlan brief")
        integrity = compute_capture_integrity_from_paths(paths)
    evidence = compute_vtp_extended_evidence(paths, integrity, {"edge": {"platform": "ios-xe"}})

    assert evidence["rows"][0]["status"] == "not_verified"
    assert evidence["summary"]["n_not_verified"] == 1
    assert evidence["coverage"][0]["status"] == "not_verified"
    assert validate_vtp_extended_evidence(evidence, require_current_run=True)["valid"] is True


def test_embedded_custody_and_closed_schema_mutations_fail_closed(tmp_path: Path):
    _protected, evidence, _paths, _integrity = _sources(tmp_path, {"edge": _spec()})
    embedded = embedded_vtp_extended_evidence(evidence)
    assert embedded["projection_custody"] == "embedded_unverified"
    assert validate_vtp_extended_evidence(embedded)["valid"] is True
    assert validate_vtp_extended_evidence(
        embedded, require_current_run=True)["valid"] is False

    renamed = copy.deepcopy(embedded)
    renamed["rows"][0]["vlan_digest"] = renamed["rows"][0].pop("vlan_database_digest")
    truncated = copy.deepcopy(embedded)
    truncated["coverage"][0]["commands"].pop("show running-config")
    malformed = copy.deepcopy(embedded)
    malformed["rows"][0]["authentication_configured"] = "yes"
    for hostile in (renamed, truncated, malformed):
        view = validate_vtp_extended_evidence(hostile)
        assert view["valid"] is False
        assert view["rows"] == [] and view["baseline"] == {}

    _protected, high, _paths, _integrity = _sources(
        tmp_path / "high-revision", {"edge": _spec(revision=100)})
    coherently_resealed = embedded_vtp_extended_evidence(high)
    row = coherently_resealed["rows"][0]
    row["status"] = "healthy"
    row["findings"] = []
    cell = coherently_resealed["coverage"][0]
    cell["status"] = "healthy"
    cell["finding_codes"] = []
    cell["projection_sha256"] = "sha256:" + vtp_ext._sha(vtp_ext._row_projection(row))
    coherently_resealed["summary"].update({
        "n_healthy": 1, "n_unsafe": 0, "n_high_revision_servers": 0,
        "by_status": {"healthy": 1, "unsafe": 0, "not_verified": 0},
        "baseline_sha256": "",
    })
    coherently_resealed["summary"]["baseline_sha256"] = vtp_ext._sha(
        vtp_ext._baseline_payload(coherently_resealed))
    assert validate_vtp_extended_evidence(coherently_resealed)["reason"] == (
        "extended_vtp_safety_findings_do_not_reconcile")


def test_legacy_only_and_after_capture_loss_abstain_instead_of_passing(tmp_path: Path):
    before_v1, before_ext, _paths, _integrity = _sources(
        tmp_path / "before", {"edge": _spec()})
    after_v1, after_ext, _paths, _integrity = _sources(
        tmp_path / "after", {"edge": _spec()})

    legacy_only = compute_vtp_safety_delta(
        {"devices": {"edge": {}}, "vtp_safety_baseline": before_v1},
        _snapshot(after_v1, after_ext, ["edge"]),
    )
    assert legacy_only["changes"][0]["transition"] == "not_comparable"
    assert legacy_only["changes"][0]["decision_effect"] == "not_verified"

    lost = compute_vtp_safety_delta(
        _snapshot(before_v1, before_ext, ["edge"]),
        {"devices": {"edge": {}}, "vtp_safety_baseline": after_v1},
    )
    assert lost["changes"][0]["transition"] == "coverage_lost"
    assert lost["changes"][0]["decision_effect"] == "not_verified"


def test_revision_reset_requires_exact_subject_cutover_intent_and_gate_reconciles(
        tmp_path: Path):
    before_v1, before_ext, _paths, _integrity = _sources(
        tmp_path / "before", {"edge": _spec(revision=9)})
    after_v1, after_ext, _paths, _integrity = _sources(
        tmp_path / "after", {"edge": _spec(revision=0)})
    native = compute_vtp_safety_delta(
        _snapshot(before_v1, before_ext, ["edge"]),
        _snapshot(after_v1, after_ext, ["edge"]),
    )
    change = native["changes"][0]
    assert change["transition"] == "intent_changed"
    assert change["decision_effect"] == "review"
    assert change["change_kind"] == "revision_decrease_observed"

    wildcard = protocol_family_change_set(_clean_ipv4(), {
        "expected_changes": [{
            "family": "vtp_safety", "transitions": ["intent_changed"],
            "subjects": [], "reason": "generic VTP movement",
        }],
    }, native_deltas=[native])
    wildcard_row = next(
        row for family in wildcard["families"] if family["family"] == "vtp_safety"
        for row in family["changes"]
    )
    assert wildcard_row["expected"] is False
    assert wildcard_row["decision_effect"] == "review"
    assert compute_cutover_gate(
        _clean_snapshot_delta(), {"verdict": "PASS", "verdict_note": "clean"},
        protocol_family_changes=wildcard,
    )["verdict"] == "REVIEW"

    exact_without_reset_intent = protocol_family_change_set(_clean_ipv4(), {
        "expected_changes": [{
            "family": "vtp_safety", "transitions": ["intent_changed"],
            "subjects": ["edge"], "reason": "generic exact-subject VTP movement",
        }],
    }, native_deltas=[native])
    untyped_row = next(
        row for family in exact_without_reset_intent["families"]
        if family["family"] == "vtp_safety"
        for row in family["changes"]
    )
    assert untyped_row["expected"] is False
    assert untyped_row["decision_effect"] == "review"

    exact = protocol_family_change_set(_clean_ipv4(), {
        "expected_changes": [{
            "family": "vtp_safety", "transitions": ["intent_changed"],
            "subjects": ["edge"], "intent_kind": "revision_reset",
            "reason": "authorized revision reset",
        }],
    }, native_deltas=[native])
    exact_row = next(
        row for family in exact["families"] if family["family"] == "vtp_safety"
        for row in family["changes"]
    )
    assert exact_row["expected"] is True
    assert exact_row["decision_effect"] == "none"
    assert exact_row["change_kind"] == "revision_decrease_observed"
    assert compute_cutover_gate(
        _clean_snapshot_delta(), {"verdict": "PASS", "verdict_note": "clean"},
        protocol_family_changes=exact,
    )["verdict"] == "PASS"

    invalid_intents = (
        {"expected_changes": [{
            "family": "vtp_safety", "transitions": ["intent_changed"],
            "subjects": [], "intent_kind": "revision_reset",
        }]},
        {"expected_changes": [{
            "family": "bgp_configured_peer", "transitions": ["intent_changed"],
            "subjects": ["edge"], "intent_kind": "revision_reset",
        }]},
        {"expected_changes": [{
            "family": "vtp_safety", "transitions": ["intent_changed"],
            "subjects": ["edge"], "intent_kind": "guessed_reset",
        }]},
    )
    for raw in invalid_intents:
        receipt = normalize_change_intent(raw, binding={})
        assert receipt["valid"] is False
        assert any("intent_kind" in failure or "revision_reset" in failure
                   for failure in receipt["failures"])

    pruning_v1, pruning_ext, _paths, _integrity = _sources(
        tmp_path / "pruning", {"edge": _spec(revision=9, pruning="vtp pruning")})
    ordinary_movement = compute_vtp_safety_delta(
        _snapshot(before_v1, before_ext, ["edge"]),
        _snapshot(pruning_v1, pruning_ext, ["edge"]),
    )
    mismatched_intent = protocol_family_change_set(_clean_ipv4(), {
        "expected_changes": [{
            "family": "vtp_safety", "transitions": ["intent_changed"],
            "subjects": ["edge"], "intent_kind": "revision_reset",
            "reason": "must not authorize unrelated pruning movement",
        }],
    }, native_deltas=[ordinary_movement])
    mismatched_row = next(
        row for family in mismatched_intent["families"]
        if family["family"] == "vtp_safety"
        for row in family["changes"]
    )
    assert mismatched_row["change_kind"] == "configuration_movement"
    assert mismatched_row["expected"] is False
    assert mismatched_row["decision_effect"] == "review"


def test_revision_reset_intent_cannot_hide_simultaneous_vtp_or_vlan_movement(
        tmp_path: Path) -> None:
    before_v1, before_ext, _paths, _integrity = _sources(
        tmp_path / "before-mixed", {"edge": _spec(revision=9)})
    changed_vlan = VLAN_BRIEF.replace("10   USERS", "10   GUESTS")
    after_v1, after_ext, _paths, _integrity = _sources(
        tmp_path / "after-mixed", {"edge": _spec(
            revision=0,
            vlan=changed_vlan,
            pruning="vtp pruning",
            password="new-auth",
        )})
    native = compute_vtp_safety_delta(
        _snapshot(before_v1, before_ext, ["edge"]),
        _snapshot(after_v1, after_ext, ["edge"]),
    )
    assert native["changes"][0]["change_kind"] == "configuration_movement"

    families = protocol_family_change_set(_clean_ipv4(), {
        "expected_changes": [{
            "family": "vtp_safety",
            "transitions": ["intent_changed"],
            "subjects": ["edge"],
            "intent_kind": "revision_reset",
            "reason": "revision reset does not authorize other changed axes",
        }],
    }, native_deltas=[native])
    row = next(
        row for family in families["families"] if family["family"] == "vtp_safety"
        for row in family["changes"]
    )
    assert row["expected"] is False
    assert row["decision_effect"] == "review"
    assert compute_cutover_gate(
        _clean_snapshot_delta(), {"verdict": "PASS", "verdict_note": "clean"},
        protocol_family_changes=families,
    )["verdict"] == "REVIEW"


def test_current_high_revision_and_appeared_unsafe_always_block_canonical_gate(
        tmp_path: Path):
    before_v1, before_ext, _paths, _integrity = _sources(
        tmp_path / "before", {"edge": _spec(revision=100)})
    after_v1, after_ext, _paths, _integrity = _sources(
        tmp_path / "after", {"edge": _spec(revision=101, pruning="vtp pruning")})
    native = compute_vtp_safety_delta(
        _snapshot(before_v1, before_ext, ["edge"]),
        _snapshot(after_v1, after_ext, ["edge"]),
    )
    row = native["changes"][0]
    assert row["transition"] == "unchanged_degraded"
    assert row["decision_effect"] == "block"
    changes = protocol_family_change_set(_clean_ipv4(), {
        "expected_changes": [{
            "family": "vtp_safety", "transitions": ["unchanged_degraded"],
            "subjects": ["edge"], "reason": "must not weaken current fault",
        }],
    }, native_deltas=[native])
    assert compute_cutover_gate(
        _clean_snapshot_delta(), {"verdict": "PASS", "verdict_note": "clean"},
        protocol_family_changes=changes,
    )["verdict"] == "REGRESSED"

    empty_v1, empty_ext, _paths, _integrity = _sources(tmp_path / "empty", {})
    appeared = compute_vtp_safety_delta(
        _snapshot(empty_v1, empty_ext, []),
        _snapshot(after_v1, after_ext, ["edge"]),
    )
    assert appeared["changes"][0]["transition"] == "appeared"
    assert appeared["changes"][0]["decision_effect"] == "block"


def test_generic_protocol_family_workbook_projects_complete_extended_vtp_state(
        tmp_path: Path):
    from openpyxl import load_workbook
    from cisco_toolkit.precert import compute_precert
    from tests.test_compare_cutover_gate_cli import _snapshot as comparison_snapshot

    before_v1, before_ext, _paths, _integrity = _sources(tmp_path / "before-workbook", {
        "R1": _spec(), "R2": _spec(nxos=True, platform="nxos"),
    })
    after_v1, after_ext, _paths, _integrity = _sources(tmp_path / "after-workbook", {
        "R1": _spec(pruning="vtp pruning"),
        "R2": _spec(nxos=True, platform="nxos"),
    })
    before = comparison_snapshot("FULL/DR", "2026-08-20T00:00:00")
    after = comparison_snapshot("FULL/DR", "2026-08-20T00:05:00")
    before.update({"vtp_safety_baseline": before_v1,
                   "vtp_extended_evidence": before_ext})
    after.update({"vtp_safety_baseline": after_v1,
                  "vtp_extended_evidence": after_ext})
    delta = compute_snapshot_delta(before, after)
    native = compute_vtp_safety_delta(before, after)
    families = protocol_family_change_set(
        delta["protocol_adjacencies"], {"expected_changes": []},
        native_deltas=[native],
    )
    expected = compute_cutover_gate(
        delta, compute_precert(before, after), protocol_family_changes=families)
    output = tmp_path / "vtp-family-changes.xlsx"

    actual = write_diff_workbook(
        before, after, str(output), precert=compute_precert(before, after),
        protocol_families=families,
    )
    assert actual == expected
    assert actual["verdict"] == "REVIEW"
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert "Protocol Adjacency Delta" in workbook.sheetnames
        rows = list(workbook["Protocol Family Changes"].iter_rows(values_only=True))
    finally:
        workbook.close()
    vtp_row = next(row for row in rows[1:] if row[0] == "vtp_safety" and row[2] == "R1")
    assert vtp_row[3] == "intent_changed" and vtp_row[5] == "review"
    assert '"pruning_state":"not_configured"' in vtp_row[7]
    assert '"pruning_state":"configured_enabled"' in vtp_row[8]
    assert '"vlan_database_digest":"sha256:' in vtp_row[7]
    assert '"authentication_configured":false' in vtp_row[8]
    totals = next(row for row in rows if row[0] == "FULL TOTALS")
    assert totals[8] == "omitted 0"
