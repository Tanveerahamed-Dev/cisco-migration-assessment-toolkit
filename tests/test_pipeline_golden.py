"""End-to-end regression test: run the real offline pipeline and assert the
JSON snapshot (the HTML explorer's contract) and the Excel sheet-name/header
schema are unchanged.

The snapshot is the single most important stability contract in the repo, so we
freeze it as a golden file. To intentionally update the goldens after a
reviewed change:

    UPDATE_GOLDEN=1 python -m pytest tests/test_pipeline_golden.py

Harness guarantees (pinned by tests/test_golden_guard.py — Plan A / Move-0.1):
- a MISSING golden FAILS (it is a broken contract, never a fresh baseline);
- UPDATE_GOLDEN=1 refuses to SHRINK the contract vs the git-HEAD baseline
  (removed snapshot sections / sheets / header cells) unless the removal is
  made explicit with ALLOW_GOLDEN_SHRINK=1.

Determinism: we run with --workers 1 (sequential) and strip the only volatile
field (`generated_at`) before comparing.
"""
import json
import os
import subprocess
import sys

import pytest
from openpyxl import Workbook, load_workbook

import synthetic_fixtures as fx

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py")
GOLDEN_DIR = os.path.join(ROOT, "tests", "golden")


def _make_template(path):
    """Minimal template workbook: the loader only needs a header row containing
    hostname/port/status; the script appends the rest of the columns itself."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Interface Data"
    ws.append(["Hostname", "Port", "Status"])
    wb.save(path)


def _run_pipeline(tmp_path, out_xlsx=None, extra_args=None):
    collection = fx.write_collection(str(tmp_path / "collection"))
    devices = tmp_path / "devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "template.xlsx"
    _make_template(str(template))
    if out_xlsx is None:
        out_xlsx = tmp_path / "out.xlsx"

    cmd = [sys.executable, SCRIPT,
           "--no-collect", "--collection-dir", collection,
           "--devices-file", str(devices), "--template", str(template),
           "--output", str(out_xlsx), "--no-html", "--workers", "1"]
    if extra_args:
        cmd += list(extra_args)
    proc = subprocess.run(cmd, cwd=str(tmp_path), capture_output=True, text=True, timeout=300)
    assert proc.returncode == 0, f"pipeline failed:\nSTDOUT\n{proc.stdout}\nSTDERR\n{proc.stderr}"
    snap_path = os.path.splitext(str(out_xlsx))[0] + ".snapshot.json"
    assert os.path.isfile(snap_path), "snapshot.json was not produced"
    with open(snap_path, encoding="utf-8") as f:
        snap = json.load(f)
    snap.pop("generated_at", None)            # volatile: wall-clock timestamp
    # collection-time provenance: now() on a live run, dir-stamp/mtime on --no-collect -> volatile like
    # generated_at; exclude it (its consumer lifecycle_risk is already excluded below). (provenance R2-1-01)
    snap.pop("collected_at", None)
    # lifecycle_risk is date-dependent (bands/years shift relative to 'today') -> exclude from the frozen
    # golden; its logic is pinned deterministically by tests/test_lifecycle.py with a fixed asof. (V3.23.117)
    snap.pop("lifecycle_risk", None)
    # executive_brief rolls up lifecycle (its EoL headline is date-relative) -> exclude too; pinned by
    # tests/test_executive_brief.py with synthetic summaries. (V3.23.120)
    snap.pop("executive_brief", None)
    # device_dossiers embeds the EoL band per asset (eol_band / exposure labels / risk_band shift as
    # dates pass) -> exclude like its lifecycle source; pinned by tests/test_device_dossiers.py with
    # synthetic axes. Its PUNCH-LIST fold stays frozen: the CR basis text is deliberately band-agnostic
    # ("past/near end-of-support"), so band transitions never reword a folded row. (V3.23.172)
    snap.pop("device_dossiers", None)
    # design_blueprint folds the date-relative lifecycle/EoL bands (its EoL decision count shifts as dates
    # pass) -> exclude like its lifecycle source; the blueprint logic is pinned deterministically by
    # tests/test_design_blueprint.py and its SSOT publish by tests/test_pipeline_inprocess.py. (design engine)
    snap.pop("design_blueprint", None)
    # design_nrfu is derived purely from design_blueprint (same date-relative folding) -> exclude it too;
    # its SSOT publish is locked by tests/test_pipeline_inprocess.py alongside the blueprint. (design engine)
    snap.pop("design_nrfu", None)
    # architecture_coverage is derived from design_blueprint (its findings shift with the date-relative blueprint)
    # -> exclude it too; its SSOT publish is locked by tests/test_pipeline_inprocess.py. (architecture coverage)
    snap.pop("architecture_coverage", None)
    # coverage_matrix (Plan-A #5) is COMPOSED from architecture_coverage -> inherits its date-relativity;
    # exclude it too, its SSOT publish is locked in tests/test_pipeline_inprocess.py. (coverage matrix)
    snap.pop("coverage_matrix", None)
    # fact_lineage (J2) embeds the canonical headline VALUES (n_past_ldos / avg_health / ... are
    # date-relative via lifecycle/executive_brief) -> exclude like its value sources; its shape +
    # coverage-honest state are pinned deterministically by tests/test_fact_lineage.py. (fact lineage)
    snap.pop("fact_lineage", None)
    # schema_census (J3) is intentionally KEPT in the golden: it records only presence/absence +
    # structural shape of the sections (no dates, no headline values), so it is non-volatile and
    # frozen here; its logic is also pinned by tests/test_schema_census.py.
    # attestation.generated_at is the one NESTED wall-clock stamp; strip just it (like the
    # top-level generated_at) — the re-derived trust CLAIMS themselves are deterministic and
    # must stay frozen in the golden. (roadmap D3)
    if isinstance(snap.get("attestation"), dict):
        snap["attestation"].pop("generated_at", None)
    return snap, str(out_xlsx)


def _sheet_schema(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    schema = {}
    for name in wb.sheetnames:
        ws = wb[name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))] if ws.max_row else []
        schema[name] = header
    wb.close()
    return schema


def _git_head_golden(name):
    """The golden as committed at git HEAD — the shrink guard's baseline. None when it
    is not tracked there (brand-new golden) or git is unavailable: nothing to shrink
    against, so the guard stands down rather than blocking legitimate first baselines."""
    try:
        proc = subprocess.run(["git", "show", f"HEAD:tests/golden/{name}"], cwd=ROOT,
                              capture_output=True, text=True, timeout=30)
        if proc.returncode != 0:
            return None
        return json.loads(proc.stdout)
    except Exception:
        return None


def _contract_shrinkage(name, baseline, produced):
    """Contract surface REMOVED between the HEAD golden and its replacement: top-level
    keys (snapshot sections / workbook sheets) always; for the sheet schema also header
    cells lost from a retained sheet. Additions and value changes are the additive norm
    and are not shrinkage (exact equality is the golden assertion's job)."""
    if not isinstance(baseline, dict) or not isinstance(produced, dict):
        return []
    lost = [f"top-level key removed: {k!r}" for k in baseline if k not in produced]
    if name == "sheet_schema.json":
        for sheet, header in baseline.items():
            new_header = produced.get(sheet)
            if isinstance(header, list) and isinstance(new_header, list):
                lost += [f"sheet {sheet!r} lost header {h!r}"
                         for h in header if h not in new_header]
    return lost


def _golden(name, produced):
    path = os.path.join(GOLDEN_DIR, name)
    if os.environ.get("UPDATE_GOLDEN") != "1":
        if not os.path.isfile(path):
            pytest.fail(
                f"golden {name} is MISSING — refusing to auto-generate a fresh baseline "
                f"(that would silently re-bless the contract). Restore it "
                f"(git checkout -- tests/golden/{name}) or regenerate deliberately with "
                f"UPDATE_GOLDEN=1.", pytrace=False)
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    # UPDATE_GOLDEN=1: a deliberate re-baseline — but refuse to silently SHRINK the
    # contract vs git HEAD; removals must be explicit (ALLOW_GOLDEN_SHRINK=1).
    baseline = _git_head_golden(name)
    if baseline is not None and os.environ.get("ALLOW_GOLDEN_SHRINK") != "1":
        lost = _contract_shrinkage(name, baseline, produced)
        if lost:
            pytest.fail(
                f"UPDATE_GOLDEN would SHRINK the {name} contract vs git HEAD:\n  - "
                + "\n  - ".join(lost)
                + "\nRemoving contract surface must be explicit: re-run with "
                  "ALLOW_GOLDEN_SHRINK=1 after review.", pytrace=False)
    os.makedirs(GOLDEN_DIR, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        # sort_keys=False: preserve meaningful order (snapshot key order is
        # code-defined; Excel sheet order is the workbook tab order).
        json.dump(produced, f, indent=1, sort_keys=False)
    return None


def test_snapshot_matches_golden(tmp_path):
    snap, _xlsx = _run_pipeline(tmp_path)
    golden = _golden("snapshot.json", snap)
    if golden is None:
        return
    # compare as objects (key order irrelevant); pinpoint drift by section
    assert set(snap) == set(golden), "snapshot top-level keys changed"
    for key in golden:
        assert snap[key] == golden[key], f"snapshot section '{key}' changed vs golden"


def test_excel_sheet_schema_matches_golden(tmp_path):
    _snap, xlsx = _run_pipeline(tmp_path)
    schema = _sheet_schema(xlsx)
    golden = _golden("sheet_schema.json", schema)
    if golden is None:
        return
    assert list(schema.keys()) == list(golden.keys()), "Excel sheet set/order changed"
    for sheet, header in golden.items():
        assert schema[sheet] == header, f"header row of sheet '{sheet}' changed"


def test_move_group_endpoint_label_is_honest_per_switch_mac_sum(tmp_path):
    """B3 (audit fix): move_groups[].endpoints is the per-SWITCH sum of learned MACs (an endpoint seen on
    N of the group's switches counts N times), so a large L2-coupled group can EXCEED the distinct fleet
    endpoint total (executive_brief.scale.n_endpoints / Endpoint Census). The workbook must label it as a
    per-switch MAC sum -- never the bare 'Endpoints'/'endpoint(s)' -- so it cannot be misread as a
    competing fleet endpoint total. Refutes the relabel silently reverting."""
    _snap, xlsx = _run_pipeline(tmp_path)
    wb = load_workbook(xlsx, read_only=True)
    try:
        mg_hdr = [c.value for c in next(wb["Move Groups"].iter_rows(min_row=1, max_row=1))]
        assert "# Endpoint MACs (per-switch sum)" in mg_hdr, f"Move Groups header not relabeled: {mg_hdr}"
        assert "# Endpoints" not in mg_hdr, "bare '# Endpoints' must not return (misreads as fleet total)"
        # Migration Scenarios carries the same per-group figure; its column header is row 2 (row 1 is the
        # fleet-recommendation banner), so scan the top rows for the honest label.
        ms_top = [str(v) for row in wb["Migration Scenarios"].iter_rows(min_row=1, max_row=3, values_only=True)
                  for v in row if v]
        assert any("Endpoint MACs (per-switch sum)" in v for v in ms_top), \
            f"Migration Scenarios endpoint column not relabeled: {ms_top}"
    finally:
        wb.close()


def test_run_manifest_emitted_and_sealed(tmp_path):
    """roadmap D2: the pipeline emits a sealed run-manifest (chain-of-custody) next to the workbook —
    a hash-chained step ledger + per-artifact sha256 + chain_root that verify_manifest() reconciles."""
    from cisco_toolkit import manifest as M
    _snap, xlsx = _run_pipeline(tmp_path)
    man_path = os.path.splitext(xlsx)[0] + ".run_manifest.json"
    assert os.path.isfile(man_path), "run_manifest.json was not produced"
    with open(man_path, encoding="utf-8") as f:
        man = json.load(f)
    assert man.get("chain_root") and man.get("artifacts"), "manifest missing seal/artifacts"
    ok, broken = M.verify_manifest(man)
    assert ok, f"manifest chain broken at rows {broken}"
    names = [a["name"] for a in man["artifacts"]]
    assert any(n.endswith(".snapshot.json") for n in names), f"snapshot not hashed: {names}"
    assert any(n.endswith(".xlsx") for n in names), f"workbook not hashed: {names}"
    assert "abstention_ledger" in man        # coverage-honest provenance is part of the seal


def test_shipped_verify_verb_passes_the_real_pipeline_manifest(tmp_path, capsys):
    """The CI check on the seal. The test above calls the library function; this runs the command an
    AUDITOR actually has — `python -m cisco_toolkit.manifest verify` — against a manifest the real
    pipeline produced, not a hand-built fixture. A verb that only ever sees synthetic manifests can
    agree with a producer bug; this is the arm that would catch the engine emitting a chain the
    shipped verifier rejects.

    Also proves the artifact hashes reconcile against the deliverables ON DISK, which is the actual
    chain-of-custody question ("is this the workbook that was sealed?") and which no test covered."""
    from cisco_toolkit import manifest as M
    _snap, xlsx = _run_pipeline(tmp_path)
    man_path = os.path.splitext(xlsx)[0] + ".run_manifest.json"

    assert M.main(["verify", man_path]) == 0, capsys.readouterr().out
    assert capsys.readouterr().out.startswith("OK: ")
    # --artifacts re-hashes every deliverable the engine listed, from the run's own output folder.
    assert M.main(["verify", man_path, "--artifacts"]) == 0, capsys.readouterr().out
    assert "hash to the seal" in capsys.readouterr().out


def test_shipped_verify_verb_rejects_a_tampered_pipeline_manifest(tmp_path, capsys):
    """Non-vacuity for the check above: prove the exit code moves. Two independent edits an auditor
    must catch — a rewritten step, and a deliverable swapped after the run."""
    from cisco_toolkit import manifest as M
    _snap, xlsx = _run_pipeline(tmp_path)
    man_path = os.path.splitext(xlsx)[0] + ".run_manifest.json"
    with open(man_path, encoding="utf-8") as f:
        man = json.load(f)

    # (1) rewrite a sealed step — the "we collected everything" edit
    tampered = os.path.join(str(tmp_path), "tampered.run_manifest.json")
    edited = json.loads(json.dumps(man))
    edited["chain"][0]["collected"] = 99999
    with open(tampered, "w", encoding="utf-8") as f:
        json.dump(edited, f)
    assert M.main(["verify", tampered]) == 4
    assert "INTEGRITY" in capsys.readouterr().out

    # (2) leave the ledger alone, swap a delivered file — caught only by --artifacts
    with open(xlsx, "ab") as f:
        f.write(b"\n<appended after sealing>")
    assert M.main(["verify", man_path]) == 0, "the ledger itself is untouched, so plain verify passes"
    capsys.readouterr()
    assert M.main(["verify", man_path, "--artifacts"]) == 4
    assert "[MISMATCH]" in capsys.readouterr().out


def test_import_inventory_reconcile(tmp_path):
    """roadmap B: --import-inventory ingests a declared inventory (CMDB/NetBox CSV) and reconciles it against
    the collected evidence -> snap['external_reconcile'] + a 'SoT Reconcile' workbook sheet (opt-in)."""
    inv = tmp_path / "cmdb.csv"
    inv.write_text("hostname,device_type\nGHOST-NOT-IN-FLEET,C9999\n", encoding="utf-8")
    snap, xlsx = _run_pipeline(tmp_path, extra_args=["--import-inventory", str(inv)])
    er = snap.get("external_reconcile")
    assert er and er.get("summary"), "external_reconcile missing from snapshot"
    assert er["summary"]["MISSING_DEVICE"] == 1          # the declared ghost is not observed
    assert er["summary"]["UNDOCUMENTED_DEVICE"] >= 1     # observed devices absent from the (ghost-only) SoT
    wb = load_workbook(xlsx, read_only=True)
    try:
        assert "SoT Reconcile" in wb.sheetnames
    finally:
        wb.close()


def test_default_run_optin_engines_absent_but_capture_integrity_present(tmp_path):
    """Opt-in engines (reconcile / what-if / path-intents / assert-pack) add nothing on a default run, so the
    golden stays untouched; the always-on capture-integrity key + sheet ARE present."""
    snap, xlsx = _run_pipeline(tmp_path)
    for key in ("external_reconcile", "whatif", "path_intents", "state_assertions"):
        assert key not in snap, f"{key} must be opt-in / absent by default"
    assert "capture_integrity" in snap            # always-on
    wb = load_workbook(xlsx, read_only=True)
    try:
        for sheet in ("SoT Reconcile", "Failure What-If", "Path Assertions"):
            assert sheet not in wb.sheetnames
        assert "Capture Integrity" in wb.sheetnames   # always-on
    finally:
        wb.close()


def test_optin_engines_emit_keys_and_sheets(tmp_path):
    """roadmap G4 / G3 / A1+H2: --scenario, --path-intents and --assert-pack each compute over the snapshot and
    emit their result (sheet and/or snapshot key) when supplied."""
    scen = tmp_path / "scen.json"
    scen.write_text(json.dumps([{"name": "edge-fail", "failures": [{"type": "node", "id": "no-such-host"}]}]), encoding="utf-8")
    intents = tmp_path / "intents.json"
    intents.write_text(json.dumps([{"id": "i1", "src": "10.0.0.1", "dst": "10.0.0.2", "expect": "REACHES"}]), encoding="utf-8")
    pack = tmp_path / "pack.json"
    pack.write_text(json.dumps({"assertions": [
        {"id": "a1", "subject": "collection_completeness", "all_of": [{"type": "contains", "value": "summary"}]}]}), encoding="utf-8")
    snap, xlsx = _run_pipeline(tmp_path, extra_args=[
        "--scenario", str(scen), "--path-intents", str(intents), "--assert-pack", str(pack)])
    assert isinstance(snap.get("whatif"), list)
    assert "results" in (snap.get("path_intents") or {})
    sa = snap.get("state_assertions") or {}
    assert "summary" in sa
    a1 = [r for r in sa.get("results", []) if r.get("id") == "a1"]
    assert a1 and a1[0]["status"] == "pass"        # 'collection_completeness' exists -> contains 'summary' -> pass
    wb = load_workbook(xlsx, read_only=True)
    try:
        assert "Failure What-If" in wb.sheetnames and "Path Assertions" in wb.sheetnames
    finally:
        wb.close()


def test_missing_output_directory_is_created(tmp_path):
    """FIX-V3.23.103: a non-existent --output directory must be created up-front,
    not crash openpyxl's save() AFTER all the heavy compute. Point --output at a
    nested directory that does not exist and assert every output lands there."""
    out_xlsx = tmp_path / "does" / "not" / "exist" / "out.xlsx"
    assert not out_xlsx.parent.exists()           # precondition: dir is missing
    _snap, xlsx = _run_pipeline(tmp_path, out_xlsx=out_xlsx)
    assert os.path.isfile(xlsx), "workbook was not written into the created directory"
    snap_path = os.path.splitext(xlsx)[0] + ".snapshot.json"
    assert os.path.isfile(snap_path), "snapshot was not written into the created directory"
