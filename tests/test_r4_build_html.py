"""Round-4 whole-repo review: cisco_toolkit/build.py + cisco_toolkit/html.py.

Every defect pinned here was DISCOVERED and MEASURED against the real 253-device offline
collection in `migration_collection_20260613_063201/` (gitignored, so it cannot be a test
input) and its shipped snapshot; the counts quoted in each test are from that run. The tests
themselves drive the same producer path CI has: `tests/synthetic_fixtures.py` writes real
device-output TEXT into a collection dir and the real parsers read it back, plus the tracked
`webapp/sample_data/sample_fleet.snapshot.json`. Nothing here hand-builds a parsed dict in the
shape the code under test expects.
"""
import json
import os

import pytest

from cisco_toolkit import build, html

import synthetic_fixtures as fx


# --------------------------------------------------------------------------------------
# helpers: write ONE device's fixture captures to a temp collection dir, optionally with a
# command's capture overridden (to model a real partial collection) or its text rewritten.
# --------------------------------------------------------------------------------------
def _write_device(root, host, overrides=None, drop=()):
    _plat, outputs = fx.COLLECTIONS[host]
    dev = os.path.join(str(root), host)
    os.makedirs(dev, exist_ok=True)
    c2f = {}
    for cmd, text in outputs.items():
        if cmd in drop:
            continue
        text = (overrides or {}).get(cmd, text)
        p = os.path.join(dev, fx.cmd_filename(cmd))
        with open(p, "w", encoding="utf-8") as f:
            f.write(text)
        c2f[cmd] = p
    return c2f


# The two running-config captures build_interfaces step 4 reads. `show running-config` (the
# plain form) is NOT one of them -- it only feeds the global-bpduguard scan.
_RUNCFG_IFACE_CMDS = ("show running-config interface", "show running-config | section ^interface")

# A verbatim Cisco refusal. _load_cmd_output screens these (cmdio._CISCO_ERRORS) and returns "",
# so a capture that exists but was refused is indistinguishable from one that was never taken --
# which is exactly the state a read-only TACACS policy produces for `show running-config`.
_AUTHZ_DENIED = "command authorization failed\n"

# Real `show vrf interface` output (the NX-OS column layout, copied from the [HISTORY-REDACTED] collection and
# re-pointed at the fixture's own VLAN ids). This is the capture that put the SVI in the model on
# the real device that reproduced BH-1: with the run-config refused, `show vrf interface` was the
# ONLY source naming Vlan1/Vlan25, and Vlan25 was also a route out-interface.
_VRF_IFACE = (
    "Interface               VRF-Name                        VRF-ID  Site-of-Origin      \n"
    "Vlan10                  default                              1  --                  \n"
    "Vlan20                  default                              1  --                  \n"
    "Vlan30                  default                              1  --                  \n"
)


# ======================================================================================
# BH-1  build.py :: build_interfaces -- an SVI with NO configured address aborted the
#       WHOLE device inside a sort key, and the multi-worker parse loop then dropped it.
# ======================================================================================
def test_svi_without_an_ip_does_not_abort_the_whole_device(tmp_path):
    """`(d.svi_ip or '').split()[0]` raises IndexError on an SVI carrying no address.

    It runs inside `svi_routes.sort(key=_rank)`, so the exception leaves build_interfaces
    entirely. COLLECT_PARSE_V3_23_0.parse_one is wrapped by a bare `except Exception` in the
    ThreadPoolExecutor branch, so the device is silently omitted from `all_interfaces` and
    lands in the snapshot with ZERO interfaces -- 'this switch has no ports', not 'the parse
    failed'. Precondition (reproduced verbatim on a real [HISTORY-REDACTED] device): the run-config channel
    did not land, so svi_ip is '' on every SVI, while `show vrf interface` still names the
    SVIs and `show ip route` still exits through one of them.
    """
    c2f = _write_device(tmp_path, "core1",
                        overrides={c: _AUTHZ_DENIED for c in _RUNCFG_IFACE_CMDS})
    p = os.path.join(str(tmp_path), "core1", fx.cmd_filename("show vrf interface"))
    with open(p, "w", encoding="utf-8") as f:
        f.write(_VRF_IFACE)
    c2f["show vrf interface"] = p

    ifaces = build.build_interfaces("core1", "ios", c2f)

    assert ifaces, "device was dropped entirely"
    # the SVIs are still discovered, still address-less, and the device did not vanish
    svis = [x for x in ifaces if x.lower().startswith("vlan")]
    assert svis, "no SVI survived the partial collection"
    assert all(not (ifaces[x].svi_ip or "").strip() for x in svis)


def test_svi_route_ranking_still_works_when_the_address_is_present(tmp_path):
    """The guard must not change the happy path: with the run-config collected, the SVI's
    connected prefix is still chosen as subnet_primary_route (over its /32 local route)."""
    c2f = _write_device(tmp_path, "core1")
    ifaces = build.build_interfaces("core1", "ios", c2f)
    assert ifaces["Vlan10"].svi_ip.startswith("10.0.10.2")
    assert ifaces["Vlan10"].subnet_primary_route == "10.0.10.0/24"


# ======================================================================================
# BH-2  build.py :: build_interfaces -- the global BPDU-Guard default was matched only in
#       its pre-IOS-XE-16 spelling, so 25 real devices / 647 access ports read 'not captured'.
# ======================================================================================
_BPDU_CLASSIC = "spanning-tree portfast bpduguard default"
_BPDU_EDGE = "spanning-tree portfast edge bpduguard default"     # IOS-XE 16.x+ (25 [HISTORY-REDACTED] devices)


def _runcfg_with(global_line):
    rc = fx.COLLECTIONS["core1"][1]["show running-config"]
    # insert the global stanza ahead of the first interface block, where IOS emits it
    head, sep, tail = rc.partition("\ninterface ")
    return head + "\n" + global_line + "\n" + sep + tail


@pytest.mark.parametrize("global_line", [_BPDU_CLASSIC, _BPDU_EDGE])
def test_global_bpduguard_default_is_seen_in_both_ios_spellings(tmp_path, global_line):
    """Both spellings are the SAME configured protection and must produce the same field.

    Reading only the classic form left stp_bpduguard EMPTY on every access port of an
    IOS-XE 16.x+ box -- which design_advisor (`_bpdu_seen`) and archreview L2-2
    (`bpduguard_state(...) is None`) both classify as NOT ASSESSED. Collected-and-protected
    therefore rendered as not-observed on 647 real access ports.
    """
    c2f = _write_device(tmp_path, "core1",
                        overrides={"show running-config": _runcfg_with(global_line)})
    ifaces = build.build_interfaces("core1", "ios", c2f)

    access = [p for p, d in ifaces.items()
              if (d.switchport_mode or "").lower() == "access"]
    assert access, "fixture has no access ports"
    assert all((ifaces[p].stp_bpduguard or "") == "Enable" for p in access), \
        f"global BPDU-Guard default not applied for {global_line!r}"


def test_no_global_bpduguard_line_still_leaves_the_field_unasserted(tmp_path):
    """Coverage honesty in the other direction: with no global default and no per-interface
    command, the field stays EMPTY (not-assessed) rather than being asserted either way."""
    c2f = _write_device(tmp_path, "core1")
    ifaces = build.build_interfaces("core1", "ios", c2f)
    access = [p for p, d in ifaces.items() if (d.switchport_mode or "").lower() == "access"]
    assert access
    assert all((ifaces[p].stp_bpduguard or "") == "" for p in access)


# ======================================================================================
# BH-3  build.py :: build_interfaces -- 'Neighbor Switch VTP Domain' was THIS switch's own
#       domain, so the mismatch check it exists for could only ever read 'they match'.
# ======================================================================================
_IDENT = {"hostname": "core1", "serial_number": "FOC0000LOCAL",
          "mgmt_ip": "10.0.99.1", "vtp_domain": "LOCAL-DOMAIN"}


def test_neighbor_vtp_domain_is_never_a_copy_of_the_local_switchs_domain(tmp_path):
    """367 of 367 rows on the [HISTORY-REDACTED] collection published a neighbour VTP domain identical to the
    local one -- because it WAS the local one, assigned from switch_identity. A reader
    comparing the two workbook columns to find a VTP-domain mismatch could only conclude
    'every trunk agrees'. Not observed must not render as agreement."""
    c2f = _write_device(tmp_path, "core1")
    ifaces = build.build_interfaces("core1", "ios", c2f, switch_identity=_IDENT)

    with_neigh = [p for p, d in ifaces.items() if (d.cdp_neighbor or "").strip()]
    assert with_neigh, "fixture has no CDP neighbours"
    assert any(ifaces[p].endpoint_type == "Switch" for p in with_neigh), \
        "fixture has no switch neighbour -- the old code path would not have fired"
    for p in with_neigh:
        assert (ifaces[p].neighbor_switch_vtp_domain or "") != _IDENT["vtp_domain"], \
            f"{p}: neighbour VTP domain is a copy of the local switch's"
    # the LOCAL column is still published (it is a real observation about this device)
    assert all(ifaces[p].current_switch_vtp_domain == _IDENT["vtp_domain"] for p in with_neigh)


def test_neighbor_switch_serial_is_never_the_neighbour_hostname(tmp_path):
    """581 rows on the [HISTORY-REDACTED] collection carried the neighbour's HOSTNAME in a column labelled
    'Neighbor Switch Serial' (no serial was ever parsed). The name is already published in
    its own `cdp_neighbor` column."""
    c2f = _write_device(tmp_path, "core1")
    ifaces = build.build_interfaces("core1", "ios", c2f, switch_identity=_IDENT)

    switches = [p for p, d in ifaces.items() if d.endpoint_type == "Switch"]
    assert switches, "fixture has no switch neighbour"
    for p in switches:
        nb = (ifaces[p].cdp_neighbor or "").strip()
        assert (ifaces[p].neighbor_switch_serial or "").strip() != nb or not nb, \
            f"{p}: 'neighbor_switch_serial' is the neighbour hostname {nb!r}"
        assert ifaces[p].cdp_neighbor, "the neighbour name must still be published"


# ======================================================================================
# BH-4  build.py :: _neighbor_is_infra -- an LLDP 'capabilities' field carrying the
#       ABSENT-TLV sentinel was read as a real advertisement, disabling the platform
#       fallback the function documents and dropping the neighbour from the
#       shadow-infrastructure candidate set.
# ======================================================================================
# Both strings are verbatim from migration_collection_20260613_063201: the sentinel appears on
# 395 of that fleet's 838 parsed LLDP neighbour records, and the description on its Catalyst
# 3850 neighbours. `_neighbor_is_infra` takes the record dict parse_neighbors_detail emits, so
# this exercises the classifier on the producer's own field shape.
_LLDP_SENTINEL = "not advertised"
_CATALYST_DESC = ("Cisco IOS Software [Gibraltar], Catalyst L3 Switch Software "
                  "(CAT3K_CAA-UNIVERSALK9-M), Version 16.12.11, RELEASE SOFTWARE (fc2)")


def test_absent_lldp_capability_tlv_falls_back_to_the_platform_family():
    """'not advertised' means the TLV was absent -- the documented platform fallback must run."""
    rec = {"device_id": "sw-unknown", "platform": _CATALYST_DESC,
           "capabilities": _LLDP_SENTINEL, "mgmt_ip": "", "local_intf": "Gi1/0/1",
           "remote_port": "Gi0/1", "proto": "lldp"}
    assert build._neighbor_is_infra(rec) is True


def test_absent_capability_tlv_on_a_non_infra_platform_stays_non_infra():
    """The fallback is still NARROW: an absent TLV does not promote a server to a switch."""
    rec = {"device_id": "esx-07", "platform": "HPE ProLiant DL380 Gen10",
           "capabilities": _LLDP_SENTINEL, "mgmt_ip": "", "local_intf": "Gi1/0/2",
           "remote_port": "vmnic0", "proto": "lldp"}
    assert build._neighbor_is_infra(rec) is False


def test_a_real_lldp_capability_advertisement_still_wins_over_the_platform():
    """Unchanged for every record that DOES advertise: an AP saying 'W' is not re-read as a
    switch via its platform string, and a bridge/router 'B, R' is still infra."""
    ap = {"device_id": "ap-12", "platform": "cisco AIR-CAP3702I-A-K9",
          "capabilities": "W", "local_intf": "Gi1/0/3", "proto": "lldp"}
    sw = {"device_id": "sw-9", "platform": "", "capabilities": "B, R",
          "local_intf": "Gi1/0/4", "proto": "lldp"}
    host = {"device_id": "pc-1", "platform": "", "capabilities": "S",
            "local_intf": "Gi1/0/5", "proto": "lldp"}
    assert build._neighbor_is_infra(ap) is False
    assert build._neighbor_is_infra(sw) is True
    assert build._neighbor_is_infra(host) is False


# ======================================================================================
# HH-1  html.py :: write_diff_workbook -- the Findings Delta 'Devices' column cut at 60
#       characters, mid-hostname, with no disclosure. 109 of 1805 real findings; 11 of 115
#       in the tracked sample fleet.
# ======================================================================================
_SAMPLE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "webapp", "sample_data", "sample_fleet.snapshot.json")


def _sample_snapshot():
    if not os.path.isfile(_SAMPLE):
        pytest.skip("sample fleet snapshot not available")
    with open(_SAMPLE, encoding="utf-8") as f:
        return json.load(f)


def test_findings_delta_devices_column_discloses_what_it_dropped(tmp_path):
    """Drive the real --compare writer over the tracked sample fleet: every punch-list
    finding is OPENED, so each one's device list is rendered. A list that does not fit must
    carry the house '(+N more)' marker and must never end mid-token."""
    new = _sample_snapshot()
    findings = [f for f in (new.get("punchlist") or []) if isinstance(f, dict)]
    over = [f for f in findings
            if len(", ".join(str(d) for d in (f.get("devices") or []))) > 60]
    assert over, "sample fleet has no over-long device list -- test would prove nothing"

    old = dict(new, punchlist=[])                      # -> every finding reads as OPENED
    out = os.path.join(str(tmp_path), "diff.xlsx")
    html.write_diff_workbook(old, new, out)

    from openpyxl import load_workbook
    ws = load_workbook(out)["Findings Delta"]
    by_title = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == "OPENED":
            by_title.setdefault(str(row[4]), []).append(str(row[3] or ""))

    checked = 0
    for f in over:
        cells = by_title.get(str(f.get("title", "")))
        if not cells:
            continue
        names = [str(d) for d in (f.get("devices") or [])]
        for cell in cells:
            assert "(+" in cell and "more)" in cell, \
                f"silent truncation, no disclosure: {cell!r}"
            shown = [t for t in cell.split(" (+")[0].split(", ") if t]
            assert all(t in names for t in shown), \
                f"a rendered device name is not a real device (cut mid-token): {cell!r}"
            hidden = int(cell.split("(+")[1].split(" ")[0])
            assert len(shown) + hidden == len(names), \
                f"disclosed count does not reconcile: {cell!r} vs {len(names)} devices"
            checked += 1
    assert checked, "no over-long finding was rendered -- test proved nothing"


def test_findings_delta_short_device_list_is_unchanged(tmp_path):
    """A list that fits is rendered exactly as before -- no marker, no reordering."""
    new = _sample_snapshot()
    old = dict(new, punchlist=[])
    out = os.path.join(str(tmp_path), "diff2.xlsx")
    html.write_diff_workbook(old, new, out)

    from openpyxl import load_workbook
    ws = load_workbook(out)["Findings Delta"]
    # a title can repeat across findings with DIFFERENT device sets, so collect every legal
    # rendering per title and require the cell to be one of them (verbatim, no marker).
    short: dict = {}
    for f in (new.get("punchlist") or []):
        if not isinstance(f, dict):
            continue
        joined = ", ".join(str(d) for d in (f.get("devices") or []))
        if joined and len(joined) <= 60:
            short.setdefault(str(f.get("title", "")), set()).add(joined)
    assert short, "sample fleet has no short device list"

    seen = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == "OPENED" and str(row[4]) in short:
            cell = str(row[3] or "")
            if "(+" in cell:                      # a same-titled LONG finding -> not this case
                continue
            assert cell in short[str(row[4])], f"short list altered: {cell!r}"
            seen += 1
    assert seen
