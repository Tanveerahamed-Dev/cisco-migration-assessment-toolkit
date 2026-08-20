"""J1 (MASTER_PLAN 3.5): the per-detector DESCRIPTOR schema. These tests pin DESCRIPTOR INTEGRITY --
they do NOT execute the detectors. The single highest-value property they enforce is the
"not-observed != healthy" schema property: every evidence-gated detector carries a non-empty
`abstains_when` (the coverage-honest guard), so a blind spot can never be declared as a schema-level
detector with no abstention state.

Grounding is asserted structurally: cited_fields must be plausible dotted snapshot keys, keys unique,
healthy_value present, threshold present-or-explicitly-null. The compute-fn shape + summary reconcile,
the workbook sheet appears with the exact headers, the sheet totals on bad input, and the real pipeline
publishes the `detector_schema` snapshot key (so the golden carries it).
"""
import json
import os
import re
import subprocess
import sys

import pytest
from openpyxl import Workbook, load_workbook

from cisco_toolkit.detector_schema import (
    SCHEMA,
    _DETECTOR_DESCRIPTORS,
    compute_detector_schema,
)
from cisco_toolkit.excel import DETECTOR_SCHEMA_SHEET_NAME, write_detector_schema_sheet

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

_DESCRIPTOR_FIELDS = {
    "key", "title", "family", "checks", "healthy_value", "threshold",
    "cited_fields", "abstains_when", "evidence_gated", "source_command",
}
_SHEET_HEADERS = ["Key", "Detector", "Family", "Checks", "Healthy Value", "Threshold",
                  "Cited Fields", "Abstains When", "Source Command"]

# A cited_field must look like a dotted snapshot key: a top-level section token followed by a dotted /
# indexed / templated path. The known top-level snapshot sections the descriptors read from.
_KNOWN_SECTIONS = {
    "physical_health", "l3_forwarding", "fhrp", "trunk_native", "stp_roots", "security",
    "config_hygiene", "lifecycle_risk", "syslog_intelligence", "qos_audit", "software_risk",
    "platform_health", "capacity", "operational_drift", "segmentation",
}


# --------------------------------------------------------------------------- registry integrity ---
def test_registry_has_enough_detectors_and_unique_keys():
    s = compute_detector_schema()
    dets = s["detectors"]
    assert len(dets) >= 20, f"expected ~20+ curated detectors, got {len(dets)}"
    keys = [d["key"] for d in dets]
    assert len(keys) == len(set(keys)), "detector keys must be unique"
    # keys are stable, lowercase-slug ids (no whitespace) so downstream references never drift
    for k in keys:
        assert isinstance(k, str) and k and re.fullmatch(r"[a-z0-9][a-z0-9-]*", k), f"bad key {k!r}"


def test_every_descriptor_has_the_full_field_set():
    for d in _DETECTOR_DESCRIPTORS:
        missing = _DESCRIPTOR_FIELDS - set(d)
        assert not missing, f"descriptor {d.get('key')!r} missing fields {missing}"


def test_evidence_gated_detectors_always_declare_abstains_when():
    """THE load-bearing property: a detector whose evidence can be absent (evidence_gated) MUST carry a
    non-empty abstains_when -- that is 'not-observed != healthy' encoded as a schema property, not a
    runtime convention. An empty guard on an evidence-gated detector is exactly the false-health hole."""
    for d in compute_detector_schema()["detectors"]:
        if d["evidence_gated"]:
            assert isinstance(d["abstains_when"], str) and d["abstains_when"].strip(), \
                f"evidence-gated detector {d['key']!r} has an EMPTY abstains_when (false-health hole)"


def test_healthy_value_present_and_threshold_present_or_explicitly_null():
    for d in compute_detector_schema()["detectors"]:
        assert isinstance(d["healthy_value"], str) and d["healthy_value"].strip(), \
            f"{d['key']!r} missing healthy_value"
        # threshold is present (a non-empty string) or EXPLICITLY null -- never absent, never empty-string
        assert d["threshold"] is None or (isinstance(d["threshold"], str) and d["threshold"].strip()), \
            f"{d['key']!r} threshold must be a non-empty string or explicit None, got {d['threshold']!r}"


def _rich_snapshot():
    """The richest committed snapshot for leaf resolution: the 23-device sample fleet (populated
    operational_drift/security/etc.), falling back to the golden. None if neither is present."""
    for p in ("webapp/sample_data/sample_fleet.snapshot.json", "tests/golden/snapshot.json"):
        path = os.path.join(ROOT, p)
        if os.path.exists(path):
            try:
                return json.load(open(path, encoding="utf-8"))
            except Exception:
                pass
    return None


# ---------------------------------------------------------- parser<->detector field-fidelity (P3-E1) ---
# A descriptor's cited_field names an evidence field by a small path grammar over the PUBLISHED snapshot.
# Binding it to a real leaf over the sample fleet makes the repo's own named recurring bug class --
# "parser<->detector format-fidelity drift" (docs/parser-format-fidelity.md), where a producing
# compute_*'s output field is renamed -- FAIL this contract instead of silently emitting a wrong/blank
# deliverable. The grammar the descriptors actually use:
#     section[].leaf            list-of-dicts          physical_health[].status
#     section.sub[].leaf        nested list-of-dicts   qos_audit.per_device[].n_trust_if
#     section.<key>.sub[].leaf  per-host / keyed map   security.<host>.findings[].id
#     section.<key>.<key>.leaf  doubly-keyed map       stp_roots.<host>.<vlan>.root_priority
#     section.sub.leaf          nested scalar          lifecycle_risk.summary.n_past_ldos
# '[]' iterates a list; '<name>' is a WILDCARD dict key (any host / vlan); a bare token is a literal key.
_SAMPLE_FLEET = os.path.join(ROOT, "webapp", "sample_data", "sample_fleet.snapshot.json")

# Cited fields that legitimately DO NOT resolve over the committed 23-device sample fleet because the
# fleet never collected/populated that section -- coverage-honest signal_absent, NOT a bug. PINNED so a
# newly-unresolved field (a real regression) OR a newly-populated one (progress) trips the test and gets
# reconciled consciously, never silently absorbed (doctrine 3: "not observed" != "healthy").
_SIGNAL_ABSENT_IN_FLEET = set()


def _tokenize_cited_field(field):
    """'a.b[].c' -> ['a', 'b', '[]', 'c']; a '<host>' segment is kept verbatim as a wildcard token."""
    toks = []
    for part in field.split("."):
        if part.endswith("[]"):
            toks.append(part[:-2])
            toks.append("[]")
        else:
            toks.append(part)
    return toks


def _walk_snapshot(node, tokens):
    """Resolve a token path into a loaded snapshot node. Returns exactly one of:
       'resolved' -- the leaf was reached on some concrete record;
       'missing'  -- a POPULATED container has no such key/leaf: the drift bug -> the test FAILS;
       'absent'   -- a container along the path is empty/absent: signal_absent, undecidable here."""
    if not tokens:
        return "resolved"
    tok, rest = tokens[0], tokens[1:]
    if tok == "[]":                                   # iterate a list-of-records
        if not isinstance(node, list) or not node:
            return "absent"
        states = [_walk_snapshot(x, rest) for x in node]
        if "resolved" in states:
            return "resolved"
        return "absent" if all(s == "absent" for s in states) else "missing"
    if tok.startswith("<") and tok.endswith(">"):     # wildcard dict key (any host / vlan)
        if not isinstance(node, dict) or not node:
            return "absent"
        states = [_walk_snapshot(v, rest) for v in node.values()]
        if "resolved" in states:
            return "resolved"
        return "absent" if all(s == "absent" for s in states) else "missing"
    if isinstance(node, dict):                        # literal key
        if tok in node:
            return _walk_snapshot(node[tok], rest)
        return "missing" if node else "absent"        # a populated dict without the key IS drift
    return "absent"


def _resolve_cited_field(rich, field):
    """Status of `field` (see _walk_snapshot) over the loaded snapshot `rich`. An entirely-absent
    top-level section is 'absent' (signal_absent), never 'missing'."""
    toks = _tokenize_cited_field(field)
    head, rest = toks[0], toks[1:]
    if not isinstance(rich, dict) or head not in rich:
        return "absent"
    return _walk_snapshot(rich[head], rest)


def test_cited_fields_name_a_field_not_a_bare_section_and_resolve():
    """Every cited_field must name the EVIDENCE FIELD inside a real section — 'section[].field',
    'section.<host>.field', 'section.sub[].field' — never a bare section token. A bare section
    ('trunk_native') tells a client "look at this whole block" without saying which field backs the
    finding, and is how the l2-native-vlan-1 descriptor once cited a DIFFERENT detector's output
    (adversarial-review finding, 2026-07-05). And where the section is POPULATED in the rich sample
    fleet, the cited leaf must actually exist on some record — across ALL the descriptor path grammars
    (P3-E1), not just the flat 'section[].leaf' shape — so a renamed producing-compute_* output field
    (parser<->detector drift) fails HERE rather than shipping a wrong/blank deliverable. A section the
    fleet never populated is signal_absent (status 'absent'), checked in the pinned test below."""
    rich = _rich_snapshot()
    for d in compute_detector_schema()["detectors"]:
        cf = d["cited_fields"]
        assert isinstance(cf, list) and cf, f"{d['key']!r} has empty cited_fields"
        for f in cf:
            assert isinstance(f, str) and f, f"{d['key']!r} has a blank cited_field"
            head = re.split(r"[.\[]", f, 1)[0]
            assert head in _KNOWN_SECTIONS, \
                f"{d['key']!r} cites {f!r} whose root {head!r} is not a known snapshot section"
            # the citation must reference a FIELD, not the bare section (the trunk_native bug class)
            assert f != head, \
                f"{d['key']!r} cites the BARE section {f!r} — name the evidence field (e.g. {head}[].field)"
            # leaf resolution across every descriptor grammar: a POPULATED section that lacks the cited
            # leaf is field drift (a renamed compute_* output) -> fail; an empty/absent section is
            # signal_absent, not a failure (pinned + reconciled in the test below).
            if rich is not None:
                status = _resolve_cited_field(rich, f)
                assert status != "missing", (
                    f"{d['key']!r} cites {f!r} but a POPULATED {head!r} has no such leaf — "
                    f"parser<->detector field drift (a producing compute_* output field was renamed)")


def test_cited_field_resolution_over_sample_fleet_and_signal_absent_is_pinned():
    """The parser<->detector field contract, made coverage-honest over the 23-device sample fleet: every
    cited_field either RESOLVES to a real leaf produced by the pipeline, or is an EXPLICITLY-PINNED
    signal_absent (a section the fleet never collected). Two ways this fails, both the point of P3-E1:
      * a POPULATED section missing the cited leaf -> a renamed compute_* output field (drift);
      * the signal_absent set drifting -> a field silently stopped/started resolving, unreconciled.
    (The sibling test above runs over the golden-fallback too; this one pins against the fleet.)"""
    if not os.path.exists(_SAMPLE_FLEET):
        pytest.skip("sample fleet snapshot not present — field-resolution pin needs the rich fleet")
    rich = json.load(open(_SAMPLE_FLEET, encoding="utf-8"))
    observed_absent = set()
    for d in compute_detector_schema()["detectors"]:
        for f in d["cited_fields"]:
            status = _resolve_cited_field(rich, f)
            assert status != "missing", (
                f"{d['key']!r} cites {f!r} but a populated section has no such leaf — "
                f"parser<->detector field drift")
            if status == "absent":
                observed_absent.add((d["key"], f))
    assert observed_absent == _SIGNAL_ABSENT_IN_FLEET, (
        "signal_absent cited-field set drifted — reconcile _SIGNAL_ABSENT_IN_FLEET (do not paper over):\n"
        f"  newly UNRESOLVED (possible regression): {sorted(observed_absent - _SIGNAL_ABSENT_IN_FLEET)}\n"
        f"  newly RESOLVED (progress; drop from pin): {sorted(_SIGNAL_ABSENT_IN_FLEET - observed_absent)}")


def test_family_and_source_command_are_read_only_show_queries():
    for d in compute_detector_schema()["detectors"]:
        assert isinstance(d["family"], str) and d["family"].strip(), f"{d['key']!r} missing family"
        sc = d["source_command"]
        # every backing command is a read-only 'show' query -- the descriptor never names a write/config verb
        assert isinstance(sc, str) and sc.startswith("show "), \
            f"{d['key']!r} source_command must be a read-only show query, got {sc!r}"


def test_source_commands_reuse_the_shared_punchlist_map_where_a_family_maps():
    """Grounding: where the engine's punch-list already owns the backing command for a category, the
    descriptor must reuse THAT command (one owner, no divergent copy)."""
    from cisco_toolkit.analyze import _PUNCH_SOURCE_COMMAND as P
    by_key = {d["key"]: d for d in compute_detector_schema()["detectors"]}
    # L1 family maps to the punch-list's L1 command; Security -> its running-config source; STP -> its command.
    assert by_key["l1-err-disabled"]["source_command"] == P["L1"]
    assert by_key["sec-no-aaa"]["source_command"] == P["Security"]
    assert by_key["stp-accidental-root"]["source_command"] == P["STP"]


def test_lifecycle_detectors_abstain_when_matched_authority_is_incomplete():
    """Unknown includes exact matches whose retained source/date evidence cannot support a date band."""
    by_key = {d["key"]: d for d in compute_detector_schema()["detectors"]}
    for key in ("lifecycle-past-ldos", "lifecycle-near-ldos", "lifecycle-past-eos"):
        detector = by_key[key]
        assert "source/date authority is incomplete" in detector["abstains_when"]
        assert "band == Unknown" in detector["abstains_when"]
        assert "not proof of contract entitlement" in detector["checks"]

    assert "asof > LDoS" in by_key["lifecycle-past-ldos"]["threshold"]
    assert "LDoS date itself is not past" in by_key["lifecycle-past-ldos"]["threshold"]
    assert "asof > EoS" in by_key["lifecycle-past-eos"]["threshold"]
    assert "more than one year away" in by_key["lifecycle-past-eos"]["threshold"]
    assert "0 <= (LDoS - asof) / 365.25 <= 1.0" in by_key["lifecycle-near-ldos"]["threshold"]
    assert "lifecycle_risk.summary.n_near" in by_key["lifecycle-near-ldos"]["cited_fields"]


# ---------------------------------------------------------------- compute-fn shape + reconcile ---
def test_compute_shape_and_summary_reconciles():
    s = compute_detector_schema()
    assert set(s) == {"schema", "detectors", "summary"}
    assert s["schema"] == SCHEMA == "detector_schema/1"
    summ = s["summary"]
    dets = s["detectors"]
    assert summ["n_detectors"] == len(dets)
    assert summ["n_with_threshold"] == sum(1 for d in dets if d["threshold"] is not None)
    assert summ["n_evidence_gated"] == sum(1 for d in dets if d["evidence_gated"])
    # the per-family histogram sums back to the detector count (no detector dropped/double-counted)
    assert sum(summ["families"].values()) == len(dets)
    assert set(summ["families"]) == {d["family"] for d in dets}
    assert "not-observed" in summ["note"] or "abstain" in summ["note"].lower()


def test_compute_is_pure_and_deterministic():
    a = compute_detector_schema()
    b = compute_detector_schema()
    assert a == b, "compute_detector_schema must be deterministic"
    # mutating the returned structure must not corrupt the next call (lists are copied out)
    a["detectors"].append({"key": "junk"})
    a["detectors"][0]["cited_fields"].append("junk.field")
    assert len(compute_detector_schema()["detectors"]) == len(b["detectors"])
    assert "junk.field" not in compute_detector_schema()["detectors"][0]["cited_fields"]


# --------------------------------------------------------------------------- workbook sheet ---
def test_sheet_appears_with_exact_headers():
    wb = Workbook()
    write_detector_schema_sheet(wb, compute_detector_schema())
    assert DETECTOR_SCHEMA_SHEET_NAME in wb.sheetnames
    ws = wb[DETECTOR_SCHEMA_SHEET_NAME]
    # row 1 is the coverage-honest note (describes, does not re-run); the column headers are row 2
    header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert header == _SHEET_HEADERS, f"header row mismatch: {header}"
    assert ws.max_row == 2 + len(compute_detector_schema()["detectors"]), "one data row per detector"
    # the numeric-threshold column renders an explicit em-dash (never a blank cell) for a null threshold
    for row in ws.iter_rows(min_row=3, values_only=True):
        assert row[5] not in (None, ""), "Threshold cell must never be blank (null renders as an em-dash)"


def test_sheet_totals_on_bad_input():
    """A total, fail-soft writer: bad/empty input yields the sheet with headers and no data rows,
    never an exception (matches the repo's other sheet writers)."""
    for bad in ({}, None, {"detectors": None}, {"detectors": [1, 2, "x"]}, "nonsense"):
        wb = Workbook()
        write_detector_schema_sheet(wb, bad)
        ws = wb[DETECTOR_SCHEMA_SHEET_NAME]
        header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        assert header == _SHEET_HEADERS, f"bad input {bad!r} broke the header row"
        assert ws.max_row == 2, f"bad input {bad!r} should yield zero data rows, got {ws.max_row - 2}"


# --------------------------------------------------------------------------- golden carries the key ---
def _make_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Interface Data"
    ws.append(["Hostname", "Port", "Status"])
    wb.save(path)


def test_pipeline_publishes_detector_schema_key_and_sheet(tmp_path):
    """End-to-end: the real offline pipeline publishes snap['detector_schema'] AND the 'Detector Schema'
    workbook sheet, so the frozen golden carries the additive key/sheet. Uses the same synthetic fixture
    and CLI shape as tests/test_pipeline_golden.py."""
    import synthetic_fixtures as fx

    collection = fx.write_collection(str(tmp_path / "collection"))
    devices = tmp_path / "devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "out.xlsx"

    cmd = [sys.executable, os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py"),
           "--no-collect", "--collection-dir", collection,
           "--devices-file", str(devices), "--template", str(template),
           "--output", str(out_xlsx), "--no-html", "--workers", "1"]
    proc = subprocess.run(cmd, cwd=str(tmp_path), capture_output=True, text=True, timeout=300)
    assert proc.returncode == 0, f"pipeline failed:\n{proc.stdout}\n{proc.stderr}"

    snap_path = os.path.splitext(str(out_xlsx))[0] + ".snapshot.json"
    snap = json.loads(open(snap_path, encoding="utf-8").read())
    ds = snap.get("detector_schema")
    assert isinstance(ds, dict) and ds.get("schema") == SCHEMA, "snapshot missing detector_schema block"
    # the PUBLISHED block equals a fresh compute (no surface recomputes / drifts the registry)
    assert ds == compute_detector_schema(), "published detector_schema must equal the canonical recompute"

    wb = load_workbook(str(out_xlsx), read_only=True)
    try:
        assert DETECTOR_SCHEMA_SHEET_NAME in wb.sheetnames, "Detector Schema sheet missing from the workbook"
    finally:
        wb.close()
