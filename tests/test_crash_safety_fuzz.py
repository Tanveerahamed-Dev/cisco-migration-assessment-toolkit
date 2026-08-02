"""[crash-safety fuzz, 2026-07-28] A malformed snapshot must DEGRADE, never raise.

Found by a systematic recursive poison sweep: every shape-path of the REAL golden snapshot
(`tests/golden/snapshot.json`) replaced by each of `inf` / `nan` / `None` / `0` / `""` / `"x"` /
`[]` / `{}` / `[1]` / `{"k":1}` / `True` / a 40k string / a huge int literal / a MIXED-type run,
driven through every offline snapshot consumer. Every poison used here is JSON-LEGAL: `json.loads`
accepts the bare `Infinity` / `-Infinity` / `NaN` tokens and integer literals of unbounded
precision, and a hand-trimmed, foreign-tool or older-schema snapshot puts a str/int/list/None
wherever a dict or list was expected.

Why a crash here is not merely a 500: the snapshot is STORED (webapp upload, `--no-collect`
re-analysis file), so every later read of it crashes IDENTICALLY -- a stored denial of service on
the one door Atlas serves in the field, and an aborted deliverable in the CLI.

The findings are grouped into FOUR classes and fixed at the shared choke point, not per traceback:

  A `or {}` / `or []` where a TRUTHY non-dict/non-list survives -> the next `.get` / `len` /
    iteration raises. Fixed with the house `_as_dict` / `_as_list` / `_dict_rows` guards.
  B numeric coercion that misses `OverflowError` (or does no coercion at all) -> the bare JSON
    `Infinity` and the unbounded-precision int literal. Fixed via `textutils._as_num` /
    `html._renderable_num` / integer arithmetic.
  C a leaf used as a dict KEY, an `in <dict>` operand, or a `sorted()` element:
    C1 an UNHASHABLE dict/list leaf -> `TypeError: unhashable type` (fixed with `_hkey`);
    C2 a MIXED-type run -> `TypeError: '<' not supported between 'str' and 'int'` (fixed with
       `_skey`). C2 needs no poison at all -- two ordinary JSON-legal rows are enough.
  D a non-finite / out-of-range NUMBER reaching a rendered cell. `10**400` raises OverflowError
    inside `wb.save()`; `inf`/`nan` are WORSE -- openpyxl saves them and Excel reads the cell back
    EMPTY, a silent false-health. Both fixed at the ONE cell/paragraph sanitizer,
    `textutils.xml_safe`, which `excel._xls_sanitize` and every docx generator already delegate to.

Every case below RAISED before its fix; reverting the named guard re-raises it.
"""
import json
import math
import os

import pytest

INF = float("inf")
NAN = float("nan")
BIG = 10 ** 400            # json.loads accepts an integer literal of unbounded precision
GOLDEN = os.path.join(os.path.dirname(__file__), "golden", "snapshot.json")


def _golden():
    with open(GOLDEN, encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------------------------
# CLASS A -- `or {}` / `or []` keeps a TRUTHY non-dict/non-list
# ---------------------------------------------------------------------------------------------
def _precert_checks():
    from cisco_toolkit import precert
    return precert.compute_readiness_freeze(
        {"migration_readiness": [{"group": "G1", "readiness": "READY", "checks": INF}]})


def _precert_cc_devices():
    from cisco_toolkit import precert
    return precert.compute_readiness_freeze(
        {"migration_readiness": [{"group": "G1"}], "collection_completeness": {"devices": True}})


def _precert_segmentation():
    """Reached from html.write_diff_workbook -- a WRITER, so this aborted the --compare deliverable."""
    from cisco_toolkit import precert
    return precert.compute_precert({"segmentation": {"domains": INF, "vrfs": True}}, {"segmentation": {}})


def _aclcheck_acls_section():
    from cisco_toolkit import aclcheck
    return aclcheck.compute_filter_line_reachability({"acls": "shown", "object_groups": 7})


def _aclcheck_per_host():
    from cisco_toolkit import aclcheck
    return aclcheck.compute_filter_line_reachability({"acls": {"h": "no acls"}, "object_groups": {"h": "x"}})


def _aclcheck_rules_nonlist():
    from cisco_toolkit import aclcheck
    return aclcheck.analyze_acl(3, {}, "h")


def _aclcheck_nondict_rule():
    """A non-dict ELEMENT in the rules list must model as an UNEVALUABLE line, never a narrow box."""
    from cisco_toolkit import aclcheck
    return aclcheck.analyze_acl(["permit ip any any", None, 7], {}, "h")


def test_aclcheck_unreadable_rule_abstains_it_is_never_proved_dead():
    """The SAFETY half of the _rule_box guard, separate from not-crashing.

    A rule row this module cannot read must degrade to the FULL, UNEVALUABLE box -> INDETERMINATE.
    Degrading it to a NARROW box instead would let the algebra 'prove' a later line dead against a
    rule it never actually parsed -- a false BLOCKING_LINES verdict, which is far worse than an
    abstention (the operator would delete a live ACE)."""
    from cisco_toolkit import aclcheck
    full = {"ip": "0.0.0.0", "wild": "255.255.255.255"}
    real = {"action": "permit", "proto": "ip", "src": full, "dst": full}
    out = aclcheck.analyze_acl([real, "permit ip any any", None, 7], {}, "h")
    unreadable = [f for f in out if f["line_index"] > 0]
    assert len(unreadable) == 3
    assert {f["reason"] for f in unreadable} == {"INDETERMINATE"}


def _aclcheck_group_body():
    from cisco_toolkit import aclcheck
    rule = {"action": "permit", "proto": "ip", "src": {"group": "G"},
            "dst": {"ip": "0.0.0.0", "wild": "255.255.255.255"}}
    return aclcheck.compute_filter_line_reachability(
        {"acls": {"h": {"A": [rule]}}, "object_groups": {"h": {"G": "not a dict"}}})


def _aclcheck_group_members():
    from cisco_toolkit import aclcheck
    rule = {"action": "permit", "proto": "ip", "src": {"group": "G"},
            "dst": {"ip": "0.0.0.0", "wild": "255.255.255.255"}}
    return aclcheck.compute_filter_line_reachability(
        {"acls": {"h": {"A": [rule]}}, "object_groups": {"h": {"G": {"members": INF}}}})


def _nrfu_move_groups():
    from cisco_toolkit import nrfu_export
    return nrfu_export.compute_nrfu_commands({"devices": {"h": {}}, "move_groups": [{"switches": INF}]})


def _nrfu_routing_neighbors():
    from cisco_toolkit import nrfu_export
    return nrfu_export.compute_nrfu_commands({"devices": {"h": {}}, "routing_neighbors": {"h": {"ospf": True}}})


def _nrfu_app_domains():
    from cisco_toolkit import nrfu_export
    return nrfu_export.compute_nrfu_commands(
        {"devices": {"h": {}}, "application_intelligence": {"domains": [{"switches": INF}]}})


def _html_macset():
    """`s or ""` kept a truthy non-str; re.split then raised, aborting the whole --compare workbook."""
    from cisco_toolkit import html
    return html._macset(INF), html._macset(7), html._macset({"k": 1}), html._macset(None)


CLASS_A = [
    ("precert.compute_readiness_freeze/checks", _precert_checks),
    ("precert.compute_readiness_freeze/collection_completeness.devices", _precert_cc_devices),
    ("precert._segmentation_invariants/domains+vrfs [diff WRITER]", _precert_segmentation),
    ("aclcheck/acls+object_groups section", _aclcheck_acls_section),
    ("aclcheck/per-host acls+object_groups", _aclcheck_per_host),
    ("aclcheck.analyze_acl/rules", _aclcheck_rules_nonlist),
    ("aclcheck._rule_box/non-dict rule element", _aclcheck_nondict_rule),
    ("aclcheck._group_prefixes/group body", _aclcheck_group_body),
    ("aclcheck._group_prefixes/members", _aclcheck_group_members),
    ("nrfu_export/move_groups.switches", _nrfu_move_groups),
    ("nrfu_export/routing_neighbors.<proto>", _nrfu_routing_neighbors),
    ("nrfu_export/application_intelligence.domains", _nrfu_app_domains),
    ("html._macset/non-str mac field [diff WRITER]", _html_macset),
]


@pytest.mark.parametrize("name,fn", CLASS_A, ids=[c[0] for c in CLASS_A])
def test_class_a_truthy_non_container_degrades(name, fn):
    """A TRUTHY non-dict / non-list where a container is expected must read as ABSENT, not raise.

    `x or {}` / `x or []` guards None and empty but keeps `5`, `True`, `"str"` and the `float('inf')`
    json.loads makes of a bare JSON `Infinity` -- and the next `.get()` / `.items()` / `for ... in`
    raises AttributeError / TypeError. Reverting any of these to `or {}` / `or []` re-raises."""
    assert fn() is not None


# ---------------------------------------------------------------------------------------------
# CLASS B -- numeric coercion missing OverflowError, or absent entirely
# ---------------------------------------------------------------------------------------------
@pytest.mark.parametrize("bad", [INF, -INF, NAN, BIG, -BIG, "x", "1e400", {"k": 1}, [1]],
                         ids=["inf", "-inf", "nan", "bigint", "-bigint", "str", "1e400", "dict", "list"])
def test_class_b_precert_readiness_counts_coerce_fail_soft(bad):
    """`int(g.get("n_fail") or 0)` had NO guard at all: OverflowError on the bare JSON Infinity,
    ValueError on NaN / a non-numeric string, TypeError on a dict/list. It runs inside
    write_diff_workbook, so one bad leaf aborted the whole --compare deliverable. Now routed through
    the shared fail-soft `textutils._as_num`, which lists (TypeError, ValueError, OverflowError)."""
    from cisco_toolkit import precert
    out = precert.compute_readiness_freeze(
        {"migration_readiness": [{"group": "G1", "readiness": "READY", "n_fail": bad, "n_warn": bad}],
         "collection_completeness": {"summary": {"not_collected": bad}}})
    g = out["groups"][0]
    assert isinstance(g["n_fail"], int) and isinstance(g["n_warn"], int)


@pytest.mark.parametrize("bad", [INF, -INF, NAN, BIG], ids=["inf", "-inf", "nan", "bigint"])
def test_class_b_trend_point_rejects_unusable_scores(bad):
    """`isinstance(score, (int, float))` admitted BOTH numbers arithmetic cannot survive: a huge int
    literal (`sum(scores)/len(scores)` -> "integer division result too large for a float") and
    inf/nan (which would render an average of inf into the campaign deck). `_renderable_num` is the
    guard; reverting it to the bare isinstance re-raises / re-renders a non-finite average."""
    from cisco_toolkit import html
    pt = html._trend_point({"health_scores": [{"switch": "a", "score": bad, "band": "Good"}]})
    avg = pt["avg_health"]
    assert avg == "" or (isinstance(avg, (int, float)) and math.isfinite(avg))


@pytest.mark.parametrize("bad", [BIG, -BIG, INF, -INF, NAN], ids=["bigint", "-bigint", "inf", "-inf", "nan"])
def test_class_b_the_isfinite_guard_itself_must_not_overflow(bad):
    """The sharpest instance of class B: the guard written to reject the JSON `Infinity` CRASHED on
    the other value json.loads accepts.

    `isinstance(v, (int, float)) and math.isfinite(v)` admits an unbounded-precision int, and
    `math.isfinite(10**400)` raises `OverflowError: int too large to convert to float` BEFORE it can
    return False. `ssot.reconcile` runs inside `docmeta.add_excellence_front`, so this aborted EVERY
    document in the docx family (and `causal` feeds the explorer's decision chips) over one leaf.
    `textutils.is_finite_num` is the one owner; it tests the int bound by COMPARISON, never by
    converting."""
    import math as _m

    from cisco_toolkit import causal, ssot
    from cisco_toolkit.textutils import is_finite_num
    if isinstance(bad, int):                       # the premise: the old idiom's own guard raises
        with pytest.raises(OverflowError):
            _m.isfinite(bad)
    assert is_finite_num(bad) is False
    snap = {"devices": {"h": {}},
            "executive_brief": {"scale": {"n_devices": 1}, "posture": {"avg_health": 50}},
            "health_scores": [{"switch": "h", "band": "Good", "score": bad}]}
    assert isinstance(ssot.reconcile(snap), list)
    assert isinstance(ssot.summary(snap), dict)
    # decision shape taken from the REAL producer (design_advisor.compute_design_blueprint over the
    # golden snapshot): status='recommended' is what makes compute_causal_flows read evidence.count.
    snap2 = dict(snap)
    snap2["design_blueprint"] = {"decisions": [
        {"id": "d", "title": "t", "domain": "l2", "priority": "High", "status": "recommended",
         "confidence": "high", "driver": "x", "recommended_action": "y",
         "evidence": {"summary": "s", "count": bad, "devices": ["h"], "fields": ["health_scores"]}}]}
    flows = causal.compute_causal_flows(snap2)
    assert isinstance(flows, dict) and isinstance(flows.get("flows"), list)


@pytest.mark.parametrize("bad", [INF, -INF, NAN, BIG], ids=["inf", "-inf", "nan", "bigint"])
def test_class_b_calibration_report_survives_an_unusable_score(bad):
    """compute_calibration_report ran statistics over `isinstance(score, (int, float))` rows -- so a
    huge int OverflowError'd and inf/nan reached statistics.mean / the int() banding. One malformed
    row aborted the calibration section for the whole fleet."""
    from cisco_toolkit import analyze
    out = analyze.compute_calibration_report(
        [{"switch": "a", "score": bad, "band": "Good"}, {"switch": "b", "score": 50, "band": "Fair"}])
    assert isinstance(out, dict) and out.get("n") == 1        # the poisoned row is EXCLUDED, not fatal


def test_class_a_search_filters_is_guarded_like_analyze_acl():
    """search_filters is the SECOND public entry into the same rule algebra and read the same
    untrusted sections with none of analyze_acl's container/element guards -- the 'enumerate every
    exit' half of the class: guarding one entry point leaves the identical crash reachable via the
    other."""
    from cisco_toolkit import aclcheck
    full = {"ip": "0.0.0.0", "wild": "255.255.255.255"}
    h = {"src": "10.0.0.1", "dst": "10.0.0.2", "proto": "tcp", "dport": 80}
    assert isinstance(aclcheck.search_filters(3, h), dict)                       # truthy non-list rules
    assert isinstance(aclcheck.search_filters(                                   # truthy non-dict ogs
        [{"action": "permit", "proto": "ip", "src": full, "dst": full}], h, object_groups=7), dict)
    assert isinstance(aclcheck.search_filters(["permit ip any any", None, 7], h), dict)  # non-dict rows


def test_is_finite_num_keeps_every_real_number():
    """Inert on real data: every in-range number still passes, so no mean, tally or cell changes."""
    from cisco_toolkit.textutils import is_finite_num
    assert all(is_finite_num(v) for v in (0, 1, -1, 42, 3.14, -2.5, 1e308, 2 ** 62, -(2 ** 62)))
    assert not any(is_finite_num(v) for v in (True, False, "5", None, {}, []))


def test_class_b_design_blueprint_firewall_limit_bigint():
    """`_peak >= 0.9 * _lim` overflowed on a huge-int limit (`0.9 * 10**400`), aborting the whole
    design blueprint and 500ing /design + /architecture_coverage. Now an INTEGER comparison."""
    from cisco_toolkit import design_advisor
    bp = design_advisor.compute_design_blueprint(
        {"devices": {"c": {}},
         "firewall": {"c": {"resource_usage": [{"resource": "conns", "limit": BIG, "peak": BIG, "current": 1}]}}}, {})
    assert isinstance(bp, dict)


def test_class_b_port_operand_must_be_a_real_int_port():
    """Every _port_intervals branch feeds val/val2 into integer arithmetic and then into `lo <= hi`.
    A str/list/float operand (foreign-tool export, older schema) raised TypeError deep in _iv_norm /
    _iv_inter, aborting the shadow proof for the WHOLE fleet over one malformed ACE. It must abstain
    (INDETERMINATE) instead -- an unreadable line must never be used to 'prove' a later line dead."""
    from cisco_toolkit import aclcheck
    full = {"ip": "0.0.0.0", "wild": "255.255.255.255"}
    for bad in ("", "80", ["x"], {"k": 1}, 80.5, True, -1, 70000):
        rules = [{"action": "permit", "proto": "tcp", "src": full, "dst": full,
                  "dport": {"op": "eq", "val": bad}},
                 {"action": "permit", "proto": "tcp", "src": full, "dst": full,
                  "dport": {"op": "range", "val": bad, "val2": bad}}]
        out = aclcheck.analyze_acl(rules, {}, "h")
        assert all(f["reason"] != "BLOCKING_LINES" for f in out), bad
    # a REAL int port still models exactly as before (the producer emits int|None -- parse._acl_portnum)
    real = aclcheck._port_intervals({"op": "eq", "val": 443})
    assert real == [(443, 443)]


# ---------------------------------------------------------------------------------------------
# CLASS C1 -- an UNHASHABLE leaf as a dict key / `in <dict>` operand
# ---------------------------------------------------------------------------------------------
@pytest.mark.parametrize("bad", [{"k": 1}, [1], {}, []], ids=["dict", "list", "emptydict", "emptylist"])
def test_class_c1_html_delta_and_trend_tolerate_unhashable_labels(bad):
    """`{r.get("switch"): r}` and `r.get("readiness") in readiness` HASH the leaf, so a dict/list
    there raised `TypeError: unhashable type` -- aborting write_diff_workbook AND
    write_campaign_workbook and 500ing the webapp diff/trend routes on every later read of a stored
    snapshot. `_hkey` keeps it a DISTINCT key rather than dropping the device."""
    from cisco_toolkit import html
    d = html.compute_snapshot_delta({"health_scores": [{"switch": bad, "band": "Good"}]},
                                    {"health_scores": [{"switch": bad, "band": "Poor"}]})
    assert isinstance(d, dict) and "verdict" in d
    pt = html._trend_point({"migration_readiness": [{"group": "G", "readiness": bad}]})
    assert isinstance(pt, dict)


@pytest.mark.parametrize("bad", [{"k": 1}, [1], {}, []], ids=["dict", "list", "emptydict", "emptylist"])
def test_class_c1_coverage_matrix_tolerates_unhashable_host_and_parser(bad):
    """`ci_by_host.setdefault(f["host"], …)` and `ev.get("parser") not in <frozenset>` both hash the
    leaf. A non-str host can never match an inventory device, and a non-str parser can never be in
    the may-be-empty exemption set -> treated as SUSPECT (the coverage-honest direction)."""
    from cisco_toolkit import coverage_matrix
    cm = coverage_matrix.compute_coverage_matrix(
        {"devices": {"h": {}},
         "capture_integrity": {"findings": [{"host": bad, "command": "show version"}]},
         "parse_yield": {"events": [{"device": "h", "parser": bad, "error": ""}]}})
    assert isinstance(cm, dict) and isinstance(cm.get("rows"), list)


# ---------------------------------------------------------------------------------------------
# CLASS C2 -- a MIXED-type run reaching sorted() (needs no poison: two JSON-legal rows suffice)
# ---------------------------------------------------------------------------------------------
def test_class_c2_mixed_type_labels_do_not_break_sorted():
    """`"switch": 10` on one row and `"switch": "core1"` on the next is ordinary, JSON-legal data,
    and `sorted()` over the union raises `'<' not supported between 'str' and 'int'`. `_skey` gives
    a TOTAL order; reverting it re-raises."""
    from cisco_toolkit import design_advisor, html
    o = {"health_scores": [{"switch": 10, "band": "Good"}, {"switch": "core1", "band": "Good"}]}
    n = {"health_scores": [{"switch": 10, "band": "Poor"}, {"switch": "core1", "band": "Poor"}]}
    assert isinstance(html.compute_snapshot_delta(o, n), dict)
    bp = design_advisor.compute_design_blueprint(
        {"devices": {"c": {}}, "move_groups": [{"switches": [1, "core1", 2.5]}],
         "aci": {"a": {"vrfs": [{"tenant": 1, "name": "v", "dn": "d"},
                                {"tenant": "T", "name": 2, "dn": "d2"}]}}}, {})
    assert isinstance(bp, dict)


def test_class_c2_skey_preserves_the_order_of_well_formed_data():
    """The guard must be INERT on real snapshots: an all-string set sorts byte-identically, so no
    deliverable's row order changes. (A guard that silently reordered valid output would be a
    regression dressed as a fix.)"""
    from cisco_toolkit import design_advisor, html
    names = ["core2", "access1", "core1", "dist-9", "dist-10"]
    for skey in (html._skey, design_advisor._skey):
        assert sorted(names, key=skey) == sorted(names)


# ---------------------------------------------------------------------------------------------
# CLASS D -- a non-finite / out-of-range number reaching a rendered cell
# ---------------------------------------------------------------------------------------------
@pytest.mark.parametrize("bad", [INF, -INF, NAN, BIG, -BIG],
                         ids=["inf", "-inf", "nan", "bigint", "-bigint"])
def test_class_d_xml_safe_bounds_unrenderable_numbers(bad):
    """The ONE cell/paragraph sanitizer (excel._xls_sanitize and every docx generator delegate here).

    Two DIFFERENT failures, one guard: an int beyond float64 raises `OverflowError: int too large to
    convert to float` inside `wb.save()` -- after every sheet is built, aborting the ONE deliverable
    produced unconditionally (there is no --no-excel); and inf/nan are WORSE than a crash -- openpyxl
    saves them and Excel reads the cell back EMPTY, so the number silently disappears from the
    delivered workbook. Both must become a VISIBLE placeholder."""
    from cisco_toolkit.textutils import xml_safe
    out = xml_safe(bad)
    assert isinstance(out, str) and out.startswith("[unrenderable")


@pytest.mark.parametrize("good", [0, 1, -1, 42, 3.14, -2.5, True, False, None, 1e308, 2 ** 62],
                         ids=["0", "1", "-1", "42", "3.14", "-2.5", "True", "False", "None", "1e308", "2**62"])
def test_class_d_xml_safe_passes_every_real_value_through_unchanged(good):
    """Coverage-honesty guard on the guard: every in-range number and both bools must render
    NATIVELY, byte-identical to before. A sanitizer that stringified real numbers would silently
    change every numeric cell in the workbook from a number to text."""
    from cisco_toolkit.textutils import xml_safe
    out = xml_safe(good)
    assert out is good or out == good
    assert not isinstance(out, str)


def test_class_d_workbook_actually_saves_with_a_poisoned_number(tmp_path):
    """End-to-end proof, not just the helper: openpyxl must accept every sanitized cell. Without the
    xml_safe numeric branch this raises OverflowError inside wb.save()."""
    import openpyxl

    from cisco_toolkit.excel import _xls_cell_value
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate([INF, -INF, NAN, BIG, -BIG, 42, "ok"], start=1):
        ws.cell(row=i, column=1, value=_xls_cell_value(v))
    out = tmp_path / "poisoned.xlsx"
    wb.save(str(out))
    back = openpyxl.load_workbook(str(out))
    # and the poisoned rows must be VISIBLE, not silently blank (the false-health half of class D)
    assert back.active.cell(1, 1).value == "[unrenderable inf]"
    assert back.active.cell(6, 1).value == 42


def test_class_d_diff_workbook_survives_a_poisoned_number(tmp_path):
    """html.py's diff/campaign workbooks wrote raw values into openpyxl cells with NO sanitizer at
    all (unlike excel.py). A huge-int leaf therefore aborted the whole --compare deliverable at
    wb.save(). Every data-row cell now routes through xml_safe."""
    from cisco_toolkit import html
    snap = {"devices": {"h": {"model": BIG, "sw_version": INF}},
            "interfaces": {"h": {"Gi0": {"port": "Gi0", "status": BIG}}},
            "health_scores": [{"switch": "h", "band": "Good", "score": BIG}],
            "punchlist": [{"severity": "High", "title": NAN}]}
    html.write_diff_workbook(snap, snap, str(tmp_path / "diff.xlsx"))
    html.write_campaign_workbook([snap, snap], str(tmp_path / "camp.xlsx"))
    assert (tmp_path / "diff.xlsx").exists() and (tmp_path / "camp.xlsx").exists()


# ---------------------------------------------------------------------------------------------
# CLASS A in the WORKBOOK sheet writers -- where a crash is SILENT, not loud
# ---------------------------------------------------------------------------------------------
@pytest.mark.parametrize("bad", [INF, NAN, BIG, True, "x", [1], {"k": 1}],
                         ids=["inf", "nan", "bigint", "True", "str", "list", "dict"])
def test_class_a_excel_sheet_writers_degrade_instead_of_dropping_a_sheet(bad):
    """The workbook is the ONE deliverable produced unconditionally (there is no --no-excel), and its
    sheet writers are wrapped PER SHEET by `_run_phase` -- so a crash here does not fail the run, it
    logs and saves a workbook with that entire sheet SILENTLY MISSING. A reader sees an absent sheet,
    not a failure: the coverage-honesty false all-clear `_xls_sanitize`'s own docstring warns about.

    Each writer now coerces its section ONCE AT ENTRY, because each reads it in several places (the
    row loop AND the trailing summary/log line) -- a per-site guard left the next read reachable."""
    import openpyxl

    from cisco_toolkit import excel
    wb = openpyxl.Workbook()
    excel.write_endpoint_intelligence_sheet(wb, bad)
    excel.write_endpoint_dependencies_sheet(wb, bad)
    excel.write_endpoint_dependencies_sheet(wb, {"clusters": bad, "dual_homed": bad, "affinity": bad})
    excel.write_subnet_reachability_sheet(wb, bad)
    excel.write_subnet_reachability_sheet(wb, {"per_device": bad})
    excel.write_architecture_review_sheet(wb, bad)
    # list SUBSECTIONS too: the element filter was present at these sites but the CONTAINER was not,
    # so `or []` handed a truthy non-list straight to `for ... in`.
    excel.write_architecture_review_sheet(wb, {"summary": {"score_pct": 50}, "domains": bad,
                                               "checks": bad, "top_actions": bad})
    excel.write_attestation_sheet(wb, {"claims": bad})
    excel.write_attestation_sheet(wb, bad)
    assert excel.ARCHREVIEW_SHEET_NAME in wb.sheetnames


@pytest.mark.parametrize("bad", [INF, BIG, True, "x", [1]], ids=["inf", "bigint", "True", "str", "list"])
def test_archreview_sheet_reports_unavailable_rather_than_a_clean_zero_scorecard(bad):
    """The SAFETY half: an UNREADABLE summary must take the UNAVAILABLE path, not be coerced to {}.

    `not A.get("summary")` caught absent/empty but not a truthy NON-dict, which crashed. Coercing it
    to {} would be worse than the crash -- it renders the clean 'grade N/A - 0 conform' scorecard the
    surrounding guard exists to prevent, i.e. 'reviewed, all fine' over a failed computation."""
    import openpyxl

    from cisco_toolkit import excel
    wb = openpyxl.Workbook()
    excel.write_architecture_review_sheet(wb, {"summary": bad})
    ws = wb[excel.ARCHREVIEW_SHEET_NAME]
    assert "unavailable" in str(ws.cell(2, 1).value or "").lower()


# ---------------------------------------------------------------------------------------------
# The golden snapshot itself must be UNAFFECTED by every guard added above (no silent behaviour change)
# ---------------------------------------------------------------------------------------------
def test_guards_are_inert_on_the_real_golden_snapshot():
    """Total-function check: the real producer artifact still flows through every hardened consumer,
    and the ACL shadow proof still finds the same findings it did before the port-operand guard."""
    from cisco_toolkit import aclcheck, coverage_matrix, html, nrfu_export, precert
    snap = _golden()
    assert isinstance(precert.compute_readiness_freeze(snap), dict)
    assert isinstance(precert.compute_precert(snap, snap), dict)
    flr = aclcheck.compute_filter_line_reachability(snap)
    assert flr["summary"]["n_findings"] == len(flr["findings"])
    assert isinstance(nrfu_export.compute_nrfu_commands(snap), dict)
    assert isinstance(coverage_matrix.compute_coverage_matrix(snap), dict)
    assert isinstance(html.compute_snapshot_delta(snap, snap), dict)
    assert isinstance(html._trend_point(snap), dict)
