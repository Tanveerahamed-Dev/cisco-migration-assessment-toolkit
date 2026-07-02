"""Tests for compute_cable_map — the EDA-style physical cable-map SSOT.

A Nokia-EDA-style cable map is a node/port/cable graph laid out in role tiers, with
op-status color derived from the underlying interface states. This engine is the single
source of truth both front-ends (explorer + webapp) render, so the model — tiering, LAG
bundling, and coverage-honest op-status ([NOT OBSERVED] neutral for uncollected devices) —
is pinned here rather than in JS/TS. (Nokia EDA physical-topology anatomy, deep-research 2026-07-01.)
"""
import json

from cisco_toolkit.analyze import compute_cable_map
from cisco_toolkit.model import InterfaceData


def _if(**kw):
    return InterfaceData(**kw)


def _nodes(cm):
    return {n["host"]: n for n in cm["nodes"]}


def test_basic_tiering_and_cable_up():
    """core seeds the top tier; a one-hop access switch lands one tier down; a link whose
    both observed ends are 'connected' is op_status 'up', with a port stub on each node."""
    all_ifaces = {
        "CORE-1": {"Gi1/0/1": _if(port="Gi1/0/1", status="connected",
                                  cdp_neighbor="ACC-1", neighbor_port="Gi1/0/24", endpoint_type="Switch")},
        "ACC-1": {"Gi1/0/24": _if(port="Gi1/0/24", status="connected",
                                  cdp_neighbor="CORE-1", neighbor_port="Gi1/0/1", endpoint_type="Switch")},
    }
    health = [{"switch": "CORE-1", "role": "core"}, {"switch": "ACC-1", "role": "access"}]
    cm = compute_cable_map(all_ifaces, health)

    assert set(cm) >= {"nodes", "cables", "tiers", "summary"}
    nodes = _nodes(cm)
    assert nodes["CORE-1"]["tier"] == 0            # core seeds the top tier
    assert nodes["ACC-1"]["tier"] == 1             # one hop down
    assert nodes["CORE-1"]["op_status"] == "up"    # collected device
    assert nodes["CORE-1"]["collected"] is True
    assert cm["tiers"][0] == ["CORE-1"]

    assert len(cm["cables"]) == 1
    c = cm["cables"][0]
    assert {c["a"], c["b"]} == {"CORE-1", "ACC-1"}
    assert c["op_status"] == "up"                  # both ends connected
    assert any(p["name"] == "Gi1/0/1" for p in nodes["CORE-1"]["ports"])
    assert any(p["name"] == "Gi1/0/24" for p in nodes["ACC-1"]["ports"])


def test_cable_down_when_either_observed_end_notconnect():
    """A 'notconnect'/'err-disabled' port makes the cable op_status 'down' — the down end wins,
    even if the far end reports connected. (EDA derives cable status from underlying interfaces.)"""
    all_ifaces = {
        "CORE-1": {"Gi1/0/1": _if(port="Gi1/0/1", status="notconnect",
                                  cdp_neighbor="ACC-1", neighbor_port="Gi1/0/24", endpoint_type="Switch")},
        "ACC-1": {"Gi1/0/24": _if(port="Gi1/0/24", status="connected",
                                  cdp_neighbor="CORE-1", neighbor_port="Gi1/0/1", endpoint_type="Switch")},
    }
    cm = compute_cable_map(all_ifaces, [{"switch": "CORE-1", "role": "core"},
                                        {"switch": "ACC-1", "role": "access"}])
    assert cm["cables"][0]["op_status"] == "down"

    # err-disabled is also 'down'
    all_ifaces["CORE-1"]["Gi1/0/1"].status = "err-disabled"
    cm2 = compute_cable_map(all_ifaces, None)
    assert cm2["cables"][0]["op_status"] == "down"


def test_uncollected_neighbor_is_not_observed_neutral():
    """An off-scan CDP neighbour (the 50 uncollected DS/CS core analogue) becomes a node with
    op_status 'unknown' ([NOT OBSERVED]) and an 'uncollected' badge — never a fake green."""
    all_ifaces = {
        "ACC-1": {"Gi1/0/24": _if(port="Gi1/0/24", status="connected",
                                  cdp_neighbor="DS-CORE", neighbor_port="Te1/1/1", endpoint_type="Switch")},
    }
    cm = compute_cable_map(all_ifaces, [{"switch": "ACC-1", "role": "access"}])
    nodes = _nodes(cm)
    assert "DS-CORE" in nodes                       # the off-scan peer is still a node in the cabling
    assert nodes["DS-CORE"]["collected"] is False
    assert nodes["DS-CORE"]["op_status"] == "unknown"
    assert "uncollected" in nodes["DS-CORE"]["badges"]


def test_cable_unknown_when_no_status_observed_either_end():
    """If neither end exposes an interface status (empty local status + off-scan far end),
    the cable is op_status 'unknown', not silently 'up'."""
    all_ifaces = {
        "ACC-1": {"Gi1/0/24": _if(port="Gi1/0/24", status="",
                                  cdp_neighbor="DS-CORE", neighbor_port="Te1/1/1", endpoint_type="Switch")},
    }
    cm = compute_cable_map(all_ifaces, None)
    assert cm["cables"][0]["op_status"] == "unknown"


def test_lag_members_bundled_into_one_cable():
    """Two physical links between the same pair that are port-channel members collapse into ONE
    bundled cable (is_pc) carrying both member port pairs — not two separate cables."""
    def leg(local, remote, peer):
        return _if(port=local, status="connected", port_channel="Po1", speed="10G",
                   cdp_neighbor=peer, neighbor_port=remote, endpoint_type="Switch")
    all_ifaces = {
        "CORE-1": {"Gi1/0/1": leg("Gi1/0/1", "Gi1/0/1", "DIST-1"),
                   "Gi1/0/2": leg("Gi1/0/2", "Gi1/0/2", "DIST-1")},
        "DIST-1": {"Gi1/0/1": leg("Gi1/0/1", "Gi1/0/1", "CORE-1"),
                   "Gi1/0/2": leg("Gi1/0/2", "Gi1/0/2", "CORE-1")},
    }
    cm = compute_cable_map(all_ifaces, [{"switch": "CORE-1", "role": "core"},
                                        {"switch": "DIST-1", "role": "distribution"}])
    pc = [c for c in cm["cables"] if c["is_pc"]]
    assert len(pc) == 1                             # bundled, not two
    assert len(pc[0]["members"]) == 2
    assert pc[0]["op_status"] == "up"
    assert pc[0]["speed"] == "10G"                  # bundle carries the member link speed


def test_degree_fallback_tiering_without_roles():
    """With no role evidence, the highest-degree node seeds the top tier (clab-io-draw pattern),
    so a hub-and-spoke fabric still tiers sensibly."""
    def leg(local, peer, remote):
        return _if(port=local, status="connected", cdp_neighbor=peer,
                   neighbor_port=remote, endpoint_type="Switch")
    all_ifaces = {
        "HUB": {f"Gi1/0/{i}": leg(f"Gi1/0/{i}", f"L{i}", "Gi1/0/1") for i in (1, 2, 3)},
        "L1": {"Gi1/0/1": leg("Gi1/0/1", "HUB", "Gi1/0/1")},
        "L2": {"Gi1/0/1": leg("Gi1/0/1", "HUB", "Gi1/0/2")},
        "L3": {"Gi1/0/1": leg("Gi1/0/1", "HUB", "Gi1/0/3")},
    }
    cm = compute_cable_map(all_ifaces, None)
    nodes = _nodes(cm)
    assert nodes["HUB"]["tier"] == 0               # highest degree seeds the top
    assert nodes["L1"]["tier"] == 1


def test_deterministic_output():
    """Same input -> byte-identical output (SSOT stability; the golden freezes this)."""
    all_ifaces = {
        "CORE-1": {"Gi1/0/1": _if(port="Gi1/0/1", status="connected",
                                  cdp_neighbor="ACC-1", neighbor_port="Gi1/0/24", endpoint_type="Switch")},
        "ACC-1": {"Gi1/0/24": _if(port="Gi1/0/24", status="connected",
                                  cdp_neighbor="CORE-1", neighbor_port="Gi1/0/1", endpoint_type="Switch")},
    }
    h = [{"switch": "CORE-1", "role": "core"}, {"switch": "ACC-1", "role": "access"}]
    a = json.dumps(compute_cable_map(all_ifaces, h), sort_keys=True)
    b = json.dumps(compute_cable_map(all_ifaces, h), sort_keys=True)
    assert a == b


def test_empty_input_is_safe():
    """Tolerant of an empty fleet (no crash, empty model)."""
    cm = compute_cable_map({}, None)
    assert cm["nodes"] == [] and cm["cables"] == [] and cm["tiers"] == []
    assert cm["summary"]["n_nodes"] == 0


def _demo_cable_map():
    return {
        "nodes": [], "tiers": [],
        "cables": [
            {"a": "CORE-1", "a_port": "Gi1/0/1", "b": "ACC-1", "b_port": "Gi1/0/24", "is_pc": False,
             "members": [{"a_port": "Gi1/0/1", "b_port": "Gi1/0/24"}], "op_status": "up", "confirmation": "Both ends",
             "speed": "1000"},
            {"a": "CORE-1", "a_port": "Gi1/0/2", "b": "ACC-2", "b_port": "Gi1/0/24", "is_pc": False,
             "members": [{"a_port": "Gi1/0/2", "b_port": "Gi1/0/24"}], "op_status": "down", "confirmation": "One end (CORE-1)",
             "speed": ""},
            {"a": "ACC-3", "a_port": "Te1/1/1", "b": "DS-CORE", "b_port": "Te1/1/1", "is_pc": False,
             "members": [{"a_port": "Te1/1/1", "b_port": "Te1/1/1"}], "op_status": "unknown", "confirmation": "One end (ACC-3)",
             "speed": "10G"},
            {"a": "CORE-1", "a_port": "Po1", "b": "DIST-1", "b_port": "Po1", "is_pc": True,
             "members": [{"a_port": "Gi1/0/3", "b_port": "Gi1/0/3"}, {"a_port": "Gi1/0/4", "b_port": "Gi1/0/4"}],
             "op_status": "up", "confirmation": "Both ends", "speed": "10G"},
        ],
        "summary": {"n_nodes": 0, "n_cables": 4, "n_tiers": 0, "op": {"up": 2, "down": 1, "unknown": 1}},
    }


def test_cabling_schedule_sheet():
    """The 'Cabling Schedule' workbook sheet is one row per cable from the cable_map SSOT: A/B host+port,
    LAG bundling, and op-status — coverage-honest (an unobserved end is '[NOT OBSERVED]', never 'Up')."""
    from openpyxl import Workbook
    from cisco_toolkit.excel import write_cabling_schedule_sheet, CABLING_SCHEDULE_SHEET_NAME
    wb = Workbook()
    write_cabling_schedule_sheet(wb, _demo_cable_map())
    assert CABLING_SCHEDULE_SHEET_NAME in wb.sheetnames
    ws = wb[CABLING_SCHEDULE_SHEET_NAME]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header[:4] == ["Switch A", "Port A", "Switch B", "Port B"]
    assert "Op-Status" in header
    assert "Speed" in header                                     # link speed from `show interface status`
    body = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert len(body) == 4                                        # one row per cable
    op_col = header.index("Op-Status")
    ops = [r[op_col] for r in body]
    assert ops.count("Up") == 2 and ops.count("Down") == 1       # derived, per-link
    assert ops.count("[NOT OBSERVED]") == 1                      # the uncollected end — never a fake 'Up'
    speeds = [r[header.index("Speed")] for r in body]
    assert "1000" in speeds and "10G" in speeds
    # the port-channel is typed as a bundle carrying 2 members
    type_col = header.index("Type")
    pc = [r for r in body if r[type_col] == "Port-channel"]
    assert len(pc) == 1 and pc[0][header.index("LAG Members")] == 2


def test_cabling_schedule_sheet_empty_is_safe():
    from openpyxl import Workbook
    from cisco_toolkit.excel import write_cabling_schedule_sheet, CABLING_SCHEDULE_SHEET_NAME
    wb = Workbook()
    write_cabling_schedule_sheet(wb, {"cables": []})
    assert CABLING_SCHEDULE_SHEET_NAME in wb.sheetnames        # sheet still emitted (with a placeholder row)


# ---------------------------------------------------------------------------
# Enhancement wave 2: speed + kind on the SSOT, and the cable-map DIFF for --compare
# ---------------------------------------------------------------------------

def test_cable_speed_and_kind_surface():
    """Cables carry the observed link speed; nodes carry an evidence-based `kind`. The load-bearing
    evidence is the advertised PLATFORM string — infer_endpoint_type classifies ANY 'cisco …'
    platform as 'Switch' (that is how cisco APs and IP phones enter the map as infra neighbours),
    so the platform is what lets the fabric-only declutter hide them without hiding real switches."""
    all_ifaces = {
        "CORE-1": {
            "Gi1/0/1": _if(port="Gi1/0/1", status="connected", speed="1000",
                           cdp_neighbor="ACC-1", neighbor_port="Gi1/0/24", endpoint_type="Switch"),
            "Gi1/0/2": _if(port="Gi1/0/2", status="connected", speed="a-1000",
                           cdp_neighbor="AP-LOBBY", neighbor_port="Gi0", endpoint_type="Switch",
                           neighbor_platform="cisco AIR-CAP3702I-E-K9"),
            "Gi1/0/3": _if(port="Gi1/0/3", status="connected",
                           cdp_neighbor="SEP001122", neighbor_port="Port 1", endpoint_type="Switch",
                           neighbor_platform="Cisco IP Phone 8845"),
            "Gi1/0/4": _if(port="Gi1/0/4", status="connected",
                           cdp_neighbor="WAN-RTR", neighbor_port="Gi0/0/1", endpoint_type="Router",
                           neighbor_platform="cisco ASR1001-X"),
            "Gi1/0/5": _if(port="Gi1/0/5", status="connected",
                           cdp_neighbor="DS-CORE", neighbor_port="Te1/1/1", endpoint_type="Switch",
                           neighbor_platform="cisco WS-C4510R+E"),
            # a positively-non-infra neighbour (camera) never enters the map at all — compute_topology_links
            # drops it via _is_infra_neighbor; document that so nobody expects a kind for it.
            "Gi1/0/6": _if(port="Gi1/0/6", status="connected",
                           cdp_neighbor="CAM-7", neighbor_port="eth0", endpoint_type="Camera"),
        },
        "ACC-1": {"Gi1/0/24": _if(port="Gi1/0/24", status="connected",
                                  cdp_neighbor="CORE-1", neighbor_port="Gi1/0/1", endpoint_type="Switch")},
    }
    cm = compute_cable_map(all_ifaces, [{"switch": "CORE-1", "role": "core"}])
    nodes = _nodes(cm)
    assert nodes["CORE-1"]["kind"] == "device" and nodes["ACC-1"]["kind"] == "device"   # scanned
    assert nodes["AP-LOBBY"]["kind"] == "ap"          # platform outranks the misclassified eptype 'Switch'
    assert nodes["SEP001122"]["kind"] == "phone"
    assert nodes["WAN-RTR"]["kind"] == "router"
    assert nodes["DS-CORE"]["kind"] == "switch"       # off-scan infra keeps its infra claim
    assert "CAM-7" not in nodes                       # non-infra endpoint: dropped upstream, no cable
    core_acc = [c for c in cm["cables"] if {c["a"], c["b"]} == {"CORE-1", "ACC-1"}][0]
    assert core_acc["speed"] == "1000"                # verbatim from the observed interface


def test_kind_infra_claim_wins_on_conflicting_platform_evidence():
    """A node observed by several switches classifies from the STRONGEST claim — if any observer's
    platform says switch, it is a switch (hiding is the risky direction; an AP-looking platform
    string from one observer never demotes a node another observer identifies as a switch)."""
    def leg(port, peer, plat):
        return _if(port=port, status="connected", cdp_neighbor=peer, neighbor_port="Te1/1/1",
                   endpoint_type="Switch", neighbor_platform=plat)
    all_ifaces = {
        "ACC-1": {"Gi1/0/25": leg("Gi1/0/25", "DS-CORE", "cisco AIR-LAP1142N")},
        "ACC-2": {"Gi1/0/25": leg("Gi1/0/25", "DS-CORE", "cisco WS-C3850-48P")},
    }
    cm = compute_cable_map(all_ifaces, None)
    assert _nodes(cm)["DS-CORE"]["kind"] == "switch"


def _cab(a, ap, b, bp, op="up", is_pc=False, members=None, speed=""):
    return {"a": a, "a_port": ap, "b": b, "b_port": bp, "is_pc": is_pc,
            "members": members or [{"a_port": ap, "b_port": bp}],
            "op_status": op, "confirmation": "Both ends", "speed": speed}


def test_cable_map_diff_added_removed_transitions_and_lag():
    """The --compare cable diff: undirected identity (a reversed record is the SAME cable), added/removed,
    op-status transitions with coverage-honest classification (up->unknown is 'no longer observed', NEVER
    'went down'), and LAG member-count changes (a pulled port-channel leg is a real cutover signal)."""
    from cisco_toolkit.analyze import compute_cable_map_diff
    old = {"cables": [
        _cab("CORE-1", "Gi1/0/1", "ACC-1", "Gi1/0/24", "up"),
        _cab("CORE-1", "Gi1/0/2", "ACC-2", "Gi1/0/24", "up"),
        _cab("CORE-1", "Gi1/0/3", "ACC-3", "Gi1/0/24", "up"),
        _cab("CORE-1", "Gi1/0/9", "OLD-SW", "Gi1/0/1", "up"),
        _cab("CORE-1", "Po1", "DIST-1", "Po1", "up", True,
             [{"a_port": "Gi1/0/5", "b_port": "Gi1/0/5"}, {"a_port": "Gi1/0/6", "b_port": "Gi1/0/6"}]),
    ]}
    new = {"cables": [
        _cab("ACC-1", "Gi1/0/24", "CORE-1", "Gi1/0/1", "up"),          # sides REVERSED -> same cable
        _cab("CORE-1", "Gi1/0/2", "ACC-2", "Gi1/0/24", "down"),        # went down
        _cab("CORE-1", "Gi1/0/3", "ACC-3", "Gi1/0/24", "unknown"),     # no longer observed (NOT down)
        _cab("CORE-1", "Gi1/0/7", "NEW-SW", "Gi1/0/1", "up"),          # added
        _cab("CORE-1", "Po1", "DIST-1", "Po1", "up", True,
             [{"a_port": "Gi1/0/5", "b_port": "Gi1/0/5"}]),            # LAG lost a leg
    ]}
    d = compute_cable_map_diff(old, new)
    assert d["assessed"] is True
    assert [c["b"] for c in d["added"]] == ["NEW-SW"]
    assert [c["b"] for c in d["removed"]] == ["OLD-SW"]
    st = {(s["a"], s["b"]): s for s in d["status_changes"]}
    assert st[("CORE-1", "ACC-2")]["from"] == "up" and st[("CORE-1", "ACC-2")]["to"] == "down"
    assert st[("CORE-1", "ACC-2")]["classification"] == "went down"
    assert st[("CORE-1", "ACC-3")]["classification"] == "no longer observed"
    assert d["summary"]["n_went_down"] == 1
    assert d["summary"]["n_no_longer_observed"] == 1
    assert len(d["members_changed"]) == 1
    assert d["members_changed"][0]["old_members"] == 2 and d["members_changed"][0]["new_members"] == 1
    # the reversed-side cable is unchanged — not an add+remove pair
    assert not any("ACC-1" in (c["a"], c["b"]) for c in d["added"] + d["removed"])


def test_cable_map_diff_not_assessed_and_tolerant():
    """Coverage-honest gating: a missing side, or zero cables on both sides, is NOT ASSESSED —
    never an implicit 'no cabling changes'. Tolerant of None/non-dict input."""
    from cisco_toolkit.analyze import compute_cable_map_diff
    assert compute_cable_map_diff(None, None)["assessed"] is False
    assert compute_cable_map_diff({}, {"cables": []})["assessed"] is False
    assert compute_cable_map_diff({"cables": [_cab("A", "1", "B", "2")]}, None)["assessed"] is False
    d = compute_cable_map_diff({"cables": [_cab("A", "1", "B", "2")]}, {"cables": []})
    assert d["assessed"] is True and d["summary"]["n_removed"] == 1


def test_cable_map_of_snapshot_prefers_stored_then_rehydrates():
    """The compare/webapp entry point: prefer the engine-computed snap['cable_map']; for a pre-feature
    snapshot, rehydrate the stored dict-interfaces (tolerating unknown legacy keys) and recompute."""
    from cisco_toolkit.analyze import cable_map_of_snapshot
    stored = {"nodes": [], "cables": [], "tiers": [], "summary": {"n_nodes": 0}}
    assert cable_map_of_snapshot({"cable_map": stored}) == stored
    snap = {"interfaces": {
        "CORE-1": {"Gi1/0/1": {"port": "Gi1/0/1", "status": "connected", "cdp_neighbor": "ACC-1",
                               "neighbor_port": "Gi1/0/24", "endpoint_type": "Switch",
                               "legacy_field_not_in_dataclass": "x"}},
        "ACC-1": {"Gi1/0/24": {"port": "Gi1/0/24", "status": "connected", "cdp_neighbor": "CORE-1",
                               "neighbor_port": "Gi1/0/1", "endpoint_type": "Switch"}},
    }, "health_scores": [{"switch": "CORE-1", "role": "core"}]}
    cm = cable_map_of_snapshot(snap)
    assert cm["summary"]["n_cables"] == 1 and cm["nodes"]
    assert cable_map_of_snapshot(None)["summary"]["n_cables"] == 0     # tolerant
    # a STALE stored section (pre-kind/speed engine schema) with evidence available -> RECOMPUTED,
    # so new front-end features (fabric filter, speed) work on old uploads (device_dossiers precedent)
    stale = {"nodes": [{"host": "X", "role": "", "tier": 0, "order": 0, "collected": True,
                        "op_status": "up", "badges": [], "ports": []}],
             "cables": [], "tiers": [["X"]], "summary": {"n_nodes": 1}}
    snap_stale = dict(snap)
    snap_stale["cable_map"] = stale
    cm2 = cable_map_of_snapshot(snap_stale)
    assert cm2["nodes"] and all("kind" in n for n in cm2["nodes"])     # fresh compute, not the stale copy
    # ...but with NO evidence to recompute from, the stale copy still beats an empty model
    assert cable_map_of_snapshot({"cable_map": stale}) == stale


def _delta_snap(acc_status="connected"):
    return {"interfaces": {
        "CORE-1": {"Gi1/0/1": {"port": "Gi1/0/1", "status": "connected", "cdp_neighbor": "ACC-1",
                               "neighbor_port": "Gi1/0/24", "endpoint_type": "Switch"}},
        "ACC-1": {"Gi1/0/24": {"port": "Gi1/0/24", "status": acc_status, "cdp_neighbor": "CORE-1",
                               "neighbor_port": "Gi1/0/1", "endpoint_type": "Switch"}},
    }}


def test_snapshot_delta_carries_cabling_and_verdict():
    """--compare: the delta gains a 'cabling' section (rehydrated when the snapshot predates the
    cable-map engine) and a physically-down cable drives the verdict to REGRESSED — a down cable
    after cutover is a hard regression, not a footnote."""
    from cisco_toolkit.html import compute_snapshot_delta
    d = compute_snapshot_delta(_delta_snap("connected"), _delta_snap("notconnect"))
    assert d["cabling"]["summary"]["n_went_down"] == 1
    assert d["verdict"] == "REGRESSED"
    assert "cable" in d["verdict_note"].lower()
    d2 = compute_snapshot_delta(_delta_snap("connected"), _delta_snap("connected"))
    assert d2["cabling"]["summary"]["n_went_down"] == 0
    assert d2["verdict"] == "CLEAN"


def test_diff_workbook_has_cabling_changes_sheet(tmp_path):
    """The --compare workbook carries a 'Cabling Changes' sheet from the same diff SSOT."""
    from openpyxl import load_workbook
    from cisco_toolkit.html import write_diff_workbook
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(_delta_snap("connected"), _delta_snap("notconnect"), str(out))
    wb = load_workbook(str(out), read_only=True)
    try:
        assert "Cabling Changes" in wb.sheetnames
        vals = [str(v) for row in wb["Cabling Changes"].iter_rows(values_only=True) for v in row if v]
        assert any("went down" in v for v in vals)
    finally:
        wb.close()
