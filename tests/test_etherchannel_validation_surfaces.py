"""EtherChannel evidence blockers remain visible and actionable on every bounded surface."""

import json
from pathlib import Path
import shutil
import subprocess

from openpyxl import Workbook
import pytest

from cisco_toolkit.excel import write_nrfu_commands_sheet, write_validation_plan_sheet


ROOT = Path(__file__).resolve().parents[1]
EXPLORER = ROOT / "cisco_toolkit" / "blast_radius_explorer.html"
NODE = shutil.which("node")


def _etherchannel_item(device: str, evidence_state: str) -> dict:
    if evidence_state == "degraded":
        expected = (
            "PRE-CUTOVER DEGRADED — BLOCKER: observed EtherChannel has 1 bundle(s), 2 member(s); "
            "1 not bundled — Gi1/0/2(I). Matching this degraded state after cutover is NOT ACCEPTANCE."
        )
        check = "EtherChannel degraded member baseline"
    elif evidence_state == "review":
        expected = (
            "PRE-CUTOVER REVIEW — BLOCKER: the member projection contradicts the current-run receipt; "
            "re-collect and verify live before acceptance."
        )
        check = "EtherChannel evidence contradiction"
    else:
        expected = (
            "ETHERCHANNEL BASELINE NOT VERIFIED — BLOCKER: the current-run receipt is not assessed; "
            "re-collect before acceptance."
        )
        check = "EtherChannel baseline not verified"
    return {
        "device": device,
        "platform": "ios",
        "wave": "Group 1",
        "category": "Link",
        "severity": "High",
        "check": check,
        "command": "show etherchannel summary",
        "expect": expected,
        "why": "A missing or degraded member baseline cannot certify bundle recovery.",
        "evidence_state": evidence_state,
        "projection_custody": "embedded_unverified",
        "source_key": (
            f"etherchannel_projection.rows[{device}] + "
            f"protocol_assessability.rows[{device},EtherChannel]"
        ),
    }


def _ordinary_item(index: int) -> dict:
    return {
        "device": "acc1",
        "platform": "ios",
        "wave": "Group 1",
        "category": "Gateway",
        "severity": "Medium",
        "check": f"ordinary check {index:02d}",
        "command": f"show interface Gi1/0/{index + 1}",
        "expect": "observed interface state is unchanged",
        "why": "Ordinary bounded validation detail.",
    }


def _checks() -> list[dict]:
    # Source-order blockers after the ordinary allowance so retention cannot pass accidentally.
    return [_ordinary_item(index) for index in range(45)] + [
        _etherchannel_item("acc1", "degraded"),
        _etherchannel_item("acc2", "review"),
        _etherchannel_item("acc3", "not_verified"),
    ]


def _plan() -> dict:
    checks = _checks()
    return {
        "items": checks,
        "by_wave": {"Group 1": checks},
        "summary": {
            "n_items": len(checks), "n_waves": 1, "n_high": 3,
            "by_category": {"Gateway": 45, "Link": 3},
        },
        "banner": "Run every cutover validation row and disposition every blocker.",
    }


def _snapshot() -> dict:
    hosts = ["acc1", "acc2", "acc3"]
    return {
        "script_version": "test",
        "devices": {host: {"platform": "ios"} for host in hosts},
        "move_groups": [{"switches": hosts, "endpoints": 8}],
        "migration_readiness": [{
            "group": "Group 1", "switches": hosts, "endpoints": 8,
            "readiness": "NOT READY", "n_fail": 1, "n_warn": 0, "checks": [],
        }],
        "wave_sequencing": [{
            "group": "Group 1", "make_before_break": hosts,
            "hard_cutover": [], "sequence": "make-before-break",
        }],
        "validation_plan": _plan(),
    }


def _tables_with_header(document, headers):
    expected = list(headers)
    return [
        table for table in document.tables
        if table.rows and [cell.text for cell in table.rows[0].cells] == expected
    ]


def test_etherchannel_workbooks_warn_and_project_evidence_custody_source():
    plan = _plan()
    wb = Workbook()
    write_validation_plan_sheet(wb, plan)
    validation = wb["Cutover Validation"]
    assert str(validation.cell(1, 1).value).startswith("⚠ ")
    assert "Evidence-state blockers apply to every protocol" in validation.cell(1, 1).value
    by_check = {
        validation.cell(row, 7).value: [validation.cell(row, col).value for col in range(9, 13)]
        for row in range(5, validation.max_row + 1)
    }
    for item in plan["items"][-3:]:
        rendered = by_check[item["check"]]
        assert rendered[1] == item["evidence_state"]
        assert "embedded projection; integrity unverified" in rendered[2]
        assert f"etherchannel_projection.rows[{item['device']}]" in rendered[3]
    assert by_check["EtherChannel baseline not verified"][0].startswith(
        "ETHERCHANNEL BASELINE NOT VERIFIED — BLOCKER:"
    )

    cases = [{
        "id": f"NRFU-W1-P2-{index:03d}", "phase": 2, "scope": "per-site",
        "command": item["command"], "expected": item["expect"],
        "source_key": item["source_key"], "evidence_state": item["evidence_state"],
        "projection_custody": item["projection_custody"],
    } for index, item in enumerate(plan["items"][-3:], start=1)]
    pack = {
        "waves": [{"wave_id": "Group 1", "devices": [
            {"host": item["device"], "platform_dialect": "ios", "cases": [case]}
            for item, case in zip(plan["items"][-3:], cases)
        ]}],
        "summary": {
            "n_cases": 3, "n_waves": 1, "n_devices": 3, "by_phase": {2: 3},
            "n_not_observed": 0, "n_human_executed": 0,
        },
        "banner": "Execute every NRFU case against its stated acceptance criterion.",
    }
    write_nrfu_commands_sheet(wb, pack)
    nrfu = wb["NRFU Commands"]
    assert str(nrfu.cell(1, 1).value).startswith("⚠ ")
    assert "Evidence-state blockers apply to every protocol" in nrfu.cell(1, 1).value
    assert [nrfu.cell(row, 10).value for row in range(5, 8)] == [
        "degraded", "review", "not_verified",
    ]
    assert all(nrfu.cell(row, 11).value == "embedded_unverified" for row in range(5, 8))


def test_runbook_retains_every_etherchannel_evidence_blocker_outside_cap(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "etherchannel-runbook.docx"
    write_runbook_docx(str(output), _snapshot(), "EtherChannel Surface Fixture")
    document = docx.Document(str(output))
    tables = _tables_with_header(
        document, ["Device", "Category", "Check", "Command", "Observed baseline / acceptance"])
    assert len(tables) == 1
    rows = [[cell.text for cell in row.cells] for row in tables[0].rows[1:]]
    assert len(rows) == 17  # 14 ordinary allowance + all three evidence blockers
    by_check = {row[2]: row for row in rows}
    for item in _plan()["items"][-3:]:
        assert item["check"] in by_check
        rendered = by_check[item["check"]][4]
        assert f"Evidence state: {item['evidence_state']}" in rendered
        assert "Projection custody: embedded_unverified" in rendered
        assert f"etherchannel_projection.rows[{item['device']}]" in rendered
        assert f"protocol_assessability.rows[{item['device']},EtherChannel]" in rendered
    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Any BASELINE NOT VERIFIED row is a blocker" in text
    assert "Every evidence-state blocker is retained above" in text


def test_mop_retains_etherchannel_blockers_and_capture_commands_outside_caps(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.mop import write_mop_docx

    output = tmp_path / "etherchannel-mop.docx"
    write_mop_docx(str(output), _snapshot(), "EtherChannel Surface Fixture")
    document = docx.Document(str(output))

    validation = _tables_with_header(
        document, ["Sev", "Category", "Device", "Check", "Command",
                   "Observed baseline / acceptance"])
    assert len(validation) == 1
    rows = [[cell.text for cell in row.cells] for row in validation[0].rows[1:]]
    assert len(rows) == 43  # 40 ordinary allowance + all three evidence blockers
    by_check = {row[3]: row for row in rows}
    assert {item["check"] for item in _plan()["items"][-3:]} <= set(by_check)

    captures = _tables_with_header(document, ["Device", "Command to capture"])
    assert len(captures) == 1
    capture_pairs = {(row.cells[0].text, row.cells[1].text) for row in captures[0].rows[1:]}
    assert len(capture_pairs) == 23  # 20 ordinary allowance + three blocker-specific pairs
    assert {(host, "show etherchannel summary") for host in ("acc1", "acc2", "acc3")} <= capture_pairs

    text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Any BASELINE NOT VERIFIED row is a blocker" in text
    assert "Every evidence-state blocker capture is retained above" in text
    assert "Every FHRP VLAN gate and evidence-state blocker is retained above" in text
    assert "A captured degraded state is evidence to resolve or disposition" in text


def test_explorer_demo_and_blocker_logic_are_etherchannel_truthful():
    html = EXPLORER.read_text(encoding="utf-8")
    assert "EtherChannel observed member baseline is degraded" in html
    assert "Gi1/0/24(I)" in html
    assert "Matching this degraded state after cutover is NOT ACCEPTANCE" in html
    assert 'evidence_state:"degraded",projection_custody:"embedded_unverified"' in html
    assert "etherchannel_projection.rows[ACC-1]" in html
    assert "Port-channel uplinks bundled" not in html

    block = html[html.index("function wavesValidationBlock(wave)"):html.index("function drawWaves()")]
    assert 'const blockers=ranked.filter(isBlocker)' in block
    assert 'const ordinary=ranked.filter(it=>!isBlocker(it))' in block
    assert 'const top=blockers.concat(ordinary.slice(0,3))' in block
    assert 'state==="degraded"?"DEGRADED"' in block
    assert '(state==="review"||state==="not_verified")?"REVIEW"' in block
    assert "isRoutingBlocker" not in block and "routingBlockers" not in block


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer blocker execution skipped")
def test_explorer_executes_all_etherchannel_blockers_outside_three_row_cap(tmp_path):
    html = EXPLORER.read_text(encoding="utf-8")
    code = html[html.index("function validationFor(wave)"):html.index("function drawWaves()")]
    payload = tmp_path / "validation.json"
    payload.write_text(json.dumps({"Group 1": _checks()}), encoding="utf-8")
    driver = tmp_path / "driver.js"
    driver.write_text(
        "const fs=require('fs');\n"
        "const byWave=JSON.parse(fs.readFileSync(process.argv[2],'utf8'));\n"
        "const SNAP={validation_plan:{by_wave:byWave}};\n"
        "const esc=value=>String(value??'');\n"
        "const shortName=value=>String(value??'');\n"
        + code
        + "\nprocess.stdout.write(wavesValidationBlock('Group 1'));\n",
        encoding="utf-8",
    )
    proc = subprocess.run(
        [NODE, str(driver), str(payload)], capture_output=True, text=True, timeout=30,
    )
    assert proc.returncode == 0, proc.stderr
    rendered = proc.stdout
    for item in _plan()["items"][-3:]:
        assert item["check"] in rendered
        assert f"<b>evidence</b> {item['evidence_state']}" in rendered
    assert "ETHERCHANNEL BASELINE NOT VERIFIED" in rendered
    assert "ordinary check 00" in rendered and "ordinary check 02" in rendered
    assert "ordinary check 03" not in rendered
