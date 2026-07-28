"""[review 2026-07-28 #24 / #53 / #85 -- cisco_toolkit/html.py] Three defects in the snapshot-reporting
layer, all of the same family: a fact that is NOT OBSERVED, or that already has an owner, is re-derived
locally and rendered as an observed number.

  * #24 ``_trend_point`` defaulted four migration metrics (``n_critical`` / ``n_punchlist`` /
    ``n_crit_high`` / ``n_not_ready``) to a hard ``0`` when their snapshot sections were simply ABSENT --
    a partial upload, an older schema, or a failed analysis phase. Every counter therefore FELL to zero,
    which the campaign trajectory reads as four metrics improving and the trend workbook paints as a green
    ``CAMPAIGN VERDICT: IMPROVING``. The went-dark guard cannot catch it: ``devices`` is unchanged and there
    are no health bands at all. ``avg_health`` and ``past_ldos`` in the SAME dict already abstained with
    ``""`` -- the coverage-honest idiom was known and applied to 2 of the 6 metrics.
  * #53 ``write_diff_workbook``'s Summary printed the fleet size as a raw ``len(devices)`` while
    ``compute_snapshot_delta`` -- called eleven lines earlier on the SAME two inputs, and what the webapp
    ``POST /api/compare`` returns -- publishes the canonical ``executive_brief.scale.n_devices``. One call,
    two fleet sizes: the AssessHub compare view and the CLI diff workbook disagree at the cutover gate.
  * #85 the same Summary re-derived the interface totals with raw ``len()`` where the delta routes the
    identical read through ``_as_dict``, so a trimmed / foreign-tool / older-schema snapshot
    (``interfaces: {"sw1": 5}``) raised ``TypeError`` -- no diff workbook and no Pre-Change Validation
    Certificate. The string variant (``{"sw1": "abc"}``) silently counted 3 phantom interfaces first.

Non-vacuity: each test below was run against the pre-fix module and fails there -- #24 reads
``n_punchlist == 0`` and ``verdict == "IMPROVING"``, #53 reads 2/3 instead of 303/305, #85 raises
``TypeError: object of type 'int' has no len()``.
"""
import pytest
from openpyxl import load_workbook

from cisco_toolkit.html import (_trend_point, compute_campaign_trend, compute_snapshot_delta,
                                write_campaign_workbook, write_diff_workbook)

_DEVS = {"sw1": {}, "sw2": {}, "sw3": {}}


def _assessed_snap(date="2026-01-01"):
    """A collection whose analysis phase ran: every migration metric has real evidence behind it."""
    return {
        "generated_at": date + "T00:00:00", "script_version": "V3.23.0",
        "devices": dict(_DEVS), "interfaces": {},
        "health_scores": [{"switch": "sw1", "band": "Critical", "score": 20},
                          {"switch": "sw2", "band": "Poor", "score": 45},
                          {"switch": "sw3", "band": "Fair", "score": 60}],
        "punchlist": [{"severity": "Critical", "category": "X", "title": "A", "devices": ["sw1"]},
                      {"severity": "High", "category": "X", "title": "B", "devices": ["sw2"]},
                      {"severity": "Low", "category": "X", "title": "C", "devices": ["sw1"]}],
        "migration_readiness": [{"readiness": "NOT READY", "switches": ["sw1"]},
                                {"readiness": "NOT READY", "switches": ["sw2"]}],
        "lifecycle_risk": {"summary": {"n_past_ldos": 4, "n_past_eos": 0}},
        "executive_brief": {"posture": {"avg_health": 41.7, "n_critical": 1}},
    }


def _uncollected_snap(date="2026-02-01"):
    """The SAME fleet, but the analysis sections never arrived. Not a statement that anything is clean."""
    return {"generated_at": date + "T00:00:00", "script_version": "V3.23.0",
            "devices": dict(_DEVS), "interfaces": {}}


# --------------------------------------------------------------------------------------------------
# #24 -- absence must abstain, not tally to zero
# --------------------------------------------------------------------------------------------------
@pytest.mark.parametrize("key", ["n_critical", "n_punchlist", "n_crit_high", "n_not_ready"])
def test_trend_point_abstains_when_the_section_was_never_collected(key):
    """A missing section is UNKNOWN, so it renders like avg_health/past_ldos already do: ''. A hard 0 is an
    observed count and this one was never observed."""
    p = _trend_point(_uncollected_snap())
    assert p[key] == "", f"{key} must abstain on an absent section, not report {p[key]!r}"
    # the two that were already right are the reference behaviour, pinned alongside
    assert p["avg_health"] == "" and p["past_ldos"] == ""


@pytest.mark.parametrize("key, expected", [("n_critical", 0), ("n_punchlist", 0),
                                           ("n_crit_high", 0), ("n_not_ready", 0)])
def test_trend_point_still_reports_a_real_observed_zero(key, expected):
    """Non-vacuity: the abstention keys on the section being ABSENT, not on the tally being zero. A
    collection that ran and found nothing must still publish 0 -- otherwise a genuinely clean fleet would
    be reported as uncollected, which is the same lie in the other direction."""
    clean = dict(_uncollected_snap(), health_scores=[], punchlist=[], migration_readiness=[])
    assert _trend_point(clean)[key] == expected


def test_trend_point_well_formed_snapshot_is_unchanged():
    """The guard must be invisible to a normal snapshot."""
    p = _trend_point(_assessed_snap())
    assert p["n_critical"] == 1 and p["n_punchlist"] == 3
    assert p["n_crit_high"] == 2 and p["n_not_ready"] == 2
    assert p["avg_health"] == 41.7 and p["past_ldos"] == 4


def test_campaign_with_an_uncollected_later_snapshot_is_not_improving():
    """The headline harm: C2 has the SAME devices as C1 but no analysis sections, so every counter fell to
    0 and the campaign read 'IMPROVING' off four metrics that were never measured."""
    t = compute_campaign_trend([_assessed_snap(), _uncollected_snap()])
    assert t["verdict"] != "IMPROVING", t["verdict_note"]
    assert not any(r["direction"] == "improving" for r in t["trajectory"]), t["trajectory"]
    assert "NOT COMPARABLE" in t["verdict_note"]
    assert "Punch-list items" in t["verdict_note"]      # the lost metrics are NAMED, not just counted


def test_campaign_partial_evidence_loss_downgrades_a_clean_improvement():
    """A campaign that keeps its health scores but LOSES its punch-list mid-way still improves on the
    metrics that remain -- it just cannot claim a clean IMPROVING over an estate that shrank its evidence,
    the same rule a device going dark already triggers."""
    c1 = _assessed_snap()
    c2 = _assessed_snap("2026-02-01")
    c2["executive_brief"] = {"posture": {"avg_health": 88.0, "n_critical": 0}}
    c2["health_scores"] = [{"switch": h, "band": "Good", "score": 88} for h in _DEVS]
    c2.pop("punchlist")                                  # the analysis phase stopped emitting it
    t = compute_campaign_trend([c1, c2])
    assert t["verdict"] == "MIXED", (t["verdict"], t["verdict_note"])
    assert "NOT COMPARABLE" in t["verdict_note"]
    metrics = {r["metric"] for r in t["trajectory"]}
    assert "Punch-list items" not in metrics             # never scored as a fall to zero
    assert "Avg health / 100" in metrics                 # the comparable metrics still trend


def test_campaign_workbook_does_not_paint_an_uncollected_campaign_green(tmp_path):
    """The client-visible surface: write_campaign_workbook stamps trend['verdict'] into Campaign Summary B2
    with a colour fill, so the abstention has to reach the cell, not just the dict."""
    out = tmp_path / "campaign.xlsx"
    write_campaign_workbook([_assessed_snap(), _uncollected_snap()], str(out))
    wb = load_workbook(str(out))
    assert wb["Campaign Summary"].cell(2, 2).value != "IMPROVING"
    assert "NOT COMPARABLE" in (wb["Campaign Summary"].cell(2, 3).value or "")
    # the uncollected collection's timeline cells are blank, not zeros presented as measurements
    tl = wb["Timeline"]
    assert [tl.cell(3, c).value for c in (5, 6, 7, 8, 9, 10)] == [None] * 6


# --------------------------------------------------------------------------------------------------
# #53 / #85 -- the diff workbook's Summary cites the delta instead of recomputing it
# --------------------------------------------------------------------------------------------------
def _scaled_snap(n_devices, hosts, interfaces):
    return {"devices": {h: {} for h in hosts}, "interfaces": interfaces,
            "health_scores": [], "punchlist": [],
            "executive_brief": {"scale": {"n_devices": n_devices}}}


def _summary_rows(path):
    ws = load_workbook(str(path))["Summary"]
    return {ws.cell(r, 1).value: (ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value)
            for r in range(2, ws.max_row + 1)}


def test_diff_workbook_switch_count_is_the_canonical_one(tmp_path):
    """SSOT (Law 1): one call must not produce two fleet sizes. The fixture is discriminating -- canonical
    303/305 vs a collected 2/3 -- so a regression to len(devices) renders 2 and trips this."""
    old = _scaled_snap(303, ["sw1", "sw2"], {"sw1": {"Gi1/0/1": {"status": "connected"}}})
    new = _scaled_snap(305, ["sw1", "sw2", "sw3"], {"sw1": {"Gi1/0/1": {"status": "connected"}}})
    delta = compute_snapshot_delta(old, new)
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(old, new, str(out))
    rows = _summary_rows(out)
    assert rows["Switches"][:2] == (delta["switches"]["old"], delta["switches"]["new"]) == (303, 305)
    assert rows["Switches"][2] == 2
    # the collected-device enumeration is labelled as such, so the two rows are not one contradiction
    assert rows["Switches added (collected)"][2] == "sw3"


def test_diff_workbook_interface_total_is_the_delta_s(tmp_path):
    old = _scaled_snap(2, ["sw1"], {"sw1": {"Gi1/0/1": {}, "Gi1/0/2": {}}})
    new = _scaled_snap(2, ["sw1"], {"sw1": {"Gi1/0/1": {}}})
    delta = compute_snapshot_delta(old, new)
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(old, new, str(out))
    rows = _summary_rows(out)
    assert rows["Interfaces (total)"][:2] == (delta["interfaces"]["old"], delta["interfaces"]["new"]) == (2, 1)
    assert rows["Interfaces (total)"][2] == -1


@pytest.mark.parametrize("poison, n_old, why", [
    (5, 0, "a hand-trimmed / older-schema snapshot storing a COUNT where the port map belongs"),
    ("abc", 0, "a string, which len() happily measures -> 3 phantom interfaces before a later crash"),
    (None, 0, "a present-but-null per-host value"),
    ({"Gi1/0/1": 7}, 1, "the per-PORT variant: a scalar where the field dict belongs -- the PORT is real"),
], ids=["int", "str", "null", "scalar_port"])
def test_diff_workbook_survives_a_malformed_interfaces_section(tmp_path, poison, n_old, why):
    """compute_snapshot_delta already routed this read through _as_dict and documented the crash it
    prevents; the Summary re-derived it raw seven lines later. A raise here means no diff workbook and no
    Pre-Change Validation Certificate -- at the cutover gate."""
    old = {"devices": {"sw1": {}}, "interfaces": {"sw1": poison}}
    new = {"devices": {"sw1": {}}, "interfaces": {"sw1": {"Gi1/0/1": {"status": "connected"}}}}
    delta = compute_snapshot_delta(old, new)
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(old, new, str(out))                       # must not raise
    rows = _summary_rows(out)
    assert rows["Interfaces (total)"][:2] == (delta["interfaces"]["old"], delta["interfaces"]["new"]), why
    assert rows["Interfaces (total)"][0] == n_old, "no phantom interfaces from an unusable port map"


def test_diff_workbook_delta_column_survives_a_non_numeric_canonical_count(tmp_path):
    """The canonical count comes from the snapshot, so a malformed one can carry a string there; the
    subtraction must degrade rather than abort the whole workbook."""
    old = _scaled_snap("lots", ["sw1"], {"sw1": {}})
    new = _scaled_snap(5, ["sw1", "sw2"], {"sw1": {}})
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(old, new, str(out))
    assert _summary_rows(out)["Switches"][2] in (None, "")
