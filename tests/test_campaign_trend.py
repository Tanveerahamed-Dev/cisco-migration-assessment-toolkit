"""NEW-V3.23.145: migration-campaign trend across a SERIES of snapshots. Pins the trajectory verdict
(IMPROVING / REGRESSING / INSUFFICIENT), the per-metric directions, the per-step findings burndown
(reusing compute_snapshot_delta), and the trend workbook's sheets."""
from openpyxl import load_workbook

from cisco_toolkit.html import compute_campaign_trend, write_campaign_workbook


def _snap(date, avg, bands, punch, not_ready, ldos):
    """bands: {band: count} -> health_scores with stable switch names; punch: list of (severity, title).
    ldos = n_past_ldos (past last-day-of-support = the engine's canonical "past end-of-support", the
    migration-critical unsupported count the brief/deck/explorer headline). n_past_eos (past end-of-SALE,
    still supported) is PINNED to 0 here so the 'Past end-of-support' trajectory metric can read as
    'improving' ONLY if it sources n_past_ldos, never n_past_eos -- guards the EoS/LDoS silent-drop class
    (a fleet can be 0 past-EoS yet have 152 past-LDoS; reading the wrong field hides every unsupported box)."""
    hs, i = [], 0
    for band, cnt in bands.items():
        for _ in range(cnt):
            hs.append({"switch": f"sw{i}", "score": avg, "band": band}); i += 1
    return {
        "generated_at": date + "T00:00:00", "script_version": "V3.23.0",
        "devices": {h["switch"]: {} for h in hs}, "interfaces": {},
        "health_scores": hs,
        "punchlist": [{"severity": s, "category": "X", "title": t, "devices": ["sw0"]} for (s, t) in punch],
        "migration_readiness": [{"readiness": "NOT READY", "switches": ["sw0"]} for _ in range(not_ready)],
        "lifecycle_risk": {"summary": {"n_past_ldos": ldos, "n_past_eos": 0}},
        "executive_brief": {"posture": {"avg_health": avg}},
    }


C1 = _snap("2026-01-01", 50, {"Critical": 2, "Fair": 1, "Good": 1},
           [("Critical", "A"), ("High", "B"), ("High", "C"), ("Low", "D")], 2, 4)
C2 = _snap("2026-02-01", 65, {"Critical": 1, "Fair": 2, "Good": 1},
           [("High", "B"), ("Low", "D")], 1, 2)
C3 = _snap("2026-03-01", 80, {"Good": 2, "Excellent": 2},
           [("Low", "D")], 0, 0)


def test_improving_campaign_verdict():
    t = compute_campaign_trend([C1, C2, C3])
    assert t["verdict"] == "IMPROVING"
    assert len(t["timeline"]) == 3
    dirs = {r["metric"]: r["direction"] for r in t["trajectory"]}
    assert dirs["Avg health / 100"] == "improving"
    assert dirs["Critical-band switches"] == "improving"
    assert dirs["Critical/High findings"] == "improving"
    assert dirs["Past end-of-support"] == "improving"
    assert t["timeline"][0]["avg_health"] == 50 and t["timeline"][-1]["avg_health"] == 80
    assert t["timeline"][0]["n_critical"] == 2 and t["timeline"][-1]["n_critical"] == 0


def test_burndown_resolves_findings_per_step():
    t = compute_campaign_trend([C1, C2, C3])
    assert len(t["steps"]) == 2
    s1 = t["steps"][0]
    assert s1["from"] == "C1" and s1["to"] == "C2"
    assert s1["resolved"] >= 2 and s1["opened"] == 0   # C1->C2 cleared the Critical A + High C findings


def test_regressing_campaign():
    t = compute_campaign_trend([C3, C1])               # reversed series = the network getting worse
    assert t["verdict"] == "REGRESSING"


def test_insufficient_single_snapshot():
    t = compute_campaign_trend([C1])
    assert t["verdict"] == "INSUFFICIENT" and t["trajectory"] == [] and t["steps"] == []


def test_campaign_workbook_sheets(tmp_path):
    out = tmp_path / "trend.xlsx"
    write_campaign_workbook([C1, C2, C3], str(out))
    wb = load_workbook(str(out))
    assert wb.sheetnames == ["Campaign Summary", "Timeline", "Burndown"]
    summ = wb["Campaign Summary"]
    assert summ.cell(2, 1).value == "CAMPAIGN VERDICT" and summ.cell(2, 2).value == "IMPROVING"
    tl = wb["Timeline"]
    assert tl.cell(1, 1).value == "Collection" and tl.max_row == 4    # header + 3 collections
    bd = wb["Burndown"]
    assert bd.cell(1, 1).value == "Step" and bd.max_row == 3          # header + 2 steps


def test_campaign_trend_not_improving_when_a_device_goes_dark():
    """[audit-2 #4] a device present then ABSENT across the campaign (gone dark) makes a rising avg-health
    survivorship-biased; the verdict must NOT read IMPROVING -- it downgrades to MIXED and discloses the dark
    device (else a failed / decommissioned-without-record device silently reads as improvement)."""
    from cisco_toolkit.html import compute_campaign_trend
    def snap(devs, scores):
        return {"devices": {d: {} for d in devs}, "interfaces": {d: {} for d in devs},
                "health_scores": [{"switch": s, "band": b, "score": sc} for s, b, sc in scores],
                "punchlist": [], "executive_brief": {"scale": {"n_devices": len(devs)}}}
    old = snap(["core1", "core2", "acc1"], [("core1", "Good", 90), ("core2", "Good", 88), ("acc1", "Fair", 70)])
    new = snap(["core1", "acc1"], [("core1", "Good", 90), ("acc1", "Good", 80)])   # core2 went dark
    t = compute_campaign_trend([old, new])
    assert t["verdict"] != "IMPROVING"
    assert "dark" in t["verdict_note"].lower()


def test_cutover_diff_insufficient_data_is_coverage_not_health_improvement():
    """[audit-4 #5 false-health] 'Insufficient Data' ranked WORSE than Critical, so a dark device re-collected and
    found Critical classified as 'improved' + CLEAN -- the single most safety-critical gate reading a newly-online
    Critical box as an improvement. Insufficient Data is a COVERAGE state, not a health-scale point."""
    from cisco_toolkit import html
    base = {"devices": {"A": {}}, "interfaces": {"A": {}}, "punchlist": []}
    dark = dict(base, health_scores=[{"switch": "A", "band": "Insufficient Data", "score": 90}])
    crit = dict(base, health_scores=[{"switch": "A", "band": "Critical", "score": 12}])
    d = html.compute_snapshot_delta(dark, crit)        # newly collected, found Critical
    assert d["health"]["n_improved"] == 0 and d["verdict"] != "CLEAN"
    d2 = html.compute_snapshot_delta(crit, dark)       # went dark = coverage loss
    assert d2["health"]["n_improved"] == 0 and d2["verdict"] != "CLEAN"
    good = dict(base, health_scores=[{"switch": "A", "band": "Good", "score": 85}])
    d3 = html.compute_snapshot_delta(crit, good)       # genuine health improvement still counts
    assert d3["health"]["n_improved"] == 1


def test_cutover_diff_survives_null_list_rows():
    """[audit-4 #15 totality] a null element in health_scores/punchlist (hand-trimmed / older-schema snapshot fed
    to --compare/--trend) crashed compute_snapshot_delta with AttributeError instead of degrading."""
    from cisco_toolkit import html
    old = {"devices": {"A": {}}, "interfaces": {"A": {}},
           "health_scores": [None, {"switch": "A", "band": "Good", "score": 80}], "punchlist": [None]}
    new = {"devices": {"A": {}}, "interfaces": {"A": {}}, "health_scores": [], "punchlist": []}
    assert isinstance(html.compute_snapshot_delta(old, new), dict)   # must not raise


def test_delta_title_rename_alias_prevents_false_churn():
    """PR-#396 review: comparing an OLD-engine baseline against a NEW-engine snapshot of an IDENTICAL
    network reported the native-VLAN-1 title rename as 1 opened + 1 resolved and flipped the cutover
    verdict CLEAN -> REVIEW. html._finding_key now aliases the pre-rename wording onto the current one
    (_TITLE_RENAMES); a GENUINE scope change (different count) must still churn honestly."""
    from cisco_toolkit.html import compute_snapshot_delta

    base = {"devices": {}, "interfaces": {}, "health_scores": []}
    row = {"severity": "Low", "category": "False-health", "devices": ["acc1", "core1"]}
    old = dict(base, punchlist=[dict(row, title="Native VLAN 1 on 4 inter-switch trunk(s)")])
    new = dict(base, punchlist=[dict(row, title="Native VLAN 1 on 4 operationally-trunking port(s)")])
    d = compute_snapshot_delta(old, new)
    assert d["findings"]["n_opened"] == 0 and d["findings"]["n_resolved"] == 0
    assert d["verdict"] == "CLEAN"
    # a genuine scope change (count moved 4 -> 5) still shows honestly as resolved+opened
    new2 = dict(base, punchlist=[dict(row, title="Native VLAN 1 on 5 operationally-trunking port(s)")])
    d2 = compute_snapshot_delta(old, new2)
    assert d2["findings"]["n_opened"] == 1 and d2["findings"]["n_resolved"] == 1


def _lc_step(day, n_past_ldos, n_unknown, health=80):
    return {"devices": {"sw1": {}}, "generated_at": f"2026-0{day}-01T00:00:00",
            "lifecycle_risk": {"summary": {"n_past_ldos": n_past_ldos, "n_unknown": n_unknown},
                               "per_device": []},
            "health_scores": [{"switch": "sw1", "band": "Good", "score": health}],
            "punchlist": [], "executive_brief": {"posture": {"avg_health": health}}}


def _past_eos_metric(trend):
    return next((m for m in trend["trajectory"] if m["metric"] == "Past end-of-support"), None)


def test_an_unassessed_campaign_step_does_not_trend_as_an_IMPROVEMENT():
    """`past_ldos` is lower-is-better, so a 0 that means 'never measured' scored as the BEST value.

    A campaign whose first snapshot found 5 past-LDoS devices and whose second could not match ANY
    platform against the offline EoX KB reported 5 -> 0 and trended as improving. Absence converted
    into a positive signal — worse than absence rendered as neutral, because someone acts on it.

    A count over partial coverage is not comparable to one over full coverage, so the metric drops
    out for that step (the dict's existing "" convention) rather than being scored.
    """
    t = compute_campaign_trend([_lc_step(1, 5, 0), _lc_step(2, 0, 4)])
    assert _past_eos_metric(t) is None, (
        f"an all-unknown step was trended: {_past_eos_metric(t)}"
    )


def test_a_genuine_fully_assessed_improvement_still_trends():
    """Non-vacuity: dropping incomplete steps must not silence real progress."""
    t = compute_campaign_trend([_lc_step(1, 5, 0), _lc_step(2, 2, 0)])
    m = _past_eos_metric(t)
    assert m is not None and m["direction"] == "improving", m


# ── W1/F4: a metric absent from BOTH ends was dropped without a word ────────────
def _full_step(day, n_past_ldos, n_unknown, health=80, lifecycle=True):
    """A collection in which EVERY `_TREND_METRICS` entry is measured, so the disclosure under
    test is attributable to the one metric the case withholds.

    `_lc_step` above carries no `migration_readiness`, which makes NOT-READY-groups abstain at both
    ends too — itself an instance of the defect, and the reason this builder is separate rather
    than an edit to a helper other tests depend on."""
    step = {
        "devices": {"sw1": {}},
        "generated_at": f"2026-0{day}-01T00:00:00",
        "health_scores": [{"switch": "sw1", "band": "Good", "score": health}],
        "punchlist": [{"severity": "High", "title": "t", "category": "c", "devices": ["sw1"]}],
        "migration_readiness": [{"group": "g1", "readiness": "READY"}],
        "executive_brief": {"posture": {"avg_health": health, "n_critical": 0}},
    }
    if lifecycle:
        step["lifecycle_risk"] = {"summary": {"n_past_ldos": n_past_ldos,
                                              "n_unknown": n_unknown}, "per_device": []}
    return step


def test_a_metric_unmeasured_at_BOTH_ends_is_disclosed_not_silently_dropped():
    """The one-sided case was disclosed; the both-sided case was not, and it is the worse one.

    `if not ok_a or not ok_b: if ok_a != ok_b: lost.append(metric)` — when NEITHER endpoint carried
    the metric it left the trajectory with no trace at all. "Never part of this campaign's
    evidence" was an assumption, not an observation: `_trend_point` abstains for a metric whose
    analysis FAILED as well as for one never collected. Measured on the pre-fix code: both ends
    all-Unknown gave verdict FLAT, no lifecycle row, and no NOT-COMPARABLE sentence — a reader
    cannot tell "we looked and found nothing" from "we never looked"."""
    t = compute_campaign_trend([_full_step(1, 0, 4), _full_step(2, 0, 4)])
    assert _past_eos_metric(t) is None, "precondition: the metric is not comparable, so it drops"
    assert t["not_comparable"]["never_measured"] == ["Past end-of-support"], t["not_comparable"]
    assert "NOT COMPARABLE" in t["verdict_note"]
    assert "never measured at EITHER end" in t["verdict_note"]
    assert "Past end-of-support" in t["verdict_note"]
    assert "NOT because there is nothing to report" in t["verdict_note"]


def test_a_fully_measured_campaign_gains_no_not_comparable_disclosure():
    """Non-vacuity: the healthy case must NOT acquire the new sentence, or the guard is always-on
    and proves nothing. Every `_TREND_METRICS` entry is numeric at both ends here."""
    t = compute_campaign_trend([_full_step(1, 5, 0), _full_step(2, 2, 0)])
    assert _past_eos_metric(t) is not None, "precondition: this metric IS comparable"
    assert t["not_comparable"] == {"lost": [], "never_measured": []}, t["not_comparable"]
    assert "NOT COMPARABLE" not in t["verdict_note"], t["verdict_note"]


def test_the_two_not_comparable_kinds_are_reported_apart():
    """One endpoint missing is a collection gap in that run; both missing means the metric was
    never measured in this campaign at all. Different actions, so they are not merged into one
    count — and the one-sided kind keeps its existing verdict downgrade while the both-sided kind
    does not (there is no evidence in either direction to downgrade on)."""
    t = compute_campaign_trend([_full_step(1, 5, 0), _full_step(2, 0, 4)])
    assert t["not_comparable"]["lost"] == ["Past end-of-support"], t["not_comparable"]
    assert t["not_comparable"]["never_measured"] == [], t["not_comparable"]
    assert "lost their evidence between the first and last collection" in t["verdict_note"]
    assert "never measured at EITHER end" not in t["verdict_note"]


# ── R8/F3: the disclosure has TWO render exits; the web adapter is the second ───
def _webapp_engine():
    """The AssessHub adapter (`webapp/backend/engine.py`) — the web campaign view's ONLY route to
    `compute_campaign_trend`, and therefore the second exit the coverage disclosure has to reach."""
    import os
    import sys
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sys.path.insert(0, os.path.join(root, "webapp"))
    import backend.engine as engine
    return engine


def test_the_web_adapter_carries_the_not_comparable_disclosure():
    """The workbook prints `verdict_note`; so does the web campaign view. Both therefore carry the
    NOT-COMPARABLE sentence today — but only because the engine happens to put it in the prose, and
    the adapter passed the dict through unexamined. It now normalises the structural twin and
    guarantees the prose the UI renders states it."""
    engine = _webapp_engine()
    out = engine.campaign_trend([_full_step(1, 0, 4), _full_step(2, 0, 4)])
    assert out["not_comparable"]["never_measured"] == ["Past end-of-support"], out["not_comparable"]
    assert out["not_comparable"]["disclosure_available"] is True
    assert "NOT COMPARABLE" in out["verdict_note"], out["verdict_note"]

    # NON-VACUITY: a fully-comparable campaign must NOT acquire an invented caveat.
    clean = engine.campaign_trend([_full_step(1, 5, 0), _full_step(2, 2, 0)])
    assert clean["not_comparable"] == {"lost": [], "never_measured": [],
                                       "disclosure_available": True}, clean["not_comparable"]
    assert "NOT COMPARABLE" not in clean["verdict_note"], clean["verdict_note"]


def test_the_web_adapter_fails_CLOSED_when_the_disclosure_is_absent():
    """The failure this guards is silent by construction: a trend dict with no comparability record
    renders a verdict with NO coverage caveat, and an absent caveat reads as "every metric was
    comparable". Missing input must fail closed — say the coverage is UNKNOWN — never open."""
    engine = _webapp_engine()
    missing = engine._carry_not_comparable({"verdict": "IMPROVING", "verdict_note": "all good",
                                            "trajectory": []})
    assert missing["not_comparable"]["disclosure_available"] is False, missing["not_comparable"]
    assert "NOT COMPARABLE" in missing["verdict_note"] and "UNKNOWN" in missing["verdict_note"]
    assert "all good" in missing["verdict_note"], "the engine's own note must survive"

    # A populated record whose prose forgot to mention it is restated, because the UI renders the
    # prose and nothing else.
    silent = engine._carry_not_comparable(
        {"not_comparable": {"lost": ["Avg health / 100"], "never_measured": []},
         "verdict_note": "Across 2 collections: 1 metric(s) improving, 0 worsening."})
    assert "NOT COMPARABLE" in silent["verdict_note"], silent["verdict_note"]
    assert "Avg health / 100" in silent["verdict_note"], silent["verdict_note"]

    # NON-VACUITY: an empty-but-PRESENT record is left exactly as the engine wrote it.
    ok = engine._carry_not_comparable({"not_comparable": {"lost": [], "never_measured": []},
                                       "verdict_note": "Across 2 collections: all comparable."})
    assert ok["verdict_note"] == "Across 2 collections: all comparable."
    assert ok["not_comparable"]["disclosure_available"] is True
