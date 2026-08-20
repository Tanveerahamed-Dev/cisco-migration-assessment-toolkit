"""Truthful single-snapshot Protocol Assurance receipt surfaces."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from cisco_toolkit.html import write_html_explorer
from cisco_toolkit.protocol_assurance import (
    bind_snapshot_json_bytes,
    bound_snapshot_source,
)
from cisco_toolkit.protocol_receipt_surfaces import (
    SHEET_NAME,
    write_protocol_assurance_receipt_sheet,
)
from webapp.backend.protocol_portfolio import (
    PERSISTED_SOURCE,
    build_protocol_single_snapshot_bundle,
)


def _bound_bundle(*, subject_cap: int = 1) -> dict:
    snapshot = json.loads(
        (Path(__file__).parent / "golden" / "snapshot.json").read_text(encoding="utf-8")
    )
    raw = json.dumps(
        snapshot,
        sort_keys=True,
        ensure_ascii=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    bound = bind_snapshot_json_bytes(raw)
    marker = bound_snapshot_source(bound)
    binding = {
        "source": PERSISTED_SOURCE,
        "sha256": marker["sha256"],
        "bytes": marker["bytes"],
        "snapshot_id": 41,
        "campaign_id": 7,
        "engagement_id": "surface-test",
        "label": "receipt fixture",
        "script_version": snapshot["script_version"],
    }
    return build_protocol_single_snapshot_bundle(
        bound,
        binding,
        subject_cap=subject_cap,
    )


def _workbook_text(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return "\n".join(
            str(cell.value)
            for row in workbook[SHEET_NAME].iter_rows()
            for cell in row
            if cell.value is not None
        )
    finally:
        workbook.close()


def _embedded_surface(path: Path) -> dict:
    html = path.read_text(encoding="utf-8")
    encoded = html.split("const EMBEDDED_PROTOCOL_ASSURANCE=", 1)[1].split(
        ";\nconst EMBEDDED_SNAPSHOT=", 1
    )[0]
    return json.loads(encoded)


def test_assessment_workbook_projects_validated_owner_receipt_and_complete_export(tmp_path):
    bundle = _bound_bundle(subject_cap=1)
    expected = bundle["receipt"]
    workbook = Workbook()
    workbook.remove(workbook.active)

    surface = write_protocol_assurance_receipt_sheet(workbook, bundle)
    output = tmp_path / "assessment.xlsx"
    workbook.save(output)
    text = _workbook_text(output)

    totals = {
        field: sum(family["subjects"][field] for family in expected["families"])
        for field in ("rendered", "total", "omitted")
    }
    assert surface["receipt_valid"] is True and surface["status"] == "BOUND"
    assert surface["subject_cap"] == totals
    assert "protocol_single_snapshot_receipt/1" in text
    assert expected["source_binding"]["sha256"] in text
    assert expected["receipt_sha256"] in text
    assert expected["complete_export"]["sha256"] in text
    assert (
        f"Rendered / total / omitted: {totals['rendered']} / "
        f"{totals['total']} / {totals['omitted']}"
    ) in text
    assert "protocol_single_snapshot_export/1 · uncapped JSON" in text
    assert "/api/snapshots/41/protocol-assurance/export" in text


def test_direct_workbook_writer_is_not_verified_and_does_not_trust_snapshot_claims(tmp_path):
    fake_digest = "sha256:" + "f" * 64
    hostile = {
        "receipt": {
            "schema": "protocol_single_snapshot_receipt/1",
            "receipt_sha256": fake_digest,
        },
        "complete_export": {"schema": "protocol_single_snapshot_export/1"},
    }
    workbook = Workbook()
    workbook.remove(workbook.active)

    surface = write_protocol_assurance_receipt_sheet(workbook, hostile)
    output = tmp_path / "legacy-assessment.xlsx"
    workbook.save(output)
    text = _workbook_text(output)

    assert surface["receipt_valid"] is False
    assert surface["status"] == "NOT VERIFIED"
    assert "This renderer did not hash the parsed snapshot or mint source custody" in text
    assert "protocol_single_snapshot_export/1 · uncapped JSON · NOT ATTACHED" in text
    assert "Rendered / total / omitted: 0 / 0 / 0" in text
    assert fake_digest not in text


def test_explorer_sidecar_is_bound_when_supplied_and_clearly_abstains_when_detached(tmp_path):
    bundle = _bound_bundle(subject_cap=1)
    bound_output = tmp_path / "bound-explorer.html"
    write_html_explorer(
        str(bound_output),
        {"devices": {"sw1": {}}, "interfaces": {"sw1": {}}},
        "bound fixture",
        protocol_assurance_bundle=bundle,
    )
    bound = _embedded_surface(bound_output)
    assert bound["receipt_valid"] is True and bound["status"] == "BOUND"
    assert bound["receipt_sha256"] == bundle["receipt"]["receipt_sha256"]
    assert bound["complete_export"]["schema"] == "protocol_single_snapshot_export/1"
    assert bound["subject_cap"]["omitted"] > 0

    fake_digest = "sha256:" + "e" * 64
    detached_output = tmp_path / "detached-explorer.html"
    write_html_explorer(
        str(detached_output),
        {
            "devices": {"sw1": {}},
            "interfaces": {"sw1": {}},
            "protocol_single_snapshot_receipt": {
                "schema": "protocol_single_snapshot_receipt/1",
                "receipt_sha256": fake_digest,
            },
        },
        "detached fixture",
    )
    detached = _embedded_surface(detached_output)
    assert detached["receipt_valid"] is False
    assert detached["status"] == "NOT VERIFIED"
    assert detached["receipt_sha256"] == ""
    assert detached["subject_cap"] == {"rendered": 0, "total": 0, "omitted": 0}
    assert detached["complete_export"]["schema"] == "protocol_single_snapshot_export/1"
    assert fake_digest not in json.dumps(detached, sort_keys=True)
