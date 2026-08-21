"""Offline --compare acquisition for one source-bound observed local L2 trial."""

from __future__ import annotations

import dataclasses
import json
import os
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

import COLLECT_PARSE_V3_23_0 as cp
from tests.test_l2_failure_rehearsal import _ether_snapshot
from tests.test_observed_l2_failure_evidence import _clean_l2_phase, _witness


ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "COLLECT_PARSE_V3_23_0.py"


def _encoded(value: dict) -> bytes:
    return json.dumps(
        dict(value),
        sort_keys=True,
        ensure_ascii=True,
        separators=(",", ":"),
        allow_nan=False,
        default=lambda item: dataclasses.asdict(item)
        if dataclasses.is_dataclass(item) else str(item),
    ).encode("utf-8")


def _trial_files(tmp_path: Path, *, reordered_mtime: bool = False) -> dict[str, Path]:
    pre_l2 = _ether_snapshot(
        tmp_path / "raw-pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    post_l2 = _ether_snapshot(
        tmp_path / "raw-post", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
    )
    recovery_l2 = _ether_snapshot(
        tmp_path / "raw-recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    values = {
        "before": _clean_l2_phase(recovery_l2, "2026-08-19T23:59:00+00:00"),
        "pre": _clean_l2_phase(pre_l2, "2026-08-20T00:00:00+00:00"),
        "post": _clean_l2_phase(post_l2, "2026-08-20T00:02:00+00:00"),
        "recovery": _clean_l2_phase(recovery_l2, "2026-08-20T00:03:00+00:00"),
    }
    paths = {name: tmp_path / f"{name}.json" for name in values}
    for name, value in values.items():
        paths[name].write_bytes(_encoded(value))
    paths["witness"] = tmp_path / "witness.json"
    paths["witness"].write_bytes(_witness(
        "etherchannel", "dist1|Po10",
        "single_observed_forwarding_member_loss", "shut_link",
        {"host": "dist1", "interface": "Gi1/0/2"},
    ))
    paths["context"] = tmp_path / "comparison-context.json"
    paths["context"].write_text(json.dumps({
        "schema": "offline_comparison_context/1",
        "engagement_id": "ENG-OFFLINE-L2",
        "campaign_id": 41,
        "before_snapshot_id": 100,
        "after_snapshot_id": 103,
        "change_intent": {"expected_changes": [], "note": "bounded local trial"},
    }), encoding="utf-8")

    custody_base = int(datetime(
        2026, 8, 20, 1, 0, 0, tzinfo=timezone.utc
    ).timestamp())
    mtimes = {
        "before": custody_base - 60,
        "pre": custody_base + (120 if reordered_mtime else 0),
        "post": custody_base + 60,
        "recovery": custody_base + 180,
    }
    for name, mtime in mtimes.items():
        os.utime(paths[name], (mtime, mtime))
    return paths


def _command(paths: dict[str, Path], output: Path) -> list[str]:
    return [
        sys.executable, str(SCRIPT),
        "--compare", str(paths["before"]), str(paths["recovery"]),
        "--comparison-context", str(paths["context"]),
        "--l2-trial-pre", str(paths["pre"]),
        "--l2-trial-post", str(paths["post"]),
        "--l2-trial-witness", str(paths["witness"]),
        "--output", str(output),
    ]


def test_cli_observed_l2_trial_reconciles_json_workbook_and_terminal(tmp_path):
    paths = _trial_files(tmp_path)
    output = tmp_path / "observed-local-trial.xlsx"

    completed = subprocess.run(
        _command(paths, output), cwd=tmp_path, text=True,
        capture_output=True, timeout=60,
    )

    assert completed.returncode == 0, completed.stdout + completed.stderr
    receipt = json.loads(output.with_suffix(".comparison.json").read_text("utf-8"))
    gate = receipt["cutover_gate"]
    assert gate["l2_observed_trial_status"] == "observed_survival"
    assert gate["l2_observed_trial_assurance"] == "local_safety_preservation"
    rehearsal = receipt["operator_evidence"]["rehearsal"]
    observed = rehearsal["observed_l2_failure_evidence"]
    assert rehearsal["assurance_level"] == "local_safety_preservation"
    assert observed["claims"] == {
        "local_scenario": "observed_survival",
        "service_path_survival": "not_verified",
        "traffic_continuity": "not_verified",
        "convergence": "not_verified",
    }
    assert observed["source_binding"]["recovery"]["sha256"] \
        == receipt["provenance"]["source_binding"]["after"]["sha256"]
    assert all(
        row["source"] == cp.OFFLINE_FILE_SOURCE
        and row["source_id"].startswith("file-sha256:")
        for row in (
            observed["source_binding"]["pre_failure"],
            observed["source_binding"]["post_failure"],
            observed["source_binding"]["recovery"],
        )
    )

    workbook = load_workbook(output, read_only=True)
    summary = workbook["Summary"]
    summary_rows = {
        summary.cell(row, 1).value: summary.cell(row, 3).value
        for row in range(2, summary.max_row + 1)
    }
    assert summary_rows["L2 OBSERVED LOCAL TRIAL"] == "LOCAL_SAFETY_PRESERVATION"
    observed_sheet = workbook["Observed L2 Trial"]
    observed_rows = {
        observed_sheet.cell(row, 1).value: observed_sheet.cell(row, 2).value
        for row in range(2, observed_sheet.max_row + 1)
    }
    workbook.close()
    assert observed_rows["Receipt status"] == "observed_survival"
    assert observed_rows["Assurance"] == "local_safety_preservation"
    assert observed_rows["Service-path survival"] == "not_verified"
    terminal = completed.stdout + completed.stderr
    assert "[OBSERVED L2 LOCAL TRIAL: OBSERVED_SURVIVAL]" in terminal
    assert "does not prove service/path survival or convergence" in terminal


def test_cli_reordered_offline_custody_abstains_in_canonical_receipt(tmp_path):
    paths = _trial_files(tmp_path, reordered_mtime=True)
    output = tmp_path / "reordered-local-trial.xlsx"

    completed = subprocess.run(
        _command(paths, output), cwd=tmp_path, text=True,
        capture_output=True, timeout=60,
    )

    assert completed.returncode == 0, completed.stdout + completed.stderr
    receipt = json.loads(output.with_suffix(".comparison.json").read_text("utf-8"))
    assert receipt["cutover_gate"]["l2_observed_trial_status"] == "not_verified"
    assert receipt["cutover_gate"]["verdict"] != "PASS"
    observed = receipt["operator_evidence"]["rehearsal"][
        "observed_l2_failure_evidence"
    ]
    assert observed["status"] == "not_verified"
    assert any("custody timestamps are not strictly" in item for item in observed["failures"])


@pytest.mark.parametrize("supplied", [
    ("--l2-trial-pre", "pre.json"),
    ("--l2-trial-post", "post.json"),
    ("--l2-trial-witness", "witness.json"),
])
def test_cli_observed_trial_inputs_are_all_or_none(tmp_path, supplied):
    output = tmp_path / "partial.xlsx"
    completed = subprocess.run(
        [
            sys.executable, str(SCRIPT), "--compare", "before.json", "after.json",
            *supplied, "--output", str(output),
        ],
        cwd=tmp_path, text=True, capture_output=True, timeout=30,
    )

    assert completed.returncode == 2
    assert "must be supplied together" in completed.stderr
    assert not output.exists()


def test_cli_observed_trial_requires_explicit_comparison_context(tmp_path):
    output = tmp_path / "unbound.xlsx"
    completed = subprocess.run(
        [
            sys.executable, str(SCRIPT), "--compare", "before.json", "after.json",
            "--l2-trial-pre", "pre.json", "--l2-trial-post", "post.json",
            "--l2-trial-witness", "witness.json", "--output", str(output),
        ],
        cwd=tmp_path, text=True, capture_output=True, timeout=30,
    )

    assert completed.returncode == 2
    assert "observed L2 trial inputs require --comparison-context" in completed.stderr
    assert not output.exists()


def test_cli_observed_witness_is_bounded_before_composition(tmp_path):
    paths = _trial_files(tmp_path)
    paths["witness"].write_bytes(b"x" * (64 * 1024 + 1))
    output = tmp_path / "oversized.xlsx"

    completed = subprocess.run(
        _command(paths, output), cwd=tmp_path, text=True,
        capture_output=True, timeout=30,
    )

    assert completed.returncode == 2
    assert "65536-byte input limit" in completed.stderr
    assert not output.exists()


def test_trial_mtime_binding_detects_touch_without_byte_change(tmp_path):
    path = tmp_path / "phase.json"
    path.write_bytes(b'{"collected_at":"2026-08-20T00:00:00+00:00"}')
    _data, _record, binding = cp._bind_input(
        str(path), role="offline_l2_trial_pre_failure", capture_mtime=True
    )
    touched = binding["mtime_ns"] + 1_000_000_000
    os.utime(path, ns=(touched, touched))

    with pytest.raises(ValueError, match="custody timestamp changed after parsing"):
        cp._require_current_bindings([binding], "observed trial")
