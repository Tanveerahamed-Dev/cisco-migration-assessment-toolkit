"""Executable CLI coverage for source-bound adjacent campaign comparison receipts."""

import hashlib
import json
import os
import subprocess
import sys

from openpyxl import load_workbook


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py")


def _snapshot(when: str, score: int) -> dict:
    return {
        "schema": "collect_parse_snapshot/1",
        "script_version": "V3.23.0",
        "generated_at": when,
        "devices": {"R1": {}},
        "interfaces": {"R1": {}},
        "health_scores": [{"switch": "R1", "score": score, "band": "Good"}],
        "punchlist": [],
        "migration_readiness": [],
        "validation_plan": {
            "items": [],
            "by_wave": {},
            "summary": {"n_items": 0, "n_waves": 0, "by_category": {}, "n_high": 0},
            "banner": "No validation evidence was supplied.",
        },
    }


def _write_snapshot(path, payload: dict) -> str:
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    path.write_bytes(raw)
    return "sha256:" + hashlib.sha256(raw).hexdigest()


def _run(*args, cwd):
    return subprocess.run(
        [sys.executable, SCRIPT, *map(str, args)],
        cwd=str(cwd),
        capture_output=True,
        text=True,
        timeout=300,
    )


def test_trend_cli_rejects_duplicate_snapshot_keys_before_artifacts(tmp_path):
    before = tmp_path / "before.snapshot.json"
    after = tmp_path / "after.snapshot.json"
    before.write_bytes(
        b'{"script_version":"V3.23.0","devices":{"R1":{}},"devices":{"R2":{}}}'
    )
    _write_snapshot(after, _snapshot("2026-08-20T00:05:00", 90))
    output = tmp_path / "trend.xlsx"

    process = _run("--trend", before, after, "--output", output, cwd=tmp_path)
    terminal = process.stdout + process.stderr

    assert process.returncode != 0, terminal
    assert "duplicate JSON object key 'devices'" in terminal
    assert not output.exists()
    assert not (tmp_path / "trend.trend-comparisons.json").exists()


def test_trend_cli_without_identity_context_exports_not_verified_receipt_set(tmp_path):
    before = tmp_path / "before.snapshot.json"
    after = tmp_path / "after.snapshot.json"
    _write_snapshot(before, _snapshot("2026-08-20T00:00:00", 88))
    _write_snapshot(after, _snapshot("2026-08-20T00:05:00", 90))
    output = tmp_path / "trend.xlsx"

    process = _run("--trend", before, after, "--output", output, cwd=tmp_path)
    terminal = process.stdout + process.stderr

    assert process.returncode == 0, terminal
    exported = json.loads(
        (tmp_path / "trend.trend-comparisons.json").read_text(encoding="utf-8")
    )
    assert exported["status"] == "not_verified"
    assert exported["assurance_level"] == "not_verified"
    assert exported["complete"] is False
    assert exported["n_pairs_total"] == 1
    assert exported["n_pairs_returned"] == 0
    assert exported["comparisons"] == []
    assert all(row["snapshot_id"] is None for row in exported["source_receipts"])

    workbook = load_workbook(output, read_only=True)
    assert workbook["Campaign Summary"].cell(4, 2).value == "NOT_VERIFIED"
    sheet = workbook["Adjacent Comparison Receipts"]
    assert sheet.cell(2, 6).value == "NOT_VERIFIED"
    assert sheet.cell(2, 9).value is False
    assert sheet.cell(2, 10).value == 0
    assert sheet.cell(2, 11).value == 0
    assert sheet.cell(2, 12).value == 0


def test_trend_context_cannot_assign_a_snapshot_id_to_different_source_bytes(tmp_path):
    before = tmp_path / "before.snapshot.json"
    after = tmp_path / "after.snapshot.json"
    before_sha = _write_snapshot(before, _snapshot("2026-08-20T00:00:00", 88))
    after_sha = _write_snapshot(after, _snapshot("2026-08-20T00:05:00", 90))
    context = tmp_path / "trend-context.json"
    context.write_text(json.dumps({
        "schema": "offline_trend_context/1",
        "engagement_id": "ENG-SWAPPED-SOURCE",
        "campaign_id": 78,
        "snapshots": [
            {"snapshot_id": 1001, "source_sha256": after_sha},
            {"snapshot_id": 1002, "source_sha256": before_sha},
        ],
    }), encoding="utf-8")
    output = tmp_path / "trend.xlsx"

    process = _run(
        "--trend", before, after, "--trend-context", context, "--output", output,
        cwd=tmp_path,
    )
    terminal = process.stdout + process.stderr

    assert process.returncode != 0, terminal
    assert "source_sha256 does not match the exact bytes" in terminal
    assert not output.exists()
    assert not (tmp_path / "trend.trend-comparisons.json").exists()


def test_trend_adjacent_receipt_is_exact_pair_compare_output(tmp_path):
    before = tmp_path / "before.snapshot.json"
    after = tmp_path / "after.snapshot.json"
    before_sha = _write_snapshot(before, _snapshot("2026-08-20T00:00:00", 88))
    after_sha = _write_snapshot(after, _snapshot("2026-08-20T00:05:00", 90))
    intent = {"expected_changes": [], "note": "same decision intent"}

    pair_context = tmp_path / "pair-context.json"
    pair_context.write_text(json.dumps({
        "schema": "offline_comparison_context/1",
        "engagement_id": "ENG-TREND-PARITY",
        "campaign_id": 77,
        "before_snapshot_id": 1001,
        "after_snapshot_id": 1002,
        "change_intent": intent,
    }), encoding="utf-8")
    pair_output = tmp_path / "pair.xlsx"
    pair_process = _run(
        "--compare", before, after,
        "--comparison-context", pair_context,
        "--output", pair_output,
        cwd=tmp_path,
    )
    assert pair_process.returncode == 0, pair_process.stdout + pair_process.stderr
    pair_receipt = json.loads(
        (tmp_path / "pair.comparison.json").read_text(encoding="utf-8")
    )

    trend_context = tmp_path / "trend-context.json"
    trend_context.write_text(json.dumps({
        "schema": "offline_trend_context/1",
        "engagement_id": "ENG-TREND-PARITY",
        "campaign_id": 77,
        "snapshots": [
            {"snapshot_id": 1001, "source_sha256": before_sha, "label": before.name},
            {"snapshot_id": 1002, "source_sha256": after_sha, "label": after.name},
        ],
        "adjacent_change_intents": [intent],
    }), encoding="utf-8")
    trend_output = tmp_path / "trend.xlsx"
    trend_process = _run(
        "--trend", before, after,
        "--trend-context", trend_context,
        "--output", trend_output,
        cwd=tmp_path,
    )
    assert trend_process.returncode == 0, trend_process.stdout + trend_process.stderr

    exported = json.loads(
        (tmp_path / "trend.trend-comparisons.json").read_text(encoding="utf-8")
    )
    assert exported["status"] == "verified"
    assert exported["complete"] is True
    assert exported["n_pairs_total"] == exported["n_pairs_returned"] == 1
    assert exported["portable_export"] == {
        "rendered": 1, "total": 1, "omitted": 0, "complete": True,
    }
    pair = exported["comparisons"][0]
    assert pair["before_sha256"] == before_sha
    assert pair["after_sha256"] == after_sha
    assert pair["comparison"] == pair_receipt
    assert pair["canonical_gate"] == pair_receipt["cutover_gate"]["verdict"]
    assert pair["comparison_receipt_sha256"] == pair_receipt["comparison_receipt"]["receipt_sha256"]

    workbook = load_workbook(trend_output, read_only=True)
    assert workbook["Campaign Summary"].cell(4, 2).value == "VERIFIED"
    sheet = workbook["Adjacent Comparison Receipts"]
    assert sheet.cell(2, 9).value is True
    assert sheet.cell(2, 10).value == 1
    assert sheet.cell(2, 11).value == 1
    assert sheet.cell(2, 12).value == 0
    assert sheet.cell(3, 2).value == 1001
    assert sheet.cell(3, 3).value == 1002
    assert sheet.cell(3, 4).value == before_sha
    assert sheet.cell(3, 5).value == after_sha
    assert sheet.cell(3, 6).value == pair_receipt["cutover_gate"]["verdict"]


def test_campaign_workbook_discloses_receipt_render_cap_and_complete_sink(tmp_path):
    from cisco_toolkit.html import write_campaign_workbook

    rows = [{
        "from": f"C{index + 1}",
        "to": f"C{index + 2}",
        "before_snapshot_id": index + 1,
        "after_snapshot_id": index + 2,
        "before_sha256": "sha256:" + ("a" * 64),
        "after_sha256": "sha256:" + ("b" * 64),
        "canonical_gate": "INDETERMINATE",
        "admission_status": "not_comparable",
        "comparison_receipt_sha256": "sha256:" + ("c" * 64),
        "comparison": {},
    } for index in range(1002)]
    receipt_set = {
        "schema": "campaign_adjacent_comparison_set/1",
        "status": "not_verified",
        "n_pairs_total": 1002,
        "n_pairs_returned": 1002,
        "complete": False,
        "note": "Synthetic projection-only cap fixture.",
        "comparisons": rows,
    }
    output = tmp_path / "capped-trend.xlsx"

    write_campaign_workbook(
        [_snapshot("2026-08-20T00:00:00", 88), _snapshot("2026-08-20T00:05:00", 90)],
        str(output),
        adjacent_comparisons=receipt_set,
    )

    assert len(receipt_set["comparisons"]) == 1002, "UI rendering must not mutate the complete set"
    sheet = load_workbook(output, read_only=True)["Adjacent Comparison Receipts"]
    assert sheet.cell(2, 10).value == 1000
    assert sheet.cell(2, 11).value == 1002
    assert sheet.cell(2, 12).value == 2
    assert "Complete uncapped export:" in sheet.cell(2, 13).value
    assert sheet.max_row == 1002  # header + status + 1000 rendered pair rows
