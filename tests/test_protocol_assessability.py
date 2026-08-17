"""Runtime protocol assessability: exact denominator, abstentions, and projections."""

from __future__ import annotations

import json

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook  # noqa: E402

from cisco_toolkit.analyze import (  # noqa: E402
    PROTOCOL_ASSESSABILITY_FAMILIES,
    PROTOCOL_ASSESSABILITY_STATES,
    compute_protocol_assessability,
    compute_protocol_health,
)
from cisco_toolkit.excel import (  # noqa: E402
    write_collection_completeness_sheet,
    write_protocol_health_sheet,
)


EXPECTED_FAMILIES = ("STP", "EtherChannel", "VTP", "OSPF", "BGP", "EIGRP", "FHRP")


def _capture(tmp_path, name: str, body: str) -> str:
    path = tmp_path / name
    path.write_text(body, encoding="utf-8")
    return str(path)


def _receipt(tmp_path) -> dict:
    usable = _capture(tmp_path, "usable.txt", "bounded protocol state\n")
    empty = _capture(tmp_path, "empty.txt", "\n")
    error = _capture(tmp_path, "error.txt", "% Invalid input detected at '^' marker.\n")
    command_files = {
        "sw1": {
            "show spanning-tree": usable,
            "show spanning-tree blockedports": empty,
            "show spanning-tree detail": error,
            "show etherchannel summary": usable,
            "show vtp status": empty,
            "show ip ospf neighbor": error,
            "show ip eigrp neighbors": usable,
            "show standby brief": usable,
            "show vrrp brief": empty,
            "show glbp brief": empty,
        }
    }
    health = [
        {"switch": "sw1", "protocol": "STP"},
        {"switch": "sw1", "protocol": "EIGRP"},
        {"switch": "sw1", "protocol": "FHRP"},
    ]
    return compute_protocol_assessability(
        ("sw2", "sw1"), {"sw1": {}}, command_files, health
    )


def test_protocol_assessability_is_exact_deterministic_and_capture_honest(tmp_path):
    receipt = _receipt(tmp_path)

    assert receipt["schema"] == "protocol_assessability/1"
    assert tuple(family["protocol"] for family in PROTOCOL_ASSESSABILITY_FAMILIES) == EXPECTED_FAMILIES
    assert tuple(family["protocol"] for family in receipt["families"]) == EXPECTED_FAMILIES
    assert {
        item["id"]: item["required"]
        for family in receipt["families"] if family["protocol"] == "STP"
        for item in family["inputs"]
    } == {
        "state": True,
        "blocked_ports": True,
        "inconsistent_ports": True,
        "topology_changes": False,
    }
    assert tuple(PROTOCOL_ASSESSABILITY_STATES) == (
        "assessed", "partial", "captured_no_record", "captured_empty",
        "capture_error", "not_collected", "analysis_unavailable",
    )
    assert [(row["switch"], row["protocol"]) for row in receipt["rows"]] == [
        (host, protocol) for host in ("sw1", "sw2") for protocol in EXPECTED_FAMILIES
    ]

    sw1 = {row["protocol"]: row for row in receipt["rows"] if row["switch"] == "sw1"}
    assert sw1["STP"]["state"] == "partial"
    assert sw1["STP"]["input_states"] == {
        "state": "usable",
        "blocked_ports": "empty",
        "inconsistent_ports": "missing",
        "topology_changes": "error",
    }
    assert sw1["EtherChannel"]["state"] == "captured_no_record"
    assert sw1["VTP"]["state"] == "captured_empty"
    assert sw1["OSPF"]["state"] == "capture_error"
    assert sw1["BGP"]["state"] == "not_collected"
    assert sw1["EIGRP"]["state"] == "assessed"
    assert sw1["FHRP"]["state"] == "partial"
    assert sw1["FHRP"]["input_states"] == {
        "hsrp_groups": "usable", "vrrp_groups": "empty", "glbp_groups": "empty"
    }
    assert all(row["state"] == "not_collected" for row in receipt["rows"] if row["switch"] == "sw2")

    assert receipt["summary"] == {
        "n_devices": 2,
        "n_families": 7,
        "n_cells": 14,
        "n_health_rows": 3,
        "n_complete_devices": 0,
        "by_state": {
            "assessed": 1,
            "partial": 2,
            "captured_no_record": 1,
            "captured_empty": 1,
            "capture_error": 1,
            "not_collected": 8,
            "analysis_unavailable": 0,
        },
    }
    # The receipt is safe to publish: neither raw bodies nor source paths cross the boundary.
    rendered = json.dumps(receipt, sort_keys=True)
    assert str(tmp_path) not in rendered
    assert "% Invalid input" not in rendered
    assert receipt == _receipt(tmp_path)


def test_protocol_assessability_unions_all_host_sources_and_fails_closed(tmp_path):
    usable = _capture(tmp_path, "bgp-good.txt", "Neighbor V AS MsgRcvd MsgSent State/PfxRcd\n")
    error = _capture(tmp_path, "bgp-error.txt", "% Command not found\n")
    receipt = compute_protocol_assessability(
        ["inventory-only"],
        {"interfaces-only": {}},
        {"captures-only": {"show ip bgp summary": error, "show bgp summary": usable,
                           "show standby brief": usable}},
        [{"switch": "health-only", "protocol": "BGP"}],
    )
    assert {row["switch"] for row in receipt["rows"]} == {
        "inventory-only", "interfaces-only", "captures-only", "health-only"
    }
    assert receipt["summary"]["n_cells"] == 4 * 7
    by_key = {(row["switch"], row["protocol"]): row for row in receipt["rows"]}
    assert by_key[("captures-only", "BGP")]["capture_state"] == "usable"
    assert by_key[("captures-only", "BGP")]["state"] == "captured_no_record"
    assert by_key[("captures-only", "FHRP")]["state"] == "partial"
    assert by_key[("captures-only", "FHRP")]["input_states"] == {
        "hsrp_groups": "usable", "vrrp_groups": "missing", "glbp_groups": "missing"
    }
    # A sparse row that cannot be reconciled to current-run evidence is never called assessed.
    assert by_key[("health-only", "BGP")]["state"] == "partial"

    unavailable = compute_protocol_assessability(
        ["inventory-only"], {}, {}, [], analysis_available=False
    )
    assert unavailable["summary"]["by_state"]["analysis_unavailable"] == 7
    assert {row["state"] for row in unavailable["rows"]} == {"analysis_unavailable"}


def test_protocol_assessability_never_lets_empty_secondary_inputs_complete_or_mask_error(tmp_path):
    usable = _capture(tmp_path, "stp-usable.txt", "Spanning tree enabled protocol rstp\n")
    empty = _capture(tmp_path, "stp-empty.txt", "\n")
    error = _capture(tmp_path, "stp-error.txt", "% Invalid input detected at '^' marker.\n")

    partial = compute_protocol_assessability(
        ["sw-partial"], {"sw-partial": {}}, {"sw-partial": {
            "show spanning-tree": usable,
            "show spanning-tree blockedports": empty,
            "show spanning-tree inconsistentports": empty,
            "show spanning-tree detail": empty,
        }}, [{"switch": "sw-partial", "protocol": "STP"}],
    )
    stp_partial = next(row for row in partial["rows"] if row["protocol"] == "STP")
    assert stp_partial["capture_state"] == "usable"
    assert stp_partial["state"] == "partial"
    assert "blocked_ports, inconsistent_ports" in stp_partial["reason"]
    assert "topology_changes" not in stp_partial["reason"]

    failed_primary = compute_protocol_assessability(
        ["sw-error"], {"sw-error": {}}, {"sw-error": {
            "show spanning-tree": error,
            "show spanning-tree blockedports": empty,
        }}, [],
    )
    stp_error = next(row for row in failed_primary["rows"] if row["protocol"] == "STP")
    assert stp_error["input_states"]["state"] == "error"
    assert stp_error["input_states"]["blocked_ports"] == "empty"
    assert stp_error["capture_state"] == "error"
    assert stp_error["state"] == "capture_error"


def test_stp_health_requires_parsed_primary_evidence_and_optional_detail_does_not_block_assessment(tmp_path):
    unknown = _capture(
        tmp_path,
        "stp-unknown.txt",
        "VLAN0010\nbounded but unrecognized command output\n",
    )
    supplemental = _capture(tmp_path, "stp-supplemental.txt", "No affected ports observed\n")
    unknown_commands = {"sw-unknown": {
        "show spanning-tree": unknown,
        "show spanning-tree blockedports": supplemental,
        "show spanning-tree inconsistentports": supplemental,
    }}
    unknown_health = compute_protocol_health({"sw-unknown": {}}, unknown_commands)
    assert not [row for row in unknown_health if row["protocol"] == "STP"]

    unknown_receipt = compute_protocol_assessability(
        ["sw-unknown"], {"sw-unknown": {}}, unknown_commands, unknown_health
    )
    unknown_stp = next(row for row in unknown_receipt["rows"] if row["protocol"] == "STP")
    assert unknown_stp["health_row_emitted"] is False
    assert unknown_stp["state"] == "captured_no_record"

    nxos_stp = _capture(tmp_path, "stp-nxos.txt", """\
VLAN0010
  Spanning tree enabled protocol rstp
  Root ID    Priority    24586
             Address     aaaa.0001.0001
             This bridge is the root
  Bridge ID  Priority    24586
             Address     aaaa.0001.0001
Interface        Role Sts Cost      Prio.Nbr Type
Eth1/1           Desg FWD 4         128.1    P2p
""")
    nxos_commands = {"sw-nxos": {
        "show spanning-tree": nxos_stp,
        "show spanning-tree blockedports": supplemental,
        "show spanning-tree inconsistentports": supplemental,
    }}
    nxos_health = compute_protocol_health({"sw-nxos": {}}, nxos_commands)
    assert [row for row in nxos_health if row["protocol"] == "STP"]

    nxos_receipt = compute_protocol_assessability(
        ["sw-nxos"], {"sw-nxos": {}}, nxos_commands, nxos_health
    )
    nxos_stp_row = next(row for row in nxos_receipt["rows"] if row["protocol"] == "STP")
    assert nxos_stp_row["input_states"] == {
        "state": "usable",
        "blocked_ports": "usable",
        "inconsistent_ports": "usable",
        "topology_changes": "missing",
    }
    assert nxos_stp_row["health_row_emitted"] is True
    assert nxos_stp_row["state"] == "assessed"


def test_protocol_assessability_projects_explicitly_into_workbook(tmp_path):
    receipt = _receipt(tmp_path)
    wb = Workbook()
    write_protocol_health_sheet(wb, [])
    write_collection_completeness_sheet(
        wb,
        {"summary": {"inventory": 2, "complete": 1, "partial": 1, "not_collected": 0},
         "devices": []},
        {"summary": {"parsers_called": 0}, "events": []},
        protocol_assessability=receipt,
    )
    protocol_health_text = "\n".join(
        str(cell.value or "") for row in wb["Protocol Health"].iter_rows() for cell in row
    )
    completeness_text = "\n".join(
        str(cell.value or "") for row in wb["Collection Completeness"].iter_rows() for cell in row
    )
    assert "absence is not healthy" in protocol_health_text
    assert "Protocol assessability — runtime family × device receipt" in completeness_text
    assert "3 of 14 host-family cells" in completeness_text
    assert "PARTIAL" in completeness_text and "NOT COLLECTED" in completeness_text
    assert ("All inventory devices satisfy the baseline essential command groups — no baseline "
            "collection blind spots." in completeness_text)
    assert ("Protocol-specific evidence gaps may remain; review the Protocol assessability receipt "
            "below." in completeness_text)
    assert "All inventory devices fully collected — no blind spots." not in completeness_text

    wb2 = Workbook()
    write_protocol_health_sheet(wb2, [{
        "switch": "sw1", "protocol": "EIGRP", "summary": "1 neighbor(s) up",
        "detail": "Gi1/0/1", "severity": "Info",
    }])
    health_values = [cell.value for row in wb2["Protocol Health"].iter_rows() for cell in row]
    assert "OBSERVED · NO SUPPORTED ISSUE" in health_values
    assert "OK" not in health_values
