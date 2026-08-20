"""Bounded L2 rehearsal composition over the existing simulation/delta owners."""

from __future__ import annotations

import copy
import dataclasses
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook

from cisco_toolkit import traffic_assurance
from cisco_toolkit.analyze import (
    compute_etherchannel_projection,
    summarize_etherchannel_baseline,
)
from cisco_toolkit.capture_integrity import compute_capture_integrity_from_paths
from cisco_toolkit.etherchannel import (
    compute_etherchannel_operational_evidence,
    embedded_etherchannel_operational_evidence,
)
from cisco_toolkit.l2_rehearsal import (
    L2_FAILURE_REHEARSAL_SCHEMA,
    compute_l2_failure_rehearsal,
    validate_l2_failure_rehearsal,
)
from cisco_toolkit.comparison import (
    bound_comparison_decision_input_authority,
    compare_bound_pair,
)
from cisco_toolkit.html import compute_cutover_gate, write_diff_workbook
from cisco_toolkit.multichassis_lag import compute_multichassis_lag_domain_baseline
from cisco_toolkit.protocol_assurance import (
    OFFLINE_FILE_SOURCE,
    bind_snapshot_json_bytes,
    cutover_operator_evidence,
)
from tests.test_compare_cutover_gate_cli import _snapshot as _clean_comparison_snapshot
from tests.test_etherchannel_cutover_truth import _receipt
from tests.test_etherchannel_operational_evidence import (
    _bind as _bind_source_paths,
    _copy_paths as _copy_etherchannel_paths,
)
from tests.test_multichassis_lag import _nxos_observation, _nxos_pair
from tests.test_stp_consistency_truth import _stp_run
from tests.test_traffic_assurance import _intent, _symmetric_snapshot


def _bind(value: dict):
    raw = json.dumps(
        value,
        sort_keys=True,
        ensure_ascii=True,
        separators=(",", ":"),
        allow_nan=False,
        default=lambda item: dataclasses.asdict(item)
        if dataclasses.is_dataclass(item) else str(item),
    ).encode("utf-8")
    return bind_snapshot_json_bytes(raw)


def _family(value: dict, name: str) -> list[dict]:
    return [row for row in value["scenarios"] if row["family"] == name]


def _assert_no_decision_owner(value) -> None:
    if isinstance(value, dict):
        assert "score" not in value
        assert "verdict" not in value
        for child in value.values():
            _assert_no_decision_owner(child)
    elif isinstance(value, list):
        for child in value:
            _assert_no_decision_owner(child)


def _ether_snapshot(tmp_path, body: str, *, host: str = "dist1") -> dict:
    desired_row = body.strip().splitlines()[-1].replace("1 Po1", "10 Po10", 1)
    paths = _copy_etherchannel_paths(
        tmp_path, "ios", host,
        {"show etherchannel summary": (
            "10     Po10(SU)        LACP       Gi1/0/1(P) Gi1/0/2(P)",
            desired_row,
        )},
    )
    if "Gi1/0/2" not in desired_row:
        second_member = (
            "interface GigabitEthernet1/0/2\n"
            " description second uplink member\n"
            " channel-group 10 mode passive\n"
            "!\n"
        )
        for command in (
                "show running-config", "show running-config | section ^interface"):
            path = Path(paths[host][command])
            path.write_text(
                path.read_text(encoding="utf-8").replace(second_member, ""),
                encoding="utf-8",
            )
    _bind_source_paths(paths)
    devices = {host: {"platform": "ios"}}
    projection = compute_etherchannel_projection(devices, paths)
    receipt = _receipt({host: {"EtherChannel": "assessed"}})
    baseline = summarize_etherchannel_baseline(
        projection, receipt, devices=devices)
    operational = compute_etherchannel_operational_evidence(
        paths,
        compute_capture_integrity_from_paths(paths),
        projection,
        devices=devices,
    )
    return {
        "devices": devices,
        "etherchannel_projection": projection,
        "protocol_assessability": receipt,
        "etherchannel_baseline": baseline,
        "etherchannel_operational_evidence":
            embedded_etherchannel_operational_evidence(operational),
    }


def _multichassis_snapshot(
        tmp_path, *, degraded: bool = False, partner: bool = True,
        orphan: bool = False) -> dict:
    if orphan:
        orphan_output = (
            "Switch# show vpc orphan-ports\n"
            "Note:\n"
            "--------::Going through port database. Please be patient.::--------\n\n"
            "VLAN           Orphan Ports\n"
            "-------        -------------------------\n"
            "10             Eth1/45\n"
        )
        raw = {"observations": [
            _nxos_observation(
                "leaf-a", "leaf-b", source_digit="a",
                orphan_output=orphan_output,
                orphan_suspend_ports=("Eth1/45",),
            ),
            _nxos_observation("leaf-b", "leaf-a", source_digit="b"),
        ]}
    else:
        raw = _nxos_pair(second_status="down" if degraded else "up")
    if not partner:
        raw["observations"][1]["legs"][0]["lacp_partner_system_id"] = ""
    baseline = compute_multichassis_lag_domain_baseline(raw)
    paths = {}
    for host in ("leaf-a", "leaf-b"):
        paths.update(_copy_etherchannel_paths(
            tmp_path / host, "nxos", host, {}))
    _bind_source_paths(paths)
    devices = {
        "leaf-a": {"platform": "nxos"},
        "leaf-b": {"platform": "nxos"},
    }
    projection = compute_etherchannel_projection(devices, paths)
    receipt = _receipt({
        "leaf-a": {"EtherChannel": "assessed"},
        "leaf-b": {"EtherChannel": "assessed"},
    })
    legacy = summarize_etherchannel_baseline(
        projection, receipt, devices=devices)
    operational = compute_etherchannel_operational_evidence(
        paths,
        compute_capture_integrity_from_paths(paths),
        projection,
        devices=devices,
    )
    return {
        "devices": devices,
        "etherchannel_projection": projection,
        "protocol_assessability": receipt,
        "etherchannel_baseline": legacy,
        "etherchannel_operational_evidence":
            embedded_etherchannel_operational_evidence(operational),
        "multichassis_lag_typed_observations": {
            "observations": [dict(row) for row in raw["observations"]],
        },
        "multichassis_lag_domain_baseline": baseline,
    }


def _traffic_failure_snapshot() -> dict:
    snapshot = _symmetric_snapshot()
    intent = {**_intent(), "id": "erp-forward"}
    snapshot["traffic_assurance"] = traffic_assurance.assess_flows(snapshot, [{
        **intent,
        "failure": {"action": "fail_node", "id": "R2"},
    }])
    return snapshot


def _canonical_pair(snapshot: dict, *, engagement: str = "ENG-L2-REHEARSAL"):
    raw = json.dumps(
        snapshot,
        ensure_ascii=True,
        separators=(",", ":"),
        allow_nan=False,
        default=lambda item: dataclasses.asdict(item)
        if dataclasses.is_dataclass(item) else str(item),
    ).encode("utf-8")
    before = bind_snapshot_json_bytes(raw)
    after = bind_snapshot_json_bytes(raw)

    def binding(snapshot_id: int) -> dict:
        return {
            "source": OFFLINE_FILE_SOURCE,
            "sha256": "sha256:" + hashlib.sha256(raw).hexdigest(),
            "bytes": len(raw),
            "snapshot_id": snapshot_id,
            "campaign_id": 91,
            "engagement_id": engagement,
            "label": f"snapshot-{snapshot_id}.json",
            "script_version": "V3.23.0",
        }

    comparison = compare_bound_pair(
        before,
        after,
        before_binding=binding(9101),
        after_binding=binding(9102),
    )
    return before, after, comparison


def _canonical_multichassis_pair(
        tmp_path: Path, *, orphan: bool = False, degraded: bool = False):
    # Retain the clean comparison's route/current-baseline/certificate surfaces,
    # but use the two typed L2 subjects as the admitted device denominator.
    base = json.loads(json.dumps(
        _clean_comparison_snapshot("FULL/DR", "2026-08-20T00:00:00")
    ).replace("R1", "leaf-a").replace("R2", "leaf-b"))
    base.pop("routing_neighbors", None)
    l2 = _multichassis_snapshot(
        tmp_path,
        orphan=orphan,
        degraded=degraded,
    )
    for key in (
            "devices", "protocol_assessability", "etherchannel_projection",
            "etherchannel_baseline", "etherchannel_operational_evidence",
            "multichassis_lag_typed_observations",
            "multichassis_lag_domain_baseline"):
        base[key] = l2[key]
    return _canonical_pair(base)


def _forged_preserved_traffic_snapshot() -> dict:
    """Build an internally tidy but producer-impossible v1 preservation claim for negatives."""
    snapshot = _traffic_failure_snapshot()
    traffic_set = snapshot["traffic_assurance"]
    result = traffic_set["results"][0]
    baseline = traffic_assurance.assess_flow(snapshot, result["intent"])
    failure = result["failure"]
    simulation = failure["cutover_evidence"]
    step = simulation["steps"][0]
    step.update({
        "newly_lost_flows": [],
        "recovered_flows": [],
        "stp_reroots": [],
        "fhrp_takeovers": [],
        "split_brain_risks": [],
        "l2_continuity": {
            "applicable": False,
            "assessed": None,
            "verdict": "not_applicable",
            "election_projection_count": 0,
            "affected_member_count": 0,
            "affected_member_hosts": [],
            "reason": "no L2-affecting mutation or election projection was requested",
        },
        "n_stp_election_candidates": 0,
        "n_fhrp_election_candidates": 0,
        "indeterminate": 0,
        "valid": True,
        "validation_errors": [],
        "is_noop": False,
        "narrative": "Synthetic failure retained the declared service path.",
    })
    simulation.update({
        "valid": True,
        "validation_errors": [],
        "worst_step": None,
        "summary": {
            "n_steps": 1,
            "total_newly_lost": 0,
            "total_recovered": 0,
            "total_stp_reroots": 0,
            "total_fhrp_takeovers": 0,
            "total_stp_election_candidates": 0,
            "total_fhrp_election_candidates": 0,
            "total_split_brain_risks": 0,
            "total_indeterminate": 0,
            "n_noop_steps": 0,
            "n_invalid_steps": 0,
            "n_input_errors": 0,
        },
    })
    result["dimensions"]["path"].update({"symmetric": True, "asymmetry": []})
    failure.update({
        "status": "preserved",
        "verdict": "proven",
        "baseline_verdict": "proven",
        "post_verdict": "proven",
        "post_dimensions": copy.deepcopy(result["dimensions"]),
        "cutover_gate": {
            "verdict": "proven",
            "n_indeterminate": 0,
            "n_split_brain_risks": 0,
            "continuity_assessed": True,
            "continuity_gaps": [],
        },
    })
    result["verdict"] = "proven"
    result["verdict_reasons"] = baseline["verdict_reasons"]
    for claim in result["claims"]:
        if claim["predicate"] in {
            "selected_rib_forwarding_projection", "stateless_acl_assurance",
        }:
            claim["verdict"] = claim["baseline_verdict"]
            claim["applicability"] = "applicable"
        elif claim["predicate"] in {
            "declared_scope_assurance", "requested_failure_assurance",
        }:
            claim["verdict"] = "proven"
    traffic_set["summary"].update({
        "proven": 1,
        "refuted": 0,
        "not_observed": 0,
        "indeterminate": 0,
    })
    return snapshot


def test_projection_reuses_existing_owners_and_never_owns_a_decision(tmp_path):
    snapshot = _ether_snapshot(
        tmp_path,
        "Group Port-channel Protocol Ports\n"
        "1 Po1(SU) LACP Gi1/0/1(P) Gi1/0/2(P)\n",
    )
    snapshot["stp_roots"] = {
        "dist1": {"10": {
            "root_priority": 4096, "root_address": "aaaa.0000.0001",
            "is_root": True, "is_mst": False, "bridge_priority": 4096,
        }},
        "dist2": {"10": {
            "root_priority": 4096, "root_address": "aaaa.0000.0001",
            "is_root": False, "is_mst": False, "bridge_priority": 8192,
        }},
    }
    result = compute_l2_failure_rehearsal(_bind(snapshot))

    assert result["schema"] == L2_FAILURE_REHEARSAL_SCHEMA
    assert result["owns_score"] is result["owns_verdict"] is False
    assert result["assurance_level"] == "not_verified"
    ether = _family(result, "etherchannel")[0]
    assert ether["disposition"] == "simulation_only"
    assert ether["evidence"]["observed_forwarding_member_count"] == 2
    assert ether["evidence"]["remaining_forwarding_members_after_loss"] == 1
    assert ether["evidence"]["configured_min_links"] == 1
    assert ether["evidence"]["service_path_survival"] == "not_verified"
    assert "min-links" in ether["note"] and "service-path survival" in ether["note"]
    stp = _family(result, "stp")[0]
    assert stp["source_owner"] == (
        "failover_readiness/1 + stp_consistency_delta/1 + stp_topology_delta/1"
    )
    assert stp["evidence"]["n_root_subjects"] == 1
    assert stp["evidence"]["n_candidate_backups"] == 1
    assert "convergence" in stp["note"] and stp["assurance_level"] == "not_verified"
    assert result["status"] == "not_verified"  # missing MCLAG/path evidence is not masked
    _assert_no_decision_owner(result)


def test_current_stp_and_etherchannel_faults_remain_explicit(tmp_path):
    (tmp_path / "stp").mkdir()
    ifaces, health, receipt = _stp_run(tmp_path / "stp", inconsistent="Gi1/0/7\n")
    snapshot = _ether_snapshot(
        tmp_path / "ether",
        "Group Port-channel Protocol Ports\n"
        "1 Po1(SU) LACP Gi1/0/1(P) Gi1/0/2(D)\n",
        host="sw1",
    )
    snapshot.update({
        "interfaces": ifaces,
        "protocol_health": health,
        "protocol_assessability": receipt,
        "stp_roots": {
            "sw1": {"10": {
                "root_priority": 4096, "root_address": "aaaa.0000.0001",
                "is_root": True, "is_mst": False, "bridge_priority": 4096,
            }},
        },
    })
    # Rebuild the protected EtherChannel owner against the final receipt. The additive typed owner
    # remains present, but a missing protected-v1 denominator must fail closed rather than be backfilled.
    snapshot["etherchannel_baseline"] = summarize_etherchannel_baseline(
        snapshot["etherchannel_projection"], receipt, devices=snapshot["devices"],
    )
    result = compute_l2_failure_rehearsal(_bind(snapshot))

    stp = _family(result, "stp")[0]
    assert stp["disposition"] == "current_fault"
    assert stp["current_fault"] is True
    assert stp["evidence"]["n_current_faults"] == 1
    assert "sw1" in stp["evidence"]["current_fault_subjects"]
    # The missing protected-v1 receipt cannot be reinterpreted through the additive owner.
    ether = _family(result, "etherchannel")[0]
    assert ether["disposition"] == "current_fault"
    assert result["summary"]["n_current_faults"] >= 2


def test_current_etherchannel_fault_is_not_laundered_by_single_member_arithmetic(tmp_path):
    snapshot = _bind(_ether_snapshot(
        tmp_path,
        "Group Port-channel Protocol Ports\n"
        "1 Po1(SU) LACP Gi1/0/1(P) Gi1/0/2(D)\n",
    ))
    result = compute_l2_failure_rehearsal(snapshot)

    row = _family(result, "etherchannel")[0]
    assert row["disposition"] == "current_fault"
    assert row["current_fault"] is True
    assert row["evidence"]["observed_forwarding_member_count"] == 1
    assert row["evidence"]["remaining_forwarding_members_after_loss"] == 0
    assert cutover_operator_evidence(snapshot)["rehearsal"]["status"] == "current_fault"


def test_single_healthy_etherchannel_member_is_projected_risk_not_survival(tmp_path):
    snapshot = _bind(_ether_snapshot(
        tmp_path,
        "Group Port-channel Protocol Ports\n"
        "1 Po1(SU) LACP Gi1/0/1(P)\n",
    ))
    result = compute_l2_failure_rehearsal(snapshot)

    row = _family(result, "etherchannel")[0]
    assert row["current_fault"] is False
    assert row["disposition"] == "projected_risk"
    assert row["evidence"]["observed_forwarding_member_count"] == 1
    assert row["evidence"]["remaining_forwarding_members_after_loss"] == 0
    assert row["assurance_level"] == "not_verified"
    assert result["status"] == "projected_risk"
    assert result["summary"]["n_projected_risks"] == 1
    assert cutover_operator_evidence(snapshot)["rehearsal"]["status"] == "projected_risk"


def test_multichassis_rehearsal_requires_typed_attachment_identity_and_retains_faults(
        tmp_path):
    healthy = compute_l2_failure_rehearsal(_bind(_multichassis_snapshot(tmp_path / "healthy")))
    row = _family(healthy, "multichassis_lag")[0]
    assert row["disposition"] == "simulation_only"
    assert row["evidence"]["transition"] == "unchanged_healthy"
    assert row["evidence"]["reconciled_local_leg_count"] == 2
    assert row["evidence"]["remaining_observed_legs_after_loss"] == 1
    assert row["evidence"]["reciprocal_peer_evidenced"] is True
    assert row["evidence"]["matching_lacp_partner_evidenced"] is True
    assert row["evidence"]["etherchannel_local_leg_join_status"] == "reconciled"
    assert len(row["evidence"]["etherchannel_local_leg_join"]) == 2
    assert all(
        leg["member_loss_status"] == "pass"
        and leg["min_links"] == 1
        and leg["forwarding_member_count"] == 2
        and leg["service_path_survival"] == "not_verified"
        for leg in row["evidence"]["etherchannel_local_leg_join"]
    )
    assert row["evidence"]["service_path_survival"] == "not_verified"
    assert "Neither owner proves" in row["note"]

    degraded = compute_l2_failure_rehearsal(_bind(
        _multichassis_snapshot(tmp_path / "degraded", degraded=True)))
    fault = _family(degraded, "multichassis_lag")[0]
    assert fault["disposition"] == "current_fault"
    assert fault["current_fault"] is True

    missing_partner = compute_l2_failure_rehearsal(_bind(
        _multichassis_snapshot(tmp_path / "missing-partner", partner=False)))
    gap = _family(missing_partner, "multichassis_lag")[0]
    assert gap["disposition"] == "not_verified"
    assert "reciprocal peers" in gap["note"] and "LACP" in gap["note"]


def test_multichassis_baseline_must_reconcile_to_complete_copublished_observations(tmp_path):
    missing = _multichassis_snapshot(tmp_path / "missing")
    missing.pop("multichassis_lag_typed_observations")
    one_sided = _multichassis_snapshot(tmp_path / "one-sided")
    one_sided["multichassis_lag_typed_observations"]["observations"].pop()

    for label, snapshot in (("missing", missing), ("one-sided", one_sided)):
        row = _family(
            compute_l2_failure_rehearsal(_bind(snapshot)), "multichassis_lag",
        )[0]
        assert row["disposition"] == "not_verified", label
        assert row["evidence"] == {}, label
        assert "co-published typed observation" in row["note"], label


def test_explicit_nxos_orphan_is_projected_risk_without_service_survival(tmp_path):
    result = compute_l2_failure_rehearsal(_bind(
        _multichassis_snapshot(tmp_path, orphan=True)))
    row = next(
        scenario for scenario in _family(result, "multichassis_lag")
        if scenario["subject"] == "multichassis_lag|orphan|leaf-a|Eth1/45"
    )

    assert row["disposition"] == "projected_risk"
    assert row["failure_scenario"] == "single_peer_or_local_device_loss"
    assert row["evidence"]["vlans"] == [10]
    assert row["evidence"]["suspend_on_peer_link_loss"] is True
    assert row["evidence"]["peer_link_loss_behavior"] == "configured_suspend"
    assert row["evidence"]["service_path_survival"] == "not_verified"


def test_canonical_orphan_risk_forces_review_and_workbook_projects_same_gate(
        tmp_path):
    before, after, comparison = _canonical_multichassis_pair(
        tmp_path / "orphan", orphan=True)
    gate = comparison["cutover_gate"]
    rehearsal = comparison["operator_evidence"]["rehearsal"][
        "l2_failure_rehearsal"]

    assert comparison["comparison_admission"]["status"] == "admitted"
    assert comparison["verdict"] == "CLEAN"
    assert comparison["precert"]["verdict"] == "PASS"
    assert comparison["protocol_families"]["summary"]["n_blocking"] == 0
    assert comparison["protocol_families"]["summary"]["n_review"] == 0
    assert gate["verdict"] == "REVIEW"
    assert gate["l2_rehearsal_status"] == "projected_risk"
    assert gate["l2_rehearsal_projected_risks"] == 1
    assert gate["l2_rehearsal_not_verified"] == 0
    assert gate["l2_rehearsal_applicable_families"] == [
        "etherchannel", "multichassis_lag",
    ]
    assert "service survival remains not verified" in gate["l2_rehearsal_note"]
    assert all(row["assurance_level"] == "not_verified"
               for row in rehearsal["scenarios"])

    output = tmp_path / "orphan-review.xlsx"
    projected = write_diff_workbook(
        before, after, str(output), comparison=comparison)

    assert projected == gate
    workbook = load_workbook(output, read_only=True)
    summary = workbook["Summary"]
    rows = {
        summary.cell(index, 1).value: (
            summary.cell(index, 3).value,
            summary.cell(index, 4).value,
        )
        for index in range(2, summary.max_row + 1)
    }
    workbook.close()
    assert rows["CUTOVER GATE VERDICT"][0] == "REVIEW"
    assert rows["L2 BOUNDED REHEARSAL"] == (
        "PROJECTED_RISK", gate["l2_rehearsal_note"])


def test_rehearsal_gate_is_applicability_aware_and_current_fault_safe(tmp_path):
    _before, _after, routed = _canonical_pair(
        _clean_comparison_snapshot("FULL/DR", "2026-08-20T00:00:00"),
        engagement="ENG-ROUTED-ONLY",
    )
    assert routed["cutover_gate"]["verdict"] == "PASS"
    assert routed["cutover_gate"]["l2_rehearsal_status"] == "not_applicable"
    assert routed["cutover_gate"]["l2_rehearsal_not_verified"] == 0

    _before, _after, healthy = _canonical_multichassis_pair(
        tmp_path / "healthy")
    assert healthy["cutover_gate"]["verdict"] == "PASS"
    assert healthy["cutover_gate"]["l2_rehearsal_status"] == "simulation_only"
    assert "do not prove" in healthy["cutover_gate"]["l2_rehearsal_note"]

    _before, _after, degraded = _canonical_multichassis_pair(
        tmp_path / "degraded", degraded=True)
    assert degraded["cutover_gate"]["verdict"] != "PASS"
    assert degraded["cutover_gate"]["l2_rehearsal_status"] == "current_fault"
    assert degraded["cutover_gate"]["l2_rehearsal_current_faults"] >= 1


def test_applicable_missing_scenario_and_mutated_supplied_receipt_fail_closed(
        tmp_path):
    _before, _after, healthy = _canonical_multichassis_pair(
        tmp_path / "missing")
    receipt = copy.deepcopy(
        healthy["operator_evidence"]["rehearsal"]["l2_failure_rehearsal"])
    receipt["scenarios"] = [
        row for row in receipt["scenarios"] if row["family"] != "etherchannel"
    ]
    by_disposition = {
        disposition: sum(row["disposition"] == disposition
                         for row in receipt["scenarios"])
        for disposition in (
            "simulation_only", "projected_risk", "current_fault", "not_verified")
    }
    receipt["summary"].update({
        "n_scenarios": len(receipt["scenarios"]),
        "n_current_faults": by_disposition["current_fault"],
        "n_projected_risks": by_disposition["projected_risk"],
        "n_not_verified": by_disposition["not_verified"],
        "by_disposition": by_disposition,
    })
    receipt["status"] = (
        "current_fault" if by_disposition["current_fault"] else
        "projected_risk" if by_disposition["projected_risk"] else
        "not_verified" if by_disposition["not_verified"] else
        "simulation_only"
    )
    validation = validate_l2_failure_rehearsal(receipt)
    assert validation["valid"] is True
    assert validation["gate_status"] == "not_verified"
    assert validation["n_not_verified"] == 1

    _before, _after, orphan = _canonical_multichassis_pair(
        tmp_path / "mutated", orphan=True)
    authority = bound_comparison_decision_input_authority(orphan)
    additive = {
        "comparison_schema", "comparison_admission", "change_intent",
        "protocol_families", "precert", "cutover_gate", "operator_evidence",
        "comparison_receipt",
    }
    delta = {key: value for key, value in orphan.items() if key not in additive}
    malformed = copy.deepcopy(orphan["operator_evidence"])
    malformed["rehearsal"]["l2_failure_rehearsal"]["summary"][
        "n_projected_risks"] = 0
    gate = compute_cutover_gate(
        delta,
        orphan["precert"],
        comparison_admission=orphan["comparison_admission"],
        protocol_family_changes=orphan["protocol_families"],
        operator_evidence=malformed,
        decision_input_authority=authority,
    )
    assert gate["verdict"] == "INDETERMINATE"
    assert gate["comparison_admission_status"] == "not_comparable"
    assert gate["l2_rehearsal_status"] == "not_verified"
    assert "detached" in gate["l2_rehearsal_note"]


def test_l2_summary_counters_reject_bool_equal_to_integer(tmp_path):
    _before, _after, orphan = _canonical_multichassis_pair(
        tmp_path / "bool-count", orphan=True)
    receipt = orphan["operator_evidence"]["rehearsal"]["l2_failure_rehearsal"]
    assert receipt["summary"]["n_projected_risks"] == 1
    assert receipt["summary"]["by_disposition"]["projected_risk"] == 1
    assert validate_l2_failure_rehearsal(receipt)["valid"] is True

    for path in (
        ("n_projected_risks",),
        ("by_disposition", "projected_risk"),
    ):
        malformed = copy.deepcopy(receipt)
        target = malformed["summary"]
        for key in path[:-1]:
            target = target[key]
        target[path[-1]] = True
        assert validate_l2_failure_rehearsal(malformed)["valid"] is False


def test_detached_or_mutated_snapshot_cannot_authorize_failure_projection(tmp_path):
    raw = _ether_snapshot(
        tmp_path,
        "Group Port-channel Protocol Ports\n"
        "1 Po1(SU) LACP Gi1/0/1(P) Gi1/0/2(P)\n",
    )
    detached = compute_l2_failure_rehearsal(raw)
    assert detached["source_bound"] is False
    assert {row["disposition"] for row in detached["scenarios"]} == {"not_verified"}

    bound = _bind(raw)
    bound["devices"]["unexpected"] = {}
    mutated = compute_l2_failure_rehearsal(bound)
    assert mutated["source_bound"] is False
    assert {row["disposition"] for row in mutated["scenarios"]} == {"not_verified"}


def test_operator_evidence_binds_complete_l2_projection_without_claiming_rehearsal(tmp_path):
    snapshot = _bind(_ether_snapshot(
        tmp_path,
        "Group Port-channel Protocol Ports\n"
        "1 Po1(SU) LACP Gi1/0/1(P) Gi1/0/2(P)\n",
    ))
    evidence = cutover_operator_evidence(snapshot)
    l2 = evidence["rehearsal"]["l2_failure_rehearsal"]

    assert l2["schema"] == L2_FAILURE_REHEARSAL_SCHEMA
    assert l2["summary"]["n_scenarios"] == len(l2["scenarios"])
    assert evidence["rehearsal"]["assurance_level"] == "not_verified"
    assert "operator rehearsal" in evidence["rehearsal"]["note"]


def test_v1_l2_affecting_failure_cannot_be_forged_into_preserved_simulation():
    snapshot = _bind(_forged_preserved_traffic_snapshot())
    post = snapshot["traffic_assurance"]["results"][0]["failure"]["post_dimensions"]
    assert "R2" in {
        hop["host"] for direction in ("forward", "reverse")
        for hop in post["path"][direction]["hops"]
    }
    result = compute_l2_failure_rehearsal(snapshot)
    row = _family(result, "service_path")[0]

    assert row["subject"] == "erp-forward"
    assert row["source_owner"] == "traffic_assurance_set/1 -> cutover_sim/1"
    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}
    assert "malformed" in row["note"]
    assert row["assurance_level"] == "not_verified"


def test_favorable_service_projection_requires_complete_outer_owner_custody_and_claims():
    mutations = (
        ("missing custody", lambda row: row.pop("custody_trust")),
        ("unverified custody", lambda row: row.__setitem__("custody_trust", "not_evaluated")),
        ("missing dimensions", lambda row: row.__setitem__("dimensions", {})),
        ("missing claims", lambda row: row.__setitem__("claims", [])),
        ("unsupported", lambda row: row.__setitem__("supported", False)),
    )
    for label, mutate in mutations:
        snapshot = _traffic_failure_snapshot()
        mutate(snapshot["traffic_assurance"]["results"][0])
        row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]
        assert row["disposition"] == "not_verified", label
        assert row["evidence"] == {}, label


def test_dimension_claim_contradiction_cannot_retain_producer_evidence():
    contradicted = _traffic_failure_snapshot()
    result = contradicted["traffic_assurance"]["results"][0]
    for dimensions in (result["dimensions"], result["failure"]["post_dimensions"]):
        for axis in ("path", "policy", "mtu", "ecmp"):
            for direction in ("forward", "reverse"):
                dimensions[axis][direction]["verdict"] = "refuted"
    row = _family(compute_l2_failure_rehearsal(_bind(contradicted)), "service_path")[0]
    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}

def test_observed_drop_path_cannot_be_relabelled_as_proven_preservation():
    snapshot = _traffic_failure_snapshot()
    result = snapshot["traffic_assurance"]["results"][0]
    for dimensions in (result["dimensions"], result["failure"]["post_dimensions"]):
        dimensions["path"]["forward"].update({
            "verdict": "proven",
            "state": "observed_drop",
            "status": "computed:unreachable",
            "reached": False,
            "drop_evidence": "observed_discard",
            "ecmp_dropping_legs": [{"drop_evidence": "observed_discard"}],
        })

    row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}


def test_reached_path_requires_hop_owners_and_claim_boundary_reconciliation():
    snapshot = _traffic_failure_snapshot()
    result = snapshot["traffic_assurance"]["results"][0]
    for dimensions in (result["dimensions"], result["failure"]["post_dimensions"]):
        dimensions["path"]["forward"]["hops"] = []

    row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}


def test_proven_acl_stage_requires_closed_native_permit_semantics():
    snapshot = _traffic_failure_snapshot()
    result = snapshot["traffic_assurance"]["results"][0]
    for dimensions in (result["dimensions"], result["failure"]["post_dimensions"]):
        stage = dimensions["policy"]["forward"]["stages"][0]
        stage.update({
            "native_result": "PROVEN_NONE",
            "detail": "the exact tuple is denied (the implicit deny is included)",
        })

    row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}


def test_traffic_axis_custody_requires_complete_owner_receipts():
    snapshot = _traffic_failure_snapshot()
    result = snapshot["traffic_assurance"]["results"][0]
    for dimensions in (result["dimensions"], result["failure"]["post_dimensions"]):
        for direction in ("forward", "reverse"):
            dimensions["mtu"][direction]["interface_custody"] = {
                "verdict": "proven", "gaps": [],
            }
            dimensions["ecmp"][direction]["route_custody"] = {
                "verdict": "proven", "gaps": [],
            }

    row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}


def test_cutover_action_and_removed_subjects_are_revalidated_by_existing_owner():
    snapshot = _traffic_failure_snapshot()
    failure = snapshot["traffic_assurance"]["results"][0]["failure"]
    failure["mutation"]["removed_hosts"] = ["R3"]
    failure["cutover_evidence"]["steps"][0]["removed_hosts"] = ["R3"]

    row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}


def test_forged_preservation_cannot_coexist_with_newly_lost_requested_flow():
    snapshot = _forged_preserved_traffic_snapshot()
    result = snapshot["traffic_assurance"]["results"][0]
    step = result["failure"]["cutover_evidence"]["steps"][0]
    step["newly_lost_flows"] = [{
        "src": result["intent"]["src"],
        "dst": result["intent"]["dst"],
        "kind": "path_lost",
    }]
    simulation = result["failure"]["cutover_evidence"]
    simulation["summary"]["total_newly_lost"] = 1
    simulation["worst_step"] = 0

    row = _family(compute_l2_failure_rehearsal(_bind(snapshot)), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}


def test_current_traffic_failure_continuity_gap_is_explicit_not_verified():
    row = _family(
        compute_l2_failure_rehearsal(_bind(_traffic_failure_snapshot())),
        "service_path",
    )[0]

    assert row["subject"] == "erp-forward"
    assert row["disposition"] == "not_verified"
    assert row["evidence"]["producer_status"] == "l2_failover_not_proven"
    assert row["evidence"]["n_steps"] == 1
    assert row["evidence"]["n_indeterminate"] >= 1


def test_malformed_stored_cutover_simulation_cannot_claim_favorable_projection():
    snapshot = _traffic_failure_snapshot()
    # One stored row cannot reconcile to an asserted 99-step simulation.
    snapshot["traffic_assurance"]["results"][0]["failure"]["cutover_evidence"][
        "summary"
    ]["n_steps"] = 99
    snapshot = _bind(snapshot)
    row = _family(compute_l2_failure_rehearsal(snapshot), "service_path")[0]

    assert row["disposition"] == "not_verified"
    assert "malformed" in row["note"]


def test_invalid_cutover_step_and_forged_zero_summary_cannot_claim_preservation():
    snapshot = _traffic_failure_snapshot()
    step = snapshot["traffic_assurance"]["results"][0]["failure"]["cutover_evidence"][
        "steps"
    ][0]
    step.update({
        "newly_lost_flows": [{}],
        "split_brain_risks": [{}],
        "indeterminate": 7,
        "valid": False,
    })
    # Preserve the forged favorable producer summary/gate: validation must derive from the closed step,
    # not trust these zero counters or the apparent top-level proven verdict.
    row = _family(
        compute_l2_failure_rehearsal(_bind(snapshot)),
        "service_path",
    )[0]

    assert row["subject"] == "erp-forward"
    assert row["disposition"] == "not_verified"
    assert row["evidence"] == {}
    assert "malformed" in row["note"]
