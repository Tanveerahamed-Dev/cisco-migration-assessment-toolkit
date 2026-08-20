"""Configured-FHRP-group operator projections.

The core receipt owns classification.  These tests pin the presentation contract:
uncapped workbook rows, uncapped blocker retention, distinct host × subtype coverage,
configured-only Explorer drill-downs, fail-closed census checks, and secret-safe copy.
"""

from __future__ import annotations

import json
import pathlib
import shutil
import subprocess

import pytest
from openpyxl import Workbook

from cisco_toolkit.excel import (
    FHRP_GROUP_INTENT_SHEET_NAME,
    write_fhrp_configured_group_sheet,
)
from cisco_toolkit.fhrp_intent import (
    _acceptance,
    _baseline_digest_payload,
    _coverage_receipt_hashes,
    _sha,
    embedded_fhrp_configured_group_baseline,
)


ROOT = pathlib.Path(__file__).resolve().parents[1]
EXPLORER = ROOT / "cisco_toolkit" / "blast_radius_explorer.html"
NODE = shutil.which("node")
PROTOCOLS = ("HSRP", "VRRP", "GLBP")
GROUP_STATUS_ORDER = (
    "degraded", "review", "not_verified", "assessed", "administratively_disabled",
)
COVERAGE_STATUS_ORDER = (
    "degraded", "review", "not_verified", "assessed", "not_applicable",
)


def _row(index: int, status: str = "assessed") -> dict:
    protocol = PROTOCOLS[index % len(PROTOCOLS)]
    runtime_observed = status == "assessed"
    disabled = status == "administratively_disabled"
    healthy_state = {"HSRP": "ACTIVE", "VRRP": "MASTER", "GLBP": "ACTIVE"}[protocol]
    runtime_command = {
        "HSRP": "show standby brief",
        "VRRP": "show vrrp brief",
        "GLBP": "show glbp brief",
    }[protocol]
    findings = (
        [{"kind": status, "code": f"fixture_{status}", "issue": f"fixture {status} evidence"}]
        if status in {"degraded", "review", "not_verified"} else []
    )
    row = {
        "switch": f"fhrp-{index % 4}",
        "protocol": protocol,
        "interface": f"Vlan{100 + index}",
        "group": str(10 + index),
        "group_key": f"{protocol}:Vlan{100 + index}:{10 + index}",
        "scope": "default/ipv4",
        "configured": True,
        "configured_vip": f"192.0.2.{(index % 240) + 1}",
        "activation": "disabled" if disabled else "active",
        "runtime_observed": runtime_observed,
        "runtime_vip": f"192.0.2.{(index % 240) + 1}" if runtime_observed else "",
        "runtime_state_raw": "Active" if runtime_observed else "",
        "runtime_state": healthy_state if runtime_observed else "NOT_OBSERVED",
        "status": status,
        "command": runtime_command,
        "acceptance": "",
        "source_key": (
            f"show running-config#line:{index + 1} + "
            f"{runtime_command}"
        ),
        "projection_custody": "current_run_source_bound",
        "findings": findings,
    }
    row["acceptance"] = _acceptance(row)
    return row


def _coverage_cell(
    switch: str,
    protocol: str,
    *,
    subject: bool,
    status: str,
    config_capture_status: str = "ok",
    config_parser_status: str = "complete",
    runtime_capture_status: str = "ok",
    runtime_parser_status: str = "complete",
    configured_group_count: int = 0,
    runtime_parsed_count: int = 0,
) -> dict:
    return {
        "switch": switch,
        "protocol": protocol,
        "platform": "IOS-XE",
        "subject": subject,
        "status": status,
        "config_command": "show running-config",
        "config_capture_status": config_capture_status,
        "config_parser_status": config_parser_status,
        "runtime_command": {
            "HSRP": "show standby brief",
            "VRRP": "show vrrp brief",
            "GLBP": "show glbp brief",
        }[protocol],
        "runtime_capture_status": runtime_capture_status,
        "runtime_parser_status": runtime_parser_status,
        "config_candidate_count": configured_group_count,
        "configured_group_count": configured_group_count,
        "config_rejected_count": 0,
        "excluded_scope_count": 0,
        "unsupported_relevant_count": 0,
        "runtime_candidate_count": runtime_parsed_count,
        "runtime_parsed_count": runtime_parsed_count,
        "runtime_rejected_count": 0,
        "config_sha256": "a" * 64,
        "runtime_sha256": "b" * 64,
        "projection_sha256": "c" * 64,
        "finding_codes": [],
    }


def _seal(baseline: dict) -> dict:
    for cell in baseline["coverage"]:
        cell_rows = [
            row for row in baseline["rows"]
            if row["switch"] == cell["switch"] and row["protocol"] == cell["protocol"]
        ]
        config_hash, runtime_hash, projection_hash = _coverage_receipt_hashes(cell, cell_rows)
        cell.update({
            "config_sha256": config_hash,
            "runtime_sha256": runtime_hash,
            "projection_sha256": projection_hash,
        })
    baseline["summary"]["baseline_sha256"] = _sha(_baseline_digest_payload(baseline))
    return baseline


def _baseline(rows: list[dict], verdict: str = "BLOCKED") -> dict:
    hosts = sorted({row["switch"] for row in rows})
    subjects = {(row["switch"], row["protocol"]) for row in rows}
    coverage = []
    for switch in hosts:
        for protocol in PROTOCOLS:
            cell_rows = [
                row for row in rows
                if row["switch"] == switch and row["protocol"] == protocol
            ]
            subject = bool(cell_rows)
            if subject:
                statuses = {row["status"] for row in cell_rows}
                status = (
                    "degraded" if "degraded" in statuses else
                    "review" if "review" in statuses else
                    "not_verified" if "not_verified" in statuses else "assessed"
                )
            else:
                status = "not_applicable"
            coverage.append(_coverage_cell(
                switch,
                protocol,
                subject=subject,
                status=status,
                configured_group_count=sum(row["configured"] is True for row in cell_rows),
                runtime_parsed_count=sum(row["runtime_observed"] is True for row in cell_rows),
            ))
    by_status = {status: sum(row["status"] == status for row in rows) for status in GROUP_STATUS_ORDER}
    by_coverage = {
        status: sum(cell["status"] == status for cell in coverage)
        for status in COVERAGE_STATUS_ORDER
    }
    subject_hosts = {switch for switch, _protocol in subjects}
    baseline = {
        "schema": "fhrp_configured_group_baseline/1",
        "scope": {
            "routing_instance": "default",
            "afi": "ipv4",
            "group_kind": "direct_literal_local",
        },
        "verdict": verdict,
        "assessed": (
            verdict in {"CLEAR", "BLOCKED"}
            and by_status["review"] == 0
            and by_status["not_verified"] == 0
        ),
        "projection_custody": "current_run_source_bound",
        "rows": rows,
        "coverage": coverage,
        "findings": [],
        "summary": {
            "by_status": by_status,
            "by_coverage_status": by_coverage,
            "n_hosts": len(hosts),
            "n_coverage_cells": len(coverage),
            "n_subject_hosts": len(subject_hosts),
            "n_subject_cells": len(subjects),
            "n_configured_groups": sum(row["configured"] is True for row in rows),
            "n_active_groups": sum(row["activation"] == "active" for row in rows),
            "n_runtime_groups": sum(row["runtime_observed"] is True for row in rows),
            "n_assessed": by_status["assessed"],
            "n_degraded": by_status["degraded"],
            "n_review": by_status["review"],
            "n_not_verified": by_status["not_verified"],
            "n_disabled": by_status["administratively_disabled"],
            "baseline_sha256": "0" * 64,
        },
        "limitations": ["bounded fixture"],
    }
    return _seal(baseline)


def _rowless_not_applicable(n_hosts: int = 23) -> dict:
    baseline = _baseline([], "NOT_APPLICABLE")
    coverage = [
        _coverage_cell(f"coverage-{index:02d}", protocol, subject=False, status="not_applicable")
        for index in range(n_hosts)
        for protocol in PROTOCOLS
    ]
    baseline["coverage"] = coverage
    baseline["assessed"] = False
    baseline["summary"].update({
        "n_hosts": n_hosts,
        "n_coverage_cells": n_hosts * len(PROTOCOLS),
        "n_subject_hosts": 0,
        "n_subject_cells": 0,
        "by_coverage_status": {
            "degraded": 0,
            "review": 0,
            "not_verified": 0,
            "assessed": 0,
            "not_applicable": n_hosts * len(PROTOCOLS),
        },
    })
    return _seal(baseline)


def _sheet_text(ws) -> str:
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    )


def _docx_text(path: pathlib.Path) -> str:
    docx = pytest.importorskip("docx")
    document = docx.Document(str(path))
    return "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )


def _explorer_section(baseline: dict, host: str | None = None) -> str:
    assert NODE is not None
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _FHRP_GROUP_SCOPE=');
const end=html.indexOf('const _BGP_INTENT_SCOPE=',start);
if(start<0||end<0)throw new Error('configured FHRP projection block not found');
const input=JSON.parse(fs.readFileSync(0,'utf8'));
let SNAP={fhrp_configured_group_baseline:input.baseline};
function esc(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
process.stdout.write(fhrpConfiguredGroupSection(input.host));
"""
    proc = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        input=json.dumps({
            "baseline": embedded_fhrp_configured_group_baseline(baseline),
            "host": host,
        }),
        timeout=20,
    )
    return proc.stdout


def test_fhrp_group_intent_workbook_is_uncapped_blocker_visible_and_secret_safe():
    rows = [_row(index) for index in range(205)]
    blocker = _row(205, "degraded")
    blocker["switch"] = "last-fhrp-device"
    blocker["configured_vip"] = "203.0.113.254"
    blocker["findings"] = [{
        "kind": "degraded",
        "code": "configured_group_not_observed",
        "issue": (
            "last configured FHRP blocker; authentication md5 visible-workbook-secret; "
            "key-string hidden-workbook-secret"
        ),
    }]
    rows.append(blocker)

    wb = Workbook()
    write_fhrp_configured_group_sheet(wb, _baseline(rows))
    ws = wb[FHRP_GROUP_INTENT_SHEET_NAME]
    text = _sheet_text(ws)

    assert ws.max_row == 4 + len(rows)
    assert ws.freeze_panes == "A5"
    assert "CONFIGURED FHRP GROUP GATE — DEFAULT/GLOBAL IPv4" in ws["A1"].value
    assert "configured-active local group absent from usable subtype runtime evidence is a blocker" in ws["A1"].value
    assert "VRFs, IPv6, templates/inheritance/dynamic constructs" in ws["A1"].value
    assert "NX-OS configured-group parsing is limited to nested HSRP" in ws["A1"].value
    assert "203.0.113.254" in text and "last configured FHRP blocker" in text
    assert "Group rows — configured: 206" in text
    assert "Host × subtype coverage (distinct from group rows): 5 host(s) · 15 HSRP/VRRP/GLBP cell(s)" in text
    assert "current_run_source_bound" in text
    assert ws.cell(4 + len(rows), 1).fill.fgColor.rgb.endswith("F4CCCC")
    assert "visible-workbook-secret" not in text
    assert "raw-workbook-secret" not in text
    assert "hidden-workbook-secret" not in text
    assert r"C:\sensitive\raw-running-config.txt" not in text
    assert "[REDACTED]" in text


def test_rowless_not_applicable_workbook_preserves_host_by_subtype_census():
    wb = Workbook()
    write_fhrp_configured_group_sheet(wb, _rowless_not_applicable())
    ws = wb[FHRP_GROUP_INTENT_SHEET_NAME]
    text = _sheet_text(ws)

    assert "Group rows — configured: 0" in ws["A2"].value
    assert "Host × subtype coverage (distinct from group rows): 23 host(s) · 69 HSRP/VRRP/GLBP cell(s)" in ws["A3"].value
    assert "coverage cells — 0 degraded, 0 review, 0 not verified, 0 assessed, and 69 not applicable" in ws["A3"].value
    assert (
        "NOT_APPLICABLE means no in-scope literal local group subject was identified; it is not "
        "proof that FHRP is absent or that configuration coverage is complete."
    ) in ws["A3"].value
    assert ws["A3"].fill.fgColor.rgb.endswith("FFF2CC")
    assert "a" * 64 not in text


def test_fhrp_workbook_fails_closed_on_non_reconciling_group_or_coverage_counts():
    baseline = _rowless_not_applicable(3)
    baseline["summary"]["n_coverage_cells"] = 8
    baseline["raw_config"] = "authentication md5 malformed-contract-secret"
    wb = Workbook()
    write_fhrp_configured_group_sheet(wb, baseline)
    ws = wb[FHRP_GROUP_INTENT_SHEET_NAME]
    assert "NOT ASSESSED" in _sheet_text(ws)
    assert "NOT_APPLICABLE" not in ws["A1"].value
    assert "malformed-contract-secret" not in _sheet_text(ws)

    baseline = _baseline([_row(0)], "CLEAR")
    baseline["summary"]["n_active_groups"] = 0
    wb = Workbook()
    write_fhrp_configured_group_sheet(wb, baseline)
    assert "NOT ASSESSED" in _sheet_text(wb[FHRP_GROUP_INTENT_SHEET_NAME])


def test_runbook_keeps_every_configured_fhrp_blocker_outside_protocol_cap(tmp_path):
    from cisco_toolkit.runbook import write_runbook_docx

    rows = [_row(index, "degraded") for index in range(19)]
    rows[-1]["switch"] = "last-fhrp-runbook"
    rows[-1]["configured_vip"] = "203.0.113.219"
    rows[-1]["findings"] = [{
        "kind": "degraded",
        "code": "configured_group_not_observed",
        "issue": "last FHRP blocker; authentication md5 visible-runbook-secret; key-string hidden-secret",
    }]
    output = tmp_path / "fhrp-runbook.docx"
    write_runbook_docx(
        str(output),
        {"devices": {"fhrp-0": {"hostname": "fhrp-0"}},
         "fhrp_configured_group_baseline": _baseline(rows)},
        "Configured FHRP test",
    )
    text = _docx_text(output)

    assert "Configured FHRP group gate — default/global IPv4 direct literal local groups" in text
    assert "Group rows: 19 configured, 19 configured-active, 0 runtime-observed" in text
    assert "Host × subtype coverage (distinct from group rows): 5 host(s) · 15 HSRP/VRRP/GLBP cell(s)" in text
    assert "all rendered; these rows are outside the ordinary protocol display cap" in text
    assert "203.0.113.219" in text and "last FHRP blocker" in text
    assert "current_run_source_bound" in text
    assert "visible-runbook-secret" not in text
    assert "raw-runbook-secret" not in text
    assert "hidden-secret" not in text
    assert "[REDACTED]" in text


def test_rowless_not_applicable_runbook_is_explicit_and_qualified(tmp_path):
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "fhrp-rowless-runbook.docx"
    write_runbook_docx(
        str(output),
        {"fhrp_configured_group_baseline": _rowless_not_applicable()},
        "Rowless configured FHRP coverage",
    )
    text = _docx_text(output)
    assert "Group rows: 0 configured, 0 configured-active, 0 runtime-observed" in text
    assert "⚠ Host × subtype coverage (distinct from group rows): 23 host(s) · 69 HSRP/VRRP/GLBP cell(s)" in text
    assert "NOT_APPLICABLE means no in-scope literal local group subject was identified" in text
    assert "a" * 64 not in text


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_explorer_renders_every_fhrp_blocker_then_first_50_ordinary_groups():
    rows = [_row(index) for index in range(51)]
    degraded = _row(80, "degraded")
    degraded["switch"] = "fhrp-degraded"
    degraded["configured_vip"] = "203.0.113.80"
    degraded["findings"] = [{
        "kind": "degraded",
        "code": "configured_group_not_observed",
        "issue": "authentication md5 visible-explorer-secret",
    }]
    review = _row(81, "review")
    review["switch"] = "fhrp-review"
    review["configured_vip"] = "203.0.113.81"
    rows.extend([degraded, review])
    output = _explorer_section(_baseline(rows))

    assert "Configured FHRP group gate — default/global IPv4" in output
    assert "53</div><div class=\"k\">group configured" in output
    assert "203.0.113.80" in output and "203.0.113.81" in output
    assert "Vlan149 group 59" in output
    assert "Vlan150 group 60" not in output
    assert "All 2 blocker row(s) are rendered" in output
    assert "50 of 51 ordinary group row(s) are shown" in output
    assert "Host × subtype coverage (distinct from group rows):</b>" in output
    assert "visible-explorer-secret" not in output
    assert "[REDACTED]" in output


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_explorer_rowless_not_applicable_and_coverage_only_subject_are_drillable():
    output = _explorer_section(_rowless_not_applicable())
    assert "23 host(s) · 69 HSRP/VRRP/GLBP cell(s)" in output
    assert "badge b-watch\">NOT APPLICABLE" in output
    assert "NOT_APPLICABLE means no in-scope literal local group subject was identified" in output

    baseline = _baseline([], "INDETERMINATE")
    coverage = []
    for protocol in PROTOCOLS:
        cell = _coverage_cell(
            "coverage-only-fhrp", protocol,
            subject=protocol == "VRRP",
            status="review" if protocol == "VRRP" else "not_applicable",
            runtime_capture_status="not_observed",
            runtime_parser_status="not_verified",
        )
        if protocol == "VRRP":
            cell["unsupported_relevant_count"] = 1
        coverage.append(cell)
    baseline["assessed"] = False
    baseline["projection_custody"] = "embedded_unverified"
    baseline["coverage"] = coverage
    baseline["summary"].update({
        "n_hosts": 1,
        "n_coverage_cells": 3,
        "n_subject_hosts": 1,
        "n_subject_cells": 1,
        "by_coverage_status": {
            "degraded": 0, "review": 1, "not_verified": 0, "assessed": 0,
            "not_applicable": 2,
        },
    })
    _seal(baseline)
    detail = _explorer_section(baseline, "coverage-only-fhrp")
    assert "Host × subtype coverage detail" in detail
    assert "coverage-only-fhrp" in detail and "VRRP · SUBJECT" in detail
    assert "No configured-group rows were published for this host" in detail
    assert "a" * 64 not in detail


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_explorer_fails_closed_on_false_clear_or_extra_raw_contract_leaves():
    false_clear = _baseline([_row(0, "degraded")])
    false_clear["verdict"] = "CLEAR"
    output = _explorer_section(false_clear)
    assert "not assessed" in output
    assert "badge b-ok" not in output

    extra_leaf = _baseline([_row(0)])
    extra_leaf["raw_config"] = "authentication md5 must-never-render"
    output = _explorer_section(extra_leaf)
    assert "not assessed" in output
    assert "must-never-render" not in output


def test_explorer_fhrp_hosts_and_sections_are_wired_into_protocols_mode():
    html = EXPLORER.read_text(encoding="utf-8")
    model = html[html.index("function buildModel(snap)"):html.index("function linkCarries(")]
    detail_hosts = html[
        html.index("function protocolDetailHosts("):html.index("function drawProtocols()")
    ]
    protocols = html[html.index("function drawProtocols()"):html.index("function drawProtocolsDetail(")]
    detail = html[html.index("function drawProtocolsDetail("):html.index("/* =====================================================================", html.index("function drawProtocolsDetail("))]

    assert "fhrpIntentRows" in model and "fhrpCoverageSubjects" in model
    assert "fhrpIntent.coverage.filter(row=>row.subject===true)" in model
    assert "fhrpConfiguredGroupRows()" in detail_hosts
    assert "fhrpConfiguredGroupCoverageRows()" in detail_hosts
    assert "fhrpConfiguredGroupSection()" in protocols
    assert "fhrpConfiguredGroupSection(host)" in detail
    assert "bounded configured BGP peer and FHRP group gates" in html
