"""Sequential exact-candidate FHRP review survives every generic operator cap."""

from __future__ import annotations

from collections import Counter
import copy
import json
from pathlib import Path
import re
import shutil
import subprocess

from openpyxl import Workbook
import pytest

from cisco_toolkit.analyze import compute_current_baseline_gate, compute_validation_plan
from cisco_toolkit.capture_integrity import compute_capture_integrity_from_paths
from cisco_toolkit.excel import (
    FHRP_GROUP_INTENT_SHEET_NAME,
    write_fhrp_configured_group_sheet,
    write_validation_plan_sheet,
)
from cisco_toolkit.fhrp_intent import (
    _acceptance,
    _baseline_digest_payload,
    _coverage_receipt_hashes,
    _sha,
    compute_fhrp_configured_group_baseline,
    embedded_fhrp_configured_group_baseline,
    validate_fhrp_configured_group_baseline,
)


ROOT = Path(__file__).resolve().parents[1]
EXPLORER = ROOT / "cisco_toolkit" / "blast_radius_explorer.html"
NODE = shutil.which("node")
HOSTS = ("edge-a", "edge-b")
ISSUE_MARKER = (
    "Exact sequential candidate scope default/ipv4, HSRP interface Vlan10 group 10, "
    "configured/runtime VIP 10.0.10.1, observed role composition ACTIVE=2 across 2 distinct hosts"
)


def _dual_leader_baseline(tmp_path: Path) -> dict:
    paths = {}
    for host in HOSTS:
        host_dir = tmp_path / host
        host_dir.mkdir(parents=True)
        config = host_dir / "running.txt"
        runtime = host_dir / "standby.txt"
        config.write_text(
            f"hostname {host}\n"
            "interface Vlan10\n"
            " standby 10 ip 10.0.10.1\n"
            "end\n",
            encoding="utf-8",
        )
        runtime.write_text(
            "Interface   Grp  Pri P State    Active          Standby         Virtual IP\n"
            "Vl10        10   110 P Active   10.0.10.2       local           10.0.10.1\n",
            encoding="utf-8",
        )
        paths[host] = {
            "show running-config": str(config),
            "show standby brief": str(runtime),
        }

    baseline = compute_fhrp_configured_group_baseline(
        paths,
        compute_capture_integrity_from_paths(paths),
        {host: {"platform": "ios"} for host in HOSTS},
    )
    assert baseline["verdict"] == "INDETERMINATE"
    assert [row["status"] for row in baseline["rows"]] == ["review", "review"]
    assert all(
        row["findings"][0]["code"] == "election_multiple_leaders_observed"
        for row in baseline["rows"]
    )
    return baseline


def _ordinary_item(index: int) -> dict:
    return {
        "device": "edge-a",
        "platform": "ios",
        "wave": "Group 1",
        "category": "Gateway",
        "severity": "Medium",
        "check": f"ordinary validation {index:02d}",
        "command": f"show ordinary {index:02d}",
        "expect": "Observed ordinary state remains stable.",
        "why": "Ordinary bounded validation detail.",
    }


def _capped_plan(baseline: dict) -> dict:
    produced = compute_validation_plan(
        {host: {} for host in HOSTS},
        move_groups=[{"switches": list(HOSTS)}],
        devices={host: {"platform": "ios"} for host in HOSTS},
        fhrp_configured_group_baseline=baseline,
    )
    blockers = [item for item in produced["items"] if item["category"] == "FHRP"]
    assert len(blockers) == 2
    assert all(item["evidence_state"] == "review" for item in blockers)
    items = [_ordinary_item(index) for index in range(45)] + blockers
    produced["items"] = items
    produced["by_wave"] = {"Group 1": list(items)}
    produced["summary"] = {
        "n_items": len(items),
        "n_waves": 1,
        "n_high": sum(item["severity"] in {"Critical", "High"} for item in items),
        "by_category": dict(Counter(item["category"] for item in items)),
    }
    gate = compute_current_baseline_gate(produced)
    assert gate["integrity"]["valid"] is True
    assert gate["verdict"] == "INDETERMINATE"
    assert gate["summary"]["by_state"] == {
        "degraded": 0,
        "review": 2,
        "not_verified": 0,
    }
    return produced


def _snapshot(baseline: dict, plan: dict) -> dict:
    return {
        "script_version": "test",
        "devices": {host: {"platform": "ios"} for host in HOSTS},
        "move_groups": [{
            "group": "Group 1",
            "switches": list(HOSTS),
            "endpoints": 2,
        }],
        # Deliberately READY: the typed current-baseline review must override scheduling readiness.
        "migration_readiness": [{
            "group": "Group 1",
            "switches": list(HOSTS),
            "endpoints": 2,
            "readiness": "READY",
            "n_fail": 0,
            "n_warn": 0,
            "checks": [],
        }],
        "wave_sequencing": [{
            "group": "Group 1",
            "make_before_break": list(HOSTS),
            "hard_cutover": [],
            "hard_cutover_endpoints": 0,
            "sequence": "make-before-break",
        }],
        "fhrp_configured_group_baseline": baseline,
        "validation_plan": plan,
    }


def _sheet_text(ws) -> str:
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )


def _document_text(document) -> str:
    return "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )


def _tables_with_header(document, headers):
    expected = list(headers)
    return [
        table for table in document.tables
        if table.rows and [cell.text for cell in table.rows[0].cells] == expected
    ]


def _row_values(document, label, column=1):
    values = []
    for table in document.tables:
        for row in table.rows:
            cells = [cell.text.strip() for cell in row.cells]
            if cells and cells[0] == label and len(cells) > column:
                values.append(cells[column])
    return values


def test_dual_leader_review_is_visible_in_workbook_and_runbook_after_caps(tmp_path):
    baseline = _dual_leader_baseline(tmp_path / "captures")
    plan = _capped_plan(baseline)

    workbook = Workbook()
    write_fhrp_configured_group_sheet(workbook, baseline)
    write_validation_plan_sheet(workbook, plan)
    intent = workbook[FHRP_GROUP_INTENT_SHEET_NAME]
    validation = workbook["Cutover Validation"]

    assert intent.max_row == 6
    assert [intent.cell(row, 12).value for row in (5, 6)] == ["REVIEW", "REVIEW"]
    assert _sheet_text(intent).count(ISSUE_MARKER) == 4  # acceptance + finding for both rows
    assert "Capture timing is not simultaneous evidence" in _sheet_text(intent)
    assert "Matching the conflicting or unresolved sequential roles is NOT ACCEPTANCE" in (
        _sheet_text(intent)
    )
    assert validation.max_row == 4 + len(plan["items"])
    validation_rows = [
        [validation.cell(row, column).value for column in range(1, 14)]
        for row in range(5, validation.max_row + 1)
    ]
    fhrp_rows = [row for row in validation_rows if row[4] == "FHRP"]
    assert len(fhrp_rows) == 2
    assert {row[2] for row in fhrp_rows} == set(HOSTS)
    assert all(row[9] == "review" and row[10] == "current_run_source_bound" for row in fhrp_rows)
    assert all(ISSUE_MARKER in row[8] for row in fhrp_rows)

    docx = pytest.importorskip("docx")
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "dual-leader-runbook.docx"
    write_runbook_docx(str(output), _snapshot(baseline, plan), "FHRP election surface fixture")
    document = docx.Document(output)
    text = _document_text(document)
    assert "Verdict: INDETERMINATE" in text
    assert "2 review" in text
    assert text.count(ISSUE_MARKER) >= 4
    assert "Capture timing is not simultaneous evidence" in text
    assert "Matching the conflicting or unresolved sequential roles is NOT ACCEPTANCE" in text

    tables = _tables_with_header(
        document,
        ["Device", "Category", "Check", "Command", "Observed baseline / acceptance"],
    )
    assert len(tables) == 1
    rows = [[cell.text for cell in row.cells] for row in tables[0].rows[1:]]
    assert len(rows) == 16  # 14 ordinary rows plus both typed review blockers.
    retained = [row for row in rows if row[1] == "FHRP"]
    assert {row[0] for row in retained} == set(HOSTS)
    assert all("Evidence state: review" in row[4] for row in retained)
    assert all(ISSUE_MARKER in row[4] for row in retained)


def test_dual_leader_review_holds_mop_and_survives_its_validation_cap(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.mop import write_mop_docx

    baseline = _dual_leader_baseline(tmp_path / "captures")
    plan = _capped_plan(baseline)
    output = tmp_path / "dual-leader-mop.docx"
    write_mop_docx(str(output), _snapshot(baseline, plan), "FHRP election surface fixture")
    document = docx.Document(output)

    gate, = _row_values(document, "Current baseline gate (all validation groups)")
    assert gate.startswith("HOLD — current baseline INDETERMINATE")
    assert "overall readiness gate is READY" not in _document_text(document)
    tables = _tables_with_header(
        document,
        ["Sev", "Category", "Device", "Check", "Command", "Observed baseline / acceptance"],
    )
    assert len(tables) == 1
    rows = [[cell.text for cell in row.cells] for row in tables[0].rows[1:]]
    assert len(rows) == 42  # 40 ordinary rows plus both typed review blockers.
    retained = [row for row in rows if row[1] == "FHRP"]
    assert {row[2] for row in retained} == set(HOSTS)
    assert all("Evidence state: review" in row[5] for row in retained)
    assert all("current_run_source_bound" in row[5] for row in retained)
    assert all(ISSUE_MARKER in row[5] for row in retained)


def _reasoning_core(html: str) -> str:
    match = re.search(r"REASONING-CORE-PORT START.*?REASONING-CORE-PORT END", html, re.S)
    assert match, "Explorer is missing its executable reasoning-core markers"
    block = match.group(0)
    return block[block.index("*/") + 2:block.rindex("/*")]


def _reseal(receipt: dict) -> dict:
    receipt["summary"]["baseline_sha256"] = _sha(
        _baseline_digest_payload(receipt)
    )
    return receipt


def _forged_dual_leader_clear(receipt: dict, *, reseal: bool) -> dict:
    forged = copy.deepcopy(receipt)
    for row in forged["rows"]:
        if row["protocol"] != "HSRP":
            continue
        row["status"] = "assessed"
        row["findings"] = []
        row["acceptance"] = _acceptance(row)
    for cell in forged["coverage"]:
        if cell["protocol"] == "HSRP":
            cell["status"] = "assessed"
            cell["finding_codes"] = []
    forged["summary"]["n_review"] = 0
    forged["summary"]["n_assessed"] = 2
    forged["summary"]["by_status"]["review"] = 0
    forged["summary"]["by_status"]["assessed"] = 2
    forged["summary"]["by_coverage_status"]["review"] = 0
    forged["summary"]["by_coverage_status"]["assessed"] = 2
    forged["verdict"] = "CLEAR"
    forged["assessed"] = True
    return _reseal(forged) if reseal else forged


def _rehash_coverage(receipt: dict) -> dict:
    for cell in receipt["coverage"]:
        cell_rows = [
            row for row in receipt["rows"]
            if row["switch"] == cell["switch"] and row["protocol"] == cell["protocol"]
        ]
        (cell["config_sha256"], cell["runtime_sha256"],
         cell["projection_sha256"]) = _coverage_receipt_hashes(cell, cell_rows)
    return receipt


def _valid_one_leader_clear(receipt: dict) -> dict:
    clear = _forged_dual_leader_clear(receipt, reseal=False)
    hsrp_rows = [row for row in clear["rows"] if row["protocol"] == "HSRP"]
    assert len(hsrp_rows) == 2
    hsrp_rows[1]["runtime_state_raw"] = "Standby"
    hsrp_rows[1]["runtime_state"] = "STANDBY"
    hsrp_rows[1]["acceptance"] = _acceptance(hsrp_rows[1])
    return _reseal(_rehash_coverage(clear))


def _port_channel_case_clear(receipt: dict) -> dict:
    forged = _valid_one_leader_clear(receipt)
    hsrp_rows = [row for row in forged["rows"] if row["protocol"] == "HSRP"]
    hsrp_rows[0]["interface"] = "Port-channel1"
    hsrp_rows[1]["interface"] = "PORT-CHANNEL1"
    for row in hsrp_rows:
        row["group_key"] = f"HSRP:Po1:{row['group']}"
        row["acceptance"] = _acceptance(row)
    return _reseal(_rehash_coverage(forged))


def _unicode_switch_casefold_clear(receipt: dict) -> dict:
    forged = _valid_one_leader_clear(receipt)
    hsrp_rows = [row for row in forged["rows"] if row["protocol"] == "HSRP"]
    old_hosts = [row["switch"] for row in hsrp_rows]
    new_hosts = ["edge-ß", "edge-ss"]
    for row, new_host in zip(hsrp_rows, new_hosts):
        row["switch"] = new_host
    second = hsrp_rows[1]
    second.update({
        "interface": "Vlan20",
        "group": "20",
        "group_key": "HSRP:Vlan20:20",
        "configured_vip": "10.0.20.1",
        "runtime_vip": "10.0.20.1",
    })
    second["acceptance"] = _acceptance(second)
    host_map = dict(zip(old_hosts, new_hosts))
    for cell in forged["coverage"]:
        cell["switch"] = host_map[cell["switch"]]
    return _reseal(_rehash_coverage(forged))


def _one_host_interface_alias(receipt: dict) -> dict:
    aliased = copy.deepcopy(receipt)
    hsrp_rows = [row for row in aliased["rows"] if row["protocol"] == "HSRP"]
    assert len(hsrp_rows) == 2
    hsrp_rows[1]["switch"] = hsrp_rows[0]["switch"]
    hsrp_rows[1]["interface"] = hsrp_rows[0]["interface"].lower()
    hsrp_rows[1]["group_key"] = (
        f"HSRP:{hsrp_rows[1]['interface']}:{hsrp_rows[1]['group']}"
    )
    host = hsrp_rows[0]["switch"]
    aliased["coverage"] = [
        cell for cell in aliased["coverage"] if cell["switch"] == host
    ]
    for cell in aliased["coverage"]:
        cell_rows = [
            row for row in aliased["rows"]
            if row["switch"] == cell["switch"] and row["protocol"] == cell["protocol"]
        ]
        cell["config_candidate_count"] = sum(row["configured"] for row in cell_rows)
        cell["configured_group_count"] = cell["config_candidate_count"]
        cell["runtime_candidate_count"] = sum(
            row["runtime_observed"] for row in cell_rows
        )
        cell["runtime_parsed_count"] = cell["runtime_candidate_count"]
        (cell["config_sha256"], cell["runtime_sha256"],
         cell["projection_sha256"]) = _coverage_receipt_hashes(cell, cell_rows)
    aliased["summary"].update({
        "n_hosts": 1,
        "n_coverage_cells": 3,
        "n_subject_hosts": 1,
        "n_subject_cells": 1,
        "by_coverage_status": {
            "degraded": 0,
            "review": 1,
            "not_verified": 0,
            "assessed": 0,
            "not_applicable": 2,
        },
    })
    return _reseal(aliased)


def _switch_case_alias(receipt: dict) -> dict:
    aliased = copy.deepcopy(receipt)
    old_host = aliased["rows"][1]["switch"]
    new_host = aliased["rows"][0]["switch"].upper()
    assert old_host.casefold() != new_host.casefold()
    for row in aliased["rows"]:
        if row["switch"] == old_host:
            row["switch"] = new_host
    for cell in aliased["coverage"]:
        if cell["switch"] == old_host:
            cell["switch"] = new_host
    return _reseal(aliased)


def _explorer_fhrp_results(cases: dict[str, dict]) -> dict:
    assert NODE
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _FHRP_GROUP_SCOPE=');
const end=html.indexOf('const _BGP_INTENT_SCOPE=',start);
if(start<0||end<0)throw new Error('configured FHRP projection block not found');
const cases=JSON.parse(fs.readFileSync(0,'utf8'));
let SNAP={};
function esc(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
const results={};
for(const [name,receipt] of Object.entries(cases)){
  SNAP={fhrp_configured_group_baseline:receipt};
  results[name]={
    accepted:fhrpConfiguredGroupBaselineFrom(SNAP)!==null,
    section:fhrpConfiguredGroupSection(),
  };
}
process.stdout.write(JSON.stringify(results));
"""
    process = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        input=json.dumps(cases),
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=30,
    )
    assert process.returncode == 0, process.stderr[:3000]
    return json.loads(process.stdout)


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer FHRP projection skipped")
def test_explorer_recomputes_dual_leader_truth_and_rejects_browser_current_run_claim(tmp_path):
    current = _dual_leader_baseline(tmp_path / "captures")
    embedded = json.loads(json.dumps(
        embedded_fhrp_configured_group_baseline(current)
    ))
    stale_clear = _forged_dual_leader_clear(embedded, reseal=False)
    resealed_clear = _forged_dual_leader_clear(embedded, reseal=True)
    one_host_alias = _one_host_interface_alias(embedded)
    switch_alias = _switch_case_alias(embedded)

    assert len(stale_clear["summary"]["baseline_sha256"]) == 64
    assert validate_fhrp_configured_group_baseline(stale_clear)["reason"] == (
        "baseline_digest_mismatch"
    )
    assert validate_fhrp_configured_group_baseline(resealed_clear)["reason"] == (
        "baseline_election_reconciliation_mismatch"
    )
    results = _explorer_fhrp_results({
        "embedded": embedded,
        "current_claim": json.loads(json.dumps(current)),
        "stale_clear": stale_clear,
        "resealed_clear": resealed_clear,
        "one_host_alias": one_host_alias,
        "switch_alias": switch_alias,
    })

    assert results["embedded"]["accepted"] is True
    assert "2</div><div class=\"k\">group review" in results["embedded"]["section"]
    assert ISSUE_MARKER in results["embedded"]["section"]
    assert "embedded_unverified" in results["embedded"]["section"]
    for name in (
        "current_claim", "stale_clear", "resealed_clear",
        "one_host_alias", "switch_alias",
    ):
        assert results[name]["accepted"] is False
        assert "not assessed" in results[name]["section"]
        assert "badge b-ok" not in results[name]["section"]
    assert validate_fhrp_configured_group_baseline(one_host_alias)["valid"] is False
    assert validate_fhrp_configured_group_baseline(switch_alias)["valid"] is False


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer FHRP projection skipped")
def test_explorer_requires_exact_election_finding_acceptance_and_coverage_codes(tmp_path):
    embedded = json.loads(json.dumps(embedded_fhrp_configured_group_baseline(
        _dual_leader_baseline(tmp_path / "captures")
    )))
    issue = copy.deepcopy(embedded)
    issue["rows"][0]["findings"][0]["issue"] += " Altered browser claim."
    issue["rows"][0]["acceptance"] = _acceptance(issue["rows"][0])
    _reseal(issue)

    code = copy.deepcopy(embedded)
    code["rows"][0]["findings"][0]["code"] = "election_no_leader_observed"
    code["rows"][0]["acceptance"] = _acceptance(code["rows"][0])
    for cell in code["coverage"]:
        if cell["switch"] == code["rows"][0]["switch"] and cell["protocol"] == "HSRP":
            cell["finding_codes"] = ["election_no_leader_observed"]
    _reseal(code)

    acceptance = copy.deepcopy(embedded)
    acceptance["rows"][0]["acceptance"] += " Forged acceptance."
    _reseal(acceptance)

    finding_codes = copy.deepcopy(embedded)
    for cell in finding_codes["coverage"]:
        if cell["switch"] == finding_codes["rows"][0]["switch"] and cell["protocol"] == "HSRP":
            cell["finding_codes"] = []
    _reseal(finding_codes)

    expected_reasons = {
        "issue": "baseline_election_reconciliation_mismatch",
        "code": "baseline_election_reconciliation_mismatch",
        "acceptance": "baseline_row_acceptance_invalid",
        "finding_codes": "baseline_finding_code_mismatch",
    }
    cases = {
        "issue": issue,
        "code": code,
        "acceptance": acceptance,
        "finding_codes": finding_codes,
    }
    for name, receipt in cases.items():
        assert validate_fhrp_configured_group_baseline(receipt)["reason"] == (
            expected_reasons[name]
        )

    results = _explorer_fhrp_results(cases)
    assert all(result["accepted"] is False for result in results.values())
    assert all("not assessed" in result["section"] for result in results.values())


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer FHRP projection skipped")
def test_explorer_recomputes_receipt_hashes_and_closes_identity_parity_gaps(tmp_path):
    embedded = json.loads(json.dumps(embedded_fhrp_configured_group_baseline(
        _dual_leader_baseline(tmp_path / "captures")
    )))
    valid_clear = _valid_one_leader_clear(embedded)

    port_channel_case = _port_channel_case_clear(embedded)
    unicode_switch_casefold = _unicode_switch_casefold_clear(embedded)

    extra_scope_key = copy.deepcopy(valid_clear)
    extra_scope_key["scope"]["unexpected"] = "forged"
    _reseal(extra_scope_key)

    stale_coverage_hash = copy.deepcopy(valid_clear)
    stale_coverage_hash["rows"][0]["configured_vip"] = "10.0.10.254"
    stale_coverage_hash["rows"][0]["runtime_vip"] = "10.0.10.254"
    stale_coverage_hash["rows"][0]["acceptance"] = _acceptance(
        stale_coverage_hash["rows"][0]
    )
    _reseal(stale_coverage_hash)

    stale_baseline_digest = copy.deepcopy(valid_clear)
    stale_baseline_digest["limitations"][0] += " Semantically benign digest mutation."

    oversized_finding_codes = copy.deepcopy(valid_clear)
    oversized_finding_codes["coverage"][0]["finding_codes"] = ["x"] * 8195
    _reseal(oversized_finding_codes)

    expected_reasons = {
        "port_channel_case": "baseline_duplicate_or_key_invalid",
        "unicode_switch_casefold": "baseline_coverage_identity_invalid",
        "extra_scope_key": "baseline_scope_or_verdict_invalid",
        "stale_coverage_hash": "baseline_coverage_receipt_hash_mismatch",
        "stale_baseline_digest": "baseline_digest_mismatch",
        "oversized_finding_codes": "baseline_coverage_findings_invalid",
    }
    cases = {
        "port_channel_case": port_channel_case,
        "unicode_switch_casefold": unicode_switch_casefold,
        "extra_scope_key": extra_scope_key,
        "stale_coverage_hash": stale_coverage_hash,
        "stale_baseline_digest": stale_baseline_digest,
        "oversized_finding_codes": oversized_finding_codes,
    }
    assert validate_fhrp_configured_group_baseline(valid_clear)["valid"] is True
    for name, receipt in cases.items():
        assert validate_fhrp_configured_group_baseline(receipt)["reason"] == (
            expected_reasons[name]
        )

    results = _explorer_fhrp_results({"valid_clear": valid_clear, **cases})
    assert results["valid_clear"]["accepted"] is True
    assert 'badge b-ok">CLEAR' in results["valid_clear"]["section"]
    for name in cases:
        assert results[name]["accepted"] is False
        assert "not assessed" in results[name]["section"]
        assert "badge b-ok" not in results[name]["section"]


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer FHRP projection skipped")
def test_explorer_canonical_json_aggregate_cap_short_circuits_under_bounded_heap():
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _FHRP_GROUP_SCOPE=');
const end=html.indexOf('const _BGP_INTENT_SCOPE=',start);
if(start<0||end<0)throw new Error('configured FHRP projection block not found');
let SNAP={};
function esc(v){return String(v??'');}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
const leaf='x'.repeat(500);
const hostileRow={findings:Array(64).fill(leaf)};
const hostile=Array(20000).fill(hostileRow);
const started=Date.now();
const canonical=_fhrpCanonicalJson(hostile);
process.stdout.write(JSON.stringify({length:canonical.length,elapsed_ms:Date.now()-started}));
"""
    process = subprocess.run(
        [NODE, "--max-old-space-size=128", "-e", script, str(EXPLORER)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=15,
    )
    assert process.returncode == 0, process.stderr[:3000]
    result = json.loads(process.stdout)
    assert result["length"] == 0
    assert result["elapsed_ms"] < 8_000


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer FHRP projection skipped")
def test_dual_leader_review_is_visible_in_explorer_group_and_waves_caps(tmp_path):
    baseline = _dual_leader_baseline(tmp_path / "captures")
    plan = _capped_plan(baseline)
    html = EXPLORER.read_text(encoding="utf-8")
    payload = tmp_path / "dual-leader-explorer.json"
    payload.write_text(json.dumps({
        "baseline": embedded_fhrp_configured_group_baseline(baseline),
        "plan": plan,
    }), encoding="utf-8")
    driver = tmp_path / "dual-leader-explorer.js"
    fhrp = html[
        html.index("const _FHRP_GROUP_SCOPE="):
        html.index("const _BGP_INTENT_SCOPE=", html.index("const _FHRP_GROUP_SCOPE="))
    ]
    workflow = html[
        html.index("function validationFor(wave)"):
        html.index("function drawWaves()")
    ]
    driver.write_text(
        "const fs=require('fs');"
        "const input=JSON.parse(fs.readFileSync(process.argv[2],'utf8'));"
        "let SNAP={fhrp_configured_group_baseline:input.baseline,validation_plan:input.plan};"
        "function esc(v){return String(v??'').replace(/[&<>\"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',\"'\":'&#39;'}[c]));}"
        "function shortName(v){return String(v??'');}\n"
        + fhrp
        + "\nconst configured=fhrpConfiguredGroupSection();\n"
        + _reasoning_core(html)
        + "\n"
        + workflow
        + "\nconst waveData=[{idx:0,g:{switches:['edge-a','edge-b']}}];"
          "const state=currentBaselineWorkflowState(waveData);"
          "process.stdout.write(JSON.stringify({configured,verdict:state.gate.verdict,"
          "panel:currentBaselineWorkflowPanel(state),validation:wavesValidationBlock('Group 1')}));\n",
        encoding="utf-8",
    )
    process = subprocess.run(
        [NODE, str(driver), str(payload)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=60,
    )
    assert process.returncode == 0, process.stderr[:3000]
    actual = json.loads(process.stdout)

    configured = actual["configured"]
    assert "2</div><div class=\"k\">group review" in configured
    assert all(host in configured for host in HOSTS)
    assert ISSUE_MARKER in configured
    assert "election_multiple_leaders_observed" in configured
    assert "Capture timing is not simultaneous evidence" in configured

    assert actual["verdict"] == "INDETERMINATE"
    assert "HOLD — current baseline INDETERMINATE" in actual["panel"]
    rendered = actual["validation"]
    assert all(host in rendered for host in HOSTS)
    assert rendered.count("<b>evidence</b> review") == 2
    assert ISSUE_MARKER in rendered
    assert "current_run_source_bound" in rendered
    assert "ordinary validation 00" in rendered and "ordinary validation 02" in rendered
    assert "ordinary validation 03" not in rendered


def test_dual_leader_review_forces_assesshub_no_go_without_losing_evidence(tmp_path):
    from webapp.backend import cutover

    baseline = _dual_leader_baseline(tmp_path / "captures")
    plan = cutover.build_plan(_snapshot(baseline, _capped_plan(baseline)))

    assert plan["summary"]["verdict"] == cutover.GATE_NOGO
    assert plan["summary"]["current_baseline"]["verdict"] == cutover.BASELINE_INDETERMINATE
    assert plan["summary"]["n_baseline_blockers"] == 2
    wave, = plan["waves"]
    assert wave["gate"] == cutover.GATE_NOGO
    assert wave["current_baseline"]["verdict"] == cutover.BASELINE_INDETERMINATE
    assert len(wave["baseline_blockers"]) == 2
    assert {row["device"] for row in wave["baseline_blockers"]} == set(HOSTS)
    assert all(row["baseline_state"] == "review" for row in wave["baseline_blockers"])
    assert all(row["projection_custody"] == "current_run_source_bound"
               for row in wave["baseline_blockers"])
    assert all(ISSUE_MARKER in row["expect"] for row in wave["baseline_blockers"])
    assert sum(row["baseline_blocker"] for row in wave["validation"]) == 2
