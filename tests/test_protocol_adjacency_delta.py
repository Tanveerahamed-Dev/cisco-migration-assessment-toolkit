"""Receipt-gated protocol-adjacency deltas for the cutover compare surface.

The pre-change neighbour table is a bounded, observed baseline.  It is not a
configured/expected-peer inventory, and the embedded ``routing_neighbors``
projection is not independently bound to the raw command capture.  These tests
therefore distinguish a definitive state regression from a peer that is merely
no longer observed, and distinguish both from lost assessability.
"""

from copy import deepcopy

from openpyxl import load_workbook

from cisco_toolkit.html import (
    compute_campaign_trend,
    compute_protocol_adjacency_delta,
    compute_snapshot_delta,
    write_campaign_workbook,
    write_diff_workbook,
)


FAMILIES = ("STP", "EtherChannel", "VTP", "OSPF", "BGP", "EIGRP", "FHRP")
ROUTING = {"OSPF": "ospf", "BGP": "bgp", "EIGRP": "eigrp"}
INPUT_IDS = {
    "STP": ("state", "blocked_ports", "inconsistent_ports", "topology_changes"),
    "EtherChannel": ("membership",),
    "VTP": ("status",),
    "OSPF": ("neighbors",),
    "BGP": ("peers",),
    "EIGRP": ("neighbors",),
    "FHRP": ("hsrp_groups", "vrrp_groups", "glbp_groups"),
}


def _capture_state(state: str) -> str:
    return {
        "assessed": "usable",
        "partial": "usable",
        "captured_no_record": "usable",
        "captured_empty": "empty",
        "capture_error": "error",
        "not_collected": "missing",
        "analysis_unavailable": "missing",
    }[state]


def _receipt(host: str, states: dict[str, str]) -> dict:
    """A complete one-device x seven-family ``protocol_assessability/1`` receipt."""
    rows = []
    by_state = {
        state: 0
        for state in (
            "assessed",
            "partial",
            "captured_no_record",
            "captured_empty",
            "capture_error",
            "not_collected",
            "analysis_unavailable",
        )
    }
    for family in FAMILIES:
        state = states.get(family, "not_collected")
        capture = _capture_state(state)
        by_state[state] += 1
        rows.append(
            {
                "switch": host,
                "protocol": family,
                "input_states": {name: capture for name in INPUT_IDS[family]},
                "capture_state": capture,
                "health_row_emitted": state == "assessed",
                "state": state,
                "reason": f"fixture: {state}",
            }
        )
    return {
        "schema": "protocol_assessability/1",
        "families": [{"protocol": family} for family in FAMILIES],
        "rows": rows,
        "summary": {
            "n_devices": 1,
            "n_families": 7,
            "n_cells": 7,
            "n_health_rows": sum(row["health_row_emitted"] for row in rows),
            "n_complete_devices": 0,
            "by_state": by_state,
        },
        "limitations": ["fixture receipt"],
    }


def _snap(
    *,
    ospf: list[dict] | None = None,
    bgp: list[dict] | None = None,
    eigrp: list[dict] | None = None,
    states: dict[str, str] | None = None,
    host: str = "core1",
) -> dict:
    neighbours = {
        "ospf": list(ospf or []),
        "bgp": list(bgp or []),
        "eigrp": list(eigrp or []),
    }
    inferred = {
        family: "assessed"
        for family, key in ROUTING.items()
        if neighbours[key]
    }
    inferred.update(states or {})
    return {
        "devices": {host: {}},
        "interfaces": {host: {}},
        "health_scores": [{"switch": host, "band": "Good", "score": 80}],
        "punchlist": [],
        "routing_neighbors": {host: neighbours},
        "protocol_assessability": _receipt(host, inferred),
    }


def _ospf(peer: str, state: str = "FULL/DR", interface: str = "Po1") -> dict:
    return {
        "neighbor": peer,
        "state": state,
        "address": peer,
        "interface": interface,
    }


def _bgp(peer: str, state: str = "12", remote_as: str = "65002") -> dict:
    return {"neighbor": peer, "as": remote_as, "state": state}


def _eigrp(peer: str, state: str = "up 12", interface: str = "Gi0/1") -> dict:
    return {"neighbor": peer, "interface": interface, "state": state}


def _change(result: dict, disposition: str) -> dict:
    return next(row for row in result["changes"] if row["result"] == disposition)


def test_assessed_unchanged_peer_is_preserved_and_custody_is_explicit():
    snap = _snap(ospf=[_ospf("10.0.0.2")])

    delta = compute_protocol_adjacency_delta(snap, deepcopy(snap))

    assert delta["schema"] == "protocol_adjacency_delta/1"
    assert delta["scope"] == "baseline_observed"
    assert delta["projection_custody"] == "embedded_unverified"
    assert delta["assessed"] is True and delta["gate"] == "PASS"
    assert delta["summary"] == {
        "n_baseline_peers": 1,
        "n_scoped_cells": 1,
        "n_comparable_cells": 1,
        "n_preserved": 1,
        "n_state_regressed": 0,
        "n_recovered": 0,
        "n_no_longer_observed": 0,
        "n_added": 0,
        "n_metadata_changed": 0,
        "n_coverage_gaps": 0,
    }
    assert delta["changes"] == [] and delta["coverage_gaps"] == []

    bound = compute_protocol_adjacency_delta(
        snap,
        deepcopy(snap),
        source_binding={
            "before": "sha256:" + "a" * 64,
            "after": "sha256:" + "b" * 64,
        },
    )
    assert bound["projection_custody"] == "source_bound_embedded_unverified"
    assert any("embedded" in str(item).lower() for item in bound["limitations"])


def test_ospf_full_to_exstart_is_a_definitive_regression():
    before = _snap(ospf=[_ospf("10.0.0.2", "FULL/DR")])
    after = _snap(ospf=[_ospf("10.0.0.2", "EXSTART/DROTHER")])

    protocol = compute_protocol_adjacency_delta(before, after)
    change = _change(protocol, "state_degraded")

    assert protocol["gate"] == "REGRESSED"
    assert protocol["summary"]["n_state_regressed"] == 1
    assert (change["switch"], change["protocol"], change["peer"]) == (
        "core1",
        "OSPF",
        "10.0.0.2",
    )
    assert change["before_state"] == "FULL/DR"
    assert change["after_state"] == "EXSTART/DROTHER"
    assert compute_snapshot_delta(before, after)["verdict"] == "REGRESSED"


def test_bgp_prefix_count_churn_is_not_session_churn_but_idle_is_regressed():
    before = _snap(bgp=[_bgp("192.0.2.2", "12")])
    prefix_churn = _snap(bgp=[_bgp("192.0.2.2", "37")])

    unchanged = compute_protocol_adjacency_delta(before, prefix_churn)
    assert unchanged["gate"] == "PASS"
    assert unchanged["summary"]["n_preserved"] == 1
    assert unchanged["summary"]["n_state_regressed"] == 0
    assert unchanged["changes"] == []

    idle = _snap(bgp=[_bgp("192.0.2.2", "Idle")])
    regressed = compute_protocol_adjacency_delta(before, idle)
    change = _change(regressed, "state_degraded")
    assert regressed["gate"] == "REGRESSED"
    assert change["before_state"] == "12" and change["after_state"] == "Idle"
    assert compute_snapshot_delta(before, idle)["verdict"] == "REGRESSED"


def test_equivalent_ipv6_bgp_peer_spellings_preserve_one_identity():
    before = _snap(bgp=[_bgp("2001:db8::1", "12")])
    after = _snap(bgp=[_bgp("2001:0db8:0000:0000:0000:0000:0000:0001", "37")])

    delta = compute_protocol_adjacency_delta(before, after)

    assert delta["gate"] == "PASS"
    assert delta["summary"]["n_baseline_peers"] == 1
    assert delta["summary"]["n_preserved"] == 1
    assert delta["summary"]["n_no_longer_observed"] == 0
    assert delta["summary"]["n_added"] == 0
    assert delta["changes"] == []


def test_eigrp_uptime_counter_churn_is_ignored():
    before = _snap(eigrp=[_eigrp("10.0.0.3", "up 12")])
    after = _snap(eigrp=[_eigrp("10.0.0.3", "up 7")])

    delta = compute_protocol_adjacency_delta(before, after)

    assert delta["gate"] == "PASS"
    assert delta["summary"]["n_preserved"] == 1
    assert delta["changes"] == []


def test_eigrp_healthy_state_grammar_is_anchored_and_unknown_is_reviewed():
    before = _snap(eigrp=[_eigrp("10.0.0.3", "up 12")])
    malformed = _snap(eigrp=[_eigrp("10.0.0.3", "upside-down")])

    delta = compute_protocol_adjacency_delta(before, malformed)

    change = _change(delta, "state_changed")
    assert delta["gate"] == "REVIEW"
    assert change["before_state"] == "up 12"
    assert change["after_state"] == "upside-down"


def test_peer_identity_is_host_protocol_and_peer_not_mutable_metadata():
    before = _snap(ospf=[_ospf("10.0.0.2", interface="Po1")])
    after = _snap(ospf=[_ospf("10.0.0.2", interface="Po2")])

    delta = compute_protocol_adjacency_delta(before, after)
    change = _change(delta, "metadata_changed")

    assert delta["gate"] == "REVIEW"
    assert delta["summary"]["n_metadata_changed"] == 1
    assert delta["summary"]["n_preserved"] == 1
    assert delta["summary"]["n_added"] == 0
    assert delta["summary"]["n_no_longer_observed"] == 0
    assert (change["switch"], change["protocol"], change["peer"]) == (
        "core1",
        "OSPF",
        "10.0.0.2",
    )


def test_one_of_multiple_assessed_peers_no_longer_observed_requires_review():
    before = _snap(ospf=[_ospf("10.0.0.2"), _ospf("10.0.0.3", interface="Po2")])
    after = _snap(ospf=[_ospf("10.0.0.2")])

    protocol = compute_protocol_adjacency_delta(before, after)
    change = _change(protocol, "no_longer_observed")

    assert protocol["gate"] == "REVIEW"
    assert protocol["summary"]["n_no_longer_observed"] == 1
    assert protocol["summary"]["n_state_regressed"] == 0
    assert change["peer"] == "10.0.0.3"
    assert "down" not in change["note"].lower()
    assert compute_snapshot_delta(before, after)["verdict"] == "REVIEW"


def test_last_peer_to_captured_no_record_is_coverage_loss_not_peer_regression():
    before = _snap(ospf=[_ospf("10.0.0.2")])
    after = _snap(ospf=[], states={"OSPF": "captured_no_record"})

    protocol = compute_protocol_adjacency_delta(before, after)

    assert protocol["gate"] == "REVIEW"
    assert protocol["summary"]["n_state_regressed"] == 0
    assert protocol["summary"]["n_no_longer_observed"] == 0
    assert protocol["summary"]["n_coverage_gaps"] == 1
    assert protocol["changes"] == []
    assert protocol["coverage_gaps"] == [
        {
            "switch": "core1",
            "protocol": "OSPF",
            "before_state": "assessed",
            "after_state": "captured_no_record",
            "reason": protocol["coverage_gaps"][0]["reason"],
        }
    ]
    reason = protocol["coverage_gaps"][0]["reason"]
    assert "after receipt state is captured_no_record" in reason
    assert "zero peers" not in reason
    outer = compute_snapshot_delta(before, after)
    assert outer["verdict"] == "REVIEW"
    assert outer["protocol_adjacencies"]["gate"] == "REVIEW"


def test_assessed_receipt_without_emitted_health_row_is_invalid():
    before = _snap(ospf=[_ospf("10.0.0.2")])
    after = deepcopy(before)
    ospf_receipt = next(
        row for row in after["protocol_assessability"]["rows"] if row["protocol"] == "OSPF"
    )
    ospf_receipt["health_row_emitted"] = False
    after["protocol_assessability"]["summary"]["n_health_rows"] -= 1

    protocol = compute_protocol_adjacency_delta(before, after)

    assert protocol["gate"] == "REVIEW"
    assert protocol["assessed"] is False
    assert protocol["summary"]["n_comparable_cells"] == 0
    assert protocol["summary"]["n_coverage_gaps"] == 1
    assert "marks a cell assessed without an emitted health row" in protocol["coverage_gaps"][0]["reason"]
    assert compute_snapshot_delta(before, after)["verdict"] == "REVIEW"


def test_assessed_emitted_cell_with_trimmed_peer_projection_is_a_coverage_gap():
    observed = _snap(ospf=[_ospf("10.0.0.2")])
    trimmed = deepcopy(observed)
    trimmed["routing_neighbors"]["core1"]["ospf"] = []

    one_sided = compute_protocol_adjacency_delta(observed, trimmed)
    assert one_sided["gate"] == "REVIEW"
    assert one_sided["assessed"] is False
    assert one_sided["summary"]["n_no_longer_observed"] == 0
    assert one_sided["summary"]["n_coverage_gaps"] == 1
    assert one_sided["changes"] == []
    assert "after receipt is assessed with an emitted health row" in one_sided["coverage_gaps"][0]["reason"]
    assert "zero peers" in one_sided["coverage_gaps"][0]["reason"]

    both_trimmed = compute_protocol_adjacency_delta(trimmed, deepcopy(trimmed))
    assert both_trimmed["gate"] == "REVIEW"
    assert both_trimmed["assessed"] is False
    assert both_trimmed["summary"]["n_baseline_peers"] == 0
    assert both_trimmed["summary"]["n_scoped_cells"] == 1
    assert both_trimmed["summary"]["n_comparable_cells"] == 0
    assert both_trimmed["summary"]["n_coverage_gaps"] == 1
    assert "before receipt is assessed with an emitted health row" in both_trimmed["coverage_gaps"][0]["reason"]
    assert "after receipt is assessed with an emitted health row" in both_trimmed["coverage_gaps"][0]["reason"]


def test_down_peer_recovery_is_reported_without_a_regression_gate():
    before = _snap(ospf=[_ospf("10.0.0.2", "EXSTART/DROTHER")])
    after = _snap(ospf=[_ospf("10.0.0.2", "FULL/DR")])

    protocol = compute_protocol_adjacency_delta(before, after)
    change = _change(protocol, "recovered")

    assert protocol["gate"] == "PASS"
    assert protocol["summary"]["n_recovered"] == 1
    assert protocol["summary"]["n_state_regressed"] == 0
    assert change["before_state"] == "EXSTART/DROTHER"
    assert change["after_state"] == "FULL/DR"
    assert compute_snapshot_delta(before, after)["verdict"] == "CLEAN"


def test_legacy_receipts_are_total_and_do_not_poison_other_axes():
    legacy = _snap(ospf=[_ospf("10.0.0.2")])
    legacy.pop("protocol_assessability")

    unavailable = compute_protocol_adjacency_delta(legacy, deepcopy(legacy))
    assert unavailable["assessed"] is False
    assert unavailable["gate"] == "NOT_ASSESSED"
    assert unavailable["changes"] == []
    assert compute_snapshot_delta(legacy, deepcopy(legacy))["verdict"] == "CLEAN"


def test_unassessed_workbook_uses_abstention_not_zero(tmp_path):
    legacy = _snap(ospf=[_ospf("10.0.0.2")])
    legacy.pop("protocol_assessability")
    out = tmp_path / "legacy-protocol-delta.xlsx"
    write_diff_workbook(
        legacy,
        deepcopy(legacy),
        str(out),
        precert={
            "verdict": "PASS", "verdict_note": "fixture certificate", "flows": {},
            "stamps": {}, "segmentation": [], "intents": [], "blind_spots": [],
            "gate_failures": [],
        },
    )
    wb = load_workbook(out, read_only=True)
    try:
        summary_rows = list(wb["Summary"].iter_rows(values_only=True))
        regressions = next(
            row for row in summary_rows if row[0] == "Protocol adjacency state regressions"
        )
        gaps = next(row for row in summary_rows if row[0] == "Protocol adjacency coverage gaps")
        assert regressions[2] == "—"
        assert gaps[2] == 1
        detail = "\n".join(
            str(value)
            for row in wb["Protocol Adjacency Delta"].iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        assert "NOT COMPARABLE" in detail
        assert "receipt is missing" in detail
    finally:
        wb.close()


def test_malformed_receipt_is_total_and_fails_closed_to_review():
    legacy = _snap(ospf=[_ospf("10.0.0.2")])
    legacy.pop("protocol_assessability")
    malformed = deepcopy(legacy)
    malformed["protocol_assessability"] = {
        "schema": "protocol_assessability/1",
        "rows": "not-a-row-list",
    }
    bad = compute_protocol_adjacency_delta(malformed, malformed)
    assert bad["assessed"] is False and bad["gate"] == "REVIEW"
    assert bad["summary"]["n_state_regressed"] == 0
    assert bad["summary"]["n_coverage_gaps"] == 1
    assert compute_snapshot_delta(malformed, malformed)["verdict"] == "REVIEW"


def test_diff_workbook_has_protocol_adjacency_sheet_summary_and_gate(tmp_path):
    before = _snap(ospf=[_ospf("10.0.0.2", "FULL/DR")])
    after = _snap(ospf=[_ospf("10.0.0.2", "EXSTART/DROTHER")])
    out = tmp_path / "protocol-delta.xlsx"
    write_diff_workbook(
        before,
        after,
        str(out),
        precert={
            "verdict": "PASS",
            "verdict_note": "fixture certificate",
            "flows": {},
            "stamps": {},
            "segmentation": [],
            "intents": [],
            "blind_spots": [],
            "gate_failures": [],
        },
    )

    wb = load_workbook(out, read_only=True)
    try:
        assert "Protocol Adjacency Delta" in wb.sheetnames
        protocol_values = [
            str(value)
            for row in wb["Protocol Adjacency Delta"].iter_rows(values_only=True)
            for value in row
            if value is not None
        ]
        assert "state_degraded" in protocol_values
        assert "10.0.0.2" in protocol_values

        summary_rows = list(wb["Summary"].iter_rows(values_only=True))
        gate = next(row for row in summary_rows if row[0] == "CUTOVER GATE VERDICT")
        assert gate[2] == "REGRESSED"
        protocol_metric = next(
            row for row in summary_rows if row[0] == "Protocol adjacency state regressions"
        )
        assert protocol_metric[2] == 1
        coverage_metric = next(
            row for row in summary_rows if row[0] == "Protocol adjacency coverage gaps"
        )
        no_longer_observed_metric = next(
            row
            for row in summary_rows
            if row[0] == "Protocol adjacencies no longer observed"
        )
        assert coverage_metric[2] == 0
        assert no_longer_observed_metric[2] == 0
    finally:
        wb.close()


def test_campaign_endpoint_gate_regresses_but_a_transient_recovered_peer_does_not():
    before = _snap(ospf=[_ospf("10.0.0.2", "FULL/DR")])
    degraded = _snap(ospf=[_ospf("10.0.0.2", "EXSTART/DROTHER")])
    recovered = _snap(ospf=[_ospf("10.0.0.2", "FULL/BDR")])
    for index, snap in enumerate((before, degraded, recovered), start=1):
        snap.update({
            "generated_at": f"2026-08-0{index}T00:00:00",
            "migration_readiness": [],
            "lifecycle_risk": {"summary": {}},
        })

    endpoint_regression = compute_campaign_trend([before, degraded])
    assert endpoint_regression["verdict"] == "REGRESSING"
    assert endpoint_regression["protocol_adjacencies"]["gate"] == "REGRESSED"
    assert endpoint_regression["steps"][0]["protocol_state_regressed"] == 1
    assert endpoint_regression["steps"][0]["protocol_projection_custody"] == "embedded_unverified"

    transient = compute_campaign_trend([before, degraded, recovered])
    assert transient["protocol_adjacencies"]["gate"] == "PASS"
    assert transient["verdict"] != "REGRESSING"
    assert [step["protocol_gate"] for step in transient["steps"]] == ["REGRESSED", "PASS"]
    assert transient["steps"][1]["protocol_recovered"] == 1


def test_campaign_workbook_threads_source_bound_projection_custody(tmp_path):
    before = _snap(bgp=[_bgp("2001:db8::1", "12")])
    after = _snap(bgp=[_bgp("2001:db8::1", "Idle")])
    for index, snap in enumerate((before, after), start=1):
        snap.update({
            "generated_at": f"2026-08-1{index}T00:00:00",
            "migration_readiness": [],
            "lifecycle_risk": {"summary": {}},
        })
    bindings = ["sha256:" + "a" * 64, "sha256:" + "b" * 64]

    trend = compute_campaign_trend([before, after], source_bindings=bindings)
    assert trend["steps"][0]["protocol_projection_custody"] == "source_bound_embedded_unverified"

    out = tmp_path / "protocol-campaign.xlsx"
    write_campaign_workbook([before, after], str(out), source_bindings=bindings)
    wb = load_workbook(out, read_only=True)
    try:
        rows = list(wb["Protocol Adjacencies"].iter_rows(values_only=True))
        assert rows[0] == (
            "Step",
            "Gate",
            "Baseline peers",
            "State regressed",
            "Recovered",
            "No longer observed",
            "Added",
            "Coverage gaps",
            "Projection custody",
            "Scope / next action",
        )
        assert rows[1][1] == "REGRESSED"
        assert rows[1][8] == "source_bound_embedded_unverified"
        assert "Protocol adjacency gate REGRESSED" in rows[1][9]
    finally:
        wb.close()
