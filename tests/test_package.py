"""PHASE 2.7 step 1: the cisco_toolkit package extraction.

Guards the extraction boundary - the leaf text helpers live in
cisco_toolkit.textutils and the monolith re-exports the *same* objects (so a
future accidental re-definition in the monolith would break the identity check).
"""


def test_textutils_importable_standalone():
    from cisco_toolkit.textutils import normalize_ifname, is_valid_iface, normalize_mac
    assert normalize_ifname("GigabitEthernet1/0/1") == "Gi1/0/1"
    assert is_valid_iface("Gi1/0/1") is True
    assert is_valid_iface("garbage") is False
    assert normalize_mac("00:11:22:33:44:55") == "0011.2233.4455"
    from cisco_toolkit.textutils import _split_macs
    assert _split_macs("aaaa.bbbb.cccc, dddd.eeee.ffff") == ["aaaa.bbbb.cccc", "dddd.eeee.ffff"]
    assert _split_macs("") == []


def test_monolith_reexports_the_package_objects(cp):
    from cisco_toolkit import textutils
    # normalize_ifname / detect_link_type went package-internal in step 28 (build_interfaces,
    # their last monolith user, moved to cisco_toolkit.build); both still live in textutils.
    assert callable(textutils.normalize_ifname) and not hasattr(cp, "normalize_ifname")
    assert callable(textutils.detect_link_type) and not hasattr(cp, "detect_link_type")
    # is_valid_iface is now used only inside the package (its monolith callers moved out)
    assert callable(textutils.is_valid_iface)
    # PHYSICAL_IFACE_RE went package-internal in step 27 (its last monolith user,
    # build_device_physical, moved to cisco_toolkit.build); IFACE_TOKEN_RE / VALID_IFACE_RE
    # went package-internal in step 17 (their last monolith users, the phy parsers, moved to parse).
    assert textutils.PHYSICAL_IFACE_RE.match("Gi1/0/1") and not hasattr(cp, "PHYSICAL_IFACE_RE")
    assert not hasattr(cp, "VALID_IFACE_RE") and not hasattr(cp, "IFACE_TOKEN_RE")
    # _split_macs went package-internal in step 21 (its last monolith user, write_vlan_census_sheet,
    # moved to excel; analyze.py + excel.py import it from textutils directly).
    assert callable(textutils._split_macs) and not hasattr(cp, "_split_macs")


def test_parse_module_reexported_and_functional(cp):
    from cisco_toolkit import parse
    # primitives live in the package; the monolith no longer references them
    # directly now that the table parsers moved out (so it doesn't import them).
    assert callable(parse.extract_fixed_cols) and callable(parse.slice_col)
    # the monolith no longer re-exports ANY parser (step 28: build_interfaces, the last parser
    # consumer, moved to cisco_toolkit.build). parse_ospf_neighbors / parse_bgp_summary were
    # already gone (step 24); parse_show_interface_status went with build_interfaces (step 28).
    for name in ("parse_ospf_neighbors", "parse_bgp_summary", "parse_show_interface_status"):
        assert hasattr(parse, name) and not hasattr(cp, name)
    # functional smoke straight from the package
    rows = parse.parse_ospf_neighbors("10.0.0.2  1  FULL/DR  00:00:35  10.0.0.2  Gi0/1")
    assert rows and rows[0]["state"] == "FULL/DR" and rows[0]["interface"] == "Gi0/1"
    # _parse_fhrp (step 16), the physical-port parsers (step 17), and parse_show_interface_counters
    # are no longer re-exported by the monolith (step 24/25: their writers moved to excel); still in parse.
    for name in ("_parse_fhrp", "_is_physical_port", "_classify_media", "parse_interface_phy",
                 "_parse_poe_watts", "parse_show_interface_counters"):
        assert hasattr(parse, name) and not hasattr(cp, name)
    assert parse._parse_fhrp("HSRP grp 1 Active VIP 10.0.10.1") == ("HSRP", "Active", "10.0.10.1", "1")
    assert parse._is_physical_port("Gi1/0/1") is True and parse._is_physical_port("Vlan10") is False
    assert parse._classify_media("media type is 10/100/1000BaseTX") == "copper"
    phy = parse.parse_interface_phy(
        "GigabitEthernet1/0/1 is up\n  Full-duplex, 1000Mb/s, media type is 10/100/1000BaseTX\n")
    assert phy["Gi1/0/1"]["duplex"] == "Full" and phy["Gi1/0/1"]["media"] == "copper"


def test_model_reexported_and_functional(cp):
    from cisco_toolkit import model
    # identity, not equality: the monolith must construct the package's classes,
    # not a re-defined copy (every layer passes these records around).
    assert cp.InterfaceData is model.InterfaceData
    assert cp.DevicePhysical is model.DevicePhysical
    # functional smoke straight from the package: fields + defaults intact.
    iface = model.InterfaceData(port="Gi1/0/1", status="connected")
    assert iface.port == "Gi1/0/1" and iface.status == "connected"
    assert iface.vlan == "" and iface.neighbor_platform == ""
    dp = model.DevicePhysical(hostname="core1")
    assert dp.hostname == "core1" and dp.num_power_supplies == 0


def test_analyze_reexported_and_functional(cp):
    from cisco_toolkit import analyze
    # identity, not equality: the compute_* functions still in the monolith must use
    # the package's ScoringConfig/SCORING/helpers, not a re-defined copy.
    # ScoringConfig / SCORING / _host_role went package-internal in step 15; _health_band in step 23
    # (its last monolith user, write_health_scores_sheet, moved to excel). All still live in analyze.
    assert callable(analyze.ScoringConfig) and analyze.SCORING is not None and callable(analyze._host_role)
    for gone in ("ScoringConfig", "SCORING", "_host_role", "_health_band"):
        assert not hasattr(cp, gone)
    # the default config reproduces the documented hard-coded tunables.
    assert analyze.SCORING.caps == {"L1": 30, "L3": 30, "XL": 45, "PROTO": 25, "SEC": 18}
    assert analyze.SCORING.l1_weights["single-fiber-uplink"] == 10
    # band thresholds + the gateway-SVI role heuristic, straight from the package.
    assert analyze._health_band(95) == ("Excellent", "36E08A")
    assert analyze._health_band(50) == ("Poor", "FF9F45")
    gw = analyze.InterfaceData(port="Vlan10", svi_ip="10.0.0.1")
    assert analyze._host_role({"Vlan10": gw}) == "distribution"
    assert analyze._host_role({"Gi1/0/1": analyze.InterfaceData(port="Gi1/0/1")}) == "access"
    # move-group computation joined the analyze layer (step 11).
    assert cp.compute_move_groups is analyze.compute_move_groups
    # MOVEGROUP_EXCLUDED_VLANS went package-internal in step 12: its last monolith user
    # (compute_findings) moved out, so the monolith no longer re-exports it.
    assert analyze.MOVEGROUP_EXCLUDED_VLANS == {1}
    assert not hasattr(cp, "MOVEGROUP_EXCLUDED_VLANS")
    # two switches sharing VLAN 20 (via access ports) collapse into one move group.
    ID = analyze.InterfaceData
    ifaces = {
        "sw1": {"Gi1/0/1": ID(port="Gi1/0/1", switchport_mode="Access", vlan="20")},
        "sw2": {"Gi1/0/1": ID(port="Gi1/0/1", switchport_mode="Access", vlan="20")},
    }
    groups = analyze.compute_move_groups(ifaces)
    assert len(groups) == 1 and groups[0]["switches"] == ["sw1", "sw2"]
    # topology-links + findings joined analyze (step 12); their monolith re-exports were dropped
    # in step 22 when write_topology_sheet / write_findings_sheet moved to excel (excel imports the
    # computes directly). They still live in analyze.
    for name in ("compute_topology_links", "compute_findings"):
        assert hasattr(analyze, name) and not hasattr(cp, name)
    # _canon_host / _canon_host_map went package-internal in step 13 (build_network_model,
    # their last monolith user, moved out, so the monolith no longer re-exports them).
    assert callable(analyze._canon_host) and callable(analyze._canon_host_map)
    assert not hasattr(cp, "_canon_host") and not hasattr(cp, "_canon_host_map")
    assert analyze._canon_host("Switch1.example.com (FOC123)") == "switch1"
    # two switches that see each other over CDP -> one link confirmed from both ends.
    tl = {
        "sw1": {"Gi1/0/1": ID(port="Gi1/0/1", cdp_neighbor="sw2", neighbor_port="Gi1/0/1", endpoint_type="Switch")},
        "sw2": {"Gi1/0/1": ID(port="Gi1/0/1", cdp_neighbor="sw1", neighbor_port="Gi1/0/1", endpoint_type="Switch")},
    }
    links = analyze.compute_topology_links(tl)
    assert len(links) == 1 and links[0]["confirmation"] == "Both ends"
    # FQDN regression: a scanned host advertised over CDP by its domain-qualified id must
    # resolve to its short scanned hostname in b_host - otherwise the diagram / Topology
    # Links sheet split it into a duplicate node (real-world bug: '*.broadcast.[HISTORY-REDACTED]' ids).
    fq = {
        "core1": {"Te1/1/1": ID(port="Te1/1/1", cdp_neighbor="acc1.example.com",
                                neighbor_port="Te1/49", endpoint_type="Switch")},
        "acc1":  {"Te1/49": ID(port="Te1/49", cdp_neighbor="core1.example.com",
                               neighbor_port="Te1/1/1", endpoint_type="Switch")},
    }
    fql = analyze.compute_topology_links(fq)
    assert len(fql) == 1
    assert {fql[0]["a_host"], fql[0]["b_host"]} == {"core1", "acc1"}   # both short, no FQDN node
    # two SVIs for VLAN 20 with no FHRP -> a High "Gateway redundancy" finding.
    fi = {"sw1": {"Vlan20": ID(port="Vlan20", svi_ip="10.0.20.1")}, "sw2": {"Vlan20": ID(port="Vlan20", svi_ip="10.0.20.2")}}
    assert any(sev == "High" and cat == "Gateway redundancy"
               for (sev, cat, scope, detail) in analyze.compute_findings(fi))
    # network-model / blast-radius cluster joined the analyze layer (step 13); their monolith
    # re-exports were dropped as the writers moved to excel (build_network_model step 25;
    # compute_causality_chains / compute_failure_impact step 24). They still live in analyze.
    for name in ("build_network_model", "compute_causality_chains", "compute_failure_impact"):
        assert hasattr(analyze, name) and not hasattr(cp, name)
    # _vlan_components (step 18) + _link_carries (step 19) went package-internal as their
    # last monolith users (build_dependency_map / _bfs_forwarding_path) moved out.
    for gone in ("_vlan_components", "_link_carries"):
        assert hasattr(analyze, gone) and not hasattr(cp, gone)
    nm = {
        "sw1": {"Vlan20": ID(port="Vlan20", svi_ip="10.0.20.1")},   # sole gateway for VLAN 20, no FHRP
        "sw2": {"Gi1/0/1": ID(port="Gi1/0/1", switchport_mode="Access", vlan="20",
                              end_host_mac="aaaa.bbbb.cccc")},
    }
    model = analyze.build_network_model(nm)
    assert model["hosts"] == ["sw1", "sw2"] and 20 in model["vlans"]
    assert model["access_presence"].get(20) == {"sw2"}
    # removing the sole gateway (sw1) hard-partitions VLAN 20's endpoints -> High severity.
    impact = {r["host"]: r for r in analyze.compute_failure_impact(nm)}
    assert impact["sw1"]["severity"] == "High"
    # scoring + readiness synthesis joined the analyze layer (step 15).
    for name in ("compute_data_quality", "compute_health_scores",
                 "compute_score_sensitivity", "compute_migration_readiness"):
        assert getattr(cp, name) is getattr(analyze, name)
    # empty inputs -> a perfect, deduction-free score.
    hs = analyze.compute_health_scores({"sw1": {}}, [], [], [], [])
    assert hs == [{"switch": "sw1", "score": 100, "band": "Excellent",
                   "role": "access", "criticality": 1.0, "deductions": []}]
    # data-quality fraction: no collected files -> 0.0; no hosts -> empty.
    assert analyze.compute_data_quality({}) == {}
    assert analyze.compute_data_quality({"sw1": {}}) == {"sw1": 0.0}
    # protocol-health joined the analyze layer (step 16); its STP/EC/VTP sub-parsers are
    # package-internal, and _parse_fhrp moved to parse.py (shared with the L3-fwd sheet).
    assert cp.compute_protocol_health is analyze.compute_protocol_health
    for internal in ("_parse_stp_mode", "_parse_stp_tcn", "_parse_etherchannel_member_states", "_parse_vtp_full"):
        assert callable(getattr(analyze, internal)) and not hasattr(cp, internal)
    ph = analyze.compute_protocol_health(
        {"sw1": {"Vlan10": ID(port="Vlan10", hsrp_behavior="HSRP grp 1 Active VIP 10.0.10.1")}}, {})
    assert any(r["protocol"] == "FHRP" and r["severity"] == "Info" for r in ph)
    # physical-health compute helpers joined the analyze layer (step 17); their monolith re-exports
    # were dropped in step 25 when write_physical_health_sheet moved to excel. They still live in analyze.
    for name in ("_poe_device_util", "_physical_uplink_index"):
        assert hasattr(analyze, name) and not hasattr(cp, name)
    DP = analyze.DevicePhysical
    assert analyze._poe_device_util([DP(hostname="sw1", power_capacity_w="1000", power_drawn_w="250")]) == {"sw1": 25.0}
    # a single non-port-channel inter-switch link -> a single-fiber uplink on both ends.
    up, sf = analyze._physical_uplink_index(
        {"links": [{"a": "sw1", "ap": "Gi1/0/1", "b": "sw2", "bp": "Gi1/0/1", "is_pc": False}]})
    assert ("sw1", "Gi1/0/1") in sf and ("sw2", "Gi1/0/1") in sf
    # dependency-map + cross-layer correlations joined the analyze layer (step 18).
    assert cp.build_dependency_map is analyze.build_dependency_map
    assert cp.compute_cross_layer_correlations is analyze.compute_cross_layer_correlations
    for internal in ("all_hosts", "_CL_RANK"):  # package-internal, not re-exported
        assert hasattr(analyze, internal) and not hasattr(cp, internal)
    cl_ai = {"sw1": {"Vlan20": ID(port="Vlan20")},
             "sw2": {"Gi1/0/1": ID(port="Gi1/0/1", switchport_mode="Access", vlan="20")}}
    dep = analyze.build_dependency_map(
        cl_ai, [], [{"vlan": 20, "switch": "sw1", "risk": "single-gateway", "fhrp": "none"}])
    assert dep["sole_gw"] == {20: "sw1"}   # sole-gateway VLAN, no FHRP
    assert any(f["id"] == "CL-03" for f in analyze.compute_cross_layer_correlations(dep))
    # flow-trace joined the analyze layer (step 19); its helpers are package-internal.
    assert cp.trace_full_flow is analyze.trace_full_flow
    for internal in ("_ip_in_prefix", "_find_endpoint_by_ip", "_find_gateways_for", "_bfs_forwarding_path"):
        assert hasattr(analyze, internal) and not hasattr(cp, internal)
    assert analyze._ip_in_prefix("10.0.10.5", "10.0.10.0/24") is True
    assert analyze._ip_in_prefix("10.0.99.5", "10.0.10.0/24") is False
    # same-subnet flow between two located endpoints in VLAN 10 -> an L2 (same subnet) path.
    flow = analyze.trace_full_flow("10.0.10.5", "10.0.10.6", {
        "sw1": {"Gi1/0/1": ID(port="Gi1/0/1", switchport_mode="Access", vlan="10", end_host_ip="10.0.10.5"),
                "Gi1/0/2": ID(port="Gi1/0/2", switchport_mode="Access", vlan="10", end_host_ip="10.0.10.6")}})
    assert flow["summary"]["flow_type"] == "L2 (same subnet)" and flow["summary"]["src_ip"] == "10.0.10.5"


def test_cmdio_reexported_and_functional(cp, tmp_path):
    from cisco_toolkit import cmdio
    # _load_cmd_output / _safe_parse went package-internal in step 28 (build_interfaces, their
    # last monolith user, moved to cisco_toolkit.build); _CISCO_ERRORS stays (platform detection
    # reuses it in detect_platform_from_files).
    assert not hasattr(cp, "_load_cmd_output") and not hasattr(cp, "_safe_parse")
    assert cp._CISCO_ERRORS is cmdio._CISCO_ERRORS
    # _safe_parse: happy path returns the value; a raising parser falls back to _default.
    assert cmdio._safe_parse(lambda x: {"k": x}, 1) == {"k": 1}
    def _boom(_):
        raise ValueError("bad section")
    assert cmdio._safe_parse(_boom, "x") == {}
    assert cmdio._safe_parse(_boom, "x", _default=[]) == []
    # _load_cmd_output: returns file content; skips an empty/Cisco-error capture + absent cmds.
    good = tmp_path / "good.txt"; good.write_text("Gi1/0/1 connected\n", encoding="utf-8")
    bad = tmp_path / "bad.txt"; bad.write_text("% Invalid input detected at '^' marker.\n", encoding="utf-8")
    c2f = {"show a": str(good), "show b": str(bad)}
    assert cmdio._load_cmd_output(c2f, "show a").strip() == "Gi1/0/1 connected"
    assert cmdio._load_cmd_output(c2f, "show b") == ""        # captured Cisco error skipped
    assert cmdio._load_cmd_output(c2f, "show missing") == ""  # absent command
    # variant fallthrough: first variant errored, second is good.
    assert cmdio._load_cmd_output(c2f, "show b", "show a").strip() == "Gi1/0/1 connected"


def test_excel_reexported_and_functional(cp):
    from cisco_toolkit import excel
    # the sheet/header helpers main() + the write_* builders use are re-exported as-is.
    for name in ("find_header_row", "ensure_headers"):
        assert getattr(cp, name) is getattr(excel, name)
    # _census_header / _census_autofit (step 24) + sortkey (step 26) went package-internal as their
    # last monolith users (the writers / append_interface_rows) moved to excel; still in excel.
    assert callable(excel._census_header) and not hasattr(cp, "_census_header")
    assert callable(excel._census_autofit) and not hasattr(cp, "_census_autofit")
    assert callable(excel.sortkey) and not hasattr(cp, "sortkey")
    # norm_header / HEADER_TO_FIELD are package-internal (only find_header_row uses them).
    assert callable(excel.norm_header) and not hasattr(cp, "norm_header")
    assert isinstance(excel.HEADER_TO_FIELD, dict) and not hasattr(cp, "HEADER_TO_FIELD")
    # functional smokes straight from the package.
    assert excel.norm_header("  Switchport  Mode ") == "switchport mode"
    assert excel.HEADER_TO_FIELD["fhrp behavior"] == "hsrp_behavior"
    # port-row sort: port-channels first, then Eth/Fa/Gi... by number (not lexical).
    ports = ["Gi1/0/2", "Po1", "Gi1/0/10", "Gi1/0/1"]
    assert sorted(ports, key=excel.sortkey) == ["Po1", "Gi1/0/1", "Gi1/0/2", "Gi1/0/10"]
    # census/inventory sheet writers joined the excel layer (step 21); re-exported for main().
    for name in ("write_device_inventory_sheet", "write_svi_gateway_sheet", "write_stp_detail_sheet",
                 "write_vlan_census_sheet", "write_endpoint_census_sheet"):
        assert getattr(cp, name) is getattr(excel, name)
    # _is_svi + the *_COLUMNS / *_SHEET_NAME constants are package-internal.
    assert excel._is_svi("Vlan10") is True and excel._is_svi("Gi1/0/1") is False
    assert not hasattr(cp, "_is_svi") and not hasattr(cp, "INVENTORY_SHEET_NAME")
    # analysis sheet writers joined the excel layer (step 22) - these call analyze compute_*.
    for name in ("write_move_group_sheet", "write_topology_sheet", "write_findings_sheet",
                 "write_capacity_sheet", "write_topology_diagram"):
        assert getattr(cp, name) is getattr(excel, name)
    # _to_float / _mermaid_id are package-internal helpers.
    assert excel._to_float("250.5 W") == 250.5 and excel._to_float("n/a") is None
    assert not hasattr(cp, "_to_float") and not hasattr(cp, "_mermaid_id")
    # health/analysis sheet writers joined the excel layer (step 23) - they render computed records.
    for name in ("write_cross_layer_sheet", "write_protocol_health_sheet", "write_health_scores_sheet",
                 "write_score_sensitivity_sheet", "write_migration_readiness_sheet"):
        assert getattr(cp, name) is getattr(excel, name)
    # _SEV_FILL was re-exported for monolith writers in step 23; step 25 moved the last of those
    # (physical/l3) to excel, so it's package-internal now. _CL_FILL/_READY_FILL/_STATUS_FILL too.
    assert excel._SEV_FILL["High"] == "F4CCCC" and not hasattr(cp, "_SEV_FILL")
    assert not hasattr(cp, "_CL_FILL") and not hasattr(cp, "_READY_FILL")
    # Tier-2 (file-reading) + intelligence sheet writers joined the excel layer (step 24).
    for name in ("write_interface_health_sheet", "write_security_posture_sheet",
                 "write_routing_adjacency_sheet", "write_causality_chains_sheet",
                 "write_failure_impact_sheet"):
        assert getattr(cp, name) is getattr(excel, name)
    # fused compute-in-writer sheets joined the excel layer (step 25).
    for name in ("write_physical_health_sheet", "write_flow_trace_sheet", "write_l3_forwarding_sheet"):
        assert getattr(cp, name) is getattr(excel, name)
    # _parse_track / _track_summary moved with write_l3_forwarding_sheet (package-internal).
    assert callable(excel._parse_track) and not hasattr(cp, "_parse_track")
    # the main interface-template filler joined the excel layer (step 26) - the LAST sheet writer.
    assert cp.append_interface_rows is excel.append_interface_rows


def test_build_reexported_and_functional(cp):
    from cisco_toolkit import build, parse
    # identity, not equality: main() must call the package's builders/enrichers, not a
    # re-defined copy. Five moved to cisco_toolkit.build in step 27; build_interfaces joined in
    # step 28 (asserted just below).
    for name in ("build_device_physical", "build_switch_identity", "collect_global_arp",
                 "apply_global_arp", "detect_cross_device_dual_connections"):
        assert getattr(cp, name) is getattr(build, name)
    # build_interfaces (the big per-device InterfaceData builder) joined the build layer in step 28;
    # re-exported for main() + conftest's `built` fixture.
    assert cp.build_interfaces is build.build_interfaces
    # step 27 took these 8 parsers package-internal; step 28 took parse_run_config_interfaces too
    # (build_interfaces, its last monolith user, moved). The monolith now re-exports NO parser.
    for gone in ("parse_show_version", "parse_show_inventory", "parse_show_environment_power",
                 "parse_show_environment", "parse_show_module_count", "parse_vtp_status",
                 "parse_switch_mgmt_ip", "parse_show_ip_arp", "parse_run_config_interfaces"):
        assert hasattr(parse, gone) and not hasattr(cp, gone)
    # functional smokes straight from the package (pure, no file I/O).
    ID = build.InterfaceData
    # apply_global_arp fills end_host_ip + arp_source_switch from the global ARP table.
    ai = {"sw1": {"Gi1/0/1": ID(port="Gi1/0/1", end_host_mac="aaaa.bbbb.cccc")}}
    build.apply_global_arp(ai, {"aaaa.bbbb.cccc": "10.0.10.5"}, {"aaaa.bbbb.cccc": "core1"})
    assert ai["sw1"]["Gi1/0/1"].end_host_ip == "10.0.10.5"
    assert ai["sw1"]["Gi1/0/1"].arp_source_switch == "core1"
    # detect_cross_device_dual_connections flags a MAC seen on two different switches.
    dc = {"sw1": {"Gi1/0/1": ID(port="Gi1/0/1", end_host_mac="dead.beef.0001")},
          "sw2": {"Gi1/0/2": ID(port="Gi1/0/2", end_host_mac="dead.beef.0001")}}
    build.detect_cross_device_dual_connections(dc)
    assert dc["sw1"]["Gi1/0/1"].dual_connection == "Yes"
    assert dc["sw2"]["Gi1/0/2"].dual_connection == "Yes"


def test_html_reexported_and_functional(cp, tmp_path):
    from openpyxl import load_workbook

    from cisco_toolkit import html
    # write_diff_workbook (the --compare diff workbook) joined the snapshot-reporting layer in
    # step 29; re-exported for main()'s --compare path. _macset / _DIFF_FIELDS are package-internal.
    assert cp.write_diff_workbook is html.write_diff_workbook
    # snapshot_state (the JSON contract) + write_html_explorer joined the layer in step 30; both
    # re-exported (main() builds the snapshot + emits the HTML; test_version calls cp.snapshot_state).
    assert cp.snapshot_state is html.snapshot_state
    assert cp.write_html_explorer is html.write_html_explorer
    # __version__ moved into cisco_toolkit/__init__.py (step 30); the monolith re-imports it.
    import cisco_toolkit
    assert cp.__version__ == cisco_toolkit.__version__ == "3.23.0"
    # snapshot_state stamps script_version from the package __version__; empty inputs -> empty maps.
    snap = html.snapshot_state({}, [])
    assert snap["schema"] == "collect_parse_snapshot/1" and snap["script_version"] == "V3.23.0"
    assert snap["devices"] == {} and snap["interfaces"] == {}
    assert callable(html._macset) and not hasattr(cp, "_macset")
    assert isinstance(html._DIFF_FIELDS, list) and not hasattr(cp, "_DIFF_FIELDS")
    # _macset splits on commas / whitespace / semicolons, dropping blanks.
    assert html._macset("aaaa.bbbb.cccc, dddd.eeee.ffff") == {"aaaa.bbbb.cccc", "dddd.eeee.ffff"}
    assert html._macset("") == set()
    # functional smoke: a status change on one port -> a Modified row in 'Interface Changes'.
    old = {"devices": {"sw1": {}}, "interfaces": {"sw1": {"Gi1/0/1": {"status": "connected"}}}}
    new = {"devices": {"sw1": {}}, "interfaces": {"sw1": {"Gi1/0/1": {"status": "notconnect"}}}}
    out = tmp_path / "diff.xlsx"
    html.write_diff_workbook(old, new, str(out))
    wb = load_workbook(str(out))
    assert {"Summary", "Interface Changes", "Endpoint Changes", "SVI Changes"} <= set(wb.sheetnames)
    rows = list(wb["Interface Changes"].iter_rows(min_row=2, values_only=True))
    assert any(r[0] == "sw1" and r[1] == "Gi1/0/1" and r[2] == "Modified"
               and "status: connected -> notconnect" in (r[3] or "") for r in rows)


def test_redact_snapshot_pseudonymizes_consistently():
    from cisco_toolkit import html
    snap = {
        "devices": {"core1": {"serial_number": "FOC1234X", "chassis_serial": "FOC1234X",
                              "system_mac": "00:11:22:33:44:55"}},
        "interfaces": {"core1": {
            "Vlan10":  {"port": "Vlan10", "svi_ip": "10.0.10.2",
                        "hsrp_behavior": "HSRP grp10 active vIP 10.0.10.1"},
            "Gi1/0/1": {"port": "Gi1/0/1", "end_host_ip": "10.0.10.5", "end_host_mac": "aa:bb:cc:dd:ee:01"},
            "Gi1/0/2": {"port": "Gi1/0/2", "end_host_ip": "10.0.10.6", "end_host_mac": "aa:bb:cc:dd:ee:01"},
        }},
        "acls": {"core1": {"PROTECT_SERVERS": [
            {"action": "permit", "proto": "tcp", "sport": None, "dport": {"op": "eq", "val": 443},
             "src": {"ip": "10.0.10.0", "wild": "0.0.0.255"}, "dst": {"ip": "10.0.30.0", "wild": "0.0.0.255"},
             "raw": "permit tcp 10.0.10.0 0.0.0.255 10.0.30.0 0.0.0.255 eq 443"}]}},
    }
    r = html.redact_snapshot(snap)
    ifs = r["interfaces"]["core1"]
    blob = str(r)
    # the real identifiers are gone
    assert "10.0.10" not in blob and "aa:bb:cc:dd:ee:01" not in blob and "FOC1234X" not in blob
    assert "00:11:22:33:44:55" not in blob
    # ACL rule IPs are redacted (no leak) but wildcard masks are preserved (post-redact L4 eval stays valid)
    assert "10.0.30" not in blob
    acl = r["acls"]["core1"]["PROTECT_SERVERS"][0]
    assert acl["src"]["ip"] != "10.0.10.0" and acl["dst"]["ip"] != "10.0.30.0"
    assert acl["src"]["wild"] == "0.0.0.255" and acl["dst"]["wild"] == "0.0.0.255"
    # consistent: the same MAC on two ports maps to the same pseudonym
    assert ifs["Gi1/0/1"]["end_host_mac"] == ifs["Gi1/0/2"]["end_host_mac"]
    # subnet preserved: two hosts in one /24 keep the same redacted /24, distinct host octets
    a, b = ifs["Gi1/0/1"]["end_host_ip"], ifs["Gi1/0/2"]["end_host_ip"]
    assert a.rsplit(".", 1)[0] == b.rsplit(".", 1)[0] and a != b
    # an IP embedded in free text is redacted too, but the surrounding words survive
    hb = ifs["Vlan10"]["hsrp_behavior"]
    assert "10.0.10.1" not in hb and hb.startswith("HSRP grp10 active vIP ")
    # the same serial in two fields maps consistently; hostnames are kept; input not mutated
    dev = r["devices"]["core1"]
    assert dev["serial_number"] == dev["chassis_serial"] and dev["serial_number"].startswith("SN")
    assert "core1" in r["interfaces"]
    assert snap["interfaces"]["core1"]["Gi1/0/1"]["end_host_ip"] == "10.0.10.5"


def test_collect_global_arp_attribution_is_order_independent(tmp_path):
    """Reproducibility (SSOT): under multi-worker collection all_cmd_to_files is built in thread-COMPLETION
    order (COLLECT_PARSE's as_completed), so collect_global_arp's first-writer-wins attribution of a MAC seen by
    >=2 switches was nondeterministic -- the chosen end_host_ip (on an IP conflict) and the arp_source_switch
    provenance could differ run-to-run for the SAME evidence, and both are persisted + displayed snapshot
    fields. The attribution must be order-INDEPENDENT: deterministic, 'lowest hostname wins'."""
    import os  # noqa: F401
    from cisco_toolkit import build

    mac = "00aa.bbcc.ddee"   # the SAME MAC, reported by two switches with DIFFERENT IPs (moved host / overlapping ARP)

    def _arp(name, ip):
        d = tmp_path / name
        d.mkdir()
        (d / "arp.txt").write_text(
            "Protocol  Address     Age   Hardware Addr   Type   Interface\n"
            f"Internet  {ip}   5     {mac}  ARPA   Vlan10\n", encoding="utf-8")
        return {"show ip arp": str(d / "arp.txt")}

    cA = _arp("access-a", "10.0.10.5")
    cB = _arp("access-b", "10.0.20.5")

    arp1, src1 = build.collect_global_arp({"access-a": cA, "access-b": cB})   # one insertion order
    arp2, src2 = build.collect_global_arp({"access-b": cB, "access-a": cA})   # the OTHER (as a different completion order would give)

    assert arp1 == arp2, (arp1, arp2)        # same MAC->IP regardless of host insertion order
    assert src1 == src2, (src1, src2)        # same provenance regardless of order
    m = next(iter(arp1))                      # the single normalized MAC key
    assert arp1[m] == "10.0.10.5" and src1[m] == "access-a"   # deterministic winner: 'access-a' < 'access-b'


def test_xml_safe_strips_xml_illegal_chars_and_deep_walks():
    """The shared sanitizer behind the excel + docx generators: strip XML-illegal chars (C0 controls except
    tab/newline/CR, U+FFFE/U+FFFF noncharacters, lone surrogates), pass non-strings through, and deep-walk a
    JSON-like structure over both keys and values."""
    from cisco_toolkit.textutils import xml_safe, xml_safe_deep
    bad = chr(0xFFFF) + chr(0xFFFE) + chr(0xD800) + chr(0xDFFF) + "\x00\x07\x1f"
    assert xml_safe("ok" + bad + "x") == "okx"
    assert xml_safe("keep\ttab\nnl\rcr") == "keep\ttab\nnl\rcr"      # legal XML whitespace preserved
    assert xml_safe(5) == 5 and xml_safe(None) is None              # non-strings pass through
    out = xml_safe_deep({"h" + bad: ["a" + bad, {"k": "v" + bad}], "n": 3})
    assert out == {"h": ["a", {"k": "v"}], "n": 3}                  # keys + values + nesting all cleaned
