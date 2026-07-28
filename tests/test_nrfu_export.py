"""Offline NRFU verification-command export (cisco_toolkit/nrfu_export.py) — orchestration-roadmap
frontier. Pins the canonical FOUR-PHASE certification pack: phase presence, the case-id grammar, the
IOS vs NX-OS command dialects, the EXPECTED-value prefill from collected evidence (STP root / HSRP
active / routing-neighbor sets), the NOT-OBSERVED abstention honesty, the READ-ONLY command guard
(grammar re-derived from tests/test_readonly_and_no_egress.py's approach), and the per-device
pack-writer file layout. COMPLEMENTS compute_validation_plan (the per-wave post-cutover spot checks);
this is the full NRFU/ATP certification pack."""
import os
import re

from cisco_toolkit.excel import NRFU_COMMANDS_SHEET_NAME, write_nrfu_commands_sheet
from cisco_toolkit.nrfu_export import (
    NOT_OBSERVED,
    UNSCHEDULED_WAVE,
    compute_nrfu_commands,
    write_nrfu_pack,
)

# READ-ONLY grammar, re-derived from tests/test_readonly_and_no_egress.py (_READ_ONLY_CMD) — the same
# verb-anchored allow-list the collection registries are held to. Extended HERE (test-local, the
# doctrine test is untouched) for the two read-only diagnostics the NRFU service phase legitimately
# issues: ping (already a doctrine read verb) and traceroute. A write verb (configure/set/reload/
# clear/debug ...) matches none of these branches and fails the guard.
_READ_ONLY_CMD = re.compile(
    r"^(show|display|get|dir|ping|moquery)\b"          # SSH/CLI read verbs (doctrine grammar)
    r"|^aws ec2 describe-"                              # AWS read-only describe (doctrine grammar)
    r"|^(api/|ers/|dataservice/)"                      # controller-REST GET paths (doctrine grammar)
    r"|^traceroute\b",                                 # NRFU extension: the 2nd read-only diagnostic
)

_CASE_ID = re.compile(r"^NRFU-W\d+-P[1-4]-\d{3}$")


# ---------------------------------------------------------------- synthetic snapshot fixture ---
def _snap():
    """A snap-shaped fixture: an IOS dist (full evidence), an NX-OS dist (partial evidence), and an
    UNSCHEDULED IOS access switch with almost no evidence (exercises the abstention paths)."""
    return {
        "devices": {
            "dist1": {"platform": "ios", "model": "WS-C3850-24T", "serial_number": "FCW111",
                      "sw_version": "16.12.4", "ps_status": "OK", "fan_status": "OK",
                      "temperature_status": "GREEN"},
            "dist2": {"platform": "nxos", "model": "N9K-C93180YC-EX", "serial_number": "FDO222",
                      "sw_version": "9.3(8)"},
            "acc1": {"platform": "ios", "model": "WS-C2960X-48", "serial_number": "FOC333",
                     "sw_version": ""},
        },
        "interfaces": {
            "dist1": {
                "Vlan10": {"port": "Vlan10", "status": "connected", "svi_ip": "10.0.10.2",
                           "hsrp_behavior": "HSRP grp 10 Active VIP 10.0.10.1"},
                "Gi1/0/1": {"port": "Gi1/0/1", "status": "connected", "port_channel": "Po1",
                            "cdp_neighbor": "dist2", "neighbor_port": "Eth1/1"},
                "Gi1/0/2": {"port": "Gi1/0/2", "status": "connected", "port_channel": "Po1"},
                "Gi1/0/3": {"port": "Gi1/0/3", "status": "notconnect"},
            },
            "dist2": {
                "Vlan20": {"port": "Vlan20", "status": "connected", "svi_ip": "10.0.20.2"},
                "Eth1/1": {"port": "Eth1/1", "status": "connected", "port_channel": "Po1",
                           "cdp_neighbor": "dist1", "neighbor_port": "Gi1/0/1"},
                # the logical bundle interface also carries port_channel — must NOT count as a member
                "Po1": {"port": "Po1", "status": "connected", "port_channel": "Po1"},
            },
            "acc1": {
                "Gi0/1": {"port": "Gi0/1", "status": "connected", "switchport_mode": "Access",
                          "vlan": "10"},
            },
        },
        "move_groups": [{"switches": ["dist1", "dist2"]}],          # acc1 stays unscheduled
        "stp_roots": {
            "dist1": {"10": {"is_root": True, "root_priority": 24586, "root_address": "aaaa.0001.0001"},
                      "20": {"is_root": False, "root_priority": 32798, "root_address": "cccc.0003.0003"}},
            "dist2": {"10": {"is_root": False, "root_priority": None, "root_address": ""}},
        },
        "routing_neighbors": {
            "dist1": {"ospf": [{"neighbor": "10.0.99.2", "state": "FULL/DR"}], "eigrp": [],
                      "bgp": [{"neighbor": "10.0.99.9"}]},
        },
        "fhrp_detail": {
            "dist1": [{"ifname": "Vlan10", "group": "10", "state": "Active", "vip": "10.0.10.1",
                       "standby_ip": "10.0.10.3"}],
        },
        "application_intelligence": {
            "domains": [{"id": "general", "domain": "General / Back-office", "tier": "Support",
                         "switches": ["dist1"]}],
        },
    }


def _all_cases(out):
    return [(w["wave_id"], d["host"], c)
            for w in out["waves"] for d in w["devices"] for c in d["cases"]]


def _cases_of(out, host, **match):
    return [c for _w, h, c in _all_cases(out)
            if h == host and all(c.get(k) == v for k, v in match.items())]


# ------------------------------------------------------------------------- shape & phases ---
def test_schema_and_four_phases_present():
    out = compute_nrfu_commands(_snap())
    assert out["schema"] == "nrfu_commands/1"
    assert out["banner"]
    phases = {c["phase"] for _w, _h, c in _all_cases(out)}
    assert phases == {1, 2, 3, 4}, "the canonical four NRFU phases must all be emitted"
    # scope vocabulary is closed
    assert {c["scope"] for _w, _h, c in _all_cases(out)} <= {"per-site", "end-to-end"}
    # every case is fully populated: command + expected + source_key are never blank
    for _w, _h, c in _all_cases(out):
        assert c["command"].strip() and c["expected"].strip() and c["source_key"].strip()
    s = out["summary"]
    assert s["n_cases"] == len(_all_cases(out)) and s["n_waves"] == len(out["waves"])


def test_case_id_format_wave_number_and_uniqueness():
    out = compute_nrfu_commands(_snap())
    ids = [c["id"] for _w, _h, c in _all_cases(out)]
    assert ids and len(ids) == len(set(ids)), "case ids must be unique"
    assert all(_CASE_ID.match(i) for i in ids), f"bad case-id format: {ids}"
    for w in out["waves"]:
        wno = "0" if w["wave_id"] == UNSCHEDULED_WAVE else re.search(r"(\d+)", w["wave_id"]).group(1)
        for d in w["devices"]:
            for c in d["cases"]:
                assert c["id"].startswith(f"NRFU-W{wno}-P{c['phase']}-"), \
                    f"{c['id']} does not encode wave {wno} / phase {c['phase']}"


def test_unscheduled_hosts_are_not_silently_skipped():
    out = compute_nrfu_commands(_snap())
    wave_ids = [w["wave_id"] for w in out["waves"]]
    assert wave_ids == ["Group 1", UNSCHEDULED_WAVE]                # numbered first, unscheduled last
    hosts = {w["wave_id"]: sorted(d["host"] for d in w["devices"]) for w in out["waves"]}
    assert hosts["Group 1"] == ["dist1", "dist2"]
    assert hosts[UNSCHEDULED_WAVE] == ["acc1"]                      # coverage-honesty: still certified


# ------------------------------------------------------------------------ platform dialect ---
def test_ios_vs_nxos_command_dialects():
    out = compute_nrfu_commands(_snap())
    dialect = {d["host"]: d["platform_dialect"] for w in out["waves"] for d in w["devices"]}
    assert dialect["dist1"] == "ios" and dialect["dist2"] == "nxos"
    ios_cmds = {c["command"] for c in _cases_of(out, "dist1")}
    nx_cmds = {c["command"] for c in _cases_of(out, "dist2")}
    assert "show inventory" in ios_cmds and "show environment all" in ios_cmds
    assert "show etherchannel summary" in ios_cmds and "show standby brief" in ios_cmds
    assert "show module" in nx_cmds and "show environment" in nx_cmds
    assert "show port-channel summary" in nx_cmds
    # the NX-OS device never gets the IOS forms and vice versa
    assert not {"show inventory", "show etherchannel summary", "show environment all"} & nx_cmds
    assert not {"show module", "show port-channel summary"} & ios_cmds


# --------------------------------------------------------------------- expected-value prefill ---
def test_expected_prefill_stp_root_hsrp_active_and_neighbor_sets():
    out = compute_nrfu_commands(_snap())
    # show version asserts the collected version, citing its source key
    ver = _cases_of(out, "dist1", command="show version")
    assert ver and "16.12.4" in ver[0]["expected"] and ver[0]["source_key"] == "devices.dist1.sw_version"
    # STP root per VLAN: the collected root identity is the assertion
    st10 = _cases_of(out, "dist1", command="show spanning-tree vlan 10")
    assert st10 and "root" in st10[0]["expected"].lower() and "24586" in st10[0]["expected"]
    assert st10[0]["source_key"] == "stp_roots.dist1.10"
    st20 = _cases_of(out, "dist1", command="show spanning-tree vlan 20")
    assert st20 and "cccc.0003.0003" in st20[0]["expected"]        # non-root: assert WHO the root is
    # HSRP active/standby from fhrp_detail
    hs = _cases_of(out, "dist1", command="show standby brief")
    assert hs and "Active" in hs[0]["expected"] and "10.0.10.1" in hs[0]["expected"]
    assert hs[0]["source_key"] == "fhrp_detail.dist1"
    # routing-neighbor sets: OSPF and BGP peers pre-filled
    ospf = _cases_of(out, "dist1", command="show ip ospf neighbor")
    assert ospf and "10.0.99.2" in ospf[0]["expected"] and ospf[0]["source_key"] == "routing_neighbors.dist1.ospf"
    bgp = _cases_of(out, "dist1", command="show ip bgp summary")
    assert bgp and "10.0.99.9" in bgp[0]["expected"] and "Established" in bgp[0]["expected"]
    # CDP adjacency set from the collected neighbors
    cdp = _cases_of(out, "dist1", command="show cdp neighbors")
    assert cdp and "dist2" in cdp[0]["expected"] and "Gi1/0/1" in cdp[0]["expected"]
    # known-up interface set
    up = _cases_of(out, "dist1", command="show interface status")
    assert up and "Gi1/0/1" in up[0]["expected"] and "Gi1/0/3" not in up[0]["expected"]
    # port-channel member count
    pc = _cases_of(out, "dist1", command="show etherchannel summary")
    assert pc and "Po1" in pc[0]["expected"] and "2 member(s)" in pc[0]["expected"]
    # the logical Po1 interface record must not inflate the member count (dist2: 1 real member)
    pc2 = _cases_of(out, "dist2", command="show port-channel summary")
    assert pc2 and "Po1: 1 member(s)" in pc2[0]["expected"]


def test_gateway_svi_ping_and_traceroute_within_move_group():
    out = compute_nrfu_commands(_snap())
    # hub = lowest (host, vlan) gateway SVI in the wave: dist1 Vlan10 (10.0.10.2); dist2 pings it
    pings = _cases_of(out, "dist2", command="ping 10.0.10.2")
    assert pings, "gateway-SVI ping between move-group members missing"
    assert pings[0]["phase"] == 3 and pings[0]["scope"] == "end-to-end"
    assert "100 percent" in pings[0]["expected"]
    assert pings[0]["source_key"] == "interfaces.dist1.Vlan10.svi_ip"
    tr = _cases_of(out, "dist2", command="traceroute 10.0.10.2")
    assert tr and tr[0]["phase"] == 3 and "10.0.10.2" in tr[0]["expected"]
    # the unscheduled lone access switch has no SVI pair -> no fabricated ping targets
    assert not [c for c in _cases_of(out, "acc1") if c["command"].startswith(("ping", "traceroute"))
                and c["phase"] == 3 and c["scope"] == "end-to-end"]


def test_phase4_application_placeholders_are_human_executed():
    out = compute_nrfu_commands(_snap())
    p4 = [c for _w, _h, c in _all_cases(out) if c["phase"] == 4]
    assert p4, "Phase IV application rows missing"
    assert all(c["scope"] == "end-to-end" for c in p4)
    dom = p4[0]
    assert "General / Back-office" in dom["expected"] and "[HUMAN-EXECUTED]" in dom["expected"]
    assert dom["source_key"].startswith("application_intelligence.domains")
    assert "<" in dom["command"], "Phase IV target is an explicit <placeholder>, never a guessed IP"


# ------------------------------------------------------------------- NOT-OBSERVED honesty ---
def test_not_observed_abstention_never_fabricates():
    out = compute_nrfu_commands(_snap())
    # acc1 has no collected sw_version -> the expected value is the abstention marker, verbatim
    ver = _cases_of(out, "acc1", command="show version")
    assert ver and ver[0]["expected"] == NOT_OBSERVED
    # dist2 has no environment evidence -> abstention
    env = _cases_of(out, "dist2", command="show environment")
    assert env and env[0]["expected"] == NOT_OBSERVED
    # dist2's STP entry has no root identity collected -> abstention, not an invented root
    st = _cases_of(out, "dist2", command="show spanning-tree vlan 10")
    assert st and st[0]["expected"] == NOT_OBSERVED
    # DHCP snooping is registered for collection but not published in the snapshot -> abstention
    ds = _cases_of(out, "acc1", command="show ip dhcp snooping binding")
    assert ds and ds[0]["expected"] == NOT_OBSERVED and ds[0]["phase"] == 3
    # no half-fabricated strings anywhere: an expected value either carries evidence or abstains
    for _w, _h, c in _all_cases(out):
        assert "None" not in c["expected"], f"fabricated/None leak in expected: {c}"
    assert out["summary"]["n_not_observed"] >= 4


# ----------------------------------------------------------------------- read-only guard ---
def test_every_emitted_command_is_read_only():
    """Every command in the pack matches the read-only grammar re-derived from the doctrine test
    (verb-anchored allow-list + the test-local ping/traceroute diagnostics extension). A future write
    verb (configure/clear/reload/copy/...) fails HERE instead of shipping in a certification pack."""
    out = compute_nrfu_commands(_snap())
    bad = [c["command"] for _w, _h, c in _all_cases(out)
           if "\n" in c["command"] or "\r" in c["command"]           # single-line IS part of read-only:
           or not _READ_ONLY_CMD.match(c["command"].strip())]       # a 2nd physical line is a 2nd command
    assert not bad, f"NON-read-only command(s) emitted in the NRFU pack: {bad}"
    # sanity: the guard is not vacuous — the pack actually contains commands of both classes
    cmds = [c["command"] for _w, _h, c in _all_cases(out)]
    assert any(c.startswith("show ") for c in cmds) and any(c.startswith("ping ") for c in cmds)


# ------------------------------------------------------------------------ empty tolerance ---
def test_empty_and_malformed_inputs_are_tolerated():
    out = compute_nrfu_commands({})
    assert out["schema"] == "nrfu_commands/1" and out["waves"] == []
    assert out["summary"]["n_cases"] == 0 and out["summary"]["n_not_observed"] == 0
    assert compute_nrfu_commands(None)["waves"] == []
    weird = compute_nrfu_commands({"devices": 5, "interfaces": "junk", "move_groups": "oops",
                                   "stp_roots": [], "fhrp_detail": 7})
    assert weird["waves"] == []


# --------------------------------------------------------------------------- pack writer ---
def test_pack_writer_file_layout_and_content(tmp_path):
    snap = _snap()
    written = write_nrfu_pack(snap, str(tmp_path))
    rel = sorted(os.path.relpath(p, str(tmp_path)).replace(os.sep, "/") for p in written)
    assert rel == ["Group_1/dist1.txt", "Group_1/dist2.txt", "unscheduled/acc1.txt"]
    body = open(os.path.join(str(tmp_path), "Group_1", "dist1.txt"), encoding="utf-8").read()
    assert "NRFU-W1-P1-001" in body and "show version" in body
    assert "expect:" in body and "source:" in body
    # every executable (non-comment, non-blank) line is read-only — the pack is paste-safe
    for p in written:
        for line in open(p, encoding="utf-8").read().splitlines():
            line = line.strip()
            if not line or line.startswith("!"):
                continue
            assert _READ_ONLY_CMD.match(line), f"non-read-only line in {p}: {line}"
    # the pack writer also works from a snapshot that already publishes nrfu_commands (SSOT reuse)
    snap["nrfu_commands"] = compute_nrfu_commands(snap)
    written2 = write_nrfu_pack(snap, str(tmp_path / "again"))
    assert len(written2) == len(written)


def _assert_pack_executable_lines_readonly(root):
    """Every non-comment, non-blank physical line in every emitted file is a single-line
    show/ping/traceroute — the executable surface of the pack, whatever the snapshot contained."""
    import glob
    for p in glob.glob(os.path.join(root, "**", "*.txt"), recursive=True):
        for line in open(p, encoding="utf-8").read().splitlines():
            if not line.strip() or line.startswith("!"):
                continue
            assert _READ_ONLY_CMD.match(line), f"executable non-read-only line in {p}: {line!r}"
            assert line.split()[0].lower() not in ("configure", "conf", "reload", "clear", "no")


def test_hostile_newline_values_cannot_inject_executable_lines(tmp_path):
    """Adversarial-review regression (2026-07-05, doctrine lens, verified-by-execution): snapshot
    strings are attacker-controllable on the --no-collect path; an embedded newline in an
    interpolated value (an stp_roots VLAN key, a device field, a CDP neighbor) must NEVER become an
    executable line in the pack — 'No writes to devices, ever' is guardrail #1."""
    payload = "10\nconfigure terminal\ninterface Gi1/0/1\nshutdown"
    snap = _snap()
    snap["stp_roots"]["dist1"][payload] = {"is_root": True, "root_priority": 24586}
    snap["devices"]["dist1"]["sw_version"] = "16.12.4\nconfigure terminal\nno logging"
    snap["interfaces"]["dist1"]["Gi1/0/1"]["cdp_neighbor"] = "evil\nreload"
    out = compute_nrfu_commands(snap)
    for _w, _h, c in _all_cases(out):
        for field in ("command", "expected", "source_key"):
            assert "\n" not in c[field] and "\r" not in c[field], \
                f"multi-line {field} escaped the chokepoint: {c[field]!r}"
    write_nrfu_pack(snap, str(tmp_path))
    _assert_pack_executable_lines_readonly(str(tmp_path))
    # belt-and-braces: a TAMPERED pre-published nrfu_commands section (which bypasses compute's
    # chokepoint entirely) is refused at the writer — the injected verb becomes a comment line.
    snap["nrfu_commands"] = {
        "waves": [{"wave_id": "Group 1", "devices": [{"host": "dist1", "platform_dialect": "ios",
                   "cases": [{"id": "NRFU-W1-P2-001", "phase": 2, "scope": "per-site",
                              "command": "show version\nconfigure terminal\nshutdown",
                              "expected": "x", "source_key": "y"},
                             {"id": "NRFU-W1-P2-002", "phase": 2, "scope": "per-site",
                              "command": "configure terminal", "expected": "x", "source_key": "y"}]}]}]}
    write_nrfu_pack(snap, str(tmp_path / "tampered"))
    _assert_pack_executable_lines_readonly(str(tmp_path / "tampered"))
    body = open(os.path.join(str(tmp_path), "tampered", "Group_1", "dist1.txt"),
                encoding="utf-8").read()
    assert "[REFUSED — not a read-only command] configure terminal" in body


def test_wave_id_cannot_escape_the_output_directory(tmp_path):
    """Same untrusted-snapshot class as the newline test above, on the PATH instead of the command
    text: `_safe_name`'s whitelist rewrites every separator but KEEPS '.', so `_safe_name("..")`
    returned '..' and `os.path.join(out_dir, "..")` resolved to the parent of --out — a crafted
    wave_id (or host) wrote the pack outside the directory the operator named."""
    from cisco_toolkit.nrfu_export import _safe_name

    for hostile in ("..", ".", "...", "../..", r"..\..", "", "   ", "/", "___"):
        got = _safe_name(hostile)
        assert got == "unnamed" or (os.sep not in got and got not in (".", "..")), (hostile, got)
    assert _safe_name("Group 1") == "Group_1"          # ordinary names are untouched
    assert _safe_name("sw1.corp.example") == "sw1.corp.example"

    out = tmp_path / "out"
    out.mkdir()
    snap = {"nrfu_commands": {"waves": [
        {"wave_id": "..", "devices": [{"host": "..", "platform_dialect": "ios",
         "cases": [{"id": "NRFU-W1-P1-001", "phase": 1, "scope": "per-site",
                    "command": "show version", "expected": "x", "source_key": "y"}]}]}]}}
    written = write_nrfu_pack(snap, str(out))
    assert written
    root = os.path.realpath(str(out))
    for p in written:
        assert os.path.realpath(p).startswith(root + os.sep), f"escaped --out: {p}"
    assert os.listdir(str(tmp_path)) == ["out"], "a file landed beside --out"


# --------------------------------------------------------------------------- sheet writer ---
def test_nrfu_commands_sheet_writer(tmp_path):
    from openpyxl import Workbook
    out = compute_nrfu_commands(_snap())
    wb = Workbook()
    write_nrfu_commands_sheet(wb, out)
    assert NRFU_COMMANDS_SHEET_NAME in wb.sheetnames
    ws = wb[NRFU_COMMANDS_SHEET_NAME]
    hdr = [c.value for c in ws[4]]
    assert hdr[:8] == ["#", "Wave", "Device", "Dialect", "Case ID", "Phase", "Scope", "Command"]
    n_rows = sum(1 for row in ws.iter_rows(min_row=5) if row[0].value is not None)
    assert n_rows == out["summary"]["n_cases"]
    # idempotent: re-writing replaces, never duplicates
    write_nrfu_commands_sheet(wb, out)
    assert wb.sheetnames.count(NRFU_COMMANDS_SHEET_NAME) == 1
    # tolerant of an empty compute result
    write_nrfu_commands_sheet(wb, {})
