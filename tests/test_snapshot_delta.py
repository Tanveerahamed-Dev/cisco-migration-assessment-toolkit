"""NEW-V3.23.106: pre/post-cutover validation — the migration-validation delta between two snapshots
(health-band shifts + consolidated punch-list findings opened/resolved + an overall verdict)."""
from openpyxl import load_workbook

from cisco_toolkit.html import _finding_key, compute_snapshot_delta, write_diff_workbook


def _snap(devices, health, punchlist):
    return {"devices": {d: {} for d in devices}, "interfaces": {d: {} for d in devices},
            "health_scores": health, "punchlist": punchlist}


def test_finding_key_distinguishes_ids_and_ignores_device_order():
    # device ORDER must not matter -- same finding on the same gateways
    a = {"category": "FHRP", "title": "Fake FHRP redundancy (VLAN 20)", "devices": ["b", "a"]}
    b = {"category": "FHRP", "title": "Fake FHRP redundancy (VLAN 20)", "devices": ["a", "b"]}
    assert _finding_key(a) == _finding_key(b)
    # a DIFFERENT embedded id is a DIFFERENT finding -- must NOT collapse (V3.23.111 fix): otherwise a
    # 'VLAN 20 fixed, VLAN 21 newly broke' swap on the same devices would read as 'no change'.
    c = {"category": "FHRP", "title": "Fake FHRP redundancy (VLAN 21)", "devices": ["a", "b"]}
    assert _finding_key(a) != _finding_key(c)
    # a different device set is also a different finding
    d = {"category": "FHRP", "title": "Fake FHRP redundancy (VLAN 20)", "devices": ["a"]}
    assert _finding_key(a) != _finding_key(d)


def test_delta_catches_same_category_same_device_id_swap():
    # the regression the over-aggressive digit-normalization used to hide
    old = _snap(["sw1"], [{"switch": "sw1", "band": "Good", "score": 80}],
                [{"severity": "High", "category": "FHRP", "title": "Fake FHRP redundancy (VLAN 20)", "devices": ["sw1"]}])
    new = _snap(["sw1"], [{"switch": "sw1", "band": "Good", "score": 80}],
                [{"severity": "High", "category": "FHRP", "title": "Fake FHRP redundancy (VLAN 21)", "devices": ["sw1"]}])
    d = compute_snapshot_delta(old, new)
    assert d["findings"]["n_opened"] == 1 and d["findings"]["opened"][0]["title"].endswith("(VLAN 21)")
    assert d["findings"]["n_resolved"] == 1 and d["findings"]["resolved"][0]["title"].endswith("(VLAN 20)")
    assert d["verdict"] == "REGRESSED"                   # a new High finding -> not silently CLEAN


def test_delta_health_findings_and_verdict():
    old = _snap(["sw1", "sw2"],
                [{"switch": "sw1", "band": "Good", "score": 80}, {"switch": "sw2", "band": "Critical", "score": 10}],
                [{"severity": "High", "category": "FHRP", "title": "Fake redundancy", "devices": ["sw1"]},
                 {"severity": "Low", "category": "STP", "title": "Accidental root", "devices": ["sw2"]}])
    new = _snap(["sw1", "sw2"],
                [{"switch": "sw1", "band": "Poor", "score": 55},    # regressed Good->Poor
                 {"switch": "sw2", "band": "Good", "score": 78}],   # improved Critical->Good
                [{"severity": "Low", "category": "STP", "title": "Accidental root", "devices": ["sw2"]},  # carried
                 {"severity": "Critical", "category": "Cross-layer", "title": "New SPOF", "devices": ["sw1"]}])  # opened
    d = compute_snapshot_delta(old, new)
    assert d["health"]["n_regressed"] == 1 and d["health"]["regressed"][0]["switch"] == "sw1"
    assert d["health"]["n_improved"] == 1 and d["health"]["improved"][0]["switch"] == "sw2"
    assert d["findings"]["n_opened"] == 1 and d["findings"]["opened"][0]["title"] == "New SPOF"
    assert d["findings"]["n_resolved"] == 1 and d["findings"]["resolved"][0]["title"] == "Fake redundancy"
    assert d["findings"]["n_opened_high"] == 1                # the Critical counts as High/Critical
    assert d["verdict"] == "REGRESSED"                       # new High/Critical + a band regression


def test_delta_clean_verdict_when_no_regressions():
    s = _snap(["sw1"], [{"switch": "sw1", "band": "Good", "score": 80}],
              [{"severity": "Low", "category": "STP", "title": "x", "devices": ["sw1"]}])
    d = compute_snapshot_delta(s, s)
    assert d["verdict"] == "CLEAN" and d["findings"]["n_opened"] == 0 and d["health"]["n_regressed"] == 0


def test_delta_tolerates_snapshots_without_computed_keys():
    bare = {"devices": {"sw1": {}}, "interfaces": {"sw1": {}}}
    d = compute_snapshot_delta(bare, bare)
    assert d["verdict"] == "CLEAN" and d["health"]["n_regressed"] == 0 and d["findings"]["n_opened"] == 0


def test_diff_workbook_has_validation_sheets(tmp_path):
    old = _snap(["sw1"], [{"switch": "sw1", "band": "Good", "score": 80}], [])
    new = _snap(["sw1"], [{"switch": "sw1", "band": "Critical", "score": 20}],
                [{"severity": "High", "category": "L3", "title": "single gateway", "devices": ["sw1"]}])
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(old, new, str(out))
    wb = load_workbook(str(out))
    assert {"Summary", "Health Shifts", "Findings Delta"} <= set(wb.sheetnames)
    # the verdict cell (Summary row 2, col C) reflects the regression
    assert wb["Summary"].cell(2, 3).value == "REGRESSED"
    hs = [tuple(c.value for c in row) for row in wb["Health Shifts"].iter_rows(min_row=2, max_row=2)]
    assert hs and hs[0][0] == "sw1" and hs[0][1] == "REGRESSED"
