"""MASTER_PLAN 2026-07-05 §4.3: the per-VLAN cutover matrix -- one row per evidenced VLAN, every
column joined from data the snapshot already carries (STP root + default-election smell, FHRP
election detail, gateway SVIs, endpoint census, application domain / criticality tier, move-group
wave + make-before-break vs hard-cutover scenario, readiness pull-through, multicast + DHCP-relay
dependency flags) plus the two deliberately-blank human fields (cutover window / rollback owner).
Coverage-honest: a VLAN with no FHRP evidence reads '[NOT OBSERVED]' -- never an implied "no
gateway redundancy" -- unless the gateway evidence positively shows a sole gateway."""
import json
import os

from openpyxl import Workbook

from cisco_toolkit.analyze import VLAN_CUTOVER_NOT_OBSERVED, compute_vlan_cutover_matrix
from cisco_toolkit.excel import VLAN_CUTOVER_SHEET_NAME, write_vlan_cutover_sheet
from cisco_toolkit.model import InterfaceData

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

EXPECTED_HEADERS = ["VLAN", "Name", "STP Root", "Root Default-Election", "FHRP",
                    "Gateway SVIs", "Endpoint MACs (per-port sum)", "App Domain", "Criticality",
                    "Dependencies", "Wave", "Scenario", "Readiness", "Cutover Window", "Rollback Owner"]


def _svi(key, ip="", fhrp="", helpers="", name=""):
    return InterfaceData(port=key, svi_ip=ip, hsrp_behavior=fhrp, dhcp_helpers=helpers, vlan_name=name)


def _access(port, vlan, mac):
    return InterfaceData(port=port, switchport_mode="Access", vlan=vlan, end_host_mac=mac)


# Three-VLAN synthetic fleet: VLAN 10 = redundant HSRP pair (distA active + detail, distB standby,
# brief only), VLAN 20 = positively-observed SOLE gateway with no FHRP, VLAN 30 = access-only (no
# gateway SVI collected anywhere -> FHRP evidence absent).
def _fabric():
    return {
        "distA": {"Vlan10": _svi("Vlan10", ip="10.0.10.2 255.255.255.0",
                                 fhrp="HSRP grp 10 Active VIP 10.0.10.1",
                                 helpers="10.0.40.10,10.0.40.11", name="USERS"),
                  "Vlan20": _svi("Vlan20", ip="10.0.20.1 255.255.255.0", name="SERVERS")},
        "distB": {"Vlan10": _svi("Vlan10", ip="10.0.10.3/24",
                                 fhrp="HSRP grp 10 Standby VIP 10.0.10.1", name="USERS")},
        "acc1": {"Gi1/0/5": _access("Gi1/0/5", "10", "aabb.ccdd.ee01 aabb.ccdd.ee02"),
                 "Gi1/0/6": _access("Gi1/0/6", "30", "aabb.ccdd.ee30")},
    }


def _stp_roots():
    # VLAN 10 root = distA at a DELIBERATE priority (24576+10); VLAN 30 root = acc1 at the DEFAULT
    # priority (32768+30) -> won on the MAC tiebreak (the accidental-root smell).
    return {"distA": {"10": {"root_priority": 24586, "root_address": "aaaa.0001.0001",
                             "is_root": True, "is_mst": False}},
            "acc1": {"10": {"root_priority": 24586, "root_address": "aaaa.0001.0001",
                            "is_root": False, "is_mst": False},
                     "30": {"root_priority": 32798, "root_address": "cccc.0003.0003",
                            "is_root": True, "is_mst": False}}}


def _fhrp_detail():
    return {"distA": [{"ifname": "Vlan10", "group": "10", "state": "Active", "priority": 110,
                       "cfg_priority": 110, "preempt": True, "preempt_delay": None,
                       "vip": "10.0.10.1", "vmac": "0000.0c07.ac0a",
                       "standby_ip": "10.0.10.3", "track": []}]}


def _ctx():
    move_groups = [{"switches": ["acc1", "distA", "distB"]}]
    wave_seq = [{"group": "Group 1", "make_before_break": ["distA", "distB"],
                 "hard_cutover": ["acc1"], "homing_unknown": [], "hard_cutover_endpoints": 3,
                 "sequence": "1 hard cutover (single-homed, schedule a window) + 2 make-before-break (dual-homed)"}]
    readiness = [{"group": "Group 1", "switches": ["acc1", "distA", "distB"], "readiness": "CAUTION"}]
    endpoint_identity = [
        {"host": "acc1", "port": "Gi1/0/5", "vlan": "10", "mac": "aabb.ccdd.ee01", "mac_count": 2},
        {"host": "acc1", "port": "Gi1/0/6", "vlan": "30", "mac": "aabb.ccdd.ee30", "mac_count": 1}]
    app_intel = {"domains": [
        {"domain": "Media Production", "tier": "On-air critical", "vlans": ["10"]},
        {"domain": "General infrastructure", "tier": "Support", "vlans": ["10", "30"]}]}
    mcast = {"querier": {"n_querier_vlans": 0, "multicast_vlans": ["10"], "gap_vlans": ["10"]}}
    return move_groups, wave_seq, readiness, endpoint_identity, app_intel, mcast


def _matrix():
    mg, ws, rd, ep, ai, mc = _ctx()
    return compute_vlan_cutover_matrix(_fabric(), _stp_roots(), _fhrp_detail(), ep, ai, mg, ws, rd, mc)


def _row(rows, vid):
    hit = [r for r in rows if r["vlan"] == vid]
    assert hit, f"no row for VLAN {vid}: {[r['vlan'] for r in rows]}"
    return hit[0]


def test_rows_carry_joined_facts():
    rows = _matrix()
    assert [r["vlan"] for r in rows] == [10, 20, 30]        # one row per evidenced VLAN, sorted
    r10 = _row(rows, 10)
    assert r10["name"] == "USERS"
    assert r10["stp_root"] == "distA" and r10["stp_root_default_election"] is False
    fhrp = r10["fhrp"]
    assert isinstance(fhrp, dict)
    assert fhrp["proto"] == "HSRP" and fhrp["group"] == "10" and fhrp["vip"] == "10.0.10.1"
    by_host = {m["host"]: m for m in fhrp["members"]}
    assert set(by_host) == {"distA", "distB"}
    assert by_host["distA"]["role"] == "active" and by_host["distA"]["priority"] == 110
    assert by_host["distA"]["preempt"] is True and by_host["distA"]["vmac"] == "0000.0c07.ac0a"
    assert by_host["distB"]["role"] == "standby"
    assert by_host["distB"]["priority"] is None             # brief-only peer: election detail not collected
    assert r10["gateway_svi_hosts"] == ["distA", "distB"]
    assert r10["endpoint_count"] == 2                       # mac_count sum for the VLAN
    assert r10["app_domain"].startswith("Media Production")  # highest tier leads
    assert r10["criticality"] == "On-air critical"
    assert "multicast active" in r10["dependencies"]
    assert "no IGMP querier" in r10["dependencies"]
    assert any(f.startswith("DHCP relay via 10.0.40.10") for f in r10["dependencies"])
    assert r10["wave"] == "Group 1" and r10["readiness"] == "CAUTION"
    # VLAN 10 spans acc1 (hard) + distA/distB (MBB) -> a mixed per-VLAN scenario
    assert "1 hard cutover" in r10["scenario"] and "2 make-before-break" in r10["scenario"]


def test_default_election_flag_matches_stp_roots():
    rows = _matrix()
    r30 = _row(rows, 30)
    assert r30["stp_root"] == "acc1"
    assert r30["stp_root_default_election"] is True         # 32798 == 32768 + 30 (default priority)
    assert _row(rows, 10)["stp_root_default_election"] is False   # 24586 is a deliberate election
    # VLAN 30 lives on acc1 only -> the group's hard-cutover switch -> hard-cutover scenario
    assert r30["scenario"].startswith("hard cutover")
    assert _row(rows, 20)["scenario"] == "make-before-break"      # distA only, dual-homed


def test_fhrp_abstention_is_honest():
    rows = _matrix()
    # VLAN 30: no gateway SVI collected anywhere -> absence of evidence is NOT "no redundancy"
    assert _row(rows, 30)["fhrp"] == VLAN_CUTOVER_NOT_OBSERVED
    assert _row(rows, 30)["gateway_svi_hosts"] == []
    # VLAN 20: the gateway evidence POSITIVELY shows a single SVI with no FHRP -> the sole-gateway
    # claim is allowed, worded like compute_migration_readiness's gateway-redundancy check
    assert _row(rows, 20)["fhrp"] == "sole gateway on distA (no FHRP)"


def test_two_gateways_without_fhrp_is_fake_redundancy_not_abstention():
    fabric = {"gwA": {"Vlan40": _svi("Vlan40", ip="10.0.40.2 255.255.255.0")},
              "gwB": {"Vlan40": _svi("Vlan40", ip="10.0.40.3 255.255.255.0")}}
    rows = compute_vlan_cutover_matrix(fabric)
    r40 = _row(rows, 40)
    assert r40["fhrp"] == "2 gateways but no FHRP — no first-hop redundancy"
    assert r40["gateway_svi_hosts"] == ["gwA", "gwB"]


def test_human_fields_are_deliberately_blank():
    rows = _matrix()
    assert rows
    for r in rows:
        assert r["cutover_window"] == "" and r["rollback_owner"] == ""


def test_unmapped_joins_stay_empty_not_invented():
    # No app / wave / readiness / multicast inputs at all: the joined columns abstain (empty),
    # they never invent a domain, wave or verdict. The DHCP-relay flag legitimately survives --
    # it rides the gateway SVI's own collected 'ip helper-address' evidence, not a join input.
    rows = compute_vlan_cutover_matrix(_fabric(), _stp_roots())
    r10 = _row(rows, 10)
    assert r10["app_domain"] == "" and r10["criticality"] == ""
    assert r10["wave"] == "" and r10["scenario"] == "" and r10["readiness"] == ""
    assert r10["endpoint_count"] == 0
    assert r10["dependencies"] == ["DHCP relay via 10.0.40.10, 10.0.40.11"]
    assert _row(rows, 30)["dependencies"] == []             # no evidence at all -> no flags


def test_mst_instance_keys_are_not_vlans():
    # MST keys are INSTANCE numbers, not VLAN ids (mirrors stp_root_findings): no phantom VLAN row.
    stp = {"coreM": {"0": {"root_priority": 32768, "root_address": "aaaa.1",
                           "is_root": True, "is_mst": True}}}
    rows = compute_vlan_cutover_matrix({}, stp)
    assert rows == []


def test_empty_inputs_are_tolerated():
    assert compute_vlan_cutover_matrix({}) == []
    assert compute_vlan_cutover_matrix(None) == []


def test_sheet_exact_headers_frozen_and_blank_human_columns():
    wb = Workbook()
    write_vlan_cutover_sheet(wb, _matrix())
    ws = wb[VLAN_CUTOVER_SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert headers == EXPECTED_HEADERS
    assert ws.freeze_panes == "A2"                          # header row frozen
    hdr_ix = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, 5):                                   # the 3 VLAN rows
        assert not ws.cell(r, hdr_ix["Cutover Window"]).value    # human fields stay blank
        assert not ws.cell(r, hdr_ix["Rollback Owner"]).value
    # row 2 = VLAN 10: FHRP dict rendered with election detail; abstention rendered verbatim on VLAN 30
    fhrp_cells = [ws.cell(r, hdr_ix["FHRP"]).value for r in range(2, 5)]
    assert "HSRP" in fhrp_cells[0] and "prio 110" in fhrp_cells[0] and "preempt" in fhrp_cells[0]
    assert fhrp_cells[2] == VLAN_CUTOVER_NOT_OBSERVED
    # default-election column is tri-state: yes / no / blank-when-root-not-observed
    de = [ws.cell(r, hdr_ix["Root Default-Election"]).value for r in range(2, 5)]
    assert de[0] == "no" and de[2] == "yes"
    assert not de[1]                                        # VLAN 20 root not observed -> no claim


def test_sheet_tolerates_empty_matrix():
    wb = Workbook()
    write_vlan_cutover_sheet(wb, [])
    ws = wb[VLAN_CUTOVER_SHEET_NAME]
    assert [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))] == EXPECTED_HEADERS


def test_golden_contract_carries_vlan_cutover_additively():
    """The re-blessed goldens must carry the new snapshot key + sheet (ADDITIVE golden discipline):
    the frozen contract is where the explorer / webapp / MOP appendix will read it from."""
    with open(os.path.join(ROOT, "tests", "golden", "snapshot.json"), encoding="utf-8") as f:
        snap = json.load(f)
    assert "vlan_cutover" in snap, "snapshot golden missing the vlan_cutover key (re-bless additively)"
    rows = snap["vlan_cutover"]
    assert isinstance(rows, list) and rows, "golden fleet evidences VLANs -> matrix must not be empty"
    for r in rows:
        assert r["cutover_window"] == "" and r["rollback_owner"] == ""
    with open(os.path.join(ROOT, "tests", "golden", "sheet_schema.json"), encoding="utf-8") as f:
        schema = json.load(f)
    assert schema.get(VLAN_CUTOVER_SHEET_NAME) == EXPECTED_HEADERS
