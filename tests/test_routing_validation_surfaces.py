"""Routing cutover evidence remains visible and executable across rendered surfaces."""

from openpyxl import Workbook
import pytest

from cisco_toolkit.excel import write_nrfu_commands_sheet, write_validation_plan_sheet


def _routing_item(name: str, evidence_state: str, command: str) -> dict:
    degraded = evidence_state == "degraded"
    marker = "PRE-CUTOVER DEGRADED" if degraded else "PRE-CUTOVER REVIEW"
    detail = (
        "observed OSPF neighbor 10.0.0.2 is EXSTART/DROTHER"
        if degraded else
        "ROUTING BASELINE NOT VERIFIED: the receipt is missing or incomplete"
    )
    return {
        "device": "distA",
        "platform": "ios",
        "wave": "Group 1",
        "category": "Routing",
        "severity": "High",
        "check": name,
        "command": command,
        "expect": f"{marker} — BLOCKER: {detail}",
        "why": "A routing baseline fault or evidence gap blocks cutover acceptance.",
        "evidence_state": evidence_state,
        "projection_custody": "embedded_unverified",
        "source_key": "routing_neighbors.distA.ospf + protocol_assessability.rows[distA,OSPF]",
    }


def _ordinary_item(index: int) -> dict:
    return {
        "device": "distA",
        "platform": "ios",
        "wave": "Group 1",
        "category": "Gateway",
        "severity": "Medium",
        "check": f"ordinary check {index:02d}",
        "command": f"show interface Gi1/0/{index + 1}",
        "expect": "observed interface state is unchanged",
        "why": "Ordinary bounded validation detail.",
    }


def _checks() -> list[dict]:
    # Put routing blockers after the ordinary rows so cap survival cannot be an accident of source order.
    return (
        [_ordinary_item(index) for index in range(45)]
        + [_routing_item("OSPF degraded baseline", "degraded", "show ip ospf neighbor")]
        + [_routing_item("BGP receipt gap", "not_verified", "show ip bgp summary")]
    )


def _plan() -> dict:
    checks = _checks()
    return {
        "items": checks,
        "by_wave": {"Group 1": checks},
        "summary": {
            "n_items": len(checks), "n_waves": 1, "n_high": 2,
            "by_category": {"Gateway": 45, "Routing": 2},
        },
        "banner": "Run every cutover validation row and disposition every blocker.",
    }


def _snapshot() -> dict:
    return {
        "script_version": "test",
        "devices": {"distA": {"platform": "ios"}},
        "move_groups": [{"switches": ["distA"], "endpoints": 8}],
        "migration_readiness": [{
            "group": "Group 1", "switches": ["distA"], "endpoints": 8,
            "readiness": "NOT READY", "n_fail": 1, "n_warn": 0, "checks": [],
        }],
        "wave_sequencing": [{
            "group": "Group 1", "make_before_break": ["distA"],
            "hard_cutover": [], "sequence": "make-before-break",
        }],
        "validation_plan": _plan(),
    }


def _tables_with_header(document, headers):
    expected = list(headers)
    return [
        table for table in document.tables
        if table.rows and [cell.text for cell in table.rows[0].cells] == expected
    ]


def test_cutover_validation_workbook_discloses_routing_evidence_and_custody():
    wb = Workbook()
    write_validation_plan_sheet(wb, _plan())
    ws = wb["Cutover Validation"]

    assert [ws.cell(4, col).value for col in range(9, 14)] == [
        "Observed baseline / acceptance", "Evidence state", "Projection custody",
        "Source key", "Why it matters",
    ]
    assert str(ws.cell(1, 1).value).startswith("⚠ ")
    assert "Evidence-state blockers apply to every protocol" in ws.cell(1, 1).value

    rows = {
        ws.cell(row, 7).value: [ws.cell(row, col).value for col in range(9, 13)]
        for row in range(5, ws.max_row + 1)
    }
    degraded = rows["OSPF degraded baseline"]
    assert "EXSTART/DROTHER" in degraded[0]
    assert degraded[1] == "degraded"
    assert "embedded projection; integrity unverified" in degraded[2]
    assert "routing_neighbors.distA.ospf" in degraded[3]

    not_verified = rows["BGP receipt gap"]
    assert "ROUTING BASELINE NOT VERIFIED" in not_verified[0]
    assert not_verified[1] == "not_verified"


def test_nrfu_workbook_discloses_routing_evidence_custody_and_blockers():
    routing_cases = [
        {
            "id": "NRFU-W1-P3-046", "phase": 3, "scope": "per-device",
            "command": "show ip ospf neighbor",
            "expected": "PRE-CUTOVER DEGRADED — BLOCKER: one observed peer is EXSTART/DROTHER",
            "source_key": "routing_neighbors.distA.ospf",
            "evidence_state": "degraded", "projection_custody": "embedded_unverified",
        },
        {
            "id": "NRFU-W1-P3-047", "phase": 3, "scope": "per-device",
            "command": "show ip bgp summary",
            "expected": "ROUTING BASELINE NOT VERIFIED: current-run evidence is incomplete",
            "source_key": "protocol_assessability.rows[distA,BGP]",
            "evidence_state": "not_verified", "projection_custody": "embedded_unverified",
        },
    ]
    pack = {
        "waves": [{
            "wave_id": "Group 1",
            "devices": [{"host": "distA", "platform_dialect": "ios", "cases": routing_cases}],
        }],
        "summary": {
            "n_cases": 2, "n_waves": 1, "n_devices": 1,
            "by_phase": {3: 2}, "n_not_observed": 0, "n_human_executed": 0,
        },
        "banner": "Execute every NRFU case against its stated acceptance criterion.",
    }

    wb = Workbook()
    write_nrfu_commands_sheet(wb, pack)
    ws = wb["NRFU Commands"]

    assert [ws.cell(4, col).value for col in range(9, 13)] == [
        "Expected (from snapshot)", "Evidence state", "Projection custody", "Source key",
    ]
    assert str(ws.cell(1, 1).value).startswith("⚠ ")
    assert "Evidence-state blockers apply to every protocol" in ws.cell(1, 1).value
    assert [ws.cell(5, col).value for col in range(10, 13)] == [
        "degraded", "embedded_unverified", "routing_neighbors.distA.ospf",
    ]
    assert ws.cell(5, 9).fill.fgColor.rgb.endswith("F4CCCC")
    assert ws.cell(5, 10).fill.fgColor.rgb.endswith("F4CCCC")
    assert ws.cell(6, 9).fill.fgColor.rgb.endswith("FFF2CC")
    assert ws.cell(6, 11).fill.fgColor.rgb.endswith("FFF2CC")


def test_runbook_keeps_every_routing_blocker_outside_ordinary_cap(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "routing-runbook.docx"
    write_runbook_docx(str(output), _snapshot(), "Routing Surface Fixture")
    document = docx.Document(str(output))
    tables = _tables_with_header(
        document, ["Device", "Category", "Check", "Command", "Observed baseline / acceptance"])
    assert len(tables) == 1
    rows = [[cell.text for cell in row.cells] for row in tables[0].rows[1:]]

    # Fourteen ordinary rows plus every routing blocker, even though both blockers were source-ordered last.
    assert len(rows) == 16
    by_check = {row[2]: row for row in rows}
    assert {"OSPF degraded baseline", "BGP receipt gap"} <= set(by_check)
    for check in ("OSPF degraded baseline", "BGP receipt gap"):
        evidence = by_check[check][4]
        assert "Evidence state:" in evidence
        assert "Projection custody: embedded_unverified" in evidence
        assert "Source: routing_neighbors.distA.ospf" in evidence

    text = "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )
    assert "Any BASELINE NOT VERIFIED row is a blocker" in text
    assert "Every evidence-state blocker is retained above" in text


def test_mop_keeps_routing_blockers_and_their_capture_commands_outside_caps(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.mop import write_mop_docx

    output = tmp_path / "routing-mop.docx"
    write_mop_docx(str(output), _snapshot(), "Routing Surface Fixture")
    document = docx.Document(str(output))

    validation = _tables_with_header(
        document, ["Sev", "Category", "Device", "Check", "Command",
                   "Observed baseline / acceptance"])
    assert len(validation) == 1
    rows = [[cell.text for cell in row.cells] for row in validation[0].rows[1:]]
    assert len(rows) == 42  # 40 ordinary allowance + both routing blockers
    by_check = {row[3]: row for row in rows}
    assert {"OSPF degraded baseline", "BGP receipt gap"} <= set(by_check)
    assert "Evidence state: degraded" in by_check["OSPF degraded baseline"][5]
    assert "Projection custody: embedded_unverified" in by_check["BGP receipt gap"][5]
    assert "routing_neighbors.distA.ospf" in by_check["BGP receipt gap"][5]

    captures = _tables_with_header(document, ["Device", "Command to capture"])
    assert len(captures) == 1
    commands = [row.cells[1].text for row in captures[0].rows[1:]]
    assert len(commands) == 22  # 20 ordinary allowance + both routing-blocker commands
    assert "show ip ospf neighbor" in commands
    assert "show ip bgp summary" in commands

    paragraph_text = "\n".join(paragraph.text for paragraph in document.paragraphs)
    assert "Any BASELINE NOT VERIFIED row is a blocker" in paragraph_text
    assert "Every evidence-state blocker capture is retained above" in paragraph_text
    assert "Every FHRP VLAN gate and evidence-state blocker is retained above" in paragraph_text
    assert "every check meets its stated acceptance criterion" in paragraph_text
    assert "reproducing a degraded baseline" in paragraph_text
    assert "every check matches its captured baseline" not in paragraph_text
    assert "A captured degraded state is evidence to resolve or disposition" in paragraph_text
