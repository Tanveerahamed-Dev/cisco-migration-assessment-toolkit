"""Pre-Change Validation Certificate (roadmap C1): the shipped fib differential packaged as the PPDIOO
gate ARTIFACT — the offline peer of Forward Networks' "Predict" / NDI pre-change analysis.

Verdict doctrine under test (coverage-honest, exact):
  * any regression (newly-blocked flow / broken intent / violated invariant) -> FAIL;
  * no regressions but >0 inconclusive flows or not_evaluable invariants -> CONDITIONAL (never PASS
    with an open blind spot);
  * everything checkable checked and clean -> PASS;
  * nothing definitively checkable -> INDETERMINATE.
Every changed flow is cited before->after; every blind spot is NAMED.
"""
import json
import os
import subprocess
import sys

from openpyxl import load_workbook

from cisco_toolkit.html import write_diff_workbook
from cisco_toolkit.precert import (CERT_SHEET_HEADERS, CERT_SHEET_NAME, READINESS_SCHEMA,
                                   compute_precert, compute_readiness_freeze)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py")


# ---------------------------------------------------------------- fixtures
def _seg_held():
    """Segmentation posture with a real claim: an isolated app domain + a dedicated VRF."""
    return {"vrfs": [{"vrf": "(global)", "gateway_count": 1}, {"vrf": "PCI", "gateway_count": 2}],
            "domains": [{"domain": "payments", "tier": "on-air-critical", "isolated": True,
                         "exposure": "Every gateway has a dedicated VRF or a gateway ACL."}]}


def _clean_routes():
    """Fully-routed symmetric pair: every sampled inter-subnet flow resolves computed:reached."""
    return {"R1": [{"prefix": "10.1.1.0/24", "source": "connected"},
                   {"prefix": "10.0.12.0/30", "source": "connected"},
                   {"prefix": "10.2.2.0/24", "source": "ospf", "next_hop": "10.0.12.2"}],
            "R2": [{"prefix": "10.2.2.0/24", "source": "connected"},
                   {"prefix": "10.0.12.0/30", "source": "connected"},
                   {"prefix": "10.1.1.0/24", "source": "ospf", "next_hop": "10.0.12.1"}]}


def _snap(routes, segmentation=None, generated_at="2026-07-01T00:00:00"):
    s = {"devices": {h: {} for h in routes}, "interfaces": {h: {} for h in routes},
         "routes": routes, "generated_at": generated_at, "script_version": "V3.23.0"}
    if segmentation is not None:
        s["segmentation"] = segmentation
    return s


def _pass_pair():
    before = _snap(_clean_routes(), _seg_held())
    after = _snap(_clean_routes(), _seg_held(), generated_at="2026-07-02T00:00:00")
    return before, after


def _fail_pair():
    """The cutover removes R1's route to 10.2.2.0/24 -> a previously-working flow definitively drops."""
    before, after = _pass_pair()
    after["routes"] = {h: [r for r in rs if not (h == "R1" and r.get("source") == "ospf")]
                       for h, rs in after["routes"].items()}
    return before, after


def _conditional_routes():
    """R1 routes via an UNCOLLECTED next hop -> flows to 10.2.2.0/24 are lower-bound inconclusive, while
    the R1-local 10.1.1.0/24 <-> 10.3.3.0/24 flows stay computed:reached (so something IS checked)."""
    return {"R1": [{"prefix": "10.1.1.0/24", "source": "connected"},
                   {"prefix": "10.3.3.0/24", "source": "connected"},
                   {"prefix": "0.0.0.0/0", "source": "static", "next_hop": "10.99.99.1"}],
            "R2": [{"prefix": "10.2.2.0/24", "source": "connected"}]}


# ---------------------------------------------------------------- verdicts
def test_pass_requires_all_checked_and_clean():
    before, after = _pass_pair()
    cert = compute_precert(before, after)
    assert cert["schema"] == "precert/1"
    assert cert["verdict"] == "PASS"
    assert cert["flows"]["preserved"] > 0 and cert["flows"]["changed"] == [] and cert["flows"]["inconclusive"] == []
    assert all(s["held"] is True for s in cert["segmentation"])
    assert cert["blind_spots"] == [] and cert["regressions"] == []


def test_newly_blocked_flow_is_fail_and_cites_before_after():
    before, after = _fail_pair()
    cert = compute_precert(before, after)
    assert cert["verdict"] == "FAIL"
    blocked = [f for f in cert["flows"]["changed"] if f["regression"]]
    assert blocked, "the newly-blocked flow must appear under flows.changed"
    for f in blocked:
        assert f["before"].startswith("computed:reached")            # each changed flow carries BOTH statuses
        assert f["after"] == "computed:unreachable"
        assert f["src"] and f["dst"]
    assert any("newly blocked" in r for r in cert["regressions"])


def test_inconclusive_flows_make_conditional_never_pass():
    routes = _conditional_routes()
    before, after = _snap(routes, _seg_held()), _snap(routes, _seg_held())
    cert = compute_precert(before, after)
    assert cert["flows"]["preserved"] > 0                            # something WAS definitively checked
    assert len(cert["flows"]["inconclusive"]) > 0
    assert cert["verdict"] == "CONDITIONAL"                          # blind spots open -> never PASS
    assert cert["verdict"] != "PASS"
    # every inconclusive flow is a NAMED blind spot with a reason
    for f in cert["flows"]["inconclusive"]:
        assert f["src"] and f["dst"] and f["reason"]
    assert any("inconclusive" in b for b in cert["blind_spots"])


def test_pass_is_impossible_with_any_inconclusive_flow():
    routes = _conditional_routes()
    cert = compute_precert(_snap(routes, _seg_held()), _snap(routes, _seg_held()))
    assert len(cert["flows"]["inconclusive"]) > 0 and cert["verdict"] != "PASS"


def test_nothing_checkable_is_indeterminate():
    cert = compute_precert({}, {})
    assert cert["verdict"] == "INDETERMINATE"
    cert2 = compute_precert(_snap({}), _snap({}))
    assert cert2["verdict"] == "INDETERMINATE"
    assert not cert2["flows"]["assessed"]


def test_compute_precert_total_on_non_dict_snapshots():
    for bad in (None, [], 42, "s", True):
        assert compute_precert(bad, {})["verdict"] == "INDETERMINATE"
        assert compute_precert({}, bad)["verdict"] == "INDETERMINATE"


# ---------------------------------------------------------------- segmentation invariants
def test_violated_segmentation_invariant_is_fail():
    before, after = _pass_pair()
    after["segmentation"] = {"vrfs": [{"vrf": "(global)", "gateway_count": 3}],
                             "domains": [{"domain": "payments", "tier": "on-air-critical", "isolated": False,
                                          "exposure": "2 of 2 gateway(s) share the global VRF with no ACL."}]}
    cert = compute_precert(before, after)
    assert cert["verdict"] == "FAIL"
    violated = [s for s in cert["segmentation"] if s["held"] is False]
    assert violated and all(s["evidence"] for s in violated)         # each violation cites evidence
    assert any("invariant violated" in r for r in cert["regressions"])


def test_segmentation_not_evaluable_is_honest_and_blocks_pass():
    # flows all clean, but the AFTER snapshot has no segmentation data: the before invariants can NOT be
    # re-verified -> not_evaluable (a first-class abstention), and the verdict is CONDITIONAL, never PASS.
    before, after = _pass_pair()
    del after["segmentation"]
    cert = compute_precert(before, after)
    assert cert["verdict"] == "CONDITIONAL"
    ne = [s for s in cert["segmentation"] if s["held"] == "not_evaluable"]
    assert ne and all("after snapshot" in s["evidence"] for s in ne)
    assert any("not evaluable" in b for b in cert["blind_spots"])    # NAMED as a blind spot


def test_no_segmentation_data_anywhere_is_a_named_blind_spot():
    # neither snapshot carries segmentation: the certificate must say so (never silently certify posture)
    before, after = _pass_pair()
    del before["segmentation"]
    del after["segmentation"]
    cert = compute_precert(before, after)
    assert cert["verdict"] == "CONDITIONAL"                          # flows clean, segmentation is a blind spot
    assert any(s["held"] == "not_evaluable" for s in cert["segmentation"])
    assert any("segmentation" in b for b in cert["blind_spots"])


def test_vrf_disappearance_is_a_violated_invariant():
    before, after = _pass_pair()
    after["segmentation"] = {"vrfs": [{"vrf": "(global)", "gateway_count": 3}],
                             "domains": [{"domain": "payments", "tier": "on-air-critical", "isolated": True,
                                          "exposure": "Every gateway has a dedicated VRF or a gateway ACL."}]}
    cert = compute_precert(before, after)
    vrf_rows = [s for s in cert["segmentation"] if "VRF 'PCI'" in s["invariant"]]
    assert vrf_rows and vrf_rows[0]["held"] is False
    assert cert["verdict"] == "FAIL"


# ---------------------------------------------------------------- path intents
def test_broken_intent_is_fail_and_intents_pass_through():
    before, after = _fail_pair()
    intents = [{"id": "app-to-db", "src": "10.1.1.5", "dst": "10.2.2.9", "expect": "REACHES"}]
    cert = compute_precert(before, after, path_intents=intents)
    assert cert["verdict"] == "FAIL"
    rows = {r["id"]: r for r in cert["intents"]}
    assert rows["app-to-db"]["old_verdict"] == "pass" and rows["app-to-db"]["new_verdict"] == "fail"
    assert rows["app-to-db"]["regressed"] is True
    assert any("app-to-db" in r for r in cert["regressions"])


def test_intent_not_observed_after_change_is_a_blind_spot_not_a_regression():
    routes = _conditional_routes()
    before, after = _snap(routes, _seg_held()), _snap(routes, _seg_held())
    intents = [{"id": "dmz-isolated", "src": "10.1.1.5", "dst": "10.2.2.9", "expect": "ISOLATED"}]
    cert = compute_precert(before, after, path_intents=intents)
    rows = {r["id"]: r for r in cert["intents"]}
    assert rows["dmz-isolated"]["new_verdict"] == "not_observed"     # a lower bound can never PROVE isolation
    assert cert["verdict"] == "CONDITIONAL"                          # blind spot, not a fabricated FAIL
    assert any("dmz-isolated" in b for b in cert["blind_spots"])


def test_no_intents_means_empty_passthrough():
    before, after = _pass_pair()
    cert = compute_precert(before, after)
    assert cert["intents"] == []
    assert compute_precert(before, after, path_intents=None)["intents"] == []


# ---------------------------------------------------------------- stamps + JSON round-trip
def test_stamps_reference_both_snapshots():
    before, after = _pass_pair()
    cert = compute_precert(before, after)
    assert cert["stamps"]["before"]["generated_at"] == "2026-07-01T00:00:00"
    assert cert["stamps"]["after"]["generated_at"] == "2026-07-02T00:00:00"
    assert cert["stamps"]["before"]["n_devices"] == 2


def test_certificate_json_round_trips():
    for pair in (_pass_pair(), _fail_pair()):
        cert = compute_precert(*pair, path_intents=[{"id": "i1", "src": "10.1.1.5", "dst": "10.2.2.9",
                                                     "expect": "REACHES"}])
        assert json.loads(json.dumps(cert)) == cert                  # pure JSON types only (no tuples/sets)


# ---------------------------------------------------------------- diff-workbook sheet + CLI emission
def test_diff_workbook_gets_certificate_sheet_with_expected_headers(tmp_path):
    before, after = _fail_pair()
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(before, after, str(out))
    wb = load_workbook(str(out))
    assert CERT_SHEET_NAME in wb.sheetnames
    ws = wb[CERT_SHEET_NAME]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == list(CERT_SHEET_HEADERS)
    # the verdict row leads, and each changed flow is cited before -> after
    assert ws.cell(2, 1).value == "VERDICT" and ws.cell(2, 5).value == "FAIL"
    rows = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    blocked = [r for r in rows if r[4] == "NEWLY BLOCKED"]
    assert blocked and blocked[0][2].startswith("computed:reached") and blocked[0][3] == "computed:unreachable"


def test_certificate_sheet_names_blind_spots(tmp_path):
    routes = _conditional_routes()
    before, after = _snap(routes, _seg_held()), _snap(routes, _seg_held())
    out = tmp_path / "diff.xlsx"
    write_diff_workbook(before, after, str(out))
    ws = load_workbook(str(out))[CERT_SHEET_NAME]
    rows = [tuple(c.value for c in row) for row in ws.iter_rows(min_row=2)]
    assert any(r[0] == "Blind spot" and r[1] for r in rows)          # each blind spot NAMED on the sheet
    assert any(r[4] == "CONDITIONAL" for r in rows if r[0] == "VERDICT")


def test_compare_cli_emits_precert_json_next_to_diff_workbook(tmp_path):
    before, after = _fail_pair()
    old_p, new_p = tmp_path / "old.snapshot.json", tmp_path / "new.snapshot.json"
    old_p.write_text(json.dumps(before), encoding="utf-8")
    new_p.write_text(json.dumps(after), encoding="utf-8")
    out = tmp_path / "Diff.xlsx"
    intents = tmp_path / "intents.json"
    intents.write_text(json.dumps([{"id": "app-to-db", "src": "10.1.1.5", "dst": "10.2.2.9",
                                    "expect": "REACHES"}]), encoding="utf-8")
    proc = subprocess.run([sys.executable, SCRIPT, "--compare", str(old_p), str(new_p),
                           "--output", str(out), "--path-intents", str(intents)],
                          cwd=str(tmp_path), capture_output=True, text=True, timeout=300)
    assert proc.returncode == 0, f"--compare failed:\nSTDOUT\n{proc.stdout}\nSTDERR\n{proc.stderr}"
    assert out.exists()
    cert_path = tmp_path / "Diff.precert.json"
    assert cert_path.exists(), "the certificate JSON must be written next to the diff workbook"
    cert = json.loads(cert_path.read_text(encoding="utf-8"))
    assert cert["schema"] == "precert/1" and cert["verdict"] == "FAIL"
    assert any(r["id"] == "app-to-db" for r in cert["intents"])      # --path-intents honored in --compare
    assert CERT_SHEET_NAME in load_workbook(str(out)).sheetnames


# ---------------------------------------------------------------- single-snapshot readiness FREEZE
def _readiness_snap(groups, completeness=None):
    s = {"schema": "collect_parse_snapshot/1", "script_version": "V3.23.0",
         "generated_at": "2026-07-01T00:00:00", "collected_at": "2026-06-13T06:32:01",
         "devices": {f"d{i}": {} for i in range(3)}, "migration_readiness": groups}
    if completeness is not None:
        s["collection_completeness"] = completeness
    return s


def test_readiness_freeze_is_worst_of_and_coverage_honest():
    groups = [{"group": "G1", "readiness": "READY", "n_fail": 0, "n_warn": 0, "checks": []},
              {"group": "G2", "readiness": "CAUTION", "n_fail": 0, "n_warn": 1,
               "checks": [{"check": "STP consistency", "status": "warn"},
                          {"check": "Redundant uplinks", "status": "pass"}]}]
    cc = {"summary": {"inventory": 5, "complete": 3, "partial": 0, "not_collected": 2},
          "devices": [{"host": "y-uncollected", "status": "not collected"},
                      {"host": "x-uncollected", "status": "not collected"},
                      {"host": "z-ok", "status": "complete"}]}
    cert = compute_readiness_freeze(_readiness_snap(groups, cc), mode="shadow")
    assert cert["schema"] == READINESS_SCHEMA and cert["mode"] == "shadow"
    assert cert["verdict"] == "CAUTION"                               # worst-of (never the better READY)
    assert cert["readiness_distribution"] == {"READY": 1, "CAUTION": 1, "NOT READY": 0, "UNRATED": 0}
    assert sum(cert["readiness_distribution"].values()) == cert["n_groups"]   # no group falls out
    assert cert["groups"][1]["blocking"] == ["STP consistency"]      # only the fail/warn check is blocking
    assert cert["coverage"]["not_collected"] == 2
    assert cert["coverage"]["not_collected_hosts"] == ["x-uncollected", "y-uncollected"]  # NAMED + sorted
    assert cert["blind_spots"] and "not collected" in cert["blind_spots"][0].lower()
    assert cert["prediction_hash"].startswith("sha256:")


def test_readiness_freeze_worst_of_promotes_not_ready():
    groups = [{"group": "G1", "readiness": "CAUTION", "checks": []},
              {"group": "G2", "readiness": "NOT READY", "n_fail": 1, "checks": []}]
    assert compute_readiness_freeze(_readiness_snap(groups))["verdict"] == "NOT READY"


def test_readiness_freeze_never_certifies_ready_over_an_unassessed_group():
    """A move-group whose readiness label is absent or unrecognised used to be silently DROPPED:
    `_canon_readiness` returned '', `_READINESS_RANK.get('', -1)` never beat the seed and '' is not a
    key of the distribution — so the group inflated `n_groups` and contributed to neither the
    worst-of verdict, the distribution, nor `blind_spots`. Measured before the fix: 2 READY + 1
    never-assessed + 1 typo'd group certified **READY** over 4 groups with an empty blind-spot list,
    and `prediction_hash` made that false READY the tamper-evident pre-cutover record."""
    groups = [{"group": "G1", "readiness": "READY", "checks": []},
              {"group": "G2", "readiness": "READY", "checks": []},
              {"group": "G3-never-assessed", "checks": []},               # no readiness key at all
              {"group": "G4-typo", "readiness": "Not-Ready!", "checks": []}]
    cert = compute_readiness_freeze(_readiness_snap(groups), mode="real")
    assert cert["verdict"] == "INDETERMINATE"                             # NOT READY-over-blind-spots
    d = cert["readiness_distribution"]
    assert d == {"READY": 2, "CAUTION": 0, "NOT READY": 0, "UNRATED": 2}
    assert sum(d.values()) == cert["n_groups"] == 4                       # the distribution accounts
    blind = " ".join(cert["blind_spots"])                                 # ...and both are NAMED
    assert "G3-never-assessed" in blind and "no readiness label" in blind
    assert "G4-typo" in blind and "Not-Ready!" in blind
    assert [g["readiness"] for g in cert["groups"]] == ["READY", "READY", "UNRATED", "UNRATED"]
    assert "UNRATED" in cert["verdict_note"]


def test_readiness_freeze_unrated_group_does_not_mask_a_worse_rating():
    """The verdict is still the WORST of what WAS assessed — an unrated group withholds READY, it
    does not erase a NOT READY (a freeze must not read better OR emptier than its evidence)."""
    groups = [{"group": "G1", "readiness": "NOT READY", "n_fail": 1, "checks": []},
              {"group": "G2", "checks": []}]
    cert = compute_readiness_freeze(_readiness_snap(groups))
    assert cert["verdict"] == "NOT READY"
    assert cert["readiness_distribution"]["UNRATED"] == 1 and cert["blind_spots"]
    # all-unrated asserts nothing at all
    allbad = compute_readiness_freeze(_readiness_snap([{"group": "X"}, {"group": "Y"}]))
    assert allbad["verdict"] == "INDETERMINATE" and allbad["n_groups"] == 2
    assert allbad["readiness_distribution"]["UNRATED"] == 2


def test_readiness_freeze_empty_is_indeterminate_never_ready():
    cert = compute_readiness_freeze(_readiness_snap([]), mode="shadow")
    assert cert["verdict"] == "INDETERMINATE" and cert["n_groups"] == 0  # nothing to score != READY


def test_readiness_freeze_hash_is_deterministic_and_tamper_evident():
    g1 = [{"group": "G1", "readiness": "CAUTION", "n_warn": 1, "checks": []}]
    g2 = [{"group": "G1", "readiness": "NOT READY", "n_fail": 1, "checks": []}]
    h1 = compute_readiness_freeze(_readiness_snap(g1))["prediction_hash"]
    h1b = compute_readiness_freeze(_readiness_snap(g1))["prediction_hash"]
    h2 = compute_readiness_freeze(_readiness_snap(g2))["prediction_hash"]
    assert h1 == h1b                                                 # deterministic: same snapshot -> same hash
    assert h1 != h2                                                  # tamper-evident: changed prediction -> new hash


def test_readiness_freeze_shadow_and_real_share_the_prediction_hash():
    g = [{"group": "G1", "readiness": "CAUTION", "n_warn": 1, "checks": []}]
    shadow = compute_readiness_freeze(_readiness_snap(g), mode="shadow")
    real = compute_readiness_freeze(_readiness_snap(g), mode="real")
    assert shadow["prediction_hash"] == real["prediction_hash"]      # mode is metadata; prediction is identical
    assert (shadow["mode"], real["mode"]) == ("shadow", "real")


def test_readiness_freeze_total_on_bad_input():
    for bad in (None, [], 42, "s", {"migration_readiness": "nope"}):
        assert compute_readiness_freeze(bad)["verdict"] == "INDETERMINATE"


def test_readiness_freeze_json_round_trips():
    g = [{"group": "G1", "readiness": "NOT READY", "n_fail": 2, "checks":
          [{"check": "Gateway redundancy", "status": "fail"}]}]
    cert = compute_readiness_freeze(_readiness_snap(g), mode="real")
    assert json.loads(json.dumps(cert)) == cert                      # pure JSON types only


def test_main_unreadable_snapshot_returns_2(capsys):
    """The readiness-freeze CLI (precert.py:387) fails soft on a bad path — exit 2, a clear message,
    never a traceback."""
    from cisco_toolkit import precert as P
    assert P.main(["no-such-snapshot.json"]) == 2
    assert "cannot read snapshot" in capsys.readouterr().out
