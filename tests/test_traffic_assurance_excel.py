"""Canonical Traffic Assurance workbook projection contract."""
import io

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook  # noqa: E402

from cisco_toolkit import traffic_assurance  # noqa: E402
from cisco_toolkit.excel import (  # noqa: E402
    TRAFFIC_ASSURANCE_SHEET_NAME,
    write_traffic_assurance_sheet,
)


def _result(flow_id: str, verdict: str, *, path_forward: str = "proven") -> dict:
    return {
        "schema": "traffic_assurance/1",
        "owner": "cisco_toolkit.traffic_assurance.assess_flow",
        "intent": {
            "id": flow_id,
            "src": "10.0.0.10",
            "dst": "10.0.1.20",
            "protocol": "tcp",
            "src_port": 49152,
            "dst_port": 443,
            "expected": "permit",
            "return_required": True,
            "required_mtu": 1500,
            "vrf": None,
        },
        "valid": True,
        "supported": True,
        "custody_trust": "current_run_verified",
        "verdict": verdict,
        "verdict_reasons": [f"producer says {verdict}"],
        "unsupported_semantics": [],
        "dimensions": {
            "path": {
                "forward": {"verdict": path_forward},
                "reverse": {"verdict": "not_observed"},
                "rpf_verdict": "INDETERMINATE",
                "symmetric": False,
            },
            "policy": {
                "forward": {"verdict": "refuted"},
                "reverse": {"verdict": "indeterminate"},
            },
            "mtu": {
                "forward": {"verdict": "proven"},
                "reverse": {"verdict": "not_requested"},
            },
            "ecmp": {
                "forward": {"verdict": "indeterminate"},
                "reverse": {"verdict": "not_observed"},
            },
        },
        "failure": {"requested": True, "status": "coverage_lost", "verdict": "indeterminate"},
        "nrfu_test_ids": [f"NRFU-{flow_id}-FORWARD", f"NRFU-{flow_id}-RETURN"],
        "limitations": ["synthetic only"],
    }


def _sheet_rows(payload):
    wb = Workbook()
    write_traffic_assurance_sheet(wb, payload)
    ws = wb[TRAFFIC_ASSURANCE_SHEET_NAME]
    headers = {cell.value: cell.column for cell in ws[1]}
    return wb, ws, headers


def test_projects_canonical_values_and_summary_without_recalculation(monkeypatch):
    def _must_not_run(*_args, **_kwargs):
        raise AssertionError("workbook projection called a traffic-assurance engine")

    monkeypatch.setattr(traffic_assurance, "assess_flow", _must_not_run)
    monkeypatch.setattr(traffic_assurance, "assess_flows", _must_not_run)
    payload = {
        "schema": "traffic_assurance_set/1",
        "owner": "cisco_toolkit.traffic_assurance.assess_flow",
        # Deliberately contradict the two rows. The producer owns this summary; the renderer must not recount it.
        "summary": {
            "n": 73,
            "proven": 71,
            "refuted": 1,
            "not_observed": 9,
            "indeterminate": 4,
            "invalid": 8,
        },
        # Deliberately non-alphabetical. Canonical declared order must survive the projection.
        "results": [_result("z-last", "refuted", path_forward="not_observed"),
                    _result("a-first", "proven", path_forward="indeterminate")],
    }

    _wb, ws, columns = _sheet_rows(payload)

    assert len(columns) == ws.max_column  # every projected value has an explicitly named allowlist column
    assert [ws.cell(row, columns["Flow ID"]).value for row in (2, 3)] == ["z-last", "a-first"]
    assert ws.cell(2, columns["Canonical Summary"]).value == (
        "n=73; proven=71; refuted=1; not_observed=9; indeterminate=4; invalid=8"
    )
    assert ws.cell(2, columns["Overall Verdict"]).value == "refuted"
    assert ws.cell(2, columns["Path Forward"]).value == "not_observed"
    assert ws.cell(2, columns["Policy Forward"]).value == "refuted"
    assert ws.cell(2, columns["MTU Forward"]).value == "proven"
    assert ws.cell(2, columns["ECMP Forward"]).value == "indeterminate"
    assert ws.cell(2, columns["Failure Verdict"]).value == "indeterminate"
    assert ws.cell(2, columns["Return Required"]).value == "true"
    assert ws.cell(2, columns["Custody Trust"]).value == "current_run_verified"


@pytest.mark.parametrize(
    ("payload", "expected_state", "marker"),
    [
        (None, "not_supplied", "[NOT ASSESSED]"),
        ({"state": "loading", "private_error": "DO-NOT-PRINT"}, "loading", "[LOADING]"),
        ({"state": "error", "error": "DO-NOT-PRINT"}, "error", "[ERROR]"),
        ({"schema": "wrong/9", "results": []}, "error", "[ERROR]"),
        ({"schema": "traffic_assurance_set/1", "results": "not-a-list"}, "error", "[ERROR]"),
        ({
            "schema": "traffic_assurance_set/1",
            "owner": "cisco_toolkit.traffic_assurance.assess_flow",
            "results": [],
            "summary": {"n": 0, "proven": 0, "refuted": 0, "not_observed": 0,
                        "indeterminate": 0, "invalid": 0},
        }, "empty", "[EMPTY]"),
    ],
)
def test_loading_empty_error_and_absent_states_are_explicit(payload, expected_state, marker):
    _wb, ws, columns = _sheet_rows(payload)
    assert ws.max_row == 2
    assert ws.cell(2, columns["Projection State"]).value == expected_state
    assert marker in ws.cell(2, columns["Coverage Disclosure"]).value
    assert "DO-NOT-PRINT" not in "\n".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
    )


def test_malformed_result_is_disclosed_in_place_not_dropped():
    payload = {
        "schema": "traffic_assurance_set/1",
        "owner": "cisco_toolkit.traffic_assurance.assess_flow",
        "summary": {"n": 2, "proven": 0, "refuted": 0, "not_observed": 0,
                    "indeterminate": 0, "invalid": 2},
        "results": [42, _result("kept-after-malformed", "indeterminate")],
    }
    _wb, ws, columns = _sheet_rows(payload)
    assert ws.max_row == 3
    assert ws.cell(2, columns["Projection State"]).value == "error"
    assert ws.cell(2, columns["Flow ID"]).value == "[MALFORMED RESULT]"
    assert ws.cell(3, columns["Flow ID"]).value == "kept-after-malformed"


def test_projection_is_allowlisted_deterministic_and_formula_safe_after_save_reload():
    secret = "RAW-ACL secret-password-never-render"
    result = _result('=HYPERLINK("http://attacker.invalid","x")', "indeterminate")
    result["raw_capture"] = secret
    result["dimensions"]["path"]["forward"]["raw_capture"] = secret
    result["failure"]["cutover_evidence"] = {"private_path": secret}
    result["unsupported_semantics"] = [{"raw": secret}]
    payload = {
        "schema": "traffic_assurance_set/1",
        "owner": "cisco_toolkit.traffic_assurance.assess_flow",
        "summary": {"n": 1, "proven": 0, "refuted": 0, "not_observed": 0,
                    "indeterminate": 1, "invalid": 0},
        "results": [result],
        "unknown_extension": {"raw": secret},
    }

    wb1, ws1, columns = _sheet_rows(payload)
    wb2, ws2, _columns2 = _sheet_rows(payload)
    assert [[cell.value for cell in row] for row in ws1.iter_rows()] == [
        [cell.value for cell in row] for row in ws2.iter_rows()
    ]
    assert secret not in "\n".join(
        str(cell.value) for row in ws1.iter_rows() for cell in row if cell.value is not None
    )
    assert ws1.cell(2, columns["Unsupported Semantics"]).value == "[MALFORMED SCALAR]"

    buffer = io.BytesIO()
    wb1.save(buffer)
    buffer.seek(0)
    saved = load_workbook(buffer)
    cell = saved[TRAFFIC_ASSURANCE_SHEET_NAME].cell(2, columns["Flow ID"])
    assert cell.data_type == "s"
    assert str(cell.value).startswith("'=")
    assert secret not in "\n".join(
        str(saved_cell.value)
        for row in saved[TRAFFIC_ASSURANCE_SHEET_NAME].iter_rows()
        for saved_cell in row
        if saved_cell.value is not None
    )


def test_missing_dimensions_render_not_observed_instead_of_blank_or_green():
    result = _result("missing-dimensions", "indeterminate")
    result["dimensions"] = {}
    payload = {
        "schema": "traffic_assurance_set/1",
        "owner": "cisco_toolkit.traffic_assurance.assess_flow",
        "summary": {"n": 1, "proven": 0, "refuted": 0, "not_observed": 0,
                    "indeterminate": 1, "invalid": 0},
        "results": [result],
    }
    _wb, ws, columns = _sheet_rows(payload)
    for header in ("Path Forward", "Policy Return", "MTU Forward", "ECMP Return"):
        assert ws.cell(2, columns[header]).value == "[NOT OBSERVED]"
