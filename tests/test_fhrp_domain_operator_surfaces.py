"""Operator projections for the closed FHRP redundancy-domain receipt."""

from __future__ import annotations

import json
from pathlib import Path
import shutil
import subprocess

import pytest
from openpyxl import Workbook

from cisco_toolkit.capture_integrity import compute_capture_integrity_from_paths
from cisco_toolkit.excel import (
    FHRP_CONSISTENCY_SHEET_NAME,
    FHRP_REDUNDANCY_DOMAIN_SHEET_NAME,
    compute_fhrp_consistency,
    write_fhrp_consistency_sheet,
    write_fhrp_redundancy_domain_sheet,
)
from cisco_toolkit.fhrp_redundancy import (
    _LIMITATIONS,
    _acceptance,
    _baseline_digest,
    _check,
    _domain_semantics,
    _finding,
    _summary,
    _why,
    compute_fhrp_redundancy_domain_baseline,
    embedded_fhrp_redundancy_domain_baseline,
    validate_fhrp_redundancy_domain_baseline,
)
from cisco_toolkit.fhrp_intent import compute_fhrp_configured_group_baseline
from cisco_toolkit.model import InterfaceData
from cisco_toolkit.runbook import write_runbook_docx


ROOT = Path(__file__).resolve().parents[1]
EXPLORER = ROOT / "cisco_toolkit" / "blast_radius_explorer.html"
NODE = shutil.which("node")
BLOCKER_STATUSES = ("degraded", "review", "not_verified")


def _member_row(index: int, side: str, status: str, *, vlan: int, vrf: str,
                subnet: str, domain_key: str, acceptance: str) -> tuple[dict, dict]:
    host = f"{status}-{index:03d}-{side}"
    interface = f"Vlan{vlan}"
    ip_octet = 2 if side == "a" else 3
    svi_ip = f"10.{index // 200}.{index % 200}.{ip_octet}/24"
    virtual_ip = f"10.{index // 200}.{index % 200}.1"
    role = "ACTIVE" if side == "a" else "STANDBY"
    candidate_key = f"HSRP:10:{virtual_ip}"
    finding_code = {
        "degraded": "local_group_degraded",
        "review": "local_group_review",
        "not_verified": "local_group_not_verified",
    }.get(status)
    findings = [_finding(finding_code)] if finding_code else []
    row = {
        "switch": host,
        "interface": interface,
        "svi_ip": svi_ip,
        "vlan": vlan,
        "vrf": vrf,
        "subnet": subnet,
        "domain_key": domain_key,
        "candidate_key": candidate_key,
        "participation": "positive",
        "protocol": "HSRP",
        "group": "10",
        "virtual_ip": virtual_ip,
        "role": role,
        "status": status,
        "check": _check(vlan, vrf, subnet),
        "command": "show standby brief",
        "acceptance": acceptance,
        "why": findings[0]["issue"] if findings else "Bounded candidate composition assessed.",
        "source_key": (
            f"fhrp_configured_group_baseline.rows[{host},HSRP,{interface},10] + "
            f"interfaces[{host},{interface}].svi_ip/vrf"
        ),
        "projection_custody": "embedded_unverified",
        "findings": findings,
    }
    member = {
        "switch": host,
        "interface": interface,
        "svi_ip": svi_ip,
        "participation": "positive",
        "protocol": "HSRP",
        "group": "10",
        "virtual_ip": virtual_ip,
        "role": role,
        "local_status": status,
        "source_group_key": f"HSRP:{interface}:10",
        "projection_custody": "embedded_unverified",
        "findings": findings,
    }
    return row, member


def _domain(index: int, status: str) -> tuple[dict, list[dict], dict | None]:
    vlan = 100 + index
    vrf = f"tenant-{index % 3}"
    subnet = f"10.{index // 200}.{index % 200}.0/24"
    domain_key = f"vlan={vlan}|vrf={vrf}|subnet={subnet}"
    rows_and_members = [
        _member_row(
            index, side, status, vlan=vlan, vrf=vrf, subnet=subnet,
            domain_key=domain_key, acceptance="",
        )
        for side in ("a", "b")
    ]
    rows = [pair[0] for pair in rows_and_members]
    members = [pair[1] for pair in rows_and_members]
    semantics = _domain_semantics(members, subnet)
    assert semantics["status"] == status
    acceptance = _acceptance(
        status, vlan=vlan, vrf=vrf, subnet=subnet,
        member_count=semantics["member_count"],
        candidate_count=semantics["candidate_count"],
    )
    domain_findings = semantics["findings"]
    why = _why(domain_findings)
    for row, member in zip(rows, members):
        row["acceptance"] = acceptance
        row["why"] = why
        row["findings"] = [dict(finding) for finding in domain_findings]
        member["findings"] = [dict(finding) for finding in domain_findings]
    domain = {
        "vlan": vlan,
        "vrf": vrf,
        "subnet": subnet,
        "domain_key": domain_key,
        "status": status,
        "assessed": status in {"assessed", "degraded"},
        "member_count": semantics["member_count"],
        "participant_count": semantics["participant_count"],
        "leader_count": semantics["leader_count"],
        "backup_count": semantics["backup_count"],
        "protocol": semantics["protocol"],
        "group": semantics["group"],
        "virtual_ip": semantics["virtual_ip"],
        "members": members,
        "findings": domain_findings,
        "acceptance": acceptance,
    }
    top_finding = (
        {**domain_findings[0], "domain_key": domain_key}
        if domain_findings else None
    )
    return domain, rows, top_finding


def _baseline(*, blocker_domains: int = 30, ordinary_domains: int = 30,
              source_valid: bool = True) -> dict:
    domains: list[dict] = []
    rows: list[dict] = []
    findings: list[dict] = []
    for index in range(blocker_domains):
        domain, domain_rows, finding = _domain(
            index, BLOCKER_STATUSES[index % len(BLOCKER_STATUSES)])
        domains.append(domain)
        rows.extend(domain_rows)
        assert finding is not None
        findings.append(finding)
    for index in range(blocker_domains, blocker_domains + ordinary_domains):
        domain, domain_rows, finding = _domain(index, "assessed")
        domains.append(domain)
        rows.extend(domain_rows)
        assert finding is None
    status_set = {domain["status"] for domain in domains}
    if not source_valid:
        verdict, assessed = "INDETERMINATE", False
    elif "degraded" in status_set:
        verdict = "BLOCKED"
        assessed = not ({"review", "not_verified"} & status_set)
    elif {"review", "not_verified"} & status_set:
        verdict, assessed = "INDETERMINATE", False
    elif domains:
        verdict, assessed = "CLEAR", True
    else:
        verdict, assessed = "NOT_APPLICABLE", False
    receipt = {
        "schema": "fhrp_redundancy_domain_baseline/1",
        "scope": {
            "domain_identity": "vlan/normalized-vrf/observed-ipv4-subnet",
            "candidate_identity": "protocol/group/virtual-ip",
            "upstream_schema": "fhrp_configured_group_baseline/1",
        },
        "verdict": verdict,
        "assessed": assessed,
        "projection_custody": "embedded_unverified",
        "source_receipt": {
            "schema": "fhrp_configured_group_baseline/1",
            "valid": source_valid,
            "source_bound": False,
            "configured_baseline_sha256": "a" * 64 if source_valid else "",
            "svi_projection_sha256": "b" * 64 if source_valid else "",
        },
        "rows": rows,
        "domains": domains,
        "findings": findings,
        "summary": _summary(domains, rows),
        "limitations": list(_LIMITATIONS),
    }
    receipt["summary"]["baseline_sha256"] = _baseline_digest(receipt)
    assert validate_fhrp_redundancy_domain_baseline(receipt)["valid"] is True
    return receipt


def _trusted_current_review(tmp_path: Path) -> tuple[dict, dict]:
    hsrp_header = (
        "Interface   Grp  Pri P State    Active          Standby         Virtual IP\n"
    )
    hosts = {
        "edge-a": {
            "config": (
                "hostname edge-a\ninterface Vlan10\n ip address 10.0.10.2 255.255.255.0\n"
                " standby 10 ip 10.0.10.1\n standby 10 priority 110\nend\n"
            ),
            "standby": hsrp_header + (
                "Vl10        10   110 P Active   local           10.0.10.3       10.0.10.1\n"
            ),
            "behavior": "HSRP grp 10 Active VIP 10.0.10.1",
        },
        "edge-independent": {
            "config": (
                "hostname edge-independent\ninterface Vlan10\n"
                " ip address 10.0.10.3 255.255.255.0\nend\n"
            ),
            "standby": "No standby groups configured\n",
            "behavior": "",
        },
    }
    paths: dict[str, dict[str, str]] = {}
    for host, spec in hosts.items():
        captures = {
            "show running-config": spec["config"],
            "show standby brief": spec["standby"],
            "show vrrp brief": "No VRRP groups configured\n",
            "show glbp brief": "No GLBP groups configured\n",
        }
        paths[host] = {}
        for index, (command, body) in enumerate(captures.items(), 1):
            path = tmp_path / f"{host}-{index}.txt"
            path.write_text(body, encoding="utf-8")
            paths[host][command] = str(path)
    integrity = compute_capture_integrity_from_paths(paths)
    configured = compute_fhrp_configured_group_baseline(
        paths, integrity, {host: {"platform": "ios"} for host in hosts},
    )
    interfaces = {
        host: {"Vlan10": InterfaceData(
            port="Vlan10",
            svi_ip=("10.0.10.2/24" if host == "edge-a" else "10.0.10.3/24"),
            hsrp_behavior=spec["behavior"],
        )}
        for host, spec in hosts.items()
    }
    current = compute_fhrp_redundancy_domain_baseline(interfaces, configured)
    view = validate_fhrp_redundancy_domain_baseline(
        current, require_current_run=True)
    assert view["valid"] is True and view["source_bound"] is True
    return current, interfaces


def _mixed_participation_baseline(participation: str) -> dict:
    """Return one exact domain with a positive member and one bounded nonparticipant."""

    assert participation in {"nonparticipant", "not_verified"}
    receipt = _baseline(blocker_domains=0, ordinary_domains=1)
    domain = receipt["domains"][0]
    member = domain["members"][1]
    row = receipt["rows"][1]
    member.update({
        "participation": participation,
        "protocol": "",
        "group": "",
        "virtual_ip": "",
        "role": "",
        "local_status": "review" if participation == "nonparticipant" else "not_verified",
        "source_group_key": "",
    })
    row.update({
        "participation": participation,
        "candidate_key": "",
        "protocol": "",
        "group": "",
        "virtual_ip": "",
        "role": "",
        "command": f"show running-config interface {row['interface']}",
        "source_key": (
            f"interfaces[{row['switch']},{row['interface']}].svi_ip/vrf + "
            f"fhrp_configured_group_baseline.coverage[{row['switch']},HSRP|VRRP|GLBP]"
        ),
    })
    semantics = _domain_semantics(domain["members"], domain["subnet"])
    acceptance = _acceptance(
        semantics["status"], vlan=domain["vlan"], vrf=domain["vrf"],
        subnet=domain["subnet"], member_count=semantics["member_count"],
        candidate_count=semantics["candidate_count"],
    )
    for field in (
        "status", "member_count", "participant_count", "leader_count",
        "backup_count", "protocol", "group", "virtual_ip",
    ):
        domain[field] = semantics[field]
    domain["assessed"] = semantics["status"] in {"assessed", "degraded"}
    domain["findings"] = [dict(finding) for finding in semantics["findings"]]
    domain["acceptance"] = acceptance
    for nested, flat in zip(domain["members"], receipt["rows"]):
        nested["findings"] = [dict(finding) for finding in semantics["findings"]]
        flat["status"] = semantics["status"]
        flat["findings"] = [dict(finding) for finding in semantics["findings"]]
        flat["acceptance"] = acceptance
        flat["why"] = _why(semantics["findings"])
    receipt["findings"] = [
        {**finding, "domain_key": domain["domain_key"]}
        for finding in semantics["findings"]
    ]
    receipt["verdict"] = "INDETERMINATE"
    receipt["assessed"] = False
    receipt["summary"] = _summary(receipt["domains"], receipt["rows"])
    receipt["summary"]["baseline_sha256"] = _baseline_digest(receipt)
    assert validate_fhrp_redundancy_domain_baseline(receipt)["valid"] is True
    return receipt


def _unobserved_subnet_baseline() -> dict:
    receipt = _baseline(blocker_domains=0, ordinary_domains=1)
    domain = receipt["domains"][0]
    domain["subnet"] = "(subnet-unobserved)"
    domain["domain_key"] = (
        f"vlan={domain['vlan']}|vrf={domain['vrf']}|subnet=(subnet-unobserved)"
    )
    for member, row in zip(domain["members"], receipt["rows"]):
        member["svi_ip"] = ""
        row.update({
            "svi_ip": "",
            "subnet": domain["subnet"],
            "domain_key": domain["domain_key"],
            "check": _check(domain["vlan"], domain["vrf"], domain["subnet"]),
        })
    semantics = _domain_semantics(domain["members"], domain["subnet"])
    acceptance = _acceptance(
        semantics["status"], vlan=domain["vlan"], vrf=domain["vrf"],
        subnet=domain["subnet"], member_count=semantics["member_count"],
        candidate_count=semantics["candidate_count"],
    )
    for field in (
        "status", "member_count", "participant_count", "leader_count",
        "backup_count", "protocol", "group", "virtual_ip",
    ):
        domain[field] = semantics[field]
    domain["assessed"] = False
    domain["findings"] = [dict(finding) for finding in semantics["findings"]]
    domain["acceptance"] = acceptance
    for member, row in zip(domain["members"], receipt["rows"]):
        member["findings"] = [dict(finding) for finding in semantics["findings"]]
        row["status"] = semantics["status"]
        row["findings"] = [dict(finding) for finding in semantics["findings"]]
        row["acceptance"] = acceptance
        row["why"] = _why(semantics["findings"])
    receipt["findings"] = [
        {**finding, "domain_key": domain["domain_key"]}
        for finding in semantics["findings"]
    ]
    receipt["verdict"] = "INDETERMINATE"
    receipt["assessed"] = False
    receipt["summary"] = _summary(receipt["domains"], receipt["rows"])
    receipt["summary"]["baseline_sha256"] = _baseline_digest(receipt)
    assert validate_fhrp_redundancy_domain_baseline(receipt)["valid"] is True
    return receipt


def _hsrp_nonbackup_baseline(role: str) -> dict:
    assert role in {"LISTEN", "SPEAK"}
    receipt = _baseline(blocker_domains=0, ordinary_domains=1)
    domain = receipt["domains"][0]
    domain["members"][1]["role"] = role
    receipt["rows"][1]["role"] = role
    semantics = _domain_semantics(domain["members"], domain["subnet"])
    assert semantics["status"] == "review" and semantics["backup_count"] == 0
    acceptance = _acceptance(
        semantics["status"], vlan=domain["vlan"], vrf=domain["vrf"],
        subnet=domain["subnet"], member_count=semantics["member_count"],
        candidate_count=semantics["candidate_count"],
    )
    for field in (
        "status", "member_count", "participant_count", "leader_count",
        "backup_count", "protocol", "group", "virtual_ip",
    ):
        domain[field] = semantics[field]
    domain["assessed"] = False
    domain["findings"] = [dict(finding) for finding in semantics["findings"]]
    domain["acceptance"] = acceptance
    for member, row in zip(domain["members"], receipt["rows"]):
        member["findings"] = [dict(finding) for finding in semantics["findings"]]
        row["status"] = semantics["status"]
        row["findings"] = [dict(finding) for finding in semantics["findings"]]
        row["acceptance"] = acceptance
        row["why"] = _why(semantics["findings"])
    receipt["findings"] = [
        {**finding, "domain_key": domain["domain_key"]}
        for finding in semantics["findings"]
    ]
    receipt["verdict"] = "INDETERMINATE"
    receipt["assessed"] = False
    receipt["summary"] = _summary(receipt["domains"], receipt["rows"])
    receipt["summary"]["baseline_sha256"] = _baseline_digest(receipt)
    assert validate_fhrp_redundancy_domain_baseline(receipt)["valid"] is True
    return receipt


def _sheet_text(ws) -> str:
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    )


def _docx_text(path: Path) -> str:
    docx = pytest.importorskip("docx")
    document = docx.Document(str(path))
    return "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )


def _explorer_result(baseline: object, host: str | None = None) -> dict:
    assert NODE is not None
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _FHRP_GROUP_SCOPE=');
const end=html.indexOf('const _BGP_INTENT_SCOPE=',start);
if(start<0||end<0)throw new Error('FHRP projection block not found');
const input=JSON.parse(fs.readFileSync(0,'utf8'));
let SNAP={fhrp_redundancy_domain_baseline:input.baseline};
function esc(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
process.stdout.write(JSON.stringify({
  valid:fhrpRedundancyDomainReceiptValid(input.baseline),
  rendered:fhrpRedundancyDomainSection(input.host),
}));
"""
    proc = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        input=json.dumps({"baseline": baseline, "host": host}),
        timeout=30,
    )
    return json.loads(proc.stdout)


def _explorer_member_cap_probe(baseline: dict) -> dict:
    assert NODE is not None
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _FHRP_GROUP_SCOPE=');
const end=html.indexOf('const _BGP_INTENT_SCOPE=',start);
const baseline=JSON.parse(fs.readFileSync(0,'utf8'));
let SNAP={};
function esc(v){return String(v??'');}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
baseline.domains=[{members:new Array(10001)},{members:new Array(10000)}];
let deepValidatorReached=false,threw=false,valid=false;
_fhrpDomainRecord=()=>{deepValidatorReached=true;throw new Error('deep walk reached');};
try{valid=fhrpRedundancyDomainReceiptValid(baseline);}catch(_error){threw=true;}
process.stdout.write(JSON.stringify({valid,deepValidatorReached,threw}));
"""
    proc = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        input=json.dumps(baseline),
        timeout=30,
    )
    return json.loads(proc.stdout)


def test_workbook_is_uncapped_and_keeps_exact_domain_candidate_member_visibility():
    receipt = _baseline()
    wb = Workbook()
    write_fhrp_redundancy_domain_sheet(wb, receipt)
    ws = wb[FHRP_REDUNDANCY_DOMAIN_SHEET_NAME]
    text = _sheet_text(ws)

    assert ws.max_row == 4 + 120
    assert ws.freeze_panes == "A5"
    assert "degraded-027-b" in text
    assert "review-028-b" in text
    assert "not_verified-029-b" in text
    assert "assessed-059-b" in text
    assert "vlan=159|vrf=tenant-2|subnet=10.0.59.0/24" in text
    assert "HSRP:10:10.0.59.1" in text
    assert "embedded_unverified" in text
    assert "current normalized in-scope SVI members" in text
    assert "off-scan" in text and "convergence" in text


def test_runbook_and_explorer_keep_every_blocker_and_only_first_50_ordinary(tmp_path: Path):
    receipt = _baseline()
    output = tmp_path / "fhrp-domain-runbook.docx"
    write_runbook_docx(
        str(output), {"fhrp_redundancy_domain_baseline": receipt}, "fixture")
    runbook = _docx_text(output)
    explorer = _explorer_result(receipt)
    rendered = explorer["rendered"]

    assert explorer["valid"] is True
    for text in (runbook, rendered):
        assert "degraded-027-b" in text
        assert "review-028-b" in text
        assert "not_verified-029-b" in text
        assert "assessed-054-b" in text
        assert "assessed-055-a" not in text
        assert "All 60 blocker row(s)" in text
        assert "50 of 60 ordinary" in text
        assert "VLAN 159" not in text
        assert "subtype + group + virtual IP" in text
        assert "not a proven unprotected gateway or failure" in text
        assert "simultaneous roles" in text and "timers" in text


def test_malformed_receipts_fail_closed_without_echo_on_every_surface(tmp_path: Path):
    hostile = {
        "schema": "fhrp_redundancy_domain_baseline/1",
        "rows": [{"switch": "HOSTILE-SWITCH", "acceptance": "HOSTILE PASS"}],
        "secret": "authentication md5 HOSTILE-SECRET",
    }
    wb = Workbook()
    write_fhrp_redundancy_domain_sheet(wb, hostile)
    workbook = _sheet_text(wb[FHRP_REDUNDANCY_DOMAIN_SHEET_NAME])
    output = tmp_path / "malformed-domain-runbook.docx"
    write_runbook_docx(
        str(output), {"fhrp_redundancy_domain_baseline": hostile}, "fixture")
    runbook = _docx_text(output)
    explorer = _explorer_result(hostile)

    assert explorer["valid"] is False
    for text in (workbook, runbook, explorer["rendered"]):
        assert "HOSTILE" not in text
        assert "NOT ASSESSED" in text.upper()
        assert "No rejected receipt leaves" in text


def test_valid_invalid_source_receipt_is_not_presented_as_empty_or_complete(tmp_path: Path):
    receipt = embedded_fhrp_redundancy_domain_baseline(None)
    assert validate_fhrp_redundancy_domain_baseline(receipt)["valid"] is True
    wb = Workbook()
    write_fhrp_redundancy_domain_sheet(wb, receipt)
    workbook = _sheet_text(wb[FHRP_REDUNDANCY_DOMAIN_SHEET_NAME])
    output = tmp_path / "unverified-domain-runbook.docx"
    write_runbook_docx(
        str(output), {"fhrp_redundancy_domain_baseline": receipt}, "fixture")
    runbook = _docx_text(output)
    explorer = _explorer_result(receipt)

    assert explorer["valid"] is True
    for text in (workbook, runbook, explorer["rendered"]):
        assert "NOT VERIFIED" in text.upper()
        assert "zero published rows is not evidence of absence or complete intent" in text.lower()
        assert "current-run configured-group/svi source receipt is not verified" in text.lower()


def test_not_applicable_copy_says_no_subject_not_absence_or_complete_intent():
    receipt = _baseline(blocker_domains=0, ordinary_domains=0)
    assert receipt["verdict"] == "NOT_APPLICABLE"
    wb = Workbook()
    write_fhrp_redundancy_domain_sheet(wb, receipt)
    explorer = _explorer_result(receipt)
    for text in (
        _sheet_text(wb[FHRP_REDUNDANCY_DOMAIN_SHEET_NAME]),
        explorer["rendered"],
    ):
        assert "no subject was identified" in text
        assert "not proof that FHRP is absent" in text
        assert "intended membership is complete" in text


@pytest.mark.parametrize(
    ("participation", "expected_status"),
    (("nonparticipant", "REVIEW"), ("not_verified", "NOT VERIFIED")),
)
def test_explorer_keeps_zero_or_unverified_member_visibility_without_failure_claim(
        participation: str, expected_status: str):
    receipt = _mixed_participation_baseline(participation)
    result = _explorer_result(receipt)
    rendered = result["rendered"]
    assert result["valid"] is True
    assert participation.replace("_", " ").upper() in rendered
    assert expected_status in rendered
    assert "no positive candidate" in rendered
    assert "intended membership is unresolved" in rendered
    assert "not a proven unprotected gateway or failure" in rendered


def test_explorer_retains_valid_unobserved_subnet_review_blockers():
    receipt = _unobserved_subnet_baseline()
    result = _explorer_result(receipt)
    assert result["valid"] is True
    assert result["rendered"].count("(subnet-unobserved)") >= 2
    assert "BLOCKER · REVIEW" in result["rendered"]
    assert "An exact observed IPv4 subnet could not be established" in result["rendered"]


@pytest.mark.parametrize("role", ("LISTEN", "SPEAK"))
def test_hsrp_listen_and_speak_are_review_not_accepted_backups(role: str):
    receipt = _hsrp_nonbackup_baseline(role)
    result = _explorer_result(receipt)
    assert result["valid"] is True
    assert f"role {role}" in result["rendered"]
    assert "BLOCKER · REVIEW" in result["rendered"]
    assert "No accepted sequential backup role was observed" in result["rendered"]


def test_legacy_consistency_is_only_a_review_projection_and_authoritative_receipt_wins(
        tmp_path: Path):
    receipt = _baseline(blocker_domains=3, ordinary_domains=1)
    current, current_interfaces = _trusted_current_review(tmp_path)
    authoritative = compute_fhrp_consistency(current_interfaces, current)
    assert {row["status"] for row in authoritative} == {"review"}
    assert all(row["authoritative"] is True for row in authoritative)
    assert {row["vrf"] for row in authoritative} == {"global"}
    assert {row["subnet"] for row in authoritative} == {"10.0.10.0/24"}

    interfaces = {
        "edge-a": {"Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.2/24")},
        "edge-b": {"Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.3/24")},
    }
    legacy = compute_fhrp_consistency(interfaces)
    assert len(legacy) == 1 and legacy[0]["status"] == "review"
    assert legacy[0]["authoritative"] is False
    assert legacy[0]["vrf"] == "" and legacy[0]["subnet"] == "10.0.10.0/24"
    copy = json.dumps(legacy).lower()
    assert "intended participation" in copy and "unresolved" in copy
    assert all(word not in copy for word in ("clean", "unprotected", "fake"))

    hostile = {"rows": [{"switch": "HOSTILE-COMPAT", "issue": "HOSTILE PASS"}]}
    rejected = compute_fhrp_consistency({}, hostile)
    assert len(rejected) == 1 and rejected[0]["status"] == "not_verified"
    assert "HOSTILE" not in json.dumps(rejected)

    embedded = compute_fhrp_consistency({}, receipt)
    assert len(embedded) == 1 and embedded[0]["status"] == "not_verified"
    assert "not current-run source-bound" in embedded[0]["issues"][0]

    invalid_source = compute_fhrp_consistency(
        {}, embedded_fhrp_redundancy_domain_baseline(None))
    assert len(invalid_source) == 1
    assert invalid_source[0]["status"] == "not_verified"
    assert "not current-run source-bound" in invalid_source[0]["issues"][0]
    assert "no domain-composition conclusion" in invalid_source[0]["issues"][0]

    wb = Workbook()
    write_fhrp_consistency_sheet(wb, interfaces)
    sheet = _sheet_text(wb[FHRP_CONSISTENCY_SHEET_NAME]).lower()
    assert "legacy fallback" in sheet and "not verified" in sheet
    assert all(word not in sheet for word in ("clean", "unprotected", "fake"))


@pytest.mark.skipif(NODE is None, reason="Node.js is required for Explorer parity")
def test_explorer_rejects_digest_and_census_tampering():
    receipt = _baseline(blocker_domains=1, ordinary_domains=1)
    digest_tamper = json.loads(json.dumps(receipt))
    digest_tamper["rows"][0]["switch"] = "tampered"
    census_tamper = json.loads(json.dumps(receipt))
    census_tamper["domains"][0]["member_count"] = 999
    census_tamper["summary"]["baseline_sha256"] = _baseline_digest(census_tamper)
    candidate_tamper = json.loads(json.dumps(receipt))
    candidate_tamper["rows"][0]["candidate_key"] = "HOSTILE-CANDIDATE"
    candidate_tamper["summary"]["baseline_sha256"] = _baseline_digest(candidate_tamper)
    source_tamper = json.loads(json.dumps(receipt))
    source_tamper["rows"][0]["source_key"] = "HOSTILE PASS"
    source_tamper["summary"]["baseline_sha256"] = _baseline_digest(source_tamper)
    semantic_tamper = json.loads(json.dumps(receipt))
    semantic_tamper["rows"][0]["role"] = "DOWN"
    semantic_tamper["domains"][0]["members"][0]["role"] = "DOWN"
    semantic_tamper["summary"]["baseline_sha256"] = _baseline_digest(semantic_tamper)
    invalid_source_with_leaves = json.loads(json.dumps(receipt))
    invalid_source_with_leaves["source_receipt"].update({
        "valid": False,
        "configured_baseline_sha256": "",
        "svi_projection_sha256": "",
    })
    invalid_source_with_leaves["summary"]["baseline_sha256"] = _baseline_digest(
        invalid_source_with_leaves)
    unicode_casefold_collision = json.loads(json.dumps(receipt))
    for index, host in enumerate(("ß", "ẞ")):
        unicode_casefold_collision["rows"][index]["switch"] = host
        unicode_casefold_collision["domains"][0]["members"][index]["switch"] = host
        row = unicode_casefold_collision["rows"][index]
        row["source_key"] = (
            f"fhrp_configured_group_baseline.rows[{host},{row['protocol']},"
            f"{row['interface']},{row['group']}] + "
            f"interfaces[{host},{row['interface']}].svi_ip/vrf"
        )
    unicode_casefold_collision["summary"]["baseline_sha256"] = _baseline_digest(
        unicode_casefold_collision)
    non_ascii_vrf = json.loads(json.dumps(receipt))
    cherokee_lowercase_a = "\uAB70"
    assert ord(cherokee_lowercase_a) == 0xAB70
    domain = non_ascii_vrf["domains"][0]
    old_domain_key = domain["domain_key"]
    domain["vrf"] = cherokee_lowercase_a
    domain["domain_key"] = (
        f"vlan={domain['vlan']}|vrf={cherokee_lowercase_a}|subnet={domain['subnet']}"
    )
    candidate_count = len({
        row["candidate_key"] for row in non_ascii_vrf["rows"]
        if row["domain_key"] == old_domain_key and row["candidate_key"]
    })
    domain["acceptance"] = _acceptance(
        domain["status"], vlan=domain["vlan"], vrf=domain["vrf"],
        subnet=domain["subnet"], member_count=domain["member_count"],
        candidate_count=candidate_count,
    )
    for row in non_ascii_vrf["rows"]:
        if row["domain_key"] != old_domain_key:
            continue
        row["vrf"] = domain["vrf"]
        row["domain_key"] = domain["domain_key"]
        row["check"] = _check(row["vlan"], row["vrf"], row["subnet"])
        row["acceptance"] = domain["acceptance"]
    for finding in non_ascii_vrf["findings"]:
        if finding["domain_key"] == old_domain_key:
            finding["domain_key"] = domain["domain_key"]
    non_ascii_vrf["summary"]["baseline_sha256"] = _baseline_digest(non_ascii_vrf)

    assert _explorer_result(receipt)["valid"] is True
    for tampered in (
        digest_tamper,
        census_tamper,
        candidate_tamper,
        source_tamper,
        semantic_tamper,
        invalid_source_with_leaves,
        unicode_casefold_collision,
        non_ascii_vrf,
    ):
        assert validate_fhrp_redundancy_domain_baseline(tampered)["valid"] is False
        result = _explorer_result(tampered)
        assert result["valid"] is False
        assert "HOSTILE" not in result["rendered"]


@pytest.mark.skipif(NODE is None, reason="Node.js is required for Explorer parity")
def test_explorer_preflights_total_nested_members_before_deep_validation():
    result = _explorer_member_cap_probe(
        _baseline(blocker_domains=0, ordinary_domains=0))
    assert result == {
        "valid": False,
        "deepValidatorReached": False,
        "threw": False,
    }
