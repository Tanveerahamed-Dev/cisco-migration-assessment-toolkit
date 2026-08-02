"""End-to-end --redact security regression (audit-3 #8).

`--redact` is documented as the SHARE-SAFE switch. It correctly pseudonymizes the snapshot.json and the
*_explorer.html, but the ALWAYS-produced .xlsx workbook is built from the raw all_interfaces / all_device_physical
dataclasses and saved BEFORE redact_snapshot ever runs on the JSON -- so it used to ship real serials / management
IPs / MACs in the primary deliverable.

This runs the REAL offline pipeline twice over the synthetic collection -- once WITHOUT --redact (to prove the real
inventory genuinely reaches the workbook) and once WITH --redact (to prove every real value is gone and pseudonyms
took their place). It is deliberately end-to-end: it would have FAILED before the redact_collected_inplace +
redact_workbook_cells passes were wired in.
"""
import json
import os
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

import synthetic_fixtures as fx
from webapp.backend.redaction_verify import certify_shareable_artifacts

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py")

# Distinctive REAL inventory values the synthetic collection embeds. Serials come from `show version`; the MAC is a
# `show mac address-table` endpoint; the IPs are SVI HOST addresses (not network/remap-target addresses). None is a
# protocol constant -- HSRP 0000.0c07.acXX and the like are intentionally PRESERVED by redaction, so they are not
# asserted here.
REAL_SERIALS = ("FCW1234A001", "FOC2233B002", "FDO12345ABC")
REAL_MAC_DOTTED = "0011.2233.4455"
REAL_MAC_COLON = "00:11:22:33:44:55"
REAL_SVI_IPS = ("10.0.10.1", "10.0.20.1", "10.0.30.1")


def _run_workbook_cells(base, redact):
    """Run the offline pipeline into `base` and return every workbook cell joined into one string."""
    os.makedirs(base, exist_ok=True)
    collection = fx.write_collection(os.path.join(base, "collection"))
    devices = os.path.join(base, "devices.json")
    with open(devices, "w", encoding="utf-8") as f:
        json.dump(fx.DEVICES, f)
    template = os.path.join(base, "template.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "Interface Data"
    ws.append(["Hostname", "Port", "Status"]); wb.save(template)
    out_xlsx = os.path.join(base, "out.xlsx")
    cmd = [sys.executable, SCRIPT, "--no-collect", "--collection-dir", collection,
           "--devices-file", devices, "--template", template, "--output", out_xlsx,
           "--no-docx", "--no-pptx", "--no-design", "--no-mop", "--no-crd",
           "--no-engagement", "--no-opshandbook", "--no-archreview", "--workers", "1"]
    if redact:
        cmd.append("--redact")
    else:
        cmd.append("--no-html")
    proc = subprocess.run(cmd, cwd=base, capture_output=True, text=True, timeout=300)
    assert proc.returncode == 0, f"pipeline failed:\nSTDOUT\n{proc.stdout}\nSTDERR\n{proc.stderr}"
    if redact:
        snapshot = Path(base, "out.snapshot.json")
        explorer = Path(base, "out_explorer.html")
        # The engine SEALS topology.mmd / topology.dot into the run manifest and ships them beside
        # the deliverables, so they are part of the shareable set whether or not the verifier used
        # to look at them. Handing them in is the point: a suffix pre-filter silently discarded
        # both and still returned a proof that read as full coverage.
        diagrams = [Path(base, "topology.mmd"), Path(base, "topology.dot")]
        assert all(d.is_file() for d in diagrams), "precondition: the engine shipped the diagrams"
        proof = certify_shareable_artifacts(snapshot, [Path(out_xlsx), explorer, *diagrams])
        assert {snapshot.name, Path(out_xlsx).name, explorer.name} <= set(proof)
        assert {d.name for d in diagrams} <= set(proof), (
            "a registered run artifact was dropped from the proof instead of being verified or "
            f"refused: {sorted(set(d.name for d in diagrams) - set(proof))}")
    wb = load_workbook(out_xlsx, read_only=True, data_only=True)
    return "\n".join(str(c) for s in wb.sheetnames
                     for row in wb[s].iter_rows(values_only=True) for c in row if c is not None)


def test_redact_workbook_does_not_leak_real_inventory(tmp_path):
    # 1) WITHOUT --redact the workbook genuinely CONTAINS the real inventory -> these are real leak vectors, so the
    #    --redact assertions below are non-vacuous.
    plain = _run_workbook_cells(str(tmp_path / "plain"), redact=False)
    assert REAL_SERIALS[0] in plain, "fixture serial should reach the unredacted workbook (test would be vacuous otherwise)"
    assert REAL_MAC_DOTTED in plain, "fixture endpoint MAC should reach the unredacted workbook"

    # 2) WITH --redact, NONE of the real serials / MACs / SVI IPs survive in ANY cell of ANY sheet, and the
    #    pseudonyms ARE present (proving redaction ran, not that the data merely vanished).
    redacted = _run_workbook_cells(str(tmp_path / "redact"), redact=True)
    for s in REAL_SERIALS:
        assert s not in redacted, f"real serial {s} leaked into the --redact workbook"
    assert REAL_MAC_DOTTED not in redacted and REAL_MAC_COLON not in redacted, \
        "real endpoint MAC leaked into the --redact workbook"
    for ip in REAL_SVI_IPS:
        assert ip not in redacted, f"real SVI host IP {ip} leaked into the --redact workbook"
    assert "serial-000001.assesshub-redacted.invalid" in redacted


def test_make_redactor_ip_map_never_reproduces_a_real_address():
    """IPv4 pseudonyms must be collision-proof and never reproduce a real input address.

    The old in-band mapping could draw a pseudonym equal to another real network. Pseudonyms now
    use a tool-owned ``.invalid`` marker namespace instead of any address-shaped value.
    """
    from cisco_toolkit.html import _make_redactor
    scrub, _ = _make_redactor()
    # 12 distinct real /24s, so 10.0.10 and 10.0.20 sit near their own would-be indices; host octet .1
    # everywhere makes any in-band collision reproduce a real address exactly.
    reals = [f"10.0.{n}.1" for n in (0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 20)]
    outs = [scrub(ip) for ip in reals]
    for real in reals:
        assert all(real not in o for o in outs), f"real address {real} reproduced by the pseudonym map"
    assert all(re.fullmatch(
        r"v4-n\d{5}-h001\.assesshub-redacted\.invalid", o
    ) for o in outs), outs
    assert scrub("240.0.0.7") != "240.0.0.7"
    # determinism within a call: the same input scrubs to the same output
    assert scrub(reals[0]) == outs[0]


def test_redact_snapshot_ip_map_never_reproduces_a_real_address():
    """The JSON/explorer path uses collision-proof markers and remains idempotent.

    Existing synthetic markers are preserved, while address-shaped input is always replaced.
    """
    from cisco_toolkit.html import redact_snapshot
    reals = [f"10.0.{n}.1" for n in (0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 20)]
    snap = {"ips": list(reals)}
    once = redact_snapshot(snap)
    outs = once["ips"]
    for real in reals:
        assert all(real not in o for o in outs), f"real address {real} reproduced by redact_snapshot's map"
    assert all(re.fullmatch(
        r"v4-n\d{5}-h001\.assesshub-redacted\.invalid", o
    ) for o in outs), outs
    r = redact_snapshot({"seed": "240.0.0.9", "real": "192.168.1.5"})
    assert r["seed"] != "240.0.0.9"
    assert r["real"].endswith(".assesshub-redacted.invalid")
    assert redact_snapshot(once) == once


def test_redact_snapshot_pseudonymizes_general_cisco_serials_in_keys_and_values():
    """Serials in generated prose and dict keys use the same stable pseudonym as schema fields."""
    from cisco_toolkit.html import redact_snapshot

    real = "FDO12345ABC"
    snapshot = {
        "nrfu_commands": {
            "expected": f"chassis N9K-C93180YC-EX, serial {real} present",
            f"device_{real.lower()}": {"echo": real},
        },
        "serial": real,
    }

    redacted = redact_snapshot(snapshot)
    pseudonym = redacted["serial"]
    assert re.fullmatch(r"serial-\d{6}\.assesshub-redacted\.invalid", pseudonym)
    assert redacted["nrfu_commands"]["expected"].endswith(f"serial {pseudonym} present")
    assert f"device_{pseudonym}" in redacted["nrfu_commands"]
    assert redacted["nrfu_commands"][f"device_{pseudonym}"]["echo"] == pseudonym
    assert real not in json.dumps(redacted)
    assert redact_snapshot(redacted) == redacted


def test_redact_snapshot_preserves_every_colliding_serial_key():
    """Case variants and a pre-existing pseudonym key must not silently overwrite one another."""
    from cisco_toolkit.html import redact_snapshot

    snapshot = {
        "serial": "FDO12345ABC",
        "FDO12345ABC": "first",
        "serial-000001.assesshub-redacted.invalid~2": "pre-existing alias",
        "fdo12345abc": "second",
        "serial-000001.assesshub-redacted.invalid": "pre-existing pseudonym",
    }

    redacted = redact_snapshot(snapshot)
    pseudonym = redacted["serial"]
    assert pseudonym == "serial-000002.assesshub-redacted.invalid"
    assert list(redacted)[1:] == [
        pseudonym,
        "serial-000001.assesshub-redacted.invalid~2",
        f"{pseudonym}~2",
        "serial-000001.assesshub-redacted.invalid",
    ]
    assert list(redacted.values())[1:] == [
        "first",
        "pre-existing alias",
        "second",
        "pre-existing pseudonym",
    ]
    assert "FDO12345ABC" not in json.dumps(redacted)
    assert "fdo12345abc" not in json.dumps(redacted)
    assert redact_snapshot(redacted) == redacted


def test_redact_snapshot_preserves_colliding_serial_keys_inside_secret_container():
    """The all-leaf secret scrub must use the same non-lossy key insertion path."""
    from cisco_toolkit.html import redact_snapshot

    snapshot = {
        "apiSecret": {
            "FDO12345ABC": "one",
            "fdo12345abc": "two",
            "serial-000001.assesshub-redacted.invalid": "three",
        }
    }

    redacted = redact_snapshot(snapshot)
    secret = redacted["apiSecret"]
    assert list(secret) == [
        "serial-000002.assesshub-redacted.invalid",
        "serial-000002.assesshub-redacted.invalid~2",
        "serial-000001.assesshub-redacted.invalid",
    ]
    assert list(secret.values()) == ["<redacted>", "<redacted>", "<redacted>"]
    assert "FDO12345ABC" not in json.dumps(redacted)
    assert "fdo12345abc" not in json.dumps(redacted)
    assert redact_snapshot(redacted) == redacted


# ── W1/F3: the shareable proof advertised coverage it did not have ─────────────
# `certify_shareable_artifacts` pre-filtered its inputs against a hand-listed
# `_SHAREABLE_SUFFIXES = {.json,.html,.xlsx,.docx,.pptx}` and DROPPED everything else in silence.
# The CLI registers topology.mmd and topology.dot as run artifacts, seals them into the run
# manifest, ships them beside the deliverables and hands them to this function; it discarded both,
# then the caller recorded `status="verified"` with `verified_artifacts` = 4 over a set of 6.
def _pytest_error():
    import pytest
    from webapp.backend.redaction_verify import RedactionVerificationError
    return pytest, RedactionVerificationError


def test_a_text_run_artifact_is_scanned_not_dropped(tmp_path):
    """The scan must be REAL. A diagram carrying an unredacted address has to be refused, or
    "covered" would just mean "counted"."""
    pytest, RedactionVerificationError = _pytest_error()
    snapshot = tmp_path / "safe.snapshot.json"
    snapshot.write_text(json.dumps({"devices": {}}), encoding="utf-8")

    clean = tmp_path / "topology.mmd"
    clean.write_text('graph LR\n    n0["core1"]\n    n0 --- n1\n', encoding="utf-8")
    proof = certify_shareable_artifacts(snapshot, [clean])
    assert clean.name in proof, "a .mmd run artifact must be verified, not silently discarded"

    leaky = tmp_path / "topology.dot"
    leaky.write_text('graph {\n  "core1" -- "acc1" [label="10.44.7.219"];\n}\n', encoding="utf-8")
    with pytest.raises(RedactionVerificationError, match="leak indicator"):
        certify_shareable_artifacts(snapshot, [leaky])


def test_an_artifact_no_scanner_can_read_is_refused_not_silently_skipped(tmp_path):
    """The other half of the shape fix. Dropping an unrecognised artifact returned a proof that
    claimed a clean set over bytes nobody read; refusing says so out loud. A verifier may narrow
    the caller's set only by telling the caller."""
    pytest, RedactionVerificationError = _pytest_error()
    snapshot = tmp_path / "safe.snapshot.json"
    snapshot.write_text(json.dumps({"devices": {}}), encoding="utf-8")
    opaque = tmp_path / "capture.pcap"
    opaque.write_bytes(b"\xd4\xc3\xb2\xa1\x00\x04\x00\x02\xff\xfe\x00\x00")
    with pytest.raises(RedactionVerificationError, match="cannot be certified"):
        certify_shareable_artifacts(snapshot, [opaque])
    # Non-vacuity: the same call over only the readable artifacts still succeeds.
    assert set(certify_shareable_artifacts(snapshot, [])) == {snapshot.name}
