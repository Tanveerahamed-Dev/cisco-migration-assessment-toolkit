"""Source-bound observed local L2 failure-trial evidence.

The fixtures retain parser-shaped IOS-XE/NX-OS output through the native STP,
EtherChannel, and multichassis owners.  The trial composer is deliberately local:
it must not turn those device states into a service, traffic, or convergence claim.
"""

from __future__ import annotations

from copy import deepcopy
import json

import pytest
from openpyxl import load_workbook

from cisco_toolkit.analyze import (
    compute_etherchannel_projection,
    summarize_etherchannel_baseline,
)
from cisco_toolkit.capture_integrity import compute_capture_integrity_from_paths
from cisco_toolkit.comparison import compare_bound_pair
from cisco_toolkit.etherchannel import (
    compute_etherchannel_operational_evidence,
    embedded_etherchannel_operational_evidence,
)
from cisco_toolkit.l2_rehearsal import (
    L2_FAILURE_WITNESS_SCHEMA,
    OBSERVED_L2_FAILURE_EVIDENCE_SCHEMA,
    compute_observed_l2_failure_evidence,
    validate_observed_l2_failure_evidence,
)
from cisco_toolkit.html import write_diff_workbook
from cisco_toolkit.multichassis_lag import compute_multichassis_lag_domain_baseline
from cisco_toolkit.protocol_assurance import (
    PERSISTED_SOURCE,
    bind_snapshot_json_bytes,
    bound_snapshot_source,
)
from cisco_toolkit.stp_topology import (
    compute_stp_topology_baseline,
    produce_stp_topology_observation,
)
from tests.test_etherchannel_cutover_truth import _receipt
from tests.test_etherchannel_operational_evidence import (
    _bind as _bind_source_paths,
    _copy_paths as _copy_etherchannel_paths,
)
from tests.test_l2_failure_rehearsal import (
    _ether_snapshot,
    _multichassis_snapshot,
)
from tests.test_multichassis_lag import _nxos_observation
from tests.test_compare_cutover_gate_cli import _snapshot as _clean_comparison_snapshot


def _bind(value: dict, collected_at: str):
    payload = deepcopy(value)
    payload["collected_at"] = collected_at
    raw = json.dumps(
        payload,
        sort_keys=True,
        ensure_ascii=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")
    return bind_snapshot_json_bytes(raw)


def _custody() -> dict:
    return {
        "pre_failure": {
            "source": PERSISTED_SOURCE,
            "source_id": "snapshot:101",
            "campaign_id": 7,
            "engagement_id": "ENG-OBSERVED-L2",
            "custody_at": "2026-08-20T01:00:00+00:00",
        },
        "post_failure": {
            "source": PERSISTED_SOURCE,
            "source_id": "snapshot:102",
            "campaign_id": 7,
            "engagement_id": "ENG-OBSERVED-L2",
            "custody_at": "2026-08-20T01:01:00+00:00",
        },
        "recovery": {
            "source": PERSISTED_SOURCE,
            "source_id": "snapshot:103",
            "campaign_id": 7,
            "engagement_id": "ENG-OBSERVED-L2",
            "custody_at": "2026-08-20T01:02:00+00:00",
        },
    }


def _witness(family: str, subject: str, scenario: str, action: str, target: dict) -> bytes:
    return json.dumps(
        {
            "schema": L2_FAILURE_WITNESS_SCHEMA,
            "family": family,
            "subject": subject,
            "failure_scenario": scenario,
            "action": action,
            "target": target,
            "induced_at": "2026-08-20T00:01:00+00:00",
        },
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _trial(pre: dict, post: dict, recovery: dict, witness: bytes, *, custody=None):
    return compute_observed_l2_failure_evidence(
        _bind(pre, "2026-08-20T00:00:00+00:00"),
        _bind(post, "2026-08-20T00:02:00+00:00"),
        _bind(recovery, "2026-08-20T00:03:00+00:00"),
        witness_bytes=witness,
        phase_custody=_custody() if custody is None else custody,
    )


def _pvst_state(*, root_address: str, bridge_address: str, is_root: bool) -> str:
    root_tail = (
        "             This bridge is the root\n"
        if is_root else
        "             Cost        4\n"
        "             Port        1 (GigabitEthernet1/0/1)\n"
    )
    forwarding_role = "Desg" if is_root else "Root"
    return f"""VLAN0010
  Spanning tree enabled protocol rstp
  Root ID    Priority    24586
             Address     {root_address}
{root_tail}  Bridge ID  Priority    32778  (priority 32768 sys-id-ext 10)
             Address     {bridge_address}
Interface        Role Sts Cost      Prio.Nbr Type
---------------- ---- --- --------- -------- ----
Gi1/0/1          {forwarding_role} FWD 4         128.1    P2p
Gi1/0/24         Altn BLK 19        128.24   P2p
"""


def _pvst_detail(counter: int) -> str:
    return f"""VLAN0010 is executing the rstp compatible Spanning Tree protocol
  Number of topology changes {counter} last change occurred 00:00:12 ago
"""


def _stp_snapshot(rows: dict[str, tuple[str, str, bool, int]]) -> dict:
    observations = {
        host: produce_stp_topology_observation(
            _pvst_state(
                root_address=root_address,
                bridge_address=bridge_address,
                is_root=is_root,
            ),
            _pvst_detail(counter),
            state_capture_state="usable",
            detail_capture_state="usable",
        )
        for host, (root_address, bridge_address, is_root, counter) in rows.items()
    }
    devices = {host: {"platform": "iosxe"} for host in rows}
    return {
        "script_version": "V3.23.0",
        "devices": devices,
        "stp_topology_observations": observations,
        "stp_topology_baseline": compute_stp_topology_baseline(observations, devices),
    }


def _stp_phases() -> tuple[dict, dict, dict]:
    pre = _stp_snapshot({
        "dist1": ("aaaa.0001.0001", "aaaa.0001.0001", True, 2),
        "dist2": ("aaaa.0001.0001", "bbbb.0002.0002", False, 2),
    })
    post = _stp_snapshot({
        "dist2": ("bbbb.0002.0002", "bbbb.0002.0002", True, 3),
    })
    recovery = _stp_snapshot({
        "dist1": ("aaaa.0001.0001", "aaaa.0001.0001", True, 4),
        "dist2": ("aaaa.0001.0001", "bbbb.0002.0002", False, 4),
    })
    return pre, post, recovery


def test_stp_observed_root_failure_is_exact_local_safety_only() -> None:
    pre, post, recovery = _stp_phases()
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "stp",
            "root|dist1|pvst_vlan|10",
            "single_proven_root_host_loss",
            "fail_node",
            {"host": "dist1"},
        ),
    )

    assert receipt["schema"] == OBSERVED_L2_FAILURE_EVIDENCE_SCHEMA
    assert receipt["status"] == "observed_survival"
    assert receipt["assurance_level"] == "local_safety_preservation"
    assert receipt["failure_witness"]["evidence"] == {
        "post_failure_root_switch": "dist2",
        "post_failure_root_address": "bbbb.0002.0002",
        "failed_root_absent_from_post_phase": True,
        "exact_pre_failure_survivor_roster": True,
        "topology_counter_witness_switches": ["dist2"],
    }
    assert receipt["claims"] == {
        "local_scenario": "observed_survival",
        "service_path_survival": "not_verified",
        "traffic_continuity": "not_verified",
        "convergence": "not_verified",
    }
    assert validate_observed_l2_failure_evidence(receipt)["valid"] is True


def test_stp_priority_change_with_target_still_present_is_not_a_failure_trial() -> None:
    pre, _post, recovery = _stp_phases()
    priority_change = _stp_snapshot({
        "dist1": ("bbbb.0002.0002", "aaaa.0001.0001", False, 3),
        "dist2": ("bbbb.0002.0002", "bbbb.0002.0002", True, 3),
    })
    receipt = _trial(
        pre,
        priority_change,
        recovery,
        _witness(
            "stp", "root|dist1|pvst_vlan|10", "single_proven_root_host_loss",
            "fail_node", {"host": "dist1"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["failure_witness"]["status"] == "not_verified"
    assert receipt["failure_witness"]["evidence"][
        "failed_root_absent_from_post_phase"
    ] is False


def test_stp_post_phase_cannot_substitute_an_unseen_replacement_root() -> None:
    pre, _post, recovery = _stp_phases()
    substituted = _stp_snapshot({
        "dist2": ("cccc.0003.0003", "bbbb.0002.0002", False, 3),
        "dist3": ("cccc.0003.0003", "cccc.0003.0003", True, 3),
    })
    receipt = _trial(
        pre,
        substituted,
        recovery,
        _witness(
            "stp", "root|dist1|pvst_vlan|10", "single_proven_root_host_loss",
            "fail_node", {"host": "dist1"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["failure_witness"]["status"] == "not_verified"
    assert receipt["failure_witness"]["evidence"][
        "exact_pre_failure_survivor_roster"
    ] is False


def test_etherchannel_observed_configured_member_loss_and_recovery(tmp_path) -> None:
    pre = _ether_snapshot(
        tmp_path / "pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    post = _ether_snapshot(
        tmp_path / "post", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
    )
    recovery = _ether_snapshot(
        tmp_path / "recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "etherchannel", "dist1|Po10",
            "single_observed_forwarding_member_loss", "shut_link",
            {"host": "dist1", "interface": "Gi1/0/2"},
        ),
    )

    assert receipt["status"] == "observed_survival"
    assert receipt["post_failure"]["evidence"]["remaining_forwarding_members"] == 1
    assert receipt["post_failure"]["evidence"]["target_remained_configured"] is True
    assert receipt["failure_witness"]["evidence"]["target_runtime_state"] \
        == "non_forwarding_observed"
    assert validate_observed_l2_failure_evidence(receipt)["valid"] is True


def test_etherchannel_config_deletion_cannot_impersonate_member_failure(tmp_path) -> None:
    pre = _ether_snapshot(
        tmp_path / "pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    deleted = _ether_snapshot(
        tmp_path / "deleted", "10 Po10(SU) LACP Gi1/0/1(P)"
    )
    recovery = _ether_snapshot(
        tmp_path / "recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    receipt = _trial(
        pre,
        deleted,
        recovery,
        _witness(
            "etherchannel", "dist1|Po10",
            "single_observed_forwarding_member_loss", "shut_link",
            {"host": "dist1", "interface": "Gi1/0/2"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["failure_witness"]["evidence"]["configured_members_preserved"] is False


def test_etherchannel_embedded_row_must_match_each_phase_device_roster(tmp_path) -> None:
    pre = _ether_snapshot(
        tmp_path / "pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    post = _ether_snapshot(
        tmp_path / "post", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
    )
    post["devices"] = {"other-switch": {"platform": "iosxe"}}
    recovery = _ether_snapshot(
        tmp_path / "recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "etherchannel", "dist1|Po10",
            "single_observed_forwarding_member_loss", "shut_link",
            {"host": "dist1", "interface": "Gi1/0/2"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert any("co-owner" in failure for failure in receipt["failures"])


def test_etherchannel_device_platform_must_match_the_closed_owner_variant(tmp_path) -> None:
    pre = _ether_snapshot(
        tmp_path / "pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    post = _ether_snapshot(
        tmp_path / "post", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
    )
    recovery = _ether_snapshot(
        tmp_path / "recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    for snapshot in (pre, post, recovery):
        snapshot["devices"] = {"dist1": {"platform": "bios-appliance"}}
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "etherchannel", "dist1|Po10",
            "single_observed_forwarding_member_loss", "shut_link",
            {"host": "dist1", "interface": "Gi1/0/2"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["precondition"]["status"] == "not_verified"
    assert receipt["precondition"]["evidence"]["subject_device_platform"] == ""


def test_etherchannel_trial_rejects_typed_and_protected_v1_owner_contradiction(
        tmp_path) -> None:
    pre = _ether_snapshot(
        tmp_path / "pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    post = _ether_snapshot(
        tmp_path / "post", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
    )
    # Leave the typed operational owner unchanged while contradicting the protected projection.
    post["etherchannel_projection"]["rows"][0]["groups"][0]["members"][1][
        "state"
    ] = "forwarding_observed"
    recovery = _ether_snapshot(
        tmp_path / "recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "etherchannel", "dist1|Po10",
            "single_observed_forwarding_member_loss", "shut_link",
            {"host": "dist1", "interface": "Gi1/0/2"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert any("co-owner" in failure for failure in receipt["failures"])


def _one_host_nxos_etherchannel(tmp_path, host: str) -> dict:
    paths = _copy_etherchannel_paths(tmp_path, "nxos", host, {})
    _bind_source_paths(paths)
    devices = {host: {"platform": "nxos"}}
    projection = compute_etherchannel_projection(devices, paths)
    assessment = _receipt({host: {"EtherChannel": "assessed"}})
    baseline = summarize_etherchannel_baseline(
        projection, assessment, devices=devices
    )
    operational = compute_etherchannel_operational_evidence(
        paths,
        compute_capture_integrity_from_paths(paths),
        projection,
        devices=devices,
    )
    return {
        "devices": devices,
        "etherchannel_projection": projection,
        "protocol_assessability": assessment,
        "etherchannel_baseline": baseline,
        "etherchannel_operational_evidence":
            embedded_etherchannel_operational_evidence(operational),
    }


def _nxos_post_failure_snapshot(
        tmp_path, *, detached: bool = False, peer: str = "leaf-a") -> dict:
    observation = _nxos_observation(
            "leaf-b",
            peer,
            source_digit="b",
            domain_state_overrides={
                "peer_status": "peer adjacency not formed",
                "keepalive_status": "peer is not alive",
                "peer_link_status": "down",
            },
        )
    if detached:
        observation = json.loads(json.dumps(observation))
    raw = {"observations": [observation]}
    snapshot = _one_host_nxos_etherchannel(tmp_path, "leaf-b")
    snapshot.update({
        "multichassis_lag_typed_observations": {
            "observations": [dict(raw["observations"][0])]
        },
        "multichassis_lag_domain_baseline":
            compute_multichassis_lag_domain_baseline(raw),
    })
    return snapshot


def test_nxos_multichassis_peer_loss_uses_closed_bad_states_and_local_leg(tmp_path) -> None:
    pre = _multichassis_snapshot(tmp_path / "pre")
    post = _nxos_post_failure_snapshot(tmp_path / "post")
    recovery = _multichassis_snapshot(tmp_path / "recovery")
    subject = pre["multichassis_lag_domain_baseline"][
        "reconciled_attachments"
    ][0]["subject_id"]
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "multichassis_lag", subject,
            "single_reciprocal_peer_or_local_leg_loss", "fail_node",
            {"host": "leaf-a"},
        ),
    )

    assert receipt["status"] == "observed_survival"
    assert receipt["failure_witness"]["evidence"] == {
        "surviving_peer": "leaf-b",
        "peer_status": "peer adjacency not formed",
        "keepalive_status": "peer is not alive",
        "peer_link_status": "down",
        "surviving_source_custody": "current_run_source_bound",
        "pair_identity_preserved": True,
    }
    assert receipt["post_failure"]["evidence"]["dual_active_status"] == "0"
    assert receipt["post_failure"]["evidence"]["service_path_survival"] \
        == "not_verified"
    assert validate_observed_l2_failure_evidence(receipt)["valid"] is True


def test_detached_multichassis_observation_cannot_authorize_local_safety(tmp_path) -> None:
    pre = _multichassis_snapshot(tmp_path / "pre")
    post = _nxos_post_failure_snapshot(tmp_path / "post", detached=True)
    recovery = _multichassis_snapshot(tmp_path / "recovery")
    subject = pre["multichassis_lag_domain_baseline"][
        "reconciled_attachments"
    ][0]["subject_id"]
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "multichassis_lag", subject,
            "single_reciprocal_peer_or_local_leg_loss", "fail_node",
            {"host": "leaf-a"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["failure_witness"]["status"] == "not_verified"
    assert receipt["failure_witness"]["evidence"][
        "surviving_source_custody"
    ] == "embedded_unverified"


def test_multichassis_failed_peer_must_leave_post_device_denominator(tmp_path) -> None:
    pre = _multichassis_snapshot(tmp_path / "pre")
    post = _nxos_post_failure_snapshot(tmp_path / "post")
    post["devices"]["leaf-a"] = {"platform": "nxos"}
    recovery = _multichassis_snapshot(tmp_path / "recovery")
    subject = pre["multichassis_lag_domain_baseline"][
        "reconciled_attachments"
    ][0]["subject_id"]

    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "multichassis_lag", subject,
            "single_reciprocal_peer_or_local_leg_loss", "fail_node",
            {"host": "leaf-a"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["failure_witness"]["status"] == "not_verified"


def test_multichassis_post_phase_cannot_substitute_a_different_peer_identity(tmp_path) -> None:
    pre = _multichassis_snapshot(tmp_path / "pre")
    post = _nxos_post_failure_snapshot(tmp_path / "post", peer="leaf-x")
    recovery = _multichassis_snapshot(tmp_path / "recovery")
    subject = pre["multichassis_lag_domain_baseline"][
        "reconciled_attachments"
    ][0]["subject_id"]

    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "multichassis_lag", subject,
            "single_reciprocal_peer_or_local_leg_loss", "fail_node",
            {"host": "leaf-a"},
        ),
    )

    assert receipt["status"] == "not_verified"
    assert receipt["failure_witness"]["status"] == "not_verified"
    assert receipt["failure_witness"]["evidence"][
        "pair_identity_preserved"
    ] is False


def test_multichassis_recovery_from_another_pair_is_not_survival(tmp_path) -> None:
    pre = _multichassis_snapshot(tmp_path / "pre")
    post = _nxos_post_failure_snapshot(tmp_path / "post")
    recovery = _multichassis_snapshot(tmp_path / "recovery")
    # Reuse the same domain ID at another site: domain ID alone must never
    # identify the reciprocal peer pair or its dual-homed attachment.
    other_pair = {"observations": [
        _nxos_observation("other-a", "other-b", source_digit="c", domain_id="10"),
        _nxos_observation("other-b", "other-a", source_digit="d", domain_id="10"),
    ]}
    recovery["devices"] = {
        "other-a": {"platform": "nxos"},
        "other-b": {"platform": "nxos"},
    }
    recovery["multichassis_lag_typed_observations"] = {
        "observations": [dict(row) for row in other_pair["observations"]]
    }
    recovery["multichassis_lag_domain_baseline"] = \
        compute_multichassis_lag_domain_baseline(other_pair)
    subject = pre["multichassis_lag_domain_baseline"][
        "reconciled_attachments"
    ][0]["subject_id"]

    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "multichassis_lag", subject,
            "single_reciprocal_peer_or_local_leg_loss", "fail_node",
            {"host": "leaf-a"},
        ),
    )

    assert receipt["status"] == "observed_failure"
    assert receipt["recovery"]["status"] == "failed"


@pytest.mark.parametrize(
    ("mutation", "failure_fragment"),
    [
        ("truncated_witness", "witness JSON bytes are malformed"),
        ("reordered", "not strictly pre < induced < post < recovery"),
        ("replayed_source", "source identities must be distinct"),
        ("engagement_mismatch", "must share one exact source owner, campaign, and engagement"),
        ("unknown_source_owner", "source owner or identity is missing"),
        ("relabelled_source_id", "source owner or identity is missing"),
        ("missing_phase_owner", "phase baselines are missing or malformed"),
    ],
)
def test_missing_malformed_reordered_and_replayed_evidence_stays_not_verified(
        mutation: str, failure_fragment: str) -> None:
    pre, post, recovery = _stp_phases()
    witness = _witness(
        "stp", "root|dist1|pvst_vlan|10", "single_proven_root_host_loss",
        "fail_node", {"host": "dist1"},
    )
    custody = _custody()
    if mutation == "truncated_witness":
        witness = witness[:-7]
    elif mutation == "reordered":
        witness_value = json.loads(witness)
        witness_value["induced_at"] = "2026-08-20T00:02:30+00:00"
        witness = json.dumps(witness_value, separators=(",", ":")).encode()
    elif mutation == "replayed_source":
        custody["post_failure"]["source_id"] = custody["pre_failure"]["source_id"]
    elif mutation == "engagement_mismatch":
        custody["post_failure"]["engagement_id"] = "ENG-CROSS-PAIR"
    elif mutation == "unknown_source_owner":
        custody["post_failure"]["source"] = "operator_supplied_claim"
    elif mutation == "relabelled_source_id":
        custody["post_failure"]["source_id"] = "operator:relabeled-post"
    elif mutation == "missing_phase_owner":
        post = {"devices": post["devices"]}

    receipt = _trial(pre, post, recovery, witness, custody=custody)

    assert receipt["status"] == "not_verified"
    assert any(failure_fragment in failure for failure in receipt["failures"])


def test_observed_receipt_is_mutation_sensitive_and_recovery_bound() -> None:
    pre, post, recovery = _stp_phases()
    receipt = _trial(
        pre,
        post,
        recovery,
        _witness(
            "stp", "root|dist1|pvst_vlan|10", "single_proven_root_host_loss",
            "fail_node", {"host": "dist1"},
        ),
    )
    recovery_source = receipt["source_binding"]["recovery"]
    recovery_binding = {
        "source": recovery_source["source"],
        "snapshot_id": 103,
        "campaign_id": recovery_source["campaign_id"],
        "engagement_id": recovery_source["engagement_id"],
        "sha256": recovery_source["sha256"],
        "bytes": recovery_source["bytes"],
    }
    assert validate_observed_l2_failure_evidence(
        receipt, expected_recovery_binding=recovery_binding
    )["valid"] is True
    for field, value in (
        ("snapshot_id", 999),
        ("campaign_id", 999),
        ("engagement_id", "ENG-SAME-BYTES-OTHER-CONTEXT"),
        ("sha256", "sha256:" + "0" * 64),
    ):
        mismatched = {**recovery_binding, field: value}
        validation = validate_observed_l2_failure_evidence(
            receipt, expected_recovery_binding=mismatched
        )
        assert validation["valid"] is False
        assert "different recovery" in validation["reason"]
    assert validate_observed_l2_failure_evidence(dict(receipt))["valid"] is False

    receipt["claims"]["service_path_survival"] = "observed"
    assert validate_observed_l2_failure_evidence(receipt)["valid"] is False


def _clean_l2_phase(l2: dict, collected_at: str):
    value = deepcopy(_clean_comparison_snapshot("FULL/DR", collected_at))
    value["devices"].update(deepcopy(l2.get("devices", {})))
    for key, item in l2.items():
        if key != "devices":
            value[key] = deepcopy(item)
    return _bind(value, collected_at)


def _comparison_binding(snapshot, snapshot_id: int) -> dict:
    source = bound_snapshot_source(snapshot)
    return {
        "source": PERSISTED_SOURCE,
        "sha256": source["sha256"],
        "bytes": source["bytes"],
        "snapshot_id": snapshot_id,
        "campaign_id": 7,
        "engagement_id": "ENG-OBSERVED-L2",
        "label": f"snapshot-{snapshot_id}.json",
        "script_version": str(snapshot.get("script_version") or "V3.23.0"),
    }


def _partner_changed_ether_snapshot(tmp_path) -> dict:
    desired = "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
    paths = _copy_etherchannel_paths(
        tmp_path,
        "ios",
        "dist1",
        {
            "show etherchannel summary": (
                "10     Po10(SU)        LACP       Gi1/0/1(P) Gi1/0/2(P)",
                desired,
            ),
            "show lacp neighbor": ("0011.2233.4455", "00aa.bbcc.ddee"),
        },
    )
    _bind_source_paths(paths)
    devices = {"dist1": {"platform": "ios"}}
    projection = compute_etherchannel_projection(devices, paths)
    assessment = _receipt({"dist1": {"EtherChannel": "assessed"}})
    baseline = summarize_etherchannel_baseline(
        projection, assessment, devices=devices
    )
    operational = compute_etherchannel_operational_evidence(
        paths,
        compute_capture_integrity_from_paths(paths),
        projection,
        devices=devices,
    )
    return {
        "devices": devices,
        "etherchannel_projection": projection,
        "protocol_assessability": assessment,
        "etherchannel_baseline": baseline,
        "etherchannel_operational_evidence":
            embedded_etherchannel_operational_evidence(operational),
    }


def _comparison_trial(tmp_path, *, refuted: bool = False):
    pre_l2 = _ether_snapshot(
        tmp_path / "pre", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    post_l2 = (
        _partner_changed_ether_snapshot(tmp_path / "post") if refuted else
        _ether_snapshot(
            tmp_path / "post", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(D)"
        )
    )
    recovery_l2 = _ether_snapshot(
        tmp_path / "recovery", "10 Po10(SU) LACP Gi1/0/1(P) Gi1/0/2(P)"
    )
    pre = _clean_l2_phase(pre_l2, "2026-08-20T00:00:00+00:00")
    post = _clean_l2_phase(post_l2, "2026-08-20T00:02:00+00:00")
    recovery = _clean_l2_phase(recovery_l2, "2026-08-20T00:03:00+00:00")
    before = _clean_l2_phase(recovery_l2, "2026-08-19T23:59:00+00:00")
    trial = compute_observed_l2_failure_evidence(
        pre,
        post,
        recovery,
        witness_bytes=_witness(
            "etherchannel", "dist1|Po10",
            "single_observed_forwarding_member_loss", "shut_link",
            {"host": "dist1", "interface": "Gi1/0/2"},
        ),
        phase_custody=_custody(),
    )
    return before, recovery, trial


def test_canonical_comparison_binds_and_renders_local_safety_without_service_claim(
        tmp_path) -> None:
    before, recovery, trial = _comparison_trial(tmp_path)
    comparison = compare_bound_pair(
        before,
        recovery,
        before_binding=_comparison_binding(before, 100),
        after_binding=_comparison_binding(recovery, 103),
        l2_failure_trial=trial,
    )
    gate = comparison["cutover_gate"]

    assert comparison["comparison_admission"]["status"] == "admitted"
    assert comparison["operator_evidence"]["rehearsal"][
        "observed_l2_failure_evidence"
    ] is trial
    # Unrelated routed/family/L2 coverage gaps remain non-PASS; the local
    # trial cannot launder them into an all-clear.
    assert gate["verdict"] == "INDETERMINATE"
    assert gate["l2_observed_trial_status"] == "observed_survival"
    assert gate["l2_observed_trial_assurance"] == "local_safety_preservation"
    assert gate["l2_observed_trial_family"] == "etherchannel"
    assert gate["l2_observed_trial_subject"] == "dist1|Po10"
    assert gate["l2_observed_trial_matched_projected_risks"] == 0
    assert "does not neutralize" in gate["l2_observed_trial_note"]
    assert trial["claims"]["service_path_survival"] == "not_verified"


def test_canonical_comparison_observed_refutation_blocks_clean_recovery(tmp_path) -> None:
    before, recovery, trial = _comparison_trial(tmp_path, refuted=True)
    assert trial["status"] == "observed_failure"
    comparison = compare_bound_pair(
        before,
        recovery,
        before_binding=_comparison_binding(before, 100),
        after_binding=_comparison_binding(recovery, 103),
        l2_failure_trial=trial,
    )

    assert comparison["comparison_admission"]["status"] == "admitted"
    assert comparison["cutover_gate"]["verdict"] == "FAIL"
    assert comparison["cutover_gate"]["l2_rehearsal_status"] == "observed_failure"
    assert comparison["cutover_gate"]["l2_observed_trial_status"] == "observed_failure"


def test_canonical_comparison_rejects_same_bytes_under_another_new_identity(tmp_path) -> None:
    before, recovery, trial = _comparison_trial(tmp_path)
    comparison = compare_bound_pair(
        before,
        recovery,
        before_binding=_comparison_binding(before, 100),
        after_binding=_comparison_binding(recovery, 999),
        l2_failure_trial=trial,
    )

    assert comparison["comparison_admission"]["status"] == "admitted"
    assert comparison["cutover_gate"]["verdict"] == "INDETERMINATE"
    assert comparison["cutover_gate"]["l2_observed_trial_status"] == "not_verified"
    assert "different recovery" in comparison["cutover_gate"]["l2_observed_trial_note"]
    assert comparison["operator_evidence"]["rehearsal"]["status"] == "not_verified"
    assert comparison["operator_evidence"]["rehearsal"][
        "assurance_level"
    ] == "not_verified"


def test_canonical_comparison_rejects_trial_that_predates_its_before_snapshot(
        tmp_path) -> None:
    before, recovery, trial = _comparison_trial(tmp_path)
    stale_context = _bind(dict(before), "2026-08-20T00:01:00+00:00")
    comparison = compare_bound_pair(
        stale_context,
        recovery,
        before_binding=_comparison_binding(stale_context, 100),
        after_binding=_comparison_binding(recovery, 103),
        l2_failure_trial=trial,
    )

    assert comparison["comparison_admission"]["status"] == "admitted"
    assert comparison["operator_evidence"]["rehearsal"]["status"] == "not_verified"
    assert comparison["cutover_gate"]["verdict"] == "INDETERMINATE"
    assert comparison["cutover_gate"]["l2_observed_trial_status"] == "not_verified"
    assert "context does not reconcile" in comparison["cutover_gate"][
        "l2_observed_trial_note"
    ]


def test_canonical_comparison_refuses_missing_before_collection_chronology(
        tmp_path) -> None:
    before, recovery, trial = _comparison_trial(tmp_path)
    missing_payload = dict(before)
    missing_payload.pop("collected_at", None)
    missing_before = bind_snapshot_json_bytes(json.dumps(
        missing_payload,
        sort_keys=True,
        ensure_ascii=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8"))
    comparison = compare_bound_pair(
        missing_before,
        recovery,
        before_binding=_comparison_binding(missing_before, 100),
        after_binding=_comparison_binding(recovery, 103),
        l2_failure_trial=trial,
    )

    assert comparison["comparison_admission"]["status"] == "admitted"
    assert comparison["operator_evidence"]["rehearsal"]["status"] == "not_verified"
    assert comparison["cutover_gate"]["verdict"] == "INDETERMINATE"
    assert comparison["cutover_gate"]["l2_observed_trial_status"] == "not_verified"


def test_diff_workbook_projects_the_same_observed_trial_gate_and_phase_receipts(
        tmp_path) -> None:
    before, recovery, trial = _comparison_trial(tmp_path / "sources")
    comparison = compare_bound_pair(
        before,
        recovery,
        before_binding=_comparison_binding(before, 100),
        after_binding=_comparison_binding(recovery, 103),
        l2_failure_trial=trial,
    )
    output = tmp_path / "observed-l2-trial.xlsx"

    projected_gate = write_diff_workbook(
        before, recovery, str(output), comparison=comparison
    )

    assert projected_gate == comparison["cutover_gate"]
    workbook = load_workbook(output, read_only=True)
    summary = workbook["Summary"]
    rows = {
        summary.cell(row, 1).value: (
            summary.cell(row, 3).value,
            summary.cell(row, 4).value,
        )
        for row in range(2, summary.max_row + 1)
    }
    assert rows["L2 OBSERVED LOCAL TRIAL"][0] == "LOCAL_SAFETY_PRESERVATION"
    assert "matched projected risks=0" in rows["L2 OBSERVED LOCAL TRIAL"][1]
    observed = workbook["Observed L2 Trial"]
    observed_rows = {
        observed.cell(row, 1).value: observed.cell(row, 2).value
        for row in range(2, observed.max_row + 1)
    }
    workbook.close()
    assert observed_rows["Receipt status"] == "observed_survival"
    assert observed_rows["Assurance"] == "local_safety_preservation"
    assert observed_rows["Service-path survival"] == "not_verified"
    assert observed_rows["recovery source ID"] == "snapshot:103"
