"""NEW-V3.23.101: the L4 service map -- ACL port references + multicast activity resolved to named
services via the offline port registry. ACL refs are design intent (Inferred), not active traffic."""
from cisco_toolkit.analyze import compute_ptp_readiness, compute_service_map
from cisco_toolkit.model import InterfaceData


def _sm_with_ptp(ptp, groups=None):
    return {"multicast": {"ptp": ptp, "classified_groups": groups or []}}


def test_ptp_readiness_all_dormant_is_flagged():
    # the real Meridian case: PTP present on several switches but none operational, PTP multicast flowing
    sm = _sm_with_ptp(
        {"sw1": {"operational": False}, "sw2": {"operational": False}},
        [{"group": "224.0.1.129", "name": "PTP-primary", "broadcast": True}])
    out = compute_ptp_readiness(sm)
    assert len(out) == 1 and out[0]["severity"] == "Medium" and out[0]["category"] == "Timing/PTP"
    assert sorted(out[0]["devices"]) == ["sw1", "sw2"]
    assert "not boundary-clocked" in out[0]["title"] and "224.0.1.129" in out[0]["detail"]


def test_ptp_readiness_partial_is_low_and_lists_dormant_only():
    sm = _sm_with_ptp({"sw1": {"operational": True}, "sw2": {"operational": False}})
    out = compute_ptp_readiness(sm)
    assert len(out) == 1 and out[0]["severity"] == "Low" and out[0]["devices"] == ["sw2"]


def test_ptp_readiness_silent_when_all_operational_or_absent():
    assert compute_ptp_readiness(_sm_with_ptp({"sw1": {"operational": True}})) == []
    assert compute_ptp_readiness(_sm_with_ptp({})) == []
    assert compute_ptp_readiness({}) == [] and compute_ptp_readiness({"multicast": {}}) == []


def _acl(action="permit", proto="udp", dport=None, sport=None, src=None, dst=None):
    return {"action": action, "proto": proto, "sport": sport, "dport": dport,
            "src": src or {"ip": "0.0.0.0", "wild": "255.255.255.255"},
            "dst": dst or {"ip": "0.0.0.0", "wild": "255.255.255.255"}, "raw": ""}


def test_service_map_resolves_acl_ports():
    acls = {
        "sw1": {"RELAY": [_acl(dport={"op": "eq", "val": 67}),
                          _acl(dport={"op": "eq", "val": 68})]},
        "sw2": {"BFD": [_acl(dport={"op": "eq", "val": 3784})],
                "MGMT": [_acl(proto="ip")]},          # ip rule, no L4 port -> ignored
    }
    sm = compute_service_map(acls, {})
    by = {(s["port"], s["proto"]): s for s in sm["services"]}
    assert by[(67, "udp")]["service"] == "bootps" and by[(67, "udp")]["category"] == "Infra"
    assert by[(3784, "udp")]["service"] == "bfd-control" and by[(3784, "udp")]["host_count"] == 1
    assert "Inferred" in by[(67, "udp")]["evidence_class"]
    assert sm["acl_rule_count"] == 4
    # categories aggregate refs
    cats = {c["category"]: c["refs"] for c in sm["categories"]}
    assert cats["Infra"] == 2 and cats["Routing"] == 1


def test_service_map_multicast_activity_and_classification():
    # PIM/mroute presence on interfaces = Confirmed multicast forwarding
    ai = {
        "sw1": {"Vlan10": InterfaceData(port="Vlan10", multicast_info="PIM Sparse / Mroute OIL"),
                "Gi1/0/1": InterfaceData(port="Gi1/0/1")},
        "sw2": {"Vlan20": InterfaceData(port="Vlan20", multicast_info="PIM enabled")},
    }
    # a multicast group referenced in an ACL gets classified
    acls = {"sw1": {"MC": [_acl(dst={"ip": "224.0.1.129", "wild": "0.0.0.0"})]}}
    sm = compute_service_map(acls, ai)
    mc = sm["multicast"]
    assert mc["active_interfaces"] == 2 and mc["active_switch_count"] == 2
    assert mc["group_level_collected"] is False          # richer collection not present
    groups = {g["group"]: g for g in mc["classified_groups"]}
    assert groups["224.0.1.129"]["name"] == "PTP-primary" and groups["224.0.1.129"]["broadcast"] is True


def test_service_map_empty_inputs():
    sm = compute_service_map({}, {})
    assert sm["services"] == [] and sm["acl_rule_count"] == 0
    assert sm["multicast"]["active_interfaces"] == 0


def test_service_map_tolerates_malformed_rules():
    # a stray non-dict rule entry must be skipped, not crash (defensive, V3.23.105)
    acls = {"sw1": {"A": ["junk", {"proto": "udp", "sport": None, "dport": {"op": "eq", "val": 67},
                                   "src": None, "dst": None}]}}
    sm = compute_service_map(acls, {})
    assert {(s["port"], s["service"]) for s in sm["services"]} == {(67, "bootps")}
    assert sm["acl_rule_count"] == 1   # only the well-formed rule counted


# ---------------------------------------------------------------- registry AUTHORITY survives the map
def test_service_entry_carries_the_registry_authority_labels():
    """`svc.setdefault(key, {"service":..., "category":..., "broadcast":..., "refs":..., "hosts":...})`
    projected THREE of portdb's fourteen record keys, so every authority label the registry stamps was
    dropped one hop from the lookup.

    Consequence measured on the shipped pack: 4440/udp resolves to the CURATED-ONLY overlay record
    "Dante-audio" (assignment_authoritative False, semantics_authoritative False, overlay_status
    'overlay-only'), and it reached the workbook with nothing to distinguish it from an IANA-assigned
    row -- while `evidence_class` can independently say "Confirmed (ACL hit-counts)", which confirms
    the TRAFFIC, not the service NAME. A renderer had no field to label it with.
    """
    acls = {"sw1": {"AV": [_acl(dport={"op": "eq", "val": 4440})]}}
    e = compute_service_map(acls, {})["services"][0]
    assert e["service"] == "Dante-audio", e
    for key in ("assignment_authoritative", "semantics_authoritative", "overlay_status"):
        assert key in e, f"authority key {key!r} was dropped: {sorted(e)}"
    assert e["assignment_authoritative"] is False and e["semantics_authoritative"] is False, e
    assert e["overlay_status"] == "overlay-only", e
    # the working-set key must NOT leak into the snapshot (a set is not JSON-serialisable)
    assert "hosts" not in e, e

    # NON-VACUITY: an IANA-assigned port must come back AUTHORITATIVE through the same code path,
    # or the labels are a constant and prove nothing.
    iana = compute_service_map({"sw1": {"M": [_acl(proto="tcp", dport={"op": "eq", "val": 22})]}},
                               {})["services"][0]
    assert iana["service"] == "ssh", iana
    assert iana["assignment_authoritative"] is True and iana["overlay_status"] != "overlay-only", iana


def test_multicast_entry_carries_authority_and_marks_the_unclassified():
    """All 21 curated multicast classifications reached output with no authority metadata at all, and
    a group with NO registry match was rendered with the placeholder category "Multicast" -- which a
    reader cannot tell apart from a real classification."""
    sm = compute_service_map({}, {}, igmp_groups=["224.0.1.129", "239.70.70.70"])
    by = {g["group"]: g for g in sm["multicast"]["classified_groups"]}
    known = by["224.0.1.129"]
    assert known["name"] == "PTP-primary"
    assert known["overlay_status"] == "curated-only", known
    assert known["semantics_authoritative"] is False, "a curated row must never claim authority"
    unknown = by["239.70.70.70"]
    # NON-VACUITY: the unmatched group must be DISTINGUISHABLE from the curated one, not merely
    # non-authoritative like it.
    assert unknown["overlay_status"] == "unclassified", unknown
    assert unknown["overlay_status"] != known["overlay_status"]
    assert unknown["assignment_source"] == "" and known["assignment_source"] == "curated-multicast"


# ------------------------------------------------- MAC-alias severity must disclose what raised it
def test_mac_alias_high_severity_discloses_that_the_av_label_is_curated():
    """`"severity": "High" if a["has_av"] else "Medium"` promoted a finding on the strength of the
    CURATED broadcast/category fields alone -- a non-authoritative classification silently deciding a
    severity that renderers colour and sort on. The overlap arithmetic is a fact; the on-air label is
    not, and the record must say which is which."""
    from cisco_toolkit.analyze import compute_multicast_intelligence
    sm = compute_service_map({}, {}, igmp_groups=["224.0.1.129", "225.0.1.129"])
    r = next(x for x in compute_multicast_intelligence(sm, {})["risks"] if x["kind"] == "mac-alias")
    assert r["severity"] == "High"                       # not weakened -- under-reporting is worse
    assert "NOT an authoritative source" in r["severity_basis"], r["severity_basis"]
    assert "curated, unverified" in r["evidence_confidence"], r["evidence_confidence"]
    assert "not in question" in r["severity_basis"], "the observed part must be stated as observed"

    # NON-VACUITY: a clash with NO on-air member stays Medium and must NOT acquire the curated-AV
    # caveat -- otherwise the disclosure is always-on and carries no information.
    sm2 = compute_service_map({}, {}, igmp_groups=["239.70.70.70", "238.70.70.70"])
    r2 = next(x for x in compute_multicast_intelligence(sm2, {})["risks"] if x["kind"] == "mac-alias")
    assert r2["severity"] == "Medium", r2
    assert "NOT an authoritative source" not in r2["severity_basis"], r2["severity_basis"]


def test_av_group_headline_count_is_paired_with_its_authoritative_subset():
    """`n_av_groups` alone presents a curated judgement as a measurement. The summary must publish how
    many of those on-air classifications rest on an authoritative source (0 on the shipped pack)."""
    from cisco_toolkit.analyze import compute_multicast_intelligence
    sm = compute_service_map({}, {}, igmp_groups=["224.0.1.129", "239.70.70.70"])
    s = compute_multicast_intelligence(sm, {})["summary"]
    assert s["n_av_groups"] == 1
    assert s["n_av_groups_authoritative"] == 0, s
    assert s["n_unclassified_groups"] == 1, s
    # NON-VACUITY: the unclassified counter must be 0 when every group DID classify.
    s2 = compute_multicast_intelligence(
        compute_service_map({}, {}, igmp_groups=["224.0.1.129"]), {})["summary"]
    assert s2["n_unclassified_groups"] == 0, s2


def test_the_on_air_basis_fields_are_the_ones_the_renderers_actually_key_off():
    """PRODUCER <-> RENDERER contract. The workbook and the explorer decide what to disclose from
    `on_air_authoritative` / `has_av_authoritative` / `n_av_groups_authoritative`, and both fail
    CLOSED when a field is absent ("basis NOT published"). If the producer ever renames or drops one,
    every on-air surface silently degrades to an un-evidenced disclaimer instead of the real basis --
    so bind the names here, at the producer, and prove the renderer's own text turns on them."""
    from cisco_toolkit.analyze import compute_multicast_intelligence
    from cisco_toolkit.excel import (mac_alias_on_air_cell, multicast_class_authority,
                                     multicast_on_air_cell)
    mi = compute_multicast_intelligence(
        compute_service_map({}, {}, igmp_groups=["224.0.1.129", "225.0.1.129"]), {})
    av = next(g for g in mi["groups"] if g["group"] == "224.0.1.129")
    alias = mi["mac_aliases"][0]
    assert "on_air_authoritative" in av and "has_av_authoritative" in alias
    assert "n_av_groups_authoritative" in mi["summary"]
    # the renderer's own strings, driven by the producer's own record
    assert multicast_on_air_cell(av) == "yes (CURATED, unverified)"
    assert mac_alias_on_air_cell(alias) == "yes (CURATED, unverified)"
    assert "CURATED, NOT authoritative" in multicast_class_authority(av)
    unmatched = next(g for g in mi["groups"] if g["group"] == "225.0.1.129")
    assert "UNCLASSIFIED" in multicast_class_authority(unmatched)
    # NON-VACUITY: a group that is NOT on air gets no cell text at all, and a record carrying no
    # authority labels gets no claim either way.
    assert multicast_on_air_cell(unmatched) == ""
    assert multicast_class_authority({"group": "224.0.1.129", "category": "Broadcast-AV"}) == ""


# ---------------------------------- the punch-list fold must not erase what RAISED a High (EXIT A)
def _media_risks(groups, authoritative=False):
    """Media risks from the REAL producer chain (compute_service_map -> compute_multicast_
    intelligence), filtered exactly the way COLLECT_PARSE_V3_23_0.py:3434 filters them before the
    fold. `authoritative=True` flips the registry's own semantics flag on the classified rows --
    the shipped pack has NO authoritative multicast row, so that branch is otherwise unreachable and
    the non-vacuity check below would be untestable."""
    from cisco_toolkit.analyze import compute_multicast_intelligence
    sm = compute_service_map({}, {}, igmp_groups=groups)
    if authoritative:
        for g in sm["multicast"]["classified_groups"]:
            g["semantics_authoritative"] = True
    mi = compute_multicast_intelligence(sm, {})
    return [r for r in mi["risks"] if r.get("kind") in ("mac-alias", "querier-gap")]


def _punchlist(media_risks=None, **kw):
    """The fleet punch-list with every other fold empty, so nothing below can be satisfied by an
    unrelated row."""
    from cisco_toolkit.analyze import compute_migration_punchlist
    return compute_migration_punchlist([], {}, {}, [], [], [], {}, [], [],
                                       media_risks=media_risks, **kw)


def _punch_detail_cell(punchlist, title_needle):
    """The Detail cell the WORKBOOK actually writes (excel.write_punchlist_sheet, column 7) -- the
    consumer this fold feeds. Asserting on the dict alone would not prove a reader sees anything."""
    from openpyxl import Workbook
    from cisco_toolkit.excel import PUNCHLIST_SHEET_NAME, write_punchlist_sheet
    wb = Workbook()
    write_punchlist_sheet(wb, punchlist)
    ws = wb[PUNCHLIST_SHEET_NAME]
    for row in range(2, ws.max_row + 1):
        if title_needle in str(ws.cell(row, 6).value or ""):
            return str(ws.cell(row, 7).value or "")
    raise AssertionError(f"no punch-list row titled ~{title_needle!r}")


def test_punchlist_fold_carries_the_curated_basis_of_a_high_media_finding():
    """review r10 EXIT A. compute_multicast_intelligence raises the mac-alias risk Medium->High
    purely on a CURATED on-air classification and publishes `severity_basis` / `evidence_confidence`
    saying so (analyze.py :2774). compute_migration_punchlist DROPPED both keys, so the fleet
    punch-list -- and every consumer downstream of it -- carried a High whose basis had vanished: an
    unverified registry hint rendered in the same voice as an IANA-assigned fact."""
    risks = _media_risks(["224.0.1.129", "225.0.1.129"])
    src = next(r for r in risks if r["kind"] == "mac-alias")
    it = next(i for i in _punchlist(risks) if i["category"] == "Multicast/Media")
    assert it["severity"] == "High"                    # NOT re-scored -- the fix is disclosure
    # the producer's own basis travels, verbatim and un-paraphrased
    assert it["severity_basis"] == src["severity_basis"]
    assert it["evidence_confidence"] == src["evidence_confidence"]
    # ...and it REACHES A READER: the workbook's Detail column is where a punch-list row is read.
    cell = _punch_detail_cell(_punchlist(risks), "MAC-address overlap")
    assert "NOT an authoritative source" in cell, cell
    assert "curated, unverified" in cell, cell
    assert "Severity basis:" in cell and "Evidence:" in cell, cell


def test_punchlist_fold_fails_closed_when_a_media_risk_publishes_no_usable_basis():
    """An older snapshot's risk row carries no basis keys at all -- and the fail-OPEN shape this
    review keeps finding is a key-PRESENCE test, so a null / empty / non-string value must be treated
    exactly like a missing one. Either way the row must say the basis is unpublished, never leave a
    High standing as though it had been measured."""
    from cisco_toolkit.analyze import PUNCH_BASIS_UNPUBLISHED, PUNCH_CONFIDENCE_UNPUBLISHED
    base = {"kind": "mac-alias", "severity": "High", "title": "Multicast MAC-address overlap on X",
            "detail": "d", "remediation": "r"}
    for bad in ({}, {"severity_basis": None, "evidence_confidence": None},
                {"severity_basis": "   ", "evidence_confidence": ""},
                {"severity_basis": {"a": 1}, "evidence_confidence": 0}):
        it = next(i for i in _punchlist([dict(base, **bad)]) if i["category"] == "Multicast/Media")
        assert it["severity_basis"] == PUNCH_BASIS_UNPUBLISHED, (bad, it)
        assert it["evidence_confidence"] == PUNCH_CONFIDENCE_UNPUBLISHED, (bad, it)
        cell = _punch_detail_cell(_punchlist([dict(base, **bad)]), "MAC-address overlap")
        assert "NOT published by this snapshot" in cell, (bad, cell)
        # Asserts the PROPERTY (absence of a basis is disclosed), not the old sentence. The wording
        # deliberately no longer claims the severity was unmeasured: this sentinel lands on EVERY
        # media risk including querier-gap, whose High IS a measurement, so asserting "not measured"
        # was false for that row and taught readers to discount a real finding.
        assert "severity basis NOT published" in cell, (bad, cell)
        assert "as measured" not in cell, (
            "the sentinel again asserts the severity was unmeasured; it lands on measured rows too")


def test_punchlist_basis_caveat_is_not_always_on():
    """NON-VACUITY, both directions. (1) A fleet whose on-air classification IS authoritative gets
    the AUTHORITATIVE basis -- no 'curated' and no 'NOT published' wording. (2) A fleet with no media
    estate at all must not acquire the caveat anywhere: no basis keys on any other fold's rows."""
    auth = _media_risks(["224.0.1.129", "225.0.1.129"], authoritative=True)
    src = next(r for r in auth if r["kind"] == "mac-alias")
    assert "NOT an authoritative source" not in src["severity_basis"], \
        "the producer itself must have taken the authoritative branch, or this proves nothing"
    it = next(i for i in _punchlist(auth) if i["category"] == "Multicast/Media")
    assert it["severity"] == "High"
    assert "authoritative source" in it["severity_basis"]
    assert "NOT an authoritative source" not in it["severity_basis"], it["severity_basis"]
    assert "NOT published by this snapshot" not in it["severity_basis"], it["severity_basis"]
    cell = _punch_detail_cell(_punchlist(auth), "MAC-address overlap")
    assert "curated, unverified" not in cell, cell
    assert "NOT published by this snapshot" not in cell, cell

    # (2) no media estate -> no Multicast/Media row, and no OTHER row grows a basis clause.
    drift = [{"severity": "High", "category": "False-health", "devices": ["sw1"],
              "title": "Interface counters frozen", "detail": "d", "remediation": "r"}]
    other = _punchlist([], drift=drift)
    assert other and not any(i["category"] == "Multicast/Media" for i in other)
    for i in other:
        assert "severity_basis" not in i and "evidence_confidence" not in i, i
        assert "NOT published by this snapshot" not in i["detail"], i
        assert "Severity basis:" not in i["detail"], i


def test_punchlist_basis_survives_a_pathologically_long_detail():
    """The disclosure is appended AFTER the 400-char pathological-length clip, so a long producer
    detail can never be the reason the basis is the part that gets truncated away."""
    it = next(i for i in _punchlist([{"kind": "mac-alias", "severity": "High", "title": "t",
                                      "detail": "x" * 5000, "remediation": "r",
                                      "severity_basis": "BASIS-SENTINEL-TOKEN",
                                      "evidence_confidence": "CONF-SENTINEL-TOKEN"}])
               if i["category"] == "Multicast/Media")
    assert "…" in it["detail"], "the long body must still be clipped"
    assert it["detail"].endswith("[Severity basis: BASIS-SENTINEL-TOKEN | "
                                 "Evidence: CONF-SENTINEL-TOKEN]"), it["detail"][-160:]
