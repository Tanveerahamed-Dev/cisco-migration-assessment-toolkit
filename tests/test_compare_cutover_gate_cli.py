"""One combined --compare verdict owns workbook, terminal, and opt-in exit behavior."""
from copy import deepcopy
import json
import os
import subprocess
import sys

import pytest
from openpyxl import load_workbook

from cisco_toolkit.comparison import compare_bound_pair
from cisco_toolkit.html import (compute_cutover_gate, compute_snapshot_delta,
                                write_diff_workbook)
from cisco_toolkit.precert import compute_precert
from cisco_toolkit.protocol_assurance import (
    OFFLINE_FILE_SOURCE,
    bind_snapshot_json_bytes,
    compute_native_protocol_deltas,
    protocol_family_change_set,
    receipt_envelope,
)


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SCRIPT = os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py")
FAMILIES = ("STP", "EtherChannel", "VTP", "OSPF", "BGP", "EIGRP", "FHRP")


def _routes():
    return {
        "R1": [
            {"prefix": "10.1.1.0/24", "source": "connected", "out_intf": "Vlan10"},
            {"prefix": "10.0.12.0/30", "source": "connected", "out_intf": "Gi0/1"},
            {"prefix": "10.2.2.0/24", "source": "ospf", "next_hop": "10.0.12.2",
             "out_intf": "Gi0/1"},
        ],
        "R2": [
            {"prefix": "10.2.2.0/24", "source": "connected", "out_intf": "Vlan20"},
            {"prefix": "10.0.12.0/30", "source": "connected", "out_intf": "Gi0/1"},
            {"prefix": "10.1.1.0/24", "source": "ospf", "next_hop": "10.0.12.1",
             "out_intf": "Gi0/1"},
        ],
    }


def _interfaces():
    return {
        "R1": {"Vlan10": {"svi_ip": "10.1.1.1/24"},
               "Gi0/1": {"svi_ip": "10.0.12.1/30"}},
        "R2": {"Gi0/1": {"svi_ip": "10.0.12.2/30"},
               "Vlan20": {"svi_ip": "10.2.2.1/24"}},
    }


def _receipt():
    rows = []
    for family in FAMILIES:
        assessed = family == "OSPF"
        rows.append({
            "switch": "R1",
            "protocol": family,
            "state": "assessed" if assessed else "not_collected",
            "health_row_emitted": assessed,
        })
    return {
        "schema": "protocol_assessability/1",
        "families": [{"protocol": family} for family in FAMILIES],
        "rows": rows,
        "summary": {"n_devices": 1, "n_families": len(FAMILIES), "n_cells": len(rows)},
    }


def _clean_validation_plan():
    item = {
        "device": "R1", "platform": "ios", "wave": "Group 1",
        "category": "Gateway", "severity": "High", "check": "Gateway remains up",
        "command": "show ip interface brief", "expect": "Vlan10 remains up/up",
        "why": "Preserve the observed gateway.",
    }
    return {
        "items": [item], "by_wave": {"Group 1": [item]},
        "summary": {"n_items": 1, "n_waves": 1, "by_category": {"Gateway": 1}, "n_high": 1},
        "banner": "bounded plan",
    }


def _degraded_validation_plan():
    item = {
        "device": "R1", "platform": "ios", "wave": "Group 1",
        "category": "Routing", "severity": "High", "check": "OSPF degraded adjacency baseline",
        "command": "show ip ospf neighbor",
        "expect": ("PRE-CUTOVER DEGRADED — BLOCKER: OSPF peer 10.0.12.2 is EXSTART/DR. "
                   "Matching this degraded state is NOT ACCEPTANCE."),
        "why": "The current observed adjacency is already degraded.",
        "evidence_state": "degraded", "projection_custody": "embedded_unverified",
        "source_key": "routing_neighbors.R1.ospf + protocol_assessability.rows[R1,OSPF]",
    }
    return {
        "items": [item], "by_wave": {"Group 1": [item]},
        "summary": {"n_items": 1, "n_waves": 1, "by_category": {"Routing": 1}, "n_high": 1},
        "banner": "bounded plan",
    }


def _snapshot(ospf_state, generated_at):
    return {
        "schema": "collect_parse_snapshot/1",
        "script_version": "V3.23.0",
        "generated_at": generated_at,
        "devices": {"R1": {}, "R2": {}},
        "interfaces": _interfaces(),
        "routes": _routes(),
        "segmentation": {
            "vrfs": [{"vrf": "(global)", "gateway_count": 1},
                     {"vrf": "PCI", "gateway_count": 2}],
            "domains": [{"domain": "payments", "tier": "on-air-critical", "isolated": True,
                         "exposure": "Every gateway has a dedicated VRF or a gateway ACL."}],
        },
        "health_scores": [
            {"switch": "R1", "band": "Good", "score": 92},
            {"switch": "R2", "band": "Good", "score": 92},
        ],
        "punchlist": [],
        "validation_plan": _clean_validation_plan(),
        "protocol_assessability": _receipt(),
        "routing_neighbors": {
            "R1": {"ospf": [{"neighbor": "10.0.12.2", "state": ospf_state,
                               "interface": "Gi0/1", "address": "10.0.12.2"}]},
        },
    }


def _workbook_gate(path):
    ws = load_workbook(path, read_only=True)["Summary"]
    rows = {ws.cell(row, 1).value: ws.cell(row, 3).value for row in range(2, ws.max_row + 1)}
    return rows["CUTOVER GATE VERDICT"]


def _comparison_context(tmp_path, name="comparison-context"):
    path = tmp_path / f"{name}.json"
    path.write_text(json.dumps({
        "schema": "offline_comparison_context/1",
        "engagement_id": "ENG-CLI-TEST",
        "campaign_id": 41,
        "before_snapshot_id": 1001,
        "after_snapshot_id": 1002,
        "change_intent": {"expected_changes": [], "note": "focused CLI test"},
    }), encoding="utf-8")
    return path


@pytest.mark.parametrize(
    "context_text",
    [
        (
            '{"schema":"offline_comparison_context/1",'
            '"engagement_id":"ENG-FIRST","engagement_id":"ENG-LAST",'
            '"campaign_id":41,"before_snapshot_id":1001,"after_snapshot_id":1002}'
        ),
        (
            '{"schema":"offline_comparison_context/1","engagement_id":"ENG-CLI-TEST",'
            '"campaign_id":41,"before_snapshot_id":1001,"after_snapshot_id":1002,'
            '"change_intent":{"expected_changes":[],"note":"first","note":"last"}}'
        ),
    ],
)
def test_compare_cli_refuses_duplicate_top_level_or_nested_context_members(
        tmp_path, context_text):
    before_path = tmp_path / "duplicate-context-before.snapshot.json"
    after_path = tmp_path / "duplicate-context-after.snapshot.json"
    context_path = tmp_path / "duplicate-context.json"
    output = tmp_path / "duplicate-context.xlsx"
    before_path.write_text(
        json.dumps(_snapshot("FULL/DR", "2026-08-15T00:00:00")), encoding="utf-8")
    after_path.write_text(
        json.dumps(_snapshot("FULL/DR", "2026-08-15T00:05:00")), encoding="utf-8")
    context_path.write_text(context_text, encoding="utf-8")

    process = subprocess.run(
        [sys.executable, SCRIPT, "--compare", str(before_path), str(after_path),
         "--comparison-context", str(context_path), "--output", str(output)],
        cwd=str(tmp_path), capture_output=True, text=True, timeout=300,
    )
    terminal = process.stdout + process.stderr

    assert process.returncode == 2, terminal
    assert "duplicate JSON object key" in terminal
    assert not output.exists()


@pytest.mark.parametrize(
    "change_intent",
    [
        {"expected_change": []},
        {
            "expected_changes": [{
                "family": "ipv4_routing_adjacency",
                "transitions": ["appeared"],
                "subjects": [],
                "reasno": "misspelled reason",
            }],
        },
    ],
)
def test_compare_cli_unknown_change_intent_fields_cannot_authorize_pass(
        tmp_path, change_intent):
    before_path = tmp_path / "unknown-intent-before.snapshot.json"
    after_path = tmp_path / "unknown-intent-after.snapshot.json"
    context_path = tmp_path / "unknown-intent-context.json"
    output = tmp_path / "unknown-intent.xlsx"
    before_path.write_text(
        json.dumps(_snapshot("FULL/DR", "2026-08-15T00:00:00")), encoding="utf-8")
    after_path.write_text(
        json.dumps(_snapshot("FULL/DR", "2026-08-15T00:05:00")), encoding="utf-8")
    context_path.write_text(json.dumps({
        "schema": "offline_comparison_context/1",
        "engagement_id": "ENG-CLI-TEST",
        "campaign_id": 41,
        "before_snapshot_id": 1001,
        "after_snapshot_id": 1002,
        "change_intent": change_intent,
    }), encoding="utf-8")

    process = subprocess.run(
        [sys.executable, SCRIPT, "--compare", str(before_path), str(after_path),
         "--comparison-context", str(context_path), "--output", str(output),
         "--fail-on-compare-gate"],
        cwd=str(tmp_path), capture_output=True, text=True, timeout=300,
    )
    terminal = process.stdout + process.stderr

    assert process.returncode == 2, terminal
    assert output.exists()
    comparison = json.loads(
        (tmp_path / "unknown-intent.comparison.json").read_text(encoding="utf-8")
    )
    assert comparison["change_intent"]["status"] == "invalid"
    assert comparison["comparison_admission"]["status"] == "not_comparable"
    assert comparison["cutover_gate"]["verdict"] == "INDETERMINATE"


@pytest.mark.parametrize(
    ("delta_verdict", "certificate_verdict", "expected"),
    [
        ("REGRESSED", "FAIL", "REGRESSED"),
        ("CLEAN", "FAIL", "FAIL"),
        ("INDETERMINATE", "FAIL", "FAIL"),
        ("INDETERMINATE", "PASS", "INDETERMINATE"),
        ("REVIEW", "CONDITIONAL", "REVIEW"),
        ("CLEAN", "CONDITIONAL", "CONDITIONAL"),
        ("CLEAN", "PASS", "PASS"),
    ],
)
def test_compute_cutover_gate_preserves_precedence(delta_verdict, certificate_verdict, expected):
    protocol_gate = ("REGRESSED" if delta_verdict == "REGRESSED" else
                     "REVIEW" if delta_verdict == "REVIEW" else "PASS")
    protocol_regressions = 2 if protocol_gate == "REGRESSED" else 0
    protocol_coverage_gaps = 1 if protocol_gate == "REVIEW" else 0
    protocol_baseline_peers = 2 if protocol_gate != "PASS" else 0
    delta = {
        "verdict": delta_verdict,
        "verdict_display": "NO DELTA REGRESSION OBSERVED" if delta_verdict == "CLEAN" else delta_verdict,
        "verdict_note": "delta note",
        "protocol_adjacencies": {
            "gate": protocol_gate,
            "summary": {"n_state_regressed": protocol_regressions,
                        "n_coverage_gaps": protocol_coverage_gaps,
                        "n_baseline_peers": protocol_baseline_peers},
        },
    }
    certificate = {"verdict": certificate_verdict, "verdict_note": "certificate note"}
    decision = compute_cutover_gate(delta, certificate)

    assert decision == {
        "schema": "cutover_gate/1",
        "verdict": expected,
        "note": (f"Delta observation: {delta['verdict_display']}. Delta basis: delta note "
                 f"Pre-Change Certificate: {certificate_verdict}. "
                 "Certificate basis: certificate note"),
        "operator_note": decision["operator_note"],
        "delta_verdict": delta_verdict,
        "delta_display": delta["verdict_display"],
        "delta_note": "delta note",
        "certificate_verdict": certificate_verdict,
        "certificate_note": "certificate note",
        "protocol_gate": protocol_gate,
        "protocol_baseline_peers": protocol_baseline_peers,
        "protocol_regressions": protocol_regressions,
        "protocol_coverage_gaps": protocol_coverage_gaps,
    }
    assert decision["operator_note"].startswith(f"Overall before/after cutover decision: {expected}.")


@pytest.mark.parametrize(
    ("protocol_gate", "regressions", "expected"),
    [("REGRESSED", 1, "REGRESSED"), ("PASS", 1, "REGRESSED"), ("REVIEW", 0, "REVIEW")],
)
def test_cutover_gate_never_passes_over_a_stricter_nested_protocol_result(
        protocol_gate, regressions, expected):
    decision = compute_cutover_gate(
        {
            "verdict": "CLEAN",
            "verdict_display": "NO DELTA REGRESSION OBSERVED",
            "verdict_note": "contradictory outer delta",
            "protocol_adjacencies": {
                "gate": protocol_gate,
                "summary": {"n_baseline_peers": 1, "n_state_regressed": regressions,
                            "n_coverage_gaps": 0},
            },
        },
        {"verdict": "PASS", "verdict_note": "bounded certificate"},
    )

    assert decision["verdict"] == expected
    assert decision["delta_verdict"] == expected
    assert decision["delta_display"] == expected


def test_cutover_gate_keeps_integrity_indeterminate_over_an_apparent_protocol_regression():
    decision = compute_cutover_gate(
        {
            "verdict": "INDETERMINATE",
            "verdict_display": "INDETERMINATE",
            "verdict_note": "Cross-schema comparison; observations are apparent only.",
            "protocol_adjacencies": {
                "gate": "REGRESSED",
                "summary": {"n_baseline_peers": 1, "n_state_regressed": 1,
                            "n_coverage_gaps": 0},
            },
        },
        {"verdict": "CONDITIONAL", "verdict_note": "schema mismatch"},
    )

    assert decision["verdict"] == "INDETERMINATE"
    assert decision["delta_verdict"] == "INDETERMINATE"
    assert "proven" not in decision["operator_note"]


def test_cutover_gate_withholds_pass_for_legacy_receipt_when_baseline_peers_exist():
    decision = compute_cutover_gate(
        {
            "verdict": "CLEAN",
            "verdict_display": "NO DELTA REGRESSION OBSERVED",
            "verdict_note": "other axes clean",
            "protocol_adjacencies": {
                "gate": "NOT_ASSESSED",
                "assessed": False,
                "summary": {"n_baseline_peers": 1, "n_state_regressed": 0,
                            "n_coverage_gaps": 1},
            },
        },
        {"verdict": "PASS", "verdict_note": "bounded certificate"},
    )

    assert decision["verdict"] == "INDETERMINATE"
    assert decision["protocol_baseline_peers"] == 1
    assert "re-collect protocol evidence" in decision["operator_note"]


def test_write_diff_workbook_returns_the_exact_public_decision(tmp_path):
    before = _snapshot("FULL/DR", "2026-08-15T00:00:00")
    after = _snapshot("EXSTART/DR", "2026-08-15T00:05:00")
    certificate = compute_precert(before, after)
    delta = compute_snapshot_delta(before, after)
    families = protocol_family_change_set(
        delta["protocol_adjacencies"], {"expected_changes": []},
        native_deltas=compute_native_protocol_deltas(before, after),
    )
    expected = compute_cutover_gate(
        delta, certificate, protocol_family_changes=families)
    out = tmp_path / "returned-gate.xlsx"

    actual = write_diff_workbook(before, after, str(out), precert=certificate)

    assert actual == expected
    assert actual["certificate_verdict"] == "PASS"
    assert actual["verdict"] == "REGRESSED"
    assert actual["protocol_baseline_peers"] == 1
    assert actual["protocol_regressions"] == 1
    assert _workbook_gate(str(out)) == actual["verdict"]
    workbook = load_workbook(out, read_only=True)
    assert "Protocol Adjacency Delta" in workbook.sheetnames
    assert "Protocol Family Changes" in workbook.sheetnames
    family_rows = list(workbook["Protocol Family Changes"].iter_rows(values_only=True))
    assert family_rows[0] == (
        "Family", "Subject kind", "Subject", "Transition", "Expected",
        "Producer decision_effect", "Assurance", "Before", "After", "Note",
    )
    assert any(row[3] == "regressed" and row[5] == "block" for row in family_rows[1:])


def test_workbook_projects_one_canonical_comparison_and_complete_receipt(tmp_path):
    before_raw = json.dumps(
        _snapshot("FULL/DR", "2026-08-15T00:00:00"), separators=(",", ":")
    ).encode("utf-8")
    after_raw = json.dumps(
        _snapshot("EXSTART/DR", "2026-08-15T00:05:00"), separators=(",", ":")
    ).encode("utf-8")
    before = bind_snapshot_json_bytes(before_raw)
    after = bind_snapshot_json_bytes(after_raw)

    def binding(raw, snapshot_id, label):
        import hashlib
        return {
            "source": OFFLINE_FILE_SOURCE,
            "sha256": "sha256:" + hashlib.sha256(raw).hexdigest(),
            "bytes": len(raw),
            "snapshot_id": snapshot_id,
            "campaign_id": 41,
            "engagement_id": "ENG-WORKBOOK-PARITY",
            "label": label,
            "script_version": "V3.23.0",
        }

    comparison = compare_bound_pair(
        before,
        after,
        before_binding=binding(before_raw, 1001, "before.json"),
        after_binding=binding(after_raw, 1002, "after.json"),
        change_intent={"expected_changes": [], "note": "workbook parity"},
    )
    output = tmp_path / "canonical-comparison.xlsx"

    workbook_gate = write_diff_workbook(
        before, after, str(output), comparison=comparison)

    assert workbook_gate == comparison["cutover_gate"]
    assert _workbook_gate(str(output)) == comparison["cutover_gate"]["verdict"]
    workbook = load_workbook(output, read_only=True)
    receipt_rows = list(workbook["Comparison Receipt"].iter_rows(values_only=True))
    assert ("Admission", "status", "admitted", "complete") in receipt_rows
    assert any(
        row[0:3] == ("Source before", "sha256",
                     comparison["comparison_admission"]["source_binding"]["before"]["sha256"])
        for row in receipt_rows
    )
    assert any(
        row[0:3] == (
            "Subjects before",
            "subjects_sha256",
            comparison["comparison_admission"]["subject_binding"]["before"][
                "subjects_sha256"
            ],
        )
        for row in receipt_rows
    )
    assert any(
        row[0:3] == (
            "Subjects after",
            "n_subjects",
            comparison["comparison_admission"]["subject_binding"]["after"]["n_subjects"],
        )
        for row in receipt_rows
    )
    assert ("Admission", "failures", "[]", "complete") in receipt_rows
    assert ("Admission", "coverage_gaps", "[]", "complete") in receipt_rows
    assert ("Receipt integrity", "status", "VERIFIED", "complete") in receipt_rows
    disclosure = next(row for row in receipt_rows if row[0:2] == ("Export", "row disclosure"))
    assert "omitted 0" in disclosure[2]
    assert disclosure[3].startswith("complete uncapped")


def test_workbook_rejects_a_rehashed_forged_canonical_gate(tmp_path):
    before_raw = json.dumps(
        _snapshot("FULL/DR", "2026-08-15T00:00:00"), separators=(",", ":")
    ).encode("utf-8")
    after_raw = json.dumps(
        _snapshot("EXSTART/DR", "2026-08-15T00:05:00"), separators=(",", ":")
    ).encode("utf-8")
    before = bind_snapshot_json_bytes(before_raw)
    after = bind_snapshot_json_bytes(after_raw)

    def binding(raw, snapshot_id, label):
        import hashlib
        return {
            "source": OFFLINE_FILE_SOURCE,
            "sha256": "sha256:" + hashlib.sha256(raw).hexdigest(),
            "bytes": len(raw),
            "snapshot_id": snapshot_id,
            "campaign_id": 42,
            "engagement_id": "ENG-WORKBOOK-FORGERY",
            "label": label,
            "script_version": "V3.23.0",
        }

    comparison = compare_bound_pair(
        before,
        after,
        before_binding=binding(before_raw, 1101, "before.json"),
        after_binding=binding(after_raw, 1102, "after.json"),
    )
    assert comparison["cutover_gate"]["verdict"] != "PASS"
    forged = deepcopy(comparison)
    forged["cutover_gate"]["verdict"] = "PASS"
    forged["cutover_gate"]["operator_note"] = "FORGED PASS"
    additive = {
        "comparison_schema", "comparison_admission", "change_intent",
        "protocol_families", "precert", "cutover_gate", "operator_evidence",
        "comparison_receipt",
    }
    forged_delta = {
        key: value for key, value in forged.items() if key not in additive
    }
    forged["comparison_receipt"] = receipt_envelope(
        admission=forged["comparison_admission"],
        change_intent=forged["change_intent"],
        protocol_families=forged["protocol_families"],
        delta=forged_delta,
        precert=forged["precert"],
        cutover_gate=forged["cutover_gate"],
        operator_evidence=forged["operator_evidence"],
    )

    with pytest.raises(ValueError, match="changed after canonical|cutover gate is detached"):
        write_diff_workbook(
            before,
            after,
            str(tmp_path / "forged.xlsx"),
            comparison=forged,
        )

    forged_identity = deepcopy(comparison)
    forged_admission = forged_identity["comparison_admission"]
    forged_admission["engagement_id"] = "ENG-FORGED"
    forged_admission["campaign_id"] = 999
    for side, snapshot_id in (("before", 9001), ("after", 9002)):
        forged_binding = forged_admission["source_binding"][side]
        forged_binding["engagement_id"] = "ENG-FORGED"
        forged_binding["campaign_id"] = 999
        forged_binding["snapshot_id"] = snapshot_id
        forged_binding["label"] = f"forged-{side}.json"
    forged_intent_binding = forged_identity["change_intent"]["binding"]
    forged_intent_binding.update({
        "engagement_id": "ENG-FORGED",
        "campaign_id": 999,
        "before_snapshot_id": 9001,
        "after_snapshot_id": 9002,
    })
    identity_delta = {
        key: value for key, value in forged_identity.items() if key not in additive
    }
    forged_identity["cutover_gate"] = compute_cutover_gate(
        identity_delta,
        forged_identity["precert"],
        comparison_admission=forged_admission,
        protocol_family_changes=forged_identity["protocol_families"],
    )
    forged_identity["comparison_receipt"] = receipt_envelope(
        admission=forged_admission,
        change_intent=forged_identity["change_intent"],
        protocol_families=forged_identity["protocol_families"],
        delta=identity_delta,
        precert=forged_identity["precert"],
        cutover_gate=forged_identity["cutover_gate"],
        operator_evidence=forged_identity["operator_evidence"],
    )
    with pytest.raises(ValueError, match="changed after canonical"):
        write_diff_workbook(
            before,
            after,
            str(tmp_path / "forged-identity.xlsx"),
            comparison=forged_identity,
        )


def test_protocol_family_changes_sheet_is_complete_uncapped_and_gate_reconciled(tmp_path):
    before = _snapshot("FULL/DR", "2026-08-15T00:00:00")
    after = _snapshot("FULL/DR", "2026-08-15T00:05:00")
    certificate = compute_precert(before, after)
    changes = [{
        "family": "stp_topology",
        "subject": f"R1|vlan|{vlan}",
        "transition": "intent_changed",
        "expected": False,
        "decision_effect": "review",
        "before_state": {"root": "R1", "vlan": vlan},
        "after_state": {"root": "R2", "vlan": vlan},
        "note": "Unexpected root movement requires operator review.",
    } for vlan in range(1, 174)]
    by_transition = {
        token: (len(changes) if token == "intent_changed" else 0)
        for token in (
            "unchanged_healthy", "unchanged_degraded", "recovered", "regressed",
            "appeared", "disappeared", "intent_changed", "coverage_lost",
            "not_comparable",
        )
    }
    native = {
        "schema": "stp_topology_delta/1",
        "family": "stp_topology",
        "owner": "stp_topology_delta/1",
        "assurance_level": "not_verified",
        "owns_score": False,
        "owns_verdict": False,
        "applicability": "applicable",
        "comparable": True,
        "assessed": True,
        "source_receipts": {
            side: {
                "present": True,
                "valid": True,
                "source_bound": True,
                "owner_source_authority": True,
                "comparison_source_bound": True,
                "comparison_source_basis": "current_run_owner_source",
                "snapshot_sha256": "",
                "projection_sha256": "sha256:" + digest * 64,
                "reason": "bounded test owner projection",
            }
            for side, digest in (("before", "a"), ("after", "b"))
        },
        "summary": {
            "n_subjects": len(changes),
            "n_comparable": len(changes),
            "by_transition": by_transition,
            "by_decision_effect": {
                "block": 0,
                "review": len(changes),
                "none": 0,
                "not_verified": 0,
            },
        },
        "changes": [
            {key: value for key, value in row.items() if key != "expected"}
            for row in changes
        ],
    }
    delta = compute_snapshot_delta(before, after)
    families = protocol_family_change_set(
        delta["protocol_adjacencies"], {"expected_changes": []},
        native_deltas=[native],
    )
    expected = compute_cutover_gate(
        delta, certificate, protocol_family_changes=families)
    out = tmp_path / "uncapped-family-changes.xlsx"

    actual = write_diff_workbook(
        before, after, str(out), precert=certificate, protocol_families=families)

    assert actual == expected
    assert actual["protocol_family_rows"] == len(changes)
    assert actual["protocol_family_review"] == len(changes)
    assert _workbook_gate(str(out)) == actual["verdict"] == "REVIEW"
    workbook = load_workbook(out, read_only=True)
    assert "Protocol Adjacency Delta" in workbook.sheetnames
    ws = workbook["Protocol Family Changes"]
    rows = list(ws.iter_rows(values_only=True))
    subjects = [row[2] for row in rows[1:] if row[0] == "stp_topology"]
    assert len(subjects) == len(changes)
    assert set(subjects) == {row["subject"] for row in changes}
    totals = next(row for row in rows if row[0] == "FULL TOTALS")
    assert totals[7] == f"rendered {len(changes)} of {len(changes)}"
    assert totals[8] == "omitted 0"
    assert f"rendered {len(changes)} of {len(changes)}" in totals[9]
    assert "omitted=0" in totals[9]
    summary = workbook["Summary"]
    summary_rows = {
        summary.cell(row, 1).value: (summary.cell(row, 3).value, summary.cell(row, 4).value)
        for row in range(2, summary.max_row + 1)
    }
    count, disclosure = summary_rows["Protocol family changes (complete)"]
    assert count == len(changes)
    assert f"rendered {len(changes)} of {len(changes)}" in disclosure
    assert "omitted=0" in disclosure


def test_protocol_family_changes_missing_or_empty_evidence_is_not_verified(tmp_path):
    before = _snapshot("FULL/DR", "2026-08-15T00:00:00")
    after = _snapshot("FULL/DR", "2026-08-15T00:05:00")
    empty_families = {
        "schema": "protocol_family_change_set/1",
        "owner": "reference_only_composition",
        "owns_score": False,
        "owns_verdict": False,
        "families": [{
            "family": "stp_topology",
            "assurance_level": "local_safety_preservation",
            "changes": [],
        }],
    }
    out = tmp_path / "empty-family-changes.xlsx"

    write_diff_workbook(
        before, after, str(out), protocol_families=empty_families)

    workbook = load_workbook(out, read_only=True)
    ws = workbook["Protocol Family Changes"]
    assert ws.cell(2, 1).value == "stp_topology"
    assert all(ws.cell(2, column).value == "NOT VERIFIED" for column in range(2, 10))
    summary = workbook["Summary"]
    row = next(
        row for row in summary.iter_rows(values_only=True)
        if row[0] == "Protocol family changes (complete)"
    )
    assert row[2] == "NOT VERIFIED"
    assert "placeholder NOT VERIFIED rows=1" in row[3]


def test_compare_cli_enforcement_rejects_an_unchanged_degraded_current_baseline(tmp_path):
    before = _snapshot("EXSTART/DR", "2026-08-15T00:00:00")
    after = _snapshot("EXSTART/DR", "2026-08-15T00:05:00")
    before["validation_plan"] = _degraded_validation_plan()
    after["validation_plan"] = _degraded_validation_plan()
    before_path = tmp_path / "unchanged-before.snapshot.json"
    after_path = tmp_path / "unchanged-after.snapshot.json"
    output = tmp_path / "unchanged-degraded.xlsx"
    before_path.write_text(json.dumps(before), encoding="utf-8")
    after_path.write_text(json.dumps(after), encoding="utf-8")
    context = _comparison_context(tmp_path)

    process = subprocess.run(
        [sys.executable, SCRIPT, "--compare", str(before_path), str(after_path),
         "--comparison-context", str(context), "--output", str(output),
         "--fail-on-compare-gate"],
        cwd=str(tmp_path), capture_output=True, text=True, timeout=300,
    )
    terminal = process.stdout + process.stderr

    assert process.returncode == 2, terminal
    assert output.exists()
    assert _workbook_gate(str(output)) == "FAIL"
    assert "[CUTOVER GATE: FAIL]" in terminal
    assert "current-baseline degradation" in terminal


def test_compare_cli_reports_combined_gate_and_enforces_only_when_requested(tmp_path):
    before = _snapshot("FULL/DR", "2026-08-15T00:00:00")
    regressed = _snapshot("EXSTART/DR", "2026-08-15T00:05:00")
    clean = _snapshot("FULL/BDR", "2026-08-15T00:05:00")
    before_path = tmp_path / "before.snapshot.json"
    regressed_path = tmp_path / "regressed.snapshot.json"
    clean_path = tmp_path / "clean.snapshot.json"
    before_path.write_text(json.dumps(before), encoding="utf-8")
    regressed_path.write_text(json.dumps(regressed), encoding="utf-8")
    clean_path.write_text(json.dumps(clean), encoding="utf-8")

    def run(after_path, name, enforce=False, before=before_path, extra_args=(),
            decision_context=True):
        out = tmp_path / f"{name}.xlsx"
        cmd = [sys.executable, SCRIPT, "--compare", str(before), str(after_path),
               "--output", str(out)]
        if decision_context:
            cmd.extend([
                "--comparison-context", str(_comparison_context(tmp_path, name)),
            ])
        cmd.extend(extra_args)
        if enforce:
            cmd.append("--fail-on-compare-gate")
        proc = subprocess.run(cmd, cwd=str(tmp_path), capture_output=True, text=True, timeout=300)
        return proc, out, tmp_path / f"{name}.precert.json"

    default_proc, default_out, default_cert_path = run(regressed_path, "default-regression")
    default_terminal = default_proc.stdout + default_proc.stderr
    assert default_proc.returncode == 0, default_terminal
    assert default_out.exists() and default_cert_path.exists()
    comparison_path = tmp_path / "default-regression.comparison.json"
    assert comparison_path.exists()
    comparison = json.loads(comparison_path.read_text(encoding="utf-8"))
    assert comparison["cutover_gate"]["verdict"] == "REGRESSED"
    assert comparison["comparison_admission"]["status"] == "admitted"
    assert json.loads(default_cert_path.read_text(encoding="utf-8"))["verdict"] == "PASS"
    assert _workbook_gate(str(default_out)) == "REGRESSED"
    assert "[CUTOVER GATE: REGRESSED]" in default_terminal
    scope = ("[CERTIFICATE SCOPE] Pre-Change Certificate: PASS. It covers bounded FIB/path-intent "
             "checks only and cannot override the overall cutover gate.")
    assert scope in default_terminal
    assert default_terminal.index("[CUTOVER GATE: REGRESSED]") < default_terminal.index(scope)

    enforced_proc, enforced_out, enforced_cert_path = run(
        regressed_path, "enforced-regression", enforce=True)
    enforced_terminal = enforced_proc.stdout + enforced_proc.stderr
    assert enforced_proc.returncode == 2, enforced_terminal
    assert enforced_out.exists() and enforced_cert_path.exists(), "enforcement happens after artifact writes"
    assert _workbook_gate(str(enforced_out)) == "REGRESSED"
    assert "[COMPARE GATE ENFORCED] Artifacts were written" in enforced_terminal

    clean_proc, clean_out, clean_cert_path = run(
        clean_path, "enforced-missing-family-coverage", enforce=True)
    clean_terminal = clean_proc.stdout + clean_proc.stderr
    assert clean_proc.returncode == 0, clean_terminal
    assert clean_out.exists() and clean_cert_path.exists()
    assert _workbook_gate(str(clean_out)) == "PASS"
    assert "[CUTOVER GATE: PASS]" in clean_terminal
    assert "missing or invalid evidence" not in clean_terminal

    legacy_before = _snapshot("FULL/DR", "2026-08-15T00:00:00")
    legacy_after = _snapshot("FULL/DR", "2026-08-15T00:05:00")
    legacy_before.pop("protocol_assessability")
    legacy_after.pop("protocol_assessability")
    legacy_before_path = tmp_path / "legacy-before.snapshot.json"
    legacy_after_path = tmp_path / "legacy-after.snapshot.json"
    legacy_before_path.write_text(json.dumps(legacy_before), encoding="utf-8")
    legacy_after_path.write_text(json.dumps(legacy_after), encoding="utf-8")
    legacy_proc, legacy_out, legacy_cert_path = run(
        legacy_after_path, "enforced-legacy", enforce=True, before=legacy_before_path)
    legacy_terminal = legacy_proc.stdout + legacy_proc.stderr
    assert legacy_proc.returncode == 2, legacy_terminal
    assert legacy_out.exists() and legacy_cert_path.exists()
    assert _workbook_gate(str(legacy_out)) == "INDETERMINATE"
    assert "[CUTOVER GATE: INDETERMINATE]" in legacy_terminal

    mismatched = _snapshot("EXSTART/DR", "2026-08-15T00:05:00")
    mismatched["script_version"] = "V9.9.9"
    mismatched_path = tmp_path / "mismatched.snapshot.json"
    mismatched_path.write_text(json.dumps(mismatched), encoding="utf-8")
    mismatch_proc, mismatch_out, mismatch_cert_path = run(
        mismatched_path, "enforced-mismatch", enforce=True,
        extra_args=("--allow-schema-mismatch",))
    mismatch_terminal = mismatch_proc.stdout + mismatch_proc.stderr
    assert mismatch_proc.returncode == 2, mismatch_terminal
    assert mismatch_out.exists() and mismatch_cert_path.exists()
    assert _workbook_gate(str(mismatch_out)) == "INDETERMINATE"
    assert "[CUTOVER GATE: INDETERMINATE]" in mismatch_terminal

    unbound_proc, unbound_out, _ = run(
        clean_path, "unbound-clean", decision_context=False)
    unbound_terminal = unbound_proc.stdout + unbound_proc.stderr
    assert unbound_proc.returncode == 0, unbound_terminal
    assert _workbook_gate(str(unbound_out)) == "INDETERMINATE"
    unbound_comparison = json.loads(
        (tmp_path / "unbound-clean.comparison.json").read_text(encoding="utf-8")
    )
    assert unbound_comparison["comparison_admission"]["status"] == "not_comparable"
    assert unbound_comparison["cutover_gate"]["verdict"] == "INDETERMINATE"

    health_regressed = _snapshot("FULL/DR", "2026-08-15T00:05:00")
    health_regressed["health_scores"][0].update({"band": "Poor", "score": 45})
    health_regressed_path = tmp_path / "health-regressed.snapshot.json"
    health_regressed_path.write_text(json.dumps(health_regressed), encoding="utf-8")
    health_proc, health_out, health_cert_path = run(
        health_regressed_path, "health-regression")
    health_terminal = health_proc.stdout + health_proc.stderr
    assert health_proc.returncode == 0, health_terminal
    assert health_out.exists() and health_cert_path.exists()
    assert _workbook_gate(str(health_out)) == "REGRESSED"
    assert "[CUTOVER GATE: REGRESSED]" in health_terminal
    assert "Delta basis:" in health_terminal
    assert "1 switch(es) dropped a health band" in health_terminal


def test_fail_on_compare_gate_requires_compare(tmp_path):
    proc = subprocess.run(
        [sys.executable, SCRIPT, "--fail-on-compare-gate"],
        cwd=str(tmp_path), capture_output=True, text=True, timeout=60)
    assert proc.returncode == 2
    assert "--fail-on-compare-gate requires --compare OLD_SNAPSHOT NEW_SNAPSHOT" in (
        proc.stdout + proc.stderr)
