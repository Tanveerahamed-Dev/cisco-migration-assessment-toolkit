"""Typed STP consistency blockers remain visible on every bounded operator surface."""

import json
from pathlib import Path
import re
import shutil
import subprocess

from openpyxl import Workbook
import pytest

from cisco_toolkit.excel import write_validation_plan_sheet


ROOT = Path(__file__).resolve().parents[1]
EXPLORER = ROOT / "cisco_toolkit" / "blast_radius_explorer.html"
NODE = shutil.which("node")


def _stp_item(device: str, evidence_state: str) -> dict:
    if evidence_state == "degraded":
        check = "STP inconsistent-port baseline"
        acceptance = (
            "PRE-CUTOVER DEGRADED — BLOCKER: current-run STP evidence reports one inconsistent "
            "port. Matching this degraded state after cutover is NOT ACCEPTANCE. "
            "Blocked-port evidence: observed separately."
        )
    elif evidence_state == "review":
        check = "STP consistency evidence incomplete"
        acceptance = (
            "PRE-CUTOVER REVIEW — BLOCKER: required current-run STP consistency evidence is "
            "incomplete (inconsistent_ports missing). Re-collect before acceptance. "
            "Blocked-port evidence: not collected; no blocked-port count is claimed."
        )
    else:
        check = "STP consistency baseline not verified"
        acceptance = (
            "STP CONSISTENCY BASELINE NOT VERIFIED — BLOCKER: current-run evidence is malformed "
            "or contradictory. Re-collect before acceptance. Blocked-port evidence: not collected; "
            "no blocked-port count is claimed."
        )
    return {
        "device": device,
        "platform": "ios",
        "wave": "Group 1",
        "category": "STP",
        "severity": "High",
        "check": check,
        "command": "show spanning-tree inconsistentports",
        "expect": acceptance,
        "why": "An inconsistent or unverified STP consistency baseline blocks cutover acceptance.",
        "evidence_state": evidence_state,
        "projection_custody": "embedded_unverified",
        "source_key": (
            f"protocol_health[{device},STP] + "
            f"protocol_assessability.rows[{device},STP]"
        ),
    }


def _ordinary_item(index: int) -> dict:
    return {
        "device": "acc1",
        "platform": "ios",
        "wave": "Group 1",
        "category": "Gateway",
        "severity": "Medium",
        "check": f"ordinary check {index:02d}",
        "command": f"show interface Gi1/0/{index + 1}",
        "expect": "observed interface state is unchanged",
        "why": "Ordinary bounded validation detail.",
    }


def _checks() -> list[dict]:
    # Source-order blockers after the ordinary allowance so retention cannot pass accidentally.
    return [_ordinary_item(index) for index in range(45)] + [
        _stp_item("dist1", "degraded"),
        _stp_item("dist2", "review"),
        _stp_item("dist3", "not_verified"),
    ]


def _plan() -> dict:
    checks = _checks()
    return {
        "items": checks,
        "by_wave": {"Group 1": checks},
        "summary": {
            "n_items": len(checks),
            "n_waves": 1,
            "n_high": 3,
            "by_category": {"Gateway": 45, "STP": 3},
        },
        "banner": "Run every cutover validation row and disposition every blocker.",
    }


def _snapshot() -> dict:
    hosts = ["dist1", "dist2", "dist3"]
    return {
        "script_version": "test",
        "devices": {host: {"platform": "ios"} for host in hosts},
        "move_groups": [{"switches": hosts, "endpoints": 8}],
        # Deliberately READY: the typed current-baseline blocker must override this scheduling signal.
        "migration_readiness": [{
            "group": "Group 1",
            "switches": hosts,
            "endpoints": 8,
            "readiness": "READY",
            "n_fail": 0,
            "n_warn": 0,
            "checks": [],
        }],
        "wave_sequencing": [{
            "group": "Group 1",
            "make_before_break": hosts,
            "hard_cutover": [],
            "sequence": "make-before-break",
        }],
        "validation_plan": _plan(),
    }


def _tables_with_header(document, headers):
    expected = list(headers)
    return [
        table for table in document.tables
        if table.rows and [cell.text for cell in table.rows[0].cells] == expected
    ]


def _row_values(document, label, column=1):
    values = []
    for table in document.tables:
        for row in table.rows:
            cells = [cell.text.strip() for cell in row.cells]
            if cells and cells[0] == label and len(cells) > column:
                values.append(cells[column])
    return values


def test_stp_blockers_project_to_workbook_and_survive_the_runbook_cap(tmp_path):
    plan = _plan()
    workbook = Workbook()
    write_validation_plan_sheet(workbook, plan)
    validation = workbook["Cutover Validation"]
    rendered = {
        validation.cell(row, 7).value: [validation.cell(row, column).value for column in range(9, 13)]
        for row in range(5, validation.max_row + 1)
    }
    for item in plan["items"][-3:]:
        row = rendered[item["check"]]
        assert row[0] == item["expect"]
        assert row[1] == item["evidence_state"]
        assert "embedded projection; integrity unverified" in row[2]
        assert f"protocol_assessability.rows[{item['device']},STP]" in row[3]
        assert "0 blocked" not in row[0]

    docx = pytest.importorskip("docx")
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "stp-consistency-runbook.docx"
    write_runbook_docx(str(output), _snapshot(), "STP Consistency Surface Fixture")
    document = docx.Document(str(output))
    tables = _tables_with_header(
        document, ["Device", "Category", "Check", "Command", "Observed baseline / acceptance"],
    )
    assert len(tables) == 1
    rows = [[cell.text for cell in row.cells] for row in tables[0].rows[1:]]
    assert len(rows) == 17  # 14 ordinary rows plus every typed STP evidence blocker.
    by_check = {row[2]: row for row in rows}
    for item in plan["items"][-3:]:
        assert item["check"] in by_check
        evidence = by_check[item["check"]][4]
        assert f"Evidence state: {item['evidence_state']}" in evidence
        assert "Projection custody: embedded_unverified" in evidence
        assert f"protocol_assessability.rows[{item['device']},STP]" in evidence
        assert "0 blocked" not in evidence
    assert "no blocked-port count is claimed" in by_check["STP consistency evidence incomplete"][4]


def test_stp_blockers_override_ready_and_survive_mop_validation_and_capture_caps(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.mop import write_mop_docx

    output = tmp_path / "stp-consistency-mop.docx"
    write_mop_docx(str(output), _snapshot(), "STP Consistency Surface Fixture")
    document = docx.Document(str(output))

    gate, = _row_values(document, "Current baseline gate (all validation groups)")
    assert gate.startswith("HOLD — current baseline BLOCKED")
    assert "overall readiness gate is READY" not in "\n".join(
        cell.text for table in document.tables for row in table.rows for cell in row.cells
    )

    validation = _tables_with_header(
        document,
        ["Sev", "Category", "Device", "Check", "Command", "Observed baseline / acceptance"],
    )
    assert len(validation) == 1
    rows = [[cell.text for cell in row.cells] for row in validation[0].rows[1:]]
    assert len(rows) == 43  # 40 ordinary rows plus all three STP blockers.
    by_check = {row[3]: row for row in rows}
    assert {item["check"] for item in _plan()["items"][-3:]} <= set(by_check)
    assert "Evidence state: degraded" in by_check["STP inconsistent-port baseline"][5]
    assert "no blocked-port count is claimed" in by_check["STP consistency baseline not verified"][5]
    assert "0 blocked" not in by_check["STP consistency baseline not verified"][5]

    captures = _tables_with_header(document, ["Device", "Command to capture"])
    assert len(captures) == 1
    pairs = {(row.cells[0].text, row.cells[1].text) for row in captures[0].rows[1:]}
    assert len(pairs) == 23  # 20 ordinary commands plus every blocker-specific device/command pair.
    assert {
        (host, "show spanning-tree inconsistentports") for host in ("dist1", "dist2", "dist3")
    } <= pairs


def _reasoning_core(html: str) -> str:
    match = re.search(r"REASONING-CORE-PORT START.*?REASONING-CORE-PORT END", html, re.S)
    assert match, "Explorer is missing its executable reasoning-core markers"
    block = match.group(0)
    return block[block.index("*/") + 2:block.rindex("/*")]


@pytest.mark.skipif(not NODE, reason="node is not installed — Explorer STP projection skipped")
def test_explorer_holds_and_retains_every_stp_blocker_outside_three_row_cap(tmp_path):
    html = EXPLORER.read_text(encoding="utf-8")
    workflow = html[
        html.index("function validationFor(wave)"):
        html.index("function drawWaves()")
    ]
    payload = tmp_path / "stp-validation-plan.json"
    payload.write_text(json.dumps(_plan()), encoding="utf-8")
    driver = tmp_path / "stp-validation-driver.js"
    driver.write_text(
        _reasoning_core(html)
        + "\nfunction esc(v){return String(v??'');} function shortName(v){return String(v??'');}"
          "const fs=require('fs');"
          "const SNAP={validation_plan:JSON.parse(fs.readFileSync(process.argv[2],'utf8'))};\n"
        + workflow
        + "\nconst waveData=[{idx:0,g:{switches:['dist1','dist2','dist3']}}];"
          "const state=currentBaselineWorkflowState(waveData);"
          "process.stdout.write(JSON.stringify({"
          "verdict:state.gate.verdict,panel:currentBaselineWorkflowPanel(state),"
          "validation:wavesValidationBlock('Group 1')}));\n",
        encoding="utf-8",
    )
    process = subprocess.run(
        [NODE, str(driver), str(payload)],
        capture_output=True,
        text=True,
        encoding="utf-8",
        timeout=60,
    )
    assert process.returncode == 0, process.stderr[:3000]
    actual = json.loads(process.stdout)
    assert actual["verdict"] == "BLOCKED"
    assert "HOLD — current baseline BLOCKED" in actual["panel"]
    rendered = actual["validation"]
    for item in _plan()["items"][-3:]:
        assert item["check"] in rendered
        assert f"<b>evidence</b> {item['evidence_state']}" in rendered
        assert f"protocol_assessability.rows[{item['device']},STP]" in rendered
    assert "no blocked-port count is claimed" in rendered
    assert "0 blocked" not in rendered
    assert "ordinary check 00" in rendered and "ordinary check 02" in rendered
    assert "ordinary check 03" not in rendered
