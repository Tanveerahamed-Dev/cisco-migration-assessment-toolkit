"""In-process end-to-end test: drive the real pipeline through ``main()`` (no subprocess) and assert
the three deliverables are produced and well-formed.

This complements ``test_pipeline_golden.py`` (which runs the pipeline in a SUBPROCESS to freeze the
snapshot/Excel-schema byte-for-byte). Running in-process instead buys two things the subprocess can't:

  * **Coverage** — the workbook writers (``excel.py``), the orchestration (``build.py``), and the
    HTML-embed path (``html.py``) are exercised IN THIS PROCESS, so ``pytest --cov`` actually credits
    them. Under the subprocess golden they run in a child the coverage tool never sees (which made
    ``excel.py`` read as ~18% covered when it is in fact exercised end to end).
  * **Debuggability** — a workbook-build regression surfaces here as a real Python traceback at the
    failing writer, instead of a stdout/stderr diff from a dead child process.

It deliberately asserts only STRUCTURAL properties (the files exist, the workbook opens and carries its
lead sheets, the snapshot carries its computed keys, the explorer embeds the snapshot) — never the
byte-exact golden, which stays the subprocess test's job, so the two don't duplicate each other.
"""
import json
import os
import sys

from openpyxl import Workbook, load_workbook

import synthetic_fixtures as fx

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

import COLLECT_PARSE_V3_23_0 as cp   # noqa: E402  (the entry module; main() is the console entry point)


def _make_template(path):
    """Minimal template workbook — the loader only needs a header row with hostname/port/status."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Interface Data"
    ws.append(["Hostname", "Port", "Status"])
    wb.save(path)


def test_empty_multichassis_producer_is_not_published_as_runtime_support():
    snapshot = {
        "multichassis_lag_typed_observations": {"stale": True},
        "multichassis_lag_domain_baseline": {"stale": True},
    }

    cp._publish_multichassis_lag_blocks(
        snapshot,
        {"observations": []},
        {"schema": "multichassis_lag_domain_baseline/1"},
    )

    assert "multichassis_lag_typed_observations" not in snapshot
    assert "multichassis_lag_domain_baseline" not in snapshot


def test_pipeline_inprocess_builds_all_three_deliverables(tmp_path, monkeypatch):
    collection = fx.write_collection(str(tmp_path / "collection"))
    devices = tmp_path / "devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "out.xlsx"
    traffic_intents = tmp_path / "traffic-intents.json"
    traffic_intents.write_text(json.dumps({"intents": [{
        "id": "users-to-voice-api",
        "src": "10.0.10.50", "dst": "10.0.20.50",
        "protocol": "tcp", "src_port": 49152, "dst_port": 443,
        "expected": "permit", "return_required": False,
    }]}), encoding="utf-8")

    # The producer is invoked once. Excel/JSON/HTML surfaces must project the returned canonical set,
    # never recompute path, policy, ECMP, MTU or failure verdicts independently.
    assurance_calls = []
    _assess_flows = cp.assess_flows

    def _counted_assess_flows(snapshot, intents):
        assurance_calls.append((snapshot, intents))
        return _assess_flows(snapshot, intents)

    monkeypatch.setattr(cp, "assess_flows", _counted_assess_flows)

    # Run from a clean working directory (main() writes its log file into cwd) with argv set as if
    # invoked from the command line. HTML is intentionally LEFT ON so write_html_explorer is exercised.
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess",
        "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--traffic-intents", str(traffic_intents),
    ])

    cp.main()   # the actual console entry point — exercises build/excel/html/runbook in-process

    # Capture-integrity custody is computed exactly once before the protocol denominator and every
    # readiness consumer; its workbook sheet may remain later without recomputing the evidence.
    timings = json.loads((tmp_path / "out.phase_timings.json").read_text(encoding="utf-8"))
    phase_names = [row["phase"] for row in timings["phases"]]
    assert phase_names.count("Capture integrity") == 1
    assert phase_names.index("Capture integrity") < phase_names.index("Protocol Health")
    assert phase_names.index("Protocol assessability") < phase_names.index(
        "BGP configured-peer baseline") < phase_names.index("Migration Readiness")
    assert phase_names.index("Protocol assessability") < phase_names.index(
        "VTP safety subject scope") < phase_names.index(
        "VTP safety baseline") < phase_names.index(
        "VTP extended evidence") < phase_names.index("Migration Readiness")
    assert phase_names.index("Protocol assessability") < phase_names.index(
        "IPv6 routing subject scope") < phase_names.index(
        "IPv6 routing adjacency baseline") < phase_names.index("Migration Readiness")
    assert phase_names.index("Protocol assessability") < phase_names.index(
        "FHRP configured-group baseline") < phase_names.index("Migration Readiness")
    assert phase_names.index("FHRP configured-group baseline") < phase_names.index(
        "FHRP redundancy-domain baseline") < phase_names.index("Migration Readiness")

    # ---- workbook ----
    assert out_xlsx.is_file(), "workbook was not written"
    wb = load_workbook(str(out_xlsx), read_only=True)
    sheets = wb.sheetnames
    wb.close()
    assert "Executive Summary" in sheets, f"Executive Summary sheet missing; got {sheets[:5]}…"
    assert "Traffic Assurance" in sheets, "canonical traffic assurance projection sheet missing"
    assert "VTP Safety" in sheets, "bounded local VTP safety receipt sheet missing"
    assert "IPv6 Routing" in sheets, "bounded observed OSPFv3/BGPv6 receipt sheet missing"
    assert "BGP Peer Intent" in sheets, "configured default/global IPv4 BGP denominator sheet missing"
    assert "FHRP Group Intent" in sheets, "configured default/global IPv4 FHRP denominator sheet missing"
    assert "FHRP Domains" in sheets, "authoritative cross-switch FHRP domain sheet missing"
    assert len(sheets) >= 20, f"expected the full multi-sheet workbook, got only {len(sheets)} sheets"

    # ---- snapshot (the data contract) ----
    snap_path = os.path.splitext(str(out_xlsx))[0] + ".snapshot.json"
    assert os.path.isfile(snap_path), "snapshot.json was not written"
    snap = json.loads(open(snap_path, encoding="utf-8").read())
    for key in ("devices", "interfaces", "health_scores", "punchlist", "causality", "executive_brief",
                "parse_yield", "unknown_evidence", "protocol_assessability",
                "bgp_configured_peer_baseline", "fhrp_configured_group_baseline",
                "fhrp_redundancy_domain_baseline", "vtp_safety_baseline",
                "vtp_extended_evidence",
                "ipv6_routing_adjacency_baseline"):
        # parser detail, governed aggregate, and the protocol coverage denominator ship together
        assert key in snap, f"snapshot missing computed key {key!r}"
    protocol_receipt = snap["protocol_assessability"]
    assert protocol_receipt["schema"] == "protocol_assessability/1"
    assert protocol_receipt["summary"]["n_families"] == 7
    assert protocol_receipt["summary"]["n_cells"] == 7 * protocol_receipt["summary"]["n_devices"]
    assert len(protocol_receipt["rows"]) == protocol_receipt["summary"]["n_cells"]
    assert {row["protocol"] for row in protocol_receipt["rows"]} == {
        "STP", "EtherChannel", "VTP", "OSPF", "BGP", "EIGRP", "FHRP"
    }
    # The canonical fixture does not collect show vtp status, so the strict owner is neutral rather
    # than inventing a fleet-wide blocker.  Publication still carries a validator-valid embedded
    # receipt and the direct workbook section, while current-run acceptance rows remain empty.
    from cisco_toolkit.vtp_safety import validate_vtp_safety_baseline
    vtp_baseline = snap["vtp_safety_baseline"]
    assert vtp_baseline["schema"] == "vtp_safety_baseline/1"
    assert vtp_baseline["projection_custody"] == "embedded_unverified"
    assert vtp_baseline["verdict"] == "NOT_APPLICABLE"
    assert vtp_baseline["rows"] == []
    assert validate_vtp_safety_baseline(vtp_baseline)["valid"] is True
    assert validate_vtp_safety_baseline(
        vtp_baseline, require_current_run=True)["valid"] is False
    from cisco_toolkit.vtp_extended import validate_vtp_extended_evidence
    vtp_extended = snap["vtp_extended_evidence"]
    assert vtp_extended["schema"] == "vtp_extended_evidence/1"
    assert vtp_extended["projection_custody"] == "embedded_unverified"
    assert validate_vtp_extended_evidence(vtp_extended)["valid"] is True
    assert validate_vtp_extended_evidence(
        vtp_extended, require_current_run=True)["valid"] is False
    assert vtp_extended["summary"]["n_not_verified"] == len(snap["devices"])
    assert all(row["status"] == "not_verified" for row in vtp_extended["rows"])
    assert "vtp password " not in json.dumps(vtp_extended).casefold()
    assert [row for row in snap["validation_plan"]["items"]
            if row.get("category") == "VTP"] == []
    assert snap["nrfu_commands"]["summary"]["n_vtp_safety_cases"] == 0
    assert snap["nrfu_commands"]["summary"]["n_vtp_safety_blockers"] == 0
    # The same in-process run contains five observed IPv6 routing adjacencies.  Two are
    # definitely degraded (OSPFv3 EXSTART and BGPv6 Active), so the source-bound owner must
    # block readiness and project every row verbatim into Validation and NRFU.  Only the
    # serialized snapshot copy loses operational authority.
    from cisco_toolkit.ipv6_routing import validate_ipv6_routing_adjacency_baseline
    ipv6_baseline = snap["ipv6_routing_adjacency_baseline"]
    assert ipv6_baseline["schema"] == "ipv6_routing_adjacency_baseline/1"
    assert ipv6_baseline["projection_custody"] == "embedded_unverified"
    assert ipv6_baseline["verdict"] == "BLOCKED"
    assert len(ipv6_baseline["rows"]) == 5
    assert len(ipv6_baseline["coverage"]) == 9
    assert ipv6_baseline["summary"]["by_status"] == {
        "degraded": 2, "review": 0, "not_verified": 0, "assessed": 3,
    }
    assert validate_ipv6_routing_adjacency_baseline(ipv6_baseline)["valid"] is True
    assert validate_ipv6_routing_adjacency_baseline(
        ipv6_baseline, require_current_run=True)["valid"] is False

    ipv6_validation = [
        row for row in snap["validation_plan"]["items"]
        if row.get("category") == "IPv6 Routing"
    ]
    ipv6_nrfu = [
        case
        for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"]
        for case in device["cases"]
        if case.get("evidence_family") == "IPv6 Routing"
    ]
    owner_projection = sorted(
        (
            row["command"], row["acceptance"], row["source_key"], row["status"],
            "current_run_source_bound",
        )
        for row in ipv6_baseline["rows"]
    )
    assert sorted(
        (
            row["command"], row["expect"], row["source_key"], row["evidence_state"],
            row["projection_custody"],
        )
        for row in ipv6_validation
    ) == owner_projection
    assert sorted(
        (
            case["command"], case["expected"], case["source_key"],
            case["evidence_state"], case["projection_custody"],
        )
        for case in ipv6_nrfu
    ) == owner_projection
    assert snap["nrfu_commands"]["summary"]["n_ipv6_routing_cases"] == 5
    assert snap["nrfu_commands"]["summary"]["n_ipv6_routing_blockers"] == 2
    assert len([row for row in snap["punchlist"]
                if row.get("category") == "IPv6 Routing"]) == 2
    ipv6_readiness = [
        (group["readiness"], check["status"])
        for group in snap["migration_readiness"]
        for check in group["checks"]
        if check["check"] == "IPv6 routing adjacencies"
    ]
    assert ipv6_readiness == [("NOT READY", "fail")]
    # STP consistency has one shared, claim-specific owner across readiness, Validation and NRFU.
    # NRFU receives the published health/receipt plus positive L2/root subject evidence; it must
    # project the owner's acceptance row exactly and never recover a zero from health summary prose.
    from cisco_toolkit.analyze import summarize_stp_consistency_baseline
    stp_consistency = summarize_stp_consistency_baseline(
        snap["protocol_health"], protocol_receipt,
        all_interfaces=snap["interfaces"], stp_roots=snap["stp_roots"],
    )
    stp_owner = {row["source_key"]: row for row in stp_consistency["rows"]}
    stp_validation = {
        row["source_key"]: row
        for row in snap["validation_plan"]["items"]
        if row.get("category") == "STP" and row.get("source_key") in stp_owner
    }
    stp_nrfu = {
        case["source_key"]: case
        for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"]
        for case in device["cases"]
        if case.get("evidence_family") == "STP"
    }
    assert stp_owner and set(stp_validation) == set(stp_nrfu) == set(stp_owner)
    for source_key, owner_row in stp_owner.items():
        validation_row = stp_validation[source_key]
        nrfu_case = stp_nrfu[source_key]
        assert validation_row["evidence_state"] == nrfu_case["evidence_state"] == (
            owner_row["status"]
        )
        assert validation_row["expect"] == nrfu_case["expected"] == owner_row["acceptance"]
        assert validation_row["command"] == nrfu_case["command"] == owner_row["command"]
        assert validation_row["projection_custody"] == nrfu_case["projection_custody"] == (
            owner_row["projection_custody"]
        )
    stp_summary = snap["nrfu_commands"]["summary"]
    assert stp_summary["n_stp_consistency_cases"] == len(stp_owner)
    assert stp_summary["n_stp_consistency_blockers"] == sum(
        row["status"] != "assessed" for row in stp_owner.values()
    )
    bgp_baseline = snap["bgp_configured_peer_baseline"]
    assert bgp_baseline["schema"] == "bgp_configured_peer_baseline/1"
    assert bgp_baseline["projection_custody"] == "embedded_unverified"
    assert all(row["projection_custody"] == "embedded_unverified"
               for row in bgp_baseline["rows"])
    fhrp_baseline = snap["fhrp_configured_group_baseline"]
    assert fhrp_baseline["schema"] == "fhrp_configured_group_baseline/1"
    assert fhrp_baseline["projection_custody"] == "embedded_unverified"
    assert fhrp_baseline["rows"]
    assert all(row["projection_custody"] == "embedded_unverified"
               for row in fhrp_baseline["rows"])
    fhrp_domain_baseline = snap["fhrp_redundancy_domain_baseline"]
    assert fhrp_domain_baseline["schema"] == "fhrp_redundancy_domain_baseline/1"
    assert fhrp_domain_baseline["projection_custody"] == "embedded_unverified"
    assert fhrp_domain_baseline["rows"]
    assert all(row["projection_custody"] == "embedded_unverified"
               for row in fhrp_domain_baseline["rows"])
    # The persisted domain receipt must bind the exact co-published embedded configured owner,
    # not the current-run digest that existed before both owners crossed the JSON boundary.
    assert fhrp_domain_baseline["source_receipt"]["configured_baseline_sha256"] == (
        fhrp_baseline["summary"]["baseline_sha256"]
    )
    from cisco_toolkit.protocol_assurance import (
        bind_snapshot_json_bytes,
        bound_snapshot_source,
    )
    from cisco_toolkit.protocol_deltas import compute_fhrp_redundancy_domain_delta
    persisted_bytes = open(snap_path, "rb").read()
    persisted = bind_snapshot_json_bytes(persisted_bytes)
    persisted_binding = bound_snapshot_source(persisted)
    persisted_domain_delta = compute_fhrp_redundancy_domain_delta(
        persisted,
        persisted,
        comparison_source_binding={
            "before": persisted_binding,
            "after": persisted_binding,
        },
    )
    assert persisted_domain_delta["summary"]["by_transition"]["not_comparable"] == 0
    assert persisted_domain_delta["changes"]
    fhrp_by_identity = {
        (row["switch"], row["protocol"], row["interface"], row["group"]): row
        for row in fhrp_baseline["rows"]
    }
    fhrp_validation = {
        (row["device"], row["protocol"], row["interface"], row["group"]): row
        for row in snap["validation_plan"]["items"]
        if row.get("category") == "FHRP" and row.get("group_key")
    }
    fhrp_nrfu = {
        (device["host"], case["protocol"], case["interface"], case["group"]): case
        for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"] for case in device["cases"]
        if case.get("evidence_family") == "FHRP" and case.get("group_key")
    }
    assert set(fhrp_validation) == set(fhrp_nrfu) == set(fhrp_by_identity)
    for identity, baseline_row in fhrp_by_identity.items():
        validation_row = fhrp_validation[identity]
        nrfu_case = fhrp_nrfu[identity]
        assert validation_row["expect"] == nrfu_case["expected"] == baseline_row["acceptance"]
        assert validation_row["source_key"] == nrfu_case["source_key"] == baseline_row["source_key"]
        assert validation_row["projection_custody"] == nrfu_case["projection_custody"] == (
            "current_run_source_bound"
        )

    # EtherChannel has a producer-owned, lossless current-summary projection and a separate
    # receipt-gated cutover baseline.  The generic interface model only carries associations;
    # it must never be used to recreate operational member flags or an all-(P) claim.
    etherchannel_projection = snap["etherchannel_projection"]
    etherchannel_baseline = snap["etherchannel_baseline"]
    assert etherchannel_projection["schema"] == "etherchannel_projection/1"
    assert etherchannel_projection["summary"] == {
        "n_devices": 3,
        "n_subject_devices": 2,
        "n_groups": 2,
        "n_members": 4,
        "n_associations": 4,
        "n_degraded_groups": 0,
        "n_review_groups": 0,
        "n_rejected_lines": 0,
        "by_capture_state": {"usable": 2, "empty": 0, "error": 0, "missing": 1},
    }
    projection_by_host = {row["switch"]: row for row in etherchannel_projection["rows"]}
    expected_members = {
        "core1": [("Gi1/0/1", "P", "assessed", "forwarding_observed"),
                  ("Gi1/0/2", "P", "assessed", "forwarding_observed")],
        "core2": [("Eth1/1", "P", "assessed", "forwarding_observed"),
                  ("Eth1/2", "P", "assessed", "forwarding_observed")],
    }
    expected_commands = {
        "core1": "show etherchannel summary",
        "core2": "show port-channel summary",
    }
    for host in ("core1", "core2"):
        projection_row = projection_by_host[host]
        assert projection_row["source_command"] == expected_commands[host]
        assert projection_row["capture_state"] == "usable"
        assert projection_row["findings"] == [] and projection_row["rejected_line_count"] == 0
        assert [(association["interface"], association["group"])
                for association in projection_row["associations"]] == [
            (interface, "Po1") for interface, *_ in expected_members[host]
        ]
        assert len(projection_row["groups"]) == 1
        group = projection_row["groups"][0]
        assert (group["group_id"], group["group"], group["group_flags"], group["protocol"],
                group["status"], group["operational_state"]) == (
            "1", "Po1", "SU", "LACP", "assessed", "up"
        )
        assert [(member["interface"], member["flags"], member["status"], member["state"])
                for member in group["members"]] == expected_members[host]
        assert group["findings"] == []

    assert etherchannel_baseline["schema"] == "etherchannel_baseline/1"
    assert etherchannel_baseline["scope"] == "baseline_observed"
    assert etherchannel_baseline["status"] == "assessed"
    assert etherchannel_baseline["assessed"] is True
    assert etherchannel_baseline["projection_custody"] == "embedded_unverified"
    assert etherchannel_baseline["projection"] == {"present": True, "valid": True, "reason": ""}
    assert etherchannel_baseline["receipt"] == {"present": True, "valid": True, "reason": ""}
    baseline_by_host = {row["switch"]: row for row in etherchannel_baseline["rows"]}
    assert set(baseline_by_host) == {"core1", "core2"}
    for host in ("core1", "core2"):
        baseline_row = baseline_by_host[host]
        assert baseline_row["status"] == "assessed"
        assert baseline_row["receipt_state"] == "assessed"
        assert baseline_row["capture_state"] == "usable"
        assert baseline_row["health_row_emitted"] is True
        assert baseline_row["group_count"] == 1 and baseline_row["member_count"] == 2
        assert baseline_row["groups"] == projection_by_host[host]["groups"]
        assert baseline_row["command"] == expected_commands[host]
        assert baseline_row["source_key"] == (
            f"etherchannel_projection.rows[{host}] + "
            f"protocol_assessability.rows[{host},EtherChannel]"
        )
        assert baseline_row["projection_custody"] == "embedded_unverified"

    # Both operator workflows consume the same baseline row verbatim, including its exact source
    # locator and custody disclosure.  Neither is allowed to independently idealize the bundle.
    etherchannel_validation = {
        row["device"]: row for row in snap["validation_plan"]["items"]
        if row.get("category") == "Link"
        and str(row.get("source_key") or "").startswith("etherchannel_projection.rows[")
    }
    etherchannel_nrfu = {
        device["host"]: case
        for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"]
        for case in device["cases"]
        if case.get("evidence_family") == "EtherChannel"
    }
    assert set(etherchannel_validation) == set(etherchannel_nrfu) == {"core1", "core2"}
    for host in ("core1", "core2"):
        baseline_row = baseline_by_host[host]
        validation_row = etherchannel_validation[host]
        nrfu_case = etherchannel_nrfu[host]
        assert validation_row["evidence_state"] == nrfu_case["evidence_state"] == "assessed"
        assert validation_row["expect"] == nrfu_case["expected"] == baseline_row["acceptance"]
        assert validation_row["command"] == nrfu_case["command"] == baseline_row["command"]
        assert validation_row["source_key"] == nrfu_case["source_key"] == baseline_row["source_key"]
        assert validation_row["projection_custody"] == nrfu_case["projection_custody"] == (
            "embedded_unverified"
        )
        assert "all-(P)" not in baseline_row["acceptance"]
    assert snap["nrfu_commands"]["summary"]["n_etherchannel_cases"] == 2
    assert snap["nrfu_commands"]["summary"]["n_etherchannel_blockers"] == 0
    assert snap["nrfu_commands"]["summary"]["etherchannel_by_evidence_state"] == {"assessed": 2}
    assert snap["nrfu_commands"]["summary"]["etherchannel_by_projection_custody"] == {
        "embedded_unverified": 2
    }

    # The synthetic fleet deliberately carries one OSPF peer stuck in EXSTART.  Validation and NRFU must
    # project the same receipt-gated observed baseline; neither may rewrite every peer as FULL.
    ospf_health = [row for row in snap["protocol_health"]
                   if row.get("switch") == "core1" and row.get("protocol") == "OSPF"]
    assert ospf_health and ospf_health[0]["severity"] == "High"
    assert "EXSTART" in ospf_health[0]["detail"]
    routing_validation = [row for row in snap["validation_plan"]["items"]
                          if row.get("device") == "core1" and row.get("category") == "Routing"]
    assert len(routing_validation) == 1
    assert routing_validation[0]["evidence_state"] == "degraded"
    assert routing_validation[0]["expect"].startswith("PRE-CUTOVER DEGRADED — BLOCKER:")
    assert "EXSTART" in routing_validation[0]["expect"] and "2 neighbor(s) in FULL" not in routing_validation[0]["expect"]
    nrfu_routing = [case for wave in snap["nrfu_commands"]["waves"]
                    for device in wave["devices"] if device.get("host") == "core1"
                    for case in device["cases"] if case.get("command") == "show ip ospf neighbor"]
    assert len(nrfu_routing) == 1
    assert nrfu_routing[0]["evidence_state"] == "degraded"
    assert nrfu_routing[0]["expected"] == routing_validation[0]["expect"]
    assert nrfu_routing[0]["projection_custody"] == "embedded_unverified"
    _pa_wb = load_workbook(str(out_xlsx), read_only=True)
    _pa_text = "\n".join(
        str(cell.value or "")
        for row in _pa_wb["Collection Completeness"].iter_rows()
        for cell in row
    )
    _pa_wb.close()
    assert "Protocol assessability — runtime family × device receipt" in _pa_text
    assert f"of {protocol_receipt['summary']['n_cells']} host-family cells" in _pa_text
    assurance = snap.get("traffic_assurance")
    assert assurance and assurance["schema"] == "traffic_assurance_set/1"
    assert assurance["owner"] == "cisco_toolkit.traffic_assurance.assess_flow"
    assert assurance["summary"]["n"] == 1 and len(assurance["results"]) == 1
    assert assurance["results"][0]["intent"]["id"] == "users-to-voice-api"
    assert len(assurance_calls) == 1, "traffic assurance must be evaluated exactly once per pipeline run"
    _ta_wb = load_workbook(str(out_xlsx), read_only=True)
    _ta_ws = _ta_wb["Traffic Assurance"]
    _ta_headers = {cell.value: index for index, cell in enumerate(next(_ta_ws.iter_rows()), 1)}
    _ta_row = next(_ta_ws.iter_rows(min_row=2, values_only=True))
    _ta_wb.close()
    _canonical = assurance["results"][0]
    assert _ta_row[_ta_headers["Flow ID"] - 1] == _canonical["intent"]["id"]
    assert _ta_row[_ta_headers["Overall Verdict"] - 1] == _canonical["verdict"]
    assert _ta_row[_ta_headers["Set Schema"] - 1] == assurance["schema"]
    # the ledger must be a REAL run's ledger (parsers were called), with its coverage-honest note
    assert snap["parse_yield"]["summary"]["parsers_called"] > 0
    assert "never a device" in snap["parse_yield"]["summary"]["note"]
    # Provenance (wave R2-1-01 / R2-3-01): the snapshot records WHEN THE EVIDENCE WAS COLLECTED, and the
    # lifecycle EoL bands are pinned to THAT instant (not wall-clock-at-regen) so a re-render reproduces
    # them deterministically. collected_at is threaded into compute_lifecycle_risk as its asof.
    assert "collected_at" in snap, "snapshot missing collection-time provenance (collected_at)"
    # lifecycle bands are date-granular, so asof is the DATE of collection (compute_lifecycle_risk
    # normalises its asof to a date); it must equal the collection date, never the regen day.
    assert snap.get("lifecycle_risk", {}).get("asof") == snap["collected_at"][:10], \
        "lifecycle bands must be dated to the collection date (asof == collected_at[:10]), not the regen day"
    # SSOT: the brief's scale block publishes the canonical VLAN count, equal to the vlan_inventory
    # derivation the deliverables + dashboards read (one source — no JS/TS recount drift).
    from cisco_toolkit.analyze import vlan_inventory
    _scale = snap["executive_brief"]["scale"]
    assert "n_vlans" in _scale, "scale.n_vlans (canonical VLAN count) missing from the brief"
    assert _scale["n_vlans"] == len(vlan_inventory(snap)), "scale.n_vlans must equal canonical vlan_inventory"
    # SSOT: the genuinely-collected subset is published canonically too, so coverage-honesty surfaces read
    # ONE "N collected of M inventoried" instead of mislabelling inventory as assessed. (SSOT-device-deliv-06)
    assert _scale.get("n_collected") == ((snap.get("collection_completeness") or {}).get("summary") or {}).get("complete"), \
        "scale.n_collected must equal collection_completeness.summary.complete"
    # SSOT (unified): every published canonical fact in the REAL assembled snapshot must still
    # reconcile to its raw-evidence derivation (scale/posture/lifecycle bands/decision count). This
    # one invariant subsumes the spot-checks above and extends them to the facts they miss — the
    # gap that let the n_past_eos/n_past_ldos conflation slip surface-by-surface. See ssot.reconcile.
    from cisco_toolkit import ssot
    assert ssot.reconcile(snap) == [], f"engine published SSOT-inconsistent facts: {ssot.reconcile(snap)}"
    # The runtime SSOT self-check and independent data-authority gates are clean:
    # OUI/port packs chain to retained hash-pinned IEEE/IANA inputs, while every
    # represented EoL scope is bound to an exact Cisco bulletin.
    assert ssot.audit(snap) is None, f"healthy run raised a false SSOT integrity alarm: {ssot.audit(snap)}"
    failed_phases = set(
        (snap.get("assessment_integrity") or {}).get("failed_phases") or []
    )
    assert "OUI registry authority" not in failed_phases
    assert "EoL knowledge-base authority" not in failed_phases
    # SCOPED AUTHORITY (handoff 5.2). The combined port pack holds source-authoritative IANA
    # assignments PLUS explicitly non-authoritative curated semantics/multicast scopes, so its
    # whole-pack flag is correctly False. This used to assert the phase FAILED -- pinning the
    # defect as intended behaviour: an intact pack whose IANA source chain verified was reported
    # as a dead registry, which also broke both Atlas self-tests.
    #
    # A pack is usable when it is INTACT and its OFFICIAL component is source-proven. It does not
    # need universal authority over every curated row, and the honest per-component labels below
    # must survive -- the phase passing is not permission to call curated hints official.
    assert "Port registry authority" not in failed_phases, (
        "the port pack is integrity-verified with a proven IANA source chain; failing the phase "
        "on its whole-pack flag treats an honestly-mixed pack as a dead registry"
    )
    assert set(snap["data_authorities"]) == {"oui", "ports", "eol"}
    oui_health = snap["data_authorities"]["oui"]
    assert oui_health["integrity_verified"] is True
    assert oui_health["source_authoritative"] is True
    assert oui_health["authoritative"] is True
    assert oui_health["source_fresh"] is True
    port_health = snap["data_authorities"]["ports"]
    assert port_health["integrity_verified"] is True
    assert port_health["source_authoritative"] is False
    assert port_health["official_source_authoritative"] is True
    assert port_health["curated_overlay_authoritative"] is False
    assert port_health["curated_multicast_authoritative"] is False
    assert port_health["authoritative"] is False
    assert port_health["source_fresh"] is True
    eol_health = snap["data_authorities"]["eol"]
    assert eol_health["schema_verified"] is True
    assert eol_health["integrity_verified"] is True
    assert eol_health["source_authoritative"] is True
    assert eol_health["authoritative"] is True
    assert eol_health["source_fresh"] is True
    assert eol_health["unresolved_reference_rows"] == 0
    # The positive self-verification badge the dashboards render must be published and pass on a
    # healthy run (executive_brief.ssot, from ssot.summary; golden-excluded with executive_brief).
    _ssot_badge = (snap.get("executive_brief") or {}).get("ssot")
    assert isinstance(_ssot_badge, dict) and _ssot_badge.get("verified") is True, \
        f"assembly must publish a passing executive_brief.ssot self-verification badge, got {_ssot_badge}"
    assert _ssot_badge.get("n_facts", 0) >= 5, "the badge must report the published canonical-fact count"
    # SSOT (workbook headline): the Executive Summary 'Scope / scale' block must STATE the canonical VLAN
    # count, never leave it blank. The brief's scale gets n_vlans injected BEFORE this sheet is written (the
    # snapshot's n_vlans is assembled later), so without that injection the workbook would miss 'VLANs in use'.
    _wb = load_workbook(str(out_xlsx), read_only=True)
    _es = _wb["Executive Summary"]
    _vlan_cell = "MISSING"; _coll_cell = "MISSING"; _engine_cell = "MISSING"
    for _row in _es.iter_rows():
        _lbl = str(_row[0].value or "").strip() if _row else ""
        if _lbl == "VLANs in use":
            _vlan_cell = _row[1].value if len(_row) > 1 else None
        elif _lbl == "Switches collected / inventoried":
            _coll_cell = _row[1].value if len(_row) > 1 else None
        elif _lbl == "Engine version":
            _engine_cell = _row[1].value if len(_row) > 1 else None
    _wb.close()
    assert _vlan_cell == _scale["n_vlans"], (
        f"Executive Summary 'VLANs in use' must state the canonical {_scale['n_vlans']}, got {_vlan_cell!r}")
    # P3-E3: the Document-control provenance section is wired end-to-end (call site -> sheet) -- the engine
    # version renders, so the landing sheet is self-traceable to the exact engine that produced it.
    assert isinstance(_engine_cell, str) and _engine_cell.startswith("V"), (
        f"Executive Summary 'Document control' must carry the engine version (P3-E3), got {_engine_cell!r}")
    # SSOT (QA F1): the Fleet-posture block reads canonical scale.n_collected ("<collected> / <inventoried>"),
    # not a local recompute that would call the full inventory "assessed". n_collected is injected into the
    # brief's scale BEFORE this sheet is written (same early-injection point as n_vlans).
    assert _coll_cell == f"{_scale['n_collected']} / {len(snap.get('health_scores') or [])}", (
        f"Exec Summary 'Switches collected / inventoried' must lead with canonical n_collected, got {_coll_cell!r}")

    # SSOT: the canonical CCDE-grounded design blueprint is PUBLISHED into the snapshot (the one source the
    # design DOCX / explorer / webapp all read), and equals a fresh compute over the same snapshot — no
    # surface recomputes design intent.
    from cisco_toolkit.design_advisor import compute_design_blueprint
    assert "design_blueprint" in snap, "snapshot must publish the canonical design_blueprint"
    _bp = snap["design_blueprint"]
    assert isinstance(_bp.get("decisions"), list) and "tradeoff_scorecard" in _bp and "summary" in _bp
    # SSOT: equals a fresh compute over the snapshot WITH the same requirements register (if one was
    # supplied via --requirements, it is stored in the snapshot so the blueprint stays reproducible).
    assert _bp == compute_design_blueprint(snap, snap.get("requirements_register")), \
        "published design_blueprint must equal the canonical recompute (with the same requirements register)"
    # UNIVERSALITY (FHRP): the fixture's core1 has an untracked active HSRP gateway. The engine must PUBLISH
    # per-device FHRP detail and ASSESS it end-to-end -- the blueprint carries the FHRP-resilience decision.
    # Meridian ran zero FHRP, so this is the first architecture coverage proven on a non-Meridian environment.
    assert isinstance(snap.get("fhrp_detail"), dict) and snap["fhrp_detail"].get("core1"), \
        "snapshot must publish per-device FHRP detail (build_fhrp_detail -> parse_hsrp_detail)"
    assert any(d.get("id") == "fhrp-resilience-tracking-and-preempt" for d in _bp.get("decisions", [])), \
        "engine must assess FHRP: an untracked active gateway must fire _d_fhrp_resilience"
    # UNIVERSALITY (VXLAN-EVPN): core2 is a VTEP with a DOWN peer. The engine must PUBLISH snap['overlay']
    # and the blueprint must carry the VXLAN NVE-peer-down decision -- its OWN target fabric, previously blind.
    assert isinstance(snap.get("overlay"), dict) and snap["overlay"].get("core2", {}).get("nve_peers"), \
        "snapshot must publish per-device VXLAN overlay (build_overlay -> parse_nve_peers)"
    assert any(d.get("id") == "vxlan-nve-peer-down" for d in _bp.get("decisions", [])), \
        "engine must assess VXLAN: a down VTEP peer must fire _d_nve_peer_health"
    assert any(d.get("id") == "vxlan-evpn-control-plane-down" for d in _bp.get("decisions", [])), \
        "engine must assess EVPN control plane: an Idle RR session must fire _d_evpn_rr_health"
    assert any(d.get("id") == "vxlan-nve-vni-down" for d in _bp.get("decisions", [])), \
        "engine must assess VXLAN VNI: a not-Up VNI must fire _d_nve_vni_health"
    assert any(d.get("id") == "copp-control-plane-policer-dropping" for d in _bp.get("decisions", [])), \
        "engine must assess CoPP: a dropping control-plane class must fire _d_copp_drops"
    # UNIVERSALITY (architecture-coverage build wave): each new axis is PUBLISHED into the snapshot and its
    # detector fires end-to-end on the synthetic fixtures (broken state only; silent companions prove no
    # over-firing). core1 = PIM-no-RP + unsynchronized NTP + QoS LLQ drops; access1 = dual-stack-no-RA-Guard +
    # port-security Secure-shutdown + toothless storm-control; core2 = an undocumented (shadow) infra router.
    assert isinstance(snap.get("pim"), dict) and snap["pim"].get("core1", {}).get("neighbors"), \
        "snapshot must publish per-device PIM (build_pim -> parse_pim_rp_mapping / parse_pim_neighbors)"
    assert any(d.get("id") == "multicast-pim-rp-resilience" for d in _bp.get("decisions", [])), \
        "engine must assess PIM: a running PIM device with no RP must fire _d_pim_rp_health"
    assert isinstance(snap.get("ipv6_fhs"), dict) and snap["ipv6_fhs"].get("access1", {}).get("dualstack"), \
        "snapshot must publish per-device IPv6 first-hop security (build_ipv6_fhs)"
    assert any(d.get("id") == "ipv6-first-hop-security-suite-at-access-edge" for d in _bp.get("decisions", [])), \
        "engine must assess IPv6 FHS: a dual-stack access switch with no RA-Guard must fire _d_ipv6_fhs"
    assert isinstance(snap.get("ntp"), dict) and snap["ntp"].get("core1", {}).get("synchronized") is False, \
        "snapshot must publish per-device NTP clock-sync STATE (build_ntp -> parse_ntp_status)"
    assert any(d.get("id") == "mgmt-time-sync-logging-baseline" for d in _bp.get("decisions", [])), \
        "engine must assess NTP: an unsynchronized clock must fire _d_ntp_sync"
    assert isinstance(snap.get("port_security"), dict) and snap["port_security"].get("access1"), \
        "snapshot must publish per-device port-security DETAIL (build_port_security_detail)"
    assert any(d.get("id") == "security-l2-access-edge-suite" for d in _bp.get("decisions", [])), \
        "engine must assess port-security: a Secure-shutdown port must fire _d_port_security_errdisable"
    assert isinstance(snap.get("storm_control"), dict) and snap["storm_control"].get("access1"), \
        "snapshot must publish per-device storm-control (build_storm_control)"
    assert any(d.get("id") == "storm-control-action-on-edge" for d in _bp.get("decisions", [])), \
        "engine must assess storm-control: a configured action-None rule must fire _d_storm_control_action"
    assert isinstance(snap.get("qos_runtime"), dict) and snap["qos_runtime"].get("core1"), \
        "snapshot must publish per-device QoS runtime (build_qos_runtime -> parse_policymap_drops)"
    assert any(d.get("id") == "qos-runtime-egress-queue-drops" for d in _bp.get("decisions", [])), \
        "engine must assess QoS runtime: an LLQ class being congestion-dropped must fire _d_qos_runtime_drops"
    assert isinstance(snap.get("shadow_infra"), dict) and snap["shadow_infra"].get("core2"), \
        "snapshot must publish per-device shadow-infra neighbours (build_undocumented_neighbors)"
    assert any(d.get("id") == "discover-undocumented-infrastructure-before-cutover" for d in _bp.get("decisions", [])), \
        "engine must assess shadow infra: an undocumented infra neighbour must fire _d_shadow_infra"
    # UNIVERSALITY (SP/MPLS): core1 acts as an MPLS PE with a Nonexistent LDP session, an Idle VPNv4
    # peer, and a DOWN pseudowire (all on 10.0.255.9 / VC 300).  The three MPLS detectors must fire
    # end-to-end; the healthy peers (Oper LDP / Established VPNv4 / UP VC 200) prove no over-firing.
    assert isinstance(snap.get("mpls"), dict) and snap["mpls"].get("core1", {}).get("ldp_neighbors"), \
        "snapshot must publish per-device MPLS state (build_mpls -> parse_mpls_ldp_neighbors)"
    assert any(d.get("id") == "mpls-ldp-session-down" for d in _bp.get("decisions", [])), \
        "engine must assess MPLS LDP underlay: a Nonexistent LDP session must fire _d_mpls_ldp_health"
    assert any(d.get("id") == "mpls-l3vpn-vpnv4-down" for d in _bp.get("decisions", [])), \
        "engine must assess MPLS L3VPN: an Idle VPNv4 peer must fire _d_mpls_l3vpn_health"
    assert any(d.get("id") == "mpls-l2vpn-pseudowire-down" for d in _bp.get("decisions", [])), \
        "engine must assess MPLS L2VPN: a DOWN pseudowire must fire _d_mpls_l2vpn_health"
    # UNIVERSALITY (Cisco ACI / JSON-ingestion channel): core2 stands in as the APIC query host with an offline
    # APIC export (moquery -o json). Two raised/unacked critical faults, a decommissioned fabric node, and a
    # degraded health score (cur 82) must each fire end-to-end; the minor + acknowledged faults prove the
    # severity/ack filter stays silent. This proves the engine ingests JSON (not just show-text) end-to-end.
    assert isinstance(snap.get("aci"), dict) and snap["aci"].get("core2", {}).get("faults"), \
        "snapshot must publish per-host ACI state (build_aci -> parse_aci_faults; the JSON-ingestion channel)"
    assert any(d.get("id") == "aci-critical-fault-raised" for d in _bp.get("decisions", [])), \
        "engine must assess ACI: a raised/unacknowledged critical fault must fire _d_aci_critical_faults"
    assert any(d.get("id") == "aci-node-not-active" for d in _bp.get("decisions", [])), \
        "engine must assess ACI inventory: a decommissioned fabric node must fire _d_aci_node_not_active"
    assert any(d.get("id") == "aci-fabric-health-degraded" for d in _bp.get("decisions", [])), \
        "engine must assess ACI health: a sub-90 fabric health score must fire _d_aci_fabric_health_degraded"
    assert any(d.get("id") == "aci-vrf-enforcement-unenforced" for d in _bp.get("decisions", [])), \
        "engine must assess ACI logical inventory: an unenforced VRF (default-permit) must fire _d_aci_vrf_unenforced"
    # ACI logical CENSUS (move-group-scoping inventory — published, not a detector): core2 carries the
    # tenant/BD/EPG inventory (the migration move-group units) for the deliverables / a future wave-planner.
    _aci_core2 = (snap.get("aci") or {}).get("core2") or {}
    assert _aci_core2.get("tenants") and _aci_core2.get("bds") and _aci_core2.get("epgs"), \
        "snapshot must publish the ACI logical census (build_aci -> parse_aci_tenants/bds/epgs)"
    # ACI move-group PLAN: the design blueprint's target_state derives tenant-by-tenant ACI move-groups from
    # the published census (the wave-planner consuming the ACI logical inventory). The LEGACY tenant's
    # unenforced VRF surfaces as a segmentation gap on its move group -- design-engine integration, end-to-end.
    _amg = (_bp.get("target_state") or {}).get("aci_move_groups") or {}
    assert _amg.get("groups") and _amg.get("n_tenants", 0) >= 2, \
        "blueprint target_state must carry ACI tenant move-groups derived from the logical census"
    assert _amg.get("n_segmentation_gaps", 0) >= 1, \
        "an unenforced-VRF tenant must be flagged as a move-group segmentation gap"
    # UNIVERSALITY (Cisco Catalyst SD-WAN / vManage JSON channel): core1 stands in as the vManage query host;
    # a DOWN vsmart control connection and an UNREACHABLE device must each fire end-to-end (the up vbond
    # connection and the reachable device prove no over-firing). The second JSON-ingestion controller fabric.
    assert isinstance(snap.get("sdwan"), dict) and snap["sdwan"].get("core1", {}).get("control_connections"), \
        "snapshot must publish per-host SD-WAN state (build_sdwan -> parse_sdwan_control_connections)"
    assert any(d.get("id") == "sdwan-control-connection-down" for d in _bp.get("decisions", [])), \
        "engine must assess SD-WAN: a down control connection must fire _d_sdwan_control_connection_down"
    assert any(d.get("id") == "sdwan-device-unreachable" for d in _bp.get("decisions", [])), \
        "engine must assess SD-WAN: an unreachable device must fire _d_sdwan_device_unreachable"
    assert any(d.get("id") == "sdwan-omp-peer-down" for d in _bp.get("decisions", [])), \
        "engine must assess SD-WAN OMP (deeper modeling): an edge with ompPeersDown>0 must fire _d_sdwan_omp_peer_down"
    # UNIVERSALITY (SD-Access LISP fabric control plane): core1 is an IOS-XE fabric node whose VRF 'red' has
    # 2 control-plane (map-server/map-resolver) sessions configured but ZERO established (both peers Down),
    # while the healthy VRF 'default' (2/2 established, peers Up) in the same output proves no over-firing.
    assert isinstance(snap.get("lisp"), dict) and snap["lisp"].get("core1", {}).get("sessions"), \
        "snapshot must publish per-device LISP fabric state (build_lisp -> parse_lisp_sessions)"
    assert any(d.get("id") == "lisp-fabric-session-down" for d in _bp.get("decisions", [])), \
        "engine must assess SD-Access LISP control plane: a VRF with total>=1/established==0 must fire _d_lisp_fabric_session_down"
    # UNIVERSALITY (Cisco TrustSec / CTS segmentation): core1 is a TrustSec node whose environment-data
    # download is stuck in WAITING_RESPONSE (not COMPLETE) -> the SGT/SGACL policy map is never downloaded,
    # so _d_cts_environment_data_health must fire end-to-end. A non-CTS device publishes no cts entry and
    # stays silent (coverage-honest).
    assert isinstance(snap.get("cts"), dict) and snap["cts"].get("core1", {}).get("environment_data"), \
        "snapshot must publish per-device CTS state (build_cts -> parse_cts_environment_data)"
    assert snap["cts"]["core1"]["environment_data"].get("state") == "WAITING_RESPONSE", \
        "core1 CTS env-data 'Current state' must be the observed non-COMPLETE value"
    assert any(d.get("id") == "cts-environment-data-not-downloaded" for d in _bp.get("decisions", [])), \
        "engine must assess TrustSec segmentation: a non-COMPLETE CTS env-data download must fire _d_cts_environment_data_health"
    # UNIVERSALITY (DMVPN WAN overlay mGRE/NHRP): core1 acts as a DMVPN hub with two spoke tunnels not in UP
    # (10.0.1.3 NHRP, 10.0.1.4 IKE) while 10.0.1.2 is UP.  The detector must fire end-to-end; the UP peer proves
    # no over-firing.
    assert isinstance(snap.get("dmvpn"), dict) and snap["dmvpn"].get("core1", {}).get("peers"), \
        "snapshot must publish per-device DMVPN state (build_dmvpn -> parse_dmvpn_peers)"
    assert any(d.get("id") == "dmvpn-tunnel-peer-down" for d in _bp.get("decisions", [])), \
        "engine must assess the DMVPN WAN overlay: a not-UP tunnel peer (NHRP/IKE) must fire _d_dmvpn_tunnel_health"
    # UNIVERSALITY (IPsec encrypted WAN): core1 acts as an IOS site-to-site IPsec hub with a DOWN-NEGOTIATING
    # crypto session (Tunnel1 -> 10.0.255.9). The detector must fire end-to-end; the healthy companion
    # (UP-ACTIVE Tunnel0 -> 10.0.255.2) proves no over-firing.
    assert isinstance(snap.get("crypto"), dict) and snap["crypto"].get("core1", {}).get("sessions"), \
        "snapshot must publish per-device IPsec crypto state (build_crypto -> parse_crypto_sessions)"
    assert any(d.get("id") == "ipsec-crypto-session-down" for d in _bp.get("decisions", [])), \
        "engine must assess IPsec encrypted WAN: a DOWN-NEGOTIATING crypto session must fire _d_crypto_session_health"
    # UNIVERSALITY (BFD fast-failover): core1 runs BFD with one session DOWN (10.0.255.9 on Gi1/0/3) and one
    # UP (10.0.255.2 on Gi1/0/1).  The detector must fire end-to-end; the healthy Up session (whose RH/RS
    # column is also 'Up') proves no over-firing and proves the parser reads State by column, not first token.
    assert isinstance(snap.get("bfd"), dict) and snap["bfd"].get("core1", {}).get("sessions"), \
        "snapshot must publish per-device BFD state (build_bfd -> parse_bfd_neighbors)"
    assert any(d.get("id") == "bfd-session-down-failover-degraded" for d in _bp.get("decisions", [])), \
        "engine must assess BFD fast-failover: a Down BFD session must fire _d_bfd_session_health"
    # UNIVERSALITY (IPv6 addressing / ND): core1 is a dual-stack distribution switch whose Vlan30 GLOBAL IPv6
    # address is in the DUPLICATE state (DAD found a clash -> IOS disabled it). The DAD detector must fire
    # end-to-end; the clean Vlan10/Gi1/0/24 addresses and the TENTATIVE Gi1/0/1 address prove no over-firing.
    assert isinstance(snap.get("ipv6_nd"), dict) and snap["ipv6_nd"].get("core1", {}).get("interfaces"), \
        "snapshot must publish per-device IPv6 ND state (build_ipv6_nd -> parse_ipv6_interface_addrs)"
    assert any(d.get("id") == "ipv6-duplicate-address-dad-failure" for d in _bp.get("decisions", [])), \
        "engine must assess IPv6 DAD: a global address in the DUPLICATE state must fire _d_ipv6_dad_duplicate"
    # UNIVERSALITY (IPv6 routing plane / dual-stack reachability): access1 is dual-stack and runs OSPFv3 + IPv6 BGP
    # with one OSPFv3 neighbor stuck EXSTART (10.0.0.9) and one IPv6 BGP peer Active (2001:DB8:0:9::9). The IPv6
    # routing-adjacency detector must fire end-to-end; the healthy companions (FULL/DR, 2WAY/DROTHER, Established
    # PfxRcd 12) prove no over-firing.
    assert isinstance(snap.get("ipv6_routing"), dict) and snap["ipv6_routing"].get("access1", {}).get("ospfv3_neighbors"), \
        "snapshot must publish per-device IPv6 routing state (build_ipv6_routing -> parse_ospfv3_neighbors)"
    assert any(d.get("id") == "ipv6-routing-adjacency-down" for d in _bp.get("decisions", [])), \
        "engine must assess the IPv6 routing plane: a stuck OSPFv3 adjacency / not-Established IPv6 BGP peer must fire _d_ipv6_routing_adjacency"
    # UNIVERSALITY (MULTI-VENDOR -- Arista EOS): core2 ALSO stands in as an Arista spine exporting 'show mlag |
    # json'. The domain is configured (state active) but DEGRADED (configSanity inconsistent + Inactive ports)
    # -> _d_arista_mlag_degraded FIRES end-to-end. This is the FIRST NON-CISCO vendor axis -- the engine now
    # assesses a non-Cisco platform's core redundancy construct (MLAG, the analogue of Cisco vPC) through the
    # same parse->build->signal->detect->coverage pipeline. A healthy / transient / 'disabled' domain stays
    # silent (proved in tests/test_arista.py). Coverage-honest on Meridian: the all-Cisco fleet has no MLAG -> silent.
    assert isinstance(snap.get("arista"), dict) and snap["arista"].get("core2", {}).get("mlag"), \
        "snapshot must publish per-device Arista MLAG state (build_arista -> parse_arista_mlag, the first non-Cisco vendor axis)"
    assert any(d.get("id") == "arista-mlag-domain-degraded" for d in _bp.get("decisions", [])), \
        "engine must assess Arista MLAG: a config-inconsistent / Inactive-port MLAG domain must fire _d_arista_mlag_degraded"
    # UNIVERSALITY (MULTI-VENDOR -- Juniper Junos, the SECOND non-Cisco vendor): core2 ALSO stands in as a
    # Juniper SRX chassis cluster ('show chassis cluster status | display json'). RG0's secondary is at PRIORITY
    # 0 (configured but not ready to accept traffic) -> _d_junos_chassis_cluster_degraded FIRES end-to-end. This
    # proves the vendor-adapter pattern generalises beyond Arista: a different NOS, a different deeply-nested
    # JSON dialect, the same coverage-honest contract. Silent on Meridian (the all-Cisco fleet runs no SRX cluster).
    assert isinstance(snap.get("juniper"), dict) and snap["juniper"].get("core2", {}).get("chassis_cluster"), \
        "snapshot must publish per-device Juniper chassis-cluster state (build_juniper -> parse_junos_chassis_cluster, the second non-Cisco vendor axis)"
    assert any(d.get("id") == "junos-chassis-cluster-ha-degraded" for d in _bp.get("decisions", [])), \
        "engine must assess Juniper SRX HA: a priority-0 redundancy-group node must fire _d_junos_chassis_cluster_degraded"
    # UNIVERSALITY (PUBLIC CLOUD -- AWS, the FIRST cloud-domain axis): core1 ALSO stands in as an AWS account
    # ('aws ec2 describe-security-groups'). sg-0bastion opens SSH(22) to 0.0.0.0/0 -> _d_cloud_sg_open_ingress
    # FIRES (CIS 5.2); sg-0pubweb's 443-to-world stays silent (no cry-wolf). This proves the engine extends
    # beyond on-prem multi-vendor to PUBLIC CLOUD via the same offline JSON pattern. Silent on Meridian (no cloud export).
    assert isinstance(snap.get("cloud"), dict) and snap["cloud"].get("core1", {}).get("security_groups"), \
        "snapshot must publish per-account cloud state (build_cloud -> parse_aws_security_groups, the first cloud-domain axis)"
    assert any(d.get("id") == "cloud-security-group-open-ingress" for d in _bp.get("decisions", [])), \
        "engine must assess cloud exposure: an SSH-open-to-0.0.0.0/0 security group must fire _d_cloud_sg_open_ingress"
    # UNIVERSALITY (MULTI-VENDOR -- Fortinet FortiGate, the THIRD non-Cisco vendor): core2 ALSO stands in as a
    # FortiGate HA cluster ('get system ha status'). The secondary is OUT-OF-SYNC (a config-checksum mismatch ->
    # the standby holds a divergent ruleset) -> _d_fortigate_ha_degraded FIRES. Three vendors now assess their
    # firewall/HA construct through the same coverage-honest pipeline. Silent on Meridian (no FortiGate captures).
    assert isinstance(snap.get("fortigate"), dict) and snap["fortigate"].get("core2", {}).get("ha"), \
        "snapshot must publish per-device FortiGate HA state (build_fortigate -> parse_fortigate_ha_status, the third non-Cisco vendor axis)"
    assert any(d.get("id") == "fortigate-ha-cluster-out-of-sync" for d in _bp.get("decisions", [])), \
        "engine must assess FortiGate HA: an out-of-sync cluster member must fire _d_fortigate_ha_degraded"
    # SSOT: the design-driven NRFU checklist is ALSO published (the one source the explorer + webapp read,
    # so neither re-derives the phased acceptance items) and equals a fresh compute over the blueprint.
    from cisco_toolkit.design_advisor import compute_design_nrfu
    assert "design_nrfu" in snap, "snapshot must publish the canonical design_nrfu"
    assert isinstance(snap["design_nrfu"].get("items"), list)
    assert snap["design_nrfu"] == compute_design_nrfu(_bp), \
        "published design_nrfu must equal the canonical recompute from the published blueprint"
    # SSOT: the architecture-coverage map is published (one source for "which architecture classes did we
    # assess, and what did we find") and equals a fresh compute. The fixtures exercise BOTH ingestion channels
    # -- ssh show-text AND json controller-REST -- so aci/sdwan (json) + mpls (ssh) are observed-with-findings;
    # this proves the engine assesses the full architecture universe across both channels, end-to-end.
    from cisco_toolkit.design_advisor import compute_architecture_coverage
    assert "architecture_coverage" in snap, "snapshot must publish the architecture_coverage SSOT"
    _cov = snap["architecture_coverage"]
    assert _cov == compute_architecture_coverage(snap), \
        "published architecture_coverage must equal the canonical recompute"
    # Coverage matrix (Plan-A #5): the composed per-(device,axis) SSOT is published + equals its recompute,
    # its abstention count is self-consistent, and no collection blind-spot reads a fake 'covered'.
    from cisco_toolkit.coverage_matrix import compute_coverage_matrix
    assert "coverage_matrix" in snap, "snapshot must publish the coverage_matrix SSOT"
    _cm = snap["coverage_matrix"]
    assert _cm == compute_coverage_matrix(snap), "published coverage_matrix must equal the canonical recompute"
    from cisco_toolkit.unknown_evidence import compute_unknown_evidence
    _ue = snap["unknown_evidence"]
    assert _ue == compute_unknown_evidence(snap), \
        "published unknown_evidence must equal the canonical aggregate-only recompute"
    assert _ue["summary"]["raw_identifiers_included"] is False
    assert _ue["summary"]["claim_scope"] == "bounded_observed_sources_only"
    assert _cm["summary"]["n_abstained"] == sum(1 for _r in _cm["rows"] if _r["is_abstention"])
    _cc_blind = {d.get("host") for d in (snap.get("collection_completeness") or {}).get("devices", [])
                 if isinstance(d, dict) and "complete" not in str(d.get("status", "")).lower()}
    assert not any(r["axis"] == "collection" and r["device"] in _cc_blind and r["state"] == "covered"
                   for r in _cm["rows"]), "a collection blind-spot must never read 'covered'"
    _covby = {c["key"]: c for c in _cov["classes"]}
    assert _covby["aci"]["status"] == "finding" and _covby["aci"]["channel"] == "json", \
        "ACI (json controller channel) must be observed-with-findings on the fixtures"
    assert _covby["sdwan"]["status"] == "finding" and _covby["mpls"]["status"] == "finding"
    assert _cov["summary"]["by_channel"] == {"ssh": 22, "json": 5}
    assert _cov["summary"]["n_with_findings"] == 27, \
        "every one of the 27 architecture classes (THREE non-Cisco vendors -- Arista + Juniper + Fortinet -- AND the public-cloud domain) fires on the synthetic fixtures (full-universe proof)"

    # ---- explorer (snapshot embedded into the single-file viewer) ----
    explorer = os.path.splitext(str(out_xlsx))[0] + "_explorer.html"
    assert os.path.isfile(explorer), "explorer HTML was not written"
    html = open(explorer, encoding="utf-8").read()
    assert "EMBEDDED_SNAPSHOT" in html, "explorer did not get the live snapshot embedded"
    # SSOT (explorer dashboard): _slim_for_embed shrinks the in-page payload, so it must PRESERVE the
    # canonical executive_brief.scale — that is the one source the explorer's censusSec() reads
    # (SNAP.executive_brief.scale ?? MODEL...). If a future slim drops it, the census silently falls back
    # to a client-side MODEL recount = the 148-vs-202 drift class. Lock that the embedded payload carries
    # the same canonical scale the snapshot publishes, so python and dashboard cannot diverge.
    _embedded_json = html.split("const EMBEDDED_SNAPSHOT=", 1)[1] \
        .split("load(EMBEDDED_SNAPSHOT,", 1)[0].rstrip().rstrip(";").rstrip()
    _embedded = json.loads(_embedded_json)
    _embed_scale = (_embedded.get("executive_brief") or {}).get("scale") or {}
    for _k in ("n_vlans", "n_endpoints", "n_devices"):
        assert _embed_scale.get(_k) == _scale.get(_k), (
            f"explorer embed dropped canonical scale.{_k} "
            f"({_embed_scale.get(_k)!r} != published {_scale.get(_k)!r}) — census would drift to a client recount")
    # SSOT (explorer ✎Design): the architecture-coverage map reaches the embedded payload AND the explorer has
    # the renderer wired (drawArchCoverage reads SNAP.architecture_coverage — no client recompute), so the
    # universal-coverage view is interactive on the dashboard, one source with the engine.
    _embed_cov = _embedded.get("architecture_coverage") or {}
    assert (_embed_cov.get("summary") or {}).get("n_classes") == _cov["summary"]["n_classes"], \
        "explorer embed must carry the engine's architecture_coverage SSOT (not slimmed away)"
    assert "drawArchCoverage" in html and "Architecture coverage" in html, \
        "explorer must render the architecture-coverage section (drawArchCoverage wired into the Design view)"

    # ---- executive deck (optional python-pptx): if the lib is present, main() must have written it ----
    try:
        import pptx  # noqa: F401
    except ImportError:
        pass
    else:
        deck = os.path.splitext(str(out_xlsx))[0] + "_executive_deck.pptx"
        assert os.path.isfile(deck), "executive deck (PPTX) was not written despite python-pptx installed"


# ------------------------------------------------------- PPDIOO gate refusal: record + exit code

def _gated_run(tmp_path, monkeypatch, eng_root, *extra):
    """One real engine run against `eng_root`'s gate ledger. Returns main()'s exit code.

    Deliberately drives the WHOLE engine rather than calling gate_state directly: the record this
    asserts on has to come from the real producer, or the test only proves that a fixture agrees
    with the parser that made it. The optional heavyweight deliverables are off (the gated ones,
    design + MOP, stay on) purely for runtime.

    `--gate-root` is passed EXPLICITLY rather than relying on the chdir below, so the test does not
    manufacture its own precondition: it exercises the flag a wrapper with a synthetic cwd must use.
    """
    collection = fx.write_collection(str(tmp_path / "collection"))
    devices = tmp_path / "devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "out.xlsx"
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess",
        "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--gate-root", str(eng_root),
        "--no-html", "--no-docx", "--no-pptx", "--no-crd", "--no-engagement",
        "--no-opshandbook", "--no-archreview",
        *extra,
    ])
    return cp.main(), out_xlsx


def test_gate_refusal_is_recorded_durably_and_exits_0_unless_asked(tmp_path, monkeypatch):
    """End-to-end, through the engine: a refused deliverable leaves a durable ledger row, is really
    absent from disk, and does NOT fail the run — the exit code is opt-in.

    The default matters as much as the flag. webapp/backend/ingest.py raises on any non-zero and
    serve.py renders that to a field engineer as "redaction FAILED ... Treat anything already
    written as UNREDACTED", so a refusal reported as a failure is a false safety alarm on a run that
    behaved correctly.
    """
    from cisco_toolkit import gate_state

    eng = tmp_path / "engagement"
    eng.mkdir()
    # A ledger that exists (so gates are ACTIVE) with design's upstream explicitly revoked and the
    # MOP's never signed: both gated deliverables must be withheld.
    gate_state.record_decision("assessment_approved", "revoked", root=str(eng), by="lead")

    rc, out_xlsx = _gated_run(tmp_path, monkeypatch, eng)
    assert rc == 0, "a correct run that withheld a gated document must not report failure"

    design = os.path.splitext(str(out_xlsx))[0] + "_design.docx"
    mop = os.path.splitext(str(out_xlsx))[0] + "_mop.docx"
    assert not os.path.exists(design), "the design was written despite a revoked assessment gate"
    assert not os.path.exists(mop), "the MOP was written despite unsigned upstream gates"
    assert out_xlsx.is_file(), "the refusal must withhold only the gated documents"

    ledger = json.loads((eng / "docs" / "engagement-state.json").read_text(encoding="utf-8"))
    refusals = [a for a in ledger["audit"] if a["event"] == "refuse"]
    assert {a["generator"] for a in refusals} == {"design", "mop"}, \
        f"the engine's refusals did not reach the ledger: {refusals}"
    assert all(a["who"] and a["at"] and a["reason"] for a in refusals)
    assert refusals[0]["missing"] == ["assessment_approved"]

    # Re-running is the natural response to a refusal — the reason this record lives in the ledger
    # and not in the per-run manifest a re-run overwrites. The first run's rows must still be there.
    rc2, _ = _gated_run(tmp_path, monkeypatch, eng, "--fail-on-gate-refusal")
    assert rc2 == 2, "--fail-on-gate-refusal did not surface the withheld deliverables"

    ledger2 = json.loads((eng / "docs" / "engagement-state.json").read_text(encoding="utf-8"))
    assert len([a for a in ledger2["audit"] if a["event"] == "refuse"]) == 4, \
        "the re-run replaced the earlier refusals instead of appending to them"


def test_gate_refusal_exit_code_reaches_the_bare_script_door(tmp_path):
    """`python COLLECT_PARSE_V3_23_0.py` swallowed main()'s RETURN VALUE entirely (a bare `main()`
    under __main__), while the cisco-assess console script and Atlas's --run-engine dispatch both
    honoured it. tests/test_pipeline_golden.py drives THAT door, so the repo's only e2e exit-code
    assertion was blind to any code main() returned.

    This must run the real pipeline to a REFUSAL, because that is the only way main() *returns* a
    non-zero code. An argparse usage error would be no proof at all: `ap.error()` raises SystemExit,
    which propagates out of the module whether or not the door wraps main() in sys.exit — that
    version of this test passed against the unfixed door.
    """
    import subprocess

    from cisco_toolkit import gate_state

    eng = tmp_path / "engagement"
    eng.mkdir()
    gate_state.record_decision("assessment_approved", "revoked", root=str(eng), by="lead")
    collection = fx.write_collection(str(tmp_path / "collection"))
    devices = tmp_path / "devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "template.xlsx"
    _make_template(str(template))

    script = os.path.join(ROOT, "COLLECT_PARSE_V3_23_0.py")
    proc = subprocess.run(
        [sys.executable, script,
         "--no-collect", "--collection-dir", collection,
         "--devices-file", str(devices), "--template", str(template),
         "--output", str(tmp_path / "out.xlsx"), "--workers", "1",
         "--gate-root", str(eng), "--fail-on-gate-refusal",
         "--no-html", "--no-docx", "--no-pptx", "--no-crd", "--no-engagement",
         "--no-opshandbook", "--no-archreview"],
        cwd=str(tmp_path), capture_output=True, text=True, timeout=300)
    assert proc.returncode == 2, (
        f"the bare-script door swallowed main()'s returned exit code (got {proc.returncode})\n"
        f"STDOUT\n{proc.stdout[-2000:]}\nSTDERR\n{proc.stderr[-2000:]}")


def test_pipeline_protocol_health_failure_publishes_analysis_unavailable_receipt(tmp_path, monkeypatch):
    """The real call site must carry a failed health phase into every protocol surface.

    Direct writer tests were insufficient: the old pipeline always passed ``[]`` and never set the
    writers' unavailable flag, making a compute failure look identical to no findings.
    """
    collection = fx.write_collection(str(tmp_path / "failed-protocol-collection"))
    devices = tmp_path / "failed-protocol-devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "failed-protocol-template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "failed-protocol-health.xlsx"

    def _fail_protocol_health(*_args, **_kwargs):
        raise RuntimeError("synthetic protocol-health failure")

    monkeypatch.setattr(cp, "compute_protocol_health", _fail_protocol_health)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess", "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--no-html", "--no-docx", "--no-pptx", "--no-design", "--no-mop",
        "--no-crd", "--no-engagement", "--no-opshandbook", "--no-archreview",
    ])

    cp.main()

    snap = json.loads((tmp_path / "failed-protocol-health.snapshot.json").read_text(encoding="utf-8"))
    receipt = snap["protocol_assessability"]
    assert snap["protocol_health"] == [] and snap["protocol_intelligence"] == []
    assert receipt["schema"] == "protocol_assessability/1"
    assert receipt["summary"]["n_cells"] == len(fx.DEVICES) * 7
    assert receipt["summary"]["by_state"]["analysis_unavailable"] == len(fx.DEVICES) * 7
    assert {row["state"] for row in receipt["rows"]} == {"analysis_unavailable"}
    assert "Protocol Health" in set(snap["assessment_integrity"]["failed_phases"])

    # The raw summary projection remains available, but the failed health phase makes its exact
    # device-family receipt non-assessed.  The shared baseline must therefore fail closed rather than
    # turning association or raw (P) tokens into an accepted operational target.
    assert snap["etherchannel_projection"]["schema"] == "etherchannel_projection/1"
    failed_etherchannel = snap["etherchannel_baseline"]
    assert failed_etherchannel["schema"] == "etherchannel_baseline/1"
    assert failed_etherchannel["projection_custody"] == "embedded_unverified"
    failed_baseline_by_host = {row["switch"]: row for row in failed_etherchannel["rows"]}
    assert set(failed_baseline_by_host) == {"core1", "core2"}
    assert failed_etherchannel["status"] == "not_verified"
    assert failed_etherchannel["assessed"] is False
    assert failed_etherchannel["summary"]["by_status"] == {
        "assessed": 0, "degraded": 0, "review": 0, "not_verified": 2,
    }
    for host, baseline_row in failed_baseline_by_host.items():
        assert baseline_row["status"] == "not_verified"
        assert baseline_row["receipt_state"] == "analysis_unavailable"
        assert baseline_row["capture_state"] == "usable"
        assert baseline_row["group_count"] == 1 and baseline_row["member_count"] == 2
        assert baseline_row["acceptance"].startswith(
            "ETHERCHANNEL BASELINE NOT VERIFIED — BLOCKER:"
        )
        assert baseline_row["source_key"] == (
            f"etherchannel_projection.rows[{host}] + "
            f"protocol_assessability.rows[{host},EtherChannel]"
        )
        assert "all-(P)" not in baseline_row["acceptance"]
        assert "all physical members are in (P)/bundled state" not in baseline_row["acceptance"]

    # A failed health phase cannot leave the independently parsed peer projection looking like an accepted
    # baseline.  Both operator workflows must emit the same explicit recapture blocker.
    routing_validation = [row for row in snap["validation_plan"]["items"]
                          if row.get("category") == "Routing"]
    assert routing_validation
    assert {row["evidence_state"] for row in routing_validation} == {"not_verified"}
    assert all(row["expect"].startswith("ROUTING BASELINE NOT VERIFIED — BLOCKER:")
               for row in routing_validation)
    nrfu_routing = [case for wave in snap["nrfu_commands"]["waves"]
                    for device in wave["devices"] for case in device["cases"]
                    if case.get("evidence_family") in {"OSPF", "BGP", "EIGRP"}]
    assert nrfu_routing
    assert {case["evidence_state"] for case in nrfu_routing} == {"not_verified"}
    assert all(case["expected"].startswith("ROUTING BASELINE NOT VERIFIED — BLOCKER:")
               for case in nrfu_routing)

    etherchannel_validation = {
        row["device"]: row for row in snap["validation_plan"]["items"]
        if row.get("category") == "Link"
        and str(row.get("source_key") or "").startswith("etherchannel_projection.rows[")
    }
    etherchannel_nrfu = {
        device["host"]: case
        for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"]
        for case in device["cases"]
        if case.get("evidence_family") == "EtherChannel"
    }
    assert set(etherchannel_validation) == set(etherchannel_nrfu) == {"core1", "core2"}
    for host in ("core1", "core2"):
        baseline_row = failed_baseline_by_host[host]
        validation_row = etherchannel_validation[host]
        nrfu_case = etherchannel_nrfu[host]
        assert validation_row["evidence_state"] == nrfu_case["evidence_state"] == "not_verified"
        assert validation_row["expect"] == nrfu_case["expected"] == baseline_row["acceptance"]
        assert validation_row["command"] == nrfu_case["command"] == baseline_row["command"]
        assert validation_row["source_key"] == nrfu_case["source_key"] == baseline_row["source_key"]
        assert validation_row["projection_custody"] == nrfu_case["projection_custody"] == (
            "embedded_unverified"
        )
        assert "all-(P)" not in validation_row["expect"]
        assert "all physical members are in (P)/bundled state" not in nrfu_case["expected"]
    assert snap["nrfu_commands"]["summary"]["n_etherchannel_cases"] == 2
    assert snap["nrfu_commands"]["summary"]["n_etherchannel_blockers"] == 2
    assert snap["nrfu_commands"]["summary"]["etherchannel_by_evidence_state"] == {
        "not_verified": 2
    }

    wb = load_workbook(str(out_xlsx), read_only=True)
    assert "UNVERIFIED" in str(wb["Protocol Health"]["A2"].value)
    assert "UNVERIFIED" in str(wb["Protocol Intelligence"]["A2"].value)
    cc_text = "\n".join(
        str(cell.value or "") for row in wb["Collection Completeness"].iter_rows() for cell in row
    )
    wb.close()
    assert "ANALYSIS UNAVAILABLE" in cc_text
    assert "no health conclusion is asserted" in cc_text


def test_pipeline_bgp_denominator_phase_failure_fails_closed_every_consumer(
        tmp_path, monkeypatch):
    """A crashed configured-peer owner is not serialized as feature absence or healthy BGP."""
    from cisco_toolkit.analyze import compute_current_baseline_gate

    collection = fx.write_collection(str(tmp_path / "failed-bgp-collection"))
    devices = tmp_path / "failed-bgp-devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "failed-bgp-template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "failed-bgp.xlsx"

    def _fail_bgp_denominator(*_args, **_kwargs):
        raise RuntimeError("synthetic configured-BGP phase failure")

    monkeypatch.setattr(cp, "compute_bgp_configured_peer_baseline", _fail_bgp_denominator)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess", "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--no-html", "--no-docx", "--no-pptx", "--no-design", "--no-mop",
        "--no-crd", "--no-engagement", "--no-opshandbook", "--no-archreview",
    ])

    cp.main()
    snap = json.loads((tmp_path / "failed-bgp.snapshot.json").read_text(encoding="utf-8"))
    assert snap["bgp_configured_peer_baseline"]["verdict"] == "INDETERMINATE"
    assert snap["bgp_configured_peer_baseline"]["projection_custody"] == "embedded_unverified"
    assert "BGP configured-peer baseline" in set(
        snap["assessment_integrity"]["failed_phases"])

    assert all(group["readiness"] != "READY" for group in snap["migration_readiness"])
    routing_checks = [
        check for group in snap["migration_readiness"] for check in group["checks"]
        if check["check"] == "Routing adjacencies up"
    ]
    assert routing_checks and all(check["status"] in {"warn", "fail"}
                                  for check in routing_checks)

    bgp_validation = [
        row for row in snap["validation_plan"]["items"]
        if row.get("evidence_state") == "not_verified"
        and row["expect"].startswith("BGP CONFIGURED PEER NOT VERIFIED — BLOCKER:")
    ]
    assert bgp_validation
    assert compute_current_baseline_gate(snap["validation_plan"])["verdict"] != "CLEAR"

    bgp_nrfu = [
        case for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"] for case in device["cases"]
        if case.get("evidence_family") == "BGP"
    ]
    assert bgp_nrfu and all(case["evidence_state"] == "not_verified"
                            for case in bgp_nrfu)
    assert snap["nrfu_commands"]["summary"]["n_routing_blockers"] >= len(bgp_nrfu)


def test_pipeline_fhrp_denominator_phase_failure_fails_closed_every_consumer(
        tmp_path, monkeypatch):
    """A crashed configured-group owner is not serialized as feature absence or healthy FHRP."""
    from cisco_toolkit.analyze import compute_current_baseline_gate

    collection = fx.write_collection(str(tmp_path / "failed-fhrp-collection"))
    devices = tmp_path / "failed-fhrp-devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "failed-fhrp-template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "failed-fhrp.xlsx"

    def _fail_fhrp_denominator(*_args, **_kwargs):
        raise RuntimeError("synthetic configured-FHRP phase failure")

    monkeypatch.setattr(cp, "compute_fhrp_configured_group_baseline", _fail_fhrp_denominator)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess", "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--no-html", "--no-docx", "--no-pptx", "--no-design", "--no-mop",
        "--no-crd", "--no-engagement", "--no-opshandbook", "--no-archreview",
    ])

    cp.main()
    snap = json.loads((tmp_path / "failed-fhrp.snapshot.json").read_text(encoding="utf-8"))
    baseline = snap["fhrp_configured_group_baseline"]
    assert baseline["verdict"] == "INDETERMINATE"
    assert baseline["projection_custody"] == "embedded_unverified"
    assert "FHRP configured-group baseline" in set(
        snap["assessment_integrity"]["failed_phases"])

    assert all(group["readiness"] != "READY" for group in snap["migration_readiness"])
    gateway_checks = [
        check for group in snap["migration_readiness"] for check in group["checks"]
        if check["check"] == "Gateway redundancy"
    ]
    assert gateway_checks and all(check["status"] in {"warn", "fail"}
                                  for check in gateway_checks)

    fhrp_validation = [
        row for row in snap["validation_plan"]["items"]
        if row.get("evidence_state") == "not_verified"
        and row["expect"].startswith("FHRP CONFIGURED GROUP NOT VERIFIED — BLOCKER:")
    ]
    assert fhrp_validation
    assert compute_current_baseline_gate(snap["validation_plan"])["verdict"] != "CLEAR"

    fhrp_nrfu = [
        case for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"] for case in device["cases"]
        if case.get("evidence_family") == "FHRP"
    ]
    assert fhrp_nrfu and all(case["evidence_state"] == "not_verified"
                            for case in fhrp_nrfu)
    assert snap["nrfu_commands"]["summary"]["n_fhrp_blockers"] == len(fhrp_nrfu)


def test_pipeline_fhrp_domain_phase_failure_publishes_static_blockers(
        tmp_path, monkeypatch):
    """A crashed cross-switch owner publishes an abstention and every safe scoped SVI survives."""
    from cisco_toolkit.analyze import compute_current_baseline_gate

    collection = fx.write_collection(str(tmp_path / "failed-fhrp-domain-collection"))
    devices = tmp_path / "failed-fhrp-domain-devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "failed-fhrp-domain-template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "failed-fhrp-domain.xlsx"

    def _fail_fhrp_domain(*_args, **_kwargs):
        raise RuntimeError("synthetic FHRP redundancy-domain phase failure")

    monkeypatch.setattr(cp, "compute_fhrp_redundancy_domain_baseline", _fail_fhrp_domain)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess", "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--no-html", "--no-docx", "--no-pptx", "--no-design", "--no-mop",
        "--no-crd", "--no-engagement", "--no-opshandbook", "--no-archreview",
    ])

    cp.main()
    snap = json.loads(
        (tmp_path / "failed-fhrp-domain.snapshot.json").read_text(encoding="utf-8"))
    published = snap["fhrp_redundancy_domain_baseline"]
    assert published["schema"] == "fhrp_redundancy_domain_baseline/1"
    assert published["verdict"] == "INDETERMINATE"
    assert published["projection_custody"] == "embedded_unverified"
    assert published["rows"] == []
    assert "FHRP redundancy-domain baseline" in set(
        snap["assessment_integrity"]["failed_phases"])

    validation = [
        row for row in snap["validation_plan"]["items"]
        if row.get("source_key") == "fhrp_redundancy_domain_baseline"
    ]
    cases = [
        case for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"] for case in device["cases"]
        if case.get("evidence_family") == "FHRP Domain"
    ]
    assert validation and len(validation) == len(cases)
    assert {row["evidence_state"] for row in validation} == {"not_verified"}
    assert {case["evidence_state"] for case in cases} == {"not_verified"}
    assert all(row["expect"].startswith(
        "FHRP REDUNDANCY DOMAIN NOT VERIFIED — BLOCKER:") for row in validation)
    assert all(case["expected"].startswith(
        "FHRP REDUNDANCY DOMAIN NOT VERIFIED — BLOCKER:") for case in cases)
    assert compute_current_baseline_gate(snap["validation_plan"])["verdict"] != "CLEAR"
    summary = snap["nrfu_commands"]["summary"]
    assert summary["n_fhrp_domain_cases"] == len(cases)
    assert summary["n_fhrp_domain_blockers"] == len(cases)
    assert summary["fhrp_domain_by_evidence_state"] == {"not_verified": len(cases)}


def test_pipeline_vtp_safety_phase_failure_publishes_subject_scoped_blocker(
        tmp_path, monkeypatch):
    """A crashed VTP owner cannot erase an independently parsed local VTP subject."""
    from cisco_toolkit.analyze import compute_current_baseline_gate

    collection = fx.write_collection(str(tmp_path / "failed-vtp-collection"))
    vtp_capture = tmp_path / "failed-vtp-collection" / "core1" / fx.cmd_filename(
        "show vtp status")
    vtp_capture.write_text(
        "VTP Version capable             : 1 to 3\n"
        "VTP version running             : 2\n"
        "VTP Domain Name                 : CAMPUS\n"
        "VTP Operating Mode              : Server\n"
        "Configuration Revision          : 150\n",
        encoding="utf-8",
    )
    devices = tmp_path / "failed-vtp-devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "failed-vtp-template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "failed-vtp.xlsx"

    def _fail_vtp_owner(*_args, **_kwargs):
        raise RuntimeError("synthetic VTP safety phase failure")

    monkeypatch.setattr(cp, "compute_vtp_safety_baseline", _fail_vtp_owner)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess", "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--no-html", "--no-docx", "--no-pptx", "--no-design", "--no-mop",
        "--no-crd", "--no-engagement", "--no-opshandbook", "--no-archreview",
    ])

    cp.main()
    snap = json.loads((tmp_path / "failed-vtp.snapshot.json").read_text(encoding="utf-8"))
    published = snap["vtp_safety_baseline"]
    assert published["schema"] == "vtp_safety_baseline/1"
    assert published["verdict"] == "INDETERMINATE"
    assert published["projection_custody"] == "embedded_unverified"
    assert published["rows"] == []
    assert "VTP safety baseline" in set(snap["assessment_integrity"]["failed_phases"])
    assert any(row.get("switch") == "core1" and row.get("protocol") == "VTP"
               for row in snap["protocol_health"])

    validation = [
        row for row in snap["validation_plan"]["items"]
        if row.get("category") == "VTP"
    ]
    cases = [
        case for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"] for case in device["cases"]
        if case.get("evidence_family") == "VTP"
    ]
    assert len(validation) == len(cases) == 1
    assert validation[0]["device"] == "core1"
    assert validation[0]["evidence_state"] == cases[0]["evidence_state"] == "not_verified"
    assert validation[0]["expect"].startswith(
        "VTP SAFETY BASELINE NOT VERIFIED — BLOCKER:")
    assert cases[0]["expected"] == validation[0]["expect"]
    assert compute_current_baseline_gate(snap["validation_plan"])["verdict"] != "CLEAR"
    assert snap["nrfu_commands"]["summary"]["n_vtp_safety_cases"] == 1
    assert snap["nrfu_commands"]["summary"]["n_vtp_safety_blockers"] == 1
    assert any(row.get("category") == "VTP" for row in snap["punchlist"])
    vtp_checks = [
        check for group in snap["migration_readiness"] for check in group["checks"]
        if check["check"] == "VTP cutover safety"
    ]
    assert vtp_checks and any(check["status"] == "warn" for check in vtp_checks)


def test_pipeline_ipv6_routing_phase_failure_publishes_subject_scoped_blockers(
        tmp_path, monkeypatch):
    """A crashed IPv6 owner cannot erase independently scoped OSPFv3/BGPv6 subjects."""
    from cisco_toolkit.analyze import compute_current_baseline_gate

    collection = fx.write_collection(str(tmp_path / "failed-ipv6-collection"))
    devices = tmp_path / "failed-ipv6-devices.json"
    devices.write_text(json.dumps(fx.DEVICES), encoding="utf-8")
    template = tmp_path / "failed-ipv6-template.xlsx"
    _make_template(str(template))
    out_xlsx = tmp_path / "failed-ipv6.xlsx"

    def _fail_ipv6_owner(*_args, **_kwargs):
        raise RuntimeError("synthetic IPv6 routing owner phase failure")

    monkeypatch.setattr(
        cp, "compute_ipv6_routing_adjacency_baseline", _fail_ipv6_owner)
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(sys, "argv", [
        "cisco-assess", "--no-collect", "--collection-dir", collection,
        "--devices-file", str(devices), "--template", str(template),
        "--output", str(out_xlsx), "--workers", "1",
        "--no-html", "--no-docx", "--no-pptx", "--no-design", "--no-mop",
        "--no-crd", "--no-engagement", "--no-opshandbook", "--no-archreview",
    ])

    cp.main()
    snap = json.loads((tmp_path / "failed-ipv6.snapshot.json").read_text(
        encoding="utf-8"))
    published = snap["ipv6_routing_adjacency_baseline"]
    assert published["schema"] == "ipv6_routing_adjacency_baseline/1"
    assert published["verdict"] == "INDETERMINATE"
    assert published["projection_custody"] == "embedded_unverified"
    assert published["rows"] == []
    assert "IPv6 routing adjacency baseline" in set(
        snap["assessment_integrity"]["failed_phases"])

    validation = [
        row for row in snap["validation_plan"]["items"]
        if row.get("category") == "IPv6 Routing"
    ]
    cases = [
        case for wave in snap["nrfu_commands"]["waves"]
        for device in wave["devices"] for case in device["cases"]
        if case.get("evidence_family") == "IPv6 Routing"
    ]
    assert len(validation) == len(cases) == 2
    assert {row["device"] for row in validation} == {"access1"}
    assert {row["evidence_state"] for row in validation} == {"not_verified"}
    assert {case["evidence_state"] for case in cases} == {"not_verified"}
    assert all(row["expect"].startswith(
        "IPV6 ROUTING BASELINE NOT VERIFIED — BLOCKER:") for row in validation)
    assert all(case["expected"].startswith(
        "IPV6 ROUTING BASELINE NOT VERIFIED — BLOCKER:") for case in cases)
    assert compute_current_baseline_gate(snap["validation_plan"])["verdict"] != "CLEAR"
    summary = snap["nrfu_commands"]["summary"]
    assert summary["n_ipv6_routing_cases"] == 2
    assert summary["n_ipv6_routing_blockers"] == 2
    assert len([row for row in snap["punchlist"]
                if row.get("category") == "IPv6 Routing"]) == 2
    ipv6_checks = [
        check for group in snap["migration_readiness"] for check in group["checks"]
        if check["check"] == "IPv6 routing adjacencies"
    ]
    assert ipv6_checks and any(check["status"] == "warn" for check in ipv6_checks)
