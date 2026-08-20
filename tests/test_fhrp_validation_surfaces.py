"""Rendered validation surfaces must not certify a degraded FHRP baseline as good."""

from openpyxl import Workbook
import pytest

from cisco_toolkit.excel import write_validation_plan_sheet


_BANNER = (
    "Run these AFTER each wave's cutover. Each row carries the observed pre-cutover baseline "
    "and its post-cutover acceptance condition. A PRE-CUTOVER DEGRADED row is a blocker to "
    "resolve or explicitly risk-accept before the window; matching that degraded state after "
    "cutover is not success."
)


def _plan() -> dict:
    item = {
        "device": "dist1",
        "platform": "ios",
        "wave": "Group 1",
        "category": "FHRP",
        "severity": "High",
        "check": "HSRP election for VLAN 10",
        "command": "show standby brief",
        "expect": (
            "PRE-CUTOVER DEGRADED — observed HSRP Init; acceptance requires one Active and "
            "one Standby before this wave can pass"
        ),
        "why": "Init is non-forwarding and cannot be treated as a successful baseline.",
    }
    return {
        "items": [item],
        "by_wave": {"Group 1": [item]},
        "summary": {"n_items": 1, "n_waves": 1, "n_high": 1, "by_category": {"FHRP": 1}},
        "banner": _BANNER,
    }


def test_cutover_validation_sheet_labels_degraded_baseline_as_acceptance_not_good():
    wb = Workbook()
    write_validation_plan_sheet(wb, _plan())
    ws = wb["Cutover Validation"]

    assert ws.cell(4, 9).value == "Observed baseline / acceptance"
    assert "PRE-CUTOVER DEGRADED" in ws.cell(5, 9).value
    assert "known-good" not in ws.cell(1, 1).value


def test_runbook_explains_that_matching_a_degraded_baseline_is_not_success(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "fhrp-validation.docx"
    write_runbook_docx(str(output), {"script_version": "test", "validation_plan": _plan()}, "FHRP")
    document = docx.Document(str(output))
    text = "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )

    assert "Observed baseline / acceptance" in text
    assert "matching that degraded state after cutover is not success" in text
    assert "known-good output" not in text
