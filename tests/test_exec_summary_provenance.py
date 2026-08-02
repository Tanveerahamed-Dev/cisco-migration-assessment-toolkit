"""P3-E3: the Executive Summary sheet carries a Document-control / provenance section (engine version,
generation time, snapshot identity) so the one-page landing sheet is self-traceable to the exact engine
+ snapshot that produced it -- the workbook twin of the deck title-footer. Absent fields are OMITTED,
never fabricated; passing no provenance renders no section (backward compatible with existing callers)."""
from openpyxl import Workbook

from cisco_toolkit.excel import EXEC_SUMMARY_SHEET_NAME, write_executive_summary_sheet

_HS = [{"band": "Good", "score": 90}]


def _cells(ws):
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def test_provenance_section_renders_when_supplied():
    wb = Workbook()
    write_executive_summary_sheet(
        wb, _HS, [], [], [],
        provenance={"script_version": "V3.23.0",
                    "generated_at": "2026-07-11T18:30:00.123456",
                    "snapshot": "MERIDIAN_Assessment"})
    cells = _cells(wb[EXEC_SUMMARY_SHEET_NAME])
    assert "Document control" in cells
    assert "V3.23.0" in cells
    assert "2026-07-11 18:30:00" in cells        # ISO 'T' -> space, trimmed to seconds
    assert "MERIDIAN_Assessment" in cells


def test_absent_fields_are_omitted_not_fabricated():
    wb = Workbook()
    write_executive_summary_sheet(wb, _HS, [], [], [],
                                  provenance={"script_version": "V3.23.0"})  # version only
    cells = _cells(wb[EXEC_SUMMARY_SHEET_NAME])
    assert "Document control" in cells and "V3.23.0" in cells
    assert "Snapshot" not in cells and "Generated" not in cells   # empty fields -> rows omitted


def test_no_provenance_arg_renders_no_section():
    wb = Workbook()
    write_executive_summary_sheet(wb, _HS, [], [], [])            # legacy call: no provenance
    assert "Document control" not in _cells(wb[EXEC_SUMMARY_SHEET_NAME])


def test_blank_provenance_values_render_no_section():
    wb = Workbook()
    write_executive_summary_sheet(wb, _HS, [], [], [],
                                  provenance={"script_version": "", "generated_at": None, "snapshot": ""})
    assert "Document control" not in _cells(wb[EXEC_SUMMARY_SHEET_NAME])
