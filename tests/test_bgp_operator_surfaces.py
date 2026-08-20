"""Configured-BGP-peer operator projections.

The core receipt owns classification.  These tests pin the presentation contract:
uncapped workbook rows, blocker retention outside bounded prose/UI budgets, exact
scope/custody copy, configured-only Explorer hosts, and no raw-secret projection.
"""

from __future__ import annotations

import json
import pathlib
import shutil
import subprocess

import pytest
from openpyxl import Workbook

from cisco_toolkit.excel import (
    BGP_PEER_INTENT_SHEET_NAME,
    write_bgp_configured_peer_sheet,
)


ROOT = pathlib.Path(__file__).resolve().parents[1]
EXPLORER = ROOT / "cisco_toolkit" / "blast_radius_explorer.html"
NODE = shutil.which("node")
COVERAGE_STATUS_ORDER = (
    "degraded", "review", "not_verified", "assessed", "not_applicable",
)


def _row(index: int, status: str = "assessed") -> dict:
    peer = f"192.0.2.{index + 1}"
    established = status == "assessed"
    return {
        "switch": f"edge-{index % 4}",
        "peer": peer,
        "peer_key": peer,
        "scope": "default/ipv4-unicast",
        "local_as": "65000",
        "configured_remote_as": str(65100 + index),
        "activation": "active",
        "runtime_observed": established,
        "runtime_remote_as": str(65100 + index) if established else "",
        "runtime_state_raw": "17" if established else "",
        "runtime_state": "ESTABLISHED" if established else "NOT OBSERVED",
        "status": status,
        "command": "show ip bgp summary",
        "acceptance": (
            "Remain Established in the bounded peer denominator."
            if established
            else "PRE-CUTOVER DEGRADED — BLOCKER: configured-active peer is absent from the usable summary."
        ),
        "source_key": f"bgp_configured_peer_baseline.rows[{index}]",
        "projection_custody": "current_run_source_bound",
        "findings": [],
    }


def _coverage_cell(
    switch: str,
    *,
    subject: bool,
    status: str,
    config_capture_status: str = "ok",
    config_parser_status: str = "complete",
    runtime_capture_status: str = "ok",
    runtime_parser_status: str = "complete",
) -> dict:
    """Presentation fixture with safe typed host detail and sensitive non-rendered leaves."""

    return {
        "switch": switch,
        "platform": "IOS-XE",
        "subject": subject,
        "status": status,
        "config_command": "show running-config",
        "config_capture_status": config_capture_status,
        "config_parser_status": config_parser_status,
        "runtime_command": "show ip bgp summary",
        "runtime_capture_status": runtime_capture_status,
        "runtime_parser_status": runtime_parser_status,
        "bgp_stanza_count": int(subject),
        "neighbor_candidate_count": int(subject),
        "supported_peer_count": int(subject),
        "rejected_candidate_count": 0,
        "excluded_scope_count": 0,
        "unsupported_relevant_count": 0,
        "runtime_candidate_count": int(subject),
        "runtime_parsed_count": int(subject),
        "runtime_rejected_count": 0,
        "runtime_local_as": "65000" if subject else "",
        "config_sha256": "a" * 64,
        "runtime_sha256": "b" * 64,
        "projection_sha256": "c" * 64,
        "finding_codes": [],
        "raw_path": r"C:\sensitive\raw-running-config.txt",
        "debug": "community coverage-row-secret",
    }


def _baseline(rows: list[dict], verdict: str = "BLOCKED") -> dict:
    by_status = {
        status: sum(row["status"] == status for row in rows)
        for status in ("assessed", "degraded", "review", "not_verified", "administratively_disabled")
    }
    n_disabled = by_status["administratively_disabled"]
    coverage_status = {
        "BLOCKED": "degraded",
        "INDETERMINATE": "review",
        "CLEAR": "assessed",
        "NOT_APPLICABLE": "not_applicable",
    }[verdict]
    coverage = [
        _coverage_cell(switch, subject=True, status=coverage_status)
        for switch in sorted({row["switch"] for row in rows})
    ]
    by_coverage_status = {
        status: sum(cell["status"] == status for cell in coverage)
        for status in COVERAGE_STATUS_ORDER
    }
    return {
        "schema": "bgp_configured_peer_baseline/1",
        "scope": {
            "routing_instance": "default",
            "afi": "ipv4",
            "safi": "unicast",
            "peer_kind": "direct_static_literal",
        },
        "verdict": verdict,
        "assessed": True,
        "projection_custody": "current_run_source_bound",
        "rows": rows,
        "coverage": coverage,
        "findings": [],
        "summary": {
            "n_hosts": len({row["switch"] for row in rows}),
            "n_subject_hosts": len({row["switch"] for row in rows}),
            "n_configured_peers": len(rows),
            "n_active_peers": len(rows) - n_disabled,
            "n_established": by_status["assessed"],
            "n_degraded": by_status["degraded"],
            "n_review": by_status["review"],
            "n_not_verified": by_status["not_verified"],
            "n_disabled": n_disabled,
            "by_status": by_status,
            "by_coverage_status": by_coverage_status,
            "baseline_sha256": "0" * 64,
        },
        "limitations": ["bounded fixture"],
    }


def _rowless_coverage_baseline(n_hosts: int = 23) -> dict:
    coverage = [
        _coverage_cell(
            f"coverage-{index:02d}",
            subject=False,
            status="not_verified",
            config_capture_status="incomplete" if index < 18 else "inspection_missing",
            config_parser_status="not_verified",
            runtime_capture_status="not_observed",
            runtime_parser_status="not_verified",
        )
        for index in range(n_hosts)
    ]
    baseline = _baseline([], verdict="NOT_APPLICABLE")
    baseline["assessed"] = False
    baseline["projection_custody"] = "embedded_unverified"
    baseline["coverage"] = coverage
    baseline["summary"].update({
        "n_hosts": n_hosts,
        "n_subject_hosts": 0,
        "by_coverage_status": {
            "degraded": 0,
            "review": 0,
            "not_verified": n_hosts,
            "assessed": 0,
            "not_applicable": 0,
        },
    })
    return baseline


def _sheet_text(ws) -> str:
    return "\n".join(
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    )


def _docx_text(path: pathlib.Path) -> str:
    docx = pytest.importorskip("docx")
    document = docx.Document(str(path))
    return "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )


def _explorer_section(baseline: dict, host: str | None = None) -> str:
    assert NODE is not None
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _BGP_INTENT_SCOPE=');
const end=html.indexOf('/* the dynamic routing protocols',start);
if(start<0||end<0)throw new Error('configured BGP projection block not found');
const input=JSON.parse(fs.readFileSync(0,'utf8'));
let SNAP={bgp_configured_peer_baseline:input.baseline};
function esc(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
process.stdout.write(bgpConfiguredPeerSection(input.host));
"""
    proc = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        input=json.dumps({"baseline": baseline, "host": host}),
        timeout=20,
    )
    return proc.stdout


def test_bgp_peer_intent_workbook_is_uncapped_blocker_visible_and_secret_safe():
    rows = [_row(index) for index in range(205)]
    blocker = _row(205, "degraded")
    blocker["peer"] = "203.0.113.254"
    blocker["peer_key"] = blocker["peer"]
    blocker["acceptance"] += " neighbor password visible-workbook-secret"
    blocker["raw_config"] = "username admin secret raw-workbook-secret"
    blocker["findings"] = [{
        "message": "Missing configured peer must be restored or dispositioned.",
        "debug": "snmp-server community hidden-workbook-secret ro",
    }]
    rows.append(blocker)

    wb = Workbook()
    write_bgp_configured_peer_sheet(wb, _baseline(rows))
    ws = wb[BGP_PEER_INTENT_SHEET_NAME]
    text = _sheet_text(ws)

    assert ws.max_row == 4 + len(rows), "the workbook has no display-row cap"
    assert ws.freeze_panes == "A5"
    assert "CONFIGURED BGP PEER GATE — DEFAULT/GLOBAL IPv4 UNICAST" in ws["A1"].value
    assert "A configured-active literal peer absent from a usable summary is a blocker" in ws["A1"].value
    assert "VRFs, IPv6, VPNv4/EVPN" in ws["A1"].value
    assert "203.0.113.254" in text and "PRE-CUTOVER DEGRADED — BLOCKER" in text
    assert "Peer rows — configured: 206" in text
    assert "Host coverage (distinct from peer rows): 4 host(s) · 4 subject host(s)" in text
    assert "current_run_source_bound" in text
    assert ws.cell(4 + len(rows), 1).fill.fgColor.rgb.endswith("F4CCCC")
    assert "visible-workbook-secret" not in text
    assert "raw-workbook-secret" not in text
    assert "hidden-workbook-secret" not in text
    assert "coverage-row-secret" not in text
    assert r"C:\sensitive\raw-running-config.txt" not in text
    assert "[REDACTED]" in text


def test_bgp_peer_intent_workbook_fails_closed_without_the_contract():
    wb = Workbook()
    write_bgp_configured_peer_sheet(wb, {"verdict": "CLEAR", "rows": []})
    text = _sheet_text(wb[BGP_PEER_INTENT_SHEET_NAME])
    assert "NOT ASSESSED" in text
    assert "No configured-peer completeness or health conclusion is asserted" in text
    assert "CLEAR" not in wb[BGP_PEER_INTENT_SHEET_NAME]["A1"].value


def test_rowless_not_applicable_workbook_exposes_host_coverage_without_peer_count_aliasing():
    wb = Workbook()
    write_bgp_configured_peer_sheet(wb, _rowless_coverage_baseline())
    ws = wb[BGP_PEER_INTENT_SHEET_NAME]
    text = _sheet_text(ws)

    assert "Peer rows — configured: 0" in ws["A2"].value
    assert "not verified: 0" in ws["A2"].value
    assert "Host coverage (distinct from peer rows): 23 host(s) · 0 subject host(s)" in ws["A3"].value
    assert "coverage cells — 0 degraded, 0 review, 23 not verified, 0 assessed" in ws["A3"].value
    assert (
        "NOT_APPLICABLE means no in-scope literal peer subject was identified; it is not proof "
        "that BGP is absent or that configuration coverage is complete."
    ) in ws["A3"].value
    assert ws["A3"].fill.fgColor.rgb.endswith("FFF2CC")
    assert "coverage-row-secret" not in text
    assert r"C:\sensitive\raw-running-config.txt" not in text
    assert "a" * 64 not in text


@pytest.mark.parametrize(
    "case",
    ("missing", "extra", "boolean", "negative", "non_reconciling"),
)
def test_bgp_peer_intent_workbook_fails_closed_on_malformed_coverage_census(case):
    baseline = _rowless_coverage_baseline(3)
    census = baseline["summary"]["by_coverage_status"]
    if case == "missing":
        census.pop("assessed")
    elif case == "extra":
        census["unknown"] = 0
    elif case == "boolean":
        census["degraded"] = True
    elif case == "negative":
        census["review"] = -1
        census["not_verified"] = 4
    else:
        census["review"] = 1
        census["not_verified"] = 2

    wb = Workbook()
    write_bgp_configured_peer_sheet(wb, baseline)
    ws = wb[BGP_PEER_INTENT_SHEET_NAME]
    assert "NOT ASSESSED" in _sheet_text(ws)
    assert "NOT_APPLICABLE" not in ws["A1"].value


def test_runbook_keeps_every_configured_bgp_blocker_outside_the_protocol_cap(tmp_path):
    docx = pytest.importorskip("docx")
    from cisco_toolkit.runbook import write_runbook_docx

    rows = [_row(index, "degraded") for index in range(19)]
    rows[-1]["peer"] = "203.0.113.219"
    rows[-1]["peer_key"] = rows[-1]["peer"]
    rows[-1]["acceptance"] += " neighbor password visible-runbook-secret"
    rows[-1]["raw_config"] = "enable secret raw-runbook-secret"
    rows[-1]["findings"] = [{"message": "last configured blocker", "debug": "community hidden-runbook-secret"}]
    snap = {
        "devices": {"edge-0": {"hostname": "edge-0"}},
        "bgp_configured_peer_baseline": _baseline(rows),
    }
    output = tmp_path / "bgp-runbook.docx"
    write_runbook_docx(str(output), snap, "Configured BGP test")
    document = docx.Document(str(output))
    text = "\n".join(
        [paragraph.text for paragraph in document.paragraphs]
        + [cell.text for table in document.tables for row in table.rows for cell in row.cells]
    )

    assert "Configured BGP peer gate — default/global IPv4 unicast" in text
    assert "Across 19 configured literal peers, 0 of 19 configured-active literal peers are Established" in text
    assert "Peer rows: 19 degraded, 0 require review, and 0 are not verified" in text
    assert "Host coverage (distinct from peer rows): 4 host(s) · 4 subject host(s)" in text
    assert "all rendered; these rows are outside the ordinary protocol display cap" in text
    assert "203.0.113.219" in text and "last configured blocker" in text
    assert "VRFs, IPv6, VPNv4/EVPN" in text
    assert "current_run_source_bound" in text
    assert "visible-runbook-secret" not in text
    assert "raw-runbook-secret" not in text
    assert "hidden-runbook-secret" not in text
    assert "coverage-row-secret" not in text
    assert r"C:\sensitive\raw-running-config.txt" not in text
    assert "[REDACTED]" in text


def test_rowless_not_applicable_runbook_exposes_amber_host_coverage_census(tmp_path):
    from cisco_toolkit.runbook import write_runbook_docx

    output = tmp_path / "bgp-rowless-runbook.docx"
    write_runbook_docx(
        str(output),
        {"bgp_configured_peer_baseline": _rowless_coverage_baseline()},
        "Rowless configured BGP coverage",
    )
    text = _docx_text(output)

    assert "Peer rows: 0 degraded, 0 require review, and 0 are not verified" in text
    assert "⚠ Host coverage (distinct from peer rows): 23 host(s) · 0 subject host(s)" in text
    assert "coverage cells — 0 degraded, 0 review, 23 not verified, 0 assessed" in text
    assert (
        "NOT_APPLICABLE means no in-scope literal peer subject was identified; it is not proof "
        "that BGP is absent or that configuration coverage is complete."
    ) in text
    assert "coverage-row-secret" not in text
    assert r"C:\sensitive\raw-running-config.txt" not in text
    assert "a" * 64 not in text


def test_runbook_fails_closed_when_host_coverage_map_is_missing(tmp_path):
    from cisco_toolkit.runbook import write_runbook_docx

    baseline = _rowless_coverage_baseline(3)
    baseline["summary"].pop("by_coverage_status")
    output = tmp_path / "bgp-malformed-map-runbook.docx"
    write_runbook_docx(
        str(output), {"bgp_configured_peer_baseline": baseline}, "Malformed BGP census",
    )
    text = _docx_text(output)
    assert "Verdict: NOT ASSESSED" in text
    assert "counts are unavailable" in text
    assert "NOT_APPLICABLE means" not in text


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_explorer_renders_every_blocker_then_first_50_ordinary_peers_and_scrubs_secrets():
    rows = [_row(index) for index in range(51)]
    degraded = _row(80, "degraded")
    degraded["peer"] = "203.0.113.80"
    degraded["peer_key"] = degraded["peer"]
    degraded["acceptance"] += " neighbor password visible-explorer-secret"
    degraded["raw_config"] = "enable secret raw-explorer-secret"
    review = _row(81, "review")
    review["peer"] = "203.0.113.81"
    review["peer_key"] = review["peer"]
    review["acceptance"] = "PRE-CUTOVER REVIEW — BLOCKER: remote AS reconciliation is ambiguous."
    rows.extend([degraded, review])
    payload = json.dumps(_baseline(rows))
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _BGP_INTENT_SCOPE=');
const end=html.indexOf('/* the dynamic routing protocols',start);
if(start<0||end<0)throw new Error('configured BGP projection block not found');
let SNAP={bgp_configured_peer_baseline:JSON.parse(fs.readFileSync(0,'utf8'))};
function esc(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
process.stdout.write(bgpConfiguredPeerSection());
"""
    proc = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        input=payload,
        timeout=20,
    )
    output = proc.stdout

    assert "Configured BGP peer gate — default/global IPv4 unicast" in output
    assert "53</div><div class=\"k\">peer configured" in output
    assert "51/53" in output
    assert "203.0.113.80" in output and "203.0.113.81" in output
    assert "192.0.2.50" in output, "the 50th ordinary peer must remain visible"
    assert "192.0.2.51" not in output, "the 51st ordinary peer is beyond the UI budget"
    assert "All 2 blocker row(s) are rendered" in output
    assert "50 of 51 ordinary peer row(s) are shown" in output
    assert "current_run_source_bound" in output
    assert "VRFs, IPv6, VPNv4/EVPN" in output
    assert "Host coverage (distinct from peer rows):</b> 4 host(s) · 4 subject host(s)" in output
    assert "visible-explorer-secret" not in output
    assert "raw-explorer-secret" not in output
    assert "coverage-row-secret" not in output
    assert r"C:\sensitive\raw-running-config.txt" not in output
    assert "[REDACTED]" in output


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_rowless_not_applicable_explorer_exposes_host_census_and_amber_qualification():
    output = _explorer_section(_rowless_coverage_baseline())

    assert "0</div><div class=\"k\">peer configured" in output
    assert "0/0" in output and "peer Established / active" in output
    assert "Host coverage (distinct from peer rows):</b> 23 host(s) · 0 subject host(s)" in output
    assert "coverage cells — 0 degraded, 0 review, 23 not verified, 0 assessed" in output
    assert "badge b-watch\">NOT APPLICABLE" in output
    assert (
        "NOT_APPLICABLE means no in-scope literal peer subject was identified; it is not proof "
        "that BGP is absent or that configuration coverage is complete."
    ) in output
    assert "inspect receipt coverage" not in output
    assert "coverage-row-secret" not in output
    assert r"C:\sensitive\raw-running-config.txt" not in output
    assert "a" * 64 not in output


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_explorer_fails_closed_when_host_coverage_map_does_not_reconcile():
    baseline = _rowless_coverage_baseline(3)
    baseline["summary"]["by_coverage_status"]["review"] = 1
    baseline["summary"]["by_coverage_status"]["not_verified"] = 2
    output = _explorer_section(baseline)
    assert "not assessed" in output
    assert "Peer-row and host-coverage counts are unavailable" in output
    assert "NOT APPLICABLE" not in output


@pytest.mark.skipif(NODE is None, reason="node is required for the Explorer projection contract")
def test_explorer_coverage_only_subject_is_drillable_without_a_peer_row():
    baseline = _baseline([], verdict="INDETERMINATE")
    baseline["assessed"] = False
    baseline["projection_custody"] = "embedded_unverified"
    cell = _coverage_cell(
        "coverage-only-edge",
        subject=True,
        status="review",
        config_parser_status="review",
        runtime_capture_status="not_observed",
        runtime_parser_status="not_verified",
    )
    cell.update({
        "supported_peer_count": 0,
        "runtime_candidate_count": 0,
        "runtime_parsed_count": 0,
        "unsupported_relevant_count": 1,
    })
    baseline["coverage"] = [cell]
    baseline["summary"].update({
        "n_hosts": 1,
        "n_subject_hosts": 1,
        "by_coverage_status": {
            "degraded": 0,
            "review": 1,
            "not_verified": 0,
            "assessed": 0,
            "not_applicable": 0,
        },
    })
    script = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const start=html.indexOf('const _BGP_INTENT_SCOPE=');
const end=html.indexOf('/* the dynamic routing protocols',start);
const detailStart=html.indexOf('function protocolDetailHosts(');
const detailEnd=html.indexOf('function drawProtocols()',detailStart);
const input=JSON.parse(fs.readFileSync(0,'utf8'));
let SNAP={bgp_configured_peer_baseline:input};
function esc(v){return String(v??'').replace(/[&<>"']/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));}
function shortName(v){return String(v??'');}
eval(html.slice(start,end));
eval(html.slice(detailStart,detailEnd));
process.stdout.write(JSON.stringify({
  hosts:protocolDetailHosts([],[],[],[]),
  section:bgpConfiguredPeerSection('coverage-only-edge')
}));
"""
    proc = subprocess.run(
        [NODE, "-e", script, str(EXPLORER)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        input=json.dumps(baseline),
        timeout=20,
    )
    result = json.loads(proc.stdout)
    assert result["hosts"] == ["coverage-only-edge"]
    assert "Host coverage detail" in result["section"]
    assert "coverage-only-edge" in result["section"]
    assert "SUBJECT" in result["section"] and "REVIEW" in result["section"]
    assert "No configured-peer rows were published for this host" in result["section"]
    assert "coverage-row-secret" not in result["section"]


def test_explorer_configured_only_host_drilldown_and_observed_only_compare_are_explicit():
    html = EXPLORER.read_text(encoding="utf-8")
    model = html[html.index("function buildModel(snap)"):html.index("function linkCarries(")]
    detail_hosts = html[
        html.index("function protocolDetailHosts("):html.index("function drawProtocols()")
    ]
    compare = html[
        html.index("function protocolCompareSection(p)"):html.index("function drawCompare()")
    ]

    assert "bgpIntentRows" in model and ".map(r=>r.switch)" in model
    assert "bgpCoverageSubjects" in model and "bgpIntent.coverage.filter(row=>row.subject===true)" in model
    assert "Object.keys(ifaces[host]||{})" in model
    assert "bgpConfiguredPeerRows()" in detail_hosts and "hosts.add(row.switch)" in detail_hosts
    assert "bgpConfiguredPeerCoverageRows()" in detail_hosts and "row.subject===true" in detail_hosts
    assert "bgpConfiguredPeerSection(host)" in html
    assert "bgpConfiguredPeerRows().filter(bgpIntentBlocker)" in html
    assert "bgpConfiguredPeerCoverageRows(host).some(row=>row.subject===true)" in html
    assert '["BGP CONFIGURED PEER NOT VERIFIED — BLOCKER:","not_verified"]' in html
    assert "BGP CONFIGURED PEER NOT VERIFIED" in html[html.index("function wavesValidationBlock("):]
    assert "observed-peer-only" in compare
    assert "current-baseline gate above still carries any typed configured-peer blocker" in compare
    assert "No expected peers are inferred here" in compare
