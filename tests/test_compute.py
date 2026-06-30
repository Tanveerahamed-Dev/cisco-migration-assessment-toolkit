"""Unit tests for the intelligence-layer pure functions:
compute_move_groups, compute_cross_layer_correlations, compute_health_scores,
compute_migration_readiness, and compute_protocol_health.

Inputs are constructed directly (not fixture-derived) so each rule/cap/verdict
is exercised in isolation and the assertions are unambiguous.
"""


def _dep(**over):
    """Full dependency-map skeleton with empty defaults; override per test."""
    d = {
        "single_fiber": set(), "uplink_ports": set(), "sole_gw": {},
        "access_by_vlan": {}, "articulation": set(), "fhrp_vlans": set(),
        "tracked_down": set(), "errored_up": set(), "halfdup_up": set(),
        "single_member_pc": set(), "errdis": set(), "gw_switches": set(),
        "orphan": set(), "model": {"hosts": set()},
    }
    d.update(over)
    return d


# --------------------------------------------------------------------------- #
# compute_move_groups
# --------------------------------------------------------------------------- #
def test_move_groups_shared_vlan_couples_switches(cp):
    def access(port, vlan):
        d = cp.InterfaceData(port=port)
        d.switchport_mode = "Access"; d.vlan = vlan
        return d

    all_interfaces = {
        "swA": {"Gi0/1": access("Gi0/1", "10"), "Vlan10": cp.InterfaceData(port="Vlan10")},
        "swB": {"Gi0/1": access("Gi0/1", "10")},
        "swC": {"Gi0/1": access("Gi0/1", "99")},  # isolated VLAN -> own group
    }
    groups = cp.compute_move_groups(all_interfaces)
    sets = [set(g["switches"]) for g in groups]
    assert {"swA", "swB"} in sets
    assert {"swC"} in sets


def test_move_groups_excludes_vlan1(cp):
    def access(port, vlan):
        d = cp.InterfaceData(port=port)
        d.switchport_mode = "Access"; d.vlan = vlan
        return d
    # VLAN 1 must NOT couple switches (it would collapse everything into one group)
    all_interfaces = {
        "swA": {"Gi0/1": access("Gi0/1", "1")},
        "swB": {"Gi0/1": access("Gi0/1", "1")},
    }
    groups = cp.compute_move_groups(all_interfaces)
    assert sorted(set(g["switches"][0] for g in groups)) == ["swA", "swB"]
    assert len(groups) == 2


# --------------------------------------------------------------------------- #
# compute_cross_layer_correlations
# --------------------------------------------------------------------------- #
def test_cross_layer_single_fiber_to_sole_gateway(cp):
    dep = _dep(
        sole_gw={30: "core1"},
        access_by_vlan={30: {"access1"}},
        single_fiber={("access1", "Gi0/1")},
        gw_switches={"core1"},
        model={"hosts": {"core1", "access1"}},
    )
    out = cp.compute_cross_layer_correlations(dep)
    by_id = {f["id"]: f for f in out}
    # CL-01: single-fiber uplink fronting a sole gateway = Critical
    assert "CL-01" in by_id and by_id["CL-01"]["severity"] == "Critical"
    assert set(by_id["CL-01"]["hosts"]) >= {"core1", "access1"}
    # CL-03: sole gateway, no FHRP = High
    assert "CL-03" in by_id and by_id["CL-03"]["severity"] == "High"


def test_cross_layer_orphan_vlan(cp):
    dep = _dep(orphan={40}, access_by_vlan={40: {"acc"}}, model={"hosts": {"acc"}})
    out = cp.compute_cross_layer_correlations(dep)
    by_id = {f["id"]: f for f in out}
    assert "CL-10" in by_id and by_id["CL-10"]["severity"] == "Medium"


def test_cross_layer_clean_dep_yields_nothing(cp):
    assert cp.compute_cross_layer_correlations(_dep()) == []


# --------------------------------------------------------------------------- #
# NEW-V3.23.90: per-VLAN finding-explosion aggregation (cry-wolf fix). An
# articulation switch / single-member-PC switch must yield ONE row per DEVICE,
# not one per (host, VLAN) -- the bug that put 13k near-identical rows into the
# real-fleet snapshot/explorer/punch-list and saturated the health-score XL cap.
# --------------------------------------------------------------------------- #
def test_cross_layer_articulation_aggregates_per_host(cp):
    dep = _dep(articulation={("swA", 10), ("swA", 20), ("swA", 30), ("swB", 40)},
               model={"hosts": {"swA", "swB"}})
    cl02 = [f for f in cp.compute_cross_layer_correlations(dep) if f["id"] == "CL-02"]
    assert len(cl02) == 2                                   # one per host, NOT four (was per-VLAN)
    a = next(f for f in cl02 if f["hosts"] == ["swA"])
    assert "3 VLAN(s)" in a["title"]                        # the VLAN COUNT is surfaced
    assert "10, 20, 30" in a["detail"]                      # and the VLANs are listed
    assert "1 VLAN(s)" in next(f for f in cl02 if f["hosts"] == ["swB"])["title"]


def test_cross_layer_single_member_pc_aggregates_per_host(cp):
    dep = _dep(single_member_pc={("swA", "Po1"), ("swA", "Po2"), ("swB", "Po3")},
               uplink_ports={("swA", "Te1/1"), ("swB", "Te1/1")},
               model={"hosts": {"swA", "swB"}})
    cl06 = [f for f in cp.compute_cross_layer_correlations(dep) if f["id"] == "CL-06"]
    assert len(cl06) == 2                                   # one per host, NOT three
    a = next(f for f in cl06 if f["hosts"] == ["swA"])
    assert "2 single-member port-channel(s)" in a["title"]
    assert "Po1" in a["detail"] and "Po2" in a["detail"]


def test_causality_aggregates_transit_and_uplink_per_host():
    """G --- T --- A line: T is the only path from A's endpoints to gateway G for VLANs 10 & 20.
    Removing T must yield ONE Chain-B row (per device) listing both VLANs, not one per VLAN; A's
    single uplink must yield ONE Chain-C row. Every trigger is unique (the aggregation invariant)."""
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit.analyze import compute_causality_chains

    def trunk(p, nb, npt):
        return InterfaceData(port=p, cdp_neighbor=nb, neighbor_port=npt,
                             endpoint_type="Switch", trunk_allowed_vlans="10,20")

    def ep(p, vlan, mac):
        return InterfaceData(port=p, switchport_mode="Access", vlan=vlan, end_host_mac=mac)

    ai = {
        "G": {"Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.1"), "Vlan20": InterfaceData(port="Vlan20", svi_ip="10.0.20.1"),
              "Gi0/1": trunk("Gi0/1", "T", "Gi0/1")},
        "T": {"Gi0/1": trunk("Gi0/1", "G", "Gi0/1"), "Gi0/2": trunk("Gi0/2", "A", "Gi0/24")},
        "A": {"Gi0/24": trunk("Gi0/24", "T", "Gi0/2"),
              "Gi0/1": ep("Gi0/1", "10", "aaaa.0000.0001"),
              "Gi0/2": ep("Gi0/2", "20", "aaaa.0000.0002")},
    }
    chains = compute_causality_chains(ai)
    triggers = [c[1] for c in chains]
    assert len(triggers) == len(set(triggers))             # aggregation invariant: no dup (host,*) rows
    transit = [c for c in chains if c[1] == "Transit switch T is removed / migrated"]
    assert len(transit) == 1                               # ONE row for T, not one per VLAN
    assert transit[0][0] == "High" and "2 VLAN(s)" in transit[0][2] and "10, 20" in transit[0][2]
    assert transit[0][5] == ("T",)                         # V3.23.92: explicit host tuple, not parsed from prose
    uplink = [c for c in chains if c[1].startswith("Uplink A ")]
    assert len(uplink) == 1 and "2 VLAN(s)" in uplink[0][2]
    assert uplink[0][5] == ("A",)


# --------------------------------------------------------------------------- #
# compute_health_scores
# --------------------------------------------------------------------------- #
def test_health_clean_host_is_perfect(cp):
    recs = cp.compute_health_scores({"clean": {}}, [], [], [], [])
    assert recs == [{"switch": "clean", "score": 100, "band": "Excellent",
                     "role": "access", "criticality": 1.0, "deductions": []}]


def test_health_l1_category_is_capped(cp):
    # 5 err-disabled ports * 8 = 40, but L1 cap is 30 -> score 70 (Fair), not 60.
    ph = [{"switch": "h", "port": f"Gi0/{i}", "risk": "err-disabled"} for i in range(5)]
    recs = cp.compute_health_scores({"h": {}}, ph, [], [], [])
    assert recs[0]["score"] == 70
    assert recs[0]["band"] == "Fair"


def test_health_cross_layer_critical_weight(cp):
    xl = [{"id": "CL-01", "severity": "Critical", "hosts": ["h"]}]
    recs = cp.compute_health_scores({"h": {}}, [], [], xl, [])
    assert recs[0]["score"] == 82          # 100 - 18
    assert recs[0]["band"] == "Good"


def test_health_scores_spread_across_bands(cp):
    all_if = {"good": {}, "mid": {}, "bad": {}}
    ph = [{"switch": "mid", "port": "Gi0/1", "risk": "single-fiber-uplink"}]   # -10
    l3 = [{"switch": "bad", "vlan": 30, "risk": "single-gateway"}]             # -10
    xl = [{"id": "CL-09", "severity": "Critical", "hosts": ["bad"]},           # -18
          {"id": "CL-01", "severity": "Critical", "hosts": ["bad"]}]           # -18 (XL cap 45)
    pr = [{"switch": "mid", "protocol": "VTP", "severity": "Medium"},          # -4  -> mid 86 (Good)
          {"switch": "bad", "protocol": "OSPF", "severity": "High"}]           # -10
    recs = {r["switch"]: r for r in cp.compute_health_scores(all_if, ph, l3, xl, pr)}
    assert recs["good"]["band"] == "Excellent"      # 100
    assert recs["mid"]["band"] == "Good"            # 100-14 = 86
    assert recs["bad"]["band"] == "Poor"            # 100-(10+36+10) = 44
    bands = {r["band"] for r in recs.values()}
    assert len(bands) == 3                           # genuine spread, not all-Critical


def test_health_security_deductions_capped(cp):
    # NEW-V3.23.60: a FAILED CIS config-hardening check deducts via sec_weights (high 8 / med 3 / low 1);
    # pass/info/na are ignored. {"h": {}} has no SVI -> access role (x1.0), so no criticality scaling.
    sec = {"h": {"findings": [
        {"id": "telnet-enabled", "severity": "high", "status": "fail"},       # -8
        {"id": "insecure-snmp", "severity": "high", "status": "fail"},        # -8
        {"id": "no-ntp", "severity": "low", "status": "fail"},                # -1
        {"id": "password-encryption", "severity": "info", "status": "pass"},  # ignored
    ]}}
    base = cp.compute_health_scores({"h": {}}, [], [], [], [])[0]["score"]
    withsec = cp.compute_health_scores({"h": {}}, [], [], [], [], security=sec)[0]["score"]
    assert base == 100 and withsec == 100 - (8 + 8 + 1)                       # only fails count


def test_health_security_missing_data_not_penalized(cp):
    # a host with no captured run-config (absent from the security dict) gets NO SEC deduction --
    # missing posture must never be scored as 'bad' (distinct from present-and-hardened).
    recs = cp.compute_health_scores({"x": {}}, [], [], [], [], security={})
    assert recs[0]["score"] == 100


def test_health_security_category_capped(cp):
    # SEC is capped at 18: six high-severity fails (6 * 8 = 48) deduct only -18.
    sec = {"h": {"findings": [{"id": f"c{i}", "severity": "high", "status": "fail"} for i in range(6)]}}
    rec = cp.compute_health_scores({"h": {}}, [], [], [], [], security=sec)[0]
    assert rec["score"] == 100 - 18


def test_score_sensitivity_sweeps_security_group(cp):
    sec = {"h": {"findings": [{"id": "telnet-enabled", "severity": "high", "status": "fail"}]}}
    rows = cp.compute_score_sensitivity({"h": {}}, [], [], [], [], security=sec)
    assert any(r["group"] == "sec_weights" for r in rows)


def test_stp_root_findings(cp):
    # VLAN30 is rooted on the default priority (32768+30) AND on a switch (acc1) that doesn't host its
    # gateway (core1 does) -> accidental + misaligned. VLAN10's root (core1) is deliberate + hosts the gateway.
    from cisco_toolkit import analyze
    from cisco_toolkit.model import InterfaceData
    stp = {"core1": {"10": {"is_root": True, "root_priority": 24586},
                     "30": {"is_root": False, "root_priority": 32798}},
           "acc1":  {"10": {"is_root": False, "root_priority": 24586},
                     "30": {"is_root": True, "root_priority": 32798}}}
    ifaces = {"core1": {"Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.1"),
                        "Vlan30": InterfaceData(port="Vlan30", svi_ip="10.0.30.1")},
              "acc1": {}}
    f = analyze.stp_root_findings(stp, ifaces)
    assert any(a["vlan"] == "30" and a["host"] == "acc1" for a in f["accidental"])   # default-priority root
    assert not any(a["vlan"] == "10" for a in f["accidental"])                       # 24586 is deliberate
    assert any(m["vlan"] == "30" and m["root"] == "acc1" and "core1" in m["gateways"]
               for m in f["misaligned"])                                             # root acc1 != gateway core1
    assert not any(m["vlan"] == "10" for m in f["misaligned"])                       # root core1 hosts the gateway


def test_compute_migration_punchlist(cp):
    # the consolidated roll-up pulls from every source, ranks Critical->Low, tags the wave, and
    # excludes non-fail security findings.
    from cisco_toolkit import analyze
    cross_layer = [{"severity": "Critical", "hosts": ["a"], "id": "CL-01", "title": "stacked SPOF", "detail": "x"}]
    security = {"a": {"findings": [
        {"id": "telnet-enabled", "severity": "high", "status": "fail",
         "title": "VTY telnet", "detail": "y", "remediation": "ssh"},
        {"id": "no-ntp", "severity": "low", "status": "pass"}]}}
    hygiene = {"a": {"undefined": [{"kind": "acl", "name": "7", "context": "nat list 7"}], "unused": []}}
    physical = [{"switch": "a", "risk": "err-disabled"}]
    l3 = [{"switch": "b", "risk": "single-gateway"}]
    proto = [{"switch": "a", "protocol": "OSPF", "severity": "High", "detail": "stuck"}]
    stp = {"accidental": [{"vlan": "30", "host": "b"}],
           "misaligned": [{"vlan": "30", "root": "b", "gateways": ["a"]}]}
    health = [{"switch": "b", "band": "Critical", "score": 20}]
    groups = [{"group": "Wave 1", "switches": ["a", "b"]}]
    pl = analyze.compute_migration_punchlist(cross_layer, security, hygiene, physical, l3,
                                             proto, stp, health, groups)
    cats = {i["category"] for i in pl}
    assert {"Cross-layer", "Security", "Config hygiene", "L1", "L3", "Protocol", "STP", "Health"} <= cats
    assert pl[0]["severity"] == "Critical" and pl[0]["priority"] == 1                 # ranked, 1-based
    assert [i["rank"] for i in pl] == sorted((i["rank"] for i in pl), reverse=True)   # Critical -> Low
    assert not any("no-ntp" in (i.get("title") or "") for i in pl)                    # pass findings excluded
    assert all(i["wave"] == "Wave 1" for i in pl)                                     # tagged from the move-group
    t = [i for i in pl if i["category"] == "Security" and "VTY" in i["title"]][0]
    assert t["remediation"] == "ssh" and t["severity"] == "High"
    # V3.23.64: the cross-switch L2 checks fold in via the `l2` arg (addressing / FHRP / trunk / link)
    l2 = {"addressing": {"dup_ip": [{"ip": "10.0.0.1", "where": [("a", "Vlan10", 10), ("b", "Vlan20", 20)]}],
                         "dup_subnet": []},
          "fhrp": [{"vid": 20, "issues": ["different FHRP groups"], "members": [{"host": "a"}, {"host": "b"}]}],
          "trunk_native": [{"a_host": "a", "a_port": "Gi0/1", "a_native": "1",
                            "b_host": "b", "b_port": "Gi0/2", "b_native": "99"}],
          "link_phy": [{"a_host": "a", "a_port": "Gi0/3", "b_host": "b", "b_port": "Gi0/4",
                        "duplex": ("full", "half"), "speed": None}]}
    pl2 = analyze.compute_migration_punchlist(cross_layer, security, hygiene, physical, l3, proto,
                                              stp, health, groups, l2=l2)
    assert {"Addressing", "FHRP", "Trunk", "Link L1"} <= {i["category"] for i in pl2}
    assert len(pl2) == len(pl) + 4                                                    # exactly the 4 L2 items added
    assert all(i["wave"] == "Wave 1" for i in pl2 if i["category"] in ("Addressing", "FHRP", "Trunk", "Link L1"))
    # V3.23.68: inventory-name vs configured-hostname mismatches fold in via hostname_mismatches
    pl3 = analyze.compute_migration_punchlist(cross_layer, security, hygiene, physical, l3, proto,
                                              stp, health, groups,
                                              hostname_mismatches=[{"inventory": "a", "reported": "a-real"}])
    inv = [i for i in pl3 if i["category"] == "Inventory"]
    assert len(inv) == 1 and inv[0]["severity"] == "Medium"
    assert "a-real" in inv[0]["title"] and inv[0]["wave"] == "Wave 1"   # device 'a' is in Wave 1


def test_punchlist_source_command_provenance():
    """[W1-3 / SmartyMe teardown] each punchlist finding whose category has ONE backing show-command carries a
    grounded source_command (a real command the engine collects); COMPOSITE / multi-source categories (Cross-layer,
    Protocol, Health, Compound risk, False-health) carry NONE -- provenance where it genuinely exists, silence
    where it doesn't, so it never becomes a global 'every claim traced' overclaim (the binding critic constraint)."""
    from cisco_toolkit import analyze
    cross_layer = [{"severity": "Critical", "hosts": ["a"], "id": "CL-01", "title": "stacked SPOF", "detail": "x"}]
    security = {"a": {"findings": [{"id": "telnet", "severity": "high", "status": "fail",
                                    "title": "VTY telnet", "detail": "y", "remediation": "ssh"}]}}
    hygiene = {"a": {"undefined": [{"kind": "acl", "name": "7", "context": "nat list 7"}], "unused": []}}
    physical = [{"switch": "a", "risk": "err-disabled"}]
    l3 = [{"switch": "b", "risk": "single-gateway"}]
    proto = [{"switch": "a", "protocol": "OSPF", "severity": "High", "detail": "stuck"}]
    stp = {"accidental": [{"vlan": "30", "host": "b"}]}
    health = [{"switch": "b", "band": "Critical", "score": 20}]
    groups = [{"group": "Wave 1", "switches": ["a", "b"]}]
    pl = analyze.compute_migration_punchlist(cross_layer, security, hygiene, physical, l3, proto, stp, health, groups)
    cmd = {}
    for f in pl:
        cmd.setdefault(f["category"], f.get("source_command"))
    # single-source categories -> a grounded show-command
    assert cmd["Security"] == "show running-config"
    assert cmd["Config hygiene"] == "show running-config"
    assert cmd["STP"] == "show spanning-tree"
    assert cmd["L3"] == "show ip route"
    assert cmd["L1"] == "show interface status"
    # composite / multi-source -> NO source_command (coverage-honest; never a single fabricated source)
    assert cmd["Cross-layer"] is None
    assert cmd["Protocol"] is None
    assert cmd["Health"] is None
    # every command the map cites is one the engine actually COLLECTS (grounded provenance, never fabricated)
    import importlib
    cp = importlib.import_module("COLLECT_PARSE_V3_23_0")
    collected = set()
    for reg in ("COMMANDS_IOS", "COMMANDS_NXOS", "COMMANDS_IOSXR", "COMMANDS_ASA", "COMMANDS_NXOS_ACI", "COMMANDS_CLOUD"):
        collected |= set(getattr(cp, reg, []) or [])
    for c in set(analyze._PUNCH_SOURCE_COMMAND.values()):
        assert c in collected, f"cited source_command not in any COMMANDS_* registry: {c}"


def test_compute_hostname_mismatches(cp):
    from cisco_toolkit import analyze
    DP = analyze.DevicePhysical
    devs = [
        DP(hostname="AS01-BC", reported_hostname="AS01-BC"),                  # match -> no finding
        DP(hostname="AS08--BC-CR03R13", reported_hostname="AS08-BC-CR03R13"), # double-dash typo -> finding
        DP(hostname="CORE1", reported_hostname="core1.example.com"),          # FQDN/case -> canon-equal, no finding
        DP(hostname="AS09", reported_hostname=""),                            # unknown reported -> skipped
    ]
    out = analyze.compute_hostname_mismatches(devs)
    assert len(out) == 1
    assert out[0] == {"inventory": "AS08--BC-CR03R13", "reported": "AS08-BC-CR03R13"}


def test_reconcile_cdp_neighbor_resolves_configured_name_split_node():
    """[audit-5 cross-artifact #1] A device collected under a suffix-SHORTER inventory name than its OWN configured
    hostname (reported_hostname from show version) is advertised by its neighbors over CDP under that configured
    name, which canon-MISSES the inventory key -> the device renders as a phantom SPLIT node and its bidirectional
    link becomes 2 records. reconcile_cdp_neighbor_names rewrites those advertisements back to the inventory key
    using the device's OWN configured hostname. Over-merge-safe: a configured name claimed by >1 device (e.g. a FEX
    module reporting its parent's hostname) is NEVER merged (would corrupt every topology/blast-radius consumer)."""
    from cisco_toolkit import analyze
    from cisco_toolkit.model import InterfaceData
    DP = analyze.DevicePhysical
    # CORE collected as 'CORE-BC' but configured 'CORE-BC-AJDOH'; ACC advertises CORE over CDP by that configured name.
    all_if = {
        "CORE-BC": {"Gi1/0/1": InterfaceData(port="Gi1/0/1", cdp_neighbor="ACC-BC", neighbor_port="Gi1/0/24",
                                             endpoint_type="Switch", speed="1G")},
        "ACC-BC": {"Gi1/0/24": InterfaceData(port="Gi1/0/24", cdp_neighbor="CORE-BC-AJDOH", neighbor_port="Gi1/0/1",
                                             endpoint_type="Switch", speed="1G")},
    }
    devs = [DP(hostname="CORE-BC", reported_hostname="CORE-BC-AJDOH"),
            DP(hostname="ACC-BC", reported_hostname="ACC-BC")]
    before = analyze.compute_topology_links(all_if)
    hosts_before = {str(r["a_host"]) for r in before} | {str(r["b_host"]) for r in before}
    assert "CORE-BC-AJDOH" in hosts_before and len(before) == 2          # phantom split node + duplicate link record
    assert analyze.reconcile_cdp_neighbor_names(all_if, devs) == 1
    after = analyze.compute_topology_links(all_if)
    hosts_after = {str(r["a_host"]) for r in after} | {str(r["b_host"]) for r in after}
    assert "CORE-BC-AJDOH" not in hosts_after                            # phantom gone
    assert len(after) == 1                                               # bidirectional link is now ONE record
    # over-merge guard: a configured name claimed by TWO devices (a FEX reporting its parent) is NOT used
    fex_if = {
        "PARENT": {"Eth1/1": InterfaceData(port="Eth1/1", cdp_neighbor="NB", endpoint_type="Switch")},
        "FEX-101": {"Eth1/2": InterfaceData(port="Eth1/2", cdp_neighbor="PARENT-DCDOH", endpoint_type="Switch")},
    }
    fex_devs = [DP(hostname="PARENT", reported_hostname="PARENT-DCDOH"),
                DP(hostname="FEX-101", reported_hostname="PARENT-DCDOH")]   # both claim 'PARENT-DCDOH'
    assert analyze.reconcile_cdp_neighbor_names(fex_if, fex_devs) == 0      # ambiguous -> no rewrite, no false merge
    assert fex_if["FEX-101"]["Eth1/2"].cdp_neighbor == "PARENT-DCDOH"       # left untouched


# --------------------------------------------------------------------------- #
# write_executive_summary_sheet (V3.23.75) — one-page landing synthesis
# --------------------------------------------------------------------------- #
def test_compute_capacity_blanks_util_when_active_ports_unobserved(cp):
    """Coverage-honesty: a device whose active-port count was NOT observed must not read 0% utilization
    (active_ports None coerced to 0) — that would rank it first as 'most consolidation headroom' in the
    runbook §8 ranking. Util/free are blank when unknown; a genuine 0 (all ports free) is distinct."""
    from cisco_toolkit.excel import compute_capacity
    from cisco_toolkit.model import DevicePhysical
    out = {r["hostname"]: r for r in compute_capacity([
        DevicePhysical(hostname="obs", total_ports=48, active_ports=10),     # observed -> 20.8%
        DevicePhysical(hostname="unobs", total_ports=48, active_ports=None),  # not observed -> blank
    ])}
    assert out["obs"]["port_util"] == 20.8
    assert out["unobs"]["port_util"] == "" and out["unobs"]["free_ports"] == ""


def test_build_device_physical_emits_none_active_ports_when_status_unobserved(cp):
    """The companion to the above: compute_capacity handles a None active count, but the PIPELINE must
    actually PRODUCE one. build_device_physical must emit active_ports=None (not 0) when no physical
    port carries an observed link status — else that None-guard is dead code and a device whose port
    status was not collected (e.g. [HISTORY-REDACTED]'s DS17/AS01: chassis ports known, status uncollected) falsely
    reads 0% utilization / all-ports-free on the Capacity sheet. A genuine all-down switch (status
    observed, 0 up) stays a real 0."""
    from cisco_toolkit.build import build_device_physical
    from cisco_toolkit.model import InterfaceData
    blind = {f"GigabitEthernet1/0/{i}": InterfaceData(port=f"GigabitEthernet1/0/{i}", status="")
             for i in range(1, 5)}
    dp = build_device_physical("blind", "ios", {}, blind)
    assert dp.total_ports == 4 and dp.active_ports is None, (dp.total_ports, dp.active_ports)
    seen = {f"GigabitEthernet1/0/{i}": InterfaceData(port=f"GigabitEthernet1/0/{i}",
            status=("connected" if i <= 3 else "notconnect")) for i in range(1, 5)}
    dp2 = build_device_physical("seen", "ios", {}, seen)
    assert dp2.active_ports == 3, dp2.active_ports


def test_physical_health_surfaces_and_flags_output_side_errors(cp, tmp_path):
    """The output-side L1 counters (output errors / late collisions) — previously parsed then discarded —
    now reach the physical_health record AND feed the 'error-rate-high' dirty-port flag, so a port that is
    clean on the input side but errored on the output side is no longer silently healthy."""
    from openpyxl import Workbook
    from cisco_toolkit.excel import write_physical_health_sheet
    from cisco_toolkit.model import InterfaceData, DevicePhysical
    shint = tmp_path / "shint.txt"
    shint.write_text("GigabitEthernet1/0/9 is up, line protocol is up\n"
                     "  0 input errors, 0 CRC, 0 frame, 0 overrun, 0 ignored\n"
                     "  7 output errors, 2 late collision, 0 deferred\n", encoding="utf-8")
    recs = write_physical_health_sheet(
        Workbook(),
        {"sw1": {"Gi1/0/9": InterfaceData(port="Gi1/0/9", status="connected")}},
        {"sw1": {"show interfaces": str(shint)}},
        [DevicePhysical(hostname="sw1")])
    r = next(x for x in recs if x["port"] == "Gi1/0/9")
    assert r["output_errors"] == 7 and r["late_collisions"] == 2
    assert "error-rate-high" in r["risk"]      # output-side errors flag a dirty port (input side is clean)


def test_executive_summary_sheet_is_first_and_synthesizes(cp):
    from openpyxl import Workbook
    health_scores = [
        {"switch": "GW-CORE", "score": 0, "band": "Critical"},
        {"switch": "ACC1", "score": 80, "band": "Good"},
        {"switch": "ACC2", "score": 55, "band": "Fair"},
    ]
    punchlist = [
        {"severity": "Critical", "category": "Cross-layer", "devices": ["GW-CORE"]},
        {"severity": "High", "category": "FHRP", "devices": ["GW-CORE"]},
        {"severity": "Medium", "category": "STP", "devices": ["ACC1"]},
    ]
    migration_readiness = [{"group": "Group 1", "readiness": "NOT READY",
                            "switches": ["GW-CORE", "ACC1", "ACC2"], "endpoints": 2,
                            "n_fail": 1, "n_warn": 1}]
    wb = Workbook(); wb.active.title = "Pre-Existing"   # prove our sheet lands ahead of it
    # V3.23.91: the 5th arg is the precomputed failure_impact (keystones); empty here.
    cp.write_executive_summary_sheet(wb, health_scores, punchlist, migration_readiness, [])
    assert wb.sheetnames[0] == "Executive Summary"      # moved to the FRONT of the workbook
    ws = wb["Executive Summary"]
    text = "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)
    # the four synthesis sections + the consumed inputs are all present
    for needle in ("Executive Summary", "Fleet posture", "Migration punch-list",
                   "Keystone devices", "Where to start", "Cross-layer", "Group 1"):
        assert needle in text, needle
    assert "3" in text   # 3 switches assessed / 3 punch-list items


# --------------------------------------------------------------------------- #
# compute_migration_readiness
# --------------------------------------------------------------------------- #
def test_readiness_not_ready_on_hard_fails(cp):
    move_groups = [{"switches": ["a", "b"], "endpoints": 3}]
    health = [{"switch": "a", "band": "Critical"}, {"switch": "b", "band": "Excellent"}]
    dep = _dep(sole_gw={30: "a"}, model={"hosts": {"a", "b"}})
    xl = [{"id": "CL-01", "severity": "Critical", "hosts": ["a"]}]
    pr = [{"switch": "a", "protocol": "OSPF", "severity": "High"}]
    out = cp.compute_migration_readiness({}, move_groups, health, [], [], xl, pr, dep)
    assert out[0]["readiness"] == "NOT READY"
    assert out[0]["n_fail"] >= 1


def test_readiness_ready_when_clean(cp):
    move_groups = [{"switches": ["c"], "endpoints": 0}]
    health = [{"switch": "c", "band": "Excellent"}]
    out = cp.compute_migration_readiness({}, move_groups, health, [], [], [], [], _dep())
    assert out[0]["readiness"] == "READY"
    assert out[0]["n_fail"] == 0 and out[0]["n_warn"] == 0


def test_readiness_caution_on_warn_only(cp):
    # single-fiber uplink is a WARN (not a fail) -> CAUTION
    move_groups = [{"switches": ["a"], "endpoints": 1}]
    health = [{"switch": "a", "band": "Good"}]
    dep = _dep(single_fiber={("a", "Gi0/1")}, model={"hosts": {"a"}})
    out = cp.compute_migration_readiness({}, move_groups, health, [], [], [], [], dep)
    assert out[0]["readiness"] == "CAUTION"


# --------------------------------------------------------------------------- #
# compute_protocol_health (integration on the synthetic fixtures)
# --------------------------------------------------------------------------- #
def test_protocol_health_flags_down_ospf(cp, built):
    all_interfaces, all_cmd_to_files = built
    recs = cp.compute_protocol_health(all_interfaces, all_cmd_to_files)
    ospf = [r for r in recs if r["switch"] == "core1" and r["protocol"] == "OSPF"]
    assert ospf and ospf[0]["severity"] == "High"
    # core2 has no OSPF neighbors collected -> no OSPF row
    assert not [r for r in recs if r["switch"] == "core2" and r["protocol"] == "OSPF"]


def test_protocol_health_etherchannel_min_links_not_met_is_not_healthy(cp, tmp_path):
    """FALSE-HEALTH: an LACP bundle that is non-forwarding because minimum-links was not met shows its members
    with the 'M' flag ('not in use, minimum links not met'); 'f' = failed-to-allocate-aggregator is likewise
    non-forwarding. These were absent from the bad/hard sets, so a down min-links bundle read as healthy Info.
    The EtherChannel row must now be High. (Also pins the parser keeping the full flag token so a combined
    'RM' can't mask the 'M'.)"""
    ec = ("Flags:  D - down        P - bundled in port-channel\n"
          "        M - not in use, minimum links not met\n"
          "Group  Port-channel  Protocol    Ports\n"
          "1      Po1(SD)        LACP        Gi1/0/5(M)   Gi1/0/6(M)\n")
    fp = tmp_path / "show_etherchannel_summary.txt"
    fp.write_text(ec, encoding="utf-8")
    recs = cp.compute_protocol_health({"SW1": {}}, {"SW1": {"show etherchannel summary": str(fp)}})
    ec_rows = [r for r in recs if r["protocol"] == "EtherChannel"]
    assert ec_rows and ec_rows[0]["severity"] == "High", ec_rows
    assert "not bundled" in ec_rows[0]["summary"]


def test_protocol_health_fhrp_stuck_init_is_not_healthy(cp):
    """FALSE-HEALTH: an FHRP group stuck in a non-forwarding Init/Learn role (interface down, auth mismatch,
    or no peer) is a real first-hop-redundancy fault -- the FHRP row must escalate to Medium, not the old
    hardcoded healthy Info. An all-Standby device stays Info (a backup with zero local actives is normal and
    can't be told apart from a fault without the peer's view -- no cry-wolf)."""
    from cisco_toolkit.model import InterfaceData
    init = {"R1": {"Vlan10": InterfaceData(port="Vlan10", hsrp_behavior="HSRP grp 1 Init VIP 10.0.10.1")}}
    f1 = [r for r in cp.compute_protocol_health(init, {"R1": {}}) if r["protocol"] == "FHRP"]
    assert f1 and f1[0]["severity"] == "Medium", f1
    standby = {"R2": {"Vlan10": InterfaceData(port="Vlan10", hsrp_behavior="HSRP grp 1 Standby VIP 10.0.10.1")}}
    f2 = [r for r in cp.compute_protocol_health(standby, {"R2": {}}) if r["protocol"] == "FHRP"]
    assert f2 and f2[0]["severity"] == "Info", f2


# --------------------------------------------------------------------------- #
# Route-aware reachability: scope_routes / inscope_subnets (build-layer helpers)
# --------------------------------------------------------------------------- #
def test_scope_routes_keeps_relevant_drops_noise():
    from cisco_toolkit.build import scope_routes
    inscope = {"10.0.10.0/24", "10.0.20.0/24"}
    route_db = {
        "0.0.0.0/0":       {"entries": [{"prefix": "0.0.0.0/0",       "source": "static",    "next_hop": "10.0.0.254",  "out_intf": ""}]},
        "10.0.0.0/16":     {"entries": [{"prefix": "10.0.0.0/16",     "source": "static",    "next_hop": "10.0.30.254", "out_intf": ""}]},
        "10.0.10.0/24":    {"entries": [{"prefix": "10.0.10.0/24",    "source": "connected", "next_hop": "",            "out_intf": "Vlan10"}]},
        "10.0.10.2/32":    {"entries": [{"prefix": "10.0.10.2/32",    "source": "local",     "next_hop": "",            "out_intf": "Vlan10"}]},
        "192.168.99.0/24": {"entries": [{"prefix": "192.168.99.0/24", "source": "static",    "next_hop": "10.0.10.254", "out_intf": ""}]},
    }
    got = {r["prefix"] for r in scope_routes(route_db, inscope)}
    assert "0.0.0.0/0" in got           # default route is always relevant
    assert "10.0.0.0/16" in got         # supernet that covers an in-scope subnet
    assert "10.0.10.0/24" in got        # connected in-scope subnet (covers itself)
    assert "10.0.10.2/32" not in got    # host (/32 local) noise dropped
    assert "192.168.99.0/24" not in got # out-of-scope prefix dropped


def test_inscope_subnets_from_svis():
    from cisco_toolkit.build import inscope_subnets
    from cisco_toolkit.model import InterfaceData
    ifaces = {"sw1": {
        "Vlan10":  InterfaceData(port="Vlan10",  svi_ip="10.0.10.1 255.255.255.0"),
        "Vlan20":  InterfaceData(port="Vlan20",  svi_ip="10.0.20.1 255.255.255.0"),
        "Gi1/0/1": InterfaceData(port="Gi1/0/1", vlan="10"),   # access port, not an SVI -> ignored
    }}
    assert inscope_subnets(ifaces) == {"10.0.10.0/24", "10.0.20.0/24"}


def test_protocol_boundaries_sheet():
    # workbook surfacing of the protocol-to-protocol analysis: per device protocols + redistribution edges,
    # with a MUTUAL-redistribution risk flag (ospf<->bgp both ways).
    from openpyxl import Workbook
    from cisco_toolkit.excel import write_protocol_boundaries_sheet, PROTOCOL_BOUNDARIES_SHEET_NAME
    wb = Workbook()
    rn = {"core1": {"ospf": [{"neighbor": "1.1.1.1", "state": "FULL"}], "eigrp": [], "bgp": []}}
    rd = {"core1": [
        {"into_proto": "ospf", "into_id": "1", "from_proto": "bgp", "from_id": "65001", "route_map": "", "raw": ""},
        {"into_proto": "bgp", "into_id": "65001", "from_proto": "ospf", "from_id": "1", "route_map": "RM", "raw": ""},
    ]}
    write_protocol_boundaries_sheet(wb, rn, rd)
    ws = wb[PROTOCOL_BOUNDARIES_SHEET_NAME]
    assert [c.value for c in ws[1]] == ["Switch", "Protocols", "Redistribution (from -> into)", "Route-map(s)", "Mutual"]
    row = [c.value for c in ws[2]]
    assert row[0] == "core1"
    assert "OSPF" in row[1] and "BGP" in row[1]
    assert row[3] == "RM"
    assert row[4] == "BGP/OSPF"   # two-way ospf<->bgp -> mutual-redistribution risk flagged


def test_compute_addressing_conflicts():
    # workbook surfacing of the addressing-integrity finding: duplicate IP + overlapping subnet (same VRF),
    # while a cross-VRF overlap and a normal FHRP pair (same VLAN, different IPs) are NOT flagged.
    from cisco_toolkit.excel import compute_addressing_conflicts
    from cisco_toolkit.model import InterfaceData
    ifaces = {
        "sw1": {
            "Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.2 255.255.255.0"),
            "Vlan99": InterfaceData(port="Vlan99", svi_ip="10.0.99.2 255.255.255.0"),   # dup IP with sw2 Vlan99
            "Vlan50": InterfaceData(port="Vlan50", svi_ip="10.0.77.1 255.255.255.0"),   # overlap w/ Vlan51 (same VRF)
            "Vlan51": InterfaceData(port="Vlan51", svi_ip="10.0.77.2 255.255.255.0"),
            "Vlan60": InterfaceData(port="Vlan60", svi_ip="10.0.88.1 255.255.255.0", vrf="RED"),   # cross-VRF overlap -> NOT flagged
            "Vlan61": InterfaceData(port="Vlan61", svi_ip="10.0.88.2 255.255.255.0", vrf="BLUE"),
        },
        "sw2": {
            "Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.3 255.255.255.0"),   # FHRP pair (same VLAN, diff IP) -> NOT flagged
            "Vlan99": InterfaceData(port="Vlan99", svi_ip="10.0.99.2 255.255.255.0"),
        },
    }
    c = compute_addressing_conflicts(ifaces)
    dup_ips = {d["ip"] for d in c["dup_ip"]}
    overlaps = {d["net"] for d in c["dup_subnet"]}
    assert dup_ips == {"10.0.99.2"}                 # same physical IP on sw1+sw2 Vlan99
    assert overlaps == {"10.0.77.0/24"}             # two VLANs share a subnet in the same VRF
    assert "10.0.88.0/24" not in overlaps           # same subnet but different VRFs -> intentional, not flagged


def test_compute_fhrp_consistency():
    # workbook surfacing of the FHRP finding: a different-group VLAN flagged; a consistent VLAN not flagged.
    from cisco_toolkit.excel import compute_fhrp_consistency
    from cisco_toolkit.model import InterfaceData
    ifaces = {
        "core1": {
            "Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.2 255.255.255.0", hsrp_behavior="HSRP grp 10 Active VIP 10.0.10.1"),
            "Vlan20": InterfaceData(port="Vlan20", svi_ip="10.0.20.2 255.255.255.0", hsrp_behavior="HSRP grp 20 Active VIP 10.0.20.1"),
        },
        "core2": {
            "Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.3 255.255.255.0", hsrp_behavior="HSRP grp 10 Standby VIP 10.0.10.1"),
            "Vlan20": InterfaceData(port="Vlan20", svi_ip="10.0.20.3 255.255.255.0", hsrp_behavior="HSRP grp 21 Standby VIP 10.0.20.1"),  # wrong group
        },
    }
    rows = compute_fhrp_consistency(ifaces)
    flagged = {r["vid"] for r in rows}
    assert 20 in flagged and 10 not in flagged          # VLAN 20 grp 20 vs 21 -> flagged; VLAN 10 consistent
    v20 = [r for r in rows if r["vid"] == 20][0]
    assert any("different FHRP groups" in i for i in v20["issues"])


def test_compute_trunk_native_mismatches():
    # workbook surfacing of the trunk native-VLAN finding: a CDP link whose two ends disagree on native VLAN.
    from cisco_toolkit.excel import compute_trunk_native_mismatches
    from cisco_toolkit.model import InterfaceData
    ifaces = {
        "core1": {"Po1": InterfaceData(port="Po1", cdp_neighbor="core2", neighbor_port="Po1",
                                       switchport_mode="Trunk", trunk_native_vlan="99")},   # native 99…
        "core2": {"Po1": InterfaceData(port="Po1", cdp_neighbor="core1", neighbor_port="Po1",
                                       switchport_mode="Trunk")},                            # …vs default 1
    }
    rows = compute_trunk_native_mismatches(ifaces)
    assert len(rows) == 1
    assert {rows[0]["a_native"], rows[0]["b_native"]} == {"99", "1"}   # explicit-on-one-end still a mismatch


def test_compute_duplex_speed_mismatches():
    # workbook surfacing of the link L1 finding: a CDP link whose two ends report a different duplex.
    from cisco_toolkit.excel import compute_duplex_speed_mismatches
    from cisco_toolkit.model import InterfaceData
    ifaces = {
        "core1": {"Gi1/0/1": InterfaceData(port="Gi1/0/1", cdp_neighbor="acc1", neighbor_port="Gi0/1",
                                           switchport_mode="Trunk", duplex="Full-duplex", speed="1000Mb/s")},
        "acc1": {"Gi0/1": InterfaceData(port="Gi0/1", cdp_neighbor="core1", neighbor_port="Gi1/0/1",
                                        switchport_mode="Trunk", duplex="Half-duplex", speed="1000Mb/s")},
    }
    rows = compute_duplex_speed_mismatches(ifaces)
    assert len(rows) == 1
    assert rows[0]["duplex"] is not None and set(rows[0]["duplex"]) == {"full", "half"}
    assert rows[0]["speed"] is None        # same speed -> no false speed flag


def test_build_routing_neighbors(tmp_path):
    # protocol-to-protocol analysis: OSPF/EIGRP/BGP adjacencies parsed from already-collected output,
    # keeping the full state token so the explorer can tell a healthy (FULL) from a stuck (EXSTART) peer.
    from cisco_toolkit.build import build_routing_neighbors
    ospf = tmp_path / "ospf.txt"
    ospf.write_text(
        "Neighbor ID     Pri   State           Dead Time   Address         Interface\n"
        "10.0.99.2         1   FULL/DR         00:00:35    10.0.99.2       Port-channel1\n"
        "10.0.99.9         1   EXSTART/DROTHER 00:00:31    10.0.40.9       Vlan40\n",
        encoding="utf-8")
    rn = build_routing_neighbors({"show ip ospf neighbor": str(ospf)})
    assert len(rn["ospf"]) == 2
    states = {n["state"] for n in rn["ospf"]}
    assert "FULL/DR" in states and "EXSTART/DROTHER" in states   # healthy + stuck adjacency both captured
    assert rn["eigrp"] == [] and rn["bgp"] == []                 # protocols not running -> empty list, never absent


# --------------------------------------------------------------------------- #
# Scoring calibration: compute_calibration_report (fleet band-discrimination)
# --------------------------------------------------------------------------- #
def test_calibration_report_flags_poor_discrimination():
    from cisco_toolkit.analyze import compute_calibration_report
    # 10 switches all banded 'Good' -> the bands don't discriminate -> poor + a re-banding suggestion
    hs = [{"switch": f"sw{i}", "score": 76 + (i % 5), "band": "Good"} for i in range(10)]
    rep = compute_calibration_report(hs)
    assert rep["n"] == 10
    assert rep["modal_band"] == "Good" and rep["modal_pct"] == 100
    assert rep["discrimination"] == 0.0                 # all one band -> zero entropy
    assert rep["discrimination_quality"] == "poor"
    assert rep["suggested_bands"] is not None
    assert rep["suggested_bands"][-1]["threshold"] == 0  # bottom band always floors at 0
    assert all(0 <= s["threshold"] <= 100 for s in rep["suggested_bands"])


def test_calibration_report_good_discrimination_excludes_insufficient_data():
    from cisco_toolkit.analyze import compute_calibration_report
    # one switch per band -> well spread; the 'Insufficient Data' switch must be excluded from stats
    hs = [{"switch": "a", "score": 95, "band": "Excellent"},
          {"switch": "b", "score": 80, "band": "Good"},
          {"switch": "c", "score": 65, "band": "Fair"},
          {"switch": "d", "score": 45, "band": "Poor"},
          {"switch": "e", "score": 20, "band": "Critical"},
          {"switch": "f", "score": 88, "band": "Insufficient Data", "data_quality": 0.3}]
    rep = compute_calibration_report(hs)
    assert rep["n"] == 5                                 # Insufficient-Data switch excluded
    assert rep["discrimination"] == 1.0                  # uniform across all 5 bands -> max entropy
    assert rep["discrimination_quality"] == "good"
    assert rep["suggested_bands"] is None                # good discrimination -> no re-banding offered


def test_calibration_report_empty():
    from cisco_toolkit.analyze import compute_calibration_report
    rep = compute_calibration_report([])
    assert rep["n"] == 0 and rep["suggested_bands"] is None


def test_link_centrality_bridge_and_betweenness():
    """A hangs off B; B-C-D form a triangle (redundant). A-B is the only bridge; its betweenness
    (serves A-B, A-C, A-D = 3 pairs) is highest, and removing it severs {A} from {B,C,D} = 3 pairs.
    The triangle links are NOT bridges (each pair has an alternate equal path)."""
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit.analyze import compute_link_centrality

    def sw(*links):
        return {p: InterfaceData(port=p, cdp_neighbor=nb, neighbor_port=npt, endpoint_type="Switch")
                for (p, nb, npt) in links}

    ai = {
        "A": sw(("Gi0/1", "B", "Gi0/1")),
        "B": sw(("Gi0/1", "A", "Gi0/1"), ("Gi0/2", "C", "Gi0/1"), ("Gi0/3", "D", "Gi0/2")),
        "C": sw(("Gi0/1", "B", "Gi0/2"), ("Gi0/2", "D", "Gi0/1")),
        "D": sw(("Gi0/1", "C", "Gi0/2"), ("Gi0/2", "B", "Gi0/3")),
    }
    recs = compute_link_centrality(ai)
    by = {frozenset((r["a_host"], r["b_host"])): r for r in recs}
    ab = by[frozenset(("A", "B"))]
    assert ab["is_bridge"] is True
    assert ab["pairs_cut"] == 3            # {A} vs {B,C,D}
    assert ab["betweenness"] == 3.0        # on the shortest path of A-B, A-C, A-D
    assert ab["rank"] == 1                 # highest betweenness
    assert by[frozenset(("C", "D"))]["is_bridge"] is False   # triangle edge is redundant
    assert sum(1 for r in recs if r["is_bridge"]) == 1       # exactly one bridge


def test_link_centrality_empty_when_no_links():
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit.analyze import compute_link_centrality
    assert compute_link_centrality({"lonely": {"Gi0/1": InterfaceData(port="Gi0/1")}}) == []


def test_operational_drift_detects_and_aggregates():
    """NEW-V3.23.93: the false-health detector finds temp L2 bridges, PoE-fault on powered ports,
    native-VLAN-1 trunks, and multi-year uptime -- and AGGREGATES the bulk ones (one row + count),
    honouring the cry-wolf doctrine."""
    from cisco_toolkit.model import InterfaceData, DevicePhysical
    from cisco_toolkit.analyze import compute_operational_drift

    ai = {
        "core1": {
            "Te1/1": InterfaceData(port="Te1/1", cdp_neighbor="core2", description="##Temp L2 connection##"),
            "Gi1/0/5": InterfaceData(port="Gi1/0/5", description="*Robotics camera*", poe_status="Fault"),
            "Te1/2": InterfaceData(port="Te1/2", trunk_status="trunking", trunk_native_vlan="1"),
        },
        "acc1": {"Te1/1": InterfaceData(port="Te1/1", trunk_status="trunking", trunk_native_vlan="1"),
                 "Gi1/0/9": InterfaceData(port="Gi1/0/9", poe_status="Fault")},   # DET-poe-002: fault on a blank-desc port
    }
    dphys = [DevicePhysical(hostname="core1", uptime="10 years, 2 weeks"),
             DevicePhysical(hostname="acc1", uptime="5 days")]
    out = compute_operational_drift(ai, dphys)
    titles = [f["title"] for f in out]
    assert any("Temporary L2 bridge on core1" in t for t in titles)
    assert any("PoE fault on core1 (powered endpoint affected)" in t for t in titles)
    # DET-poe-002: a PoE fault on a port WITHOUT a powered-endpoint description is no longer silently dropped
    _acc_poe = next((f for f in out if "PoE fault on acc1" in f["title"]), None)
    assert _acc_poe is not None and "powered endpoint affected" not in _acc_poe["title"] and _acc_poe["severity"] == "Medium"
    # native VLAN 1 -> ONE aggregated finding across 2 trunks / 2 switches (not one row per trunk)
    nat = next(f for f in out if "Native VLAN 1" in f["title"])
    assert "2 inter-switch trunk(s)" in nat["title"] and set(nat["devices"]) == {"core1", "acc1"}
    # multi-year uptime -> ONE aggregated finding; acc1 (<3y) excluded
    up = next(f for f in out if "Multi-year uptime" in f["title"])
    assert "1 device" in up["title"] and "10 years" in up["title"] and up["devices"] == ["core1"]
    assert all(f["category"] == "False-health" for f in out)


def test_operational_drift_folds_into_punchlist(cp):
    """The drift findings reach the executive punch-list (so the runbook / explorer / exec summary
    surface them) via the new `drift` parameter."""
    drift = [{"severity": "High", "category": "False-health", "devices": ["core1"],
              "title": "Temporary L2 bridge on core1", "detail": "x", "remediation": "y"}]
    pl = cp.compute_migration_punchlist([], {}, {}, [], [], [], {}, [], [], drift=drift)
    assert any(i["category"] == "False-health" and "Temporary L2 bridge" in i["title"] for i in pl)


def test_link_centrality_pairs_cut_excludes_other_components():
    """NEW-V3.23.91: a bridge's pairs_cut is the product of the TWO components it actually separates,
    not size_a*(n-size_a) -- which over-counted switches in OTHER disconnected components AND was
    non-deterministic (tuple(frozenset) flips the BFS start side by hash order, so pairs_cut swung
    302<->22082 for one bridge on the real fleet). A--B bridge, B-C-D triangle, plus a separate E-F
    island: removing A-B severs {A} from {B,C,D} = 1*3 = 3; the unrelated island must NOT count."""
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit.analyze import compute_link_centrality

    def sw(*links):
        return {p: InterfaceData(port=p, cdp_neighbor=nb, neighbor_port=npt, endpoint_type="Switch")
                for (p, nb, npt) in links}

    ai = {
        "A": sw(("Gi0/1", "B", "Gi0/1")),
        "B": sw(("Gi0/1", "A", "Gi0/1"), ("Gi0/2", "C", "Gi0/1"), ("Gi0/3", "D", "Gi0/2")),
        "C": sw(("Gi0/1", "B", "Gi0/2"), ("Gi0/2", "D", "Gi0/1")),
        "D": sw(("Gi0/1", "C", "Gi0/2"), ("Gi0/2", "B", "Gi0/3")),
        "E": sw(("Gi0/1", "F", "Gi0/1")),     # separate island, unreachable from A/B/C/D
        "F": sw(("Gi0/1", "E", "Gi0/1")),
    }
    recs = compute_link_centrality(ai)
    ab = next(r for r in recs if {r["a_host"], r["b_host"]} == {"A", "B"})
    assert ab["is_bridge"] is True
    assert ab["pairs_cut"] == 3            # {A} x {B,C,D}; the {E,F} island excluded (old gave 5 or 9)


def test_wave_sequencing_classifies_cutover():
    """A dual-homed switch (>=2 uplinks) is make-before-break; a single-homed switch (1 uplink) is a hard
    cutover, and its endpoints are counted as at-risk during the window."""
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit.analyze import compute_wave_sequencing

    def link(p, nb, npt):
        return InterfaceData(port=p, cdp_neighbor=nb, neighbor_port=npt, endpoint_type="Switch")

    def ep(p, mac):
        return InterfaceData(port=p, switchport_mode="Access", vlan="10", end_host_mac=mac)

    ai = {
        "DIST1": {"Gi0/1": link("Gi0/1", "ACC1", "Gi0/24"), "Gi0/2": link("Gi0/2", "ACC2", "Gi0/24")},
        "ACC1": {"Gi0/24": link("Gi0/24", "DIST1", "Gi0/1"),
                 "Gi0/1": ep("Gi0/1", "aaaa.0000.0001"), "Gi0/2": ep("Gi0/2", "aaaa.0000.0002")},
        "ACC2": {"Gi0/24": link("Gi0/24", "DIST1", "Gi0/2"), "Gi0/1": ep("Gi0/1", "aaaa.0000.0003")},
    }
    seq = compute_wave_sequencing(ai, [{"switches": ["ACC1", "ACC2", "DIST1"]}])[0]
    assert seq["make_before_break"] == ["DIST1"]          # degree 2 -> dual-homed
    assert seq["hard_cutover"] == ["ACC1", "ACC2"]        # degree 1 -> single-homed
    assert seq["hard_cutover_endpoints"] == 3             # 2 on ACC1 + 1 on ACC2


def test_stp_root_findings_skips_mst_instances_not_vlan0(cp):
    """ANALY-02: parse_spanning_tree_root keys MST records by INSTANCE number, not VLAN id. stp_root_findings
    did `int(vlan)` and tested prio==32768+int(vlan), so MST instance 0 at default priority fired a phantom
    'accidental root' for a non-existent VLAN 0, while the gateway-misalignment join (keyed on real VLAN ids)
    was silently dead for MST. MST records must be skipped; PVST default-priority roots must still fire."""
    from cisco_toolkit import parse
    mst = ("MST0\n  Spanning tree enabled protocol mstp\n  Root ID    Priority    32768\n"
           "             Address     0011.2233.4455\n             This bridge is the root\n")
    rm = parse.parse_spanning_tree_root(mst)
    assert rm["0"]["is_mst"] is True
    assert cp.stp_root_findings({"SW-MST": rm}, {})["accidental"] == []           # no phantom 'VLAN 0'
    pvst = ("VLAN0010\n  Root ID    Priority    32778\n             Address     0011.2233.4455\n"
            "             This bridge is the root\n")
    rp = parse.parse_spanning_tree_root(pvst)
    acc = cp.stp_root_findings({"SW1": rp}, {})["accidental"]
    assert acc == [{"vlan": "10", "host": "SW1", "priority": 32778}]               # genuine PVST accidental still fires


def test_inscope_subnets_handles_nxos_slash_form_svi():
    """[multi-domain audit #9] NX-OS / IOS-XE SVIs use 'ip address 10.1.20.1/24' (slash form); inscope_subnets
    parsed only the IOS 'ip <addr> <mask>' space form and dropped the slash form -> route-scope coverage gap for
    NX-OS L3 segments."""
    from cisco_toolkit.build import inscope_subnets
    from cisco_toolkit.model import InterfaceData
    def one(addr):
        return {"sw": {"Vlan20": InterfaceData(port="Vlan20", svi_ip=addr)}}
    assert "10.1.20.0/24" in inscope_subnets(one("10.1.20.1 255.255.255.0"))   # IOS space form (regression)
    assert "10.1.20.0/24" in inscope_subnets(one("10.1.20.1/24"))              # NX-OS slash form (the fix)


def test_endpoint_2110_token_anchored_to_smpte_standard():
    """[audit-2 #2] the bare '2110' token matched ANY substring (room/rack/asset numbers 'RM2110','AP-2110'),
    misclassifying ordinary endpoints as Broadcast A/V at the first-match break. It must match only the
    SMPTE ST-2110 standard forms."""
    from cisco_toolkit.analyze import _classify_endpoint
    for desc in ("Dell-PC RM2110", "HP Printer Rm2110", "Avaya VoIP phone rm2110", "AP-2110-ceiling", "Desk-2110"):
        assert _classify_endpoint("", desc, "", "", False)[0] != "Broadcast A/V", desc
    for desc in ("ST-2110 encoder", "SMPTE 2110 gateway", "2110-20 video flow"):
        assert _classify_endpoint("", desc, "", "", False)[0] == "Broadcast A/V", desc


def test_wave_sequencing_endpoint_count_not_exceed_move_group_total():
    """[audit-2 L1] compute_wave_sequencing counted MACs on ANY Access port, but compute_move_groups (the SSOT
    group endpoint total) requires Access AND a numeric VLAN -> 'hard_cutover_endpoints' could EXCEED the group's
    own endpoint total (self-contradictory deliverables). Align the predicate so it never exceeds the SSOT."""
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit import analyze
    def acc(vlan, mac):
        d = InterfaceData(); d.switchport_mode = "Access"; d.vlan = vlan; d.end_host_mac = mac
        return d
    ifaces = {"sw1": {"Gi1/0/1": acc("10", "aaaa.bbbb.0001"), "Gi1/0/2": acc("", "aaaa.bbbb.0002")}}
    mg = analyze.compute_move_groups(ifaces)
    ws = analyze.compute_wave_sequencing(ifaces, mg)
    g_ep = mg[0]["endpoints"]
    assert sum(w["hard_cutover_endpoints"] for w in ws) <= g_ep and g_ep == 1


def test_find_bridges_iterative_long_chain_matches_bruteforce():
    """[audit-3 L4 totality] _find_bridges used recursive DFS -> RecursionError on a ~1000+-node simple path
    (long daisy-chain). Iterative now; verified against an independent brute-force bridge oracle."""
    from cisco_toolkit.analyze import _find_bridges
    def brute(adj):
        edges = {frozenset((u, v)) for u in adj for v in adj[u] if u != v}
        out = set()
        for e in edges:
            a, b = tuple(e)
            seen, stack = {a}, [a]
            while stack:
                x = stack.pop()
                for w in adj[x]:
                    if (x == a and w == b) or (x == b and w == a):
                        continue
                    if w not in seen:
                        seen.add(w); stack.append(w)
            if b not in seen:
                out.add(e)
        return out
    N = 1500
    chain = {str(i): set() for i in range(N)}
    for i in range(N - 1):
        chain[str(i)].add(str(i + 1)); chain[str(i + 1)].add(str(i))
    assert len(_find_bridges(chain)) == N - 1        # no RecursionError; every chain edge is a bridge
    cycle = {"a": {"b", "c"}, "b": {"a", "c"}, "c": {"a", "b"}}
    tadpole = {"a": {"b", "c"}, "b": {"a", "c"}, "c": {"a", "b", "d"}, "d": {"c", "e"}, "e": {"d"}}
    for g in (cycle, tadpole, chain):
        assert _find_bridges(g) == brute(g)


def test_exec_summary_sheet_no_fabricated_coverage_on_brief_crash():
    """[audit-3 #2 HIGH false-health] when compute_executive_brief raises, _run_phase wires {'_unavailable':True}.
    The Exec Summary sheet then fell back to the raw recompute -> claimed full collection coverage (n/n) + an
    inflated all-rows average (50 vs true 41). It must disclose 'unavailable', like the Architecture Review sheet."""
    from openpyxl import Workbook
    import cisco_toolkit.excel as X
    hs = [{"switch": "a", "band": "Critical", "score": 20}, {"switch": "b", "band": "Poor", "score": 40},
          {"switch": "c", "band": "Good", "score": 80},
          {"switch": "d", "band": "Insufficient Data", "score": 90}, {"switch": "e", "band": "Insufficient Data", "score": 90}]
    wb = Workbook(); X.harden_workbook(wb)
    X.write_executive_summary_sheet(wb, hs, [], [], [], brief={"_unavailable": True})
    ws = wb[X.EXEC_SUMMARY_SHEET_NAME]
    def _val(label):
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c not in (None, "")]
            if cells and str(cells[0]).strip() == label:
                return " ".join(str(c) for c in cells[1:])
        return None
    collected, avg = _val("Switches collected / inventoried"), _val("Average health score")
    assert collected is not None and "5 / 5" not in collected            # must NOT fabricate full coverage
    assert "unavailable" in collected.lower() or "—" in collected
    assert avg is None or "unavailable" in avg.lower() or "—" in avg     # no fabricated clean average on a crash


def test_readiness_sheet_discloses_endpoint_mac_is_per_switch_sum():
    """[audit-3 L3 scale-ssot] the Migration Readiness sheet printed a bare 'N endpoint-MAC(s)' per group whose
    sum across groups (5160 on [HISTORY-REDACTED]) exceeds the canonical distinct n_endpoints (5127) -- a multi-homed MAC counts
    once per switch. The dedicated move-group sheets disclose '(per-switch sum)'; this sheet must too."""
    from openpyxl import Workbook
    import cisco_toolkit.excel as X
    wb = Workbook(); X.harden_workbook(wb)
    readiness = [{"group": "Group 1", "switches": ["sw1", "sw2"], "endpoints": 40, "readiness": "READY",
                  "checks": [], "n_fail": 0, "n_warn": 0}]
    X.write_migration_readiness_sheet(wb, readiness)
    blob = " ".join(str(c) for ws in wb.worksheets for row in ws.iter_rows(values_only=True) for c in row if c)
    assert "40 endpoint-MAC(s)" in blob          # the per-group figure is present
    assert "per-switch sum" in blob              # ...disclosed as a per-switch sum, not the distinct census


def test_exec_brief_fleet_health_assessed_excludes_uncollected():
    """[audit-3 #6 false-health] 'N switch(es) assessed' counted the Insufficient-Data (uncollected) rows that the
    average excludes -> 'assessed' overclaimed coverage. It must be the scored count + disclose the uncollected."""
    from cisco_toolkit.analyze import compute_executive_brief
    hs = [{"switch": "a", "band": "Critical", "score": 20}, {"switch": "b", "band": "Poor", "score": 40},
          {"switch": "c", "band": "Good", "score": 80},
          {"switch": "d", "band": "Insufficient Data", "score": 90}, {"switch": "e", "band": "Insufficient Data", "score": 90}]
    ax = {a["axis"]: a for a in compute_executive_brief(health_scores=hs)["axes"]}["Fleet health"]
    assert "3 switch(es) assessed" in ax["detail"] and "5 switch(es) assessed" not in ax["detail"]
    assert "2 not collected" in ax["detail"]


def test_exec_brief_eol_axis_distinguishes_blind_from_verified_clean():
    """[audit-3 #5 false-health] a fully-unknown-model (blind) fleet reported the SAME 'Low / 0%' as a verified-
    clean fleet -- the unknowns diluted the denominator. % must be over assessable devices, blind -> Info."""
    from cisco_toolkit.analyze import compute_lifecycle_risk, compute_executive_brief
    blind = {a["axis"]: a for a in compute_executive_brief(
        health_scores=[{"switch": f"CORE-{i:02d}", "band": "Insufficient Data", "score": 90} for i in range(30)],
        lifecycle_risk=compute_lifecycle_risk({f"CORE-{i:02d}": {"model": "", "sw_version": ""} for i in range(30)}))["axes"]}["Hardware lifecycle (EoL)"]
    clean = {a["axis"]: a for a in compute_executive_brief(
        health_scores=[{"switch": f"ACC-{i:02d}", "band": "Good", "score": 85} for i in range(10)],
        lifecycle_risk=compute_lifecycle_risk({f"ACC-{i:02d}": {"model": "C9300-48P", "sw_version": "17.09.04"} for i in range(10)}))["axes"]}["Hardware lifecycle (EoL)"]
    assert blind["severity"] == "Info" and "unknown model" in blind["headline"]
    assert not (blind["headline"] == clean["headline"] and blind["severity"] == clean["severity"])


def test_exec_brief_multicast_querier_blind_when_no_mcast_svi():
    """[audit-3 #12 false-health] AV multicast known present but NO multicast SVI collected -> querier coverage is
    blind, must NOT read '0 querier gap(s) / Low' (indistinguishable from verified-clean)."""
    from cisco_toolkit.analyze import compute_multicast_intelligence, compute_executive_brief
    from cisco_toolkit.model import InterfaceData
    groups = [{"group": "239.1.1.1", "name": "AV", "broadcast": True, "category": "Broadcast-AV"}]
    mi = compute_multicast_intelligence({"multicast": {"classified_groups": groups, "igmp_queriers": [], "ptp": {}}},
                                        {"ACC": {"Vlan10": InterfaceData(port="Vlan10", svi_ip="10.0.10.1")}})
    ax = {a["axis"]: a for a in compute_executive_brief(health_scores=[{"band": "Good", "score": 85}],
                                                        multicast_intelligence=mi)["axes"]}["Multicast / timing"]
    assert ax["severity"] == "Info" and "not assessable" in ax["headline"]


def test_stp_accidental_root_flagged_when_ext_id_disabled():
    """[audit-3 #14 false-health] a bare default-priority 32768 root (ext-id OFF, legacy IOS) was missed; the
    detector hard-assumed 32768 + sys-id-ext(vlan)."""
    from cisco_toolkit.parse import parse_spanning_tree_root
    from cisco_toolkit import analyze
    legacy = ("VLAN0010\n  Spanning tree enabled protocol ieee\n  Root ID    Priority    32768\n"
              "             Address     0019.0011.2233\n             This bridge is the root\n"
              "  Bridge ID  Priority    32768\n             Address     0019.0011.2233\n")
    f = analyze.stp_root_findings({"core": parse_spanning_tree_root(legacy)}, {"core": {}})
    assert any(a["vlan"] == "10" for a in f["accidental"])             # bare-32768 now flagged
    eng = ("VLAN0010\n  Root ID    Priority    4106\n             Address     0019.0011.2233\n"
           "             This bridge is the root\n  Bridge ID  Priority    4106\n")
    f2 = analyze.stp_root_findings({"core": parse_spanning_tree_root(eng)}, {"core": {}})
    assert not f2["accidental"]                                        # engineered 4096+10 -> NOT accidental


def test_failure_impact_off_scan_gateway_is_indeterminate_not_no_impact():
    """[audit-3 #7 false-health] a transit SPOF whose VLAN gateway is on an UNCOLLECTED device was skipped and the
    switch reported 'No reachability impact' -- a clean bill. It must disclose the off-scan-gateway coverage gap."""
    from cisco_toolkit.model import InterfaceData
    from cisco_toolkit.analyze import compute_failure_impact
    def trunk(p, nb, npt, v="10"):
        return InterfaceData(port=p, cdp_neighbor=nb, neighbor_port=npt, endpoint_type="Switch", trunk_allowed_vlans=v)
    def ep(p, vl, mac):
        return InterfaceData(port=p, switchport_mode="Access", vlan=vl, end_host_mac=mac)
    # GW (gateway) UNCOLLECTED: DIST is the sole transit path for VLAN 10 endpoints on ACC
    ai = {"DIST": {"Gi0/1": trunk("Gi0/1", "GW", "Gi1/1"), "Gi0/2": trunk("Gi0/2", "ACC", "Gi0/24")},
          "ACC": {"Gi0/24": trunk("Gi0/24", "DIST", "Gi0/2"), "Gi0/1": ep("Gi0/1", "10", "aaaa.0000.0001"),
                  "Gi0/2": ep("Gi0/2", "10", "bbbb.0000.0002")}}
    dist = [r for r in compute_failure_impact(ai) if r["host"] == "DIST"][0]
    assert dist["off_scan_gw_vlans"] >= 1
    assert "No reachability impact" not in dist["detail"] and "INDETERMINATE" in dist["detail"]


def test_endpoint_classify_apc_legal_name_and_truthful_evidence():
    """[audit-4 #13 format-fidelity] APC's IEEE-registry legal name 'American Power Conversion Corp' (block
    00C0B7) contains none of the UPS/PDU rule substrings, so real APC UPS endpoints fell through to Unknown with a
    FALSE 'no vendor signal' evidence string -- contradicted by the row's own vendor column."""
    from cisco_toolkit.analyze import _classify_endpoint
    cls, _conf, _ev = _classify_endpoint("American Power Conversion Corp", "", "", "", False)
    assert cls == "UPS/PDU"                                                   # legal name now classifies
    _c2, _q2, ev2 = _classify_endpoint("Some Unlisted Maker Inc", "", "", "", False)
    assert _c2 == "Unknown" and "no vendor" not in ev2.lower() and "Some Unlisted Maker" in ev2   # truthful evidence
    _c3, _q3, ev3 = _classify_endpoint("", "", "", "", False)
    assert _c3 == "Unknown" and "no vendor" in ev3.lower()                    # genuinely no signal still says so


def test_endpoint_census_discloses_basis_vs_canonical_n_endpoints():
    """[audit-4 #17 scale-ssot] the Endpoint Census sheet counts learned MACs on ALL non-trunk ports (a superset),
    while executive_brief.scale.n_endpoints counts only Access-tagged-port endpoints -> the sheet's row count (5326
    on [HISTORY-REDACTED]) silently contradicted the Exec Summary headline (5127). The sheet must disclose its basis."""
    from openpyxl import Workbook
    import cisco_toolkit.excel as X
    from cisco_toolkit.model import InterfaceData
    def port(mode, mac):
        d = InterfaceData(); d.switchport_mode = mode; d.end_host_mac = mac; d.vlan = "10"; return d
    ai = {"sw1": {"Gi1/0/1": port("Access", "aaaa.0000.0001"),     # canonical n_endpoints
                  "Gi1/0/2": port("", "bbbb.0000.0002"),           # superset extra (empty switchport, has a MAC)
                  "Gi1/0/3": port("Trunk", "cccc.0000.0003")}}     # excluded (trunk)
    wb = Workbook(); wb.remove(wb.active); X.harden_workbook(wb)
    X.write_endpoint_census_sheet(wb, ai)
    blob = " ".join(str(c) for row in wb[X.ENDPOINT_CENSUS_SHEET_NAME].iter_rows(values_only=True) for c in row if c)
    assert "n_endpoints" in blob and "Access" in blob              # discloses the basis difference vs the canonical


def test_nrfu_routing_adjacencies_warns_when_routing_not_collected():
    """[audit-4 #7 false-health] 'Routing adjacencies up' is a hard fail-gate, but it could only fire when an
    OSPF/BGP protocol_health row exists -> a group whose routing was NEVER collected silently PASSED, reading
    identical to a verified-up group (the bare-show-logging false-health class at the cutover gate). Must WARN
    (not assessable) on no routing evidence."""
    from cisco_toolkit.analyze import compute_migration_readiness
    dm = {"single_fiber": set(), "errdis": set(), "halfdup_up": set(), "sole_gw": {}, "orphan": set(),
          "access_by_vlan": {}, "model": {"hosts": ["CORE1", "ACC1"]}}
    mg = [{"switches": ["CORE1", "ACC1"], "endpoints": 100}]
    hs = [{"switch": "CORE1", "band": "Good"}, {"switch": "ACC1", "band": "Good"}]
    def rcheck(ph):   # ph = protocol_health (the 7th positional arg)
        g = compute_migration_readiness({}, mg, hs, [], [], [], ph, dm)[0]
        return next(c for c in g["checks"] if c["check"] == "Routing adjacencies up")
    # blind: routing WAS collected this run (host OTHER, outside the group) but NOT for this group's switches ->
    # the gate can't certify this group -> warn (mirrors the Baseline-capture coverage pattern).
    blind = rcheck([{"switch": "OTHER", "protocol": "OSPF", "severity": "Info"}])
    verified = rcheck([{"switch": "CORE1", "protocol": "OSPF", "severity": "Info"}])  # routing collected, up
    down = rcheck([{"switch": "CORE1", "protocol": "OSPF", "severity": "High"}])      # routing collected, down
    none_run = rcheck([])                                                             # no routing anywhere -> not in scope
    assert blind["status"] == "warn" and verified["status"] == "pass" and down["status"] == "fail"
    assert none_run["status"] == "pass"                                              # pure-L2 run: no cry-wolf
    assert blind != verified                                                         # blind distinguishable from verified
