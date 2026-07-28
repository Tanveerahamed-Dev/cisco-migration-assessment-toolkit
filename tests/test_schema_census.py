"""Feature J3 — the per-section coverage census (SuzieQ `describe` analog).

`ssot.compute_schema_census(snap)` turns a snapshot into a coverage MAP: for every top-level
section it records the coverage-honest 3-state token (published / collected_but_empty /
not_collected), the section's cardinality, its structural shape, and an honest note. This is the
map that answers, for an access-only collection (the [HISTORY-REDACTED] fleet, where a whole distribution/core
tier is UN-collected), exactly what was SEEN vs what is a blind spot — directly addressing the
real "outputs feel filler" field problem (an uncollected tier, never a code bug).

Coverage-honesty (Law 3): a present-but-empty section is `collected_but_empty` (collected,
nothing found — NOT a blind spot); an absent section is `not_collected` (a real blind spot).
Nothing is ever labelled "ok"/"healthy": absence of evidence is never health.
"""
import json
import os

import pytest
from openpyxl import Workbook

from cisco_toolkit import ssot
from cisco_toolkit.excel import COVERAGE_SCHEMA_SHEET_NAME, write_coverage_schema_sheet


# --- a small synthetic snapshot exercising all three states + all kinds ----------------------
SNAP = {
    "routes": [{"prefix": "10.0.0.0/24"}],          # published list
    "vpc": {"sw1": {"role": "primary"}},            # published dict
    "acls": [],                                     # collected_but_empty (present, empty list)
    "overlay": {},                                  # collected_but_empty (present, empty dict)
    "script_version": "3.23.0",                     # published scalar (kind=scalar, count=None)
    "punchlist": 0,                                 # collected_but_empty scalar (falsy, no len)
    # `mpls` is deliberately ABSENT -> not_collected (a blind spot); never label it healthy.
}


def _census(snap):
    return ssot.compute_schema_census(snap)


def _row(census, key):
    return next((s for s in census["sections"] if s["key"] == key), None)


def test_every_section_is_classified_into_a_row():
    census = _census(SNAP)
    keys_in = set(SNAP)
    keys_out = {s["key"] for s in census["sections"]}
    assert keys_out == keys_in, "every top-level section must appear exactly once in the census"
    assert census["schema"] == "schema_census/1"


def test_states_are_the_three_honest_tokens_only():
    census = _census(SNAP)
    allowed = {"published", "collected_but_empty", "not_collected"}
    for s in census["sections"]:
        assert s["state"] in allowed, f"{s['key']} carries a non-honest state {s['state']!r}"
    # NON-VACUITY (mutation-proved, 2026-07-28): this used to close with
    #   assert s["state"] == ssot.abstention_reason(SNAP, s["key"])
    # while compute_schema_census's own body is literally `state = abstention_reason(snap, key)` —
    # an f(x) == f(x) identity that NO mutation of abstention_reason can fail, because both sides
    # move together. (Changing its `return "collected_but_empty"` to `return "published"` — a
    # coverage-honesty lie — left this test green; only the sibling below caught it.) The set check
    # above is the same tautology one level out: those three tokens ARE the function's codomain.
    # Pin the expected state PER KEY instead, so a token that flips is a red test here too.
    assert {s["key"]: s["state"] for s in census["sections"]} == {
        "script_version": "published",
        "routes": "published",
        "vpc": "published",
        "acls": "collected_but_empty",
        "overlay": "collected_but_empty",
        "punchlist": "collected_but_empty",
    }


def test_present_but_empty_reads_collected_but_empty_not_blind_spot():
    """The load-bearing coverage-honesty distinction: an EMPTY section was collected — genuinely
    nothing of this kind found — and must NOT be mislabelled a blind spot."""
    census = _census(SNAP)
    not_collected_note = ssot._CENSUS_NOTE["not_collected"]
    for empty_key in ("acls", "overlay", "punchlist"):
        row = _row(census, empty_key)
        assert row["state"] == "collected_but_empty", empty_key
        # the empty note must NOT be the blind-spot note (it explicitly says "not a blind spot")
        assert row["note"] != not_collected_note, f"{empty_key} empty mislabelled as blind spot"
        assert "not a blind spot" in row["note"].lower(), f"{empty_key} note must disclaim blind spot"
        assert row["note"] != "ok" and "healthy" not in row["note"].lower()


def test_absent_section_reads_not_collected_blind_spot():
    """An absent section is a real blind spot (not_collected) — never silently a clean result."""
    # `mpls` is not in SNAP; the census only rows keys that ARE present, so absence is proven via
    # abstention_reason directly (the census never fabricates a healthy row for a missing key).
    assert ssot.abstention_reason(SNAP, "mpls") == "not_collected"
    assert _row(_census(SNAP), "mpls") is None
    # and a section explicitly present with value None is not_collected in the census too
    census = _census({**SNAP, "mpls": None})
    row = _row(census, "mpls")
    assert row["state"] == "not_collected"
    assert row["kind"] == "absent"
    assert "blind spot" in row["note"].lower()


def test_count_is_len_for_containers_and_none_for_scalars():
    census = _census(SNAP)
    assert _row(census, "routes")["count"] == 1 and _row(census, "routes")["kind"] == "list"
    assert _row(census, "vpc")["count"] == 1 and _row(census, "vpc")["kind"] == "dict"
    assert _row(census, "acls")["count"] == 0 and _row(census, "acls")["kind"] == "list"
    assert _row(census, "overlay")["count"] == 0 and _row(census, "overlay")["kind"] == "dict"
    # a scalar has no cardinality -> count is None (never a fabricated 0)
    assert _row(census, "script_version")["count"] is None
    assert _row(census, "script_version")["kind"] == "scalar"
    assert _row(census, "punchlist")["count"] is None      # 0 is a falsy scalar, not a container


def test_summary_reconciles_with_the_rows():
    census = _census(SNAP)
    summ = census["summary"]
    sections = census["sections"]
    assert summ["n_sections"] == len(sections)
    assert summ["n_published"] == sum(1 for s in sections if s["state"] == "published")
    assert summ["n_collected_but_empty"] == sum(1 for s in sections if s["state"] == "collected_but_empty")
    assert summ["n_not_collected"] == sum(1 for s in sections if s["state"] == "not_collected")
    # the three buckets partition the section set (no row is double-counted or dropped)
    assert (summ["n_published"] + summ["n_collected_but_empty"]
            + summ["n_not_collected"]) == summ["n_sections"]


@pytest.mark.parametrize("bad", [None, [], "garbage", 42, {"": None}])
def test_total_on_bad_input(bad):
    """Deterministic + total: any non-dict (or degenerate) snapshot yields a well-formed census,
    never an exception."""
    census = ssot.compute_schema_census(bad)
    assert census["schema"] == "schema_census/1"
    assert isinstance(census["sections"], list)
    assert census["summary"]["n_sections"] == len(census["sections"])
    # buckets still partition
    s = census["summary"]
    assert s["n_published"] + s["n_collected_but_empty"] + s["n_not_collected"] == s["n_sections"]


def test_census_is_deterministic_and_non_volatile():
    """Presence/absence + structural shape only — no dates, no model — so two runs on the same
    snapshot are byte-identical (this is what keeps the golden stable)."""
    a = json.dumps(_census(SNAP), sort_keys=True)
    b = json.dumps(_census(dict(SNAP)), sort_keys=True)
    assert a == b


# --- the workbook sheet -----------------------------------------------------------------------
_EXPECTED_HEADER = ["Section", "State", "Kind", "Count", "Note"]


def test_sheet_appears_with_exact_headers():
    wb = Workbook()
    census = _census(SNAP)
    write_coverage_schema_sheet(wb, census)
    assert COVERAGE_SCHEMA_SHEET_NAME in wb.sheetnames
    ws = wb[COVERAGE_SCHEMA_SHEET_NAME]
    # row 1 is the coverage-honesty caveat banner; the header row is row 2
    header = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert header == _EXPECTED_HEADER, header


def test_row1_banner_is_static_no_live_counts():
    """The frozen sheet-schema golden captures row 1 as the header cell. Row 1 must therefore be
    STATIC — no data-dependent counts — or every new snapshot section churns the golden and
    desensitises the additive-only shrink guard (the exact cry-wolf this design avoids). The live
    roll-up lives in the '(all sections)' totals DATA row instead."""
    wb = Workbook()
    write_coverage_schema_sheet(wb, _census(SNAP))
    ws = wb[COVERAGE_SCHEMA_SHEET_NAME]
    banner = ws.cell(1, 1).value or ""
    assert not any(ch.isdigit() for ch in banner), f"row-1 banner must carry no live counts: {banner!r}"
    # the counts still appear, but as a data row (row 3) so the header stays byte-stable
    totals = [ws.cell(3, col).value for col in range(1, 6)]
    assert totals[0] == "(all sections)" and isinstance(totals[3], int)
    assert "published" in str(totals[4]) and "blind spots" in str(totals[4])


def test_sheet_is_total_on_empty_census():
    wb = Workbook()
    write_coverage_schema_sheet(wb, {})           # missing schema/sections/summary -> no crash
    assert COVERAGE_SCHEMA_SHEET_NAME in wb.sheetnames


# --- the golden fixture, where present ---------------------------------------------------------
_GOLDEN = os.path.join(os.path.dirname(__file__), "golden", "snapshot.json")


@pytest.mark.skipif(not os.path.isfile(_GOLDEN), reason="golden snapshot not present in worktree")
def test_census_over_golden_fixture_is_coverage_honest():
    with open(_GOLDEN, encoding="utf-8") as f:
        snap = json.load(f)
    census = ssot.compute_schema_census(snap)
    # every real snapshot section is rowed, and each state is one of the three honest tokens
    assert census["summary"]["n_sections"] == len(snap)
    for s in census["sections"]:
        assert s["state"] in {"published", "collected_but_empty", "not_collected"}
