"""Current-snapshot cutover baseline: unchanged blockers can never become PASS."""
from collections import Counter, defaultdict
from copy import deepcopy

import pytest
from openpyxl import load_workbook

from cisco_toolkit.analyze import (
    classify_current_baseline_item,
    compute_current_baseline_gate,
    compute_validation_plan,
)
from cisco_toolkit.html import (compute_campaign_trend, compute_cutover_gate, compute_snapshot_delta,
                                write_campaign_workbook, write_diff_workbook)
from cisco_toolkit.model import InterfaceData
from cisco_toolkit.protocol_assurance import (
    compute_native_protocol_deltas,
    current_baseline_blocker_export,
    protocol_family_change_set,
)


def _row(number=1, *, state="assessed", marker="", wave="Group 1", category="Routing"):
    row = {
        "device": f"r{number}",
        "platform": "ios",
        "wave": wave,
        "category": category,
        "severity": "High",
        "check": f"check {number}",
        "command": "show evidence",
        "expect": marker or "Preserve the observed baseline.",
        "why": "Cutover acceptance requires this observation.",
    }
    if state is not None:
        row.update({
            "evidence_state": state,
            "projection_custody": "embedded_unverified",
            "source_key": f"routing_neighbors.r{number}.ospf",
        })
    return row


def _plan(rows):
    by_wave = defaultdict(list)
    for row in rows:
        by_wave[row["wave"]].append(row)
    return {
        "items": rows,
        "by_wave": dict(by_wave),
        "summary": {
            "n_items": len(rows),
            "n_waves": len(by_wave),
            "by_category": dict(Counter(row["category"] for row in rows)),
            "n_high": sum(row["severity"] in ("Critical", "High") for row in rows),
        },
        "banner": "bounded validation plan",
    }


@pytest.mark.parametrize(
    ("row", "expected"),
    [
        (_row(state="assessed"), "clear"),
        (_row(state="degraded"), "degraded"),
        (_row(state="review"), "review"),
        (_row(state="not_verified"), "not_verified"),
        (_row(state=None, marker="PRE-CUTOVER DEGRADED — BLOCKER: legacy"), "degraded"),
        (_row(state=None, marker="PRE-CUTOVER REVIEW — BLOCKER: legacy"), "review"),
        (_row(state=None, marker="ROUTING BASELINE NOT VERIFIED — BLOCKER: legacy"), "not_verified"),
        (_row(state=None, marker="ETHERCHANNEL BASELINE NOT VERIFIED — BLOCKER: legacy"), "not_verified"),
        (_row(state=None, marker="prefix PRE-CUTOVER DEGRADED — BLOCKER: not exact"), "clear"),
        (_row(state="assessed", marker="PRE-CUTOVER DEGRADED — BLOCKER: conflict"), "invalid"),
        ({"evidence_state": ["degraded"], "expect": ""}, "invalid"),
        (None, "invalid"),
    ],
)
def test_row_classifier_prefers_typed_state_and_keeps_exact_legacy_fallback(row, expected):
    assert classify_current_baseline_item(row) == expected


def test_gate_distinguishes_absence_clean_blocked_and_uncertain_baselines():
    absent = compute_current_baseline_gate(None)
    empty = compute_current_baseline_gate(_plan([]))
    clear = compute_current_baseline_gate(_plan([_row()]))
    blocked = compute_current_baseline_gate(_plan([
        _row(1, state="degraded"), _row(2, state="review", wave="Group 2"),
    ]))
    uncertain = compute_current_baseline_gate(_plan([
        _row(1, state="review"), _row(2, state="not_verified"),
    ]))

    assert absent["verdict"] == empty["verdict"] == "NOT_ASSESSED"
    assert clear["verdict"] == "CLEAR" and clear["assessed"] is True
    assert blocked["verdict"] == "BLOCKED"
    assert blocked["summary"] == {
        "n_items": 2,
        "n_blockers": 2,
        "n_blockers_returned": 2,
        "blockers_capped": False,
        "by_state": {"degraded": 1, "review": 1, "not_verified": 0},
        "by_wave": {"Group 1": 1, "Group 2": 1},
    }
    assert [row["evidence_state"] for row in blocked["blockers"]] == ["degraded", "review"]
    assert uncertain["verdict"] == "INDETERMINATE"
    assert uncertain["integrity"]["valid"] is True
    assert uncertain["summary"]["by_state"] == {
        "degraded": 0, "review": 1, "not_verified": 1,
    }


@pytest.mark.parametrize("mutation", ["wave", "summary", "row", "category", "conflict"])
def test_gate_reconciles_all_plan_views_and_never_echoes_rows_from_an_invalid_plan(mutation):
    plan = deepcopy(_plan([_row(state="degraded")]))
    if mutation == "wave":
        forged = dict(plan["by_wave"]["Group 1"][0])
        forged["check"] = "forged blocker"
        plan["by_wave"]["Group 1"][0] = forged
    elif mutation == "summary":
        plan["summary"]["n_items"] = 99
    elif mutation == "row":
        plan["items"][0]["expect"] = ["hostile"]
        plan["by_wave"]["Group 1"][0]["expect"] = ["hostile"]
    elif mutation == "category":
        plan["items"][0]["category"] = ["hostile"]
        plan["by_wave"]["Group 1"][0]["category"] = ["hostile"]
    else:
        marker = "PRE-CUTOVER REVIEW — BLOCKER: contradict typed degradation"
        plan["items"][0]["expect"] = marker
        plan["by_wave"]["Group 1"][0]["expect"] = marker

    gate = compute_current_baseline_gate(plan)

    assert gate["verdict"] == "INDETERMINATE"
    assert gate["integrity"]["valid"] is False
    assert gate["blockers"] == []
    assert gate["summary"]["n_blockers"] == 0
    assert gate["summary"]["by_state"] == {
        "degraded": 0, "review": 0, "not_verified": 0,
    }
    assert "hostile" not in str(gate)


def test_gate_bounds_returned_rows_and_device_controlled_text_without_hiding_counts(tmp_path):
    rows = []
    for number in range(55):
        row = _row(number, state="degraded", wave="Group 10")
        row["check"] = "x" * 500
        row["expect"] = "y" * 900
        rows.append(row)

    gate = compute_current_baseline_gate(_plan(rows))

    assert gate["verdict"] == "BLOCKED"
    assert gate["summary"]["n_blockers"] == 55
    assert gate["summary"]["n_blockers_returned"] == 50
    assert gate["summary"]["blockers_capped"] is True
    assert gate["summary"]["by_wave"] == {"Group 10": 55}
    assert len(gate["blockers"]) == 50
    assert max(map(len, (row["check"] for row in gate["blockers"]))) == 240
    assert max(map(len, (row["expect"] for row in gate["blockers"]))) == 600

    complete = current_baseline_blocker_export({"validation_plan": _plan(rows)})
    assert complete["status"] == "available"
    assert complete["owns_verdict"] is False
    assert complete["summary"]["n_blockers_total"] == 55
    assert complete["summary"]["n_rows_returned"] == 55
    assert complete["summary"]["omitted"] == 0
    assert complete["summary"]["complete"] is True
    assert len(complete["rows"]) == 55
    assert complete["rows"][54]["device"] == "r54"

    snapshot = {
        "devices": {f"r{number}": {} for number in range(55)},
        "interfaces": {},
        "health_scores": [],
        "punchlist": [],
        "validation_plan": _plan(rows),
    }
    output = tmp_path / "capped-baseline.xlsx"
    write_diff_workbook(snapshot, deepcopy(snapshot), str(output))
    workbook = load_workbook(output, read_only=True)
    try:
        capped_rows = list(workbook["Current Baseline Gate"].iter_rows(values_only=True))
        export_rows = list(workbook["Current Baseline Export"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert len([row for row in capped_rows[1:] if row[0] == "degraded"]) == 50
    assert len([row for row in export_rows[1:] if row[0] == "degraded"]) == 55
    assert any(row[1] == "r54" for row in export_rows)
    assert "rendered=50, total=55, omitted=5" in capped_rows[-1][5]
    assert "rendered=55, total=55, omitted=0, complete=YES" in export_rows[-1][5]

    long_wave = "Group 20 " + "z" * 400
    long_gate = compute_current_baseline_gate(_plan([_row(state="review", wave=long_wave)]))
    bounded_wave = next(iter(long_gate["summary"]["by_wave"]))
    assert len(bounded_wave) == 120 and bounded_wave.endswith("...")


def test_fhrp_validation_rows_publish_typed_state_custody_and_exact_source_leaf():
    interfaces = {
        "core1": {
            "Vlan10": InterfaceData(
                port="Vlan10", svi_ip="10.0.10.1/24",
                hsrp_behavior="HSRP grp 10 Init VIP 10.0.10.254",
            ),
            "Vlan20": InterfaceData(
                port="Vlan20", svi_ip="10.0.20.1/24",
                hsrp_behavior="VRRP grp 20 Master VIP 10.0.20.254",
            ),
        },
    }
    plan = compute_validation_plan(
        interfaces,
        move_groups=[{"switches": ["core1"]}],
        devices={"core1": {"platform": "ios"}},
    )
    rows = {row["device"] + row["check"]: row for row in plan["items"]
            if row["category"] == "FHRP"}
    degraded = next(row for row in rows.values() if "VLAN 10" in row["check"])
    healthy = next(row for row in rows.values() if "VLAN 20" in row["check"])

    assert degraded["evidence_state"] == "degraded"
    assert healthy["evidence_state"] == "assessed"
    for row, interface in ((degraded, "Vlan10"), (healthy, "Vlan20")):
        assert row["projection_custody"] == "embedded_unverified"
        assert row["source_key"] == f"interfaces.core1.{interface}.hsrp_behavior"
    assert compute_current_baseline_gate(plan)["verdict"] == "BLOCKED"


def test_unchanged_degraded_current_state_fails_combined_gate_without_changing_delta_semantics(tmp_path):
    validation = _plan([_row(state="degraded")])
    snapshot = {
        "devices": {"r1": {}},
        "interfaces": {"r1": {}},
        "health_scores": [{"switch": "r1", "band": "Good", "score": 90}],
        "punchlist": [],
        "validation_plan": validation,
    }

    delta = compute_snapshot_delta(snapshot, deepcopy(snapshot))
    gate = compute_cutover_gate(
        delta,
        {"verdict": "PASS", "verdict_note": "bounded certificate passed"},
    )

    assert delta["verdict"] == "CLEAN", "the before/after owner remains delta-only"
    assert delta["current_baseline"]["verdict"] == "BLOCKED"
    assert gate["verdict"] == "FAIL"
    assert gate["current_baseline_verdict"] == "BLOCKED"
    assert gate["current_baseline_degraded"] == 1
    assert "resolve or explicitly disposition 1 definite current-baseline degradation" in gate["operator_note"]

    output = tmp_path / "unchanged-degraded.xlsx"
    workbook_gate = write_diff_workbook(
        snapshot, deepcopy(snapshot), str(output),
        precert={
            "verdict": "PASS", "verdict_note": "bounded certificate passed",
            "flows": {}, "stamps": {}, "segmentation": [], "intents": [], "blind_spots": [],
        },
    )
    family_changes = protocol_family_change_set(
        delta["protocol_adjacencies"], {"expected_changes": []},
        native_deltas=compute_native_protocol_deltas(snapshot, deepcopy(snapshot)))
    additive_gate = compute_cutover_gate(
        delta,
        {"verdict": "PASS", "verdict_note": "bounded certificate passed"},
        protocol_family_changes=family_changes,
    )
    assert workbook_gate == additive_gate
    assert workbook_gate["verdict"] == gate["verdict"]
    workbook = load_workbook(output, read_only=True)
    try:
        assert "Current Baseline Gate" in workbook.sheetnames
        rows = list(workbook["Current Baseline Gate"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows[1][:5] == ("degraded", "r1", "Group 1", "Routing", "check 1")
    assert rows[1][6:] == ("embedded_unverified", "routing_neighbors.r1.ospf")
    assert rows[2][0] == "DISCLOSURE"
    assert "rendered=1, total=1, omitted=0 (NOT CAPPED)" in rows[2][5]
    assert "Current Baseline Export sheet" in rows[2][5]


def test_missing_current_validation_plan_withholds_pass_for_a_computed_delta():
    snapshot = {
        "schema": "collect_parse_snapshot/1",
        "devices": {}, "interfaces": {}, "health_scores": [], "punchlist": [],
    }
    delta = compute_snapshot_delta(snapshot, snapshot)
    gate = compute_cutover_gate(
        delta,
        {"verdict": "PASS", "verdict_note": "bounded certificate passed"},
    )

    assert delta["verdict"] == "CLEAN"
    assert delta["current_baseline_required"] is True
    assert delta["current_baseline"]["verdict"] == "NOT_ASSESSED"
    assert gate["verdict"] == "INDETERMINATE"
    assert gate["current_baseline_verdict"] == "NOT_ASSESSED"


@pytest.mark.parametrize(
    "by_state",
    [
        {"degraded": 0, "review": 0},
        {"degraded": 0, "review": 0, "not_verified": 0, "forged_clear": 99},
    ],
)
def test_combined_gate_totality_rejects_noncanonical_baseline_state_counts(by_state):
    malformed = {
        "schema": "current_baseline_gate/1",
        "verdict": "CLEAR",
        "assessed": True,
        "note": "forged clear",
        "summary": {
            "n_items": 1, "n_blockers": 0, "n_blockers_returned": 0,
            "blockers_capped": False, "by_state": by_state, "by_wave": {},
        },
        "blockers": [],
        "integrity": {"valid": True, "failures": []},
    }
    decision = compute_cutover_gate(
        {
            "verdict": "CLEAN", "verdict_display": "NO DELTA REGRESSION OBSERVED",
            "verdict_note": "delta only", "current_baseline": malformed,
        },
        {"verdict": "PASS", "verdict_note": "bounded certificate"},
    )

    assert decision["verdict"] == "INDETERMINATE"
    assert decision["current_baseline_verdict"] == "INDETERMINATE"
    assert decision["current_baseline_blockers"] == 0
    assert "missing or malformed" in decision["current_baseline_note"]


def test_combined_gate_totality_rejects_an_unhashable_blocker_state():
    malformed = {
        "schema": "current_baseline_gate/1", "verdict": "BLOCKED", "assessed": True,
        "note": "forged blocker",
        "summary": {
            "n_items": 1, "n_blockers": 1, "n_blockers_returned": 1,
            "blockers_capped": False,
            "by_state": {"degraded": 1, "review": 0, "not_verified": 0}, "by_wave": {"G": 1},
        },
        "blockers": [{"evidence_state": ["degraded"]}],
        "integrity": {"valid": True, "failures": []},
    }

    decision = compute_cutover_gate(
        {"verdict": "CLEAN", "verdict_note": "delta only", "current_baseline": malformed},
        {"verdict": "PASS", "verdict_note": "certificate"},
    )

    assert decision["verdict"] == "INDETERMINATE"
    assert decision["current_baseline_verdict"] == "INDETERMINATE"


def test_campaign_exposes_last_collection_baseline_without_rewriting_trajectory_semantics(tmp_path):
    before = {
        "generated_at": "2026-01-01T00:00:00", "devices": {"r1": {}}, "interfaces": {},
        "health_scores": [{"switch": "r1", "band": "Fair", "score": 65}],
        "punchlist": [], "migration_readiness": [], "lifecycle_risk": {"summary": {"n_past_ldos": 0}},
        "validation_plan": _plan([_row(state="assessed")]),
    }
    after = deepcopy(before)
    after["generated_at"] = "2026-02-01T00:00:00"
    after["health_scores"] = [{"switch": "r1", "band": "Good", "score": 85}]
    after["validation_plan"] = _plan([_row(state="degraded")])

    trend = compute_campaign_trend([before, after])

    assert trend["verdict"] == "IMPROVING", "trajectory remains direction-of-travel, not acceptance"
    assert trend["current_baseline"]["verdict"] == "BLOCKED"
    assert trend["current_baseline"]["summary"]["n_blockers"] == 1
    assert trend["steps"][-1]["current_baseline_verdict"] == "BLOCKED"
    assert trend["steps"][-1]["current_baseline_blockers"] == 1

    output = tmp_path / "campaign-blocked.xlsx"
    write_campaign_workbook([before, after], str(output))
    workbook = load_workbook(output, read_only=True)
    try:
        summary = workbook["Campaign Summary"]
        assert summary.cell(2, 2).value == "IMPROVING"
        assert summary.cell(3, 1).value == "CURRENT BASELINE GATE"
        assert summary.cell(3, 2).value == "BLOCKED"
    finally:
        workbook.close()
