#!/usr/bin/env python3
"""
CISCO MIGRATION EXTRACTOR - COLLECT + PARSE + AUTO-POPULATE EXCEL (V3.23.0)

V3.23.0 (additive; health scoring + migration readiness - final intelligence layer):
  * compute_health_scores() + 'Health Scores' sheet (Phase 28): a per-switch 0-100 score with
    weighted deductions synthesised from the physical (L1), L3-forwarding, cross-layer and
    protocol-health findings already computed this run; banded Excellent/Good/Fair/Poor/Critical.
  * compute_migration_readiness() + 'Migration Readiness' sheet (Phase 29): per move-group
    (reusing compute_move_groups) a READY / CAUTION / NOT READY verdict from a 10-check
    pre-migration checklist (uplink + gateway redundancy, cross-layer criticals, err-disable,
    STP, port-channels, routing adjacencies, orphan VLANs, degraded uplinks, device health).
  * Both embedded in the snapshot ('health_scores', 'migration_readiness') and rendered by the
    HTML explorer's new health mode (nodes coloured by score band). Cross-layer findings now
    carry a 'hosts' field for per-device attribution. Pure derivation; no new collection.

V3.22.0 (additive; protocol behavior analysis):
  * compute_protocol_health() + 'Protocol Health' sheet (Phase 27) - one row per
    (switch, protocol) for STP / EtherChannel / VTP / OSPF / BGP / EIGRP / FHRP, each with a
    derived health severity. Re-parses already-collected raw output (STP blocked/inconsistent
    ports, etherchannel member-state flags, VTP mode/revision, OSPF/BGP neighbor states) plus
    the FHRP info on InterfaceData. Adds 'show spanning-tree detail' to collection for STP
    topology-change counts (tolerant: absent output -> TCN omitted). Embedded in the snapshot
    under 'protocol_health'. No new InterfaceData fields, no new imports.

V3.21.0 (additive; cross-layer correlation):
  * build_dependency_map() + compute_cross_layer_correlations() + 'Cross-Layer Analysis'
    sheet (Phase 26). Correlates findings ACROSS layers - it ingests the physical-health (L1)
    and L3-forwarding (L3) records already computed this run, plus topology facts derived from
    the model (articulation via _vlan_components, single-fiber uplinks, single-member port-
    channels, orphan VLANs), and emits CL-01..CL-10 compounded-risk findings (e.g. a single
    fiber that is the only path to a sole gateway = total isolation with no L1 or L3 backup).
    Embedded in the snapshot under 'cross_layer'. Pure derivation; no new collection, no new
    InterfaceData fields, no new imports.

V3.20.0 (additive; L3 forwarding map):
  * 'L3 Forwarding Map' sheet (Phase 25) - one row per gateway SVI: VLAN, SVI IP, FHRP
    protocol/role/virtual-IP (parsed from hsrp_behavior), routing source + next hop +
    primary/secondary subnets, device object-tracking summary, and a derived L3 Risk flag
    (single-gateway / tracked-object-down / no-FHRP / ok). Adds 'show track' to collection
    for object/IP-SLA tracking (tolerant: absent output -> blank). Pure derivation otherwise;
    reuses the existing FHRP parsers and per-SVI routing enrichment. Embedded in the snapshot
    under 'l3_forwarding'. No new InterfaceData fields, no new imports.
    Note: tracked objects are device-level (show track is not bound to a specific SVI without
    parsing HSRP 'standby ... track' config), so the Tracking column is per-device context.

V3.19.0 (additive; full-stack flow trace):
  * trace_full_flow(src_ip, dst_ip, all_interfaces) + optional 'Flow Trace' sheet (written
    only when BOTH --flow-src and --flow-dst are given). Traces an L1->L3 path between two
    endpoint IPs: locates each endpoint's access switch/port/VLAN (end_host_ip), resolves the
    gateway SVI(s) per subnet (svi_ip + subnet_primary_route containment, FHRP peers), decides
    L2-same-subnet vs L3-routed, and walks the fabric over STP-forwarding links (reuses
    _link_carries / build_network_model). Flags single-points-of-failure on the path
    (single-fiber uplinks via _physical_uplink_index, single gateway, VLAN partition) and
    scores risk LOW/MEDIUM/HIGH/CRITICAL. Embedded in the snapshot under 'flow_trace' for the
    HTML explorer's Flow mode. L4 note: no ACL/service-policy data is collected, so transport
    filtering is NOT simulated (path reachability is reported as L1-L3, L4 assumed open).
    Routing is DERIVED from InterfaceData (no standalone routing_data dict exists). No new
    collection, no new InterfaceData fields, no new top-level imports.

V3.18.0 (additive; L1 physical health):
  * 'Physical Health' sheet (Phase 23) - one row per physical port: negotiated speed/duplex/
    media, input/CRC errors, output drops, port-channel, PoE, and a derived Physical Risk flag
    (err-disabled / half-duplex-on-trunk / error-rate-high / single-fiber-uplink / ok). Pure
    derivation from already-parsed data: error counters via parse_show_interface_counters,
    uplink topology via build_network_model (single-fiber), PoE budget via DevicePhysical;
    negotiated duplex/media read from the 'show interfaces' DETAIL block (InterfaceData.duplex
    is normalized and drops half-duplex). Records embedded in the snapshot under
    'physical_health'. No new collection, no new InterfaceData fields, no new imports.

V3.17.0 (additive; consolidation):
  * One run now also emits a self-contained Blast-Radius Explorer ('<output>_explorer.html')
    with the live topology baked in - no more hand-loading the snapshot into a separate HTML
    tool. write_html_explorer() patches a copy of the read-only 'blast_radius_explorer.html'
    template that ships beside this script; '--no-html' suppresses it; a missing template is a
    warning, not a crash. New: Phase 22 (post-save). No new imports, no new InterfaceData fields.

V3.16.0 intelligence layer (additive; deterministic graph reasoning over already-parsed
data - no new collection, no external calls):
  * 'Causality Chains' sheet - root cause -> propagation mechanism -> blast radius, e.g. a
    sole-gateway VLAN with no FHRP, a transit switch that is the only L2 path from some
    endpoints to their gateway, or a single non-bundled uplink that isolates a VLAN.
  * 'Failure Impact' sheet - migration blast-radius simulation: each switch is removed from
    the CDP/LLDP-derived graph and per-VLAN endpoint->gateway reachability is recomputed,
    classifying each VLAN as Hard partition / Backup-covered (STP reconverges via a named
    blocked link) / FHRP-covered (standby gateway) / no impact.
  Limits: graph completeness == CDP/LLDP discovery (off-scan links invisible -> impact is a
  lower bound); reachability simulation, not bandwidth (no traffic-volume telemetry);
  post-failure path is approximate (backup links are proven and named, STP trees are not
  recomputed). See the NEW-V3.16 block above main() for details.

V3.15.0 feature batch (Tier 1 + Tier 2). All additive; the existing interface-build
pipeline (build_interfaces / InterfaceData / the main interface sheet) is unchanged.
  Tier 1 (aggregations of already-parsed data, no new collection):
   * 'Findings' sheet - a risk register cross-referenced from the interface DB:
     no-gateway VLANs, multi-SVI VLANs without FHRP, err-disabled ports, STP-inconsistent
     ports, and native-VLAN/duplex mismatches across confirmed inter-switch links.
   * 'Capacity' sheet - per-switch port and PoE headroom (from DevicePhysical) with
     port-bound / PoE-bound flags, for consolidation planning.
   * Topology diagram - writes topology.mmd (Mermaid) and topology.dot (Graphviz) next to
     the workbook; one-end-only links are drawn dashed. (compute_topology_links() is now the
     shared link builder used by the Topology Links sheet, Findings, and the diagram.)
   * Pre/post-cutover validation - a snapshot JSON (<output>.snapshot.json) is written next to every
     workbook; run with '--compare OLD.json NEW.json' to emit a validation workbook: a Summary CUTOVER
     VERDICT (CLEAN / REVIEW / REGRESSED) + Interface / Endpoint / SVI changes + Health Shifts (per-switch
     band regressions/improvements) + Findings Delta (punch-list findings opened vs resolved by the
     cutover). --compare skips SSH and the template.
  Tier 2 (new commands collected on both platforms; parsers are best-effort and blank when a
  platform/feature is absent - _load_cmd_output() already skips command output that errored):
   * 'Interface Health' sheet - input/CRC errors, output drops and last-input/last-output from
     'show interfaces' (IOS) / 'show interface' (NX-OS); flags error ports and up-but-idle ports.
   * 'Security Posture' sheet - port-security, 802.1X/MAB sessions and DHCP-snooping binding
     counts ('show port-security', 'show authentication sessions', 'show ip dhcp snooping binding').
   * 'Routing Adjacencies' sheet - OSPF/EIGRP neighbors and BGP peers ('show ip ospf neighbor',
     'show ip eigrp neighbors', 'show ip bgp summary' / 'show bgp summary').

V3.14.13 fix: per-SVI subnet matching now prefers the CONNECTED route whose prefix contains
  the SVI's own IP, so a coexisting /32 host route or redistributed loopback can no longer be
  picked as the SVI's subnet (seen on a 4500-X VSS where Vlan167/169 matched 10.0.0.60/32
  instead of their connected /24). Connected-and-containing > connected > less-specific.

V3.14.12 platform-resilience hardening (a wrong 'platform' in devices.json was silently
gutting the SVI/routing/HSRP/run-config columns - it ran the NX-OS command set against IOS
Catalyst gear, so NX-only command forms errored and that data was never collected):
FIX-V14-19 collect() now gathers the UNION of IOS+NX-OS command variants (ordered by the
  detected platform). The two lists differ only in platform-divergent forms (show standby vs
  show hsrp, the two run-config forms, show interfaces vs show interface, etc.); collecting
  both means a wrong label no longer drops data - _load_cmd_output() already skips the form
  that errors and keeps the one that worked.
FIX-V14-20 detect_platform_from_files() now treats a command that errored on the wrong
  platform as NOT a valid marker (checks all _CISCO_ERRORS, not just 'invalid'); an IOS
  '% Incomplete command' file was being miscounted as an NX-OS marker.
FIX-V14-21 After collection the platform is re-derived from the actual (non-error) output and
  overrides a wrong devices.json value, so parsing and the Switch Inventory label are correct.
  NOTE: the netmiko driver is still chosen at connect time from the device-file platform, so
  for best results set 'platform' to 'ios'/'auto' (not a wrong value) in devices.json.

V3.14.11 cleanup/quality pass:
CLEAN-V14-16 Removed parse_management_ip_from_neighbors() - orphaned since V3.14.4 (FIX-V14-8
  re-sourced mgmt IP from the switch's own interfaces); it had zero callers.
FIX-V14-17 parse_multicast_info() rewritten: parse the real 'show ip pim interface' table
  (Address/Interface/Ver-Mode) instead of an 'X is up' assumption that never matched it, and
  capture VlanN interfaces in mroute IIF/OIL via a function-local Vlan-aware token regex (the
  global IFACE_TOKEN_RE intentionally excludes Vlan so it can't be widened safely). The
  Multicast Info column now populates on SVIs running PIM instead of staying blank.
FIX-V14-18 HSRP non-brief fallback (m1/m2) now captures the virtual IP via last-IP-on-line,
  the same heuristic as the brief branch, instead of a weak trailing optional that usually
  missed it. (These lines were a live fallback for 'show standby all'-style output, not dead
  code as a prior changelog note implied.)

V3.14.10 adds over V3.14.9:
FEAT-V14-15 New "Topology Links" sheet: physical inter-switch link map built from CDP/LLDP.
  The two parsers now also capture the remote port; each link is de-duplicated across both
  endpoints' views into one canonical undirected row and flagged "Both ends" (seen from both
  switches) or "One end" (only one side reported). Host endpoints (phones/APs/servers) are
  excluded - infrastructure links only. This is the physical map for cutover sequencing and
  also informs the transit-coupling gap noted on the Move Groups sheet.

V3.14.9 adds over V3.14.8:
FEAT-V14-14 New "Move Groups" sheet: proposes migration waves by computing connected
  components of the shared-VLAN graph (switches that share an L2 broadcast domain must move
  together). Captures TRANSITIVE coupling - if VLAN10 spans A+B and VLAN20 spans B+C, then
  A/B/C are one group even though A and C share nothing directly (easy to miss by eye). Each
  row: the switches in the wave, the spanning VLANs that force the coupling, endpoint count,
  gateways, and any redundant STP-blocked paths to verify before cutover. Ordered
  largest-blast-radius first. Coupling uses ACTUAL VLAN presence (access port or SVI),
  consistent with the VLAN Census; trunk-allowed-only transit is not modeled and VLAN 1
  (default) is excluded from grouping (flagged in Notes where it still spans a group).

V3.14.8 adds over V3.14.7:
FEAT-V14-13 Two assessment sheets aggregated from the interface DB (no new parsing risk):
  - "VLAN Census": one row per in-use VLAN (has access ports and/or an SVI) with VLAN id +
    name, switch spread, access-port count, distinct-endpoint (MAC) count, and the gateway
    (which switch holds the SVI, its IP, subnet, FHRP state). The migration move-group unit.
  - "Endpoint Census": flat cross-switch list, one row per port with a learned MAC -
    hostname, port, VLAN(+name), MAC, IP, endpoint type, location, neighbor.
  Also collects 'show vlan brief' and uses it to fill authoritative VLAN names (enriching the
  interface sheet's VLAN Name column too).

V3.14.7 adds over V3.14.6:
FEAT-V14-12 New "STP Detail" Excel sheet: one row per STP port showing WHICH VLANs it
  forwards vs blocks (compressed ranges, e.g. "10,20-23"), instead of only the collapsed
  state on the interface sheet. Surfaces per-trunk blocked VLANs - the redundant/standby
  path per VLAN, key for migration move-group planning. The summary parser was refactored
  to derive from the new per-VLAN detail parser, so the interface-sheet STP state and this
  sheet can never disagree. Under MST the VLAN columns hold MST instance numbers.

V3.14.6 adds/fixes over V3.14.5:
FEAT-V14-10 VRRP and GLBP gateways are now parsed ('show vrrp brief', 'show glbp brief',
  collected on IOS + NX-OS). Previously only HSRP was recognized, so VRRP/GLBP gateways were
  blank. The column is renamed "HSRP Behavior" -> "FHRP Behavior" (the umbrella term for
  HSRP/VRRP/GLBP) and each value is prefixed with the actual protocol (e.g. "VRRP grp 1
  Master VIP ..."). HSRP takes precedence when multiple are present on one interface. The old
  "hsrp behavior" header is still recognized on existing sheets (no duplicate column on re-run).
CLEAN-V14-11 Removed 3 stray backspace (\x08) bytes that had corrupted the detailed-format
  HSRP regexes in parse_hsrp_summary, restoring the 'show standby all' / 'show hsrp all'
  fallback path. Updated stale log-file/logger/usage-example version strings to V3.14.6.

V3.14.5 fixes over V3.14.4:
FIX-V14-9 STP state column no longer claims "Forwarding" from link-up alone (a false-health
  signal: a port can be up but STP-blocked). "Forwarding" is now asserted only when confirmed
  by the per-VLAN Role/Sts table in 'show spanning-tree' (newly collected). Blocked ports come
  from 'show spanning-tree blockedports'; the previously-collected-but-unused
  'show spanning-tree inconsistentports' is now honored as "Inconsistent". When STP state
  cannot be confirmed the cell is left blank (honest "unknown") instead of a false "Forwarding".

V3.14.4 fixes over V3.14.3:
FIX-V14-8 'Switch Mgmt IP' / current_switch_ip now reflects the switch's OWN management IP.
  Previously it was taken from parse_management_ip_from_neighbors(), i.e. a CDP/LLDP
  NEIGHBOR's IP, so the column was wrong on both the interface and SVI sheets. Now sourced
  from the switch's own 'show ip interface brief' (default VRF + NX-OS 'vrf management' for
  mgmt0), with a running-config 'ip address' fallback. Priority: OOB mgmt port > loopback >
  sole management SVI > lowest up SVI. Blank (honest) if nothing qualifies, instead of wrong.
  neighbor_switch_ip is unchanged (correctly CDP-sourced).

V3.14.3 adds over V3.14.2:
FEAT-V14-7 New "SVI Gateway" Excel sheet: one row per SVI (Vlan interface) across all
  devices, carrying the per-SVI routing / HSRP / multicast data that the interface sheet
  excludes by design (it drops Vlan/loopback/tunnel rows). Adds a 'svi_ip' field captured
  from each SVI's 'ip address' line (IOS dotted-mask and NX-OS prefix forms). The interface
  sheet and physical-port rows are unchanged.

V3.14.2 Fixes over V3.14.1 (make the V3.14 feature set actually function end-to-end):
FIX-V14-2 build_interfaces() now accepts switch_identity. It was declared with a dead
  global_bpduguard parameter while the caller passed switch_identity=..., so EVERY parse
  raised TypeError. On the default parallel run the worker pool swallowed it and the
  interface sheet came out empty; with --workers 1 it hard-crashed.
FIX-V14-3 Added the 12 V3.14 fields to InterfaceData (current/neighbor switch identity,
  per-SVI routing, HSRP, multicast). They were read and written but never defined on the
  model, raising AttributeError as soon as FIX-V14-2 was applied.
FIX-V14-4 Collected the commands the V3.14 enrichment consumes: show ip route,
  show vtp status, show standby brief / show hsrp brief, show ip mroute,
  show ip pim interface. They were referenced by parsers but absent from the collection
  lists, so routing/HSRP/multicast/VTP enrichment was always empty.
FIX-V14-5 Wired the 12 fields through to Excel (HEADER_TO_FIELD, ensure_headers, row writer).
FIX-V14-6 Corrected per-SVI subnet matching: now matches the connected route whose outgoing
  interface is the SVI, instead of attaching the first /24-/28 in the table to every SVI.

V3.13.0 Enhancements over V3.11.0:
NEW-V12-1 New "Switch Inventory" Excel sheet (site survey): one row per switch.
  Columns: Hostname, Platform, Model/PID, Serial Number, Chassis Serial,
  SW Version, Uptime, System MAC, # Power Supplies, PS Status,
  Power Capacity (W), Power Drawn (W), Power Remaining (W), # Modules,
  Total Ports, Active Ports, Fan Status, Temperature Status.
NEW-V12-2 New parsers: parse_show_version, parse_show_inventory,
  parse_show_environment_power, parse_show_environment, parse_show_module_count.
NEW-V12-3 New commands: show version, show inventory, show environment power,
  show environment (both platforms), show module (NX-OS).
NEW-V12-4 DevicePhysical dataclass + build_device_physical() builder.
Backward compatible: interface sheet and all existing columns unchanged.

V3.14.1 Fixes over V3.14.0:
FIX-V14-1 Guarded Phase 6 interface row writing against missing per-host interface dictionaries to prevent KeyError and continue processing.

V3.14.0 Enhancements over V3.13.0:
NEW-V14-1 Added current-switch identity fields to interface model: current switch serial, current switch mgmt IP, current switch VTP domain.
NEW-V14-2 Added neighbor-switch identity fields from CDP/LLDP: neighbor switch serial, neighbor switch mgmt IP, neighbor switch VTP domain.
NEW-V14-3 Added per-SVI/routed subnet enrichment: primary route, secondary route(s), next-hop, routing source, HSRP behavior, multicast information.
NEW-V14-4 Kept backward compatibility by only appending new commands, parsers, dataclass fields, and optional Excel columns.

V3.11.0 Enhancements over V3.10.0:
NEW-V11-1 "ARP Source Switch" column: switch whose ARP resolved the MAC->IP.
NEW-V11-2 "CDP/LLDP Neighbor" column: CDP/LLDP device-id on that port.

V3.10.0 Fixes over V3.9.0:
FIX-V10-1  Expanded Cisco error rejection in _load_cmd_output.
FIX-V10-2  collect_global_arp: additive merge across all ARP commands.
FIX-V10-3  Added "show ip arp detail" as third ARP source.
FIX-V10-4  Per-device ARP contribution logged at INFO.

V3.9.0 Fixes: FIX-R1..FIX-R10 (see git history for details).

Requirements: pip install netmiko openpyxl
Run:
  python3 COLLECT_PARSE_V3_14_6.py --devices-file devices.json \
          --template Migration_Assessment_Template_Updated.xlsx
Optional:
  --output OUT.xlsx  --debug-headers  --no-collect
  --collection-dir DIR  --workers N  --debug-arp
"""

import os, sys, json, re, logging, warnings, argparse, time  # CHANGED-V3.23.1: +time (connect backoff)
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
# 'from dataclasses import dataclass, field' moved to cisco_toolkit.analyze with
# ScoringConfig (PHASE 2.7 step 10) - the monolith's last dataclass user.
from datetime import datetime
from typing import Callable, Dict, List, NamedTuple, Optional   # Callable/NamedTuple/Optional: the Tier-2 #8 axis registry

# 'from openpyxl.cell.cell import MergedCell' dropped (step 26): its only user,
# append_interface_rows, moved to cisco_toolkit.excel.
# NEW-V3.23.11 (PHASE 2.7 step 1): pure text/interface-name helpers + regex
# constants extracted to the cisco_toolkit package; imported back so every
# existing reference keeps working unchanged (behaviour byte-identical).
# normalize_ifname / detect_link_type dropped step 28 (build_interfaces, their last monolith
# user, moved to build); PHYSICAL_IFACE_RE step 27; _TRUNK_STATUS_WORDS step 26;
# VALID_IFACE_RE / IFACE_TOKEN_RE / normalize_speed (steps 17/25). Only safe_fs_name remains.
from cisco_toolkit.textutils import safe_fs_name
# _split_macs re-import dropped in step 21 (its last monolith user, write_vlan_census_sheet,
# moved to excel; analyze.py + excel.py now import it from textutils directly).
# NEW-V3.23.38 (PHASE 2.7 step 28): the monolith no longer imports ANY parser from
# cisco_toolkit.parse - build_interfaces, the last parser consumer, moved to
# cisco_toolkit.build. The ~30 parsers all live in cisco_toolkit/parse.py; the parser
# unit tests import them from there directly (test_parsers / test_platform_variants).
# NEW-V3.23.19 (PHASE 2.7 step 9): the passed-around data model (InterfaceData /
# DevicePhysical) extracted to the package leaf; imported back so every type hint
# and constructor call keeps working unchanged (behaviour byte-identical).
from cisco_toolkit.model import InterfaceData, DevicePhysical
# NEW-V3.23.40 (PHASE 2.7 step 30): the version string, hoisted into cisco_toolkit/__init__.py
# (single source of truth); imported back here so LOG_FILE / argparse keep working + snapshot_state
# (now in cisco_toolkit.html) reads the same value.
from cisco_toolkit import __version__
# NEW-V3.23.20-.25 (PHASE 2.7 steps 10-15): analyze-layer symbols imported back so the
# Excel writers + the physical/L3/flow functions + main() still in this file keep working.
# ScoringConfig / SCORING / _host_role are NOT re-exported anymore (step 15 moved their last
# monolith users, the scoring compute_*); _health_band stays (write_health_scores_sheet uses it).
from cisco_toolkit.analyze import (
    vlan_inventory,        # NEW: canonical VLAN inventory — single source of truth for the dashboards' VLAN count
    compute_move_groups,   # _health_band dropped (step 23): write_health_scores_sheet moved to excel
    # compute_topology_links / compute_findings dropped (step 22): their last monolith users
    # (write_topology_sheet / write_findings_sheet) moved to excel.
    # build_network_model dropped (step 25): write_physical_health_sheet (its last monolith user) moved.
    # _link_carries (step 19) / _vlan_components (step 18) / compute_causality_chains / compute_failure_impact
    # (step 24) were dropped earlier for the same reason.
    compute_data_quality, compute_health_scores,
    compute_score_sensitivity, compute_calibration_report, compute_migration_readiness,
    compute_protocol_health,   # _poe_device_util / _physical_uplink_index dropped (step 25): phy writer moved
    build_dependency_map, compute_cross_layer_correlations, trace_full_flow,
    stp_root_findings, compute_migration_punchlist, compute_framework_coverage,   # NEW-V3.23.62/.63 + W2-3 (framework-mapping matrix)
    compute_hostname_mismatches, reconcile_cdp_neighbor_names,   # NEW-V3.23.68 + audit-5 #1 (split-node reconcile)
    compute_operational_drift, compute_trunk_capture_gaps,   # NEW-V3.23.93 (false-health) + trunk-capture abstention (l2-native-vlan-1 abstains_when, mechanical)
    compute_endpoint_identity,                         # NEW-V3.23.95 (endpoint vendor + class intelligence)
    compute_flow_paths,                                # NEW-V3.23.126 (representative end-to-end flow paths; workbook/runbook twin of the Flow Simulator)
    compute_endpoint_dependencies,                     # NEW-V3.23.96 (clusters / dependencies / per-switch validation)
    compute_subnet_intelligence,                       # NEW-V3.23.97 (subnet / routing reachability)
    compute_migration_scenarios,                       # NEW-V3.23.98 (per-group cutover scenario framework)
    compute_protocol_intelligence,                     # NEW-V3.23.100 (protocol-state doctrine: cause + remediation)
    compute_service_map,                               # NEW-V3.23.101 (L4 service map from ACL ports + multicast activity)
    compute_ptp_readiness,                             # NEW-V3.23.108 (PTP / media-timing readiness -> punch-list)
    compute_multicast_intelligence,                    # NEW-V3.23.115 (media-fabric deep-dive: MAC-alias / querier / PTP tree)
    compute_remediation_plan,                          # NEW-V3.23.116 (assess->act: per-device config snippets, review-only)
    compute_validation_plan,                           # NEW-V3.23.143 (assess->verify: per-wave post-cutover checks)
    compute_golden_drift,                              # NEW-V3.23.146 (per-device config drift vs a baseline)
    compute_syslog_intelligence,                       # NEW-V3.23.164 (NOS-style operational log analysis)
    compute_qos_audit,                                 # NEW-V3.23.165 (configured QoS posture + doctrine findings)
    compute_software_risk,                             # NEW-V3.23.166 (advisory-surface screening + train lifecycle)
    compute_platform_health,                           # NEW-V3.23.167 (control-plane CPU/memory capacity screening)
    compute_lifecycle_risk,                            # NEW-V3.23.117 (hardware EoL / end-of-support band per device)
    compute_segmentation,                              # NEW-V3.23.118 (L3 isolation posture: VRF / gateway-ACL per domain)
    compute_executive_brief,                           # NEW-V3.23.120 (cross-axis migration brief synthesis)
    compute_device_dossiers,                           # NEW-V3.23.172 (per-asset compound-risk synthesis -- the Device Risk Register)
    compute_collection_completeness,                   # NEW-V3.23.109 (pre-assessment blind-spot / collection report)
    compute_application_intelligence,                  # NEW-V3.23.112 (application-domain synthesis + migration risk)
    compute_vlan_cutover_matrix,                       # MASTER_PLAN §4.3 (per-VLAN cutover matrix: the maintenance-window sheet)
    compute_cable_map,                                 # NEW: EDA-style physical cable-map SSOT (explorer + webapp cable-map views)
)
from cisco_toolkit.design_advisor import compute_design_blueprint, compute_design_nrfu, compute_architecture_coverage   # NEW: CCDE-grounded target-state design blueprint + design-driven NRFU + architecture-coverage SSOT (single source of truth)
from cisco_toolkit.attestation import compute_attestation   # roadmap D3: zero-egress attestation — trust claims RE-DERIVED at build time (falsifiable, never a badge)
from cisco_toolkit.nrfu_export import compute_nrfu_commands   # NEW: four-phase NRFU certification pack (READ-ONLY commands + expected values pre-filled from the snapshot)
# NEW-V3.23.24 (PHASE 2.7 step 14): the command-output I/O glue. _load_cmd_output +
# _safe_parse dropped step 28 (build_interfaces, their last monolith user, moved to
# build); only _CISCO_ERRORS stays (platform detection reuses it). Plan A / Tier-1 #3
# adds the zero-parse yield ledger (reset at run start, published as snap['parse_yield']).
from cisco_toolkit.cmdio import _CISCO_ERRORS, parse_yield_report, reset_parse_ledger
# NEW-V3.23.30 (PHASE 2.7 step 20): the Excel layer's shared sheet/header helpers; imported
# back so the write_* sheet builders + main()'s template-fill keep working unchanged.
from cisco_toolkit.excel import (
    find_header_row, ensure_headers,   # sortkey dropped step 26 (append_interface_rows moved);
    # _census_header/_census_autofit were dropped step 24 (their writers moved to excel).
    write_device_inventory_sheet, write_svi_gateway_sheet, write_stp_detail_sheet,  # step 21
    write_vlan_census_sheet, write_endpoint_census_sheet,                            # step 21
    write_move_group_sheet, write_topology_sheet, write_findings_sheet,             # step 22
    write_capacity_sheet, compute_capacity, write_topology_diagram,                 # step 22 (compute_capacity: V3.23.87 capacity embedded in snapshot)
    write_cross_layer_sheet, write_protocol_health_sheet,   # _SEV_FILL dropped step 25 (last writers moved)
    write_health_scores_sheet, write_score_sensitivity_sheet, write_calibration_sheet,  # step 23 (+calibration V3.23.47)
    write_nat_sheet,   # NAT inventory V3.23.50
    write_security_sheet,   # Security Posture (CIS-aligned) V3.23.59
    write_framework_coverage_sheet,   # Framework Coverage (CIS/NIST/PCI/STIG mapping over existing checks) W2-3
    write_detector_schema_sheet,   # Detector Schema (per-detector descriptors; not-observed != healthy) J1
    write_config_hygiene_sheet,   # Config Hygiene (undefined refs / unused structures) V3.23.61
    write_stp_roots_sheet,   # STP Root Bridges (accidental root / gateway misalignment) V3.23.62
    write_punchlist_sheet,   # Migration Punch-List (consolidated, severity-ranked) V3.23.63
    write_protocol_boundaries_sheet,   # protocol-to-protocol analysis (workbook surfacing)
    write_addressing_conflicts_sheet,  # reachability findings -> workbook surfacing
    write_fhrp_consistency_sheet,      # reachability findings -> workbook surfacing
    write_trunk_native_sheet,          # reachability findings -> workbook surfacing
    write_link_phy_sheet,              # reachability findings -> workbook surfacing
    compute_addressing_conflicts, compute_fhrp_consistency,   # NEW-V3.23.64 (L2 checks -> punch-list)
    compute_trunk_native_mismatches, compute_duplex_speed_mismatches,
    write_migration_readiness_sheet,
    write_interface_health_sheet, write_security_posture_sheet, write_routing_adjacency_sheet,  # step 24
    write_causality_chains_sheet, write_failure_impact_sheet, write_link_centrality_sheet,      # step 24 (+Link Centrality V3.23.88)
    write_cabling_schedule_sheet,                                                               # EDA-style cable schedule (op-status + LAG) from the cable-map SSOT
    write_wave_sequencing_sheet,                                                                # NEW-V3.23.89 (cutover sequencing)
    write_endpoint_intelligence_sheet,                          # NEW-V3.23.95 (endpoint vendor + class)
    write_endpoint_dependencies_sheet,                          # NEW-V3.23.96 (clusters / dependencies)
    write_subnet_reachability_sheet,                            # NEW-V3.23.97 (subnet / routing reachability)
    write_migration_scenarios_sheet,                            # NEW-V3.23.98 (per-group cutover scenario)
    write_protocol_intelligence_sheet,                          # NEW-V3.23.100 (protocol-state cause + remediation)
    write_service_map_sheet,                                     # NEW-V3.23.101 (L4 service map + multicast activity)
    write_multicast_intelligence_sheet,                          # NEW-V3.23.115 (media-fabric multicast intelligence)
    write_remediation_plan_sheet,                                # NEW-V3.23.116 (generated config snippets, review-only)
    write_validation_plan_sheet,                                 # NEW-V3.23.143 (per-wave post-cutover validation checklist)
    write_nrfu_commands_sheet,                                   # NEW (four-phase NRFU certification pack; orchestration-roadmap frontier)
    write_golden_drift_sheet,                                    # NEW-V3.23.146 (per-device config drift vs baseline)
    write_feature_compliance_sheet,                             # roadmap I2 (per-feature ConfigCompliance decomposition)
    write_acl_shadow_sheet,                                     # roadmap G1 (offline ACL line-reachability / shadow proof)
    write_external_reconcile_sheet,                            # roadmap B (declared source-of-truth vs collected evidence)
    write_capture_integrity_sheet,                             # roadmap K1 (truncation / pager / CLI-error guard)
    write_attestation_sheet,                                   # roadmap D3 (re-derived zero-egress attestation panel)
    write_whatif_sheet,                                        # roadmap G4 (single-snapshot failure-injection)
    write_path_intents_sheet,                                  # roadmap G3 (named segmentation/path intents)
    write_syslog_intelligence_sheet,                             # NEW-V3.23.164 (NOS-style operational log analysis)
    write_qos_audit_sheet,                                       # NEW-V3.23.165 (configured QoS posture + doctrine findings)
    write_software_risk_sheet,                                   # NEW-V3.23.166 (advisory-surface screening + train lifecycle)
    write_platform_health_sheet,                                 # NEW-V3.23.167 (control-plane CPU/memory capacity screening)
    write_device_risk_sheet,                                     # NEW-V3.23.172 (per-asset compound-risk register)
    write_lifecycle_risk_sheet,                                  # NEW-V3.23.117 (hardware EoL / end-of-support)
    write_segmentation_sheet,                                    # NEW-V3.23.118 (L3 segmentation / isolation posture)
    write_architecture_review_sheet,                             # NEW-V3.23.161 (leading-practice conformance scorecard)
    write_collection_completeness_sheet,                         # NEW-V3.23.109 (pre-assessment blind-spot report)
    write_coverage_schema_sheet,                                 # J3 (per-section coverage census — SuzieQ `describe` analog)
    write_application_intelligence_sheet,                        # NEW-V3.23.112 (application-domain synthesis)
    write_vlan_cutover_sheet,                                    # MASTER_PLAN §4.3 (per-VLAN cutover matrix)
    write_executive_summary_sheet,                              # NEW-V3.23.75 (one-page landing synthesis)
    write_physical_health_sheet, write_flow_trace_sheet, write_flow_paths_sheet, write_l3_forwarding_sheet,             # step 25
    append_interface_rows,                                                                       # step 26
)
from cisco_toolkit.feature_compliance import compute_feature_compliance   # roadmap I2 (per-feature ConfigCompliance)
from cisco_toolkit.aclcheck import compute_filter_line_reachability       # roadmap G1 (offline ACL shadow proof)
from cisco_toolkit.external_import import read_inventory_file, reconcile_external   # roadmap B (external SoT reconcile)
# roadmap K1 guard, widened Tier-2 #6: the entry feeds the STREAMING path API (all captures);
# the in-memory dict API (compute_capture_integrity) remains the module's direct-use surface.
from cisco_toolkit.capture_integrity import (compute_capture_integrity_from_paths,
                                             load_capture_meta, CAPTURE_META_FILENAME)
from cisco_toolkit.whatif import run_scenarios                          # roadmap G4 (failure-injection what-if)
from cisco_toolkit.path_assertions import evaluate_path_assertions      # roadmap G3 (named path/segmentation intents)
from cisco_toolkit.precert import compute_precert, schema_compat_status  # roadmap C1 (Pre-Change Cert); P3-E2 schema gate
from cisco_toolkit.assertions import evaluate_pack                      # roadmap A1/H1/H2 (checks-as-data)
# NEW-V3.23.37-.38 (PHASE 2.7 steps 27-28): the model-construction layer - the per-device
# InterfaceData builder + switch-level DevicePhysical / switch-identity records + global-ARP
# enrichment. Homed in cisco_toolkit/build.py; imported back so main() keeps calling them.
from cisco_toolkit.build import (
    build_interfaces,   # step 28: the big per-device InterfaceData builder
    build_device_physical, build_switch_identity, collect_global_arp,
    apply_global_arp, detect_cross_device_dual_connections, build_acls, build_object_groups,
    read_run_config,   # NEW-V3.23.146 (raw running-config text for golden-config drift)
    read_syslog_log,   # NEW-V3.23.164 (raw 'show logging' text for syslog intelligence)
    build_platform_metrics,   # NEW-V3.23.167 (CPU/memory/system-resources facts for platform health)
    build_routes, inscope_subnets, scope_routes, build_bgp_received, build_nat, build_security, build_config_hygiene, build_stp_roots, build_vpc, build_fhrp_detail, build_overlay, build_copp, build_mpls, build_routing_neighbors,
    build_lisp, build_cts, build_dmvpn, build_crypto, build_bfd, build_ipv6_nd, build_ipv6_routing,   # universal arch coverage
    build_aci,   # Cisco ACI (APIC JSON-ingestion channel)
    build_sdwan,   # Cisco Catalyst SD-WAN (vManage JSON-ingestion channel)
    build_firewall,   # Cisco firewall (ASA / Secure Firewall Threat Defense) HA -- SSH show-text channel
    build_ise,   # Cisco ISE (Identity Services Engine deployment) -- JSON controller-REST channel
    build_fmc,   # Cisco Secure Firewall Mgmt Center (FMC) -- JSON controller-REST channel
    build_arista,   # MULTI-VENDOR: Arista EOS MLAG -- the FIRST non-Cisco vendor channel (device-native JSON)
    build_juniper,   # MULTI-VENDOR: Juniper Junos SRX chassis-cluster HA -- the SECOND non-Cisco vendor channel ('| display json')
    build_cloud,   # PUBLIC CLOUD: AWS security-group exposure -- the FIRST cloud-domain channel (account-as-device JSON export)
    build_fortigate,   # MULTI-VENDOR: Fortinet FortiGate HA cluster sync -- the THIRD non-Cisco vendor channel ('get system ha status')
    build_mroute,   # CISCO DEPTH: multicast RPF integrity -- (S,G) Null incoming-interface from 'show ip mroute'
    build_pim, build_ipv6_fhs, build_ntp, build_port_security_detail, build_storm_control, build_qos_runtime, build_undocumented_neighbors,
    build_igmp_groups, build_igmp_queriers, build_ptp, build_acl_hits,   # NEW-V3.23.102 (multicast / PTP / ACL-hit collection)
    build_redistribution,
)
# NEW-V3.23.39-.40 (PHASE 2.7 steps 29-30): the snapshot-reporting layer - snapshot_state (the JSON
# contract), write_html_explorer (bakes the snapshot into the Blast-Radius Explorer), and the
# '--compare OLD NEW' diff workbook. Homed in cisco_toolkit/html.py; imported back so main() keeps
# building/serializing the snapshot + emitting the HTML + diff outputs.
from cisco_toolkit.html import (snapshot_state, sparsify_interfaces, write_html_explorer, write_diff_workbook,
                                write_campaign_workbook, redact_snapshot,
                                redact_collected_inplace, redact_workbook_cells,   # campaign trend; audit-3 #8 workbook redact
                                redact_collection_dir)                             # Plan A Tier-1 #5 (raw-capture secret scrub)
from cisco_toolkit.context import AnalysisContext                    # Plan A #15 (typed pipeline carrier / strangler)
from cisco_toolkit.coverage_matrix import compute_coverage_matrix   # Plan A #5 (coverage-as-a-first-class-row SSOT)
from cisco_toolkit.detector_schema import compute_detector_schema   # J1 (per-detector descriptors; not-observed != healthy)
from cisco_toolkit.runbook import write_runbook_docx                 # NEW-V3.23.93 (DOCX runbook deliverable)
from cisco_toolkit.deck import write_executive_deck_pptx             # NEW-V3.23.144 (executive PPTX deck deliverable)
from cisco_toolkit.design import write_design_doc_docx               # NEW-V3.23.148 (As-Built HLD/LLD design document)
from cisco_toolkit.mop import write_mop_docx                         # NEW-V3.23.149 (per-wave Method of Procedure)
from cisco_toolkit.gate_state import enforce as gate_enforce         # P0-3/DEC-003 (PPDIOO document-gate refusal)
from cisco_toolkit.gate_state import reset_verdicts as gate_reset_verdicts   # ...and its per-run verdict ledger
from cisco_toolkit.crd import write_crd_docx                         # NEW-V3.23.156 (Plan-phase requirements capture)
from cisco_toolkit.engagement import write_engagement_docx           # NEW-V3.23.157 (engagement workflow / plan of record)
from cisco_toolkit.archreview import (compute_architecture_review,   # NEW-V3.23.160 (leading-practice design review)
                                      write_archreview_docx)
from cisco_toolkit.ops import write_ops_handbook_docx                # NEW-V3.23.168 (Operate-phase handbook)
# FIX-V3.23.8 (P4): scope the warnings filter to DeprecationWarning (netmiko /
# paramiko / cryptography churn + Netmiko 5.x deprecations) instead of suppressing
# EVERYTHING - so genuine UserWarning / RuntimeWarning signals surface.
warnings.filterwarnings("ignore", category=DeprecationWarning)

try:
    from netmiko import ConnectHandler
except ImportError:
    print("ERROR: netmiko not installed. Run: python3 -m pip install netmiko paramiko")
    sys.exit(1)

try:
    from netmiko.ssh_autodetect import SSHDetect
except Exception:
    SSHDetect = None

try:
    from openpyxl import load_workbook
    # Font / PatternFill / Alignment (styles) + get_column_letter (utils) dropped step 29:
    # their last monolith user, write_diff_workbook, moved to cisco_toolkit.html.
except ImportError:
    print("ERROR: openpyxl not installed. Run: python3 -m pip install openpyxl")
    sys.exit(1)

# =============================================================================
# CONFIG
# =============================================================================
# __version__ moved to cisco_toolkit/__init__.py (PHASE 2.7 step 30); imported back near the top of this file.
LOG_FILE              = f"cisco_migration_autofill_v{__version__.replace('.', '_')}.log"
COLLECTION_DIR        = "migration_collection_{}"
DEFAULT_TEMPLATE_FILE = "Migration_Assessment_Template_Updated.xlsx"
DEFAULT_OUTPUT_FILE   = "Migration_Assessment_AUTOFILLED_{}.xlsx"

NETMIKO_TYPE = {"nxos": "cisco_nxos", "ios": "cisco_ios"}

COMMANDS_NXOS = [
    "show interface brief",
    "show interface status",
    "show interface switchport",
    "show vlan brief",           # NEW-V14.8 (authoritative VLAN id->name for census)
    "show interface trunk",
    "show interface transceiver",
    "show port-channel summary",
    "show etherchannel summary",
    "show vpc",                          # NEW-V3.23.125 (vPC / MLAG peer confirmation for the flow simulator)
    "show nve peers",                    # VXLAN-EVPN VTEP peer state -> build_overlay / _d_nve_peer_health
    "show bgp l2vpn evpn summary",       # BGP-EVPN control plane (MAC/IP route exchange) -> _d_evpn_rr_health
    "show nve vni",                      # VXLAN VNI (VLAN/VRF<->VNI) binding state -> _d_nve_vni_health
    "show policy-map interface control-plane",  # CoPP drop counters (NX-OS) -> build_copp / _d_copp_drops
    "show mpls ldp neighbor",            # SP/MPLS: LDP transport-underlay session state -> build_mpls / _d_mpls_ldp_health
    "show bgp vpnv4 unicast summary",    # SP/MPLS: L3VPN VPNv4 PE-PE control-plane -> build_mpls / _d_mpls_l3vpn_health
    "show mpls l2transport vc",          # SP/MPLS: L2VPN/AToM pseudowire VC state -> build_mpls / _d_mpls_l2vpn_health
    "show bfd neighbors",                # BFD fast-failover (both platforms) -> build_bfd / _d_bfd_session_health
    "show ipv6 route summary",           # IPv6 routing-active gate (both platforms) -> build_ipv6_routing
    "show bgp ipv6 unicast summary",     # IPv6 BGP peers (both platforms) -> build_ipv6_routing / _d_ipv6_routing_adjacency
    "moquery -c faultInst",              # Cisco ACI (APIC JSON export) -> build_aci / _d_aci_critical_faults
    "moquery -c fabricNode",             # Cisco ACI fabric inventory (APIC JSON) -> build_aci / _d_aci_node_not_active
    "moquery -c fabricHealthTotal",      # Cisco ACI fabric health score (APIC JSON) -> build_aci / _d_aci_fabric_health_degraded
    "moquery -c fvCtx",                  # Cisco ACI VRF inventory + enforcement (APIC JSON) -> build_aci / _d_aci_vrf_unenforced
    "moquery -c fvTenant",               # Cisco ACI tenant census (APIC JSON) -> build_aci (move-group scoping)
    "moquery -c fvBD",                   # Cisco ACI bridge-domain census (APIC JSON) -> build_aci (move-group scoping)
    "moquery -c fvAEPg",                 # Cisco ACI EPG census (APIC JSON) -> build_aci (finest move-group unit)
    "show failover",                     # Cisco firewall (ASA / Secure Firewall Threat Defense) HA -> build_firewall / _d_firewall_ha_degraded
    "show resource usage",               # Cisco firewall (ASA/FTD) resource capacity -> build_firewall / _d_firewall_resource_exhaustion
    "api/v1/deployment/node",            # Cisco ISE (Identity Services Engine) deployment nodes -> build_ise / _d_ise_*
    "ers/config/node",                   # Cisco ISE ERS-channel fallback export -> build_ise (registered so --no-collect re-reads a collected ERS-only cluster, not 'not-observed')
    "api/fmc_config/v1/devices/devicerecords",            # Cisco FMC device inventory + health -> build_fmc / _d_fmc_device_disconnected
    "api/fmc_config/v1/devicehapairs/ftddevicehapairs",   # Cisco FMC FTD HA pairs -> build_fmc / _d_ftd_ha_degraded
    "api/fmc_config/v1/deployment/deployabledevices",     # Cisco FMC staged-deploy state -> build_fmc / _d_fmc_deployment_pending
    "api/fmc_config/v1/integration/fmchastatuses",        # Cisco FMC manager-HA -> build_fmc / _d_fmc_manager_ha_degraded
    "api/fmc_platform/v1/info/serverversion",             # Cisco FMC server version (vs managed FTD) -> build_fmc / _d_fmc_version_inversion
    "show policy-map interface",         # QoS RUNTIME egress queue/policer drops -> build_qos_runtime / _d_qos_runtime_drops
    "show ip pim rp mapping",            # PIM-SM learned RP -> build_pim / _d_pim_rp_health
    "show ip pim neighbor",              # PIM-SM neighbor adjacency (proof sparse-mode is live) -> build_pim
    "show ipv6 nd raguard policy",       # IPv6 first-hop security: RA-Guard -> build_ipv6_fhs / _d_ipv6_fhs
    "show ipv6 dhcp guard policy",       # IPv6 first-hop security: DHCPv6-Guard -> build_ipv6_fhs
    "show ntp status",                   # NTP clock-sync STATE (IOS form) -> build_ntp / _d_ntp_sync
    "show ntp peer-status",              # NTP clock-sync STATE (NX-OS peer table) -> build_ntp
    "show port-security interface",      # access-edge port-security DETAIL (Secure-shutdown) -> build_port_security_detail / _d_port_security_errdisable
    "show storm-control",                # storm-control action audit (toothless 'None') -> build_storm_control / _d_storm_control_action
    "show spanning-tree",                # NEW-V14.5 (confirmed per-VLAN STP state)
    "show spanning-tree blockedports",
    "show spanning-tree inconsistentports",
    "show cdp neighbors detail",
    "show lldp neighbors detail",
    "show mac address-table",
    "show vrf interface",
    "show ip vrf interface",
    "show running-config interface",
    "show running-config",
    "show power inline",
    "show ip arp",
    "show ip arp vrf all",       # FIX-R1
    "show version",              # NEW-V12
    "show inventory",            # NEW-V12
    "show environment power",    # NEW-V12
    "show environment",          # NEW-V12
    "show module",               # NEW-V12 NX-OS
    "show ip route",             # NEW-V14.2 wiring (per-SVI routing enrichment)
    "show vtp status",           # NEW-V14.2 wiring (VTP domain identity)
    "show hsrp brief",           # NEW-V14.2 wiring (gateway / HSRP behavior)
    "show vrrp brief",           # NEW-V14.6 (VRRP gateways)
    "show glbp brief",           # NEW-V14.6 (GLBP gateways)
    "show ip mroute",            # NEW-V14.2 wiring (multicast info)
    "show ip pim interface",     # NEW-V14.2 wiring (multicast info)
    "show ip interface brief",            # NEW-V14.4 (switch's own mgmt IP)
    "show ip interface brief vrf management",  # NEW-V14.4 (NX-OS mgmt0 lives here)
    "show interface",            # NEW-V15 (interface health counters - NX-OS)
    "show port-security",        # NEW-V15 (security posture)
    "show ip dhcp snooping binding",  # NEW-V15 (security posture)
    "show ip ospf neighbor",     # NEW-V15 (routing adjacencies)
    "show ip eigrp neighbors",   # NEW-V15 (routing adjacencies)
    "show bgp summary",          # NEW-V15 (routing adjacencies - NX-OS)
    "show ip bgp summary",       # NEW-V15 (routing adjacencies)
    "show bgp ipv4 unicast",     # NEW-V3.23.97 (BGP RIB / received prefixes - NX-OS)
    "show ip igmp groups",          # NEW-V3.23.102 (multicast membership census - broadcast fabric)
    "show ip igmp snooping groups", # NEW-V3.23.102 (L2 multicast forwarding)
    "show ip igmp snooping querier",# NEW-V3.23.102 (L2 querier per VLAN)
    "show ptp clock",               # NEW-V3.23.102 (IEEE 1588 PTP clock health)
    "show ptp parent",              # NEW-V3.23.102 (PTP grandmaster)
    "show ip access-lists",         # NEW-V3.23.102 (ACL hit-counts = active-traffic evidence)
    "show logging",                 # NEW-V3.23.164 (buffered log = syslog-intelligence evidence)
    "show logging logfile",         # NEW-V3.23.170 (NX-OS log buffer -- the bare form is '% Incomplete command' on Nexus)
    "show processes cpu",           # NEW-V3.23.167 (control-plane CPU sample - platform health)
    "show system resources",        # NEW-V3.23.167 (NX-OS CPU/memory/load - platform health)
]

COMMANDS_IOS = [
    "show interface status",
    "show interfaces switchport",
    "show vlan brief",           # NEW-V14.8 (authoritative VLAN id->name for census)
    "show interfaces trunk",
    "show interfaces transceiver",
    "show etherchannel summary",
    "show spanning-tree",                # NEW-V14.5 (confirmed per-VLAN STP state)
    "show spanning-tree blockedports",
    "show spanning-tree inconsistentports",
    "show cdp neighbors detail",
    "show lldp neighbors detail",
    "show mac address-table",
    "show vrf interface",
    "show ip vrf interfaces",
    "show running-config | section ^interface",
    "show running-config",
    "show power inline",
    "show ip arp",
    "show ip arp vrf all",       # FIX-R1
    "show version",              # NEW-V12
    "show inventory",            # NEW-V12
    "show environment power",    # NEW-V12
    "show environment",          # NEW-V12
    "show environment all",      # NEW-V3.23.67 (IOS-XE 9300/3850: bare 'show environment' -> '% Incomplete command')
    "show ip route",             # NEW-V14.2 wiring (per-SVI routing enrichment)
    "show vtp status",           # NEW-V14.2 wiring (VTP domain identity)
    "show standby brief",        # NEW-V14.2 wiring (gateway / HSRP behavior)
    "show standby all",          # FHRP DETAIL (election/preempt/tracking) -> build_fhrp_detail / _d_fhrp_resilience
    "show policy-map control-plane",  # CoPP drop counters (IOS / IOS-XE) -> build_copp / _d_copp_drops
    "show mpls ldp neighbor",     # SP/MPLS: LDP transport-underlay session state -> build_mpls / _d_mpls_ldp_health
    "show bgp vpnv4 unicast summary",  # SP/MPLS: L3VPN VPNv4 PE-PE control-plane -> build_mpls / _d_mpls_l3vpn_health
    "show mpls l2transport vc",   # SP/MPLS: L2VPN/AToM pseudowire VC state -> build_mpls / _d_mpls_l2vpn_health
    "show lisp session",  # universal arch coverage
    "show cts environment-data",  # universal arch coverage
    "show dmvpn",  # universal arch coverage
    "show crypto session",  # universal arch coverage
    "show bfd neighbors",  # universal arch coverage
    "show ipv6 interface",  # universal arch coverage
    "show ipv6 route summary",  # universal arch coverage
    "show ospfv3 neighbor",  # universal arch coverage
    "show ipv6 ospf neighbor",  # classic-IOS OSPFv3 form (build_ipv6_routing falls back to it; collect it or OSPFv3 is blind on classic IOS)
    "show bgp ipv6 unicast summary",  # universal arch coverage
    "dataservice/device/control/connections",  # Cisco Catalyst SD-WAN (vManage JSON) -> build_sdwan / _d_sdwan_control_connection_down
    "dataservice/device",             # Cisco Catalyst SD-WAN device reachability (vManage JSON) -> build_sdwan / _d_sdwan_device_unreachable
    "dataservice/device/counters",    # Cisco Catalyst SD-WAN OMP peer counters (vManage JSON) -> build_sdwan / _d_sdwan_omp_peer_down
    "show policy-map interface",  # QoS RUNTIME egress queue/policer drops -> build_qos_runtime / _d_qos_runtime_drops
    "show ip pim rp mapping",     # PIM-SM learned RP -> build_pim / _d_pim_rp_health
    "show ip pim neighbor",       # PIM-SM neighbor adjacency (proof sparse-mode is live) -> build_pim
    "show ipv6 nd raguard policy",  # IPv6 first-hop security: RA-Guard -> build_ipv6_fhs / _d_ipv6_fhs
    "show ipv6 dhcp guard policy",  # IPv6 first-hop security: DHCPv6-Guard -> build_ipv6_fhs
    "show ntp status",            # NTP clock-sync STATE (IOS form) -> build_ntp / _d_ntp_sync
    "show port-security interface",  # access-edge port-security DETAIL (Secure-shutdown) -> build_port_security_detail / _d_port_security_errdisable
    "show storm-control",         # storm-control action audit (toothless 'None') -> build_storm_control / _d_storm_control_action
    "show vrrp brief",           # NEW-V14.6 (VRRP gateways)
    "show glbp brief",           # NEW-V14.6 (GLBP gateways)
    "show ip mroute",            # NEW-V14.2 wiring (multicast info)
    "show ip pim interface",     # NEW-V14.2 wiring (multicast info)
    "show ip interface brief",   # NEW-V14.4 (switch's own mgmt IP)
    "show interfaces",                # NEW-V15 (interface health counters - IOS)
    "show port-security",             # NEW-V15 (security posture)
    "show authentication sessions",   # NEW-V15 (security posture - 802.1X/MAB)
    "show ip dhcp snooping binding",  # NEW-V15 (security posture)
    "show ip ospf neighbor",          # NEW-V15 (routing adjacencies)
    "show ip eigrp neighbors",        # NEW-V15 (routing adjacencies)
    "show ip bgp summary",            # NEW-V15 (routing adjacencies)
    "show ip bgp",                    # NEW-V3.23.97 (BGP RIB / received prefixes - IOS/IOS-XE)
    "show track",                     # NEW-V3.20 (L3 forwarding: object / IP-SLA tracking)
    "show spanning-tree detail",      # NEW-V3.22 (STP topology-change counters)
    "show ip igmp groups",            # NEW-V3.23.102 (multicast membership census - broadcast fabric)
    "show ip igmp snooping groups",   # NEW-V3.23.102 (L2 multicast forwarding)
    "show ip igmp snooping querier",  # NEW-V3.23.102 (L2 querier per VLAN)
    "show ptp clock",                 # NEW-V3.23.102 (IEEE 1588 PTP clock health)
    "show ptp parent",                # NEW-V3.23.102 (PTP grandmaster)
    "show ip access-lists",           # NEW-V3.23.102 (ACL hit-counts = active-traffic evidence)
    "show logging",                   # NEW-V3.23.164 (buffered log = syslog-intelligence evidence)
    "show processes cpu",             # NEW-V3.23.167 (control-plane CPU sample - platform health)
    "show processes memory",          # NEW-V3.23.167 (processor-pool memory - platform health)
]

# MULTI-VENDOR (Arista EOS): device-native JSON commands ('show ... | json'). Deliberately kept OFF the Cisco
# live-collection lists (COMMANDS_NXOS/IOS) -- they are folded into the OFFLINE analysis union (the --no-collect
# path) so a captured Arista 'show mlag' is read and assessed, WITHOUT making the Cisco collector issue
# Arista-only commands. Live Arista collection (an arista_eos netmiko driver) is a follow-on; offline is the
# universal entry point and needs only this.
COMMANDS_ARISTA = [
    "show mlag",                         # Arista MLAG (multi-chassis link agg) state -> build_arista / _d_arista_mlag_degraded
    "show bgp evpn summary",             # Arista BGP-EVPN overlay control plane -> build_arista / _d_arista_evpn_degraded
]
# MULTI-VENDOR (Juniper Junos): device-native JSON via 'show ... | display json'. Same offline-only fold as
# COMMANDS_ARISTA (kept off the Cisco live lists); a captured 'show chassis cluster status' is read and assessed.
COMMANDS_JUNIPER = [
    "show chassis cluster status",       # Juniper SRX chassis-cluster HA state -> build_juniper / _d_junos_chassis_cluster_degraded
]
# PUBLIC CLOUD (AWS): a cloud account is added as a 'device' (like the ACI/ISE/FMC controllers); its read-only
# export drops into the SAME offline pipeline. Folded into the offline union only (no live collection here).
COMMANDS_CLOUD = [
    "aws ec2 describe-security-groups",  # AWS security-group exposure -> build_cloud / _d_cloud_sg_open_ingress
]
# MULTI-VENDOR (Fortinet FortiGate): CLI show-text, captured per-device; folded into the offline union only.
COMMANDS_FORTINET = [
    "get system ha status",              # FortiGate HA cluster-sync state -> build_fortigate / _d_fortigate_ha_degraded
]
COMMANDS_ALL = list(dict.fromkeys(COMMANDS_NXOS + COMMANDS_IOS))

# =============================================================================
# SSH-WIRE ELIGIBILITY  (whole-repo review 2026-07-28, finding #9)
# =============================================================================
# The COMMANDS_* registries serve TWO channels. The live SSH collector types them at a device exec
# prompt; the OFFLINE (--no-collect) reader uses the SAME strings as capture-FILENAME keys -- and
# cisco_toolkit.rest_collect writes a controller export under its REST ENDPOINT PATH as the "command"
# (`api/…` FMC, `ers/…` ISE, `dataservice/…` vManage), with the APIC managed-object query keyed as
# `moquery -c …`. Those 17 strings are not device CLI, yet collect() sent the UNION of both Cisco
# registries to EVERY device, so all 17 were typed at the exec prompt of live production gear. On IOS
# with `ip domain-lookup` on (the default) an unknown first word is treated as a HOSTNAME to resolve:
# each one costs a DNS lookup plus a Telnet attempt FROM the production switch and stalls that VTY for
# the resolver timeout. COMMANDS_ARISTA/JUNIPER/CLOUD/FORTINET are already kept off the live lists for
# exactly this reason (see the comment above them); these were not. cisco_toolkit.attestation's shared
# read-only grammar already classifies `api/`|`ers/`|`dataservice/` as controller-REST GET paths, i.e.
# a different channel from the SSH read verbs.
#
# The rule is a POSITIVE grammar, not a denylist of the 17: only the CLI `show` read grammar goes on
# the wire, so a registry entry added later for some new controller/cloud channel is excluded BY
# DEFAULT instead of having to be remembered. The strings STAY in the registries -- deleting them would
# blind the offline re-analysis that reads a controller export collected by
# `python -m cisco_toolkit.rest_collect`. (A future APIC-over-SSH channel would be its own platform +
# registry with its own driver, not a moquery smuggled into every switch's command list.)
TERMINAL_SETUP_CMDS = ("terminal length 0", "terminal width 511")   # the ONLY non-`show` wire strings
_SSH_WIRE_CMD = re.compile(r"^show\s+\S")

# The `^show` prefix above only constrains the FIRST WORD. Everything after it was unchecked, so the
# gate admitted strings that START with a read verb and go on to MUTATE the device or egress from it --
# the exact outcomes guardrail 1 forbids, waved through by the guard that exists to forbid them:
#   'show running-config | redirect bootflash:pwn.txt'  -> NX-OS WRITES that file on the device
#   'show running-config | tftp://10.0.0.9/pwn.txt'     -> ships the running-config OFF the device
#   a read verb followed by an embedded NEWLINE and a config-mode line -> netmiko writes the string to
#       the channel verbatim, so every embedded line is typed at the exec prompt, on production gear
#   'show version ; reload' / 'show version && write memory'
# No registry entry does any of this today (the union of characters across all 100 wire-eligible
# commands is letters/digits/space/'-'/'^'/'|', and the single pipe is the read filter
# `show running-config | section ^interface`), so this is a latent hole, not a live defect. It matters
# anyway because the whole POINT of the positive grammar, per the block above, is that a registry entry
# added LATER is excluded BY DEFAULT instead of having to be remembered -- and a write tacked onto a
# read verb is both the most dangerous such entry and the one the prefix check cannot see. It is also
# what `cisco_toolkit.attestation`'s published `read_only_command_surface` claim rests on.
#
# So the check is now WHOLE-STRING: no chaining/redirection/substitution metacharacter anywhere, and a
# pipe may only introduce an output FILTER (an allowlist), never a redirect/tee/append/URL sink.
# A BACKSLASH is in the set for a second reason: collect() derives each capture's FILENAME from the
# command (`cmd.replace(" ","_").replace("|","_").replace("^","").replace("/","_")`). That sanitises the
# POSIX separator but not the Windows one, and the path is then fed to os.makedirs(os.path.dirname(p)) --
# so a backslash-bearing command would write its capture outside the collection directory, creating the
# traversal path on the way. Excluding it here fixes both surfaces at the one gate.
_WIRE_FORBIDDEN = re.compile(r"[\r\n;&`$><\\]")
_WIRE_PIPE_FILTER_OK = re.compile(
    r"^(section|include|exclude|begin|grep|egrep|count|json|xml|no-more|last|head|diff|sort|uniq|in)\b")


def is_ssh_wire_command(cmd) -> bool:
    """True when `cmd` is a device-CLI read this collector may type at an SSH exec prompt.

    A read VERB is necessary but not sufficient: the whole string must carry no way to chain,
    redirect or substitute another command (see the block above)."""
    s = str(cmd).strip()
    if not _SSH_WIRE_CMD.match(s):
        return False
    if _WIRE_FORBIDDEN.search(s):
        return False
    # every pipe stage must be a read-only output filter
    return all(_WIRE_PIPE_FILTER_OK.match(seg.strip()) for seg in s.split("|")[1:])


def ssh_wire_commands(cmds) -> List[str]:
    """`cmds` filtered to those that may go on the SSH wire (order preserved). The complement is
    OFFLINE-only: controller-REST endpoint paths / controller-native queries that a device CLI cannot
    execute and must never be typed at one."""
    return [c for c in cmds if is_ssh_wire_command(c)]


# Interface regex constants moved to cisco_toolkit.textutils (PHASE 2.7 step 1);
# imported near the top of this file.

# =============================================================================
# LOGGING
# =============================================================================
def _attach_package_logging(fh, ch, level):
    """DELIBERATE (and the answer to "desirable or noisy?"): route the whole ``cisco_toolkit.*``
    package tree into the ENGINE's handlers, so library log records reach the run's log file.

    Before this, ``setup_logging`` configured ONLY the ``CiscoMigrationAutofillV3_14_6`` logger. Every
    ``cisco_toolkit`` module uses ``logging.getLogger(__name__)``, which is a different tree, and the
    root logger has no handlers — so package records fell through to ``logging.lastResort``: WARNING+
    went to bare stderr with no timestamp/level, and INFO was DISCARDED ENTIRELY. Consequences:

    1. Gate verdicts were unauditable. ``cisco_toolkit.gate_state`` logs ``[GATE REFUSED]`` /
       ``[GATE OVERRIDDEN]``; grepping the run log for them after a refused run returned nothing,
       for a control whose overrides DEC-003 says are reviewed weekly.
    2. It was a silent REGRESSION, not a design. ``git log -S`` puts the per-sheet ``[OK] '<Sheet>'
       sheet: N row(s)`` lines in this file, on this logger, from the initial commit until 8aa9a4e
       ("PHASE 2.7 step 21") moved the sheet writers to ``cisco_toolkit.excel`` — the same commit
       that removed them from the log. Attaching here RESTORES that output rather than inventing it.

    Volume, MEASURED rather than assumed: at the 2026-07-28 review, ~150 call sites across the package
    (the split was ~95 info / ~32 warning / ~17 error / 5 debug). Treat those as a SHAPE, not a pin --
    an exact count was recorded here once ("132 / 93 / 22 / 12 / 5") and had drifted on every figure but
    `debug` by this review, because a bare number in a docstring has no owner and no reconcile guard
    (SSOT Law 1). Re-measure with an AST walk over ``cisco_toolkit/`` rather than trusting this line.
    The floor is ~80 records a run and CONSTANT — mostly excel's one line per sheet written. Two sites
    in ``cisco_toolkit.build`` are per-device but CONDITIONAL: the per-device ARP line in
    ``collect_global_arp`` fires only when the device returned ARP entries (``if device_total > 0``) and
    the ``[STP]`` line in ``build_interfaces`` only when the config carries
    ``spanning-tree portfast bpduguard default``. So on the synthetic fixtures neither fires and the
    count is flat 81 at 3, 30 and 60 devices; on a real ARP-bearing fleet it grows ~1/device, i.e.
    ~380 at the canonical 303. Either way this is a few hundred lines a run, none per-interface, and
    it is dwarfed by the engine's OWN pre-existing per-device console output (~13 lines/device) — at
    303 devices the package tree is ~2% of the console. That is the right granularity for a
    chain-of-custody log, so both handlers are attached rather than the file alone.
    (File-only would also have SILENCED today's stderr warnings: attaching any handler stops
    ``lastResort`` firing.) ``propagate = False`` stops a later ``basicConfig()``, or a root handler
    that is not stderr, from duplicating or reformatting these records; note that attaching handlers
    here is what stops them being SWALLOWED, since that was only possible while ``lastResort`` was the
    sole path. Because pytest's ``caplog`` captures at the ROOT logger, the repo-root ``conftest.py``
    restores propagation for the duration of each test rather than letting the suite depend on whether
    a given pytest version force-attaches to non-propagating loggers (requirements-dev allows
    >=8,<10). Scope is engine runs only — the webapp drives the engine as a SUBPROCESS
    (``webapp/backend/ingest.py``), so AssessHub's own logging is unaffected.

    NB: this log is half of the audit trail, not a lesser copy of it. The sealed manifest step (see
    ``build_run_manifest``) is tamper-evident but is written only in the final stage; these lines are
    plain text but are written at the INSTANT of the verdict, so they are what survives a crash
    between the two. Neither alone suffices.
    """
    pkg = logging.getLogger("cisco_toolkit")
    pkg.setLevel(level)
    if pkg.handlers:                       # idempotent: re-running setup_logging must not double-log
        pkg.handlers.clear()
    pkg.addHandler(fh)
    pkg.addHandler(ch)
    pkg.propagate = False
    return pkg


def setup_logging(level=logging.INFO):
    logger = logging.getLogger("CiscoMigrationAutofillV3_14_6")
    logger.setLevel(level)
    # REUSE an already-open handler for this log file instead of opening a second mode="w" one.
    # `cisco_toolkit.attestation :: _claim_read_only` does importlib.import_module(_COLLECTOR_MODULE),
    # i.e. import_module("COLLECT_PARSE_V3_23_0") (cited by SYMBOL: the line number this named until the
    # 2026-07-28 review, `attestation.py:122`, had rotted to an unrelated function); under
    # `python COLLECT_PARSE_V3_23_0.py` this module's identity is __main__, so that import RE-EXECUTES
    # the body -- including `logger = setup_logging()` at the bottom -- and a fresh mode="w" handler
    # TRUNCATED the log near the end of every script-path run, discarding most of what it had already
    # recorded. (A module-level "already configured" flag cannot fix this: the re-executed copy gets
    # its own globals. The logging registry is process-global, so the open handler is the reliable
    # witness.) Console-script runs (`cisco-assess`, Atlas's --run-engine) import by real name and were
    # never affected -- which is why this cost the log file its tail without failing anything.
    # `stream is not None` matters: FileHandler.close() sets stream=None/_closed=True, and a closed
    # mode="w" handler REFUSES to reopen on emit -- reusing one would silently discard every
    # subsequent record, the exact failure class this guard exists to prevent. Bound worth knowing:
    # LOG_FILE is relative, so this matches per-CWD. Nothing in the engine chdirs, but a host that
    # does gets a separate log per directory (and, on returning, a fresh mode="w" file).
    fh = next((h for h in logger.handlers
               if isinstance(h, logging.FileHandler)
               and getattr(h, "stream", None) is not None
               and os.path.abspath(getattr(h, "baseFilename", "")) == os.path.abspath(LOG_FILE)), None)
    if logger.handlers:
        logger.handlers.clear()
    if fh is None:
        fh = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
    fh.setLevel(level)
    logger.addHandler(fh)
    ch = logging.StreamHandler()
    ch.setLevel(level)
    ch.setFormatter(logging.Formatter("%(levelname)s - %(message)s"))
    logger.addHandler(ch)
    _attach_package_logging(fh, ch, level)   # the cisco_toolkit.* tree shares this run's handlers
    return logger

logger = setup_logging()

# DATA MODEL - InterfaceData / DevicePhysical moved to cisco_toolkit.model
# (PHASE 2.7 step 9); imported back near the top of this file. ScoringConfig
# stays here - it travels with the analyze layer, not the data model.

# Normalization helpers moved to cisco_toolkit.textutils (PHASE 2.7 step 1);
# imported near the top of this file.

# =============================================================================
# EXCEL HEADER DEBUG
# =============================================================================
def debug_scan_headers(ws, max_rows=120, max_cols=80):
    print("\n=== DEBUG HEADER SCAN START ===")
    for r in range(1, max_rows + 1):
        vals = [(c, ws.cell(row=r, column=c).value.strip()[:60])
                for c in range(1, min(ws.max_column, max_cols)+1)
                if isinstance(ws.cell(row=r, column=c).value, str)
                and ws.cell(row=r, column=c).value.strip()]
        if vals:
            print(f"ROW {r}: {vals[:18]}")
    print("=== DEBUG HEADER SCAN END ===\n")

# =============================================================================
# PLATFORM AUTODETECT
# =============================================================================
def _close_detect_session(guesser) -> None:
    """Close the SSH session SSHDetect opened, fail-soft (review 2026-07-28 #89).

    ``SSHDetect.__init__`` CONNECTS (it builds a ConnectHandler and does a channel read), and
    ``autodetect()`` disconnects on each of its own return paths -- but not when it RAISES, which is
    the ordinary outcome of a slow/odd device: a read timeout mid-probe. The bare ``except`` below
    then returned 'ios' and dropped the last reference to an ESTABLISHED session, leaking a VTY line
    on production gear until the device's exec-timeout reaps it. On the default `platform: auto` path
    that is one leaked line per device, per run, before the real collection even opens its own.
    netmiko's ``disconnect()`` is idempotent and swallows its own errors, so calling it on the
    already-closed success path is safe. (If the CONSTRUCTOR raised there is no object to close --
    netmiko's own ConnectHandler cleans up on a failed establish.)"""
    conn = getattr(guesser, "connection", None)
    if conn is None:
        return
    try:
        conn.disconnect()
    except Exception as e:                                             # noqa: BLE001
        logger.debug(f"autodetect session close failed (ignored): {e}")


def autodetect_platform(ip: str, username: str, password: str) -> str:
    if SSHDetect is None: return "ios"
    guesser = None
    try:
        guesser = SSHDetect(device_type="autodetect", host=ip,
                            username=username, password=password)
        best = (guesser.autodetect() or "").strip().lower()
        if "nxos" in best: return "nxos"
        return "ios"
    except Exception as e:
        logger.debug(f"autodetect_platform({ip}) failed; defaulting to ios: {e}")  # NEW-V3.23.1
        return "ios"
    finally:
        _close_detect_session(guesser)

def detect_platform_from_files(dev_dir: str) -> str:
    nxos_markers = ["show_interface_brief.txt","show_interface_switchport.txt",
                    "show_port-channel_summary.txt","show_running-config_interface.txt"]
    ios_markers  = ["show_interfaces_switchport.txt","show_etherchannel_summary.txt",
                    "show_running-config___section_interface.txt"]
    def has_real(fname):
        p = os.path.join(dev_dir, fname)
        if not os.path.isfile(p): return False
        try:
            with open(p, "r", encoding="utf-8", errors="ignore") as f:
                c = f.read(512).strip()
            # V14.12: a command that errored on the wrong platform is NOT a valid marker.
            low = c[:200].lower()
            if any(pat in low for pat in _CISCO_ERRORS): return False
            return len(c) > 10
        except Exception as e:
            logger.debug(f"detect_platform_from_files: unreadable marker {fname}: {e}")  # NEW-V3.23.1
            return False
    nxos_score = sum(1 for f in nxos_markers if has_real(f))
    ios_score  = sum(1 for f in ios_markers  if has_real(f))
    return "nxos" if nxos_score > ios_score else "ios"

# =============================================================================
# SSH COLLECTION
# =============================================================================
# NEW-V3.23.1: bounded connection retry. A one-shot audit shouldn't lose a live
# switch to a momentary SSH timeout. Transient/timeout failures get a few backoff
# retries; AUTH failures are never retried (a bad credential won't fix itself and
# repeated tries can lock the account). Set CONNECT_MAX_ATTEMPTS = 1 to restore the
# original single-shot behaviour.
CONNECT_MAX_ATTEMPTS = 3        # total attempts per device (1 = no retry)
CONNECT_BACKOFF_BASE = 2.0      # seconds between attempts: 2s, 4s, ...

try:  # netmiko moved its exception classes across versions - stay tolerant
    from netmiko.exceptions import (NetmikoAuthenticationException,
                                    NetmikoTimeoutException)
except Exception:
    try:
        from netmiko.ssh_exception import (NetmikoAuthenticationException,
                                           NetmikoTimeoutException)
    except Exception:
        NetmikoAuthenticationException = NetmikoTimeoutException = None

def _is_auth_error(exc) -> bool:
    """True for an authentication failure (these are NOT retried)."""
    if NetmikoAuthenticationException is not None and \
       isinstance(exc, NetmikoAuthenticationException):
        return True
    m = str(exc).lower()  # fallback if the class isn't importable
    return "authentication" in m or "auth failed" in m or "bad password" in m

def connect_device(ip, hostname, username, password, platform):
    resolved = platform
    if platform in ("auto", ""):                       # CHANGED-V3.23.1: detect once,
        resolved = autodetect_platform(ip, username, password)  # not per retry
    attempts = max(1, CONNECT_MAX_ATTEMPTS)
    last_err = None
    for attempt in range(1, attempts + 1):
        try:
            logger.info(f"Connecting to {hostname} ({ip}) [{resolved}] "
                        f"(attempt {attempt}/{attempts}) ...")
            dev = ConnectHandler(
                device_type=NETMIKO_TYPE.get(resolved, "cisco_ios"),
                host=ip, username=username, password=password,
                timeout=120, conn_timeout=30, auth_timeout=30,
                global_delay_factor=3, fast_cli=False,
            )
            # Session setup. A FAILURE here is not cosmetic (review 2026-07-28 #11): it was swallowed
            # by a bare `except Exception: pass` with no log and no record, yet it changes what every
            # later capture MEANS. Under a TACACS+ command-authorization policy that denies `terminal *`
            # -- exactly the narrow read-only account this repo's own security doctrine recommends --
            # paging stays on and the width stays at 80, so captures come back pager-truncated and
            # line-wrapped, and the fixed-column parsers (extract_fixed_cols & friends) return
            # plausible-but-wrong values from the wrapped text. Log it loudly and hand the record to
            # collect(), which persists it into the _capture_meta.json sidecar so an offline
            # re-analysis can distrust the run instead of reading clean-looking text as clean state.
            setup_failures: Dict[str, str] = {}
            for cmd in TERMINAL_SETUP_CMDS:
                try:
                    dev.send_command(cmd, read_timeout=30)
                except Exception as e:                                 # noqa: BLE001
                    setup_failures[cmd] = f"{type(e).__name__}: {e}"
            if setup_failures:
                logger.warning(
                    f"  [TERMINAL-SETUP] {hostname}: {len(setup_failures)} of {len(TERMINAL_SETUP_CMDS)} "
                    f"session-setup command(s) FAILED "
                    f"({'; '.join(f'{c!r} -> {r}' for c, r in setup_failures.items())}). Paging and/or the "
                    f"80-column wrap may still be ON (a command-authorization policy denying 'terminal *' "
                    f"does exactly this), so this device's captures can be truncated or wrapped and the "
                    f"fixed-column parsers can yield plausible-but-wrong values -- treat every capture from "
                    f"{hostname} as unproven, not as observed state.")
            try:
                setattr(dev, TERMINAL_SETUP_ATTR, setup_failures)
            except Exception as e:                                     # noqa: BLE001
                logger.warning(f"  [TERMINAL-SETUP] could not record the setup result on the "
                               f"{hostname} connection ({e}); the sidecar will not carry it")
            logger.info(f"[OK] Connected to {hostname}")
            return dev, resolved
        except Exception as e:
            last_err = e
            if _is_auth_error(e):                       # don't retry bad credentials
                logger.error(f"[FAIL] Authentication failed to {hostname}: {e} (not retrying)")
                return None, resolved
            if attempt < attempts:
                wait = CONNECT_BACKOFF_BASE * attempt
                logger.warning(f"[RETRY] Connect to {hostname} failed "
                               f"(attempt {attempt}/{attempts}): {e}; retrying in {wait:.0f}s ...")
                time.sleep(wait)
    logger.error(f"[FAIL] Connection failed to {hostname} after {attempts} attempt(s): {last_err}")
    return None, resolved

_SLOW_CMDS = {"show running-config", "show cdp neighbors detail",
              "show lldp neighbors detail", "show mac address-table",
              "show interface trunk", "show interfaces trunk",
              "show interfaces", "show interface"}  # NEW-V15 (full counter dumps)

# Plan A / Tier-2 #6: per-thread flags for the LAST send_cmd call.
#   .fallback — True when the read fell back to send_command_timing (prompt never confirmed =>
#               the capture's completeness is unproven).
#   .failed   — True when BOTH transports raised, i.e. NOTHING was read from the device
#               (review 2026-07-28 #10); "" is then a transport failure, not an answer.
# collect() reads both after each command and persists them as the _capture_meta.json sidecar,
# so an offline re-analysis can flag those captures `unverified_prompt` / not-collected.
_SEND_TRANSPORT = threading.local()

# connect_device stashes {terminal-setup command: reason} here on the live connection object so
# collect() can persist it; see the [TERMINAL-SETUP] block in connect_device (review #11).
TERMINAL_SETUP_ATTR = "_assess_terminal_setup_failures"

def send_cmd(dev, cmd: str) -> str:
    # FIX (C2): prefer pattern-based send_command() (waits for the device prompt)
    # over send_command_timing(), which can return early/truncated output that is
    # then parsed as if complete. read_timeout is sized up for slow/large dumps.
    # Fall back to the timing-based read only if the prompt pattern isn't matched,
    # so behaviour is no worse than before on edge cases.
    timeout = 300 if any(s in cmd for s in _SLOW_CMDS) else 120
    _SEND_TRANSPORT.fallback = False
    _SEND_TRANSPORT.failed = False
    try:
        return dev.send_command(cmd, read_timeout=timeout) or ""
    except Exception as e:
        logger.warning(f"Command '{cmd}' pattern-read failed ({e}); retrying timing-based")
        # Review 2026-07-28 #90: a read timeout does NOT stop the device -- the rest of THIS
        # execution is still arriving on the channel. Re-sending immediately made
        # send_command_timing return that tail PREPENDED to the second execution's output: one
        # capture spliced from two executions (and when the tail belonged to a `show running-config`,
        # config text landing in an unrelated command's file, parsed as that command's evidence).
        # Drain the channel first so the retry reads only what it asked for; the discarded bytes are
        # LOGGED, never silently folded into a capture.
        try:
            residue = dev.clear_buffer() or ""
            if residue:
                logger.warning(f"  [wire] discarded {len(residue)} byte(s) of unread output pending on "
                               f"the channel before re-sending '{cmd}' — they belong to the timed-out "
                               f"execution, not to this capture")
        except Exception as e3:                                        # noqa: BLE001
            logger.warning(f"  [wire] could not drain the channel before re-sending '{cmd}' ({e3}); "
                           f"the timing-based capture may splice the timed-out execution's output")
        try:
            out = dev.send_command_timing(cmd, read_timeout=timeout) or ""
            _SEND_TRANSPORT.fallback = True     # only a SUCCESSFUL timing read is an unverified capture
            return out
        except Exception as e2:
            logger.warning(f"Command failed '{cmd}': {e2}")
            _SEND_TRANSPORT.failed = True       # BOTH transports failed: "" is no answer at all
            return ""

def collect(hostname: str, platform: str, dev, out_dir: str,
            archive_all_output: bool = False,
            include_config_backup: bool = False,
            include_heavy_commands: bool = False) -> Dict[str, str]:
    os.makedirs(out_dir, exist_ok=True)
    # V14.12: collect the UNION of command variants, ordered by the detected platform.
    # The two lists differ only in platform-divergent forms (show standby vs show hsrp,
    # the two run-config forms, etc.); collecting both makes the run resilient to a wrong
    # platform label - _load_cmd_output() already skips the form that errors. Extra commands
    # on the wrong platform just return a quick error and are ignored.
    if platform == "nxos":
        cmds = list(dict.fromkeys(COMMANDS_NXOS + COMMANDS_IOS))
    elif platform == "ios":
        cmds = list(dict.fromkeys(COMMANDS_IOS + COMMANDS_NXOS))
    else:
        cmds = COMMANDS_ALL

    extra_cmds: List[str] = []
    if include_config_backup:
        extra_cmds.append("show startup-config")
    if include_heavy_commands:
        extra_cmds.extend(["show logging", "show tech-support"])

    cmds = list(dict.fromkeys(cmds + extra_cmds))
    # Review 2026-07-28 #9: the registries are ALSO the offline capture-filename keys, so they carry
    # controller-REST endpoint paths / controller-native queries. Those must never be typed at a device
    # exec prompt (see the SSH-WIRE ELIGIBILITY block above). Filter here, at the one place that talks
    # to the wire — the registries stay whole for --no-collect.
    wire_cmds = ssh_wire_commands(cmds)
    withheld = [c for c in cmds if not is_ssh_wire_command(c)]
    if withheld:
        logger.info(f"  [wire] {len(withheld)} non-CLI registry string(s) withheld from the SSH session "
                    f"on {hostname} (controller-REST / controller-native; collected read-only by "
                    f"cisco_toolkit.rest_collect and read back offline): {', '.join(withheld)}")
    cmds = wire_cmds
    paths: Dict[str, str] = {}
    command_index: List[dict] = []
    fallback_cmds: Dict[str, str] = {}   # Tier-2 #6: prompt-unverified captures this device
    # Review #11: a denied/failed `terminal length 0` / `terminal width 511` changes what EVERY capture
    # below means (paging on / 80-col wrap -> truncated + wrapped text through fixed-column parsers).
    # connect_device already logged it loudly; persist it so an offline re-analysis can distrust the run.
    for _tcmd, _twhy in (getattr(dev, TERMINAL_SETUP_ATTR, None) or {}).items():
        fallback_cmds[_tcmd] = "terminal_setup_failed"
        logger.warning(f"  [capture-meta] {hostname}: recording terminal_setup_failed for {_tcmd!r} "
                       f"({_twhy}) — captures from this device are unproven, not observed state")
    started = datetime.now().isoformat()

    for cmd in cmds:
        logger.info(f"  Executing: {cmd}")
        out = send_cmd(dev, cmd)
        if getattr(_SEND_TRANSPORT, "failed", False):
            # Review 2026-07-28 #10: BOTH transports raised — nothing was read from the device. Writing
            # "" here produced a capture FILE with paths[cmd] set, and cmd_capture_state ranks `empty`
            # ABOVE `error`, so a device whose SSH session had died read as one that positively reported
            # nothing to report. Record the failure and omit the path: the command is then `missing` — a
            # blind spot — which is the truth. ("Not observed" must never become "healthy".)
            fallback_cmds[cmd] = "send_failed"
            logger.warning(f"  [capture] '{cmd}' produced NO capture on {hostname} (both transports "
                           f"failed) — recorded as send_failed and reported NOT COLLECTED, never empty")
            continue
        if getattr(_SEND_TRANSPORT, "fallback", False):
            fallback_cmds[cmd] = "timing_fallback"
        fn  = cmd.replace(" ","_").replace("|","_").replace("^","").replace("/","_") + ".txt"
        p   = os.path.join(out_dir, fn)
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            f.write(out)
        paths[cmd] = p

        if archive_all_output:
            size = 0
            sha = ""
            try:
                size = os.path.getsize(p)
                sha = file_sha256(p)
            except Exception:
                pass
            command_index.append({
                "command": cmd,
                "file": fn,
                "path": os.path.abspath(p),
                "bytes": size,
                "sha256": sha,
                "collected_at": datetime.now().isoformat(),
                "has_output": bool((out or "").strip()),
            })

    if archive_all_output:
        device_info = {
            "hostname": hostname,
            "platform": platform,
            "output_dir": os.path.abspath(out_dir),
            "collection_started": started,
            "collection_finished": datetime.now().isoformat(),
            "command_count": len(command_index),
        }
        write_json_file(os.path.join(out_dir, "device_info.json"), device_info)
        write_json_file(os.path.join(out_dir, "command_index.json"), {"commands": command_index})

    # Tier-2 #6 (+ review 2026-07-28 #10/#11): persist what is KNOWN to be wrong with this
    # device's collection — `timing_fallback` (prompt-unverified capture, offline-flagged
    # `unverified_prompt`), `send_failed` (both transports failed, no capture written at all),
    # `terminal_setup_failed` (paging/width setup denied, so every capture is unproven). Written
    # only when non-empty; the offline loader keys captures by KNOWN command names, so this
    # sidecar can never be read as a capture. Fail-soft: a sidecar write error never
    # loses the collection.
    if fallback_cmds:
        try:
            write_json_file(os.path.join(out_dir, CAPTURE_META_FILENAME), fallback_cmds)
            _by_reason = {r: sum(1 for v in fallback_cmds.values() if v == r)
                          for r in sorted(set(fallback_cmds.values()))}
            logger.info(f"  [capture-meta] {len(fallback_cmds)} collection defect(s) recorded in "
                        f"{CAPTURE_META_FILENAME}: {_by_reason}")
        except Exception as e:
            logger.warning(f"  capture-meta write failed (non-fatal): {e}")

    return paths

# =============================================================================
# SAFE FILE LOADER
# =============================================================================
# _CISCO_ERRORS / _load_cmd_output / _safe_parse moved to cisco_toolkit.cmdio
# (PHASE 2.7 step 14); all three imported back near the top of this file (nearly
# every build_*/compute_*/write_* reads collected output via _load_cmd_output, and
# platform detection reuses _CISCO_ERRORS).

# =============================================================================
# PARSERS - interface commands
# =============================================================================
# extract_fixed_cols / slice_col moved to cisco_toolkit.parse (PHASE 2.7 step 2);
# imported near the top of this file.

# parse_show_interface_status / _switchport / _trunk_table moved to
# cisco_toolkit.parse (PHASE 2.7 step 4); imported at the top of this file.

# MAC / STP / vlan_brief / _compress_vlans moved to cisco_toolkit.parse
# (PHASE 2.7 step 5); imported at the top of this file.

# VRF / PoE / CDP / LLDP / infer_endpoint_type moved to cisco_toolkit.parse
# (PHASE 2.7 step 5); imported at the top of this file.

# run-config / etherchannel proto+members / ip-arp parsers moved to
# cisco_toolkit.parse (PHASE 2.7 step 6); imported at the top of this file.

# vtp_status / _parse_ip_int_brief / parse_switch_mgmt_ip moved to
# cisco_toolkit.parse (PHASE 2.7 step 6); imported at the top of this file.

# parse_ip_routes moved to cisco_toolkit.parse (PHASE 2.7 step 3); imported at top.

# parse_hsrp_summary / parse_vrrp_summary / parse_glbp_summary moved to
# cisco_toolkit.parse (PHASE 2.7 step 3); imported at the top of this file.

# parse_multicast_info moved to cisco_toolkit.parse (PHASE 2.7 step 6);
# imported at the top of this file.

# =============================================================================
# PHYSICAL PARSERS (NEW-V12)
# =============================================================================

# parse_show_version / _inv_commit / parse_show_inventory moved to
# cisco_toolkit.parse (PHASE 2.7 step 7); imported at the top of this file.


# parse_show_environment_power / parse_show_environment / parse_show_module_count
# moved to cisco_toolkit.parse (PHASE 2.7 step 7); imported at the top of this file.

# =============================================================================
# BUILD PHYSICAL DEVICE SUMMARY (NEW-V12)
# =============================================================================
# build_device_physical moved to cisco_toolkit.build (PHASE 2.7 step 27); imported
# back near the top of this file (main() calls it per device).

# =============================================================================
# SITE SURVEY SHEET WRITER (NEW-V12)
# =============================================================================
# Census/inventory sheet writers moved to cisco_toolkit.excel (PHASE 2.7 step 21):
# write_device_inventory_sheet / write_svi_gateway_sheet / write_stp_detail_sheet /
# write_vlan_census_sheet / write_endpoint_census_sheet (+ their *_COLUMNS / *_SHEET_NAME
# constants + _is_svi). All five writers imported back near the top of this file (main()
# calls them); the constants + _is_svi are package-internal.

# (SVI / STP Detail / VLAN Census / Endpoint Census writers also moved to
# cisco_toolkit.excel in step 21 - see the note above.)


# =============================================================================
# NEW-V14.9: Move-group planning. Switches that share an L2 broadcast domain must
# migrate together; connected components over "shared-VLAN" edges = migration waves.
# Captures transitive coupling (A-VLAN10-B, B-VLAN20-C => A,B,C are one group) that
# is easy to miss by eye. Coupling uses ACTUAL VLAN presence (access port or SVI),
# consistent with the VLAN Census; trunk-allowed-only transit is NOT modeled, and
# VLAN 1 (default, present everywhere) is excluded from grouping by design.
# =============================================================================
# Analysis sheet writers (move-group / topology / findings / capacity / topology-diagram)
# moved to cisco_toolkit.excel (PHASE 2.7 step 22); the five writers are imported back near
# the top of this file (main() calls them). They call the analyze compute_* internally now,
# so compute_move_groups / compute_topology_links / compute_findings are no longer used here.


# build_switch_identity moved to cisco_toolkit.build (PHASE 2.7 step 27); imported
# back near the top of this file (main() calls it per device).

# =============================================================================
# BUILD INTERFACE DB (per device)
# =============================================================================
# build_interfaces (the per-device InterfaceData builder) moved to cisco_toolkit.build
# (PHASE 2.7 step 28); imported back near the top of this file (main() calls it per device).

# =============================================================================
# GLOBAL ARP COLLECTION (FIX-R1, FIX-R3, FIX-R5, NEW-V11)
# =============================================================================
# collect_global_arp / apply_global_arp / detect_cross_device_dual_connections moved to
# cisco_toolkit.build (PHASE 2.7 step 27); imported back near the top of this file
# (main() calls them during the global-ARP enrichment + dual-connection pass).

# =============================================================================
# EXCEL WRITER
# =============================================================================
# norm_header / HEADER_TO_FIELD / find_header_row / ensure_headers / sortkey moved to
# cisco_toolkit.excel (PHASE 2.7 step 20); find_header_row / ensure_headers (main()) and
# sortkey (append_interface_rows below) are imported back near the top of this file.
# norm_header / HEADER_TO_FIELD are package-internal.

# append_interface_rows (the main 'Migration Assessment' interface-sheet filler) moved to
# cisco_toolkit.excel (PHASE 2.7 step 26); imported back near the top of this file (main()
# calls it per device). It's the last sheet writer - the whole excel layer now lives in
# cisco_toolkit/excel.py.

# =============================================================================
# DEVICES LOADER
# =============================================================================
def load_devices(devices_file: str, allow_prompt: bool = True) -> List[dict]:
    # FIX-V3.23.177: allow_prompt gates ONLY the interactive getpass fallback (chain step 4).
    # main() passes allow_prompt=not args.no_collect: an offline re-analysis never opens SSH,
    # so blocking the whole pipeline on a TTY password prompt for credentials that are never
    # used hung real runs (password-less devices.json + TTY stdin, perf harness 2026-07-02).
    # The non-blocking env chain (steps 1-3) always runs; default True keeps every other
    # caller -- and live collection -- on the full V3.23.1 behaviour.
    if not os.path.isfile(devices_file):
        raise FileNotFoundError(f"Devices file not found: {devices_file}")
    with open(devices_file, "r", encoding="utf-8-sig") as f:
        raw = f.read().strip()
    if not raw: raise ValueError(f"Devices file is empty: {devices_file}")
    data = None
    try:
        parsed = json.loads(raw)
        data = parsed if isinstance(parsed, list) else [parsed]
    except json.JSONDecodeError:
        pass
    if data is None:
        try:
            items = [json.loads(l) for l in raw.splitlines() if l.strip()]
            if items: data = items
        except json.JSONDecodeError:
            pass
    if data is None:
        decoder = json.JSONDecoder()
        idx, items = 0, []
        while idx < len(raw):
            mv = re.search(r"[^\s,]", raw[idx:])
            if not mv: break
            idx += mv.start()
            try:
                obj, end_idx = decoder.raw_decode(raw, idx)
                items.append(obj); idx = end_idx
            except json.JSONDecodeError:
                break
        if items: data = items
    if data is None:
        raise ValueError(f"Could not parse {devices_file} as valid JSON.")
    plat_map = {
        "cisco_nxos":"nxos","nxos":"nxos","nexus":"nxos","nx-os":"nxos",
        "cisco_ios":"ios","ios":"ios","ios-xe":"ios","iosxe":"ios",
        "cisco_xe":"ios","xe":"ios","auto":"auto","autodetect":"auto","":"auto",
    }
    for d in data:
        if not isinstance(d, dict):
            # Same reasoning as the missing-key raise below: the value is caller-supplied and can
            # carry a credential (a bare "user:pass@host" string), so report its TYPE, not its bytes.
            raise ValueError(f"Each devices.json entry must be a JSON object, got {type(d).__name__}")
        for alias, key in [("host","ip"),("address","ip"),("name","hostname"),
                           ("user","username"),("pass","password"),("secret","password")]:
            if alias in d and key not in d: d[key] = d[alias]
        for k in ("ip","hostname","username"):     # CHANGED-V3.23.1: 'password' no longer
            if k not in d:                          # hard-required - it may come from env/getpass
                # Name the entry by its IDENTIFYING keys only, never by dumping `d`. The alias pass
                # three lines above has already copied `pass`/`secret` into d['password'], so
                # formatting the whole entry put a production device credential in cleartext into an
                # uncaught traceback -- and, driven from AssessHub, into the web UI's error text
                # (ingest.py concatenates the engine's stdout+stderr into an EngineRunError message).
                ident = ", ".join(f"{a}={d[a]!r}" for a in ("hostname", "ip") if a in d) or "<unnamed>"
                raise ValueError(f"Missing '{k}' in devices.json entry ({ident}). "
                                 f"Entry keys present: {sorted(d)}")
        for k in ("ip","hostname","username"):
            d[k] = (d.get(k) or "").strip()
        # NEW-V3.23.1: resolve the password WITHOUT forcing plaintext into the file.
        #   1) explicit "password" in the entry  (unchanged - fully back-compatible)
        #   2) per-entry  "password_env": "VAR"   -> os.environ["VAR"]
        #   3) global      $CISCO_PASS
        #   4) (after the loop) a secure getpass prompt, only on an interactive TTY
        #      AND only when prompting is allowed (live collection; never --no-collect)
        pw = d.get("password") or ""
        if not pw and d.get("password_env"):
            pw = os.environ.get(d["password_env"], "")
        if not pw:
            pw = os.environ.get("CISCO_PASS", "")
        d["password"] = pw
        plat_raw = (d.get("platform") or d.get("device_type") or d.get("os") or d.get("nos") or "auto")
        d["platform"] = plat_map.get(plat_raw.strip().lower(), "auto")

    # NEW-V3.23.1: prompt securely for any device still lacking a password - but ONLY when
    # attached to a terminal. An unattended/batch run must not block on input, so it keeps the
    # previous blank-password behaviour (those devices then fail auth visibly: connect_device
    # logs the auth failure and, per V3.23.1, does not retry it). Prompt once per username.
    need_pw = [d for d in data if not d["password"]]
    if need_pw:
        if not allow_prompt:
            # FIX-V3.23.177: offline run -- credentials are never used, so neither block on a
            # prompt nor nag about $CISCO_PASS; leave them blank with a debug breadcrumb only.
            logger.debug(f"  {len(need_pw)} device(s) without a password; offline run "
                         f"(--no-collect) never uses credentials - not prompting.")
        elif sys.stdin.isatty():
            import getpass
            by_user: Dict[str, str] = {}
            for d in need_pw:
                u = d["username"]
                if u not in by_user:
                    by_user[u] = getpass.getpass(f"  SSH password for '{u}': ")
                d["password"] = by_user[u]
        else:
            hosts = ", ".join(sorted({d["hostname"] for d in need_pw}))
            logger.warning(f"  [WARN] No password for {len(need_pw)} device(s) "
                           f"(no entry, no password_env, no $CISCO_PASS): {hosts}. "
                           f"Set $CISCO_PASS or run interactively to be prompted.")
    return data


# os.replace on Windows must DELETE the destination, so it fails while ANY process holds a handle
# to it -- an on-access AV scan, the search indexer, a viewer the engineer left open on the stick.
# A short retry absorbs the transient scanners without making the caller wait meaningfully.
_ATOMIC_REPLACE_ATTEMPTS = 4
_ATOMIC_REPLACE_BACKOFF_S = 0.1


def _write_json_atomic(path: str, dump) -> None:
    """Publish `path` atomically: same-directory tmp -> fsync -> os.replace. Raises on any failure,
    and never leaves the temp file behind (including on KeyboardInterrupt/SIGTERM)."""
    # Same directory keeps os.replace atomic (it is not, across filesystems); pid-stamped so two
    # concurrent runs sharing an output dir cannot collide on the temp name.
    tmp = "%s.%d.tmp" % (path, os.getpid())
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            dump(f)
            f.flush()
            os.fsync(f.fileno())          # bytes reach the medium BEFORE the rename publishes them
        for attempt in range(_ATOMIC_REPLACE_ATTEMPTS):
            try:
                os.replace(tmp, path)
                return
            except OSError:
                if attempt == _ATOMIC_REPLACE_ATTEMPTS - 1:
                    raise
                time.sleep(_ATOMIC_REPLACE_BACKOFF_S)
    except BaseException:
        # BaseException, not Exception: KeyboardInterrupt / SIGTERM is precisely the unplug-and-Ctrl-C
        # case this exists for, and a leaked .tmp is itself a confusing audit artifact (it would
        # otherwise be hashed into the next run's sealed artifact set).
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def write_json_file(path: str, data: dict, compact: bool = False) -> None:
    """Crash-safe JSON write: ATOMIC where the filesystem allows it, with a LOUD degrade where it
    does not -- never a hard failure, because a plain truncate-write is what this used to do and it
    succeeds in cases os.replace cannot.

    Why atomic at all: the plain ``open(path, "w")`` this replaced truncates the good file BEFORE
    writing the new bytes, so a crash in that window (ADR-0004 P3: a USB unplug mid-write) left
    neither the old file nor a complete new one. That matters most for the two files here that are
    EVIDENCE, not scratch: the snapshot, and the sealed run manifest whose whole purpose is
    tamper-evidence -- a torn write destroys exactly the guarantee ``chain_root`` provides.

    Why the fallback is NOT optional: on Windows ``os.replace`` must delete the destination, so it
    raises PermissionError while any process holds a handle -- where the old truncate-write simply
    succeeded (a reader opened via the CRT shares read+write). Failing there would be a REGRESSION,
    and worse than the bug being fixed: the caller logs a warning and ships the run, leaving the
    PREVIOUS manifest in place beside freshly rewritten artifacts -- a stale seal that still passes
    verify_manifest while every hash in it is wrong. A torn write at least announces itself; a stale
    one certifies a lie. So we retry briefly, then fall back to the in-place write and say so.

    The fallback is deliberately noisy: silent degradation of a durability guarantee is the failure
    mode --selftest exists to prevent elsewhere in this codebase.
    """
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    def _dump(fh):
        # NEW-Tier3#14: the SNAPSHOT is written compact (a pure size win — every consumer json.loads it
        # and compares PARSED objects, so byte formatting is invisible to the golden/--compare contract);
        # the run manifest / phase timings stay indented (human-readable) by default.
        if compact:
            json.dump(data, fh, separators=(",", ":"), ensure_ascii=False)
        else:
            json.dump(data, fh, indent=2, ensure_ascii=False)

    try:
        _write_json_atomic(path, _dump)
        return
    except (KeyboardInterrupt, SystemExit):
        # Nothing was published and the temp is gone, so the existing file is intact -- exactly the
        # outcome this function exists to guarantee. Never downgrade an interrupt into a retry.
        raise
    except Exception as e:
        logger.warning(
            "  Atomic write unavailable for %s (%s: %s) -- falling back to an in-place write. "
            "The file is NOT crash-safe for this write: a failure now can leave it truncated. "
            "Usual cause on Windows is another process holding the file open (AV scan, indexer, an "
            "open viewer); close it and re-run to restore the atomic path.",
            os.path.basename(path), type(e).__name__, e)

    with open(path, "w", encoding="utf-8") as f:
        _dump(f)


def file_sha256(path: str) -> str:
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def build_run_manifest(out_xlsx: str, snap_dict: dict) -> dict:
    """roadmap D2 + J4: a sealed, deterministic chain-of-custody manifest for one assessment run — a
    hash-chained ledger of the pipeline stages PLUS a per-artifact sha256 of every produced deliverable
    PLUS the coverage-honest abstention ledger, sealed by ``chain_root`` (no Git dependency). Pure
    stdlib via cisco_toolkit.manifest; the GAIT-style audit trail, offline.

    The pipeline EMITS this seal and never re-checks it: verification is the receiver's deliberate act
    (``python -m cisco_toolkit.manifest verify`` / ``Atlas.exe --verify-manifest``). Scope the claim —
    the chain is unkeyed and ``manifest.build_manifest`` is public, so it detects a careless edit or a
    truncation, NOT a forger who re-seals; only a ``chain_root`` recorded out of band pins a delivered
    file to this run.

    SCOPE — this is a PER-RUN seal, not a cross-run archive. It describes the artifact set currently
    on disk beside ``out_xlsx``, and a re-run to the same ``--output`` replaces it exactly as it
    replaces every sibling artifact it hashes; a manifest outliving the artifacts it seals would
    describe files that no longer exist. The hash chain is append-only WITHIN one file. The durable
    cross-run record of human gate decisions is the gate-state ledger
    (``docs/engagement-state.json``), which is append-only ACROSS runs by design.
    """
    import glob as _glob
    from cisco_toolkit import gate_state as _gate_state, manifest as _manifest, ssot as _ssot
    try:
        from cisco_toolkit import __version__ as _schema_version
    except Exception:
        _schema_version = "unknown"
    base = os.path.splitext(os.path.abspath(out_xlsx))[0]
    out_dir = os.path.dirname(base) or "."
    base_name = os.path.basename(base)
    artifacts = {}
    for path in sorted(_glob.glob(os.path.join(out_dir, base_name + "*"))):
        # Skip .tmp: a write killed uncatchably (TerminateProcess / power loss) leaves the atomic
        # writer's temp behind, and a CONCURRENT run's temp is visible for the duration of its write.
        # Either would otherwise be sealed as a delivered artifact -- publishing, under the seal,
        # bytes the pipeline deliberately never published.
        if os.path.isfile(path) and not path.endswith((".run_manifest.json", ".tmp")):
            try:
                with open(path, "rb") as _fh:
                    artifacts[os.path.basename(path)] = _manifest.artifact_sha256(_fh.read())
            except OSError:
                continue
    cc = (snap_dict.get("collection_completeness") or {}).get("summary") or {}
    steps = [
        {"stage": "collect", "inventory": cc.get("inventory"), "collected": cc.get("complete")},
        {"stage": "analyze", "punchlist_findings": len(snap_dict.get("punchlist") or [])},
        # P0-3/DEC-003: the PPDIOO document-gate verdicts, as DATA. Deliberately a sealed step and not
        # a loose top-level key: `build_manifest` hash-chains `steps` but silently DROPS any meta key
        # outside its fixed allowlist, so a top-level key would be both unsealed and discarded. Sealed
        # here, editing or deleting a refusal breaks `chain_root`. Scope that honestly: the chain is
        # UNKEYED and `manifest.build_manifest` is public, so this detects careless edits, not a
        # forger who re-seals -- and nothing in the product or CI verifies a delivered manifest today.
        # Placed before "deliver" because a gate decides whether that step's artifact exists at all.
        # Do NOT read the two steps as a cross-check: the artifact glob hashes whatever is on disk,
        # so a stale *_design.docx from an earlier run to the same --output appears beside a "refused"
        # verdict. An empty list means NO gate ran (e.g. --no-design --no-mop); coverage-honest, it is
        # not "gates passed".
        #
        # Log and seal are COMPLEMENTARY, not ranked -- each covers the other's blind spot. The log
        # line is written at the INSTANT of the verdict but is plain text anyone can edit; this row is
        # tamper-evident but is only written here, in the last stage. Everything between the gates and
        # this point (four DOCX renders, ~4s on a 3-device fixture) is a window where a crash, a
        # SIGTERM or AssessHub's engine timeout loses the seal and leaves ONLY the log. That is why
        # the change attaches the package logger as well as sealing the ledger; neither alone suffices.
        {"stage": "gate", "verdicts": _gate_state.verdicts()},
        {"stage": "deliver", "artifacts": sorted(artifacts)},
    ]
    ledger = {subj: _ssot.abstention_reason(snap_dict, subj) for subj in
              ("fhrp", "vpc", "golden_drift", "feature_compliance", "acl_line_reachability", "framework_coverage")}
    meta = {"schema_version": _schema_version, "generated_at": datetime.now().isoformat(),
            "collected_at": snap_dict.get("collected_at"), "abstention_ledger": ledger}
    return _manifest.build_manifest(meta, artifacts, steps)


def _derive_collected_at(no_collect: bool, collection_dir: str, root_dir: str):
    """Provenance: the instant the EVIDENCE was collected (NOT wall-clock-at-regen). Returns
    (iso_datetime, defaulted). A live run stamps now() -- collection IS happening now. A
    `--no-collect` re-analysis instead RECOVERS the original collection time so re-rendering old
    evidence reproduces the original lifecycle bands + cover date byte-for-byte: first from the
    `YYYYMMDD_HHMMSS` stamp in the collection-dir name (how every run names its dir), else the
    earliest member-file mtime, else -- last resort -- now() flagged `defaulted=True` so the caller
    can disclose that the collection date was unknown. Pure read; no side effects."""
    if not no_collect:
        return datetime.now().isoformat(), False
    base = os.path.basename(os.path.normpath(collection_dir or root_dir or ""))
    m = re.search(r"(\d{8})_(\d{6})", base)
    if m:
        try:
            return datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S").isoformat(), False
        except ValueError:
            pass
    d = collection_dir or root_dir or ""
    try:
        mtimes = [os.path.getmtime(os.path.join(dp, f))
                  for dp, _dirs, files in os.walk(d) for f in files]
        if mtimes:
            return datetime.fromtimestamp(min(mtimes)).isoformat(), False
    except OSError:
        pass
    return datetime.now().isoformat(), True

# =============================================================================
# V3.15.0 ADDITIONS - new analysis outputs (Tier 1) + new collection (Tier 2).
# Tier 1 sheets aggregate already-parsed data (InterfaceData / DevicePhysical).
# Tier 2 sheets parse raw command output directly (writer takes the per-device
# cmd_to_file map) so the existing interface-building pipeline is untouched.
# =============================================================================

# -----------------------------------------------------------------------------
# compute_topology_links moved to cisco_toolkit.analyze (PHASE 2.7 step 12); imported
# back near the top of this file (write_topology_sheet / write_topology_diagram use it).
# -----------------------------------------------------------------------------


# -----------------------------------------------------------------------------
# Tier 1 #1: Findings / Risk Register  (pure cross-reference of InterfaceData)
# -----------------------------------------------------------------------------
# (Findings / Capacity / Topology-diagram writers also moved to cisco_toolkit.excel in
# step 22 - see the note above. _to_float / _mermaid_id moved with them as package-internal.)


# -----------------------------------------------------------------------------
# Tier 1 #2: Pre/post-cutover snapshot + diff. A snapshot JSON is written next
# to every output workbook; '--compare OLD NEW' produces a diff workbook.
# -----------------------------------------------------------------------------
# snapshot_state + write_html_explorer moved to cisco_toolkit.html (PHASE 2.7 step 30);
# both imported back near the top of this file (main() builds the snapshot + emits the HTML).


# _DIFF_FIELDS / _macset / write_diff_workbook moved to cisco_toolkit.html (PHASE 2.7 step 29);
# write_diff_workbook imported back near the top of this file (main()'s --compare path calls it).
# _DIFF_FIELDS / _macset are package-internal. (snapshot_state + write_html_explorer follow in step 30.)


# -----------------------------------------------------------------------------
# Tier 2 #1: Interface health counters  ('show interfaces' / 'show interface')
# -----------------------------------------------------------------------------
# Tier-2 (file-reading) + intelligence sheet writers moved to cisco_toolkit.excel
# (PHASE 2.7 step 24): write_interface_health_sheet / write_security_posture_sheet /
# write_routing_adjacency_sheet / write_causality_chains_sheet / write_failure_impact_sheet.
# All five imported back near the top of this file (main() calls them). excel.py now imports
# _load_cmd_output + the parse parsers + compute_causality_chains/compute_failure_impact itself.


# (Causality Chains + Failure Impact writers also moved to cisco_toolkit.excel in step 24 -
# see the note above. The intelligence-layer design doc lives with compute_causality_chains /
# compute_failure_impact in cisco_toolkit/analyze.py + the .md.)


# -----------------------------------------------------------------------------
# NEW-V3.18: Physical Health (L1). One row per physical port - negotiated
# speed/duplex/media, error counters, port-channel, PoE, and a derived physical
# risk flag. Pure derivation: reuses parse_show_interface_counters (errors),
# build_network_model (uplink topology -> single-fiber), and DevicePhysical
# (PoE budget). Negotiated duplex/media/speed come from the 'show interfaces'
# DETAIL block because InterfaceData.duplex is normalized and loses half-duplex.
# No new collection, no new InterfaceData fields, no new imports.
# -----------------------------------------------------------------------------
# Fused compute-in-writer sheets (write_physical_health_sheet / write_l3_forwarding_sheet /
# write_flow_trace_sheet) moved to cisco_toolkit.excel (PHASE 2.7 step 25). The physical-health
# + l3-forwarding writers compute their L1/L3 records inline and RETURN them; main() captures
# those via _run_phase. All three imported back near the top of this file.


# -----------------------------------------------------------------------------
# NEW-V3.19: Full-stack flow trace (L1->L3). Given two endpoint IPs, derive the
# path: locate each endpoint (end_host_ip), resolve gateway SVIs per subnet
# (svi_ip + subnet_primary_route containment + FHRP peers), decide L2 vs L3, walk
# the fabric over STP-forwarding links (_link_carries), flag single points of
# failure (single-fiber uplinks, single gateway, VLAN partition) and score risk.
# Pure derivation - routing comes from InterfaceData (no routing_data dict exists).
# L4 is NOT simulated (no ACL/service-policy is collected). No new imports.
# -----------------------------------------------------------------------------
# (Flow Trace + L3 Forwarding writers also moved to cisco_toolkit.excel in step 25 - see the
# note above. _parse_track / _track_summary moved with write_l3_forwarding_sheet (package-internal).)


# -----------------------------------------------------------------------------
# NEW-V3.21: Cross-layer correlation. Combines the L1 (physical_health) and L3
# (l3_forwarding) records already computed this run with topology facts from the
# model (articulation, single-fiber uplinks, single-member port-channels, orphan
# VLANs) to surface compounded-risk findings CL-01..CL-10 that no single-layer
# view shows. Pure derivation; no new collection.
# -----------------------------------------------------------------------------
# Health / analysis sheet writers (write_cross_layer_sheet / write_protocol_health_sheet /
# write_health_scores_sheet / write_score_sensitivity_sheet / write_migration_readiness_sheet)
# moved to cisco_toolkit.excel (PHASE 2.7 step 23); imported back near the top of this file
# (main() runs them via _run_phase). Their fill maps + sheet-name constants + _SEV_FILL moved too.


# =============================================================================
# MAIN
# =============================================================================
def _empty_dep_map() -> dict:
    """Empty dependency-map skeleton with every key the cross-layer and
    migration-readiness consumers read. Used as the resilient fallback if
    build_dependency_map() raises, so one analysis failure can't sink the run."""
    return {"single_fiber": set(), "uplink_ports": set(), "sole_gw": {},
            "access_by_vlan": {}, "articulation": set(), "fhrp_vlans": set(),
            "tracked_down": set(), "errored_up": set(), "halfdup_up": set(),
            "single_member_pc": set(), "errdis": set(), "gw_switches": set(),
            "orphan": set(), "model": {"hosts": set()}}


# Measure-first perf harness (Plan A / Tier-2 #11): every _run_phase call is timed here,
# chronologically, and main() writes the ledger as a .phase_timings.json SIDECAR next to
# the workbook — golden-neutral by construction (never a snapshot key). This is the hard
# predecessor of any perf work: no restructure without before/after numbers from THIS ledger.
_PHASE_TIMINGS: List[dict] = []


def _run_phase(label, fn, *args, _default=None, **kwargs):
    """FIX (resilience): run a pre-save phase so a failure LOGS AND CONTINUES
    instead of aborting the whole run before wb.save() (extends the FIX-V14-1
    'guard and continue' idea from Phase 6 to every phase). Returns the phase's
    result, or _default if it raised. Happy-path behaviour is unchanged.
    Tier-2 #11: each call is timed into _PHASE_TIMINGS (ok=False on a failure,
    so a fast-because-it-crashed phase can never read as a fast phase)."""
    _t0 = time.perf_counter()
    _ok = True
    try:
        return fn(*args, **kwargs)
    except Exception as e:
        _ok = False
        logger.error(f"  [SKIP] Phase '{label}' failed: {e!r}; "
                     f"continuing so the workbook still saves.")
        return _default
    finally:
        _PHASE_TIMINGS.append({"phase": str(label),
                               "seconds": round(time.perf_counter() - _t0, 4), "ok": _ok})


# =============================================================================
# PER-DEVICE AXIS REGISTRY (Plan A / Tier-2 #8)
# =============================================================================
# ONE entry wires a per-device evidence axis through Phase 5.6. The accumulator
# declaration + build-loop stanza (+ its log line) used to be hand-parallel code
# sites per axis — the drift generator that kept finished axes dark for months
# (see cisco_toolkit/summary.py's NOS-quartet note). Follows the _DETECTORS
# precedent (design_advisor). The snap-assign site stays explicit at snapshot
# assembly; tests/test_axis_registry.py enforces that every registry key ships
# in the golden snapshot, so forgetting that site fails the suite.
#   key         store name AND (except the three transient feeds) the snap key
#   builder     per-device fn(cmd_to_file) -> value
#   tag/log     the Phase-5.6 log line, preserved verbatim (log=None: silent)
#   keep        None = keep truthy; callable = custom predicate (pim / PROTO)
#   keep_always True = record EVERY host ('' / all-empty declares not-collected)
class _AxisSpec(NamedTuple):
    key: str
    builder: Callable
    tag: str
    log: Optional[Callable]
    keep: Optional[Callable]
    keep_always: bool


_PER_DEVICE_AXES: List[_AxisSpec] = [
    _AxisSpec("run_configs", read_run_config, "RC", None, None, False),                  # NEW-V3.23.146 (golden-config drift feed)
    _AxisSpec("syslogs", read_syslog_log, "SYSLOG", None, None, True),                   # NEW-V3.23.164 (every host recorded so not-collected is declared)
    _AxisSpec("platform_metrics", build_platform_metrics, "PLAT", None, None, True),     # NEW-V3.23.167 (all-empty members = not collected)
    _AxisSpec("acls", build_acls, "ACL",
              lambda v: f"{len(v)} access-list(s) parsed", None, False),
    _AxisSpec("object_groups", build_object_groups, "ACL",
              lambda v: f"{len(v)} object-group(s) parsed", None, False),
    _AxisSpec("nat", build_nat, "NAT",
              lambda v: f"{len(v.get('static', []))} static, "
                        f"{len(v.get('dynamic', []))} dynamic rule(s)", None, False),
    _AxisSpec("security", build_security, "SEC",
              lambda v: f"{v.get('summary', {}).get('fail', 0)} fail / "
                        f"{len(v.get('findings', []))} check(s) ({v.get('summary', {}).get('grade', '')})",
              None, False),
    _AxisSpec("config_hygiene", build_config_hygiene, "HYGIENE",
              lambda v: f"{v.get('summary', {}).get('undefined', 0)} undefined ref(s), "
                        f"{v.get('summary', {}).get('unused', 0)} unused structure(s)", None, False),
    _AxisSpec("stp_roots", build_stp_roots, "STP",
              lambda v: f"root bridge info for {len(v)} VLAN(s)", None, False),
    _AxisSpec("vpc", build_vpc, "vPC",
              lambda v: f"domain {v.get('domain_id')} role {v.get('role') or '?'}, "
                        f"{len(v.get('vpcs', []))} vPC(s)", None, False),
    _AxisSpec("fhrp_detail", build_fhrp_detail, "FHRP",
              lambda v: f"{len(v)} HSRP group(s) (detail)", None, False),
    _AxisSpec("overlay", build_overlay, "VXLAN",
              lambda v: f"{len(v.get('nve_peers', []))} NVE peer(s)", None, False),
    _AxisSpec("copp", build_copp, "CoPP",
              lambda v: f"{len(v)} control-plane class(es)", None, False),
    _AxisSpec("mpls", build_mpls, "MPLS",
              lambda v: f"{len(v.get('ldp_neighbors', []))} LDP session(s), "
                        f"{len(v.get('vpnv4_neighbors', []))} VPNv4 peer(s), "
                        f"{len(v.get('l2vpn_vcs', []))} L2VPN VC(s)", None, False),
    # universal architecture coverage + multi-vendor + controller-REST axes (silent by
    # design: their evidence is optional exports, absent on most fleets)
    _AxisSpec("lisp", build_lisp, "LISP", None, None, False),
    _AxisSpec("cts", build_cts, "CTS", None, None, False),
    _AxisSpec("dmvpn", build_dmvpn, "DMVPN", None, None, False),
    _AxisSpec("crypto", build_crypto, "CRYPTO", None, None, False),
    _AxisSpec("bfd", build_bfd, "BFD", None, None, False),
    _AxisSpec("aci", build_aci, "ACI", None, None, False),
    _AxisSpec("sdwan", build_sdwan, "SDWAN", None, None, False),
    _AxisSpec("firewall", build_firewall, "FW", None, None, False),
    _AxisSpec("ise", build_ise, "ISE", None, None, False),
    _AxisSpec("fmc", build_fmc, "FMC", None, None, False),
    _AxisSpec("arista", build_arista, "ARISTA", None, None, False),                      # MULTI-VENDOR ({} on a Cisco device)
    _AxisSpec("juniper", build_juniper, "JUNIPER", None, None, False),                   # MULTI-VENDOR ({} on a non-Juniper device)
    _AxisSpec("cloud", build_cloud, "CLOUD", None, None, False),                         # PUBLIC CLOUD ({} when no cloud export)
    _AxisSpec("fortigate", build_fortigate, "FGT", None, None, False),                   # MULTI-VENDOR ({} on a non-FortiGate device)
    _AxisSpec("mroute", build_mroute, "MROUTE", None, None, False),                      # CISCO DEPTH ({} when no mroute table)
    _AxisSpec("ipv6_nd", build_ipv6_nd, "V6ND", None, None, False),
    _AxisSpec("ipv6_routing", build_ipv6_routing, "V6RT", None, None, False),
    _AxisSpec("pim", build_pim, "PIM",
              lambda v: f"{len(v.get('neighbors', []))} neighbor(s), "
                        f"{(v.get('rp_mapping') or {}).get('rp_count', 0)} RP(s)",
              lambda v: bool(v.get("rp_mapping") or v.get("neighbors")), False),
    _AxisSpec("ipv6_fhs", build_ipv6_fhs, "V6FHS",
              lambda v: f"dualstack={v.get('dualstack')} ra_guard={v.get('ra_guard_present')}",
              None, False),
    _AxisSpec("ntp", build_ntp, "NTP",
              lambda v: f"synchronized={v.get('synchronized')} stratum={v.get('stratum')}",
              None, False),
    _AxisSpec("port_security", build_port_security_detail, "PSEC",
              lambda v: f"{len(v)} secured port(s)", None, False),
    _AxisSpec("storm_control", build_storm_control, "STORM",
              lambda v: f"{len(v)} storm-control rule(s)", None, False),
    _AxisSpec("qos_runtime", build_qos_runtime, "QOSRT",
              lambda v: f"{len(v)} egress class(es)", None, False),
    _AxisSpec("shadow_infra", build_undocumented_neighbors, "SHADOW",
              lambda v: f"{len(v)} infra neighbour(s) seen", None, False),
    _AxisSpec("routing_neighbors", build_routing_neighbors, "PROTO",
              lambda v: f"{len(v['ospf'])} OSPF, {len(v['eigrp'])} EIGRP, "
                        f"{len(v['bgp'])} BGP adjacency(ies)",
              lambda v: any(v.values()), False),
    _AxisSpec("redistribution", build_redistribution, "PROTO",
              lambda v: f"{len(v)} redistribution edge(s)", None, False),
]


def _run_per_device_axes(stores: Dict[str, dict], hostname: str, cmd_to_file: Dict[str, str],
                         specs: Optional[List[_AxisSpec]] = None) -> None:
    """Phase-5.6 driver: run every registered builder for ONE device in registry order,
    honoring each axis's keep semantics; the per-axis log line fires only when kept."""
    for sp in (_PER_DEVICE_AXES if specs is None else specs):
        v = sp.builder(cmd_to_file)
        kept = True if sp.keep_always else (bool(sp.keep(v)) if sp.keep else bool(v))
        if not kept:
            continue
        stores[sp.key][hostname] = v
        if sp.log is not None:
            logger.info(f"  [{sp.tag}] {hostname}: {sp.log(v)}")


def main():
    ap = argparse.ArgumentParser(description=f"Cisco Migration Extractor V{__version__}")
    ap.add_argument("--devices-file",   default=None,
                    help="Devices JSON (required unless --compare is used)")
    ap.add_argument("--template",        default=DEFAULT_TEMPLATE_FILE)
    ap.add_argument("--output",          default="")
    ap.add_argument("--no-collect",      action="store_true")
    ap.add_argument("--collection-dir",  default="")
    ap.add_argument("--requirements",    default="",
                    help="Path to a design requirements register (JSON: availability_tier, critical_apps, "
                         "convergence_budget_ms, growth_horizon, fabric_operating_model, constraints, "
                         "data_classification, address_space, vlan_zones). Supplied, it right-sizes the "
                         "published design_blueprint (the WHY) so every deliverable reflects the requirements; "
                         "absent, the blueprint surfaces the open requirement questions.")
    ap.add_argument("--debug-headers",   action="store_true")
    ap.add_argument("--workers",         type=int, default=5,
                    help="Parallel SSH workers (default 5; use 1 for sequential)")
    ap.add_argument("--debug-arp",       action="store_true",
                    help="Enable DEBUG logging for per-device ARP counts")
    ap.add_argument("--compare",         nargs=2, default=None,
                    metavar=("OLD_SNAPSHOT", "NEW_SNAPSHOT"),
                    help="NEW-V15: compare two snapshot JSONs and write a pre/post-cutover validation "
                         "workbook (Summary verdict + Interface/Endpoint/SVI changes + Health Shifts + "
                         "Findings Delta + Pre-Change Certificate) plus a <output>.precert.json "
                         "Pre-Change Validation Certificate (roadmap C1; --path-intents is honored); "
                         "skips collection and the template.")
    ap.add_argument("--no-html",         action="store_true",
                    help="NEW-V3.17: skip Blast-Radius Explorer (HTML) generation "
                         "(default: HTML is always written beside the workbook).")
    ap.add_argument("--no-docx",         action="store_true",
                    help="NEW-V3.23.93: skip the Assessment & Migration Runbook (DOCX). The runbook is "
                         "the narrative twin of the workbook; it needs python-docx (a missing library is "
                         "a warning, not an error).")
    ap.add_argument("--no-pptx",         action="store_true",
                    help="NEW-V3.23.144: skip the Executive presentation deck (PPTX). The deck is the "
                         "stakeholder twin of the workbook; it needs python-pptx (a missing library is "
                         "a warning, not an error).")
    ap.add_argument("--no-design",       action="store_true",
                    help="NEW-V3.23.148: skip the As-Built Network Design Document (HLD/LLD, DOCX). It "
                         "reconstructs the current design + target-state recommendations from the "
                         "snapshot; needs python-docx (a missing library is a warning, not an error).")
    ap.add_argument("--no-mop",          action="store_true",
                    help="NEW-V3.23.149: skip the per-wave Method of Procedure (MOP, DOCX) — a "
                         "maintenance-window cutover template per migration wave; needs python-docx "
                         "(a missing library is a warning, not an error).")
    ap.add_argument("--override-gate",   default=None, metavar="REASON",
                    help="P0-3/DEC-003: proceed past a REFUSED PPDIOO document gate. When "
                         "docs/engagement-state.json exists (see python -m cisco_toolkit.gate_state), "
                         "the design doc requires an approved assessment and the MOP an approved LLD "
                         "+ captured baseline; a missing approval refuses that deliverable. This flag "
                         "overrides the refusal, recording WHO/WHEN/WHY as an audit line in the "
                         "gate-state store (reason must be non-empty). Without the store, generation "
                         "proceeds ungated (brownfield).")
    ap.add_argument("--gate-root",       default=".", metavar="DIR",
                    help="Engagement root holding docs/engagement-state.json for the PPDIOO "
                         "document gates above (default: the working directory). Set this when the "
                         "engine is launched with a SYNTHETIC cwd — a wrapper that re-homes the "
                         "child to a scratch dir would otherwise resolve the store to nothing and "
                         "silently downgrade every gate to brownfield warn-and-proceed. See "
                         "cisco_toolkit/gate_state.py 'Root resolution'.")
    ap.add_argument("--engagement",      default=None, metavar="ID",
                    help="ADR-0006: the engagement this run is FOR. When given, the gate ledger "
                         "found under --gate-root must declare the SAME engagement or the gated "
                         "deliverables are refused (non-overridably) — ownership is VERIFIED "
                         "rather than inferred from where the process happens to be running. "
                         "Bind a ledger once with 'python -m cisco_toolkit.gate_state bind <ID>'. "
                         "Omit it and behaviour is exactly as before (proximity, unverified).")
    ap.add_argument("--fail-on-gate-refusal", action="store_true",
                    help="Exit 2 when a PPDIOO document gate REFUSED a deliverable (design/MOP). "
                         "OPT-IN: by default a refusal is a successful run that correctly withheld "
                         "a document, and the exit code stays 0 — the wrappers that drive this "
                         "engine treat any non-zero as a hard failure (AssessHub's redaction path "
                         "tells the field engineer to treat everything already written as "
                         "UNREDACTED), so a refusal must not masquerade as one. Use this in CI or a "
                         "pipeline that must STOP when governance withholds a document. NB it fires "
                         "on ANY refusal, including a mis-set --gate-root or an unreadable ledger, "
                         "not only on a missing approval. A refusal over a missing approval is "
                         "additionally recorded in the gate ledger's audit array (python -m "
                         "cisco_toolkit.gate_state show); the other refusal classes have no ledger "
                         "to write to, and the run says which is which as it exits.")
    ap.add_argument("--no-crd",          action="store_true",
                    help="NEW-V3.23.156: skip the Customer Requirements Document (CRD, DOCX) — the "
                         "Plan-phase requirements-capture instrument primed with the assessment "
                         "evidence; needs python-docx (a missing library is a warning, not an error).")
    ap.add_argument("--no-engagement",   action="store_true",
                    help="NEW-V3.23.157: skip the Engagement Workflow & Plan of Record (DOCX) — the "
                         "phase-gated engagement plan (verdict, gate calendar, RAID log, next actions) "
                         "synthesized from the assessment; needs python-docx (a missing library is a "
                         "warning, not an error).")
    ap.add_argument("--no-opshandbook",  action="store_true",
                    help="NEW-V3.23.168: skip the Operations Handbook (DOCX) — the Operate-phase "
                         "Day-2 handbook (monitoring baseline from the fleet's own logs/capacity, "
                         "drift control, software governance, escalation pack); needs python-docx "
                         "(a missing library is a warning, not an error).")
    ap.add_argument("--no-archreview",   action="store_true",
                    help="NEW-V3.23.160: skip the Architecture Review & Conformance Report (DOCX) — "
                         "the leading-practice design review (8-domain conformance scorecard + "
                         "availability analysis); needs python-docx (a missing library is a warning, "
                         "not an error).")
    ap.add_argument("--golden-config",   default=None, metavar="FILE",
                    help="NEW-V3.23.146: a golden-config baseline file (one required directive per line; "
                         "'#' comments and 're:<regex>' supported) for the Golden-Config Drift sheet. "
                         "Omit to auto-derive the baseline from the fleet's majority (de-facto standard).")
    ap.add_argument("--import-inventory", default=None, metavar="FILE",
                    help="roadmap B: a declared inventory (CMDB/NetBox/Nautobot CSV or XLSX export) to "
                         "reconcile against the collected evidence -> a 'SoT Reconcile' sheet + snapshot key. "
                         "Opt-in (read-only file ingest; never a live API). Off by default.")
    ap.add_argument("--scenario",        default=None, metavar="FILE",
                    help="roadmap G4: a JSON failure-injection catalog (a list of {name, failures:[{type:node|site, id}]}) "
                         "-> a 'Failure What-If' sheet + snapshot key. Opt-in; read-only (the snapshot is mutated in memory).")
    ap.add_argument("--path-intents",    default=None, metavar="FILE",
                    help="roadmap G3: a JSON catalog of named path/segmentation intents (a list of {id, src, dst, expect: "
                         "REACHES|ISOLATED}) evaluated over the computed FIB -> a 'Path Assertions' sheet + snapshot key. "
                         "With --compare, the catalog is re-validated across the pair into the Pre-Change Certificate. Opt-in.")
    ap.add_argument("--assert-pack",     default=None, metavar="FILE",
                    help="roadmap A1/H2: a JSON state-assertion check-pack ({assertions:[...]}) evaluated over the "
                         "snapshot -> a 'state_assertions' snapshot key. Opt-in; coverage-honest ([NOT OBSERVED] abstention).")
    ap.add_argument("--flow-src",        default=None, metavar="IP",
                    help="NEW-V3.19: source endpoint IP for the optional Flow Trace sheet "
                         "(requires --flow-dst).")
    ap.add_argument("--flow-dst",        default=None, metavar="IP",
                    help="NEW-V3.19: destination endpoint IP for the optional Flow Trace sheet "
                         "(requires --flow-src).")
    ap.add_argument("--trend",           nargs="+", default=None, metavar="SNAPSHOT",
                    help="NEW-V3.23.145: migration-campaign trend across a SERIES of snapshot JSONs (>=2, "
                         "oldest first) -> a trend workbook (Campaign Summary verdict + Timeline w/ chart + "
                         "Burndown). Skips collection and the template.")
    ap.add_argument("--allow-schema-mismatch", action="store_true",
                    help="P3-E2: permit --compare/--trend across snapshots produced by DIFFERENT engine "
                         "schema versions (script_version). OFF by default, so a cross-schema diff -- which "
                         "can misreport a schema change as a real network change -- is REFUSED rather than "
                         "emitted silently; passing this downgrades the refusal to a logged warning "
                         "(both versions are recorded in the certificate provenance either way).")
    ap.add_argument("--redact",          action="store_true",
                    help="NEW-V3.23.41: pseudonymize IPs / MACs / serial numbers across the WHOLE output "
                         "bundle -- the snapshot JSON, the HTML explorer, AND the always-produced .xlsx "
                         "workbook (consistent, subnet-preserving; hostnames kept) -- so every deliverable "
                         "can be shared without leaking real addressing. Default off.")
    ap.add_argument("--redact-collection", action="store_true",
                    help="Plan A Tier-1 #5: scrub SECRET VALUES (passwords / SNMP communities / keys) "
                         "IN PLACE across the raw collection dir's .txt captures, AFTER analysis "
                         "completes (this run reads the originals). Values only -- IPs / hostnames are "
                         "kept so the dir stays analyzable and remains the --compare/--trend source; "
                         "nothing is deleted. Idempotent. Rewritten captures will no longer match "
                         "archive hashes recorded at collection time (deliberate). Default off.")
    args = ap.parse_args()

    # This run's gate ledger starts empty. Matters for hosts that call main() twice in one process --
    # in-repo that is tests/test_pipeline_inprocess.py and tests/test_pipeline_failopen.py (the webapp
    # uses a subprocess). Without it, run 2's manifest would seal run 1's verdicts: a false audit
    # record, which is a worse failure than the missing one this change fixes. Placed before the
    # --compare/--trend early returns so every path that can reach build_run_manifest starts clean.
    gate_reset_verdicts()

    if args.debug_arp:
        logger.setLevel(logging.DEBUG)
        for h in logger.handlers: h.setLevel(logging.DEBUG)
        # The per-device ARP debug lines this flag advertises moved to cisco_toolkit.build (dcf53a8,
        # PHASE 2.7 step 27), so lowering only the engine logger left the flag a documented no-op --
        # the package tree filters at ITS logger before the (shared, already-lowered) handlers see
        # anything. Lower it too, or --debug-arp still cannot reach the lines it is named after.
        logging.getLogger("cisco_toolkit").setLevel(logging.DEBUG)

    # NEW-V15: diff mode - compare two snapshots, no SSH/template needed.
    if args.compare:
        old_p, new_p = args.compare
        try:                                              # NEW-V3.23.1: clean message, not a raw traceback
            with open(old_p, encoding="utf-8") as f: old_snap = json.load(f)
            with open(new_p, encoding="utf-8") as f: new_snap = json.load(f)
        except FileNotFoundError as e:
            ap.error(f"--compare: snapshot file not found: {e.filename}")
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            ap.error(f"--compare: could not parse snapshot JSON ({e})")
        # P3-E2: refuse to diff across a schema/script_version change unless explicitly allowed -- a
        # cross-schema diff can misreport a schema difference as a real change (never emit it silently).
        _sc_status, _sc_msg = schema_compat_status([old_snap, new_snap], labels=[old_p, new_p])
        if _sc_status == "mismatch" and not args.allow_schema_mismatch:
            ap.error(f"--compare: {_sc_msg}. Re-run with --allow-schema-mismatch to diff across schema "
                     f"versions anyway (both versions are recorded in the certificate provenance).")
        if _sc_status in ("mismatch", "unverifiable"):
            logger.warning(f"  [schema] {_sc_msg}"
                           + (" -- proceeding under --allow-schema-mismatch" if _sc_status == "mismatch" else ""))
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        diff_out = args.output or f"Migration_Diff_{stamp}.xlsx"
        os.makedirs(os.path.dirname(os.path.abspath(diff_out)) or ".", exist_ok=True)  # FIX-V3.23.103: same missing-output-dir guard
        # roadmap C1: the Pre-Change Validation Certificate — the fib differential packaged as the PPDIOO
        # gate artifact (<output>.precert.json + a 'Pre-Change Certificate' sheet). An optional
        # --path-intents catalog is re-validated across the pair and folded into the certificate.
        _intents = None
        if args.path_intents:
            try:
                with open(args.path_intents, encoding="utf-8") as _pf:
                    _intents = json.load(_pf)
            except Exception as e:
                _intents = None
                logger.warning(f"  --path-intents not read ({e}); certificate proceeds without intents.")
            else:
                _intents = _intents if isinstance(_intents, list) else (_intents.get("assertions") or _intents.get("intents") or [])
        cert = compute_precert(old_snap, new_snap, path_intents=_intents)
        write_diff_workbook(old_snap, new_snap, diff_out, precert=cert)
        precert_path = os.path.splitext(os.path.abspath(diff_out))[0] + ".precert.json"
        write_json_file(precert_path, cert)
        logger.info(f"[OK] Saved diff workbook: {os.path.abspath(diff_out)}")
        logger.info(f"[OK] Pre-Change Validation Certificate: {precert_path}  (verdict: {cert.get('verdict')})")
        return

    # NEW-V3.23.145: trend mode - migration-campaign trajectory across a SERIES of snapshots, no SSH/template.
    if args.trend:
        if len(args.trend) < 2:
            ap.error("--trend: need at least two snapshot JSON files (oldest first).")
        snaps = []
        try:                                              # clean message, not a raw traceback
            for p in args.trend:
                with open(p, encoding="utf-8") as f:
                    snaps.append(json.load(f))
        except FileNotFoundError as e:
            ap.error(f"--trend: snapshot file not found: {e.filename}")
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            ap.error(f"--trend: could not parse snapshot JSON ({e})")
        # P3-E2: same schema/script_version gate for the campaign series (an engine upgrade mid-campaign
        # would otherwise trend across incompatible schemas silently).
        _sc_status, _sc_msg = schema_compat_status(snaps, labels=list(args.trend))
        if _sc_status == "mismatch" and not args.allow_schema_mismatch:
            ap.error(f"--trend: {_sc_msg}. Re-run with --allow-schema-mismatch to trend across schema versions.")
        if _sc_status in ("mismatch", "unverifiable"):
            logger.warning(f"  [schema] {_sc_msg}"
                           + (" -- proceeding under --allow-schema-mismatch" if _sc_status == "mismatch" else ""))
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        trend_out = args.output or f"Migration_Trend_{stamp}.xlsx"
        os.makedirs(os.path.dirname(os.path.abspath(trend_out)) or ".", exist_ok=True)
        write_campaign_workbook(snaps, trend_out)
        logger.info(f"[OK] Saved campaign trend workbook: {os.path.abspath(trend_out)}")
        return

    if not args.devices_file:
        ap.error("--devices-file is required (unless using --compare OLD NEW)")

    if not os.path.exists(args.template):
        raise FileNotFoundError(f"Template not found: {args.template}")

    devices = load_devices(args.devices_file, allow_prompt=not args.no_collect)  # FIX-V3.23.177: --no-collect must never block on a TTY password prompt
    stamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = args.output or DEFAULT_OUTPUT_FILE.format(stamp)

    # FIX-V3.23.103: create the --output directory up-front, the moment the path is
    # resolved -- BEFORE the ~31 phases of multi-minute compute. The workbook, snapshot,
    # topology diagram, HTML explorer and runbook are all emitted only at the very end of
    # the run, and they all derive their directory from this same out_xlsx (see the
    # post-save block). A missing output directory would otherwise crash openpyxl's
    # wb.save() with FileNotFoundError AFTER all the heavy compute, wasting the whole run.
    # One guard here covers every output. No behaviour change when the directory exists.
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)) or ".", exist_ok=True)

    wb = load_workbook(args.template)
    from cisco_toolkit.excel import harden_workbook
    harden_workbook(wb)   # every sheet sanitizes control chars -> one device-text field can't abort the workbook
    ws = wb.active
    logger.info(f"[OK] Loaded template: {args.template}")

    if args.debug_headers:
        debug_scan_headers(ws, max_rows=150, max_cols=90)
        print("Exiting because --debug-headers was used.")
        return

    header_row, col_map = find_header_row(ws)
    logger.info(f"[OK] Header row at row {header_row}")
    col_map = ensure_headers(ws, header_row, col_map, {
        "TRUNK Native VLAN":    "trunk_native_vlan",
        "TRUNK Allowed VLANs":  "trunk_allowed_vlans",
        "TRUNK Status":         "trunk_status",
        "ARP Source Switch":    "arp_source_switch",  # NEW-V11
        "CDP/LLDP Neighbor":    "cdp_neighbor",       # NEW-V11
        # NEW-V14.2 (header text normalizes to the HEADER_TO_FIELD keys above)
        "Current Switch Serial":      "current_switch_serial",
        "Current Switch Mgmt IP":     "current_switch_ip",
        "Current Switch VTP Domain":  "current_switch_vtp_domain",
        "Neighbor Switch Serial":     "neighbor_switch_serial",
        "Neighbor Switch Mgmt IP":    "neighbor_switch_ip",
        "Neighbor Switch VTP Domain": "neighbor_switch_vtp_domain",
        "Subnet Primary Route":       "subnet_primary_route",
        "Subnet Secondary Routes":    "subnet_secondary_routes",
        "Route Next Hop":             "route_next_hop",
        "Routing Source":             "routing_source",
        "FHRP Behavior":              "hsrp_behavior",
        "Multicast Info":             "multicast_info",
    })

    root_dir = args.collection_dir or COLLECTION_DIR.format(stamp)
    os.makedirs(root_dir, exist_ok=True)
    logger.info(f"[OK] Collection directory: {root_dir}")

    # Provenance (wave R2-1-01 / R2-3-01): pin the lifecycle EoL bands + the deliverable "Snapshot
    # captured" date to WHEN THE EVIDENCE WAS COLLECTED, not the wall-clock day this pipeline runs. On a
    # --no-collect re-analysis collected_at is recovered from the collection-dir stamp, so a "frozen"
    # assessment reproduces its bands + cover date byte-for-byte instead of drifting against today.
    collected_at, _collected_at_defaulted = _derive_collected_at(args.no_collect, args.collection_dir, root_dir)
    if _collected_at_defaulted:
        logger.warning("  [PROVENANCE] collection date could not be determined from %r; "
                       "'Snapshot captured' falls back to generation time.", root_dir)
    else:
        logger.info(f"  [PROVENANCE] evidence collected_at = {collected_at}")

    _progress_lock = threading.Lock()
    _done_count    = [0]

    def collect_one(devinfo):
        ip       = devinfo["ip"]
        hostname = devinfo["hostname"]
        username = devinfo["username"]
        password = devinfo["password"]
        platform = devinfo["platform"]
        safe_host = safe_fs_name(hostname)
        dev_dir   = os.path.join(root_dir, safe_host)
        cmd_to_file: Dict[str,str] = {}

        if args.no_collect:
            if platform in ("auto",""):
                platform = detect_platform_from_files(dev_dir)
                logger.info(f"  [{hostname}] Auto-detected platform: {platform}")
            all_cmds = list(dict.fromkeys(COMMANDS_NXOS + COMMANDS_IOS + COMMANDS_ARISTA + COMMANDS_JUNIPER + COMMANDS_CLOUD + COMMANDS_FORTINET))
            for cmd in all_cmds:
                fn    = cmd.replace(" ","_").replace("|","_").replace("^","").replace("/","_")+".txt"
                fpath = os.path.join(dev_dir, fn)
                if os.path.isfile(fpath): cmd_to_file[cmd] = fpath
        else:
            logger.info(f"  Connecting to {hostname} ({ip}) ...")
            dev, platform = connect_device(ip, hostname, username, password, platform)
            if not dev:
                logger.error(f"  [FAIL] Skipped {hostname}")
                with _progress_lock:
                    _done_count[0] += 1
                    logger.info(f"  Progress: {_done_count[0]}/{len(devices)} devices done")
                return hostname, platform, None
            try:
                cmd_to_file = collect(hostname, platform, dev, dev_dir)
            finally:
                try: dev.disconnect()
                except Exception: pass

        # V14.12: correct the platform from the actually-collected output. Union collection
        # means both command forms are present; whichever returned real (non-error) output
        # reveals the true OS. Fixes a wrong 'platform' in devices.json (e.g. nxos on an IOS
        # Catalyst) so parsing and the Switch Inventory label are right.
        if cmd_to_file:
            detected = detect_platform_from_files(dev_dir)
            if detected != platform:
                logger.info(f"  [{hostname}] Platform corrected '{platform}' -> '{detected}' from collected output")
                platform = detected

        with _progress_lock:
            _done_count[0] += 1
            logger.info(f"  [{_done_count[0]}/{len(devices)}] Finished collecting {hostname}")
        return hostname, platform, cmd_to_file

    # Phase 1: Collect
    # Fresh zero-parse ledger per run (webapp ingest reuses this pipeline in-process, so
    # the ledger must never carry a previous run's events into this snapshot).
    reset_parse_ledger()
    del _PHASE_TIMINGS[:]        # fresh perf ledger per run, same reasoning (Tier-2 #11)
    workers = max(1, min(args.workers, len(devices)))
    logger.info(f"\n[Phase 1] Collecting {len(devices)} device(s) with {workers} parallel worker(s) ...")
    all_cmd_to_files: Dict[str, Dict[str,str]] = {}
    all_devices_meta = []

    if workers == 1:
        results = [collect_one(d) for d in devices]
    else:
        results = []
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="collector") as pool:
            futures = {pool.submit(collect_one, d): d["hostname"] for d in devices}
            for fut in as_completed(futures):
                try:
                    results.append(fut.result())
                except Exception as e:
                    logger.error(f"  [FAIL] Worker exception for {futures[fut]}: {e}")

    for hostname, platform, cmd_to_file in results:
        if cmd_to_file is None: continue
        all_cmd_to_files[hostname] = cmd_to_file
        all_devices_meta.append((hostname, platform, cmd_to_file))

    logger.info(f"  Collection complete: {len(all_devices_meta)}/{len(devices)} succeeded.")
    # Plan A / Tier-1 #5 — say it on EVERY run, loudly: the raw captures are the one output
    # --redact never touches, and they hold running-configs with cleartext secrets.
    logger.warning(f"  [SENSITIVE] Raw collection dir holds CLEARTEXT device output "
                   f"(running-configs can embed passwords / SNMP communities / keys): {root_dir} "
                   f"-- --redact protects the DELIVERABLES only; once analysis is final, scrub the "
                   f"raw captures in place with --redact-collection (kept, never auto-deleted -- "
                   f"it is the --compare/--trend source).")

    # Phase 2: Global ARP
    logger.info("\n[Phase 2] Building global ARP table ...")
    global_arp, global_arp_source = collect_global_arp(all_cmd_to_files)  # NEW-V11 tuple

    # Phase 3: Parallel Parse
    logger.info(f"\n[Phase 3] Parsing {len(all_devices_meta)} device(s) ...")
    all_interfaces: Dict[str, Dict[str, InterfaceData]] = {}

    def parse_one(args_tuple):
        hostname, platform, cmd_to_file = args_tuple
        switch_identity = build_switch_identity(hostname, platform, cmd_to_file)
        ifaces = build_interfaces(hostname, platform, cmd_to_file, switch_identity=switch_identity)
        logger.info(f"  Parsed {hostname}: {len(ifaces)} interfaces")
        return hostname, ifaces

    if workers == 1:
        parse_results = [parse_one(t) for t in all_devices_meta]
    else:
        parse_results = []
        with ThreadPoolExecutor(max_workers=workers, thread_name_prefix="parser") as pool:
            futures = {pool.submit(parse_one, t): t[0] for t in all_devices_meta}
            for fut in as_completed(futures):
                try:
                    parse_results.append(fut.result())
                except Exception as e:
                    logger.error(f"  [FAIL] Parse exception for {futures[fut]}: {e}")

    for hostname, ifaces in parse_results:
        all_interfaces[hostname] = ifaces

    # Phase 4: Cross-device dual-connection detection
    logger.info("\n[Phase 4] Cross-device dual-connection detection ...")
    detect_cross_device_dual_connections(all_interfaces)

    # Phase 5: Apply global ARP -> IP + arp_source_switch
    logger.info("\n[Phase 5] Applying global ARP -> IP addresses ...")
    apply_global_arp(all_interfaces, global_arp, global_arp_source)  # NEW-V11

    # Phase 5.5: Build device physical inventory (NEW-V12)
    logger.info("\n[Phase 5.5] Building device physical inventory ...")
    all_device_physical: List[DevicePhysical] = []
    for hostname, platform, cmd_to_file in all_devices_meta:
        dp = build_device_physical(hostname, platform, cmd_to_file,
                                   all_interfaces.get(hostname, {}))
        all_device_physical.append(dp)
        logger.info(f"  [OK] Physical: {hostname}  model={dp.model}  "
                    f"PSU={dp.num_power_supplies}  "
                    f"cap={dp.power_capacity_w}  draw={dp.power_drawn_w}")

    # Phase 5.6: per-device evidence axes — REGISTRY-DRIVEN (Plan A / Tier-2 #8). Each
    # axis is ONE _PER_DEVICE_AXES entry above main() (the old accumulator decl + loop
    # stanza + log line were 3 hand-parallel sites per axis — the drift class that kept
    # finished axes dark); the per-axis log lines are preserved verbatim via the specs'
    # formatters. The aliases below keep every downstream consumer (computes / sheet
    # writers / snapshot assigns) reference-identical — this swap is contract-invisible,
    # proven by the golden suite passing with NO UPDATE_GOLDEN.
    _axis_stores: Dict[str, dict] = {sp.key: {} for sp in _PER_DEVICE_AXES}
    for hostname, platform, cmd_to_file in all_devices_meta:
        _run_per_device_axes(_axis_stores, hostname, cmd_to_file)
    all_run_configs = _axis_stores["run_configs"]            # NEW-V3.23.146 (raw running-config for golden-config drift)
    all_syslogs = _axis_stores["syslogs"]                    # NEW-V3.23.164 (raw 'show logging'; '' = not collected)
    all_platform_metrics = _axis_stores["platform_metrics"]  # NEW-V3.23.167 (CPU/memory/system facts)
    all_acls = _axis_stores["acls"]
    all_object_groups = _axis_stores["object_groups"]
    all_nat = _axis_stores["nat"]
    all_security = _axis_stores["security"]
    all_config_hygiene = _axis_stores["config_hygiene"]
    all_stp_roots = _axis_stores["stp_roots"]
    all_vpc = _axis_stores["vpc"]
    all_fhrp_detail = _axis_stores["fhrp_detail"]
    all_overlay = _axis_stores["overlay"]
    all_copp = _axis_stores["copp"]
    all_mpls = _axis_stores["mpls"]
    all_lisp = _axis_stores["lisp"]
    all_cts = _axis_stores["cts"]
    all_dmvpn = _axis_stores["dmvpn"]
    all_crypto = _axis_stores["crypto"]
    all_bfd = _axis_stores["bfd"]
    all_aci = _axis_stores["aci"]
    all_sdwan = _axis_stores["sdwan"]
    all_firewall = _axis_stores["firewall"]
    all_ise = _axis_stores["ise"]
    all_fmc = _axis_stores["fmc"]
    all_arista = _axis_stores["arista"]
    all_juniper = _axis_stores["juniper"]
    all_cloud = _axis_stores["cloud"]
    all_fortigate = _axis_stores["fortigate"]
    all_mroute = _axis_stores["mroute"]
    all_ipv6_nd = _axis_stores["ipv6_nd"]
    all_ipv6_routing = _axis_stores["ipv6_routing"]
    all_pim = _axis_stores["pim"]
    all_ipv6_fhs = _axis_stores["ipv6_fhs"]
    all_ntp = _axis_stores["ntp"]
    all_port_security = _axis_stores["port_security"]
    all_storm_control = _axis_stores["storm_control"]
    all_qos_runtime = _axis_stores["qos_runtime"]
    all_shadow_infra = _axis_stores["shadow_infra"]
    all_routing_neighbors = _axis_stores["routing_neighbors"]
    all_redistribution = _axis_stores["redistribution"]

    # Phase 5.7: routing tables (route-aware reachability) - parsed from the already-collected
    # 'show ip route', scoped to the in-scope gateway subnets so the embedded snapshot stays small.
    all_routes: Dict[str, list] = {}
    all_routes_full: Dict[str, dict] = {}                            # NEW-V3.23.97 (FULL table for subnet intelligence; transient, not embedded)
    all_bgp_received: Dict[str, list] = {}                           # NEW-V3.23.97 (BGP RIB when 'show ip bgp' was collected)
    # NEW-V3.23.102: multicast/PTP/ACL-hit collection (inert until the new commands are collected).
    _igmp_groups: set = set()                                        # distinct multicast groups, fleet-wide
    all_igmp_queriers: list = []                                     # [{switch, vlan, querier}]
    all_ptp: Dict[str, dict] = {}                                   # {host: ptp clock health}
    all_acl_hits: Dict[str, int] = {}                              # {"port:proto": total matches}
    _inscope = inscope_subnets(all_interfaces)
    for hostname, platform, cmd_to_file in all_devices_meta:
        rdb = build_routes(cmd_to_file)
        if rdb:
            all_routes_full[hostname] = rdb
        scoped = scope_routes(rdb, _inscope)
        if scoped:
            all_routes[hostname] = scoped
            logger.info(f"  [ROUTE] {hostname}: {len(scoped)} in-scope route(s) embedded")
        bgpr = build_bgp_received(cmd_to_file)
        if bgpr:
            all_bgp_received[hostname] = bgpr
        _igmp_groups.update(build_igmp_groups(cmd_to_file))
        for q in build_igmp_queriers(cmd_to_file):
            all_igmp_queriers.append({"switch": hostname, **q})
        ptp = build_ptp(cmd_to_file)
        if ptp:
            all_ptp[hostname] = ptp
        for k, v in build_acl_hits(cmd_to_file).items():
            all_acl_hits[k] = all_acl_hits.get(k, 0) + v
    all_igmp_groups = sorted(_igmp_groups, key=lambda ip: tuple(int(o) for o in ip.split(".")))

    # --redact: pseudonymize the COLLECTED dataclasses NOW, before the first sheet writer consumes them, so the
    # always-produced .xlsx (built from these, before redact_snapshot touches the JSON) is share-safe too. The
    # snapshot + explorer are redacted later from snap_dict; without this the workbook leaked real serials/IPs/
    # MACs from the same --redact run (audit-3 #8). Opt-in path only; never on a normal run.
    if getattr(args, "redact", False):
        _run_phase("redact collected dataclasses", redact_collected_inplace, all_interfaces, all_device_physical)
        logger.info("[redact] collected inventory (workbook) serials / IPs / MACs pseudonymized (--redact)")

    # Reconcile phantom split-nodes BEFORE any topology / sheet build: a device collected under a suffix-shorter
    # inventory name than its OWN configured hostname is advertised by neighbors over CDP under that configured
    # name; rewrite those advertisements to the inventory key so it renders as ONE node (audit-5 cross-artifact #1).
    _n_reconciled = _run_phase("reconcile CDP split-node names", reconcile_cdp_neighbor_names,
                               all_interfaces, all_device_physical, _default=0)
    if _n_reconciled:
        logger.info(f"[topology] reconciled {_n_reconciled} CDP advertisement(s) to inventory key(s) (split-node fix)")

    # Phase 6: Write interface rows to template sheet (each host guarded so one
    # host's write failure can't abort the run before wb.save()).
    logger.info("\n[Phase 6] Writing interface rows ...")
    for hostname, platform, cmd_to_file in all_devices_meta:
        iface_rows = all_interfaces.get(hostname, {})
        if not iface_rows:
            logger.warning(f"  [SKIP] No parsed interfaces for {hostname}")
            continue
        _run_phase(f"interface rows ({hostname})", append_interface_rows,
                   ws, header_row, col_map, hostname, iface_rows)
        logger.info(f"  [OK] {len(iface_rows)} rows for {hostname}")

    # Phases 7-20 are pure sheet writers; each is guarded so a single sheet's
    # failure logs and is skipped rather than sinking the whole workbook.
    # Phase 7: Write Switch Inventory sheet (NEW-V12)
    logger.info("\n[Phase 7] Writing Switch Inventory sheet ...")
    _run_phase("Switch Inventory sheet", write_device_inventory_sheet, wb, all_device_physical)

    # Phase 8: Write SVI / Gateway sheet (NEW-V14.3)
    logger.info("\n[Phase 8] Writing SVI / Gateway sheet ...")
    _run_phase("SVI / Gateway sheet", write_svi_gateway_sheet, wb, all_interfaces)

    # Phase 9: Write STP Detail sheet (NEW-V14.7)
    logger.info("\n[Phase 9] Writing STP Detail sheet ...")
    _run_phase("STP Detail sheet", write_stp_detail_sheet, wb, all_interfaces)

    # Phase 10: Write VLAN Census sheet (NEW-V14.8)
    logger.info("\n[Phase 10] Writing VLAN Census sheet ...")
    _run_phase("VLAN Census sheet", write_vlan_census_sheet, wb, all_interfaces)

    # Phase 11: Write Endpoint Census sheet (NEW-V14.8)
    logger.info("\n[Phase 11] Writing Endpoint Census sheet ...")
    _run_phase("Endpoint Census sheet", write_endpoint_census_sheet, wb, all_interfaces)

    # Phase 12: Write Move Groups sheet (NEW-V14.9)
    logger.info("\n[Phase 12] Writing Move Groups sheet ...")
    _run_phase("Move Groups sheet", write_move_group_sheet, wb, all_interfaces)

    # Phase 13: Write Topology Links sheet (NEW-V14.10)
    logger.info("\n[Phase 13] Writing Topology Links sheet ...")
    _run_phase("Topology Links sheet", write_topology_sheet, wb, all_interfaces)

    # Phase 14: Findings / Risk Register (NEW-V15)
    logger.info("\n[Phase 14] Writing Findings sheet ...")
    _run_phase("Findings sheet", write_findings_sheet, wb, all_interfaces)

    # Phase 15: Capacity / Consolidation (NEW-V15)
    logger.info("\n[Phase 15] Writing Capacity sheet ...")
    _run_phase("Capacity sheet", write_capacity_sheet, wb, all_device_physical)

    # Phase 15b: Endpoint Intelligence (NEW-V3.23.95). Compute ONCE here (vendor via offline OUI +
    # inferred class); reused by the sheet AND the snapshot (one source of truth -> explorer + runbook).
    logger.info("\n[Phase 15b] Writing Endpoint Intelligence sheet ...")
    endpoint_identity = _run_phase("Endpoint identity", compute_endpoint_identity, all_interfaces, _default=[])
    _run_phase("Endpoint Intelligence sheet", write_endpoint_intelligence_sheet, wb, endpoint_identity)

    # Phase 16: Interface Health counters (NEW-V15)
    logger.info("\n[Phase 16] Writing Interface Health sheet ...")
    _run_phase("Interface Health sheet", write_interface_health_sheet, wb, all_cmd_to_files)

    # Phase 17: Security Posture (NEW-V15)
    logger.info("\n[Phase 17] Writing Security Posture sheet ...")
    _run_phase("Security Posture sheet", write_security_posture_sheet, wb, all_cmd_to_files)

    # Phase 18: Routing Adjacencies (NEW-V15)
    logger.info("\n[Phase 18] Writing Routing Adjacencies sheet ...")
    _run_phase("Routing Adjacencies sheet", write_routing_adjacency_sheet, wb, all_cmd_to_files)

    # Phase 19-20: Causality Chains + Failure Impact (NEW-V3.16 intelligence layer).
    # NEW-V3.23.91: compute each ONCE here and reuse for the sheet, the executive summary, and the
    # snapshot. The heavy model rebuild + per-(host,VLAN) articulation walk / per-switch removal
    # simulation previously ran 2-3x per report (the snapshot-build phase recomputed both), which was
    # ~2.5 min of the run on the real fleet. Local import keeps these out of this module's namespace
    # (the test_package re-export contract asserts they are NOT attributes of this module).
    from cisco_toolkit.analyze import compute_causality_chains, compute_failure_impact
    logger.info("\n[Phase 19] Writing Causality Chains sheet ...")
    causality_chains = _run_phase("Causality Chains", compute_causality_chains, all_interfaces, _default=[])
    _run_phase("Causality Chains sheet", write_causality_chains_sheet, wb, causality_chains)

    # Phase 20: Failure Impact simulation (NEW-V3.16 intelligence layer)
    logger.info("\n[Phase 20] Writing Failure Impact sheet ...")
    failure_impact = _run_phase("Failure Impact", compute_failure_impact, all_interfaces, _default=[])
    _run_phase("Failure Impact sheet", write_failure_impact_sheet, wb, failure_impact)
    _run_phase("Link Centrality sheet", write_link_centrality_sheet, wb, all_interfaces)   # NEW-V3.23.88 (chokepoint links)

    # Phase 23: Physical Health (L1) - NEW-V3.18. Sheet writers must precede wb.save(); the
    # phase number is a logical label (21/22 are post-save: snapshot/diagram + HTML). The
    # returned records are injected into snap_dict below so they reach both JSON and HTML.
    logger.info("\n[Phase 23] Writing Physical Health sheet ...")
    physical_health = _run_phase("Physical Health", write_physical_health_sheet,
                                 wb, all_interfaces, all_cmd_to_files, all_device_physical,
                                 _default=[])

    # Phase 24: Flow Trace (L1->L3) - NEW-V3.19. Optional; only when BOTH endpoint IPs given.
    flow_trace = None
    if args.flow_src and args.flow_dst:
        logger.info(f"\n[Phase 24] Tracing flow {args.flow_src} -> {args.flow_dst} ...")
        try:
            flow_trace = trace_full_flow(args.flow_src, args.flow_dst, all_interfaces)
            write_flow_trace_sheet(wb, flow_trace)
        except Exception as e:
            logger.warning(f"  [WARN] Flow Trace failed: {e}")
            flow_trace = None
    elif args.flow_src or args.flow_dst:
        logger.warning("  [WARN] --flow-src and --flow-dst must be given together; skipping Flow Trace.")

    # Phase 24b: Representative Flow Paths - NEW-V3.23.126. The static twin of the explorer's Flow
    # Simulator: one representative endpoint per VLAN, traced L1->L3 (reuses trace_full_flow). Always
    # written; NOT embedded in the snapshot (the hop tie-break is topology-dependent -> not golden-stable).
    flow_paths = _run_phase("Flow paths", compute_flow_paths, all_interfaces, endpoint_identity,
                            _default={"flows": [], "summary": {}})
    _run_phase("Flow Paths sheet", write_flow_paths_sheet, wb, flow_paths)

    # Phase 25: L3 Forwarding Map - NEW-V3.20 (pre-save; records reused in snapshot)
    logger.info("\n[Phase 25] Writing L3 Forwarding Map sheet ...")
    l3_forwarding = _run_phase("L3 Forwarding Map", write_l3_forwarding_sheet,
                               wb, all_interfaces, all_cmd_to_files, _default=[])

    # Phase 26: Cross-Layer Analysis - NEW-V3.21 (consumes physical_health + l3_forwarding).
    # A failed dependency map falls back to an empty skeleton so Phases 28-29 still run.
    logger.info("\n[Phase 26] Writing Cross-Layer Analysis sheet ...")
    dep_map = _run_phase("dependency map", build_dependency_map,
                         all_interfaces, physical_health, l3_forwarding,
                         _default=_empty_dep_map()) or _empty_dep_map()
    cross_layer = _run_phase("Cross-Layer correlations", compute_cross_layer_correlations,
                             dep_map, _default=[])
    _run_phase("Cross-Layer Analysis sheet", write_cross_layer_sheet, wb, cross_layer)

    # Phase 27: Protocol Health - NEW-V3.22 (re-parses collected protocol output)
    logger.info("\n[Phase 27] Writing Protocol Health sheet ...")
    protocol_health = _run_phase("Protocol Health", compute_protocol_health,
                                 all_interfaces, all_cmd_to_files, _default=[])
    _run_phase("Protocol Health sheet", write_protocol_health_sheet, wb, protocol_health)
    # Phase 27b: Protocol Intelligence (NEW-V3.23.100). Join the protocol_health rows against the
    # offline protocol-state doctrine -> meaning + likely cause + remediation. Derived from
    # protocol_health (no re-parse); compute once -> sheet + snapshot (one source of truth).
    protocol_intelligence = _run_phase("Protocol intelligence", compute_protocol_intelligence,
                                       protocol_health, _default=[])
    _run_phase("Protocol Intelligence sheet", write_protocol_intelligence_sheet, wb, protocol_intelligence)
    # Phase 27c: Service Map (NEW-V3.23.101). Resolve ACL L4 port references + fleet multicast activity to
    # named services via the offline port registry (portdb). ACL refs are design intent (Inferred), not
    # active traffic. Compute once -> sheet + snapshot (one source of truth).
    service_map = _run_phase("Service map", compute_service_map, all_acls, all_interfaces, _default={},
                             igmp_groups=all_igmp_groups, ptp=all_ptp,
                             igmp_queriers=all_igmp_queriers, acl_hits=all_acl_hits)
    _run_phase("Service Map sheet", write_service_map_sheet, wb, service_map)
    # Phase 27c-bis: Multicast Intelligence (NEW-V3.23.115). Media-fabric deep-dive over service_map.multicast
    # (MAC-address aliasing / IGMP querier coverage / PTP timing tree). Compute once -> sheet + snapshot +
    # punch-list fold (one source of truth).
    multicast_intelligence = _run_phase("Multicast intelligence", compute_multicast_intelligence,
                                        service_map, all_interfaces, _default={})
    _run_phase("Multicast Intelligence sheet", write_multicast_intelligence_sheet, wb, multicast_intelligence)
    # Phase 27c-ter: Hardware Lifecycle Risk (NEW-V3.23.117). Per-device EoL / end-of-support band from the
    # offline eoldb KB (a top reason orgs migrate). asof defaults to now (the assessment date). Compute once
    # -> sheet + snapshot + punch-list fold.
    _dev_lifecycle = {dp.hostname: {"model": dp.model, "sw_version": dp.sw_version} for dp in all_device_physical}
    lifecycle_risk = _run_phase("Lifecycle risk", compute_lifecycle_risk, _dev_lifecycle,
                                asof=collected_at, _default={})   # provenance: bands as-of collection, not regen-day
    _run_phase("Lifecycle Risk sheet", write_lifecycle_risk_sheet, wb, lifecycle_risk)

    # Phase 27d: Collection completeness (NEW-V3.23.109). The pre-assessment blind-spot report -- which
    # INVENTORY (devices.json) devices were not / only partially collected, so the gaps are explicit
    # rather than buried in band counts. Compute once -> sheet + snapshot (one source of truth).
    _inventory_hosts = [d.get("hostname", "") for d in devices]
    collection_completeness = _run_phase("Collection completeness", compute_collection_completeness,
                                         _inventory_hosts, all_cmd_to_files, _default={})
    # parse_yield_report() here and at snapshot assembly return IDENTICAL content: all parsing
    # is complete before either runs, and the sheet writers call no _safe_parse (excel.py only
    # loads raw text for its 3 re-parse sheets), so the ledger is quiescent between the two.
    _run_phase("Collection Completeness sheet", write_collection_completeness_sheet, wb,
               collection_completeness, parse_yield_report())

    # Phase 28: Health Scores - NEW-V3.23 (synthesises L1/L3/cross-layer/protocol findings)
    logger.info("\n[Phase 28] Writing Health Scores sheet ...")
    data_quality = _run_phase("data quality", compute_data_quality, all_cmd_to_files, _default={}) or {}
    health_scores = _run_phase("Health Scores", compute_health_scores,
                               all_interfaces, physical_health, l3_forwarding,
                               cross_layer, protocol_health, _default=[],
                               data_quality=data_quality, security=all_security)
    _run_phase("Health Scores sheet", write_health_scores_sheet, wb, health_scores)

    # Phase 29: Migration Readiness - NEW-V3.23 (per move-group 10-check verdict)
    logger.info("\n[Phase 29] Writing Migration Readiness sheet ...")
    move_groups = _run_phase("move groups", compute_move_groups, all_interfaces, _default=[])
    migration_readiness = _run_phase(
        "Migration Readiness", compute_migration_readiness,
        all_interfaces, move_groups, health_scores, physical_health, l3_forwarding,
        cross_layer, protocol_health, dep_map, _default=[])
    _run_phase("Migration Readiness sheet", write_migration_readiness_sheet, wb, migration_readiness)
    _run_phase("Wave Sequencing sheet", write_wave_sequencing_sheet, wb, all_interfaces, move_groups)   # NEW-V3.23.89 (make-before-break vs hard cutover; needs move_groups computed just above)
    # Phase 29b: Endpoint Dependencies (NEW-V3.23.96). Needs both endpoint_identity (Phase 15b) and
    # move_groups (just above); compute once -> sheet + snapshot (one source of truth).
    endpoint_dependencies = _run_phase("Endpoint dependencies", compute_endpoint_dependencies,
                                       endpoint_identity, move_groups, _default={})
    _run_phase("Endpoint Dependencies sheet", write_endpoint_dependencies_sheet, wb, endpoint_dependencies)
    # Phase 29c: Subnet & routing reachability (NEW-V3.23.97). Full route table (transient) + SVIs +
    # move_groups -> per-device destination/reachable + per-group source<->destination; compute once.
    subnet_intelligence = _run_phase("Subnet intelligence", compute_subnet_intelligence,
                                     all_interfaces, all_routes_full, move_groups, all_bgp_received,
                                     _default={})
    _run_phase("Subnet Reachability sheet", write_subnet_reachability_sheet, wb, subnet_intelligence)
    # Phase 29d: Migration scenarios (NEW-V3.23.98). Synthesise readiness + wave sequencing + health
    # bands into a per-group cutover-scenario recommendation; compute once -> sheet + snapshot. Local
    # import of compute_wave_sequencing keeps it out of this module's namespace (test_package contract).
    from cisco_toolkit.analyze import compute_wave_sequencing as _compute_ws
    _ws_seq = _compute_ws(all_interfaces, move_groups)
    migration_scenarios = _run_phase("Migration scenarios", compute_migration_scenarios,
                                     migration_readiness, _ws_seq, health_scores, _default={})
    _run_phase("Migration Scenarios sheet", write_migration_scenarios_sheet, wb, migration_scenarios)

    # Phase 30: Score Sensitivity - NEW-V3.23.5 (OAT robustness sweep over scoring weights)
    logger.info("\n[Phase 30] Writing Score Sensitivity sheet ...")
    score_sensitivity = _run_phase("Score Sensitivity", compute_score_sensitivity,
                                   all_interfaces, physical_health, l3_forwarding,
                                   cross_layer, protocol_health, _default=[],
                                   security=all_security)
    _run_phase("Score Sensitivity sheet", write_score_sensitivity_sheet, wb, score_sensitivity)

    # Phase 30b: Score Calibration - NEW-V3.23.47 (fleet band-discrimination diagnostic + re-banding suggestion)
    calibration = _run_phase("Score Calibration", compute_calibration_report, health_scores, _default={})
    _run_phase("Score Calibration sheet", write_calibration_sheet, wb, calibration)

    # Phase 30c: NAT Inventory - NEW-V3.23.50 (every static/dynamic NAT rule the migration must recreate)
    _run_phase("NAT Inventory sheet", write_nat_sheet, wb, all_nat)
    _run_phase("Config Compliance sheet", write_security_sheet, wb, all_security)
    framework_cov = _run_phase("Framework coverage", compute_framework_coverage, all_security, _default={})  # W2-3: computed once here; reused for snap_dict below
    _run_phase("Framework Coverage sheet", write_framework_coverage_sheet, wb, framework_cov)
    detector_schema = _run_phase("Detector schema", compute_detector_schema, _default={}) or {}  # J1: computed once here; reused for snap_dict below
    _run_phase("Detector Schema sheet", write_detector_schema_sheet, wb, detector_schema)
    _run_phase("Config Hygiene sheet", write_config_hygiene_sheet, wb, all_config_hygiene)
    _run_phase("STP Root Bridges sheet", write_stp_roots_sheet, wb, all_stp_roots, all_interfaces)

    # Phase 30d: Migration Punch-List - NEW-V3.23.63 (consolidated, severity-ranked fix-this-first roll-up);
    # V3.23.64 folds in the cross-switch L2 checks (addressing / FHRP / trunk-native / link duplex-speed).
    _stp_findings = stp_root_findings(all_stp_roots, all_interfaces)
    _l2 = {"addressing": compute_addressing_conflicts(all_interfaces),
           "fhrp": compute_fhrp_consistency(all_interfaces),
           "trunk_native": compute_trunk_native_mismatches(all_interfaces),
           "link_phy": compute_duplex_speed_mismatches(all_interfaces)}
    _hostname_mismatches = compute_hostname_mismatches(all_device_physical)   # NEW-V3.23.68
    # trunk-capture gaps computed first (fail-soft by construction) so the drift detector can DISCLOSE
    # devices whose native-VLAN-1 exposure is not assessable (l2-native-vlan-1 abstains_when, mechanical).
    _trunk_gaps = _run_phase("Trunk-capture gaps", compute_trunk_capture_gaps,
                             all_interfaces, all_cmd_to_files, _default=[])
    _drift = _run_phase("Operational drift", compute_operational_drift,
                        all_interfaces, all_device_physical, _trunk_gaps, _default=[])  # NEW-V3.23.93 (false-health)
    _ptp_readiness = _run_phase("PTP readiness", compute_ptp_readiness, service_map, _default=[])  # NEW-V3.23.108
    # NEW-V3.23.115: fold the media-fabric findings (MAC-aliasing / IGMP querier gaps) into the punch-list;
    # PTP is already folded via _ptp_readiness, so exclude the ptp-dormant risk to avoid a duplicate item.
    _media_risks = [r for r in ((multicast_intelligence or {}).get("risks") or [])
                    if r.get("kind") in ("mac-alias", "querier-gap")]
    # NEW-V3.23.169: the V3.23.164-.167 axes are COMPUTED here (hoisted from their original Phase 30d-*
    # slots) so the punch-list and the executive brief can consume their findings; their workbook SHEETS
    # are still written at the original phase slots below (compute-once, sheet-order unchanged).
    # _dev_platform is hoisted with them (software_risk needs it; the remediation phase reuses it).
    _dev_platform = {dp.hostname: {"platform": dp.platform} for dp in all_device_physical}
    syslog_intelligence = _run_phase("Syslog intelligence", compute_syslog_intelligence,
                                     all_syslogs, _default={})
    qos_audit = _run_phase("QoS audit", compute_qos_audit,
                           all_run_configs, sorted(all_syslogs), _default={})
    software_risk = _run_phase("Software risk screening", compute_software_risk,
                               all_run_configs, _dev_lifecycle, _dev_platform,
                               sorted(all_syslogs), _default={})
    platform_health = _run_phase("Platform health", compute_platform_health,
                                 all_platform_metrics, _default={})
    # NEW-V3.23.172: golden-drift is COMPUTED here (hoisted from its Phase 30d-quinquies slot, the
    # same compute-once hoist V3.23.169 applied to the quartet above) so the Device Risk Register
    # can join it; its workbook sheet is still written at the original phase slot below.
    _golden_lines = None
    if args.golden_config:
        try:
            with open(args.golden_config, encoding="utf-8") as f:
                _golden_lines = f.read().splitlines()
            logger.info(f"[Phase 30d] Golden-config baseline: {len(_golden_lines)} line(s) from {args.golden_config}")
        except OSError as e:
            logger.warning(f"  --golden-config not read ({e}); falling back to the auto-derived majority baseline.")
    golden_drift = _run_phase("Golden-config drift", compute_golden_drift,
                              all_run_configs, _golden_lines, _default={})
    # roadmap I2: decompose the golden-config drift per policy FEATURE (aaa/ntp/snmp/...). Pure projection of golden_drift.
    feature_compliance = _run_phase("Feature compliance", compute_feature_compliance, golden_drift, _default={})
    # roadmap G1: offline Batfish-style ACL line-reachability / shadow proof over the already-parsed ACLs + object-groups.
    acl_line_reachability = _run_phase("ACL line-reachability", compute_filter_line_reachability,
                                       {"acls": all_acls, "object_groups": all_object_groups}, _default={})
    # NEW-V3.23.172: the Device Risk Register -- the per-ASSET synthesis. Joins the 11 per-device
    # axes already computed above into one dossier per box (risk_index = topology impact x stacked
    # exposure + named compound patterns CR-01..CR-06), BEFORE the punch-list so the compound
    # findings fold into the decision layer (and reach every downstream surface via snap["punchlist"]).
    # Plan-A #15 strangler (Stage-3 positional collapse): thread the analyze-stage synthesis feeds
    # through the typed AnalysisContext and read them back via the thin ctx-adapters, so the three
    # WIDE compute_* below stop threading 15/19/14 locals positionally through _run_phase (a silent
    # reorder could corrupt them). _actx is populated per-consumer as each feed becomes available;
    # the public compute_* keep their explicit-arg signatures (see the adapters just after main()).
    _actx = AnalysisContext(
        health_scores=health_scores, failure_impact=failure_impact, lifecycle_risk=lifecycle_risk,
        software_risk=software_risk, platform_health=platform_health,
        syslog_intelligence=syslog_intelligence, qos_audit=qos_audit, golden_drift=golden_drift,
        all_security=all_security, all_config_hygiene=all_config_hygiene,
        all_stp_roots=all_stp_roots, all_vpc=all_vpc, physical_health=physical_health,
        protocol_health=protocol_health, move_groups=move_groups)
    device_dossiers = _run_phase("Device risk register", _device_dossiers, _actx, _default={})
    # EDA-style physical cable map (SSOT for the explorer + webapp cable-map views): a node/port/cable
    # graph, role-tiered lanes, LAG bundled, op-status DERIVED from interface state (coverage-honest --
    # uncollected devices/ports are [NOT OBSERVED] neutral, never a fake green).
    cable_map = _run_phase("Cable map", compute_cable_map, all_interfaces, health_scores, _default={})
    _run_phase("Cabling Schedule sheet", write_cabling_schedule_sheet, wb, cable_map)   # EDA-style cable schedule (op-status + LAG) from the cable-map SSOT
    # NEW-V3.23.117: lifecycle risk is kept as its OWN axis (sheet / cockpit / runbook §4.1), NOT folded into
    # the punch-list -- its band is date-relative, which would make the frozen golden punch-list date-dependent.
    # (The V3.23.172 compound patterns MAY reference the EoL band, but only when stacked with a second
    # independent risk axis -- the synthetic golden fleet's models are outside the EoL KB, so the frozen
    # golden punch-list stays date-free.)
    _actx.device_dossiers = device_dossiers          # feed the just-computed dossiers into the punch-list
    _actx.cross_layer = cross_layer
    _actx.l3_forwarding = l3_forwarding
    _actx.stp_findings = _stp_findings
    _actx.l2 = _l2
    _actx.hostname_mismatches = _hostname_mismatches
    _actx.drift = _drift
    _actx.ptp_readiness = _ptp_readiness
    _actx.media_risks = _media_risks
    punchlist = _run_phase("Migration Punch-List", _punchlist, _actx, _default=[])
    _run_phase("Migration Punch-List sheet", write_punchlist_sheet, wb, punchlist)
    _run_phase("Device Risk Register sheet", write_device_risk_sheet, wb, device_dossiers)

    # Phase 30d-ter: Remediation Plan - NEW-V3.23.116. The assess->act layer: turn the SAME structured finding
    # sources into per-device, platform-tagged, copy-pasteable Cisco config snippets (review-only). Reuses
    # the _l2 bundle / _stp_findings / hygiene / security / multicast_intelligence already computed above.
    # (_dev_platform is built above the punch-list since V3.23.169.)
    remediation_plan = _run_phase("Remediation plan", compute_remediation_plan,
                                  _dev_platform, _l2, _stp_findings, all_config_hygiene, all_security,
                                  multicast_intelligence, move_groups, _default={})
    _run_phase("Remediation Plan sheet", write_remediation_plan_sheet, wb, remediation_plan)

    # Phase 30d-quater: Cutover VALIDATION plan - NEW-V3.23.143 (assess->ACT->verify). From the current-state
    # topology, generate the per-wave checks an engineer runs after each cutover to PROVE it worked -- gateway
    # up, FHRP healthy, endpoints reach their gateway, routing adjacencies re-established, STP root unchanged,
    # port-channels bundled - each with the expected good result captured from the pre-cutover state. Pure
    # synthesis of already-collected data; compute once -> sheet + snapshot (one source of truth).
    validation_plan = _run_phase("Validation plan", compute_validation_plan,
                                 all_interfaces, move_groups, all_routing_neighbors, all_stp_roots,
                                 _dev_platform, protocol_health=protocol_health, _default={})   # NRFU #18: bundle baseline reflects real (D)/down members, not a blanket all-(P)
    _run_phase("Cutover Validation sheet", write_validation_plan_sheet, wb, validation_plan)

    # Phase 30d-quinquies: Golden-config DRIFT - NEW-V3.23.146. Per-device running-config drift vs a baseline:
    # the user's --golden-config file (required directives) if supplied, else the fleet's auto-derived MAJORITY
    # baseline (de-facto standard) -> flag the outliers. Compute once -> sheet + snapshot (one source of truth).
    # Computed ABOVE the punch-list since V3.23.172 (Device Risk Register join); the sheet keeps its slot.
    _run_phase("Golden-Config Drift sheet", write_golden_drift_sheet, wb, golden_drift)
    _run_phase("Feature Compliance sheet", write_feature_compliance_sheet, wb, feature_compliance)   # roadmap I2
    _run_phase("ACL Shadow sheet", write_acl_shadow_sheet, wb, acl_line_reachability)                # roadmap G1
    # roadmap B: opt-in external source-of-truth reconcile (--import-inventory FILE). Off by default -> golden-safe.
    external_reconcile = None
    if args.import_inventory:
        _declared = read_inventory_file(args.import_inventory)
        external_reconcile = reconcile_external(
            {"lifecycle_risk": lifecycle_risk, "health_scores": health_scores,
             "collection_completeness": collection_completeness}, _declared)
        _run_phase("SoT Reconcile sheet", write_external_reconcile_sheet, wb, external_reconcile)
        logger.info(f"[Phase 30g] SoT reconcile: {len(_declared)} declared vs "
                    f"{external_reconcile['summary'].get('n_observed', 0)} observed -> "
                    f"{external_reconcile['summary'].get('n_rows', 0)} drift row(s) from {args.import_inventory}")
    # roadmap K1, WIDENED in Plan A / Tier-2 #6: capture-integrity now inspects EVERY collected capture
    # (all ~160 commands/device via all_cmd_to_files), not just run-config — streamed one body at a time
    # so the fleet's raw output is never held in memory at once. The per-device _capture_meta.json
    # sidecar (live-collect) adds `unverified_prompt` for timing-fallback captures; absent meta (every
    # older collection) makes no prompt claim. Together with snap['parse_yield'] this separates
    # "collection problem" (zero-yield + non-ok capture) from "parser format gap" (zero-yield + clean).
    _ci_meta: Dict[str, Dict[str, str]] = {}
    for _h, _c2f in (all_cmd_to_files or {}).items():
        for _p in (_c2f or {}).values():
            _m = load_capture_meta(os.path.dirname(_p))
            if _m:
                _ci_meta[_h] = _m
            break                                   # one device dir per host — first path suffices
    capture_integrity = _run_phase("Capture integrity", compute_capture_integrity_from_paths,
                                   all_cmd_to_files, _ci_meta, _default={})
    _run_phase("Capture Integrity sheet", write_capture_integrity_sheet, wb, capture_integrity)
    # roadmap G4: opt-in failure-injection what-if (--scenario FILE). Golden-safe (off by default).
    whatif_result = None
    if args.scenario:
        try:
            with open(args.scenario, encoding="utf-8") as _sf:
                _scen = json.load(_sf)
        except Exception as e:
            _scen = []
            logger.warning(f"  --scenario not read ({e}); skipping what-if.")
        _scen = _scen if isinstance(_scen, list) else (_scen.get("scenarios") or [])
        whatif_result = run_scenarios({"routes": all_routes}, _scen)
        _run_phase("Failure What-If sheet", write_whatif_sheet, wb, whatif_result)
    # roadmap G3: opt-in named path/segmentation intents (--path-intents FILE). Golden-safe.
    path_intents = None
    if args.path_intents:
        try:
            with open(args.path_intents, encoding="utf-8") as _pf:
                _intents = json.load(_pf)
        except Exception as e:
            _intents = []
            logger.warning(f"  --path-intents not read ({e}); skipping.")
        _intents = _intents if isinstance(_intents, list) else (_intents.get("assertions") or _intents.get("intents") or [])
        path_intents = evaluate_path_assertions({"routes": all_routes}, _intents)
        _run_phase("Path Assertions sheet", write_path_intents_sheet, wb, path_intents)

    # Phase 30d-sexies: Syslog intelligence - NEW-V3.23.164. The NOS-style operational log analysis
    # (Cisco's Network Optimization Service names syslog analysis as an analytic pillar): per-device
    # severity profile + top messages + deterministic detections (MAC flap / err-disable / link flap /
    # environmental / TCAM / crash / login failures) from the already-collected 'show logging' buffers.
    # Devices without log output are DECLARED not-collected, never scored. Computed ABOVE the
    # punch-list since V3.23.169 (decision-layer fold); the sheet keeps its original slot.
    _run_phase("Syslog Intelligence sheet", write_syslog_intelligence_sheet, wb, syslog_intelligence)

    # Phase 30d-septies: QoS audit - NEW-V3.23.165. QoS is a named HLD/LLD design domain (trust
    # boundary at the access edge, per-hop consistency fleet-wide) that had zero coverage. Audits the
    # CONFIGURED posture from the captured running-configs (mechanisms, trust, voice-edge policy,
    # inert/dangling policy, fleet consistency); a device without a full capture is DECLARED not
    # assessable, never scored. all_syslogs carries EVERY inventory host (recorded unconditionally in
    # the collection loop), so it doubles as the fleet roster. Computed ABOVE the punch-list since
    # V3.23.169 (decision-layer fold); the sheet keeps its original slot.
    _run_phase("QoS Audit sheet", write_qos_audit_sheet, wb, qos_audit)

    # Phase 30d-octies: Software risk screening - NEW-V3.23.166. The last NOS analytic pillar
    # (software risk analysis): config-evidenced attack surfaces joined to a curated landmark-advisory
    # KB (exploited-in-the-wild surfaces first) + cautious software-TRAIN lifecycle bands. SCREENING,
    # not a vulnerability scan -- the note tells the reader to validate releases with the Cisco PSIRT
    # Software Checker. Reuses _dev_lifecycle ({host:{model,sw_version}}, Phase 27c-ter) and
    # _dev_platform; all_syslogs doubles as the full fleet roster. Computed ABOVE the punch-list
    # since V3.23.169 (decision-layer fold); the sheet keeps its original slot.
    _run_phase("Software Risk sheet", write_software_risk_sheet, wb, software_risk)

    # Phase 30d-nonies: Platform health - NEW-V3.23.167. The pre-migration control-plane capacity
    # question ("is this control plane already stressed before we add protocol churn?"): CPU 5-min
    # average + processor-pool memory (IOS) / system resources (NX-OS), banded Hot/Elevated/OK with
    # doctrine findings. SINGLE point-in-time sample (the note says so -- correlate with the syslog
    # axis and re-sample before the window); devices without capacity output are DECLARED not
    # collected. Computed ABOVE the punch-list since V3.23.169 (decision-layer fold); the sheet
    # keeps its original slot.
    _run_phase("Platform Health sheet", write_platform_health_sheet, wb, platform_health)

    # Phase 30d-bis: Application Intelligence - NEW-V3.23.112. Synthesize endpoint_identity +
    # endpoint_dependencies + service_map + health_scores + move_groups + punchlist into named
    # application DOMAINS (workloads) with footprint, criticality tier, health rollup, migration-wave
    # span, and standards-grounded migration risks (PTP boundary-clock / ST 2059, IGMP querier
    # continuity / RFC 4541, ST 2022-7 dual-path, NEXIS dual-leg). Pure synthesis of already-computed
    # records; compute once -> sheet + snapshot (one source of truth).
    application_intelligence = _run_phase("Application intelligence", compute_application_intelligence,
                                          all_interfaces, endpoint_identity, endpoint_dependencies,
                                          service_map, health_scores, move_groups, punchlist,
                                          subnet_intelligence=subnet_intelligence, _default={})
    _run_phase("Application Intelligence sheet", write_application_intelligence_sheet, wb, application_intelligence)
    # Phase 30d-bis2: Segmentation / isolation audit (NEW-V3.23.118). L3 isolation posture per application
    # domain (VRF separation + gateway-ACL coverage) from the gateway SVIs + the domain map. Compute once ->
    # sheet + snapshot.
    segmentation = _run_phase("Segmentation audit", compute_segmentation,
                              all_interfaces, application_intelligence, _default={})
    _run_phase("Segmentation sheet", write_segmentation_sheet, wb, segmentation)

    # Phase 30d-bis3: Per-VLAN cutover matrix - MASTER_PLAN 2026-07-05 §4.3 ("the single most
    # valuable artifact"): the sheet a cutover team runs the maintenance window from -- one row per
    # evidenced VLAN, every column joined from the axes computed above (STP root + default-election
    # smell, FHRP brief+detail, gateway SVIs, endpoint census, application domains, move-group wave
    # + make-before-break vs hard-cutover scenario, readiness pull-through, multicast / DHCP-relay
    # dependency flags), with the window / rollback-owner columns deliberately left to the human.
    # Pure synthesis of already-computed records; compute once -> sheet + snapshot (one source of
    # truth). Sits AFTER its inputs exist: move_groups / _ws_seq (Phase 29d) / migration_readiness /
    # multicast_intelligence (Phase 27c-bis) / application_intelligence (Phase 30d-bis).
    vlan_cutover = _run_phase("VLAN cutover matrix", compute_vlan_cutover_matrix,
                              all_interfaces, all_stp_roots, all_fhrp_detail,
                              endpoint_identity, application_intelligence,
                              move_groups, _ws_seq, migration_readiness,
                              multicast_intelligence, _default=[])
    _run_phase("VLAN Cutover Matrix sheet", write_vlan_cutover_sheet, wb, vlan_cutover)

    # Phase 30e: Executive Summary - NEW-V3.23.75 (one-page synthesis, landed as the FIRST workbook tab)
    logger.info("\n[Phase 30e] Writing Executive Summary sheet ...")
    # NEW-V3.23.120: the cross-axis migration brief -- one headline per assessment axis (the 7 new axes +
    # health + punch-list + readiness) rolled up into one decision-grade synthesis. Compute once -> the
    # Executive Summary sheet leads with it + snapshot.
    _actx.punchlist = punchlist                      # feed the just-computed punch-list into the brief
    _actx.migration_readiness = migration_readiness
    _actx.application_intelligence = application_intelligence
    _actx.segmentation = segmentation
    _actx.multicast_intelligence = multicast_intelligence
    _actx.remediation_plan = remediation_plan
    _actx.endpoint_identity = endpoint_identity      # SSOT: n_endpoints == len(endpoint_identity)
    executive_brief = _run_phase("Executive brief", _executive_brief, _actx,
                                 _default={"_unavailable": True})   # sentinel: a CRASH != a legit-empty brief (wave R2-4-01)
    # SSOT: inject the canonical VLAN count (vlan_inventory) into the brief's scale HERE, BEFORE the
    # Executive Summary sheet reads it. At this point interfaces are still InterfaceData OBJECTS (the dict
    # serialization happens during snapshot assembly, after the workbook), so build the minimal vlan/
    # vlan_name dict vlan_inventory reads. The snapshot re-injects the same value at assembly (idempotent).
    if isinstance(executive_brief, dict) and isinstance(executive_brief.get("scale"), dict):
        _ifd = {h: {p: {"vlan": getattr(d, "vlan", ""), "vlan_name": getattr(d, "vlan_name", "")}
                    for p, d in (ports or {}).items()}
                for h, ports in (all_interfaces or {}).items()}
        executive_brief["scale"]["n_vlans"] = len(vlan_inventory(
            {"interfaces": _ifd, "l3_forwarding": l3_forwarding, "service_map": service_map}))
        # ...and the genuinely-COLLECTED count HERE too (QA F1), so the Exec Summary Fleet-posture block reads
        # canonical n_collected (253 of 303 inventoried) instead of a local recompute. collection_completeness
        # is already computed (Phase 27d). Assembly re-injects the same value at snapshot time (idempotent).
        executive_brief["scale"]["n_collected"] = ((collection_completeness or {}).get("summary") or {}).get("complete")
    _run_phase("Executive Summary sheet", write_executive_summary_sheet, wb,
               health_scores, punchlist, migration_readiness, failure_impact,   # NEW-V3.23.91: reuse precomputed fi
               brief=executive_brief,
               provenance={"script_version": f"V{__version__}",                 # P3-E3: self-trace the landing sheet
                           "generated_at": datetime.now().isoformat(),
                           "snapshot": os.path.splitext(os.path.basename(str(out_xlsx or "")))[0]})
    _run_phase("Protocol Boundaries sheet", write_protocol_boundaries_sheet, wb, all_routing_neighbors, all_redistribution)
    _run_phase("Addressing Conflicts sheet", write_addressing_conflicts_sheet, wb, all_interfaces)
    _run_phase("FHRP Consistency sheet", write_fhrp_consistency_sheet, wb, all_interfaces)
    _run_phase("Trunk Native-VLAN sheet", write_trunk_native_sheet, wb, all_interfaces)
    _run_phase("Link Duplex-Speed sheet", write_link_phy_sheet, wb, all_interfaces)

    # Phase 30f: Architecture Review - NEW-V3.23.161. The senior-engineer design review (V3.23.160)
    # gets its workbook evidence twin. The review is computed ONCE here, from a snap-shaped view of
    # the sections already computed above (it reads exactly these keys), then the SAME object is
    # attached to the snapshot below -- one source of truth: sheet, snapshot JSON and the DOCX
    # report all carry identical verdicts. capacity_rows is hoisted for the same reason (the
    # snapshot reuses it instead of recomputing).
    capacity_rows = compute_capacity(all_device_physical)
    architecture_review = _run_phase(
        "Architecture review", compute_architecture_review,
        {**snapshot_state(all_interfaces, all_device_physical),
         "l3_forwarding": l3_forwarding, "routing_neighbors": all_routing_neighbors,
         "stp_roots": all_stp_roots, "fhrp": _l2["fhrp"], "capacity": capacity_rows,
         "lifecycle_risk": lifecycle_risk, "failure_impact": failure_impact,
         "security": all_security, "config_hygiene": all_config_hygiene,
         "operational_drift": _drift, "redistribution": all_redistribution,
         "collection_completeness": collection_completeness, "punchlist": punchlist},
        _default={"_unavailable": True})   # sentinel: a CRASH != a legit-empty review (cf. executive_brief)
    _run_phase("Architecture Review sheet", write_architecture_review_sheet, wb, architecture_review)

    # roadmap D3: zero-egress attestation — RE-DERIVE the read-only / no-egress / GET-only /
    # no-LLM trust claims NOW, from this installation's actual registries + sources (same
    # mechanics as tests/test_readonly_and_no_egress.py). Computed ONCE here for the 'Trust &
    # Sovereignty' sheet, then the SAME object is attached to the snapshot below (one source of
    # truth). A compute crash yields the _unavailable sentinel, which the sheet renders as
    # UNVERIFIED — a failed proof is disclosed, never defaulted to a pass (cf. architecture_review).
    attestation = _run_phase("Zero-egress attestation", compute_attestation,
                             _default={"_unavailable": True})
    _run_phase("Trust & Sovereignty sheet", write_attestation_sheet, wb, attestation)

    # Phase 30f-bis: NRFU COMMAND EXPORT (orchestration-roadmap frontier). The four-phase NRFU
    # certification pack: per-wave, per-device READ-ONLY verification commands with the EXPECTED values
    # pre-filled from the snapshot evidence (assertive, not informational; missing evidence stays the
    # explicit NOT-OBSERVED abstention). COMPLEMENTS the Phase 30d-quater validation plan (post-cutover
    # spot checks) with the full certification pack. Computed once from a snap-shaped view (the Phase 30f
    # pattern) -> sheet + snapshot (one source of truth); per-device .txt command files are available via
    # cisco_toolkit.nrfu_export.write_nrfu_pack (a --nrfu-pack CLI flag is deferred).
    nrfu_commands = _run_phase(
        "NRFU commands", compute_nrfu_commands,
        {**snapshot_state(all_interfaces, all_device_physical),
         "move_groups": move_groups, "stp_roots": all_stp_roots,
         "routing_neighbors": all_routing_neighbors, "fhrp_detail": all_fhrp_detail,
         "application_intelligence": application_intelligence},
        _default={})
    _run_phase("NRFU Commands sheet", write_nrfu_commands_sheet, wb, nrfu_commands)

    if getattr(args, "redact", False):    # final --redact net: scrub IP/MAC/secrets from computed + raw-config
        _run_phase("redact workbook cells", redact_workbook_cells, wb)   # sheets the dataclass pass can't reach
    wb.save(out_xlsx)
    logger.info(f"\n[OK] Saved: {out_xlsx}")
    logger.info(f"[OK] Log:   {LOG_FILE}")

    # Phase 21: Topology diagram files + state snapshot (NEW-V15). abspath() guarantees a
    # directory component so the diagram dir and snapshot path are always well-formed.
    diagram_dir = os.path.dirname(os.path.abspath(out_xlsx)) or "."
    try:
        write_topology_diagram(all_interfaces, diagram_dir)
    except Exception as e:
        logger.warning(f"  Topology diagram write failed: {e}")
    snap_path = os.path.splitext(os.path.abspath(out_xlsx))[0] + ".snapshot.json"
    snap_dict = snapshot_state(all_interfaces, all_device_physical)  # CHANGED-V3.17: capture for reuse (JSON + HTML)
    snap_dict["collected_at"] = collected_at                         # provenance: evidence collection instant (vs generated_at = regen time)
    snap_dict["physical_health"] = physical_health                   # NEW-V3.18
    snap_dict["l3_forwarding"] = l3_forwarding                       # NEW-V3.20
    snap_dict["cross_layer"] = cross_layer                           # NEW-V3.21
    snap_dict["protocol_health"] = protocol_health                   # NEW-V3.22
    snap_dict["protocol_intelligence"] = protocol_intelligence       # NEW-V3.23.100 (per-(switch,protocol,state) cause + remediation; reused from Phase 27b)
    snap_dict["service_map"] = service_map                           # NEW-V3.23.101 (L4 services from ACL ports + multicast activity; reused from Phase 27c)
    snap_dict["multicast_intelligence"] = multicast_intelligence     # NEW-V3.23.115 (media-fabric deep-dive; reused from Phase 27c-bis)
    snap_dict["lifecycle_risk"] = lifecycle_risk                     # NEW-V3.23.117 (hardware EoL band per device; reused from Phase 27c-ter)
    snap_dict["remediation_plan"] = remediation_plan                 # NEW-V3.23.116 (generated config snippets, review-only; reused from Phase 30d-ter)
    snap_dict["validation_plan"] = validation_plan                   # NEW-V3.23.143 (per-wave post-cutover validation checklist; reused from Phase 30d-quater)
    snap_dict["nrfu_commands"] = nrfu_commands                       # NEW (four-phase NRFU certification pack, READ-ONLY + expected-from-evidence; reused from Phase 30f-bis)
    snap_dict["golden_drift"] = golden_drift                         # NEW-V3.23.146 (per-device config drift vs baseline; reused from Phase 30d-quinquies)
    snap_dict["syslog_intelligence"] = syslog_intelligence           # NEW-V3.23.164 (NOS-style operational log analysis; reused from Phase 30d-sexies)
    snap_dict["qos_audit"] = qos_audit                               # NEW-V3.23.165 (configured QoS posture + doctrine findings; reused from Phase 30d-septies)
    snap_dict["software_risk"] = software_risk                       # NEW-V3.23.166 (advisory-surface screening + train lifecycle; reused from Phase 30d-octies)
    snap_dict["platform_health"] = platform_health                   # NEW-V3.23.167 (control-plane CPU/memory capacity screening; reused from Phase 30d-nonies)
    snap_dict["collection_completeness"] = collection_completeness   # NEW-V3.23.109 (pre-assessment blind-spot report; reused from Phase 27d)
    snap_dict["health_scores"] = health_scores                       # NEW-V3.23
    snap_dict["cable_map"] = cable_map                               # NEW: EDA-style physical cable map (explorer + webapp cable-map views)
    snap_dict["migration_readiness"] = migration_readiness           # NEW-V3.23
    snap_dict["score_sensitivity"] = score_sensitivity               # NEW-V3.23.5
    snap_dict["nat"] = all_nat                                       # NEW-V3.23.50 (NAT inventory: {host:{static,dynamic,pools,inside,outside}})
    snap_dict["security"] = all_security                            # NEW-V3.23.59 (CIS-aligned security posture: {host:{findings,summary}})
    snap_dict["framework_coverage"] = framework_cov                 # W2-3 (CIS/NIST/PCI/STIG mapping; computed once at the workbook stage above)
    snap_dict["detector_schema"] = detector_schema                   # J1 (per-detector descriptors; not-observed != healthy as a schema property; computed once at the workbook stage above)
    snap_dict["config_hygiene"] = all_config_hygiene                # NEW-V3.23.61 (undefined refs / unused structures: {host:{undefined,unused,summary}})
    snap_dict["stp_roots"] = all_stp_roots                          # NEW-V3.23.62 (per-VLAN STP root bridge: {host:{vlan:{root_priority,root_address,is_root}}})
    snap_dict["vpc"] = all_vpc                                       # NEW-V3.23.125 (vPC / MLAG status: {host:{domain_id,role,peer_status,vpcs}}) -> confirms MLAG peers in the flow simulator
    snap_dict["fhrp_detail"] = all_fhrp_detail                       # full HSRP election/preempt/tracking detail -> _d_fhrp_resilience (first non-[HISTORY-REDACTED] coverage)
    snap_dict["overlay"] = all_overlay                               # VXLAN-EVPN overlay (NVE peers) -> _d_nve_peer_health (engine's own target fabric, was blind)
    snap_dict["copp"] = all_copp                                     # CoPP drop counters -> _d_copp_drops (control-plane policing health)
    snap_dict["mpls"] = all_mpls                                     # SP/MPLS service-plane (LDP/VPNv4/L2VPN) -> _d_mpls_ldp_health/_d_mpls_l3vpn_health/_d_mpls_l2vpn_health
    snap_dict["aci"] = all_aci                                         # Cisco ACI (APIC JSON-ingestion channel) -> _d_aci_critical_faults/_d_aci_node_not_active/_d_aci_fabric_health_degraded
    snap_dict["sdwan"] = all_sdwan                                     # Cisco Catalyst SD-WAN (vManage JSON channel) -> _d_sdwan_control_connection_down/_d_sdwan_device_unreachable
    snap_dict["firewall"] = all_firewall                              # Cisco firewall (ASA/FTD failover) HA -> _d_firewall_ha_degraded (SSH show-text channel)
    snap_dict["ise"] = all_ise                                        # Cisco ISE (Identity Services Engine deployment) -> _d_ise_* (JSON controller-REST channel)
    snap_dict["fmc"] = all_fmc                                        # Cisco Secure Firewall Mgmt Center (FMC) -> _d_ftd_*/_d_fmc_* (JSON controller-REST channel)
    snap_dict["arista"] = all_arista                                  # MULTI-VENDOR: Arista EOS MLAG -> _d_arista_mlag_degraded (device-native JSON; the first non-Cisco vendor axis)
    snap_dict["juniper"] = all_juniper                               # MULTI-VENDOR: Juniper Junos SRX chassis-cluster -> _d_junos_chassis_cluster_degraded (the second non-Cisco vendor axis)
    snap_dict["cloud"] = all_cloud                                   # PUBLIC CLOUD: AWS security-group exposure -> _d_cloud_sg_open_ingress (the first cloud-domain axis)
    snap_dict["fortigate"] = all_fortigate                           # MULTI-VENDOR: Fortinet FortiGate HA cluster sync -> _d_fortigate_ha_degraded (the third non-Cisco vendor axis)
    snap_dict["mroute"] = all_mroute                                 # CISCO DEPTH: multicast RPF integrity -> _d_mcast_rpf_failure ((S,G) Null incoming-interface)
    snap_dict["lisp"] = all_lisp                                       # universal arch coverage
    snap_dict["cts"] = all_cts                                       # universal arch coverage
    snap_dict["dmvpn"] = all_dmvpn                                       # universal arch coverage
    snap_dict["crypto"] = all_crypto                                       # universal arch coverage
    snap_dict["bfd"] = all_bfd                                       # universal arch coverage
    snap_dict["ipv6_nd"] = all_ipv6_nd                                       # universal arch coverage
    snap_dict["ipv6_routing"] = all_ipv6_routing                                       # universal arch coverage
    snap_dict["pim"] = all_pim                                       # PIM-SM RP/neighbor -> _d_pim_rp_health (multicast resilience)
    snap_dict["ipv6_fhs"] = all_ipv6_fhs                             # IPv6 first-hop security -> _d_ipv6_fhs (rogue-RA gateway hijack)
    snap_dict["ntp"] = all_ntp                                       # NTP clock-sync STATE -> _d_ntp_sync (unsynchronized clock)
    snap_dict["port_security"] = all_port_security                   # port-security DETAIL -> _d_port_security_errdisable (Secure-shutdown)
    snap_dict["storm_control"] = all_storm_control                   # storm-control action -> _d_storm_control_action (toothless rule)
    snap_dict["qos_runtime"] = all_qos_runtime                       # QoS runtime egress drops -> _d_qos_runtime_drops
    snap_dict["shadow_infra"] = all_shadow_infra                     # undocumented infra neighbours -> _d_shadow_infra
    snap_dict["punchlist"] = punchlist                              # NEW-V3.23.63 (consolidated severity-ranked migration punch-list)
    snap_dict["operational_drift"] = _drift                         # NEW-V3.23.93 (false-health / operational-drift findings; also folded into the punch-list)
    snap_dict["calibration"] = calibration                           # NEW-V3.23.47 (fleet band-discrimination diagnostic)
    snap_dict["acls"] = all_acls                                     # NEW (L4 ACL sim): {host:{name:[rule,...]}}
    snap_dict["object_groups"] = all_object_groups                  # NEW (L4 depth): {host:{name:{kind,members}}}
    snap_dict["feature_compliance"] = feature_compliance             # roadmap I2 (golden-drift decomposed per policy feature)
    snap_dict["acl_line_reachability"] = acl_line_reachability       # roadmap G1 (offline ACL shadow / dead-line proof)
    if external_reconcile is not None:                               # roadmap B (opt-in: only when --import-inventory supplied)
        snap_dict["external_reconcile"] = external_reconcile         # declared source-of-truth vs collected evidence
    snap_dict["capture_integrity"] = capture_integrity              # roadmap K1 (per-capture truncation/pager/error guard)
    snap_dict["attestation"] = attestation                          # roadmap D3 (re-derived trust claims; 'Trust & Sovereignty' sheet twin — computed at the workbook stage above)
    snap_dict["parse_yield"] = parse_yield_report()                 # Plan A / Tier-1 #3: content-in/0-entities-out ledger (collected-but-unparsed ≠ feature-absent; K1's sibling)
    if whatif_result is not None:                                   # roadmap G4 (opt-in: only when --scenario supplied)
        snap_dict["whatif"] = whatif_result                         # failure-injection what-if results
    if path_intents is not None:                                    # roadmap G3 (opt-in: only when --path-intents supplied)
        snap_dict["path_intents"] = path_intents                    # named path/segmentation intent verdicts
    snap_dict["routes"] = all_routes                                # NEW (route-aware reachability): {host:[{prefix,source,next_hop,out_intf}]}
    snap_dict["routing_neighbors"] = all_routing_neighbors          # NEW (protocol-to-protocol analysis): {host:{ospf:[...],eigrp:[...],bgp:[...]}}
    snap_dict["redistribution"] = all_redistribution                # NEW (protocol-to-protocol analysis): {host:[{into_proto,into_id,from_proto,from_id,route_map,raw}]}
    # NEW-V3.23.81: embed the causality chains + failure-impact rollup the workbook computes, so the
    # explorer's Causality mode + cockpit keystones consume the SAME analysis instead of re-deriving in JS
    # (one source of truth). NEW-V3.23.91: REUSE the values already computed for the Phase 19/20 sheets
    # (causality_chains / failure_impact) instead of recomputing here -- this snapshot phase was the main
    # source of the duplicate heavy walks. Tuples -> dicts for a self-describing JSON contract. Local
    # import (link_centrality / wave_sequencing only) keeps those out of this module's public surface.
    from cisco_toolkit.analyze import compute_link_centrality, compute_wave_sequencing
    snap_dict["causality"] = [{"severity": s, "trigger": t, "mechanism": m, "impact": i, "mitigation": mt,
                               "hosts": list(hosts)}                            # NEW-V3.23.92: explicit host list (one source of truth; explorer highlights these directly)
                              for (s, t, m, i, mt, hosts) in (causality_chains or [])]
    snap_dict["failure_impact"] = failure_impact
    snap_dict["link_centrality"] = compute_link_centrality(all_interfaces)   # NEW-V3.23.88 (chokepoint links: betweenness + bridges; LINK twin of failure_impact)
    snap_dict["wave_sequencing"] = compute_wave_sequencing(all_interfaces, move_groups)   # NEW-V3.23.89 (per-group make-before-break vs hard cutover)
    # NEW-V3.23.84: embed the 4 cross-switch L2/L3 consistency checks the workbook computes, so the
    # explorer consumes the SAME analysis instead of re-deriving in JS (one source of truth). These
    # compute_* fns are already imported at module level (V3.23.64, also used to build the punch-list
    # above), so no local re-import is needed — and adding one would shadow that earlier use.
    snap_dict["addressing_conflicts"] = compute_addressing_conflicts(all_interfaces)
    snap_dict["fhrp"] = compute_fhrp_consistency(all_interfaces)
    snap_dict["trunk_native"] = compute_trunk_native_mismatches(all_interfaces)
    snap_dict["link_phy"] = compute_duplex_speed_mismatches(all_interfaces)
    snap_dict["move_groups"] = move_groups                           # NEW-V3.23.86 (Migration Waves mode: the move-group / shared-VLAN-domain structure the readiness verdicts attach to; already computed above for migration_readiness)
    snap_dict["capacity"] = capacity_rows                            # NEW-V3.23.87 (port + PoE headroom); V3.23.161: reuses the rows computed for the architecture review (Phase 30f)
    snap_dict["endpoint_identity"] = endpoint_identity               # NEW-V3.23.95 (per-endpoint vendor + inferred class; reused from the Phase 15b compute)
    snap_dict["endpoint_dependencies"] = endpoint_dependencies       # NEW-V3.23.96 (clusters / dual-homed / VLAN tiers / per-switch validation; reused from Phase 29b)
    snap_dict["subnet_intelligence"] = subnet_intelligence           # NEW-V3.23.97 (per-device subnet source/destination + move-group reachability; reused from Phase 29c)
    snap_dict["migration_scenarios"] = migration_scenarios           # NEW-V3.23.98 (per-group cutover scenario recommendation; reused from Phase 29d)
    snap_dict["application_intelligence"] = application_intelligence  # NEW-V3.23.112 (application-domain synthesis + migration risk; reused from Phase 30d-bis)
    snap_dict["segmentation"] = segmentation                         # NEW-V3.23.118 (L3 isolation posture; reused from Phase 30d-bis2)
    snap_dict["vlan_cutover"] = vlan_cutover                         # MASTER_PLAN §4.3 (per-VLAN cutover matrix; reused from Phase 30d-bis3)
    snap_dict["executive_brief"] = executive_brief                   # NEW-V3.23.120 (cross-axis migration brief; reused from Phase 30e)
    # Single source of truth: publish the canonical VLAN count (the same `vlan_inventory` derivation the
    # design / CRD deliverables read) into the brief's scale block, so every consumer — workbook, explorer,
    # webapp — reads ONE value instead of re-counting VLANs in JS/TS (which drifts: explorer ~176 vs 172).
    _ebscale = snap_dict["executive_brief"].get("scale") if isinstance(snap_dict.get("executive_brief"), dict) else None
    if isinstance(_ebscale, dict):
        _ebscale["n_vlans"] = len(vlan_inventory(snap_dict))
        # publish the genuinely-COLLECTED subset alongside scale so coverage-honesty surfaces (deck /
        # engagement / ops / workbook) read ONE canonical "253 of 303 collected" instead of re-deriving it
        # or mislabelling the 303 inventory as "assessed / evidence scope". (SSOT-device-deliv-06)
        _ebscale["n_collected"] = ((snap_dict.get("collection_completeness") or {}).get("summary") or {}).get("complete")
    # Integrity (wave R2-4-01): if the cross-axis brief synthesis FAILED, _run_phase returned the sentinel
    # default and scale/posture/axes are absent. Disclose it as a machine-readable flag so consumers
    # (deck / explorer / webapp) surface "cross-axis synthesis unavailable" instead of rendering the
    # missing numbers as healthy zeros -- the assembly-level analogue of the device false-health class.
    if isinstance(executive_brief, dict) and executive_brief.get("_unavailable"):
        logger.warning("  [INTEGRITY] executive-brief synthesis FAILED; stamping assessment_integrity so "
                       "consumers disclose 'cross-axis synthesis unavailable' rather than zeros.")
        snap_dict["assessment_integrity"] = {"executive_brief": "compute_failed"}
    if isinstance(architecture_review, dict) and architecture_review.get("_unavailable"):
        logger.warning("  [INTEGRITY] architecture-review computation FAILED; stamping assessment_integrity so "
                       "consumers disclose 'architecture review unavailable' rather than a clean grade.")
        snap_dict.setdefault("assessment_integrity", {})["architecture_review"] = "compute_failed"
    # ... and the same disclosure for EVERY other fail-soft phase (review 2026-07-28). The two stamps
    # above cover 2 of the 41 fail-soft sections main() publishes into the snapshot; the other 39 fall
    # back to `[]` / `{}` on a compute crash, which on disk is BYTE-IDENTICAL to "computed fine, found
    # nothing". So a crashed `cross_layer` publishes no cross-layer criticals (and migration-readiness'
    # cross-layer check then passes, grading a move group READY off evidence that never ran), a crashed
    # `punchlist` publishes a clean punch-list, a crashed `operational_drift` publishes no drift, a
    # crashed `software_risk` publishes no advisory exposure. That is "not observed" silently becoming
    # "healthy" — the false-health class guardrail 3 forbids, at assembly level rather than device level.
    # The `_unavailable` sentinel can't be pushed down into the other 39 defaults without changing the
    # shape every consumer indexes, so disclose ALONGSIDE instead: _run_phase already records ok=False
    # per phase, so publish that list and let consumers tell a crash from a clean result.
    # Golden-neutral by construction: a clean run has no failed phase, so the key is never written.
    _failed_phases = list(dict.fromkeys(p["phase"] for p in _PHASE_TIMINGS if not p.get("ok", True)))
    if _failed_phases:
        logger.warning(f"  [INTEGRITY] {len(_failed_phases)} analysis phase(s) FAILED and fell back to an "
                       f"empty result that is indistinguishable from a clean one; stamping "
                       f"assessment_integrity.failed_phases so consumers do not read them as healthy: "
                       f"{_failed_phases}")
        snap_dict.setdefault("assessment_integrity", {})["failed_phases"] = _failed_phases
    snap_dict["device_dossiers"] = device_dossiers                   # NEW-V3.23.172 (per-asset compound-risk register; reused from the Phase 30d synthesis)
    # NEW-V3.23.160: the senior-engineer design review. V3.23.161: REUSES the object computed in
    # Phase 30f for the workbook sheet (one source of truth — sheet, snapshot and DOCX agree),
    # instead of recomputing from the assembled snap_dict (the Phase 30f view carries exactly the
    # keys the review reads, so the result is identical).
    snap_dict["architecture_review"] = architecture_review
    if flow_trace is not None:                                       # NEW-V3.19
        snap_dict["flow_trace"] = flow_trace
    if args.redact:                                                  # NEW-V3.23.41
        snap_dict = redact_snapshot(snap_dict)
        logger.info("  [redact] snapshot IPs / MACs / serials pseudonymized (--redact)")
    # NEW: the canonical CCDE-grounded target-state DESIGN BLUEPRINT (the senior-network-design-engineer
    # layer). Computed LAST, from the fully-assembled (and, under --redact, already-redacted) snap_dict, so
    # the design DOCX / explorer / webapp all read ONE blueprint instead of re-deriving design intent. It
    # folds the date-relative lifecycle/EoL evidence, so it is excluded from the frozen golden like
    # executive_brief / device_dossiers; the SSOT publish is locked by tests/test_pipeline_inprocess.py.
    try:
        from cisco_toolkit.design_advisor import load_requirements
        _req = load_requirements(getattr(args, "requirements", "") or "")
        if _req:
            snap_dict["requirements_register"] = _req                 # stored so the blueprint stays reproducible from the snapshot alone
            logger.info(f"  [design] requirements register supplied ({', '.join(sorted(_req))}) -> blueprint right-sized")
        # design_blueprint INCLUDES the gated EVPN-migration guardrails (compute_design_blueprint folds them in
        # so the published blueprint stays reproducible from the snapshot + the same requirements register).
        snap_dict["design_blueprint"] = compute_design_blueprint(snap_dict, _req or None)
        # design-driven NRFU/ATP checklist, published canonically so the OFFLINE explorer and the webapp
        # read ONE set of phased acceptance items (and their phases) instead of re-deriving them in JS.
        # Derived purely from design_blueprint -> excluded from the golden + publish-locked alongside it.
        snap_dict["design_nrfu"] = compute_design_nrfu(snap_dict["design_blueprint"])
        # Architecture-coverage SSOT: which architecture CLASSES were observed vs not, and what fired -- one
        # source the deliverables/explorer/webapp read instead of re-deriving coverage. Reads design_blueprint
        # + the per-axis snapshot keys; excluded from the golden (date-relative via the blueprint) like it.
        snap_dict["architecture_coverage"] = compute_architecture_coverage(snap_dict)
        # Coverage matrix (Plan-A #5): compose the four coverage sources -- collection / capture / parse /
        # architecture -- into ONE per-(device, axis) first-class table (recomputes no device state, just
        # projects the published verdicts). Reads architecture_coverage above, so it inherits the
        # blueprint's date-relativity -> excluded from the golden + publish-locked alongside it.
        snap_dict["coverage_matrix"] = compute_coverage_matrix(snap_dict)
    except Exception as e:                                            # fail-soft: never break the snapshot write
        logger.warning(f"  design_blueprint compute failed (non-fatal): {e}")
    # SSOT self-check (the field-data safety net): the suite only proves SSOT-consistency on its own
    # fixtures, never on a real customer snapshot. Reconcile the published canonical facts against
    # their raw-evidence derivation NOW, on the fully-assembled snap_dict, and ONLY on drift disclose
    # it via the existing assessment_integrity channel (which the deck/explorer already surface) +
    # an [INTEGRITY] warning. A clean run stamps nothing -> no false alarm, no golden churn. (ssot.audit)
    try:
        from cisco_toolkit import ssot as _ssot
        _ssot_drift = _ssot.audit(snap_dict)
        # Positive self-verification badge for the dashboards (golden-excluded with executive_brief):
        # an always-present client-facing trust signal that the published headline facts were
        # reconciled to the raw evidence, distinct from the drift-only assessment_integrity channel.
        if isinstance(snap_dict.get("executive_brief"), dict):
            snap_dict["executive_brief"]["ssot"] = _ssot.summary(snap_dict)
        if _ssot_drift:
            logger.warning(f"  [INTEGRITY] {_ssot_drift['n_violations']} published fact(s) do not reconcile "
                           f"to the raw evidence; disclosing in assessment_integrity: {_ssot_drift['violations'][:3]}")
            snap_dict.setdefault("assessment_integrity", {}).update(_ssot_drift)
    except Exception as e:                                            # fail-soft: a self-check must never break the write
        logger.warning(f"  SSOT self-check skipped (non-fatal): {e}")
    # J3 + J2: coverage-schema census + fact-lineage — the coverage-honesty doctrine made QUERYABLE.
    # Both are computed HERE, from the fully-assembled snap_dict (so the census sees every section and
    # the lineage reads the final canonical blocks), reusing the SSOT contract (abstention_reason /
    # CANONICAL_FACTS) rather than a parallel one. schema_census is presence/absence + structural shape
    # only -> non-volatile -> golden-stable; fact_lineage embeds the date-relative headline VALUES, so
    # it is golden-EXCLUDED (stripped like executive_brief/lifecycle_risk) and pinned in its own test.
    try:
        from cisco_toolkit import ssot as _ssot2
        snap_dict["schema_census"] = _ssot2.compute_schema_census(snap_dict)   # per-section coverage map (SuzieQ `describe` analog)
        snap_dict["fact_lineage"] = _ssot2.compute_fact_lineage(snap_dict)      # attribute-level provenance for the canonical facts
        # The census self-describes the FINAL snapshot; the two keys it adds above (schema_census /
        # fact_lineage) are themselves absent from the map it computed, so re-census once so the
        # published census is complete + self-consistent with the JSON it ships in (one source of truth).
        snap_dict["schema_census"] = _ssot2.compute_schema_census(snap_dict)
        # Coverage Schema sheet: the workbook twin of the census. The workbook was saved above (the
        # census needs the fully-assembled snap_dict, which exists only now), so add this last sheet
        # and re-save. It carries only section names/states/counts/notes — no IP/MAC/serial — so it is
        # share-safe under --redact without a further scrub. Fail-soft: never break the snapshot write.
        write_coverage_schema_sheet(wb, snap_dict["schema_census"])
        wb.save(out_xlsx)
    except Exception as e:                                            # fail-soft: census/lineage must never break the write
        logger.warning(f"  schema-census / fact-lineage skipped (non-fatal): {e}")
    # roadmap A1/H2: opt-in state-assertion check-pack (--assert-pack FILE), evaluated over the FULL snapshot
    # (coverage-honest: a subject that was not collected -> [NOT OBSERVED], excluded from the pass/fail denominator).
    if args.assert_pack:
        try:
            with open(args.assert_pack, encoding="utf-8") as _af:
                _pack = json.load(_af)
            snap_dict["state_assertions"] = evaluate_pack(snap_dict, _pack if isinstance(_pack, dict) else {"assertions": _pack})
        except Exception as e:
            logger.warning(f"  --assert-pack not evaluated ({e}); skipping.")
    try:
        write_json_file(snap_path, sparsify_interfaces(snap_dict), compact=True)   # Tier3#14 minified + #14-Phase2 sparse interfaces on disk
        logger.info(f"[OK] Snapshot: {snap_path}  (use --compare OLD NEW for pre/post diff)")
    except Exception as e:
        logger.warning(f"  Snapshot write failed: {e}")

    # Phase 22: HTML Explorer (self-contained) - NEW-V3.17. Embeds the same snap_dict the
    # snapshot.json carries directly into a copy of blast_radius_explorer.html, so one run
    # yields both the workbook and a ready-to-open, air-gapped topology explorer. The JSON
    # write (above) and this HTML write are the last emits, so any future per-version
    # enrichment of snap_dict lands in both outputs.
    if not args.no_html:
        html_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_explorer.html"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_html_explorer(html_out, snap_dict, label)
        except Exception as e:
            logger.warning(f"  HTML Explorer write failed: {e}")

    # Phase 31: Assessment & Migration Runbook (DOCX) - NEW-V3.23.93. The narrative sign-off twin of
    # the workbook, built from the SAME snap_dict (one source of truth -> every number agrees), to the
    # cisco-network-assessment 12-section, evidence-disciplined standard. Optional (python-docx); a
    # missing library is a warning, never a crash (the workbook / explorer / JSON already saved).
    if not args.no_docx:
        docx_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_runbook.docx"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_runbook_docx(docx_out, snap_dict, label, flow_paths=flow_paths)
        except Exception as e:
            logger.warning(f"  Runbook (DOCX) write failed: {e}")

    # NEW-V3.23.144: Executive presentation deck (PPTX) - the stakeholder twin of the workbook, rendered
    # from the same snapshot. Optional python-pptx; a missing library is a warning, never a crash.
    if not args.no_pptx:
        pptx_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_executive_deck.pptx"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_executive_deck_pptx(pptx_out, snap_dict, label)
        except Exception as e:
            logger.warning(f"  Executive deck (PPTX) write failed: {e}")

    # Phase 33: As-Built Network Design Document (HLD/LLD, DOCX) - NEW-V3.23.148. The Design-phase twin
    # of the workbook: reconstructs the current (as-built) design + target-state recommendations from the
    # SAME snapshot (one source of truth). Optional python-docx; a missing library is a warning, never a crash.
    # P0-3/DEC-003: design follows an APPROVED assessment (PPDIOO gate). With a gate-state store
    # present (docs/engagement-state.json) and the marker unapproved, this REFUSES (skips the
    # write, loudly); --override-gate proceeds and appends a who/when/why audit line. No store at
    # all = warn-and-proceed (brownfield).
    # The verdicts are COLLECTED (not just tested) so the run can answer "was anything withheld?"
    # at the end without re-deriving the decision from the flags -- see cisco_toolkit.gate_state
    # .GateVerdict. Each gate is still evaluated ONLY when its deliverable was actually requested:
    # --no-design short-circuits before gate_enforce, so a suppressed document neither consults the
    # ledger nor writes a refusal row about a document nobody asked for.
    gate_verdicts = []
    if not args.no_design:
        design_gate = gate_enforce("design", override_reason=args.override_gate,
                                   root=args.gate_root, engagement=args.engagement)
        gate_verdicts.append(design_gate)
        if design_gate:
            design_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_design.docx"
            label = os.path.splitext(os.path.basename(out_xlsx))[0]
            try:
                write_design_doc_docx(design_out, snap_dict, label)
            except Exception as e:
                logger.warning(f"  Design document (DOCX) write failed: {e}")

    # Phase 34: per-wave Method of Procedure (MOP, DOCX) - NEW-V3.23.149. The Implement-phase cutover
    # template, one section per migration wave, reusing the validation plan as the post-cutover checks.
    # Optional python-docx; a missing library is a warning, never a crash.
    # P0-3/DEC-003: the MOP follows an APPROVED LLD + a captured current-state baseline (PPDIOO
    # gate; mop-change-author charter). Same refusal/override/brownfield semantics as the design
    # gate above — the refusal skips ONLY this deliverable; the workbook/snapshot already saved.
    if not args.no_mop:
        mop_gate = gate_enforce("mop", override_reason=args.override_gate,
                                root=args.gate_root, engagement=args.engagement)
        gate_verdicts.append(mop_gate)
        if mop_gate:
            mop_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_mop.docx"
            label = os.path.splitext(os.path.basename(out_xlsx))[0]
            try:
                write_mop_docx(mop_out, snap_dict, label)
            except Exception as e:
                logger.warning(f"  MOP (DOCX) write failed: {e}")

    # Phase 35: Customer Requirements Document (CRD, DOCX) - NEW-V3.23.156. The Plan-phase
    # requirements-capture instrument: evidence-primed current-environment summary + REQ-ID capture
    # tables + traceability skeleton. Optional python-docx; a missing library is a warning, never a crash.
    if not args.no_crd:
        crd_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_crd.docx"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_crd_docx(crd_out, snap_dict, label)
        except Exception as e:
            logger.warning(f"  CRD (DOCX) write failed: {e}")

    # Phase 36: Engagement Workflow & Plan of Record (DOCX) - NEW-V3.23.157. The engagement-management
    # layer over the whole document set: evidence-led verdict, phase tracker with gates, next-action
    # queue, T-minus wave calendar, and a RAID log seeded from the assessment's own findings.
    # Optional python-docx; a missing library is a warning, never a crash.
    if not args.no_engagement:
        eng_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_engagement.docx"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_engagement_docx(eng_out, snap_dict, label)
        except Exception as e:
            logger.warning(f"  Engagement workflow (DOCX) write failed: {e}")

    # Phase 37: Architecture Review & Conformance Report (DOCX) - NEW-V3.23.160. The senior-engineer
    # design review: 8-domain leading-practice conformance scorecard + availability analysis rendered
    # from the architecture_review section attached to the snapshot above. Optional python-docx; a
    # missing library is a warning, never a crash.
    if not args.no_archreview:
        ar_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_archreview.docx"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_archreview_docx(ar_out, snap_dict, label)
        except Exception as e:
            logger.warning(f"  Architecture review (DOCX) write failed: {e}")

    # Phase 38: Operations Handbook (DOCX) - NEW-V3.23.168. The PPDIOO Operate-phase deliverable:
    # the Day-2 handbook whose monitoring/capacity baselines, drift standard and software-governance
    # cadence are derived from THIS fleet's assessed evidence (syslog/platform-health/golden-drift/
    # software-risk axes), with evidence-gated sections for older snapshots. Optional python-docx;
    # a missing library is a warning, never a crash.
    if not args.no_opshandbook:
        oh_out = os.path.splitext(os.path.abspath(out_xlsx))[0] + "_ops_handbook.docx"
        label = os.path.splitext(os.path.basename(out_xlsx))[0]
        try:
            write_ops_handbook_docx(oh_out, snap_dict, label)
        except Exception as e:
            logger.warning(f"  Operations handbook (DOCX) write failed: {e}")

    # Pipeline stage 5 (Plan-A #15 strangler): the leaf finalize phases. Reuse the ONE analyze-stage
    # carrier (_actx, unconditionally built above) instead of a second fresh AnalysisContext -- the
    # strangler converging on a single typed carrier threaded through the pipeline. Set the
    # finalize-only handles it reads (config + the assembled snapshot); golden-neutral, same values.
    _actx.args = args
    _actx.out_xlsx = out_xlsx
    _actx.root_dir = root_dir
    _actx.all_devices_meta = all_devices_meta
    _actx.workers = workers
    _actx.snap_dict = snap_dict
    _stage_finalize(_actx)

    # Exit disposition. A gate refusal is a CORRECT run that withheld a document, so the default
    # stays 0 -- every wrapper here reads non-zero as "the run failed" (webapp/backend/ingest.py
    # raises, and serve.py renders that to a field engineer as "redaction FAILED ... Treat anything
    # already written as UNREDACTED"), and a false alarm on a correct run is its own safety defect.
    # --fail-on-gate-refusal opts a pipeline into stopping instead. 2, not 1, matches this repo's
    # refusal code (holdout.py, scorecard.py); 1 stays "the run itself broke".
    refused = [v for v in gate_verdicts if v.refused]
    if refused and args.fail_on_gate_refusal:
        # Report what the ledger ACTUALLY holds, by reading each verdict's `recorded` flag rather
        # than assuming the write happened. Five of the six refusal statuses record nothing (a
        # mis-set --gate-root, an unreadable ledger, an ownership mismatch), and claiming otherwise
        # would send the operator to `gate_state show` to look for rows that were never written --
        # a false statement about an audit trail, emitted by the audit feature itself.
        written = [v for v in refused if v.recorded]
        unwritten = [v for v in refused if not v.recorded]
        detail = f"recorded in the gate ledger: {', '.join(v.generator for v in written)}" \
            if written else "nothing was recorded in a ledger"
        if unwritten:
            detail += ("; NOT recorded (%s) -- see the [GATE REFUSED] lines above, which are the "
                       "only trace" % ", ".join(f"{v.generator}: {v.status}" for v in unwritten))
        logger.error("[GATE] exiting 2 (--fail-on-gate-refusal): withheld %s. %s "
                     "(python -m cisco_toolkit.gate_state show)",
                     ", ".join(v.generator for v in refused), detail)
        return 2
    return 0


def _stage_finalize(ctx: "AnalysisContext") -> None:
    """Pipeline stage 5 (Plan-A #15 strangler): the leaf finalize phases -- run-manifest
    chain-of-custody (39), opt-in raw-capture scrub (40), perf-timings sidecar (41), closing gate
    summary (42 -- LAST, so a withheld deliverable is the final thing the operator sees). A true LEAF:
    nothing downstream reads its outputs, which is why it is the safest first extraction. Reads the
    context's out_xlsx / snap_dict / root_dir / args / all_devices_meta / workers + the module-global
    _PHASE_TIMINGS. The body was byte-identical to the inline block it replaced until the gate-audit
    change, which reworked the manifest write's error arm and appended the closing gate summary."""
    out_xlsx, snap_dict, root_dir, args = ctx.out_xlsx, ctx.snap_dict, ctx.root_dir, ctx.args
    all_devices_meta, workers = ctx.all_devices_meta, ctx.workers

    # Phase 39: sealed run-manifest chain-of-custody (roadmap D2 + J4). The offline, deterministic answer to a
    # GAIT-style audit trail: a hash-chained ledger of the pipeline stages + per-artifact sha256 + the
    # coverage-honest abstention ledger, sealed by chain_root. LAST emit so it hashes every produced artifact.
    # Observed, never re-derived: the closing gate summary below MUST NOT claim the seal from the
    # manifest file merely EXISTING. write_json_file leaves the PREVIOUS manifest intact when it
    # fails, so a re-run to the same --output whose seal failed leaves an earlier run's manifest on
    # disk -- and an os.path.isfile() check then reported "Sealed" while pointing the operator at a
    # manifest that can carry the OPPOSITE verdict. Reproduced end-to-end: an approved run followed
    # by a revoked run whose manifest write failed ended on "Sealed in the run manifest's 'gate'
    # step" three lines below "manifest write FAILED", with proceed=true still on disk for both
    # deliverables this run withheld. This is the repo's "existence != delivery" class.
    _manifest_sealed = False
    try:
        _run_manifest = build_run_manifest(out_xlsx, snap_dict)
        _manifest_path = os.path.splitext(os.path.abspath(out_xlsx))[0] + ".run_manifest.json"
        write_json_file(_manifest_path, _run_manifest)
        _manifest_sealed = True
        logger.info(f"[OK] Run manifest (chain-of-custody): {_manifest_path}  "
                    f"(chain_root {str(_run_manifest.get('chain_root', ''))[:12]}…, "
                    f"{len(_run_manifest.get('artifacts') or [])} artifact(s))")
    except Exception as e:
        # A lost manifest is normally cosmetic; when this run reached a gate verdict it is evidence,
        # so escalate rather than warn. What it is NOT is the only record: a refusal over missing
        # approvals appends a durable `refuse` row to the gate ledger (gate_state's audit array),
        # which is the cross-run owner the manifest only caches. Saying "the only durable record is
        # now lost" while that row sits in the ledger this same run wrote is a false alarm -- it
        # would send an auditor hunting for evidence that is exactly where it should be.
        from cisco_toolkit import gate_state as _gs
        _v = _gs.verdicts()
        if _v:
            _in_ledger = [v for v in _v if v.get("recorded")]
            _only_here = [v for v in _v if not v.get("recorded")]
            _where = ("; the gate ledger still holds the durable row(s) for %s"
                      % ", ".join(v["generator"] for v in _in_ledger) if _in_ledger else "")
            _lost = ("; NOTHING durable was recorded for %s -- the [GATE] log lines above are the "
                     "only trace" % ", ".join(v["generator"] for v in _only_here)) if _only_here else ""
            logger.error(f"  Run manifest write FAILED and this run has gate verdicts "
                         f"{[v['verdict'] for v in _v]} -- the tamper-evident per-run seal of them "
                         f"is lost{_where}{_lost} (non-fatal to the run): {e}")
        else:
            logger.warning(f"  Run manifest write failed (non-fatal): {e}")

    # Phase 40: opt-in raw-capture secret scrub (Plan A / Tier-1 #5). Deliberately the VERY
    # last step: every deliverable above was built from the ORIGINAL captures, and the sealed
    # manifest is already written. Secret VALUES only — the dir stays analyzable/--compare-able.
    if getattr(args, "redact_collection", False):
        try:
            _scanned, _changed = redact_collection_dir(root_dir)
            logger.info(f"[OK] redact-collection: scrubbed secret values in {_changed} of "
                        f"{_scanned} raw capture file(s) under {root_dir} (in place, idempotent; "
                        f"IPs/hostnames kept)")
        except Exception as e:
            logger.warning(f"  redact-collection failed (non-fatal; raw dir unchanged): {e}")

    # Phase 41: perf sidecar (Plan A / Tier-2 #11 — the measure-first harness). A SEPARATE
    # .phase_timings.json next to the workbook, chronological, golden-neutral (never a
    # snapshot key; written after the manifest seal like the Phase-40 scrub). Sweep runner:
    # `python tests/perf_scale.py 100 300 600 1000`.
    try:
        _pt_path = os.path.splitext(os.path.abspath(out_xlsx))[0] + ".phase_timings.json"
        _pt = {"n_devices": len(all_devices_meta), "workers": workers,
               "total_seconds": round(sum(p["seconds"] for p in _PHASE_TIMINGS), 3),
               "phases": list(_PHASE_TIMINGS)}
        write_json_file(_pt_path, _pt)
        _slowest = max(_PHASE_TIMINGS, key=lambda p: p["seconds"], default={"phase": "-", "seconds": 0})
        logger.info(f"[OK] Phase timings: {_pt_path}  ({len(_PHASE_TIMINGS)} timed phase(s), "
                    f"slowest: {_slowest['phase']} {_slowest['seconds']}s)")
    except Exception as e:
        logger.warning(f"  phase-timings write failed (non-fatal): {e}")

    # Phase 42: closing gate summary. Every phase above ends in "[OK] ..." whether or not a gate
    # refused, so a run that produced NEITHER the design nor the MOP still signed off with five
    # consecutive [OK] lines -- "not observed" reading as healthy, on the last thing the operator
    # sees. The per-verdict ERROR is emitted thousands of lines earlier (a 303-device run prints
    # ~4,000). Restate the refusals LAST, at ERROR, naming the deliverables that were withheld.
    # Silent when nothing was refused: a summary that always prints is one nobody reads.
    try:
        from cisco_toolkit import gate_state as _gs
        _refused = [v for v in _gs.verdicts() if v.get("proceed") is False]
        if _refused:
            # THIS run's seal outcome, threaded from the write itself -- never os.path.isfile(). A
            # stale manifest from an earlier run to the same --output satisfies an existence check
            # while carrying the OPPOSITE verdict, which made the operator's LAST line contradict
            # the failure logged three lines above it. See the comment at the manifest write.
            _sealed = _manifest_sealed
            _stale = (not _sealed) and os.path.isfile(
                os.path.splitext(os.path.abspath(out_xlsx))[0] + ".run_manifest.json")
            # Remediation is per-STATUS, not one-size-fits-all, and not collapsed to the WORST
            # status either: a run can withhold the design over a missing approval (fixable by
            # `approve`) and the MOP over an unreadable ledger (fixable only by repairing the file).
            # Advice that names one and not the other is unactionable for whichever it dropped, on
            # the last line the operator reads.
            _fix_for = {
                "pending": "record the approval ('python -m cisco_toolkit.gate_state approve "
                           "<gate>') or re-run with --override-gate \"<reason>\" (audited)",
                "bad_root": "fix --gate-root -- it does not point at an existing directory, so no "
                            "ledger was located (NOT overridable: an override is consent to skip a "
                            "KNOWN gate)",
                "unreadable": "repair or remove the gate ledger -- it exists but could not be "
                              "parsed, and it is deliberately never rewritten so the corruption "
                              "stays as evidence (NOT overridable)",
                "ownership_mismatch": "this ledger governs a DIFFERENT engagement -- point "
                                      "--gate-root at the right one, or correct --engagement (NOT "
                                      "overridable: no gate for this engagement was located)",
                "ownership_unbound": "bind the ledger to this engagement once ('python -m "
                                     "cisco_toolkit.gate_state bind <ID>') so the declaration can "
                                     "be verified (NOT overridable)",
            }
            _statuses = sorted({str(v.get("verdict")) for v in _refused})
            _how = " ".join(f"[{s}] {_fix_for.get(s, 'no remediation is known for this status.')}."
                            for s in _statuses)
            if _sealed:
                _seal_note = "Sealed in the run manifest's 'gate' step."
            elif _stale:
                # The dangerous case: a file IS there, so an operator who checks will find one.
                _seal_note = ("NOT SEALED -- THIS run's manifest write failed. A manifest from an "
                              "EARLIER run is still on disk at that path and does NOT describe this "
                              "run; do not read it as this run's record.")
            else:
                _seal_note = ("NOT SEALED -- the run manifest was not written, so these log lines "
                              "and the gate ledger's audit rows are the only record.")
            # Name where the durable record actually is, per deliverable: `recorded` is observed at
            # the moment of the write, so this cannot claim a ledger row that was never appended.
            _rec = [v["generator"] for v in _refused if v.get("recorded")]
            _unrec = [v["generator"] for v in _refused if not v.get("recorded")]
            _durable = (("Durable ledger row(s) written for %s. " % ", ".join(_rec)) if _rec else "")
            if _unrec:
                _durable += ("NO durable ledger row for %s (see the [GATE REFUSED] line(s) above "
                             "for why). " % ", ".join(_unrec))
            logger.error("[GATE] %d deliverable(s) WITHHELD by the PPDIOO document gates: %s -- %s %s%s",
                         len(_refused),
                         "; ".join(f"{v['generator']} (missing: "
                                   f"{', '.join(v.get('missing') or []) or 'not evaluated'})"
                                   for v in _refused),
                         _how,
                         _durable,
                         _seal_note)
    except Exception as e:
        logger.warning(f"  gate summary failed (non-fatal): {e}")


# =============================================================================
# STAGE-3 (analyze) CTX-ADAPTERS  (Plan-A #15 strangler -- the positional collapse)
# =============================================================================
# The three WIDE analyze-stage syntheses used to thread 15/19/14 locals positionally through
# _run_phase -- the fragile call-shape a silent reorder could corrupt (the format-fidelity /
# positional-drift bug class). Each adapter reads the typed AnalysisContext and forwards to the
# UNCHANGED public compute_* by KEYWORD (named -> reorder-proof; it also documents the
# main-local -> parameter renames such as all_security -> security). The public functions keep
# their explicit-arg signatures, so the webapp / excel / the ~40 direct-call tests are untouched;
# ONLY main() reads from ctx. Module-private (leading underscore) so test_package's surface
# contract is unchanged. Golden-neutral: the same values reach compute_* in the same order.
def _device_dossiers(ctx: "AnalysisContext") -> dict:
    """Ctx-adapter for the Device Risk Register (was 15 positional args at 'Device risk register')."""
    return compute_device_dossiers(
        health_scores=ctx.health_scores, failure_impact=ctx.failure_impact,
        lifecycle_risk=ctx.lifecycle_risk, software_risk=ctx.software_risk,
        platform_health=ctx.platform_health, syslog_intelligence=ctx.syslog_intelligence,
        qos_audit=ctx.qos_audit, golden_drift=ctx.golden_drift,
        security=ctx.all_security, config_hygiene=ctx.all_config_hygiene,
        stp_roots=ctx.all_stp_roots, vpc=ctx.all_vpc,
        physical_health=ctx.physical_health, protocol_health=ctx.protocol_health,
        move_groups=ctx.move_groups)


def _punchlist(ctx: "AnalysisContext") -> list:
    """Ctx-adapter for the Migration Punch-List (was 9 positional + 10 keyword args)."""
    return compute_migration_punchlist(
        cross_layer=ctx.cross_layer, security=ctx.all_security, config_hygiene=ctx.all_config_hygiene,
        physical_health=ctx.physical_health, l3_forwarding=ctx.l3_forwarding,
        protocol_health=ctx.protocol_health, stp_findings=ctx.stp_findings,
        health_scores=ctx.health_scores, move_groups=ctx.move_groups,
        l2=ctx.l2, hostname_mismatches=ctx.hostname_mismatches, drift=ctx.drift,
        ptp_readiness=ctx.ptp_readiness, media_risks=ctx.media_risks,
        syslog_intelligence=ctx.syslog_intelligence, qos_audit=ctx.qos_audit,
        software_risk=ctx.software_risk, platform_health=ctx.platform_health,
        device_dossiers=ctx.device_dossiers)


def _executive_brief(ctx: "AnalysisContext") -> dict:
    """Ctx-adapter for the cross-axis Executive Brief (was 8 positional + 6 keyword args)."""
    return compute_executive_brief(
        health_scores=ctx.health_scores, punchlist=ctx.punchlist,
        migration_readiness=ctx.migration_readiness,
        application_intelligence=ctx.application_intelligence, lifecycle_risk=ctx.lifecycle_risk,
        segmentation=ctx.segmentation, multicast_intelligence=ctx.multicast_intelligence,
        remediation_plan=ctx.remediation_plan, syslog_intelligence=ctx.syslog_intelligence,
        qos_audit=ctx.qos_audit, software_risk=ctx.software_risk, platform_health=ctx.platform_health,
        device_dossiers=ctx.device_dossiers, endpoint_identity=ctx.endpoint_identity)


if __name__ == "__main__":
    # sys.exit(), not a bare main(): this door SWALLOWED main()'s return code, so `python
    # COLLECT_PARSE_V3_23_0.py ...` exited 0 no matter what it returned, while the `cisco-assess`
    # console script (which wraps main() in sys.exit) and Atlas's --run-engine dispatch (int(rc or
    # 0)) both honoured it. The repo's own e2e exit-code assertion runs through THIS door
    # (tests/test_pipeline_golden.py), so it was blind to any code main() returned. No behaviour
    # change today beyond that: main() returns 0 on every path unless --fail-on-gate-refusal is
    # passed and a gate refused.
    sys.exit(main())
